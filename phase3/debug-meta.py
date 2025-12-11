import cv2
import numpy as np
import os
import glob
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
from io import BytesIO

# --------------------------------------------------
# TUNABLE PARAMETERS
# --------------------------------------------------
GAUSSIAN_SIGMA = 15
MORPH_ITERATIONS = 1
DILATE_ITERATIONS = 1
ERODE_ITERATIONS = 1
MIN_OBJECT_AREA = 100
MAX_OBJECT_AREA = 5000

# Red channel visualization parameters
RED_NORMALIZE = True
RED_BRIGHTNESS = 1.2
RED_GAMMA = 0.7

# Scale bar parameters
SCALE_BAR_LENGTH_UM = 20  # Length of scale bar in micrometers
SCALE_BAR_HEIGHT = 5      # Height of scale bar in pixels
SCALE_BAR_MARGIN = 15     # Margin from bottom-right corner
SCALE_BAR_COLOR = (255, 255, 255)  # White
SCALE_BAR_BG_COLOR = (0, 0, 0)     # Black background
SCALE_BAR_TEXT_COLOR = (255, 255, 255)  # White text
SCALE_BAR_FONT_SCALE = 0.6
SCALE_BAR_FONT_THICKNESS = 2

# Error bar plotting parameters
ERROR_PERCENTAGE = 0.1  # 10% error for demonstration
PLOT_DPI = 150  # Resolution for embedded plots

# --------------------------------------------------
# DEBUG FOLDER SETUP
# --------------------------------------------------
DEBUG_DIR = "debug"
os.makedirs(DEBUG_DIR, exist_ok=True)

for f in glob.glob(os.path.join(DEBUG_DIR, "*")):
    try:
        os.remove(f)
    except OSError:
        pass

def add_scale_bar(img, pixel_size, unit='um', length_um=50):
    """
    Add a scale bar to the image
    
    Parameters:
    - img: Image to add scale bar to (will be modified in place)
    - pixel_size: Physical size per pixel in micrometers
    - unit: Unit string (default 'um' for micrometers)
    - length_um: Desired length of scale bar in micrometers
    
    Returns:
    - Modified image with scale bar
    """
    if pixel_size is None or pixel_size <= 0:
        return img  # No calibration available
    
    # Calculate scale bar length in pixels
    bar_length_px = int(round(length_um / pixel_size))
    
    # Debug: print actual scale bar calculation
    print(f"Scale bar calculation: {length_um} um / {pixel_size:.6f} um/pixel = {bar_length_px} pixels")
    
    if bar_length_px < 10:
        return img  # Scale bar would be too small
    
    # Get image dimensions
    h, w = img.shape[:2]
    
    # Position scale bar in bottom-right corner
    bar_x = w - bar_length_px - SCALE_BAR_MARGIN
    bar_y = h - SCALE_BAR_HEIGHT - SCALE_BAR_MARGIN
    
    # Create text label - use 'um' instead of µ for better compatibility
    if unit == 'µm' or unit == 'um':
        label = f"{length_um} um"
    else:
        label = f"{length_um} {unit}"
    
    font = cv2.FONT_HERSHEY_SIMPLEX
    
    # Get text size
    (text_w, text_h), baseline = cv2.getTextSize(
        label, font, SCALE_BAR_FONT_SCALE, SCALE_BAR_FONT_THICKNESS
    )
    
    # Position text above the bar
    text_x = bar_x + (bar_length_px - text_w) // 2
    text_y = bar_y - 8
    
    # Calculate background rectangle for better visibility
    bg_padding = 5
    bg_x1 = min(bar_x, text_x) - bg_padding
    bg_y1 = text_y - text_h - bg_padding
    bg_x2 = max(bar_x + bar_length_px, text_x + text_w) + bg_padding
    bg_y2 = bar_y + SCALE_BAR_HEIGHT + bg_padding
    
    # Draw semi-transparent background
    overlay = img.copy()
    cv2.rectangle(overlay, (bg_x1, bg_y1), (bg_x2, bg_y2), 
                  SCALE_BAR_BG_COLOR, -1)
    cv2.addWeighted(overlay, 0.7, img, 0.3, 0, img)
    
    # Draw the scale bar
    cv2.rectangle(img, 
                  (bar_x, bar_y), 
                  (bar_x + bar_length_px, bar_y + SCALE_BAR_HEIGHT),
                  SCALE_BAR_COLOR, -1)
    
    # Draw text label
    cv2.putText(img, label, (text_x, text_y), 
                font, SCALE_BAR_FONT_SCALE, SCALE_BAR_TEXT_COLOR, 
                SCALE_BAR_FONT_THICKNESS, cv2.LINE_AA)
    
    return img

def save_debug(name, img, pixel_size=None, unit='um'):
    """
    Save debug image with optional scale bar
    
    Parameters:
    - name: Filename for the debug image
    - img: Image to save
    - pixel_size: Physical size per pixel (if None, no scale bar added)
    - unit: Unit string for scale bar
    """
    path = os.path.join(DEBUG_DIR, name)
    
    # Create a copy to avoid modifying original
    img_with_scale = img.copy()
    
    # Add scale bar if calibration is available
    if pixel_size is not None and pixel_size > 0:
        img_with_scale = add_scale_bar(img_with_scale, pixel_size, unit, SCALE_BAR_LENGTH_UM)
    
    cv2.imwrite(path, img_with_scale)

def create_error_bar_plot(object_data, unit_str='um', error_percentage=0.1):
    """
    Create error bar plot for fluorescence intensity
    
    Parameters:
    - object_data: List of dictionaries with particle measurements
    - unit_str: Unit string for axis labels
    - error_percentage: Percentage of value to use as error (default 10%)
    
    Returns:
    - BytesIO object containing PNG image
    """
    particle_ids = [obj['object_id'] for obj in object_data]
    intensities = [obj['intensity_per_area_orig'] for obj in object_data]
    
    # Calculate errors (using percentage of value or standard deviation from data)
    errors = [intensity * error_percentage for intensity in intensities]
    
    # Create figure
    plt.figure(figsize=(12, 6))
    plt.errorbar(particle_ids, intensities, yerr=errors, 
                 fmt='o-', capsize=5, capthick=2, 
                 ecolor='red', markersize=6, color='blue',
                 linewidth=1.5, label=f'Intensity ± {int(error_percentage*100)}% error')
    
    plt.xlabel('Particle ID', fontsize=12, fontweight='bold')
    plt.ylabel(f'Total Intensity per {unit_str}²', fontsize=12, fontweight='bold')
    plt.title('Fluorescence Intensity by Particle with Error Bars', 
              fontsize=14, fontweight='bold')
    plt.grid(True, alpha=0.3, linestyle='--')
    plt.legend(fontsize=10)
    plt.tight_layout()
    
    # Save to BytesIO object
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', dpi=PLOT_DPI, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()
    
    return img_buffer

def create_bar_chart_with_errors(object_data, unit_str='um', error_percentage=0.1):
    """
    Create bar chart with error bars for fluorescence intensity
    
    Parameters:
    - object_data: List of dictionaries with particle measurements
    - unit_str: Unit string for axis labels
    - error_percentage: Percentage of value to use as error
    
    Returns:
    - BytesIO object containing PNG image
    """
    particle_ids = [obj['object_id'] for obj in object_data]
    intensities = [obj['intensity_per_area_orig'] for obj in object_data]
    errors = [intensity * error_percentage for intensity in intensities]
    
    # Create figure
    plt.figure(figsize=(14, 6))
    bars = plt.bar(particle_ids, intensities, yerr=errors, 
                   capsize=5, color='skyblue', edgecolor='navy', linewidth=1.5,
                   error_kw={'elinewidth': 2, 'capthick': 2, 'ecolor': 'darkred'})
    
    # Color gradient for bars
    colors = plt.cm.get_cmap('RdYlGn_r')(np.linspace(0.2, 0.8, len(bars)))
    for bar, color in zip(bars, colors):
        bar.set_color(color)
    
    plt.xlabel('Particle ID', fontsize=12, fontweight='bold')
    plt.ylabel(f'Total Intensity per {unit_str}²', fontsize=12, fontweight='bold')
    plt.title('Fluorescence Intensity Distribution with Error Bars', 
              fontsize=14, fontweight='bold')
    plt.xticks(particle_ids)
    plt.grid(True, axis='y', alpha=0.3, linestyle='--')
    plt.tight_layout()
    
    # Save to BytesIO object
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', dpi=PLOT_DPI, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()
    
    return img_buffer

def create_statistics_plot(object_data, unit_str='um'):
    """
    Create summary statistics plot with mean and SEM
    
    Parameters:
    - object_data: List of dictionaries with particle measurements
    - unit_str: Unit string for axis labels
    
    Returns:
    - BytesIO object containing PNG image
    """
    intensities = np.array([obj['intensity_per_area_orig'] for obj in object_data])
    
    # Calculate statistics
    mean_intensity = np.mean(intensities)
    std_intensity = np.std(intensities, ddof=1)
    sem_intensity = std_intensity / np.sqrt(len(intensities))
    
    # Create figure
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
    
    # Plot 1: Mean with SEM
    ax1.bar(['Mean Intensity'], [mean_intensity], 
            yerr=[sem_intensity], capsize=20, 
            color='lightcoral', edgecolor='darkred', linewidth=2,
            error_kw={'elinewidth': 3, 'capthick': 3, 'ecolor': 'black'})
    ax1.set_ylabel(f'Total Intensity per {unit_str}²', fontsize=11, fontweight='bold')
    ax1.set_title(f'Average Fluorescence Intensity\n(n={len(intensities)} particles)', 
                  fontsize=12, fontweight='bold')
    ax1.grid(True, axis='y', alpha=0.3)
    ax1.text(0, mean_intensity + sem_intensity + 200, 
             f'Mean: {mean_intensity:.2f}\nSEM: ±{sem_intensity:.2f}\nSD: ±{std_intensity:.2f}',
             ha='center', fontsize=9, 
             bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
    
    # Plot 2: Histogram with statistics
    ax2.hist(intensities, bins=15, color='steelblue', edgecolor='black', alpha=0.7)
    ax2.axvline(mean_intensity, color='red', linestyle='--', linewidth=2, label='Mean')
    ax2.axvline(mean_intensity + std_intensity, color='orange', linestyle=':', linewidth=2, label='Mean ± SD')
    ax2.axvline(mean_intensity - std_intensity, color='orange', linestyle=':', linewidth=2)
    ax2.set_xlabel(f'Total Intensity per {unit_str}²', fontsize=11, fontweight='bold')
    ax2.set_ylabel('Frequency', fontsize=11, fontweight='bold')
    ax2.set_title('Intensity Distribution', fontsize=12, fontweight='bold')
    ax2.legend()
    ax2.grid(True, alpha=0.3)
    
    plt.tight_layout()
    
    # Save to BytesIO object
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', dpi=PLOT_DPI, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()
    
    return img_buffer

def parse_metadata(xml_path):
    """
    Parse metadata XML to extract physical pixel size and bit depth
    Returns: (pixel_size_x, pixel_size_y, unit, bit_depth) or (None, None, None, None)
    """
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        
        pixel_size_x = None
        pixel_size_y = None
        unit = None
        bit_depth = None
        
        # Method 1: Look for DimensionDescription tags (Leica LAS format)
        for dim in root.iter('DimensionDescription'):
            dim_id = dim.get('DimID')
            length = dim.get('Length')
            num_elements = dim.get('NumberOfElements')
            dim_unit = dim.get('Unit')
            
            if length and num_elements:
                pixel_size = float(length) / float(num_elements)
                
                if dim_id == 'X':
                    pixel_size_x = pixel_size
                    if dim_unit:
                        unit = dim_unit
                elif dim_id == 'Y':
                    pixel_size_y = pixel_size
                    if dim_unit and not unit:
                        unit = dim_unit
        
        # Method 2: Look for ChannelDescription with OpticalResolutionXY
        if not pixel_size_x or not pixel_size_y:
            for channel in root.iter('ChannelDescription'):
                optical_res = channel.get('OpticalResolutionXY')
                if optical_res:
                    parts = optical_res.split()
                    if len(parts) >= 2:
                        pixel_size_x = pixel_size_y = float(parts[0])
                        unit = parts[1]
                
                # Get bit depth from Resolution attribute
                if not bit_depth:
                    resolution = channel.get('Resolution')
                    if resolution:
                        bit_depth = int(resolution)
        
        # Convert unit to 'um' for consistency
        if unit in ['µm', 'μm']:
            unit = 'um'
        
        # Default unit if not found
        if pixel_size_x and pixel_size_y and not unit:
            unit = 'um'
        
        if pixel_size_x and pixel_size_y:
            print(f"✓ Metadata parsed successfully:")
            print(f"  - X dimension: {pixel_size_x:.6f} {unit}/pixel")
            print(f"  - Y dimension: {pixel_size_y:.6f} {unit}/pixel")
            if bit_depth:
                print(f"  - Bit depth: {bit_depth}-bit")
            return pixel_size_x, pixel_size_y, unit, bit_depth
        else:
            print("⚠ Pixel size not found in metadata")
            return None, None, None, bit_depth
            
    except Exception as e:
        print(f"⚠ Error parsing metadata: {e}")
        return None, None, None, None

def adjust_red_channel(img, normalize=True, brightness=1.0, gamma=1.0):
    """
    Adjust red channel with normalization, brightness, and gamma
    """
    img_float = img.astype(np.float32)
    
    if normalize:
        min_val = np.min(img_float)
        max_val = np.max(img_float)
        if max_val > min_val:
            img_float = (img_float - min_val) * 255.0 / (max_val - min_val)
        print(f"Red channel normalized: [{min_val:.1f}, {max_val:.1f}] -> [0, 255]")
    
    if brightness != 1.0:
        img_float = img_float * brightness
        img_float = np.clip(img_float, 0, 255)
    
    if gamma != 1.0:
        img_normalized = img_float / 255.0
        img_normalized = np.power(img_normalized, gamma)
        img_float = img_normalized * 255.0
    
    return img_float.astype(np.uint8)

def style_header(ws, row_num, fill_color='4472C4'):
    """Apply header styling to a row"""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(bold=True, color='FFFFFF', size=11)
    alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment

def style_summary_header(cell):
    """Apply summary section header styling"""
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# --------------------------------------------------
# LOAD IMAGES AND METADATA
# --------------------------------------------------
grey_path = "source/12/12 N NO 1_ch00.tif"
red_path = "source/12/12 N NO 1_ch01.tif"
meta_path = "source/12/MetaData/12 N NO 1_Properties.xml"

print("="*60)
print("LOADING METADATA AND IMAGES")
print("="*60)

# Parse metadata for physical units AND bit depth
pixel_size_x, pixel_size_y, unit, metadata_bit_depth = parse_metadata(meta_path)

# Calculate conversion factors for physical dimensions
if pixel_size_x is not None and pixel_size_y is not None:
    pixel_size = (pixel_size_x + pixel_size_y) / 2.0
    area_factor = pixel_size ** 2
    unit_str = unit if unit else 'um'
    has_calibration = True
    print(f"✓ Using physical scale: {pixel_size:.6f} {unit_str}/pixel")
    print(f"✓ Area conversion factor: {area_factor:.10f} {unit_str}²/pixel²")
    print(f"✓ Scale bars will show {SCALE_BAR_LENGTH_UM} {unit_str}")
else:
    pixel_size = None
    area_factor = 1.0
    unit_str = 'pixels'
    has_calibration = False
    print("⚠ Using pixel units (no physical calibration found)")
    print("⚠ Scale bars will NOT be added to images")

# --------------------------------------------------
# LOAD BRIGHTFIELD (for contour detection)
# --------------------------------------------------
img_bf = cv2.imread(grey_path, cv2.IMREAD_UNCHANGED)
if img_bf is None:
    raise FileNotFoundError(grey_path)

if img_bf.ndim == 3:
    img_bf = cv2.cvtColor(img_bf, cv2.COLOR_BGR2GRAY)

print(f"\nBrightfield loaded: dtype={img_bf.dtype}, shape={img_bf.shape}, range=[{img_bf.min()}, {img_bf.max()}]")

# --------------------------------------------------
# LOAD RED FLUORESCENCE CHANNEL (KEEP ORIGINAL)
# --------------------------------------------------
img_red_original = cv2.imread(red_path, cv2.IMREAD_UNCHANGED)
if img_red_original is None:
    raise FileNotFoundError(red_path)

if img_red_original.ndim == 3:
    img_red_original = cv2.cvtColor(img_red_original, cv2.COLOR_BGR2GRAY)

original_dtype = img_red_original.dtype
original_min = img_red_original.min()
original_max = img_red_original.max()

print(f"Red channel loaded: dtype={original_dtype}, shape={img_red_original.shape}, range=[{original_min}, {original_max}]")

# --------------------------------------------------
# DETERMINE BIT DEPTH AND SCALING FACTOR
# --------------------------------------------------
if original_dtype == np.uint16:
    # Detect actual bit depth from data
    if original_max <= 4095:
        actual_bit_depth = 12
        max_possible_value = 4095
    elif original_max <= 16383:
        actual_bit_depth = 14
        max_possible_value = 16383
    else:
        actual_bit_depth = 16
        max_possible_value = 65535
    
    # Cross-check with metadata if available
    if metadata_bit_depth and metadata_bit_depth != actual_bit_depth:
        print(f"⚠ Note: Metadata says {metadata_bit_depth}-bit, but data suggests {actual_bit_depth}-bit")
        print(f"  Using actual data bit depth: {actual_bit_depth}-bit")
    
    bit_depth = actual_bit_depth
    
    # Calculate the scaling factor for 12-bit to 8-bit conversion
    bit_conversion_factor = max_possible_value / 255.0
    
    print(f"✓ Detected {bit_depth}-bit image (max value: {max_possible_value})")
    print(f"✓ Bit conversion factor: {bit_conversion_factor:.4f} ({bit_depth}-bit units per 8-bit unit)")
    
else:
    bit_depth = 8
    max_possible_value = 255
    bit_conversion_factor = 1.0
    print(f"✓ Detected 8-bit image (no conversion needed)")

# --------------------------------------------------
# CONVERT TO 8-BIT FOR PROCESSING
# --------------------------------------------------
print("\n" + "="*60)
print("BIT DEPTH CONVERSION STRATEGY")
print("="*60)

if img_bf.dtype == np.uint16:
    img_bf_8bit = np.zeros_like(img_bf, dtype=np.uint8)
    cv2.normalize(img_bf, img_bf_8bit, 0, 255, cv2.NORM_MINMAX)
    img_bf = img_bf_8bit
    save_debug("01a_bf_converted_8bit.png", img_bf, pixel_size, unit_str)
    print(f"Brightfield: Converted to 8-bit for segmentation")

# Create 8-bit version of red channel for visualization
if img_red_original.dtype == np.uint16:
    img_red_8bit = np.zeros_like(img_red_original, dtype=np.uint8)
    cv2.normalize(img_red_original, img_red_8bit, 0, 255, cv2.NORM_MINMAX)
    save_debug("01b_red_converted_8bit.png", img_red_8bit, pixel_size, unit_str)
    print(f"Red channel: Created 8-bit version for visualization")
    print(f"Red channel: Original {bit_depth}-bit data PRESERVED for measurements")
else:
    img_red_8bit = img_red_original.copy()
    print(f"Red channel: Already 8-bit, using as-is")

print(f"\n✓ Fluorescence measurements will use: Original {bit_depth}-bit data")
print(f"✓ Conversion factor: {bit_conversion_factor:.4f}")
print("="*60 + "\n")

# --------------------------------------------------
# APPLY ENHANCEMENT TO 8-BIT VERSION (FOR VISUALIZATION)
# --------------------------------------------------
img_red_enhanced = adjust_red_channel(
    img_red_8bit,
    normalize=RED_NORMALIZE,
    brightness=RED_BRIGHTNESS,
    gamma=RED_GAMMA
)
save_debug("01c_red_enhanced.png", img_red_enhanced, pixel_size, unit_str)

# --------------------------------------------------
# BRIGHTFIELD PROCESSING (SEGMENTATION)
# --------------------------------------------------
bg = cv2.GaussianBlur(img_bf, (0, 0), sigmaX=GAUSSIAN_SIGMA, sigmaY=GAUSSIAN_SIGMA)
enhanced = cv2.subtract(bg, img_bf)
enhanced_blur = cv2.GaussianBlur(enhanced, (3, 3), 0)

save_debug("02_enhanced.png", enhanced, pixel_size, unit_str)
save_debug("03_enhanced_blur.png", enhanced_blur, pixel_size, unit_str)

_, thresh = cv2.threshold(
    enhanced_blur, 0, 255,
    cv2.THRESH_BINARY + cv2.THRESH_OTSU
)

save_debug("04_thresh_raw.png", thresh, pixel_size, unit_str)

kernel = np.ones((3, 3), np.uint8)
closed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=MORPH_ITERATIONS)
closed = cv2.dilate(closed, kernel, iterations=DILATE_ITERATIONS)
closed = cv2.erode(closed, kernel, iterations=ERODE_ITERATIONS)

save_debug("05_closed.png", closed, pixel_size, unit_str)

num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(
    closed, connectivity=8
)

solid = np.where(labels > 0, 255, 0).astype(np.uint8)
save_debug("06_solid.png", solid, pixel_size, unit_str)

contours, hierarchy = cv2.findContours(
    solid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
)
print(f"Clumps found: {len(contours)}")

# Apply area filtering
min_area = MIN_OBJECT_AREA
max_area = MAX_OBJECT_AREA
filtered = [c for c in contours if min_area <= cv2.contourArea(c) <= max_area]
print(f"Filtered objects (by area {min_area}-{max_area} pixels): {len(filtered)}")

# --------------------------------------------------
# MEASURE RED FLUORESCENCE INTENSITY
# --------------------------------------------------
print(f"\n{'='*60}")
print(f"FLUORESCENCE MEASUREMENT PROTOCOL")
print(f"{'='*60}")
print(f"Source Data: {bit_depth}-bit original data")
print(f"Value Range: 0 to {max_possible_value}")
print(f"Spatial Unit: {unit_str}")
if has_calibration:
    print(f"Area Unit: {unit_str}²")
    print(f"Pixel Size: {pixel_size:.6f} {unit_str}/pixel")
    print(f"Area Factor: {area_factor:.10f} {unit_str}²/pixel²")
if bit_conversion_factor != 1.0:
    print(f"Bit Conversion: {bit_conversion_factor:.4f} ({bit_depth}-bit/8-bit)")
print(f"{'='*60}\n")

object_data = []
objects_without_red = 0

for c in filtered:
    area_px = cv2.contourArea(c)
    perimeter_px = cv2.arcLength(c, True)
    
    # Convert to physical units
    if has_calibration and pixel_size is not None:
        area_physical = area_px * area_factor
        perimeter_physical = perimeter_px * pixel_size
    else:
        area_physical = area_px
        perimeter_physical = perimeter_px
    
    # Calculate centroid
    M = cv2.moments(c)
    if M["m00"] != 0:
        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])
    else:
        cx, cy = 0, 0
    
    # Create mask
    mask = np.zeros_like(img_red_8bit, dtype=np.uint8)
    cv2.drawContours(mask, [c], -1, 255, -1)
    
    # ===== MEASUREMENTS FROM ORIGINAL BIT DEPTH =====
    red_pixels_orig = img_red_original[mask == 255]
    if len(red_pixels_orig) > 0:
        red_pixels_float = red_pixels_orig.astype(np.float64)
        red_total_orig = float(np.sum(red_pixels_float))
        red_mean_orig = float(np.mean(red_pixels_float))
        red_std_orig = float(np.std(red_pixels_float))
    else:
        red_total_orig = 0.0
        red_mean_orig = 0.0
        red_std_orig = 0.0
    
    # Filter out objects with no red intensity
    if red_total_orig == 0.0:
        objects_without_red += 1
        continue
    
    # Calculate intensity per physical area
    intensity_per_area_orig = red_total_orig / area_physical if area_physical > 0 else 0.0
    mean_intensity_per_area_orig = red_mean_orig / area_factor if area_factor > 0 else red_mean_orig
    
    # ===== MEASUREMENTS FROM ENHANCED 8-BIT (FOR COMPARISON) =====
    red_pixels_enh = img_red_enhanced[mask == 255]
    if len(red_pixels_enh) > 0:
        red_pixels_float = red_pixels_enh.astype(np.float64)
        red_total_enh = float(np.sum(red_pixels_float))
        red_mean_enh = float(np.mean(red_pixels_float))
    else:
        red_total_enh = 0.0
        red_mean_enh = 0.0
    
    intensity_per_area_enh = red_total_enh / area_physical if area_physical > 0 else 0.0
    
    # ===== ALSO MEASURE FROM NORMALIZED 8-BIT (TO SHOW CONVERSION) =====
    red_pixels_8bit = img_red_8bit[mask == 255]
    if len(red_pixels_8bit) > 0:
        red_pixels_float = red_pixels_8bit.astype(np.float64)
        red_total_8bit = float(np.sum(red_pixels_float))
        red_mean_8bit = float(np.mean(red_pixels_float))
        
        # Scale back to original bit depth
        red_total_8bit_scaled = red_total_8bit * bit_conversion_factor
        red_mean_8bit_scaled = red_mean_8bit * bit_conversion_factor
    else:
        red_total_8bit = 0.0
        red_mean_8bit = 0.0
        red_total_8bit_scaled = 0.0
        red_mean_8bit_scaled = 0.0
    
    intensity_per_area_8bit_scaled = red_total_8bit_scaled / area_physical if area_physical > 0 else 0.0
    
    object_data.append({
        'contour': c,
        'area_px': area_px,
        'area_physical': area_physical,
        'perimeter_px': perimeter_px,
        'perimeter_physical': perimeter_physical,
        'centroid_x': cx,
        'centroid_y': cy,
        # Original bit-depth measurements (GROUND TRUTH)
        'red_total_orig': red_total_orig,
        'red_mean_orig': red_mean_orig,
        'red_std_orig': red_std_orig,
        'intensity_per_area_orig': intensity_per_area_orig,
        'mean_intensity_per_area_orig': mean_intensity_per_area_orig,
        # 8-bit normalized (with scaling back)
        'red_total_8bit': red_total_8bit,
        'red_mean_8bit': red_mean_8bit,
        'red_total_8bit_scaled': red_total_8bit_scaled,
        'red_mean_8bit_scaled': red_mean_8bit_scaled,
        'intensity_per_area_8bit_scaled': intensity_per_area_8bit_scaled,
        # Enhanced 8-bit measurements (FOR VISUALIZATION REFERENCE)
        'red_total_enh': red_total_enh,
        'red_mean_enh': red_mean_enh,
        'intensity_per_area_enh': intensity_per_area_enh,
    })

print(f"Objects with red intensity: {len(object_data)}")
print(f"Objects without red intensity (removed): {objects_without_red}")

# --------------------------------------------------
# SORT BY INTENSITY PER AREA (ORIGINAL)
# --------------------------------------------------
object_data.sort(key=lambda x: x['intensity_per_area_orig'], reverse=True)

for i, obj in enumerate(object_data, 1):
    obj['object_id'] = i

print(f"Objects sorted by Fluorescence Intensity per Area (descending)")

# --------------------------------------------------
# CREATE PLOTS
# --------------------------------------------------
print("\n" + "="*60)
print("GENERATING ERROR BAR PLOTS")
print("="*60)

excel_unit = 'um' if unit_str in ['um', 'µm', 'μm'] else unit_str

# Create error bar plot
print("Creating line plot with error bars...")
errorbar_plot = create_error_bar_plot(object_data, excel_unit, ERROR_PERCENTAGE)

# Create bar chart with error bars
print("Creating bar chart with error bars...")
barchart_plot = create_bar_chart_with_errors(object_data, excel_unit, ERROR_PERCENTAGE)

# Create statistics plot
print("Creating statistics summary plot...")
statistics_plot = create_statistics_plot(object_data, excel_unit)

print("✓ All plots generated successfully")

# --------------------------------------------------
# EXPORT STATISTICS TO EXCEL
# --------------------------------------------------
excel_file = os.path.join(DEBUG_DIR, "fluorescence_statistics.xlsx")
wb = Workbook()

if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Determine precision
if has_calibration:
    area_precision = 4
    perimeter_precision = 2
else:
    area_precision = 2
    perimeter_precision = 2

# --------------------------------------------------
# SHEET 1: PRIMARY DATA WITH ERROR BAR PLOTS
# --------------------------------------------------
ws_orig = wb.create_sheet(f"Primary Data ({bit_depth}-bit)")

headers_orig = [
    'Particle ID',
    f'Area ({excel_unit}²)' if has_calibration else 'Area (pixels²)',
    f'Perimeter ({excel_unit})' if has_calibration else 'Perimeter (pixels)',
    f'Total Intensity ({bit_depth}-bit)',
    f'Mean Intensity ({bit_depth}-bit)',
    f'Std Dev',
    f'Total Int. per {excel_unit}²' if has_calibration else 'Total Int. per pixel²',
    f'Mean Int. per {excel_unit}²' if has_calibration else 'Mean Int. per pixel²'
]
ws_orig.append(headers_orig)
style_header(ws_orig, 1)

for obj in object_data:
    ws_orig.append([
        obj['object_id'],
        round(obj['area_physical'], area_precision),
        round(obj['perimeter_physical'], perimeter_precision),
        round(obj['red_total_orig'], 1),
        round(obj['red_mean_orig'], 2),
        round(obj['red_std_orig'], 2),
        round(obj['intensity_per_area_orig'], 2),
        round(obj['mean_intensity_per_area_orig'], 2)
    ])

# Summary
summary_start = len(object_data) + 3
ws_orig.append([])
ws_orig.append(['SUMMARY STATISTICS'])
style_summary_header(ws_orig.cell(summary_start, 1))

ws_orig.append(['Total Particles:', len(object_data)])
ws_orig.append(['Particles Excluded (no fluorescence):', objects_without_red])
ws_orig.append([])
ws_orig.append(['Avg Area:', round(np.mean([obj['area_physical'] for obj in object_data]), area_precision), 
                f'{excel_unit}²' if has_calibration else 'pixels²'])
ws_orig.append(['Avg Perimeter:', round(np.mean([obj['perimeter_physical'] for obj in object_data]), perimeter_precision), 
                f'{excel_unit}' if has_calibration else 'pixels'])
ws_orig.append([])
ws_orig.append(['Avg Total Intensity:', round(np.mean([obj['red_total_orig'] for obj in object_data]), 1)])
ws_orig.append(['Avg Mean Intensity:', round(np.mean([obj['red_mean_orig'] for obj in object_data]), 2)])
ws_orig.append(['Avg Intensity per Area:', round(np.mean([obj['intensity_per_area_orig'] for obj in object_data]), 2), 
                f'per {excel_unit}²' if has_calibration else 'per pixel²'])

ws_orig.append([])
ws_orig.append(['MEASUREMENT DETAILS'])
style_summary_header(ws_orig.cell(summary_start + 11, 1))
ws_orig.append(['Data Source:', f'Original {bit_depth}-bit uncompressed data'])
ws_orig.append(['Bit Depth:', f'{bit_depth}-bit'])
ws_orig.append(['Value Range:', f'0 to {max_possible_value}'])
ws_orig.append(['Data Type:', 'Ground truth measurements'])
ws_orig.append(['Error Bar %:', f'{int(ERROR_PERCENTAGE * 100)}% (for visualization)'])

if has_calibration and pixel_size is not None:
    ws_orig.append([])
    ws_orig.append(['PHYSICAL CALIBRATION'])
    style_summary_header(ws_orig.cell(summary_start + 18, 1))
    ws_orig.append(['Status:', 'Active'])
    ws_orig.append(['Pixel Size:', round(pixel_size, 6), f'{excel_unit}/pixel'])
    ws_orig.append(['Area Factor:', round(area_factor, 10), f'{excel_unit}²/pixel²'])
    ws_orig.append(['Linear Unit:', excel_unit])
    ws_orig.append(['Area Unit:', f'{excel_unit}²'])
    ws_orig.append(['Scale Bar:', f'{SCALE_BAR_LENGTH_UM} {excel_unit}'])

# Add plots to the sheet
print("Embedding error bar plot (line) in Excel...")
img_errorbar = XLImage(errorbar_plot)
img_errorbar.width = 900
img_errorbar.height = 450
ws_orig.add_image(img_errorbar, 'K2')

print("Embedding error bar plot (bar chart) in Excel...")
img_barchart = XLImage(barchart_plot)
img_barchart.width = 1050
img_barchart.height = 450
ws_orig.add_image(img_barchart, 'K28')

print("Embedding statistics plot in Excel...")
img_statistics = XLImage(statistics_plot)
img_statistics.width = 900
img_statistics.height = 375
ws_orig.add_image(img_statistics, 'K54')

auto_adjust_column_width(ws_orig)

# --------------------------------------------------
# SHEET 2: 8-BIT CONVERTED WITH SCALING
# --------------------------------------------------
ws_8bit = wb.create_sheet("8-bit Converted (Scaled)")

headers_8bit = [
    'Particle ID',
    f'Area ({excel_unit}²)' if has_calibration else 'Area (pixels²)',
    f'Perimeter ({excel_unit})' if has_calibration else 'Perimeter (pixels)',
    'Total (8-bit)',
    'Mean (8-bit)',
    f'Total (scaled to {bit_depth}-bit)',
    f'Mean (scaled to {bit_depth}-bit)',
    f'Total Int. per {excel_unit}²' if has_calibration else 'Total Int. per pixel²',
    'Recovery Accuracy (%)'
]
ws_8bit.append(headers_8bit)
style_header(ws_8bit, 1)

for obj in object_data:
    # Calculate recovery accuracy
    if obj['red_total_orig'] > 0:
        accuracy = (obj['red_total_8bit_scaled'] / obj['red_total_orig']) * 100
    else:
        accuracy = 0.0
    
    ws_8bit.append([
        obj['object_id'],
        round(obj['area_physical'], area_precision),
        round(obj['perimeter_physical'], perimeter_precision),
        round(obj['red_total_8bit'], 1),
        round(obj['red_mean_8bit'], 2),
        round(obj['red_total_8bit_scaled'], 1),
        round(obj['red_mean_8bit_scaled'], 2),
        round(obj['intensity_per_area_8bit_scaled'], 2),
        round(accuracy, 1)
    ])

# Summary
ws_8bit.append([])
ws_8bit.append(['SUMMARY'])
style_summary_header(ws_8bit.cell(summary_start, 1))

ws_8bit.append(['Total Particles:', len(object_data)])
ws_8bit.append(['Avg 8-bit Total:', round(np.mean([obj['red_total_8bit'] for obj in object_data]), 1)])
ws_8bit.append(['Avg Scaled Total:', round(np.mean([obj['red_total_8bit_scaled'] for obj in object_data]), 1)])
ws_8bit.append(['Avg Recovery Accuracy:', 
                round(np.mean([(obj['red_total_8bit_scaled']/obj['red_total_orig']*100) 
                               for obj in object_data if obj['red_total_orig'] > 0]), 1), '%'])

ws_8bit.append([])
ws_8bit.append(['CONVERSION DETAILS'])
style_summary_header(ws_8bit.cell(summary_start + 6, 1))
ws_8bit.append(['Conversion:', f'{bit_depth}-bit → 8-bit → scaled back to {bit_depth}-bit'])
ws_8bit.append(['Scaling Factor:', round(bit_conversion_factor, 4)])
ws_8bit.append(['Formula:', f'Original ≈ (8-bit value) × {bit_conversion_factor:.4f}'])
ws_8bit.append(['Note:', 'This demonstrates information loss from bit-depth reduction'])

auto_adjust_column_width(ws_8bit)

# --------------------------------------------------
# SHEET 3: ENHANCED 8-BIT (VISUALIZATION REFERENCE)
# --------------------------------------------------
ws_enh = wb.create_sheet("Enhanced 8-bit (Visual)")

headers_enh = [
    'Particle ID',
    f'Area ({excel_unit}²)' if has_calibration else 'Area (pixels²)',
    f'Perimeter ({excel_unit})' if has_calibration else 'Perimeter (pixels)',
    'Total (enhanced)',
    'Mean (enhanced)',
    f'Total Int. per {excel_unit}²' if has_calibration else 'Total Int. per pixel²'
]
ws_enh.append(headers_enh)
style_header(ws_enh, 1)

for obj in object_data:
    ws_enh.append([
        obj['object_id'],
        round(obj['area_physical'], area_precision),
        round(obj['perimeter_physical'], perimeter_precision),
        round(obj['red_total_enh'], 1),
        round(obj['red_mean_enh'], 2),
        round(obj['intensity_per_area_enh'], 2)
    ])

# Summary
ws_enh.append([])
ws_enh.append(['SUMMARY'])
style_summary_header(ws_enh.cell(summary_start, 1))

ws_enh.append(['Total Particles:', len(object_data)])
ws_enh.append(['Avg Total Fluorescence:', round(np.mean([obj['red_total_enh'] for obj in object_data]), 1)])
ws_enh.append(['Avg Intensity per Area:', round(np.mean([obj['intensity_per_area_enh'] for obj in object_data]), 2)])

ws_enh.append([])
ws_enh.append(['ENHANCEMENT SETTINGS'])
style_summary_header(ws_enh.cell(summary_start + 5, 1))
ws_enh.append(['Normalize:', RED_NORMALIZE])
ws_enh.append(['Brightness:', RED_BRIGHTNESS])
ws_enh.append(['Gamma:', RED_GAMMA])

ws_enh.append([])
ws_enh.append(['IMPORTANT NOTE'])
style_summary_header(ws_enh.cell(summary_start + 10, 1))
ws_enh.append(['', 'This data is for VISUALIZATION purposes only'])
ws_enh.append(['', f'Use "Primary Data ({bit_depth}-bit)" sheet for quantitative analysis'])
ws_enh.append(['', 'Enhancement applied for better visual contrast'])

auto_adjust_column_width(ws_enh)

# --------------------------------------------------
# SHEET 4: COMPREHENSIVE COMPARISON
# --------------------------------------------------
ws_comp = wb.create_sheet("Comparison Analysis")

ws_comp.append(['', '', '', f'{bit_depth}-BIT ORIGINAL', '', '8-BIT SCALED', '', 'ENHANCED 8-BIT'])
ws_comp.merge_cells('D1:E1')
ws_comp.merge_cells('F1:G1')
ws_comp.merge_cells('H1:I1')
style_header(ws_comp, 1, 'ED7D31')

unit_label = f'{excel_unit}²' if has_calibration else 'pixels²'
length_label = f'{excel_unit}' if has_calibration else 'pixels'

headers_comp = [
    'Particle ID',
    f'Area ({unit_label})',
    f'Perimeter ({length_label})',
    'Total',
    'Mean',
    'Total (scaled)',
    'Recovery %',
    'Total (enhanced)',
    'Visual Ratio'
]
ws_comp.append(headers_comp)
style_header(ws_comp, 2)

for obj in object_data:
    if obj['red_total_orig'] > 0:
        recovery = (obj['red_total_8bit_scaled'] / obj['red_total_orig']) * 100
        visual_ratio = obj['red_total_enh'] / obj['red_total_orig']
    else:
        recovery = 0.0
        visual_ratio = 0.0
    
    ws_comp.append([
        obj['object_id'],
        round(obj['area_physical'], area_precision),
        round(obj['perimeter_physical'], perimeter_precision),
        round(obj['red_total_orig'], 1),
        round(obj['red_mean_orig'], 2),
        round(obj['red_total_8bit_scaled'], 1),
        f"{recovery:.1f}%",
        round(obj['red_total_enh'], 1),
        f"{visual_ratio:.3f}x"
    ])

summary_start = len(object_data) + 4
ws_comp.append([])
ws_comp.append(['ANALYSIS SUMMARY'])
style_summary_header(ws_comp.cell(summary_start, 1))

avg_recovery = np.mean([(obj['red_total_8bit_scaled']/obj['red_total_orig']*100) 
                        for obj in object_data if obj['red_total_orig'] > 0])

ws_comp.append(['Total Particles:', len(object_data)])
ws_comp.append(['Particles Excluded:', objects_without_red])
ws_comp.append([])
ws_comp.append(['Average Recovery Accuracy:', f'{avg_recovery:.2f}%'])
ws_comp.append(['Conversion Factor Used:', round(bit_conversion_factor, 4)])
ws_comp.append(['Information Loss:', f'{100 - avg_recovery:.2f}%'])

ws_comp.append([])
ws_comp.append(['KEY FINDINGS'])
style_summary_header(ws_comp.cell(summary_start + 8, 1))
ws_comp.append(['', f'✓ Original {bit_depth}-bit data provides highest accuracy'])
ws_comp.append(['', f'✓ 8-bit conversion with scaling recovers ~{avg_recovery:.1f}% of information'])
ws_comp.append(['', '✓ Enhanced 8-bit useful for visualization, not quantification'])
if has_calibration and pixel_size is not None:
    ws_comp.append(['', f'✓ All spatial measurements in {excel_unit} (physical units)'])
    ws_comp.append(['', f'✓ Scale bars show {SCALE_BAR_LENGTH_UM} {excel_unit} on all images'])

ws_comp.append([])
ws_comp.append(['RECOMMENDATION'])
style_summary_header(ws_comp.cell(summary_start + 14, 1))
ws_comp.append(['', f'Always use "{bit_depth}-bit Original Data" for scientific measurements'])
ws_comp.append(['', 'Bit-depth conversion causes irreversible information loss'])
ws_comp.append(['', 'Enhanced versions are for visualization only'])

auto_adjust_column_width(ws_comp)

# Save workbook
wb.save(excel_file)

print(f"\n{'='*60}")
print(f"✓ Excel file saved: {excel_file}")
print(f"{'='*60}")
print(f"SHEETS CREATED:")
print(f"  1. Primary Data ({bit_depth}-bit) - WITH ERROR BAR PLOTS")
print(f"  2. 8-bit Converted (Scaled) - Shows conversion accuracy")
print(f"  3. Enhanced 8-bit (Visual) - Visualization reference")
print(f"  4. Comparison Analysis - Complete comparison")
print(f"\n✓ PLOTS EMBEDDED:")
print(f"  - Line plot with error bars")
print(f"  - Bar chart with error bars")
print(f"  - Statistics summary (mean ± SEM + histogram)")
print(f"  - Error percentage: {int(ERROR_PERCENTAGE * 100)}%")
print(f"\n✓ CALIBRATION:")
if has_calibration and pixel_size is not None:
    print(f"  - Physical scale: {pixel_size:.6f} {unit_str}/pixel")
    print(f"  - Area factor: {area_factor:.10f} {unit_str}²/pixel²")
    print(f"  - Scale bars: {SCALE_BAR_LENGTH_UM} {unit_str} (added to all images)")
else:
    print(f"  - No physical calibration (using pixel units)")
    print(f"  - Scale bars not added")
print(f"\n✓ BIT DEPTH:")
print(f"  - Source: {bit_depth}-bit (range: 0-{max_possible_value})")
print(f"  - Conversion factor: {bit_conversion_factor:.4f}")
if avg_recovery < 100:
    print(f"  ⚠ 8-bit conversion loses ~{100-avg_recovery:.1f}% of information")
print(f"{'='*60}\n")

# --------------------------------------------------
# VISUALIZATIONS (ALL WITH SCALE BARS)
# --------------------------------------------------
print("Creating visualizations with scale bars...")

# Brightfield with contours
vis_bf = cv2.cvtColor(img_bf, cv2.COLOR_GRAY2BGR)
for obj in object_data:
    cv2.drawContours(vis_bf, [obj['contour']], -1, (0, 0, 255), 1)
    label = str(obj['object_id'])
    cv2.putText(vis_bf, label, (obj['centroid_x'], obj['centroid_y']), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1, cv2.LINE_AA)
save_debug("07a_bf_contours.png", vis_bf, pixel_size, unit_str)

# Red channel with contours
vis_red_display = np.zeros((img_red_8bit.shape[0], img_red_8bit.shape[1], 3), dtype=np.uint8)
vis_red_display[:, :, 2] = img_red_enhanced
vis_red_display[:, :, 0] = img_red_enhanced // 8
vis_red_display[:, :, 1] = img_red_enhanced // 8

for obj in object_data:
    cv2.drawContours(vis_red_display, [obj['contour']], -1, (0, 255, 0), 1)
    label = str(obj['object_id'])
    cv2.putText(vis_red_display, label, (obj['centroid_x'], obj['centroid_y']), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1, cv2.LINE_AA)

save_debug("07b_red_contours.png", vis_red_display, pixel_size, unit_str)

# Overlay
vis_overlay = cv2.cvtColor(img_bf, cv2.COLOR_GRAY2BGR)
red_overlay = np.zeros_like(vis_overlay)
red_overlay[:, :, 2] = img_red_8bit
vis_overlay = cv2.addWeighted(vis_overlay, 0.7, red_overlay, 0.3, 0)

all_contours = [obj['contour'] for obj in object_data]
cv2.drawContours(vis_overlay, all_contours, -1, (0, 255, 0), 1)
save_debug("07c_overlay_contours.png", vis_overlay, pixel_size, unit_str)

# Intensity-coded visualization
vis_intensity = cv2.cvtColor(img_bf, cv2.COLOR_GRAY2BGR)

intensities_per_area = [obj['intensity_per_area_orig'] for obj in object_data]
if intensities_per_area:
    min_int, max_int = min(intensities_per_area), max(intensities_per_area)
    
    for obj in object_data:
        intensity = obj['intensity_per_area_orig']
        if max_int > min_int:
            normalized = (intensity - min_int) / (max_int - min_int)
        else:
            normalized = 0.5
        
        color = (int(255 * normalized), 0, int(255 * (1 - normalized)))
        cv2.drawContours(vis_intensity, [obj['contour']], -1, color, 2)

save_debug("07d_intensity_coded.png", vis_intensity, pixel_size, unit_str)

print("✓ Processing complete")
print(f"Enhancement settings: Normalize={RED_NORMALIZE}, Brightness={RED_BRIGHTNESS}, Gamma={RED_GAMMA}")
print(f"Total particles analyzed: {len(object_data)}")
if has_calibration and pixel_size is not None:
    print(f"\n✓ All images saved with {SCALE_BAR_LENGTH_UM} {unit_str} scale bars")
    print(f"✓ All measurements use proper physical units ({unit_str}) and original {bit_depth}-bit data")
else:
    print(f"\n✓ Images saved without scale bars (no calibration available)")
    print(f"✓ All measurements in pixel units using original {bit_depth}-bit data")