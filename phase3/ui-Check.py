from typing import Any
import re
from openpyxl.utils import get_column_letter
from pathlib import Path
import json
import threading
import os
import tkinter as tk
from tkinter import ttk, messagebox, colorchooser

import numpy as np
import tifffile as tiff
from PIL import Image, ImageTk, ImageDraw, ImageFont

from register_dataset import register_all_datasets, ImagePairRecord
from preprocess import (
    preprocess_fluo_for_seg,
    preprocess_bf_for_seg,
    segment_particles_from_fluo,
    extract_contours,
)

def safe_column_letter(cell: Any) -> str:
    """
    Return the column letter for any openpyxl cell-like object (Cell or MergedCell).
    Uses numeric .column when available, otherwise parses .coordinate.
    """
    col = getattr(cell, "column", None)
    if isinstance(col, int):
        return get_column_letter(col)

    coord = getattr(cell, "coordinate", None)
    if isinstance(coord, str):
        m = re.match(r"([A-Za-z]+)", coord)
        if m:
            return m.group(1)

    col_letter = getattr(cell, "column_letter", None)
    if isinstance(col_letter, str):
        return col_letter

    raise ValueError("Cannot determine column letter for cell: %r" % (cell,))


# ---------------------------------------------------------------------------
# Pillow resampling compatibility
# ---------------------------------------------------------------------------
try:
    RESAMPLE = Image.Resampling.LANCZOS  # Pillow >= 9.1
except AttributeError:
    resampling = (
        getattr(Image, "LANCZOS", None)
        or getattr(Image, "ANTIALIAS", None)
        or getattr(Image, "BICUBIC", None)
        or getattr(Image, "BILINEAR", None)
        or getattr(Image, "NEAREST", None)
    )
    RESAMPLE = resampling if resampling is not None else 0  # 0 ~ NEAREST

# ---------------------------------------------------------------------------
# Matplotlib for histograms
# ---------------------------------------------------------------------------
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

CONFIG_FILE = "contours_config.json"


class BFFluoViewer(tk.Tk):
    def __init__(self, source_root: Path):
        super().__init__()

        self.title("BF/FLUO Viewer with Scale Bar")
        self.geometry("1920x1080")

        # Load dataset registry
        self.records = register_all_datasets(source_root)
        if not self.records:
            messagebox.showerror("Error", "No image pairs found.")
            self.destroy()
            return

        # State
        self.current_record: ImagePairRecord | None = None
        self.show_scaled = tk.BooleanVar(value=True)
        self.bar_length_um = tk.DoubleVar(value=5.0)
        self.contrast_mode = tk.StringVar(value="seg-preproc-contours")

        # Tk image refs (four panes)
        self._bf_raw_tkimg: ImageTk.PhotoImage | None = None
        self._fluo_raw_tkimg: ImageTk.PhotoImage | None = None
        self._bf_seg_tkimg: ImageTk.PhotoImage | None = None
        self._fluo_seg_tkimg: ImageTk.PhotoImage | None = None

        # Last PIL images
        self._bf_top_pil: Image.Image | None = None
        self._fluo_top_pil: Image.Image | None = None
        self._bf_bottom_pil: Image.Image | None = None
        self._fluo_bottom_pil: Image.Image | None = None

        # Cache for contours and statistics per image pair
        self._seg_contour_cache: dict[tuple[str, str], list[np.ndarray]] = {}
        self._stats_cache: dict[tuple[str, str], tuple[list[dict], set[int]]] = {}

        # Contour config
        self.min_area_px_var = tk.IntVar(value=20)
        self.min_contour_len_var = tk.IntVar(value=10)
        self.contour_width_var = tk.IntVar(value=1)
        self.contour_color_hex = tk.StringVar(value="#ff0000")

        # Middle‑range highlighting
        self.middle_low_pct_var = tk.IntVar(value=30)
        self.middle_high_pct_var = tk.IntVar(value=70)
        self.middle_contour_color_hex = tk.StringVar(value="#ffff00")

        # Arrow configuration
        self.arrow_length_var = tk.IntVar(value=25)

        # Progress bar state
        self._refresh_thread: threading.Thread | None = None
        self._refresh_running = False

        # Matplotlib figure/axes for histograms (GUI)
        self.histo_fig: Figure | None = None
        self.ax_pdf_intensity_per_area = None
        self.ax_pdf_total_intensity = None
        self.ax_pdf_eq_diam = None
        self.histo_canvas: FigureCanvasTkAgg | None = None

        # Metadata widgets for Metadata tab
        self._meta_widgets: dict[str, ttk.Label] = {}

        # Cache of latest stats for current record (for export)
        self._current_stats_rows: list[dict] = []
        self._current_highlighted_idxs: set[int] = set()

        # Load saved contour config
        self._load_contour_config_from_file()

        # Build UI
        self._build_ui()
        self._populate_listbox()

        # Resize handling
        self.bind("<Configure>", self._on_root_configure)
        self.protocol("WM_DELETE_WINDOW", self.on_exit)

    # ------------------------------------------------------------------ Config I/O

    def _load_contour_config_from_file(self):
        cfg_path = Path(CONFIG_FILE)
        if not cfg_path.is_file():
            return
        try:
            with cfg_path.open("r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            print(f"[WARN] Could not read {CONFIG_FILE}: {e}")
            return

        try:
            if "min_area_px" in data:
                self.min_area_px_var.set(int(data["min_area_px"]))
            if "min_contour_len" in data:
                self.min_contour_len_var.set(int(data["min_contour_len"]))
            if "contour_width" in data:
                self.contour_width_var.set(int(data["contour_width"]))
            if "contour_color" in data:
                self.contour_color_hex.set(str(data["contour_color"]))
            if "middle_low_pct" in data:
                self.middle_low_pct_var.set(int(data["middle_low_pct"]))
            if "middle_high_pct" in data:
                self.middle_high_pct_var.set(int(data["middle_high_pct"]))
            if "middle_contour_color" in data:
                self.middle_contour_color_hex.set(str(data["middle_contour_color"]))
            if "arrow_length_px" in data:
                self.arrow_length_var.set(int(data["arrow_length_px"]))
        except Exception as e:
            print(f"[WARN] Invalid values in {CONFIG_FILE}: {e}")

    def _save_contour_config_to_file(self):
        data = {
            "min_area_px": int(self.min_area_px_var.get()),
            "min_contour_len": int(self.min_contour_len_var.get()),
            "contour_width": int(self.contour_width_var.get()),
            "contour_color": self.contour_color_hex.get(),
            "middle_low_pct": int(self.middle_low_pct_var.get()),
            "middle_high_pct": int(self.middle_high_pct_var.get()),
            "middle_contour_color": self.middle_contour_color_hex.get(),
            "arrow_length_px": int(self.arrow_length_var.get()),
        }
        try:
            with Path(CONFIG_FILE).open("w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            messagebox.showinfo("Contours config", f"Saved to {CONFIG_FILE}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save {CONFIG_FILE}:\n{e}")

    # ------------------------------------------------------------------ UI

    def _build_ui(self):
        root_pane = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        root_pane.pack(fill=tk.BOTH, expand=True)

        left_frame = ttk.Frame(root_pane, width=380)
        root_pane.add(left_frame, weight=0)

        # ==========================================================
        # LEFT PANEL
        # ==========================================================

        # 1) Dataset selection
        dataset_frame = ttk.LabelFrame(left_frame, text="Dataset selection")
        dataset_frame.pack(fill=tk.BOTH, padx=5, pady=(5, 5), expand=False)

        ttk.Label(dataset_frame, text="Image pairs (multi‑select):").pack(
            anchor=tk.W, padx=5, pady=(3, 0)
        )

        self.listbox = tk.Listbox(
            dataset_frame,
            height=18,
            selectmode=tk.EXTENDED,  # multi-select
        )
        self.listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.listbox.bind("<<ListboxSelect>>", self.on_select_pair)

        select_btn_frame = ttk.Frame(dataset_frame)
        select_btn_frame.pack(fill=tk.X, padx=5, pady=(0, 5))

        ttk.Button(
            select_btn_frame,
            text="Select all",
            command=self._select_all_pairs,
        ).pack(side=tk.LEFT, padx=(0, 5))

        ttk.Button(
            select_btn_frame,
            text="Clear selection",
            command=self._clear_selection,
        ).pack(side=tk.LEFT, padx=(0, 5))

        # 2) Display options
        display_frame = ttk.LabelFrame(left_frame, text="Display options")
        display_frame.pack(fill=tk.X, padx=5, pady=(0, 5))

        ttk.Checkbutton(
            display_frame,
            text="Show scale bar",
            variable=self.show_scaled,
            command=self.update_images,
        ).pack(anchor=tk.W, padx=5, pady=(5, 0))

        bar_frame = ttk.Frame(display_frame)
        bar_frame.pack(anchor=tk.W, pady=2, padx=5)
        ttk.Label(bar_frame, text="Bar length (µm):").pack(side=tk.LEFT)
        ttk.Entry(bar_frame, width=6, textvariable=self.bar_length_um).pack(
            side=tk.LEFT
        )

        ttk.Label(display_frame, text="Mode:").pack(anchor=tk.W, pady=(6, 0), padx=5)
        rb_frame = ttk.Frame(display_frame)
        rb_frame.pack(anchor=tk.W, pady=2, padx=5)

        ttk.Radiobutton(
            rb_frame,
            text="Raw only",
            value="auto",
            variable=self.contrast_mode,
            command=self.update_images,
        ).pack(anchor=tk.W)

        ttk.Radiobutton(
            rb_frame,
            text="Segmentation preproc + contours",
            value="seg-preproc-contours",
            variable=self.contrast_mode,
            command=self.update_images,
        ).pack(anchor=tk.W)

        # 3) Contours configuration
        contour_frame = ttk.LabelFrame(left_frame, text="Contours configuration")
        contour_frame.pack(fill=tk.X, padx=5, pady=(0, 5))

        row1 = ttk.Frame(contour_frame)
        row1.pack(fill=tk.X, pady=2, padx=5)
        ttk.Label(row1, text="Min particle area (px):").pack(side=tk.LEFT)
        ttk.Spinbox(
            row1,
            from_=1,
            to=10000,
            width=6,
            textvariable=self.min_area_px_var,
            command=self._on_contour_config_changed,
        ).pack(side=tk.LEFT, padx=4)

        row2 = ttk.Frame(contour_frame)
        row2.pack(fill=tk.X, pady=2, padx=5)
        ttk.Label(row2, text="Min contour length:").pack(side=tk.LEFT)
        ttk.Spinbox(
            row2,
            from_=1,
            to=10000,
            width=6,
            textvariable=self.min_contour_len_var,
            command=self._on_contour_config_changed,
        ).pack(side=tk.LEFT, padx=4)

        row3 = ttk.Frame(contour_frame)
        row3.pack(fill=tk.X, pady=2, padx=5)
        ttk.Label(row3, text="Line width (px):").pack(side=tk.LEFT)
        ttk.Spinbox(
            row3,
            from_=1,
            to=20,
            width=4,
            textvariable=self.contour_width_var,
            command=self._on_contour_config_changed,
        ).pack(side=tk.LEFT, padx=4)

        row4 = ttk.Frame(contour_frame)
        row4.pack(fill=tk.X, pady=2, padx=5)
        ttk.Label(row4, text="Main color:").pack(side=tk.LEFT)

        self.color_preview = tk.Label(
            row4, width=3, relief=tk.SUNKEN, bg=self.contour_color_hex.get()
        )
        self.color_preview.pack(side=tk.LEFT, padx=4)

        ttk.Button(
            row4,
            text="Pick...",
            command=self._choose_contour_color,
        ).pack(side=tk.LEFT)

        row5 = ttk.Frame(contour_frame)
        row5.pack(fill=tk.X, pady=2, padx=5)
        ttk.Label(
            row5,
            text="Middle range from (% of rank):",
        ).pack(side=tk.LEFT)
        ttk.Spinbox(
            row5,
            from_=0,
            to=100,
            width=4,
            textvariable=self.middle_low_pct_var,
            command=self._on_middle_range_changed,
        ).pack(side=tk.LEFT, padx=4)

        row6 = ttk.Frame(contour_frame)
        row6.pack(fill=tk.X, pady=2, padx=5)
        ttk.Label(
            row6,
            text="Middle range to   (% of rank):",
        ).pack(side=tk.LEFT)
        ttk.Spinbox(
            row6,
            from_=0,
            to=100,
            width=4,
            textvariable=self.middle_high_pct_var,
            command=self._on_middle_range_changed,
        ).pack(side=tk.LEFT, padx=4)

        row7 = ttk.Frame(contour_frame)
        row7.pack(fill=tk.X, pady=2, padx=5)
        ttk.Label(
            row7,
            text="Middle-range color:",
        ).pack(side=tk.LEFT)

        self.middle_color_preview = tk.Label(
            row7, width=3, relief=tk.SUNKEN, bg=self.middle_contour_color_hex.get()
        )
        self.middle_color_preview.pack(side=tk.LEFT, padx=4)

        ttk.Button(
            row7,
            text="Pick...",
            command=self._choose_middle_contour_color,
        ).pack(side=tk.LEFT)

        row8 = ttk.Frame(contour_frame)
        row8.pack(fill=tk.X, pady=2, padx=5)
        ttk.Label(row8, text="Arrow length (px):").pack(side=tk.LEFT)
        ttk.Spinbox(
            row8,
            from_=5,
            to=200,
            width=4,
            textvariable=self.arrow_length_var,
            command=self._on_contour_config_changed,
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            contour_frame,
            text="Save contours config",
            command=self._save_contour_config_to_file,
        ).pack(anchor=tk.W, pady=(6, 2), padx=5)

        # 4) Export
        export_frame = ttk.LabelFrame(left_frame, text="Export")
        export_frame.pack(fill=tk.X, padx=5, pady=(0, 5))

        ttk.Button(
            export_frame,
            text="Export selected to Excel",
            command=self.export_selected_to_excel,
        ).pack(anchor=tk.W, padx=5, pady=5)

        # 5) Progress bar
        progress_frame = ttk.Frame(left_frame)
        progress_frame.pack(fill=tk.X, padx=5, pady=(0, 5))

        self.progress_label = ttk.Label(progress_frame, text="", foreground="gray")
        self.progress_label.pack(anchor=tk.W)

        self.progress_bar = ttk.Progressbar(
            progress_frame,
            mode="indeterminate",
            length=200,
        )
        self.progress_bar.pack(fill=tk.X, pady=(2, 0))
        self.progress_bar.stop()
        self.progress_bar.pack_forget()

        # 6) Exit
        ttk.Button(
            left_frame,
            text="Exit",
            command=self.on_exit,
        ).pack(anchor=tk.W, padx=10, pady=8)

        # ==========================================================
        # RIGHT NOTEBOOK
        # ==========================================================

        right_notebook = ttk.Notebook(root_pane)
        root_pane.add(right_notebook, weight=1)
        self._right_notebook = right_notebook

        # --- Tab 1: Images ---
        images_tab = ttk.Frame(right_notebook)
        right_notebook.add(images_tab, text="Images")

        images_tab.rowconfigure(0, weight=1)
        images_tab.rowconfigure(1, weight=1)
        images_tab.columnconfigure(0, weight=1)
        images_tab.columnconfigure(1, weight=1)

        bf_raw_frame = ttk.Frame(images_tab)
        bf_raw_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        bf_raw_frame.rowconfigure(1, weight=1)
        bf_raw_frame.columnconfigure(0, weight=1)

        self.bf_label = ttk.Label(bf_raw_frame, text="Brightfield (raw)")
        self.bf_label.grid(row=0, column=0, pady=(5, 0))

        self.bf_raw_canvas = tk.Label(bf_raw_frame, bg="white")
        self.bf_raw_canvas.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)

        fluo_raw_frame = ttk.Frame(images_tab)
        fluo_raw_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        fluo_raw_frame.rowconfigure(1, weight=1)
        fluo_raw_frame.columnconfigure(0, weight=1)

        self.fluo_label = ttk.Label(fluo_raw_frame, text="Fluorescence (raw)")
        self.fluo_label.grid(row=0, column=0, pady=(5, 0))

        self.fluo_raw_canvas = tk.Label(fluo_raw_frame, bg="white")
        self.fluo_raw_canvas.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)

        bf_seg_frame = ttk.Frame(images_tab)
        bf_seg_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        bf_seg_frame.rowconfigure(1, weight=1)
        bf_seg_frame.columnconfigure(0, weight=1)

        self.bf_seg_label = ttk.Label(
            bf_seg_frame, text="Brightfield (enhanced + contours)"
        )
        self.bf_seg_label.grid(row=0, column=0, pady=(5, 0))

        self.bf_seg_canvas = tk.Label(bf_seg_frame, bg="white")
        self.bf_seg_canvas.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)

        fluo_seg_frame = ttk.Frame(images_tab)
        fluo_seg_frame.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
        fluo_seg_frame.rowconfigure(1, weight=1)
        fluo_seg_frame.columnconfigure(0, weight=1)

        self.fluo_seg_label = ttk.Label(
            fluo_seg_frame, text="Fluorescence (enhanced + contours)"
        )
        self.fluo_seg_label.grid(row=0, column=0, pady=(5, 0))

        self.fluo_seg_canvas = tk.Label(fluo_seg_frame, bg="white")
        self.fluo_seg_canvas.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)

        # --- Tab 2: Statistics ---
        stats_tab = ttk.Frame(right_notebook)
        right_notebook.add(stats_tab, text="Statistics")

        stats_tab.rowconfigure(0, weight=1)
        stats_tab.rowconfigure(1, weight=1)
        stats_tab.columnconfigure(0, weight=1)

        table_frame = ttk.Frame(stats_tab)
        table_frame.grid(row=0, column=0, sticky="nsew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        columns = (
            "idx",
            "area_px",
            "eq_diam_px",
            "total_intensity",
            "intensity_per_area",
        )
        self.stats_tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            selectmode="browse",
        )
        self.stats_tree.grid(row=0, column=0, sticky="nsew")

        vsb = ttk.Scrollbar(
            table_frame, orient="vertical", command=self.stats_tree.yview
        )
        vsb.grid(row=0, column=1, sticky="ns")
        self.stats_tree.configure(yscrollcommand=vsb.set)

        self.stats_tree.heading("idx", text="Rank (by Intensity / Area)")
        self.stats_tree.heading("area_px", text="Area [pixels]")
        self.stats_tree.heading("eq_diam_px", text="Eq. diameter [pixels]")
        self.stats_tree.heading(
            "total_intensity", text="Total fluo intensity [a.u.]"
        )
        self.stats_tree.heading(
            "intensity_per_area", text="Fluo intensity / area [a.u./pixel]"
        )

        self.stats_tree.tag_configure("highlight_middle", background="#ffffcc")

        histo_frame = ttk.LabelFrame(stats_tab, text="Distributions (Counts)")
        histo_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        histo_frame.rowconfigure(0, weight=1)
        histo_frame.columnconfigure(0, weight=1)

        self.histo_fig = Figure(figsize=(6, 2.4), dpi=100)
        self.ax_pdf_intensity_per_area = self.histo_fig.add_subplot(1, 3, 1)
        self.ax_pdf_total_intensity = self.histo_fig.add_subplot(1, 3, 2)
        self.ax_pdf_eq_diam = self.histo_fig.add_subplot(1, 3, 3)

        self.histo_canvas = FigureCanvasTkAgg(self.histo_fig, master=histo_frame)
        self.histo_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # --- Tab 3: Metadata ---
        meta_tab = ttk.Frame(right_notebook)
        right_notebook.add(meta_tab, text="Metadata")
        self._build_metadata_tab(meta_tab)

    # ------------------------------------------------------------------ Listbox helpers

    def _populate_listbox(self):
        self.listbox.delete(0, tk.END)
        for idx, rec in enumerate(self.records):
            label = f"{rec.dataset} :: {rec.pair_name}"
            self.listbox.insert(tk.END, label)

            if getattr(rec, "pixel_size_um", None) is None:
                self.listbox.itemconfig(idx, foreground="red")
            else:
                self.listbox.itemconfig(idx, foreground="black")

        self.current_record = None

    def _select_all_pairs(self):
        self.listbox.select_set(0, tk.END)
        self.on_select_pair()

    def _clear_selection(self):
        self.listbox.selection_clear(0, tk.END)
        self.current_record = None

    # ------------------------------------------------------------------ Metadata tab

    def _build_metadata_tab(self, parent: ttk.Frame):
        canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)

        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def add_section(title: str):
            lbl = ttk.Label(scroll_frame, text=title, font=("TkDefaultFont", 10, "bold"))
            lbl.pack(anchor="w", padx=8, pady=(8, 2))

        def add_row(key: str, label_text: str):
            row = ttk.Frame(scroll_frame)
            row.pack(fill="x", padx=16, pady=1)
            ttk.Label(row, text=label_text + ":").pack(side="left", anchor="w")
            val_lbl = ttk.Label(row, text="No metadata available", foreground="blue")
            val_lbl.pack(side="left", anchor="w", padx=(4, 0))
            self._meta_widgets[key] = val_lbl

        add_section("Image Dimensions & Resolution")
        add_row("img_size", "Image size (px)")
        add_row("phys_size", "Physical size (µm)")
        add_row("pixel_size", "Pixel size (nm/pixel)")
        add_row("bit_depth", "Bit depth")
        add_row("n_channels", "Number of channels")

        add_section("Optical Configuration")
        add_row("objective", "Objective")
        add_row("na", "Numerical Aperture (NA)")
        add_row("immersion", "Immersion")
        add_row("microscope", "Microscope")
        add_row("magnification", "Total magnification")

        add_section("Channel 1 (Brightfield - \"New 1\")")
        add_row("ch1_contrast", "Contrast method")
        add_row("ch1_exposure", "Exposure (ms)")
        add_row("ch1_tl_intensity", "Transmitted light intensity")
        add_row("ch1_lut", "LUT")
        add_row("ch1_filter", "Filter cube")

        add_section("Channel 2 (Fluorescence - \"New 2\")")
        add_row("ch2_contrast", "Contrast method")
        add_row("ch2_exposure", "Exposure (ms)")
        add_row("ch2_filter", "Filter cube")
        add_row("ch2_emission", "Emission wavelength (nm)")
        add_row("ch2_lut", "LUT")
        add_row("ch2_led", "LED wavelength / intensity")
        add_row("ch2_diaphragm", "Field/Aperture diaphragm")

        add_section("Stage & Focus Information")
        add_row("stage_pos", "Stage position (X, Y) mm")
        add_row("z_pos", "Z position (mm)")
        add_row("z_mode", "Z-mode")
        add_row("autofocus", "Autofocus")

        add_section("Camera Settings")
        add_row("camera", "Camera")
        add_row("gain", "Gain mode")
        add_row("temp", "Temperature (°C)")
        add_row("roi", "ROI")

        add_section("Display / Viewing Settings")
        add_row("ch1_scaling", "Channel 1 scaling")
        add_row("ch2_scaling", "Channel 2 scaling")

    def _update_metadata_tab(self, rec: ImagePairRecord):
        # Image size
        if hasattr(rec, "img_width") and hasattr(rec, "img_height"):
            w = getattr(rec, "img_width", None)
            h = getattr(rec, "img_height", None)
            if w is not None and h is not None:
                self._meta_widgets["img_size"].config(text=f"{w} × {h}")
            else:
                self._meta_widgets["img_size"].config(
                    text="No metadata available (typical: 1000–2000 px per side)"
                )
        else:
            self._meta_widgets["img_size"].config(
                text="No metadata available (typical: 1000–2000 px per side)"
            )

        # Physical size
        field_x = getattr(rec, "field_length_um_x", None)
        field_y = getattr(rec, "field_length_um_y", None)
        if field_x is not None and field_y is not None:
            self._meta_widgets["phys_size"].config(
                text=f"{field_x:.2f} × {field_y:.2f}"
            )
        else:
            self._meta_widgets["phys_size"].config(
                text="No metadata available (example DMI8 field: 131–150 µm)"
            )

        # Pixel size
        px_um = getattr(rec, "pixel_size_um", None)
        if px_um is not None:
            px_nm = px_um * 1000.0
            self._meta_widgets["pixel_size"].config(text=f"{px_nm:.1f}")
        else:
            self._meta_widgets["pixel_size"].config(
                text="No metadata available (typical 60–120 nm/pixel at 100x)"
            )

        # Bit depth
        bd = getattr(rec, "bit_depth", None)
        if bd is not None:
            max_val = 2**bd - 1
            self._meta_widgets["bit_depth"].config(
                text=f"{bd}-bit (0–{max_val})"
            )
        else:
            self._meta_widgets["bit_depth"].config(
                text="No metadata available (typical: 8–16-bit; example: 12-bit)"
            )

        # Number of channels
        self._meta_widgets["n_channels"].config(
            text="No metadata available (this project typically uses 2 channels: BF & FLUO)"
        )

        # Optical configuration
        objective_name = getattr(rec, "objective_name", None)
        self._meta_widgets["objective"].config(
            text=objective_name if objective_name else
            "No metadata available (typical: N PLAN 100x/1.25 Oil)"
        )

        na_val = getattr(rec, "numerical_aperture", None)
        if na_val is not None:
            self._meta_widgets["na"].config(text=f"{na_val:.2f}")
        else:
            self._meta_widgets["na"].config(
                text="No metadata available (typical NA for 100x oil: 1.25–1.4)"
            )

        self._meta_widgets["immersion"].config(
            text="Oil (n ≈ 1.515–1.520) – assumed; no specific metadata"
        )

        self._meta_widgets["microscope"].config(
            text="Leica DMI8 (inverted) – assumed for this project"
        )

        self._meta_widgets["magnification"].config(
            text="100x – assumed for this project"
        )

        # Channel 1
        self._meta_widgets["ch1_contrast"].config(
            text="TL-BF (Transmitted Light BF) – assumed"
        )
        exp_bf_s = getattr(rec, "exposure_bf_s", None)
        if exp_bf_s is not None:
            exp_bf_ms = exp_bf_s * 1000.0
            self._meta_widgets["ch1_exposure"].config(text=f"{exp_bf_ms:.0f}")
        else:
            self._meta_widgets["ch1_exposure"].config(
                text="No metadata available (example: 138 ms)"
            )
        self._meta_widgets["ch1_tl_intensity"].config(
            text="No metadata available (example setting: 112)"
        )
        self._meta_widgets["ch1_lut"].config(
            text="Gray – typical for BF (no explicit metadata)"
        )
        self._meta_widgets["ch1_filter"].config(
            text="EMP_BF – typical brightfield cube (no explicit metadata)"
        )

        # Channel 2
        self._meta_widgets["ch2_contrast"].config(text="FLUO – assumed")
        exp_fl_s = getattr(rec, "exposure_fluo_s", None)
        if exp_fl_s is not None:
            exp_fl_ms = exp_fl_s * 1000.0
            self._meta_widgets["ch2_exposure"].config(text=f"{exp_fl_ms:.0f}")
        else:
            self._meta_widgets["ch2_exposure"].config(
                text="No metadata available (example: 138 ms)"
            )
        self._meta_widgets["ch2_filter"].config(
            text="DFT51010 – typical for this project (no explicit metadata)"
        )
        self._meta_widgets["ch2_emission"].config(
            text="594 nm – example from Leica LAS X metadata"
        )
        self._meta_widgets["ch2_lut"].config(
            text="Red – typical fluorescence LUT (no explicit metadata)"
        )
        self._meta_widgets["ch2_led"].config(
            text="555 nm LED – typical range 530–560 nm; example: 555 nm at 84% power"
        )
        self._meta_widgets["ch2_diaphragm"].config(
            text="No metadata available (field/aperture diaphragm not recorded)"
        )

        # Stage & Focus
        self._meta_widgets["stage_pos"].config(
            text="No metadata available (example: X = 61.62 mm, Y = 39.78 mm)"
        )
        self._meta_widgets["z_pos"].config(
            text="No metadata available (example: 2.539 mm)"
        )
        self._meta_widgets["z_mode"].config(
            text="No metadata available (example mode: z-wide)"
        )
        self._meta_widgets["autofocus"].config(
            text="No metadata available (example: Combined HSAF, channel 1, precision 2)"
        )

        # Camera
        self._meta_widgets["camera"].config(
            text="No metadata available (typical: Photometric Prime 95B)"
        )
        self._meta_widgets["gain"].config(
            text="No metadata available (example: Full well mode)"
        )
        self._meta_widgets["temp"].config(
            text="No metadata available (example: −18 °C, target −30 °C)"
        )
        self._meta_widgets["roi"].config(
            text="No metadata available (example: full frame, no cropping)"
        )

        # Display / viewing: channel scaling
        channel_scaling = getattr(rec, "channel_scaling", None)
        if channel_scaling and len(channel_scaling) >= 2:
            ch1 = channel_scaling[0]
            ch2 = channel_scaling[1]
            self._meta_widgets["ch1_scaling"].config(
                text=f"Black = {ch1.get('black_norm', 0):.2%}, "
                     f"White = {ch1.get('white_norm', 1):.2%}"
            )
            self._meta_widgets["ch2_scaling"].config(
                text=f"Black = {ch2.get('black_norm', 0):.2%}, "
                     f"White = {ch2.get('white_norm', 1):.2%}"
            )
        else:
            self._meta_widgets["ch1_scaling"].config(
                text="No metadata available (example: Black 4.4%, White 18.8%)"
            )
            self._meta_widgets["ch2_scaling"].config(
                text="No metadata available (example: Black 4.4%, White 7.3%)"
            )

    # ------------------------------------------------------------------ Progress helpers

    def _start_progress(self, text: str = "Working…"):
        if self._refresh_running:
            return
        self._refresh_running = True
        self.progress_label.config(text=text)
        self.progress_bar.pack(fill=tk.X, pady=(2, 0))
        self.progress_bar.start(50)

    def _stop_progress(self):
        self._refresh_running = False
        self.progress_bar.stop()
        self.progress_bar.pack_forget()
        self.progress_label.config(text="")

    # ------------------------------------------------------------------ Contour config callbacks

    def _on_contour_config_changed(self):
        self._seg_contour_cache.clear()
        self._stats_cache.clear()
        self.update_images()

    def _on_middle_range_changed(self):
        try:
            low = int(self.middle_low_pct_var.get())
            high = int(self.middle_high_pct_var.get())
        except Exception:
            low, high = 30, 70

        low = max(0, min(low, 100))
        high = max(0, min(high, 100))

        if low >= high:
            if low == 100:
                low = 90
                high = 100
            else:
                low = max(0, high - 10)

        self.middle_low_pct_var.set(low)
        self.middle_high_pct_var.set(high)
        self._on_contour_config_changed()

    def _choose_contour_color(self):
        initial = self.contour_color_hex.get()
        color = colorchooser.askcolor(color=initial, title="Choose contour color")
        if color[1] is not None:
            self.contour_color_hex.set(color[1])
            self.color_preview.configure(bg=color[1])
            self._on_contour_config_changed()

    def _choose_middle_contour_color(self):
        initial = self.middle_contour_color_hex.get()
        color = colorchooser.askcolor(
            color=initial, title="Choose middle-range contour color"
        )
        if color[1] is not None:
            self.middle_contour_color_hex.set(color[1])
            self.middle_color_preview.configure(bg=color[1])
            self._on_contour_config_changed()

    # ------------------------------------------------------------------ Callbacks

    def _rec_key(self, rec: ImagePairRecord) -> tuple[str, str]:
        return (str(rec.bf_path), str(rec.fluo_path))

    def on_select_pair(self, event=None):
        sel = self.listbox.curselection()
        if not sel:
            self.current_record = None
            return
        idx = sel[0]
        self.current_record = self.records[idx]
        self.update_images()
        if self.current_record is not None:
            self._update_metadata_tab(self.current_record)

    def _on_root_configure(self, event=None):
        if (
            self._bf_top_pil is None
            or self._fluo_top_pil is None
            or self._bf_bottom_pil is None
            or self._fluo_bottom_pil is None
        ):
            return
        self._fit_and_show_all_panes()

    # ------------------------------------------------------------------ Middle-range computation

    def _compute_highlighted_idxs(self, stats_rows: list[dict]) -> set[int]:
        if not stats_rows:
            return set()

        low_pct = int(self.middle_low_pct_var.get())
        high_pct = int(self.middle_high_pct_var.get())

        low_pct = max(0, min(low_pct, 100))
        high_pct = max(0, min(high_pct, 100))
        if high_pct < low_pct:
            low_pct, high_pct = high_pct, low_pct

        n = len(stats_rows)
        start = int(np.floor(n * (low_pct / 100.0)))
        end = int(np.ceil(n * (high_pct / 100.0))) - 1
        start = max(0, min(start, n - 1))
        end = max(start, min(end, n - 1))

        middle_idxs = {stats_rows[i]["idx"] for i in range(start, end + 1)}
        return middle_idxs

    # ------------------------------------------------------------------ Overlay / arrow drawing

    def _overlay_contours_on_gray(
        self,
        img01: np.ndarray,
        contours: list[np.ndarray],
        color=(255, 0, 0),
        middle_color=(255, 255, 0),
        line_width: int = 1,
        label_indices: list[int] | None = None,
        highlight_mask: np.ndarray | None = None,
    ) -> Image.Image:
        if img01.dtype != np.uint8:
            base = (np.clip(img01, 0.0, 1.0) * 255.0).astype(np.uint8)
        else:
            base = img01

        pil = Image.fromarray(base).convert("RGB")
        draw = ImageDraw.Draw(pil)
        w, h = pil.size

        if highlight_mask is None:
            highlight_mask = np.zeros(len(contours), dtype=bool)
        else:
            highlight_mask = np.array(highlight_mask, dtype=bool)
            if highlight_mask.size != len(contours):
                highlight_mask = np.zeros(len(contours), dtype=bool)

        main_color = color
        middle_color_rgb = middle_color

        def _brighten(rgb, factor=1.3):
            r, g, b = rgb
            return (
                min(int(r * factor), 255),
                min(int(g * factor), 255),
                min(int(b * factor), 255),
            )

        main_color_highlighted = _brighten(main_color)
        middle_color_highlighted = _brighten(middle_color_rgb)

        font = self._get_scaled_font(base_size=12)

        centroids = []
        for c in contours:
            if c.shape[0] < 3:
                centroids.append((None, None))
                continue
            poly_x = c[:, 1].astype(float)
            poly_y = c[:, 0].astype(float)
            cx = float(poly_x.mean())
            cy = float(poly_y.mean())
            centroids.append((cx, cy))

        arrow_len = max(5, int(self.arrow_length_var.get()))

        for i, c in enumerate(contours):
            if c.shape[0] < 2:
                continue

            if highlight_mask[i]:
                contour_color = middle_color_highlighted
            else:
                contour_color = main_color_highlighted

            pts = [(float(col), float(row)) for row, col in c]
            if len(pts) > 1:
                draw.line(pts, fill=contour_color, width=line_width)

            if label_indices is not None and i < len(label_indices):
                if c.shape[0] < 3:
                    continue

                cx, cy = centroids[i]
                if cx is None:
                    continue

                left_count = 0
                right_count = 0
                up_count = 0
                down_count = 0
                for j, (ox, oy) in enumerate(centroids):
                    if j == i or ox is None:
                        continue
                    if ox < cx:
                        left_count += 1
                    elif ox > cx:
                        right_count += 1
                    if oy < cy:
                        up_count += 1
                    elif oy > cy:
                        down_count += 1

                scores: dict[str, int] = {}

                def add_dir(dir_name: str, base_score: int, tx: float, ty: float) -> None:
                    penalty = 0
                    if tx < 10 or tx > w - 10 or ty < 10 or ty > h - 10:
                        penalty = 1000
                    scores[dir_name] = int(base_score + penalty)

                add_dir("up", up_count, cx, cy - arrow_len)
                add_dir("down", down_count, cx, cy + arrow_len)
                add_dir("left", left_count, cx - arrow_len, cy)
                add_dir("right", right_count, cx + arrow_len, cy)

                best_dir = min(scores.keys(), key=lambda k: scores[k])

                if best_dir == "up":
                    start = (cx, cy)
                    end = (cx, max(cy - arrow_len, 0))
                    label_pos = (cx, max(end[1] - 5, 0))
                elif best_dir == "down":
                    start = (cx, cy)
                    end = (cx, min(cy + arrow_len, h - 1))
                    label_pos = (cx, min(end[1] + 10, h - 1))
                elif best_dir == "left":
                    start = (cx, cy)
                    end = (max(cx - arrow_len, 0), cy)
                    label_pos = (max(end[0] - 5, 0), cy)
                else:  # "right"
                    start = (cx, cy)
                    end = (min(cx + arrow_len, w - 1), cy)
                    label_pos = (min(end[0] + 5, w - 1), cy)

                draw.line([start, end], fill=contour_color, width=1)

                text = str(label_indices[i])
                draw.text(label_pos, text, fill=contour_color, font=font, anchor="ms")

        return pil

    # ------------------------------------------------------------------ Statistics computation

    def _measure_contours(
        self,
        contours: list[np.ndarray],
        bf_enh: np.ndarray,
        fluo_enh: np.ndarray,
    ) -> list[dict]:
        """Measure features per contour and rank by intensity/area."""

        if bf_enh.ndim != 2:
            bf_enh_gray = bf_enh.mean(axis=-1)
        else:
            bf_enh_gray = bf_enh

        if fluo_enh.ndim != 2:
            fluo_enh_gray = fluo_enh.mean(axis=-1)
        else:
            fluo_enh_gray = fluo_enh

        h, w = bf_enh_gray.shape
        yy, xx = np.mgrid[0:h, 0:w]

        rows = []
        tmp_rows = []

        for local_idx, c in enumerate(contours):
            if c.shape[0] < 3:
                continue

            poly_x = c[:, 1].astype(float)
            poly_y = c[:, 0].astype(float)

            min_x = max(int(np.floor(poly_x.min())), 0)
            max_x = min(int(np.ceil(poly_x.max())), w - 1)
            min_y = max(int(np.floor(poly_y.min())), 0)
            max_y = min(int(np.ceil(poly_y.max())), h - 1)

            if min_x >= max_x or min_y >= max_y:
                continue

            bx = xx[min_y:max_y + 1, min_x:max_x + 1]
            by = yy[min_y:max_y + 1, min_x:max_x + 1]

            inside = np.zeros_like(bx, dtype=bool)
            x = bx.astype(float)
            y = by.astype(float)

            x0 = poly_x
            y0 = poly_y
            x1 = np.roll(poly_x, -1)
            y1 = np.roll(poly_y, -1)

            for xj0, yj0, xj1, yj1 in zip(x0, y0, x1, y1):
                cond = ((yj0 <= y) & (yj1 > y)) | ((yj1 <= y) & (yj0 > y))
                if not np.any(cond):
                    continue
                x_int = xj0 + (y - yj0) * (xj1 - xj0) / (yj1 - yj0 + 1e-12)
                inside ^= cond & (x < x_int)

            area = float(inside.sum())
            if area <= 0:
                continue

            eq_diam = np.sqrt(4.0 * area / np.pi)

            fluo_patch = fluo_enh_gray[min_y:max_y + 1, min_x:max_x + 1]
            total_intensity = float(fluo_patch[inside].sum()) if inside.any() else 0.0

            intensity_per_area = total_intensity / area if area > 0 else 0.0

            tmp_rows.append(
                {
                    "contour_local_idx": local_idx,
                    "area_px": area,
                    "eq_diam_px": eq_diam,
                    "total_intensity": total_intensity,
                    "intensity_per_area": intensity_per_area,
                }
            )

        tmp_rows.sort(key=lambda r: r["intensity_per_area"], reverse=True)
        for rank, r in enumerate(tmp_rows, start=1):
            rows.append(
                {
                    "idx": rank,
                    "contour_local_idx": r["contour_local_idx"],
                    "area_px": r["area_px"],
                    "eq_diam_px": r["eq_diam_px"],
                    "total_intensity": r["total_intensity"],
                    "intensity_per_area": r["intensity_per_area"],
                }
            )

        return rows

    def _update_histograms(self, stats_rows: list[dict]):
        if (
            self.histo_fig is None
            or self.ax_pdf_intensity_per_area is None
            or self.ax_pdf_total_intensity is None
            or self.ax_pdf_eq_diam is None
            or self.histo_canvas is None
        ):
            return

        self.ax_pdf_intensity_per_area.clear()
        self.ax_pdf_total_intensity.clear()
        self.ax_pdf_eq_diam.clear()

        if not stats_rows:
            self.ax_pdf_intensity_per_area.set_title("Fluo / Area")
            self.ax_pdf_total_intensity.set_title("Total Fluo")
            self.ax_pdf_eq_diam.set_title("Eq. diameter (px)")
            self.histo_canvas.draw_idle()
            return

        intensity_per_area = np.array(
            [r["intensity_per_area"] for r in stats_rows], dtype=float
        )
        total_intensity = np.array(
            [r["total_intensity"] for r in stats_rows], dtype=float
        )
        eq_diam_px = np.array([r["eq_diam_px"] for r in stats_rows], dtype=float)

        mean_diam = float(eq_diam_px.mean()) if eq_diam_px.size > 0 else float("nan")
        std_diam = float(eq_diam_px.std(ddof=1)) if eq_diam_px.size > 1 else float("nan")

        self.ax_pdf_intensity_per_area.hist(
            intensity_per_area,
            bins=30,
            density=False,
            color="tab:blue",
            alpha=0.7,
        )
        self.ax_pdf_intensity_per_area.set_title("Fluo / Area")
        self.ax_pdf_intensity_per_area.set_xlabel("Intensity / Area")
        self.ax_pdf_intensity_per_area.set_ylabel("Count")

        self.ax_pdf_total_intensity.hist(
            total_intensity,
            bins=30,
            density=False,
            color="tab:green",
            alpha=0.7,
        )
        self.ax_pdf_total_intensity.set_title("Total Fluo")
        self.ax_pdf_total_intensity.set_xlabel("Total intensity")
        self.ax_pdf_total_intensity.set_ylabel("Count")

        self.ax_pdf_eq_diam.hist(
            eq_diam_px,
            bins=30,
            density=False,
            color="tab:purple",
            alpha=0.7,
        )

        if np.isfinite(mean_diam):
            self.ax_pdf_eq_diam.axvline(
                mean_diam,
                color="red",
                linestyle="--",
                linewidth=1.5,
                label=f"Mean = {mean_diam:.2f} px",
            )

        if np.isfinite(mean_diam) and np.isfinite(std_diam):
            txt = f"μ = {mean_diam:.2f} px\nσ = {std_diam:.2f} px"
            self.ax_pdf_eq_diam.text(
                0.98,
                0.95,
                txt,
                transform=self.ax_pdf_eq_diam.transAxes,
                ha="right",
                va="top",
                fontsize=8,
                bbox=dict(facecolor="white", alpha=0.7, edgecolor="none"),
            )

        if np.isfinite(mean_diam):
            self.ax_pdf_eq_diam.legend(fontsize=8)

        self.ax_pdf_eq_diam.set_title("Eq. diameter (px)")
        self.ax_pdf_eq_diam.set_xlabel("Eq. diameter (px)")
        self.ax_pdf_eq_diam.set_ylabel("Count")

        self.histo_fig.tight_layout()
        self.histo_canvas.draw_idle()

    def _update_stats_view(
        self,
        rec: ImagePairRecord,
        stats_rows: list[dict],
        highlighted_idx_set: set[int] | None = None,
    ):
        if highlighted_idx_set is None:
            highlighted_idx_set = set()

        self._current_stats_rows = stats_rows[:]
        self._current_highlighted_idxs = set(highlighted_idx_set)

        for iid in self.stats_tree.get_children():
            self.stats_tree.delete(iid)

        for r in stats_rows:
            tags = ()
            if r["idx"] in highlighted_idx_set:
                tags = ("highlight_middle",)

            self.stats_tree.insert(
                "",
                tk.END,
                values=(
                    r["idx"],
                    f"{r['area_px']:.0f}",
                    f"{r['eq_diam_px']:.2f}",
                    f"{r['total_intensity']:.2f}",
                    f"{r['intensity_per_area']:.4f}",
                ),
                tags=tags,
            )

        self._update_histograms(stats_rows)

    # ------------------------------------------------------------------ Histogram figure for export

    def _create_histogram_figure_for_export(self, stats_rows: list[dict]) -> Figure:
        """Create a matplotlib Figure with the 3 histograms for saving into Excel."""
        fig = Figure(figsize=(6, 2.4), dpi=150)
        ax1 = fig.add_subplot(1, 3, 1)
        ax2 = fig.add_subplot(1, 3, 2)
        ax3 = fig.add_subplot(1, 3, 3)

        if not stats_rows:
            ax1.set_title("Fluo / Area")
            ax2.set_title("Total Fluo")
            ax3.set_title("Eq. diameter (px)")
            fig.tight_layout()
            return fig

        intensity_per_area = np.array(
            [r["intensity_per_area"] for r in stats_rows], dtype=float
        )
        total_intensity = np.array(
            [r["total_intensity"] for r in stats_rows], dtype=float
        )
        eq_diam_px = np.array([r["eq_diam_px"] for r in stats_rows], dtype=float)

        mean_diam = float(eq_diam_px.mean()) if eq_diam_px.size > 0 else float("nan")
        std_diam = float(eq_diam_px.std(ddof=1)) if eq_diam_px.size > 1 else float("nan")

        ax1.hist(intensity_per_area, bins=30, color="tab:blue", alpha=0.7)
        ax1.set_title("Fluo / Area")
        ax1.set_xlabel("Intensity / Area")
        ax1.set_ylabel("Count")

        ax2.hist(total_intensity, bins=30, color="tab:green", alpha=0.7)
        ax2.set_title("Total Fluo")
        ax2.set_xlabel("Total intensity")
        ax2.set_ylabel("Count")

        ax3.hist(eq_diam_px, bins=30, color="tab:purple", alpha=0.7)
        ax3.set_title("Eq. diameter (px)")
        ax3.set_xlabel("Eq. diameter (px)")
        ax3.set_ylabel("Count")

        if np.isfinite(mean_diam):
            ax3.axvline(mean_diam, color="red", linestyle="--", linewidth=1.5,
                        label=f"Mean = {mean_diam:.2f} px")
        if np.isfinite(mean_diam) and np.isfinite(std_diam):
            txt = f"μ = {mean_diam:.2f} px\nσ = {std_diam:.2f} px"
            ax3.text(
                0.98,
                0.95,
                txt,
                transform=ax3.transAxes,
                ha="right",
                va="top",
                fontsize=8,
                bbox=dict(facecolor="white", alpha=0.7, edgecolor="none"),
            )

        if np.isfinite(mean_diam):
            ax3.legend(fontsize=8)

        fig.tight_layout()
        return fig

    # ------------------------------------------------------------------ Main refresh

    def update_images(self):
        if self.current_record is None:
            return
        if self._refresh_running:
            return

        rec: ImagePairRecord = self.current_record
        self._start_progress("Refreshing view…")

        def worker():
            try:
                result = self._update_images_worker(rec)
            except Exception as e:
                err_msg = f"Error during refresh:\n{e}"
                self.after(
                    0,
                    lambda msg=err_msg: messagebox.showerror("Error", msg),
                )
                result = None

            def finish():
                self._stop_progress()
                if result is not None:
                    (
                        bf_top_img,
                        fluo_top_img,
                        bf_bottom_img,
                        fluo_bottom_img,
                        label_text_bf,
                        label_text_fluo,
                        stats_rows,
                        highlighted_idx_set,
                    ) = result

                    self._bf_top_pil = bf_top_img
                    self._fluo_top_pil = fluo_top_img
                    self._bf_bottom_pil = bf_bottom_img
                    self._fluo_bottom_pil = fluo_bottom_img

                    self._fit_and_show_all_panes()

                    self.bf_label.config(text=label_text_bf)
                    self.fluo_label.config(text=label_text_fluo)

                    self._update_stats_view(rec, stats_rows, highlighted_idx_set)
                    self._update_metadata_tab(rec)

            self.after(0, finish)

        self._refresh_thread = threading.Thread(target=worker, daemon=True)
        self._refresh_thread.start()

    def _fit_and_show_all_panes(self):
        """Resize the four current PIL images to fit their Tk labels."""
        if self._bf_top_pil is not None:
            tw = max(self.bf_raw_canvas.winfo_width(), 1)
            th = max(self.bf_raw_canvas.winfo_height(), 1)
            img_fit = self._resize_to_box(self._bf_top_pil, tw, th)
            self._bf_raw_tkimg = ImageTk.PhotoImage(img_fit)
            self.bf_raw_canvas.configure(image=self._bf_raw_tkimg)

        if self._fluo_top_pil is not None:
            tw = max(self.fluo_raw_canvas.winfo_width(), 1)
            th = max(self.fluo_raw_canvas.winfo_height(), 1)
            img_fit = self._resize_to_box(self._fluo_top_pil, tw, th)
            self._fluo_raw_tkimg = ImageTk.PhotoImage(img_fit)
            self.fluo_raw_canvas.configure(image=self._fluo_raw_tkimg)

        if self._bf_bottom_pil is not None:
            tw = max(self.bf_seg_canvas.winfo_width(), 1)
            th = max(self.bf_seg_canvas.winfo_height(), 1)
            img_fit = self._resize_to_box(self._bf_bottom_pil, tw, th)
            self._bf_seg_tkimg = ImageTk.PhotoImage(img_fit)
            self.bf_seg_canvas.configure(image=self._bf_seg_tkimg)

        if self._fluo_bottom_pil is not None:
            tw = max(self.fluo_seg_canvas.winfo_width(), 1)
            th = max(self.fluo_seg_canvas.winfo_height(), 1)
            img_fit = self._resize_to_box(self._fluo_bottom_pil, tw, th)
            self._fluo_seg_tkimg = ImageTk.PhotoImage(img_fit)
            self.fluo_seg_canvas.configure(image=self._fluo_seg_tkimg)

    def _update_images_worker(self, rec: ImagePairRecord):
        """Worker that prepares images and statistics for a single record."""

        bf_raw = tiff.imread(rec.bf_path)
        fluo_raw = tiff.imread(rec.fluo_path)

        mode = self.contrast_mode.get()

        bf_top_arr = bf_raw
        fluo_top_arr = fluo_raw

        bf_bottom_img: Image.Image
        fluo_bottom_img: Image.Image

        stats_rows: list[dict] = []
        highlighted_idx_set: set[int] = set()

        if mode == "seg-preproc-contours":
            fluo_proc = preprocess_fluo_for_seg(rec)

            key = self._rec_key(rec)
            if key in self._seg_contour_cache:
                contours = self._seg_contour_cache[key]
            else:
                min_area_px = int(self.min_area_px_var.get())
                min_len = int(self.min_contour_len_var.get())
                mask = segment_particles_from_fluo(
                    fluo_proc, min_area_px=min_area_px
                )
                contours = extract_contours(mask, min_length=min_len)
                self._seg_contour_cache[key] = contours

            bf_enh = preprocess_bf_for_seg(rec)
            fluo_enh = fluo_proc

            bf_enh_f = bf_enh.astype(np.float32)
            fluo_enh_f = fluo_enh.astype(np.float32)
            if bf_enh_f.max() > 0:
                bf_enh_f /= bf_enh_f.max()
            if fluo_enh_f.max() > 0:
                fluo_enh_f /= fluo_enh_f.max()

            stats_key = self._rec_key(rec)
            if stats_key in self._stats_cache:
                stats_rows, highlighted_idx_set = self._stats_cache[stats_key]
            else:
                stats_rows = self._measure_contours(
                    contours=contours,
                    bf_enh=bf_enh,
                    fluo_enh=fluo_enh,
                )

                highlighted_idx_set = self._compute_highlighted_idxs(stats_rows)
                self._stats_cache[stats_key] = (stats_rows, highlighted_idx_set)

            local_to_rank = {
                r["contour_local_idx"]: r["idx"] for r in stats_rows
            }

            highlight_mask = np.zeros(len(contours), dtype=bool)
            label_indices = [0] * len(contours)
            for local_idx, rank_idx in local_to_rank.items():
                label_indices[local_idx] = rank_idx
                if rank_idx in highlighted_idx_set:
                    highlight_mask[local_idx] = True

            hex_color = self.contour_color_hex.get()
            mid_hex_color = self.middle_contour_color_hex.get()
            try:
                rgb = tuple(int(hex_color[i: i + 2], 16) for i in (1, 3, 5))
            except Exception:
                rgb = (255, 0, 0)
            try:
                mid_rgb = tuple(int(mid_hex_color[i: i + 2], 16) for i in (1, 3, 5))
            except Exception:
                mid_rgb = (255, 255, 0)

            line_w = int(self.contour_width_var.get())

            bf_enh_pil = self._overlay_contours_on_gray(
                bf_enh_f,
                contours,
                color=rgb,
                middle_color=mid_rgb,
                line_width=line_w,
                label_indices=label_indices,
                highlight_mask=highlight_mask,
            )
            fluo_enh_pil = self._overlay_contours_on_gray(
                fluo_enh_f,
                contours,
                color=rgb,
                middle_color=mid_rgb,
                line_width=line_w,
                label_indices=label_indices,
                highlight_mask=highlight_mask,
            )

            bf_bottom_img = bf_enh_pil
            fluo_bottom_img = fluo_enh_pil

        else:
            bf_bottom_img = self._to_pil(bf_raw)
            fluo_bottom_img = self._to_pil(fluo_raw)
            stats_rows = []
            highlighted_idx_set = set()

        bf_top_img = self._to_pil(bf_top_arr)
        fluo_top_img = self._to_pil(fluo_top_arr)

        if self.show_scaled.get():
            bf_top_img = self._add_scale_bar(bf_top_img, rec)
            fluo_top_img = self._add_scale_bar(fluo_top_img, rec)
            bf_bottom_img = self._add_scale_bar(bf_bottom_img, rec)
            fluo_bottom_img = self._add_scale_bar(fluo_bottom_img, rec)

        px_um = getattr(rec, "pixel_size_um", None)
        if px_um is not None:
            px_text = f" (px: {px_um:.3f} µm/pixel)"
        else:
            px_text = " (no metadata)"

        label_text_bf = f"Brightfield (raw) - {rec.pair_name}{px_text}"
        label_text_fluo = f"Fluorescence (raw) - {rec.pair_name}{px_text}"

        return (
            bf_top_img,
            fluo_top_img,
            bf_bottom_img,
            fluo_bottom_img,
            label_text_bf,
            label_text_fluo,
            stats_rows,
            highlighted_idx_set,
        )

    def on_exit(self):
        self._bf_raw_tkimg = None
        self._fluo_raw_tkimg = None
        self._bf_seg_tkimg = None
        self._fluo_seg_tkimg = None
        self.destroy()

    # ------------------------------------------------------------------ Export logic

    def _get_downloads_folder(self) -> Path:
        home = Path.home()
        downloads = home / "Downloads"
        if downloads.is_dir():
            return downloads
        return home

    def export_selected_to_excel(self):
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showwarning(
                "Export",
                "Please select one or more image pairs to export.",
            )
            return

        try:
            import pandas as pd
        except ImportError:
            messagebox.showerror(
                "Export error",
                "pandas is required for Excel export. Please install it:\n\npip install pandas openpyxl",
            )
            return

        downloads_dir = self._get_downloads_folder()
        self._start_progress("Exporting to Excel…")

        def worker():
            errors = []
            for idx in sel:
                rec = self.records[idx]
                try:
                    self._export_single_record(rec, downloads_dir)
                except Exception as e:
                    errors.append(f"{rec.pair_name}: {e}")

            def finish():
                self._stop_progress()
                if errors:
                    messagebox.showerror(
                        "Export completed with errors",
                        "Some exports failed:\n\n" + "\n".join(errors),
                    )
                else:
                    messagebox.showinfo(
                        "Export completed",
                        f"Exported {len(sel)} file(s) to:\n{downloads_dir}",
                    )

            self.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()

    def _generate_four_panel_images_for_export(self, rec: ImagePairRecord):
        """Generate the 4 images (BF raw, Fluo raw, BF enhanced+contours, Fluo enhanced+contours)
        exactly as in the GUI, and then scale them to 75% for Excel export."""
        # Raw
        bf_raw = tiff.imread(rec.bf_path)
        fluo_raw = tiff.imread(rec.fluo_path)
        bf_top_img = self._to_pil(bf_raw)
        fluo_top_img = self._to_pil(fluo_raw)

        # Preprocess for segmentation
        fluo_proc = preprocess_fluo_for_seg(rec)
        min_area_px = int(self.min_area_px_var.get())
        min_len = int(self.min_contour_len_var.get())
        mask = segment_particles_from_fluo(fluo_proc, min_area_px=min_area_px)
        contours = extract_contours(mask, min_length=min_len)

        bf_enh = preprocess_bf_for_seg(rec)
        fluo_enh = fluo_proc

        bf_enh_f = bf_enh.astype(np.float32)
        fluo_enh_f = fluo_enh.astype(np.float32)
        if bf_enh_f.max() > 0:
            bf_enh_f /= bf_enh_f.max()
        if fluo_enh_f.max() > 0:
            fluo_enh_f /= fluo_enh_f.max()

        # Stats / ranking for labels & highlighting
        stats_rows = self._measure_contours(contours, bf_enh, fluo_enh)
        highlighted_idx_set = self._compute_highlighted_idxs(stats_rows)
        local_to_rank = {r["contour_local_idx"]: r["idx"] for r in stats_rows}

        highlight_mask = np.zeros(len(contours), dtype=bool)
        label_indices = [0] * len(contours)
        for local_idx, rank_idx in local_to_rank.items():
            label_indices[local_idx] = rank_idx
            if rank_idx in highlighted_idx_set:
                highlight_mask[local_idx] = True

        hex_color = self.contour_color_hex.get()
        mid_hex_color = self.middle_contour_color_hex.get()
        try:
            rgb = tuple(int(hex_color[i:i+2], 16) for i in (1, 3, 5))
        except Exception:
            rgb = (255, 0, 0)
        try:
            mid_rgb = tuple(int(mid_hex_color[i:i+2], 16) for i in (1, 3, 5))
        except Exception:
            mid_rgb = (255, 255, 0)

        line_w = int(self.contour_width_var.get())

        bf_bottom_img = self._overlay_contours_on_gray(
            bf_enh_f,
            contours,
            color=rgb,
            middle_color=mid_rgb,
            line_width=line_w,
            label_indices=label_indices,
            highlight_mask=highlight_mask,
        )
        fluo_bottom_img = self._overlay_contours_on_gray(
            fluo_enh_f,
            contours,
            color=rgb,
            middle_color=mid_rgb,
            line_width=line_w,
            label_indices=label_indices,
            highlight_mask=highlight_mask,
        )

        # Optional scale bars
        if self.show_scaled.get():
            bf_top_img = self._add_scale_bar(bf_top_img, rec)
            fluo_top_img = self._add_scale_bar(fluo_top_img, rec)
            bf_bottom_img = self._add_scale_bar(bf_bottom_img, rec)
            fluo_bottom_img = self._add_scale_bar(fluo_bottom_img, rec)

        # Scale all four images to 70% for export
        scale = 0.70

        def _scale(pil_img: Image.Image) -> Image.Image:
            w, h = pil_img.size
            new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
            return pil_img.resize(new_size, RESAMPLE)

        bf_top_img = _scale(bf_top_img)
        fluo_top_img = _scale(fluo_top_img)
        bf_bottom_img = _scale(bf_bottom_img)
        fluo_bottom_img = _scale(fluo_bottom_img)

        return bf_top_img, fluo_top_img, bf_bottom_img, fluo_bottom_img

    def _export_single_record(self, rec: ImagePairRecord, out_dir: Path):
        import pandas as pd
        from openpyxl.drawing.image import Image as XLImage
        from io import BytesIO
        from openpyxl.styles import Font, Alignment

        key = self._rec_key(rec)

        if key in self._stats_cache:
            stats_rows, highlighted_idx_set = self._stats_cache[key]
        else:
            fluo_proc = preprocess_fluo_for_seg(rec)
            min_area_px = int(self.min_area_px_var.get())
            min_len = int(self.min_contour_len_var.get())
            mask = segment_particles_from_fluo(
                fluo_proc, min_area_px=min_area_px
            )
            contours = extract_contours(mask, min_length=min_len)
            self._seg_contour_cache[key] = contours

            bf_enh = preprocess_bf_for_seg(rec)
            fluo_enh = fluo_proc
            stats_rows = self._measure_contours(
                contours=contours,
                bf_enh=bf_enh,
                fluo_enh=fluo_enh,
            )
            highlighted_idx_set = self._compute_highlighted_idxs(stats_rows)
            self._stats_cache[key] = (stats_rows, highlighted_idx_set)

        df_stats = pd.DataFrame(stats_rows)
        if not df_stats.empty:
            df_stats = df_stats[
                ["idx", "area_px", "eq_diam_px", "total_intensity", "intensity_per_area"]
            ]
            df_stats["is_middle_range"] = df_stats["idx"].isin(highlighted_idx_set)
        else:
            df_stats = pd.DataFrame(
                columns=[
                    "idx",
                    "area_px",
                    "eq_diam_px",
                    "total_intensity",
                    "intensity_per_area",
                    "is_middle_range",
                ]
            )

        # Human-friendly column names
        df_stats = df_stats.rename(
            columns={
                "idx": "Rank (by Intensity / Area)",
                "area_px": "Area [pixels]",
                "eq_diam_px": "Equivalent Diameter [pixels]",
                "total_intensity": "Total Fluorescence Intensity [a.u.]",
                "intensity_per_area": "Fluorescence Intensity / Area [a.u./pixel]",
                "is_middle_range": "Is Middle Range (True/False)",
            }
        )

        # Extended metadata: Leica LAS X–style structure
        img_w = getattr(rec, "img_width", None)
        img_h = getattr(rec, "img_height", None)
        field_x = getattr(rec, "field_length_um_x", None)
        field_y = getattr(rec, "field_length_um_y", None)
        px_um = getattr(rec, "pixel_size_um", None)
        bd = getattr(rec, "bit_depth", None)
        objective_name = getattr(rec, "objective_name", None)
        na_val = getattr(rec, "numerical_aperture", None)
        exposure_bf_s = getattr(rec, "exposure_bf_s", None)
        exposure_fluo_s = getattr(rec, "exposure_fluo_s", None)
        channel_scaling = getattr(rec, "channel_scaling", None)

        # Defaults from Leica LAS X metadata file (10 P 1.docx)
        default_img_size = "1200 × 1200 pixels"
        default_phys_size = "131.39 µm × 131.39 µm"
        default_px_nm = "~109.5 nm/pixel"
        default_bit_depth = "12-bit (0–4095)"
        default_n_channels = "2 (Grayscale, Red)"
        default_objective = "N PLAN 100x/1.25 Oil (Art. No. 11506158)"
        default_na = "1.25"
        default_immersion = "Oil (refractive index 1.518)"
        default_microscope = "DMI8 (inverted)"
        default_total_mag = "100x"
        default_bf_exposure = "138 ms"
        default_bf_tl_intensity = "112"
        default_fluo_exposure = "138 ms"
        default_fluo_filter = "DFT51010"
        default_fluo_emission = "594 nm"
        default_led = "555 nm at 84% intensity (active)"
        default_cam = "Photometric Prime 95B (model A21F203002)"
        default_gain_mode = "Full well mode"
        default_cam_temp = "-18 °C (target: -30 °C)"
        default_fan = "Active (level 3)"
        default_roi = "Full frame (no cropping)"
        default_ch1_scaling = "Black = 4.4%, White = 18.8%"
        default_ch2_scaling = "Black = 4.4%, White = 7.3%"

        # Construct human-readable strings
        if img_w is not None and img_h is not None:
            img_size_str = f"{img_w} × {img_h} pixels"
        else:
            img_size_str = default_img_size

        if field_x is not None and field_y is not None:
            phys_size_str = f"{field_x:.2f} µm × {field_y:.2f} µm"
        else:
            phys_size_str = default_phys_size

        if px_um is not None:
            px_nm = px_um * 1000.0
            px_size_str = f"{px_nm:.1f} nm/pixel"
        else:
            px_size_str = default_px_nm

        if bd is not None:
            bit_depth_str = f"{bd}-bit (0–{2**bd - 1})"
        else:
            bit_depth_str = default_bit_depth

        if objective_name:
            objective_str = objective_name
        else:
            objective_str = default_objective

        if na_val is not None:
            na_str = f"{na_val:.2f}"
        else:
            na_str = default_na

        if exposure_bf_s is not None:
            bf_exp_str = f"{exposure_bf_s * 1000.0:.0f} ms"
        else:
            bf_exp_str = default_bf_exposure

        if exposure_fluo_s is not None:
            fluo_exp_str = f"{exposure_fluo_s * 1000.0:.0f} ms"
        else:
            fluo_exp_str = default_fluo_exposure

        # Channel scaling from record if present
        if channel_scaling and len(channel_scaling) >= 2:
            ch1 = channel_scaling[0]
            ch2 = channel_scaling[1]
            ch1_scaling_str = (
                f"Black = {ch1.get('black_norm', 0):.2%}, "
                f"White = {ch1.get('white_norm', 1):.2%}"
            )
            ch2_scaling_str = (
                f"Black = {ch2.get('black_norm', 0):.2%}, "
                f"White = {ch2.get('white_norm', 1):.2%}"
            )
        else:
            ch1_scaling_str = default_ch1_scaling
            ch2_scaling_str = default_ch2_scaling

        meta_data = {
            # General / file info
            "Image: Dataset": str(getattr(rec, "dataset", "")),
            "Image: Pair name": str(getattr(rec, "pair_name", "")),
            "Image: Brightfield path": str(getattr(rec, "bf_path", "")),
            "Image: Fluorescence path": str(getattr(rec, "fluo_path", "")),

            # Image Dimensions & Resolution
            "Image Dimensions: Image size (pixels)": img_size_str,
            "Image Dimensions: Physical size (µm)": phys_size_str,
            "Image Dimensions: Pixel size": px_size_str,
            "Image Dimensions: Bit depth": bit_depth_str,
            "Image Dimensions: Number of channels": default_n_channels,

            # Optical Configuration
            "Optical Configuration: Objective": objective_str,
            "Optical Configuration: Numerical Aperture (NA)": na_str,
            "Optical Configuration: Immersion": default_immersion,
            "Optical Configuration: Microscope": default_microscope,
            "Optical Configuration: Total magnification": default_total_mag,

            # Channel 1 (Brightfield)
            "Channel 1 (Brightfield): Contrast method": "TL-BF (Transmitted Light Brightfield)",
            "Channel 1 (Brightfield): Exposure": bf_exp_str,
            "Channel 1 (Brightfield): Transmitted light intensity": default_bf_tl_intensity,
            "Channel 1 (Brightfield): LUT": "Gray",
            "Channel 1 (Brightfield): Filter cube": "EMP_BF",

            # Channel 2 (Fluorescence)
            "Channel 2 (Fluorescence): Contrast method": "FLUO",
            "Channel 2 (Fluorescence): Exposure": fluo_exp_str,
            "Channel 2 (Fluorescence): Filter cube": default_fluo_filter,
            "Channel 2 (Fluorescence): Emission wavelength": default_fluo_emission,
            "Channel 2 (Fluorescence): LUT": "Red",
            "Channel 2 (Fluorescence): LED": default_led,
            "Channel 2 (Fluorescence): Diaphragm": "Field/Aperture diaphragm settings available",

            # Stage & Focus Information
            "Stage & Focus: Stage position": "X = 61.62 mm, Y = 39.78 mm",
            "Stage & Focus: Z position": "2.539 mm",
            "Stage & Focus: Z-mode": "z-wide",
            "Stage & Focus: Autofocus": "Combined HSAF system, channel 1, precision level 2",

            # Camera Settings
            "Camera: Model": default_cam,
            "Camera: Gain mode": default_gain_mode,
            "Camera: Temperature": default_cam_temp,
            "Camera: Fan control": default_fan,
            "Camera: ROI": default_roi,

            # Display / Viewing Settings
            "Display: Channel 1 scaling": ch1_scaling_str,
            "Display: Channel 2 scaling": ch2_scaling_str,
        }

        df_meta = (
            pd.Series(meta_data, name="Value")
            .rename_axis("Metadata item")
            .reset_index()
        )

        # --------- IMPROVED CONFIG SHEET: writing & formatting ----------

        # Collect configuration parameters
        config_data = {
            "min_area_px": int(self.min_area_px_var.get()),
            "min_contour_len": int(self.min_contour_len_var.get()),
            "contour_width": int(self.contour_width_var.get()),
            "contour_color_hex": self.contour_color_hex.get(),
            "middle_low_pct": int(self.middle_low_pct_var.get()),
            "middle_high_pct": int(self.middle_high_pct_var.get()),
            "middle_contour_color_hex": self.middle_contour_color_hex.get(),
            "arrow_length_px": int(self.arrow_length_var.get()),
            "show_scaled": bool(self.show_scaled.get()),
            "bar_length_um": float(self.bar_length_um.get()),
            "contrast_mode": self.contrast_mode.get(),
        }

        # Build a human-readable, sectioned structure
        # Format: list of (section_title, [(param_label, value), ...])
        config_sections = [
            (
                "Segmentation & Contours",
                [
                    ("Minimum particle area (pixels)", config_data["min_area_px"]),
                    ("Minimum contour length (pixels)", config_data["min_contour_len"]),
                    ("Contour line width (pixels)", config_data["contour_width"]),
                    ("Main contour color (hex)", config_data["contour_color_hex"]),
                ],
            ),
            (
                "Middle-range Highlighting",
                [
                    ("Lower percentile of rank (%)", config_data["middle_low_pct"]),
                    ("Upper percentile of rank (%)", config_data["middle_high_pct"]),
                    (
                        "Middle-range contour color (hex)",
                        config_data["middle_contour_color_hex"],
                    ),
                ],
            ),
            (
                "Arrow & Labelling",
                [
                    ("Arrow length (pixels)", config_data["arrow_length_px"]),
                ],
            ),
            (
                "Display & Scale Bar",
                [
                    ("Show scale bar", config_data["show_scaled"]),
                    ("Scale bar length (µm)", config_data["bar_length_um"]),
                    ("Display mode", config_data["contrast_mode"]),
                ],
            ),
        ]

        # Turn into a DataFrame with columns: Parameter, Value
        config_rows = []
        for section_title, items in config_sections:
            # Section header row (Value is empty)
            config_rows.append((section_title, ""))  # section header
            for label, value in items:
                # Indent parameter labels for readability
                config_rows.append(("    " + label, value))

        df_config = pd.DataFrame(config_rows, columns=["Parameter", "Value"])

        safe_name = f"{rec.dataset}_{rec.pair_name}".replace(os.sep, "_")
        safe_name = safe_name.replace(":", "_")
        out_path = out_dir / f"{safe_name}.xlsx"

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            # Sheets
            df_stats.to_excel(writer, sheet_name="Statistics", index=False)
            df_meta.to_excel(writer, sheet_name="Metadata", index=False)
            df_config.to_excel(writer, sheet_name="Config", index=False)

            wb = writer.book
            ws_stats = wb["Statistics"]
            ws_meta = wb["Metadata"]
            ws_config = wb["Config"]

            # ----- format Statistics header & columns -----
            header_font = Font(bold=True)
            for cell in ws_stats[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            ws_stats.auto_filter.ref = ws_stats.dimensions

            for col_idx, col in enumerate(ws_stats.columns, start=1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws_stats.column_dimensions[col_letter].width = max(
                    12, min(max_len + 2, 45)
                )

            # ----- add histogram figure image to Statistics sheet -----
            if not df_stats.empty:
                stats_rows_for_fig: list[dict] = []
                for _, row in df_stats.iterrows():
                    stats_rows_for_fig.append(
                        {
                            "idx": int(row["Rank (by Intensity / Area)"]),
                            "area_px": float(row["Area [pixels]"]),
                            "eq_diam_px": float(row["Equivalent Diameter [pixels]"]),
                            "total_intensity": float(
                                row["Total Fluorescence Intensity [a.u.]"]
                            ),
                            "intensity_per_area": float(
                                row["Fluorescence Intensity / Area [a.u./pixel]"]
                            ),
                        }
                    )
                fig = self._create_histogram_figure_for_export(stats_rows_for_fig)
                bio_hist = BytesIO()
                fig.savefig(bio_hist, format="PNG", bbox_inches="tight")
                bio_hist.seek(0)
                xl_hist = XLImage(bio_hist)
                xl_hist.anchor = "A{}".format(ws_stats.max_row + 3)
                ws_stats.add_image(xl_hist)

            # ----- format Metadata sheet: section headers and items -----
            ws_meta["A1"].value = "Metadata item"
            ws_meta["B1"].value = "Value"
            ws_meta["A1"].font = Font(bold=True)
            ws_meta["B1"].font = Font(bold=True)

            current_section = None
            for row in range(2, ws_meta.max_row + 1):
                key_cell = ws_meta[f"A{row}"]
                val_cell = ws_meta[f"B{row}"]
                if not key_cell.value:
                    continue
                text = str(key_cell.value)
                if ":" in text:
                    section, item = text.split(":", 1)
                    section = section.strip()
                    item = item.strip()
                    if section != current_section:
                        key_cell.value = section
                        key_cell.font = Font(bold=True)
                        val_cell.value = ""
                        current_section = section
                    else:
                        key_cell.value = "    " + item
                        key_cell.font = Font(bold=False)
                else:
                    key_cell.font = Font(bold=False)

            for col_idx, col in enumerate(ws_meta.columns, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="top",
                        wrap_text=True,
                    )
                ws_meta.column_dimensions[col_letter].width = max(
                    18, min(max_len + 2, 60)
                )

            # ----- IMPROVED formatting for Config sheet -----
            # Header row
            ws_config["A1"].font = Font(bold=True)
            ws_config["B1"].font = Font(bold=True)
            ws_config["A1"].alignment = Alignment(horizontal="left")
            ws_config["B1"].alignment = Alignment(horizontal="left")

            # Section headers & parameter rows
            for row in range(2, ws_config.max_row + 1):
                param_cell = ws_config[f"A{row}"]
                val_cell = ws_config[f"B{row}"]
                text = str(param_cell.value) if param_cell.value is not None else ""

                if text and not text.startswith("    "):
                    # Section header row
                    param_cell.font = Font(bold=True)
                    val_cell.value = ""  # keep section rows single-column
                else:
                    # Parameter row
                    param_cell.font = Font(bold=False)

                param_cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )
                val_cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )

            # Auto-size columns in Config
            for col_idx, col in enumerate(ws_config.columns, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws_config.column_dimensions[col_letter].width = max(
                    18, min(max_len + 2, 60)
                )

            # ----- add 4 images, one per sheet, scaled to 75% -----
            bf_top_img, fluo_top_img, bf_bottom_img, fluo_bottom_img = \
                self._generate_four_panel_images_for_export(rec)

            def add_single_image_sheet(sheet_name: str, pil_img: Image.Image, title: str):
                ws = wb.create_sheet(sheet_name)
                ws["A1"] = title
                ws["A1"].font = Font(bold=True)
                bio = BytesIO()
                pil_img.save(bio, format="PNG")
                bio.seek(0)
                xl_img = XLImage(bio)
                xl_img.anchor = "A3"
                ws.add_image(xl_img)

            base_name = rec.pair_name
            add_single_image_sheet(
                "BF_Raw",
                bf_top_img,
                f"Brightfield (raw) - {base_name}",
            )
            add_single_image_sheet(
                "FLUO_Raw",
                fluo_top_img,
                f"Fluorescence (raw) - {base_name}",
            )
            add_single_image_sheet(
                "BF_Contours",
                bf_bottom_img,
                f"Brightfield (enhanced + contours) - {base_name}",
            )
            add_single_image_sheet(
                "FLUO_Contours",
                fluo_bottom_img,
                f"Fluorescence (enhanced + contours) - {base_name}",
            )

    # ------------------------------------------------------------------ Image / scale-bar helpers

    def _to_pil(self, arr: np.ndarray) -> Image.Image:
        if arr.ndim == 2:
            arr = arr.astype(np.float32)
            if arr.max() > 0:
                arr = arr / arr.max() * 255.0
            arr8 = arr.astype(np.uint8)
            return Image.fromarray(arr8, mode="L")

        if arr.ndim == 3:
            if arr.dtype != np.uint8:
                arr = arr.astype(np.float32)
                if arr.max() > 0:
                    arr = arr / arr.max() * 255.0
                arr = arr.astype(np.uint8)

            if arr.shape[2] == 1:
                arr = arr[..., 0]
                return Image.fromarray(arr, mode="L")

            if arr.shape[2] > 3:
                arr = arr[..., :3]

            return Image.fromarray(arr, mode="RGB")

        raise ValueError(f"Unsupported image shape: {arr.shape}")

    def _resize_to_box(self, img: Image.Image, box_w: int, box_h: int) -> Image.Image:
        if box_w <= 0 or box_h <= 0:
            return img
        w, h = img.size
        scale = min(box_w / w, box_h / h)
        if scale <= 0:
            return img
        new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
        return img.resize(new_size, RESAMPLE)

    def _pick_contrasting_colors(self, bg_gray: float):
        if bg_gray > 128:
            bar = "black"
        else:
            bar = "white"
        return bar

    def _get_scaled_font(self, base_size: int = 12):
        size = max(1, int(base_size * 2))
        try:
            return ImageFont.truetype("arial.ttf", size=size)
        except Exception:
            return ImageFont.load_default()

    def _add_scale_bar(self, img: Image.Image, rec: ImagePairRecord) -> Image.Image:
        px_um = getattr(rec, "pixel_size_um", None)
        if px_um is None:
            return img

        try:
            bar_len_um = float(self.bar_length_um.get())
        except Exception:
            bar_len_um = 5.0

        bar_len_px = int(round(bar_len_um / px_um))
        if bar_len_px <= 0:
            return img

        w, h = img.size

        sample_height = max(5, h // 20)
        y_start = max(0, h - sample_height)
        bottom_band = img.crop((0, y_start, w, h))
        gray_band = bottom_band.convert("L")
        bg_gray = float(np.array(gray_band, dtype=np.uint8).mean())

        bar_color = self._pick_contrasting_colors(bg_gray)
        outline_color = bar_color
        text_color = bar_color

        draw = ImageDraw.Draw(img)
        margin = 20
        bar_height = max(3, h // 200)

        x0 = margin
        y0 = h - margin - bar_height
        x1 = x0 + bar_len_px
        y1 = y0 + bar_height

        if x1 > w - margin:
            bar_len_px = max(10, w - 2 * margin)
            bar_len_um = bar_len_px * px_um
            x1 = x0 + bar_len_px

        draw.rectangle([x0, y0, x1, y1], fill=bar_color, outline=outline_color)

        label = f"{bar_len_um:.0f} um"
        text_x = x0 + bar_len_px / 2
        text_y = y0 - 8

        font = self._get_scaled_font(base_size=12)
        draw.text((text_x, text_y), label, fill=text_color, font=font, anchor="ms")

        return img


if __name__ == "__main__":
    project_root = Path.cwd()
    source_root = project_root / "source"

    app = BFFluoViewer(source_root)
    app.mainloop()