import cv2
import numpy as np
import os
import glob
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
from io import BytesIO
from pathlib import Path
import argparse
import shutil
from datetime import datetime
from PIL import Image as PILImage

# --------------------------------------------------
# TUNABLE PARAMETERS
# --------------------------------------------------
GAUSSIAN_SIGMA = 15
MORPH_ITERATIONS = 1
DILATE_ITERATIONS = 1
ERODE_ITERATIONS = 1
MIN_OBJECT_AREA = 100
MAX_OBJECT_AREA = 5000

# Red channel visualization parameters
RED_NORMALIZE = True
RED_BRIGHTNESS = 1.2
RED_GAMMA = 0.7

# Scale bar parameters
SCALE_BAR_LENGTH_UM = 10
SCALE_BAR_HEIGHT = 4
SCALE_BAR_MARGIN = 15
SCALE_BAR_COLOR = (255, 255, 255)
SCALE_BAR_BG_COLOR = (0, 0, 0)
SCALE_BAR_TEXT_COLOR = (255, 255, 255)
SCALE_BAR_FONT_SCALE = 0.5
SCALE_BAR_FONT_THICKNESS = 1

# Error bar plotting parameters
ERROR_PERCENTAGE = 0.1
PLOT_DPI = 150

# --------------------------------------------------
# DIRECTORY MANAGEMENT
# --------------------------------------------------

def get_downloads_folder() -> Path:
    """Get the current user's Downloads folder."""
    if os.name == 'nt':  # Windows
        downloads = Path.home() / "Downloads"
    else:  # macOS/Linux
        downloads = Path.home() / "Downloads"
    
    downloads.mkdir(exist_ok=True)
    return downloads


def create_output_directory() -> Path:
    """Create timestamped output directory in Downloads folder."""
    downloads = get_downloads_folder()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = downloads / f"particle_analysis_{timestamp}"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def get_debug_directory() -> Path:
    """Get debug directory (cleaned up each run)."""
    debug_dir = Path("debug")
    if debug_dir.exists():
        shutil.rmtree(debug_dir)
    debug_dir.mkdir(parents=True, exist_ok=True)
    return debug_dir


# --------------------------------------------------
# HELPER FUNCTIONS FROM debug-meta.py
# --------------------------------------------------

def add_scale_bar(img, pixel_size, unit='um', length_um=50):
    """Add a scale bar to the image"""
    if pixel_size is None or pixel_size <= 0:
        return img
    
    bar_length_px = int(round(length_um / pixel_size))
    
    if bar_length_px < 10:
        return img
    
    h, w = img.shape[:2]
    bar_x = w - bar_length_px - SCALE_BAR_MARGIN
    bar_y = h - SCALE_BAR_HEIGHT - SCALE_BAR_MARGIN
    
    if unit == 'µm' or unit == 'um':
        label = f"{length_um} um"
    else:
        label = f"{length_um} {unit}"
    
    font = cv2.FONT_HERSHEY_SIMPLEX
    (text_w, text_h), baseline = cv2.getTextSize(
        label, font, SCALE_BAR_FONT_SCALE, SCALE_BAR_FONT_THICKNESS
    )
    
    text_x = bar_x + (bar_length_px - text_w) // 2
    text_y = bar_y - 8
    
    bg_padding = 5
    bg_x1 = min(bar_x, text_x) - bg_padding
    bg_y1 = text_y - text_h - bg_padding
    bg_x2 = max(bar_x + bar_length_px, text_x + text_w) + bg_padding
    bg_y2 = bar_y + SCALE_BAR_HEIGHT + bg_padding
    
    overlay = img.copy()
    cv2.rectangle(overlay, (bg_x1, bg_y1), (bg_x2, bg_y2), SCALE_BAR_BG_COLOR, -1)
    cv2.addWeighted(overlay, 0.7, img, 0.3, 0, img)
    
    cv2.rectangle(img, (bar_x, bar_y), 
                  (bar_x + bar_length_px, bar_y + SCALE_BAR_HEIGHT),
                  SCALE_BAR_COLOR, -1)
    
    cv2.putText(img, label, (text_x, text_y), 
                font, SCALE_BAR_FONT_SCALE, SCALE_BAR_TEXT_COLOR, 
                SCALE_BAR_FONT_THICKNESS, cv2.LINE_AA)
    
    return img


def numpy_to_excel_image(img_array: np.ndarray, format: str = 'PNG') -> XLImage:
    """Convert numpy array to Excel-compatible image."""
    # Convert to PIL Image
    if len(img_array.shape) == 2:  # Grayscale
        pil_img = PILImage.fromarray(img_array)
    else:  # Color (BGR to RGB)
        pil_img = PILImage.fromarray(cv2.cvtColor(img_array, cv2.COLOR_BGR2RGB))
    
    # Save to bytes buffer
    img_buffer = BytesIO()
    pil_img.save(img_buffer, format=format)
    img_buffer.seek(0)
    
    # Create Excel image
    excel_img = XLImage(img_buffer)
    # Scale down for Excel (max width ~600 pixels)
    scale_factor = min(1.0, 600 / pil_img.width)
    excel_img.width = int(pil_img.width * scale_factor)
    excel_img.height = int(pil_img.height * scale_factor)
    
    return excel_img


def parse_metadata(xml_path):
    """Parse metadata XML to extract physical pixel size and bit depth"""
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        
        pixel_size_x = None
        pixel_size_y = None
        unit = None
        bit_depth = None
        
        for dim in root.iter('DimensionDescription'):
            dim_id = dim.get('DimID')
            length = dim.get('Length')
            num_elements = dim.get('NumberOfElements')
            dim_unit = dim.get('Unit')
            
            if length and num_elements:
                pixel_size = float(length) / float(num_elements)
                
                if dim_id == 'X':
                    pixel_size_x = pixel_size
                    if dim_unit:
                        unit = dim_unit
                elif dim_id == 'Y':
                    pixel_size_y = pixel_size
                    if dim_unit and not unit:
                        unit = dim_unit
        
        if not pixel_size_x or not pixel_size_y:
            for channel in root.iter('ChannelDescription'):
                optical_res = channel.get('OpticalResolutionXY')
                if optical_res:
                    parts = optical_res.split()
                    if len(parts) >= 2:
                        pixel_size_x = pixel_size_y = float(parts[0])
                        unit = parts[1]
                
                if not bit_depth:
                    resolution = channel.get('Resolution')
                    if resolution:
                        bit_depth = int(resolution)
        
        if unit in ['µm', 'μm']:
            unit = 'um'
        
        if pixel_size_x and pixel_size_y and not unit:
            unit = 'um'
        
        if pixel_size_x and pixel_size_y:
            return pixel_size_x, pixel_size_y, unit, bit_depth
        else:
            return None, None, None, bit_depth
            
    except Exception as e:
        print(f"⚠ Error parsing metadata: {e}")
        return None, None, None, None


def adjust_red_channel(img, normalize=True, brightness=1.0, gamma=1.0):
    """Adjust red channel with normalization, brightness, and gamma"""
    img_float = img.astype(np.float32)
    
    if normalize:
        min_val = np.min(img_float)
        max_val = np.max(img_float)
        if max_val > min_val:
            img_float = (img_float - min_val) * 255.0 / (max_val - min_val)
    
    if brightness != 1.0:
        img_float = img_float * brightness
        img_float = np.clip(img_float, 0, 255)
    
    if gamma != 1.0:
        img_normalized = img_float / 255.0
        img_normalized = np.power(img_normalized, gamma)
        img_float = img_normalized * 255.0
    
    return img_float.astype(np.uint8)


def style_header(ws, row_num, fill_color='4472C4'):
    """Apply header styling to a row"""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(bold=True, color='FFFFFF', size=11)
    alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def style_summary_header(cell):
    """Apply summary section header styling"""
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')


def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def trim_outliers(data, trim_percentage=0.20):
    """
    Remove top and bottom percentages of data based on intensity_per_area_orig
    
    Args:
        data: List of particle dictionaries
        trim_percentage: Percentage to trim from each end (default 0.20 = 20%)
    
    Returns:
        Trimmed list of particles (middle 60% by default)
    """
    if not data:
        return []
    
    # Sort by intensity per area
    sorted_data = sorted(data, key=lambda x: x['intensity_per_area_orig'])
    
    n = len(sorted_data)
    lower_cutoff = int(n * trim_percentage)
    upper_cutoff = int(n * (1 - trim_percentage))
    
    # Keep middle portion
    trimmed = sorted_data[lower_cutoff:upper_cutoff]
    
    print(f"  Original particles: {n}")
    print(f"  Removed bottom {trim_percentage*100:.0f}%: {lower_cutoff} particles")
    print(f"  Removed top {trim_percentage*100:.0f}%: {n - upper_cutoff} particles")
    print(f"  Retained: {len(trimmed)} particles ({len(trimmed)/n*100:.1f}%)")
    
    return trimmed


# --------------------------------------------------
# BATCH PROCESSING FUNCTIONS
# --------------------------------------------------

def find_image_pairs(directory):
    """
    Find all matching BF/FL image pairs in a directory
    
    Returns:
    - List of tuples: (base_name, bf_path, fl_path, xml_path)
    """
    directory = Path(directory)
    
    bf_files = sorted(directory.glob("*_ch00.tif"))
    
    pairs = []
    for bf_path in bf_files:
        base_name = bf_path.stem.replace('_ch00', '')
        
        fl_path = directory / f"{base_name}_ch01.tif"
        xml_path = directory / "MetaData" / f"{base_name}_Properties.xml"
        
        if fl_path.exists():
            pairs.append((base_name, str(bf_path), str(fl_path), str(xml_path)))
        else:
            print(f"⚠ Warning: No FL image for {base_name}")
    
    return pairs


def process_image_pair(base_name, bf_path, fl_path, xml_path, debug_dir):
    """
    Process a single BF/FL image pair
    
    Returns:
    - Dictionary with all measurements and metadata
    """
    print(f"\n{'='*60}")
    print(f"PROCESSING: {base_name}")
    print(f"{'='*60}")
    
    img_debug_dir = os.path.join(debug_dir, base_name)
    os.makedirs(img_debug_dir, exist_ok=True)
    
    def img_save_debug(name, img, pixel_size=None, unit='um'):
        path = os.path.join(img_debug_dir, name)
        img_with_scale = img.copy()
        if pixel_size is not None and pixel_size > 0:
            img_with_scale = add_scale_bar(img_with_scale, pixel_size, unit, SCALE_BAR_LENGTH_UM)
        cv2.imwrite(path, img_with_scale)
    
    pixel_size_x, pixel_size_y, unit, metadata_bit_depth = parse_metadata(xml_path)
    
    if pixel_size_x is not None and pixel_size_y is not None:
        pixel_size = (pixel_size_x + pixel_size_y) / 2.0
        area_factor = pixel_size ** 2
        unit_str = unit if unit else 'um'
        has_calibration = True
    else:
        pixel_size = None
        area_factor = 1.0
        unit_str = 'pixels'
        has_calibration = False
    
    img_bf = cv2.imread(bf_path, cv2.IMREAD_UNCHANGED)
    if img_bf is None:
        raise FileNotFoundError(f"Cannot load brightfield image: {bf_path}")
    if img_bf.ndim == 3:
        img_bf = cv2.cvtColor(img_bf, cv2.COLOR_BGR2GRAY)
    
    img_red_original = cv2.imread(fl_path, cv2.IMREAD_UNCHANGED)
    if img_red_original is None:
        raise FileNotFoundError(f"Cannot load fluorescence image: {fl_path}")
    if img_red_original.ndim == 3:
        img_red_original = cv2.cvtColor(img_red_original, cv2.COLOR_BGR2GRAY)
    
    original_dtype = img_red_original.dtype
    original_max = img_red_original.max()
    
    if original_dtype == np.uint16:
        if original_max <= 4095:
            bit_depth = 12
            max_possible_value = 4095
        elif original_max <= 16383:
            bit_depth = 14
            max_possible_value = 16383
        else:
            bit_depth = 16
            max_possible_value = 65535
        bit_conversion_factor = max_possible_value / 255.0
    else:
        bit_depth = 8
        max_possible_value = 255
        bit_conversion_factor = 1.0
    
    if img_bf.dtype == np.uint16:
        img_bf_8bit = np.zeros_like(img_bf, dtype=np.uint8)
        cv2.normalize(img_bf, img_bf_8bit, 0, 255, cv2.NORM_MINMAX)
        img_bf = img_bf_8bit
    
    if img_red_original.dtype == np.uint16:
        img_red_8bit = np.zeros_like(img_red_original, dtype=np.uint8)
        cv2.normalize(img_red_original, img_red_8bit, 0, 255, cv2.NORM_MINMAX)
    else:
        img_red_8bit = img_red_original.copy()
    
    img_red_enhanced = adjust_red_channel(img_red_8bit, RED_NORMALIZE, RED_BRIGHTNESS, RED_GAMMA)
    
    bg = cv2.GaussianBlur(img_bf, (0, 0), sigmaX=GAUSSIAN_SIGMA, sigmaY=GAUSSIAN_SIGMA)
    enhanced = cv2.subtract(bg, img_bf)
    enhanced_blur = cv2.GaussianBlur(enhanced, (3, 3), 0)
    
    _, thresh = cv2.threshold(enhanced_blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    kernel = np.ones((3, 3), np.uint8)
    closed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=MORPH_ITERATIONS)
    closed = cv2.dilate(closed, kernel, iterations=DILATE_ITERATIONS)
    closed = cv2.erode(closed, kernel, iterations=ERODE_ITERATIONS)
    
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(closed, connectivity=8)
    solid = np.where(labels > 0, 255, 0).astype(np.uint8)
    
    contours, hierarchy = cv2.findContours(solid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    filtered = [c for c in contours if MIN_OBJECT_AREA <= cv2.contourArea(c) <= MAX_OBJECT_AREA]
    
    print(f"  Contours found: {len(contours)}")
    print(f"  After filtering: {len(filtered)}")
    
    object_data = []
    objects_without_red = 0
    
    for c in filtered:
        area_px = cv2.contourArea(c)
        perimeter_px = cv2.arcLength(c, True)
        
        area_physical = area_px * area_factor if has_calibration else area_px
        perimeter_physical = perimeter_px * pixel_size if has_calibration and pixel_size else perimeter_px
        
        M = cv2.moments(c)
        cx = int(M["m10"] / M["m00"]) if M["m00"] != 0 else 0
        cy = int(M["m01"] / M["m00"]) if M["m00"] != 0 else 0
        
        mask = np.zeros_like(img_red_8bit, dtype=np.uint8)
        cv2.drawContours(mask, [c], -1, 255, -1)
        
        red_pixels_orig = img_red_original[mask == 255]
        if len(red_pixels_orig) > 0:
            red_total_orig = float(np.sum(red_pixels_orig.astype(np.float64)))
            red_mean_orig = float(np.mean(red_pixels_orig.astype(np.float64)))
            red_std_orig = float(np.std(red_pixels_orig.astype(np.float64)))
        else:
            red_total_orig = red_mean_orig = red_std_orig = 0.0
        
        if red_total_orig == 0.0:
            objects_without_red += 1
            continue
        
        intensity_per_area_orig = red_total_orig / area_physical if area_physical > 0 else 0.0
        
        object_data.append({
            'contour': c,
            'area_px': area_px,
            'area_physical': area_physical,
            'perimeter_px': perimeter_px,
            'perimeter_physical': perimeter_physical,
            'centroid_x': cx,
            'centroid_y': cy,
            'red_total_orig': red_total_orig,
            'red_mean_orig': red_mean_orig,
            'red_std_orig': red_std_orig,
            'intensity_per_area_orig': intensity_per_area_orig,
        })
    
    object_data.sort(key=lambda x: x['intensity_per_area_orig'], reverse=True)
    for i, obj in enumerate(object_data, 1):
        obj['object_id'] = i
    
    print(f"  Particles with fluorescence: {len(object_data)}")
    print(f"  Particles excluded: {objects_without_red}")
    
    # Create contour visualization image
    vis_bf = cv2.cvtColor(img_bf, cv2.COLOR_GRAY2BGR)
    for obj in object_data:
        cv2.drawContours(vis_bf, [obj['contour']], -1, (0, 0, 255), 1)
        label = str(obj['object_id'])
        cv2.putText(vis_bf, label, (obj['centroid_x'], obj['centroid_y']), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1, cv2.LINE_AA)
    
    # Add scale bar to contour image
    vis_bf_with_scale = add_scale_bar(vis_bf.copy(), pixel_size, unit_str, SCALE_BAR_LENGTH_UM)
    
    # Save to debug directory
    img_save_debug("contours.png", vis_bf, pixel_size, unit_str)
    
    return {
        'image_name': base_name,
        'object_data': object_data,
        'contour_image': vis_bf_with_scale,  # Store for Excel embedding
        'metadata': {
            'pixel_size': pixel_size,
            'unit': unit_str,
            'has_calibration': has_calibration,
            'area_factor': area_factor,
            'bit_depth': bit_depth,
            'max_value': max_possible_value,
            'bit_conversion_factor': bit_conversion_factor,
            'objects_excluded': objects_without_red
        }
    }


def process_group(group_dir, group_name, debug_dir):
    """Process all images in a group directory"""
    print(f"\n{'#'*60}")
    print(f"# PROCESSING GROUP: {group_name}")
    print(f"# Directory: {group_dir}")
    print(f"{'#'*60}\n")
    
    group_debug_dir = os.path.join(debug_dir, group_name)
    os.makedirs(group_debug_dir, exist_ok=True)
    
    pairs = find_image_pairs(group_dir)
    
    if not pairs:
        print(f"⚠ No image pairs found in {group_dir}")
        return []
    
    print(f"Found {len(pairs)} image pairs")
    
    all_results = []
    for base_name, bf_path, fl_path, xml_path in pairs:
        try:
            result = process_image_pair(base_name, bf_path, fl_path, xml_path, group_debug_dir)
            result['group'] = group_name
            all_results.append(result)
        except Exception as e:
            print(f"❌ Error processing {base_name}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    print(f"\n✓ Processed {len(all_results)}/{len(pairs)} images successfully")
    
    return all_results


def save_group_excel(results, group_name, output_dir, apply_trimming=True, trim_percentage=0.20):
    """Save Excel file for a single group with optional trimming and embedded images"""
    if not results:
        print(f"⚠ No results to save for {group_name}")
        return
    
    excel_file = os.path.join(output_dir, f"{group_name}_analysis.xlsx")
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    meta = results[0]['metadata']
    bit_depth = meta['bit_depth']
    unit_str = meta['unit']
    has_calibration = meta['has_calibration']
    
    # Aggregate all particles
    all_particles = []
    for result in results:
        all_particles.extend(result['object_data'])
    
    # Apply trimming if requested
    if apply_trimming and len(all_particles) > 0:
        print(f"\n📊 Trimming outliers for {group_name}:")
        trimmed_particles = trim_outliers(all_particles, trim_percentage)
    else:
        trimmed_particles = all_particles
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    ws_summary.append(['GROUP SUMMARY'])
    style_summary_header(ws_summary.cell(1, 1))
    ws_summary.append(['Group Name:', group_name])
    ws_summary.append(['Total Images:', len(results)])
    ws_summary.append([])
    
    # Statistics - Original data
    ws_summary.append(['ORIGINAL DATA STATISTICS'])
    style_summary_header(ws_summary.cell(5, 1))
    ws_summary.append(['Total Particles:', len(all_particles)])
    
    if all_particles:
        intensities = [p['intensity_per_area_orig'] for p in all_particles]
        ws_summary.append(['Mean Intensity per Area:', f"{np.mean(intensities):.2f}"])
        ws_summary.append(['Std Dev:', f"{np.std(intensities, ddof=1):.2f}"])
        ws_summary.append(['SEM:', f"{np.std(intensities, ddof=1) / np.sqrt(len(intensities)):.2f}"])
        ws_summary.append(['Min:', f"{np.min(intensities):.2f}"])
        ws_summary.append(['Max:', f"{np.max(intensities):.2f}"])
    
    ws_summary.append([])
    
    # Statistics - Trimmed data
    if apply_trimming:
        ws_summary.append(['TRIMMED DATA STATISTICS (Middle 60%)'])
        style_summary_header(ws_summary.cell(ws_summary.max_row, 1))
        ws_summary.append(['Total Particles:', len(trimmed_particles)])
        
        if trimmed_particles:
            trimmed_intensities = [p['intensity_per_area_orig'] for p in trimmed_particles]
            ws_summary.append(['Mean Intensity per Area:', f"{np.mean(trimmed_intensities):.2f}"])
            ws_summary.append(['Std Dev:', f"{np.std(trimmed_intensities, ddof=1):.2f}"])
            ws_summary.append(['SEM:', f"{np.std(trimmed_intensities, ddof=1) / np.sqrt(len(trimmed_intensities)):.2f}"])
            ws_summary.append(['Min:', f"{np.min(trimmed_intensities):.2f}"])
            ws_summary.append(['Max:', f"{np.max(trimmed_intensities):.2f}"])
    
    auto_adjust_column_width(ws_summary)
    
    # Create Images sheet with contours
    ws_images = wb.create_sheet("Contour Images")
    
    current_row = 1
    for result in results:
        img_name = result['image_name']
        contour_img = result['contour_image']
        
        # Add image title
        cell = ws_images.cell(row=current_row, column=1)
        cell.value = f"Image: {img_name}"
        cell.font = Font(bold=True, size=12)
        current_row += 1
        
        # Add particle count
        cell = ws_images.cell(row=current_row, column=1)
        cell.value = f"Particles detected: {len(result['object_data'])}"
        current_row += 1
        
        # Embed contour image
        excel_img = numpy_to_excel_image(contour_img)
        ws_images.add_image(excel_img, f'A{current_row}')
        
        # Calculate rows needed for image (approximate: height in pixels / 15)
        rows_needed = int(excel_img.height / 15) + 2
        current_row += rows_needed
        
        # Add spacing between images
        current_row += 2
    
    # Individual image data sheets
    for result in results:
        img_name = result['image_name']
        object_data = result['object_data']
        
        ws = wb.create_sheet(img_name[:31])
        
        headers = [
            'Particle ID',
            f'Area ({unit_str}²)',
            f'Perimeter ({unit_str})',
            f'Total Intensity ({bit_depth}-bit)',
            f'Mean Intensity',
            f'Std Dev',
            f'Intensity per {unit_str}²'
        ]
        ws.append(headers)
        style_header(ws, 1)
        
        for obj in object_data:
            ws.append([
                obj['object_id'],
                round(obj['area_physical'], 4 if has_calibration else 2),
                round(obj['perimeter_physical'], 2),
                round(obj['red_total_orig'], 1),
                round(obj['red_mean_orig'], 2),
                round(obj['red_std_orig'], 2),
                round(obj['intensity_per_area_orig'], 2)
            ])
        
        auto_adjust_column_width(ws)
    
    wb.save(excel_file)
    print(f"✓ Saved: {excel_file}")
    
    # Return trimmed data for comparison
    return trimmed_particles


# --------------------------------------------------
# MAIN EXECUTION
# --------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Batch process Control and Sample groups')
    parser.add_argument('--sample-group', type=str, help='Sample group directory name')
    parser.add_argument('--control-group', type=str, default='Control group',
                        help='Control group directory name')
    parser.add_argument('--source-dir', type=str, default='source',
                        help='Source directory containing group folders')
    parser.add_argument('--trim', action='store_true',
                        help='Enable outlier trimming (disabled by default)')
    parser.add_argument('--trim-percentage', type=float, default=0.20,
                        help='Percentage to trim from each end (default: 0.20)')
    args = parser.parse_args()
    
    source_dir = Path(args.source_dir)
    
    # Create output directory in Downloads
    output_dir = create_output_directory()
    
    # Create debug directory (cleaned up each run)
    debug_dir = get_debug_directory()
    
    if not args.sample_group:
        print("\nAvailable directories:")
        dirs = [d.name for d in source_dir.iterdir() if d.is_dir()]
        for i, d in enumerate(dirs, 1):
            print(f"  {i}. {d}")
        
        choice = input("\nEnter sample group name (or number): ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(dirs):
            args.sample_group = dirs[int(choice) - 1]
        else:
            args.sample_group = choice
    
    apply_trimming = args.trim
    
    print(f"\n{'='*60}")
    print(f"PARTICLE ANALYSIS - CONFIGURATION")
    print(f"{'='*60}")
    print(f"  Control group: {args.control_group}")
    print(f"  Sample group: {args.sample_group}")
    print(f"  Output directory: {output_dir}")
    print(f"  Debug directory: {debug_dir}")
    print(f"  Outlier trimming: {'Enabled' if apply_trimming else 'Disabled'}")
    if apply_trimming:
        print(f"  Trim percentage: {args.trim_percentage*100:.0f}% each end")
    print(f"{'='*60}\n")
    
    # Process Control group
    control_dir = source_dir / args.control_group
    control_results = process_group(str(control_dir), "Control", str(debug_dir))
    control_trimmed = None
    if control_results:
        control_trimmed = save_group_excel(control_results, "Control", str(output_dir), 
                                          apply_trimming, args.trim_percentage)
    
    # Process Sample group
    sample_dir = source_dir / args.sample_group
    sample_results = process_group(str(sample_dir), f"Sample_{args.sample_group}", str(debug_dir))
    sample_trimmed = None
    if sample_results:
        sample_trimmed = save_group_excel(sample_results, f"Sample_{args.sample_group}", 
                                         str(output_dir), apply_trimming, args.trim_percentage)
    
    print(f"\n{'#'*60}")
    print("# BATCH PROCESSING COMPLETE")
    print(f"{'#'*60}")
    print(f"✓ Control images processed: {len(control_results)}")
    print(f"✓ Sample images processed: {len(sample_results)}")
    if apply_trimming:
        if control_trimmed:
            print(f"✓ Control particles (trimmed): {len(control_trimmed)}")
        if sample_trimmed:
            print(f"✓ Sample particles (trimmed): {len(sample_trimmed)}")
    print(f"\n📁 Excel outputs saved to: {output_dir}")
    print(f"🔍 Debug files saved to: {debug_dir}")
    print(f"{'#'*60}\n")


if __name__ == "__main__":
    main()