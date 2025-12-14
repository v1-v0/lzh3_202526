#!/usr/bin/env python3
"""
Particle Scout - Advanced Particle Analysis Tool
Analyzes fluorescence microscopy images to detect and quantify particles
"""

import cv2
import numpy as np
import argparse
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from io import BytesIO
import os
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend

# Global constants
PLOT_DPI = 100
DEBUG_ENABLED = True


def get_debug_directory():
    """Get or create debug directory in Downloads"""
    downloads = Path.home() / 'Downloads'
    debug_dir = downloads / 'particle_scout_debug'
    debug_dir.mkdir(exist_ok=True)
    return debug_dir


def create_output_directory():
    """Create timestamped output directory in Downloads"""
    downloads = Path.home() / 'Downloads'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = downloads / f'particle_analysis_{timestamp}'
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def load_image(image_path):
    """
    Load image and extract metadata
    
    Args:
        image_path: Path to image file
        
    Returns:
        tuple: (image_array, metadata_dict) or (None, None) if failed
    """
    try:
        # Try loading with PIL first to get metadata
        pil_img = PILImage.open(image_path)
        
        # Extract metadata
        metadata = {
            'filename': Path(image_path).name,
            'format': pil_img.format,
            'mode': pil_img.mode,
            'size': pil_img.size,
            'info': pil_img.info
        }
        
        # Get calibration info from TIFF tags if available
        has_calibration = False
        pixel_size = 1.0
        unit = 'pixels'
        
        # Use type: ignore for tag_v2 access (exists at runtime for TIFF images)
        if hasattr(pil_img, 'tag_v2'):
            # Resolution tags (282=XResolution, 283=YResolution, 296=ResolutionUnit)
            x_res = pil_img.tag_v2.get(282, None)  # type: ignore
            y_res = pil_img.tag_v2.get(283, None)  # type: ignore
            res_unit = pil_img.tag_v2.get(296, 1)  # type: ignore  # 1=None, 2=inch, 3=cm
            
            if x_res and y_res:
                # Handle IFDRational objects - convert to float
                try:
                    # IFDRational objects have numerator and denominator
                    if hasattr(x_res, 'numerator') and hasattr(x_res, 'denominator'):
                        x_res_value = float(x_res.numerator) / float(x_res.denominator)
                    else:
                        x_res_value = float(x_res)
                    
                    if hasattr(y_res, 'numerator') and hasattr(y_res, 'denominator'):
                        y_res_value = float(y_res.numerator) / float(y_res.denominator)
                    else:
                        y_res_value = float(y_res)
                    
                    # x_res_value is "pixels per unit"
                    # We need to convert this to "micrometers per pixel"
                    
                    if res_unit == 2:  # inches
                        # pixels_per_inch -> convert to um_per_pixel
                        # 1 inch = 25400 micrometers
                        # pixel_size = micrometers_per_inch / pixels_per_inch
                        pixel_size = 25400.0 / x_res_value
                        unit = 'um'
                        has_calibration = True
                    elif res_unit == 3:  # cm
                        # pixels_per_cm -> convert to um_per_pixel
                        # 1 cm = 10000 micrometers
                        pixel_size = 10000.0 / x_res_value
                        unit = 'um'
                        has_calibration = True
                    
                    print(f"  Resolution: {x_res_value:.2f} pixels per {'inch' if res_unit == 2 else 'cm' if res_unit == 3 else 'unit'}")
                    
                    # Safety check: If pixel size is unrealistically large (>50 um), 
                    # it's likely just a screen resolution default, not real microscopy calibration
                    if pixel_size > 50:
                        print(f"  ⚠ Warning: Unrealistic pixel size ({pixel_size:.2f} um/pixel)")
                        print(f"  → Falling back to pixel-based analysis")
                        has_calibration = False
                        pixel_size = 1.0
                        unit = 'pixels'
                    
                except (AttributeError, TypeError, ZeroDivisionError) as e:
                    # If calibration extraction fails, use default values
                    print(f"  Warning: Could not extract calibration: {e}")
                    has_calibration = False
                    pixel_size = 1.0
                    unit = 'pixels'
        
        # Convert to numpy array
        img_array = np.array(pil_img)
        
        # Detect bit depth
        if img_array.dtype == np.uint8:
            bit_depth = 8
        elif img_array.dtype == np.uint16:
            bit_depth = 16
        else:
            bit_depth = 8
            img_array = img_array.astype(np.uint8)
        
        metadata['bit_depth'] = bit_depth
        metadata['has_calibration'] = has_calibration
        metadata['pixel_size'] = pixel_size
        metadata['unit'] = unit
        
        print(f"✓ Loaded: {metadata['filename']}")
        print(f"  Size: {metadata['size']}, Mode: {metadata['mode']}, Bit depth: {bit_depth}")
        print(f"  Calibration: {has_calibration}, Pixel size: {pixel_size:.6f} {unit}/pixel")
        
        return img_array, metadata
        
    except Exception as e:
        print(f"✗ Error loading {image_path}: {e}")
        import traceback
        traceback.print_exc()
        return None, None




def extract_red_channel(image):
    """
    Extract red channel from image
    
    Args:
        image: Input image (grayscale or RGB)
        
    Returns:
        Red channel as 2D array
    """
    if len(image.shape) == 2:
        # Already grayscale
        return image
    elif len(image.shape) == 3:
        if image.shape[2] == 3:
            # RGB image - extract red channel
            return image[:, :, 0]
        elif image.shape[2] == 4:
            # RGBA image - extract red channel
            return image[:, :, 0]
    
    return image


def create_mask(red_channel, method='otsu', invert=True):
    """
    Create binary mask using thresholding
    
    Args:
        red_channel: Red channel image
        method: Thresholding method ('otsu' or 'adaptive')
        invert: If True, detect dark objects; if False, detect bright objects
        
    Returns:
        Binary mask
    """
    # Normalize to 8-bit if needed
    if red_channel.dtype == np.uint16:
        red_8bit = (red_channel / 256).astype(np.uint8)
    else:
        red_8bit = red_channel
    
    if method == 'otsu':
        # Otsu's thresholding
        _, mask = cv2.threshold(red_8bit, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    else:
        # Adaptive thresholding
        mask = cv2.adaptiveThreshold(red_8bit, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                     cv2.THRESH_BINARY, 11, 2)
    
    # Invert mask if we want to detect dark objects
    if invert:
        mask = cv2.bitwise_not(mask)
    
    return mask

def clean_mask(mask, min_size=50):
    """
    Clean mask using morphological operations
    
    Args:
        mask: Binary mask
        min_size: Minimum object size to keep
        
    Returns:
        Cleaned mask
    """
    # Remove small noise
    kernel = np.ones((3,3), np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel, iterations=2)
    
    # Fill holes
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=2)
    
    # Remove small objects
    num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(mask, connectivity=8)
    
    cleaned_mask = np.zeros_like(mask)
    for i in range(1, num_labels):
        if stats[i, cv2.CC_STAT_AREA] >= min_size:
            cleaned_mask[labels == i] = 255
    
    return cleaned_mask


def analyze_objects(red_channel_orig, mask, pixel_size=1.0, unit='pixels'):
    """
    Analyze detected objects and extract measurements
    
    Args:
        red_channel_orig: Original red channel (preserving bit depth)
        mask: Binary mask
        pixel_size: Size of one pixel in physical units
        unit: Unit of measurement
        
    Returns:
        List of object dictionaries with measurements
    """
    # Find contours
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    objects = []
    for idx, contour in enumerate(contours, 1):
        # Create mask for this object
        obj_mask = np.zeros_like(mask)
        cv2.drawContours(obj_mask, [contour], -1, 255, -1)
        
        # Extract pixel values (original bit depth)
        pixels = red_channel_orig[obj_mask == 255]
        
        if len(pixels) == 0:
            continue
        
        # Measurements in pixels
        area_pixels = cv2.contourArea(contour)
        perimeter_pixels = cv2.arcLength(contour, True)
        
        # Convert to physical units
        area_physical = area_pixels * (pixel_size ** 2)
        perimeter_physical = perimeter_pixels * pixel_size
        
        # Intensity measurements (original values)
        total_intensity_orig = float(np.sum(pixels))
        mean_intensity_orig = float(np.mean(pixels))
        std_intensity_orig = float(np.std(pixels))
        
        # Normalized intensity per unit area
        intensity_per_area_orig = total_intensity_orig / area_physical if area_physical > 0 else 0
        
        obj_data = {
            'object_id': idx,
            'contour': contour,
            'area_pixels': area_pixels,
            'area_physical': area_physical,
            'perimeter_pixels': perimeter_pixels,
            'perimeter_physical': perimeter_physical,
            'red_total_orig': total_intensity_orig,
            'red_mean_orig': mean_intensity_orig,
            'red_std_orig': std_intensity_orig,
            'intensity_per_area_orig': intensity_per_area_orig,
            'pixel_count': len(pixels)
        }
        
        objects.append(obj_data)
    
    return objects


def create_contour_overlay(image, objects, show_labels=True):
    """
    Create visualization with contours and labels
    
    Args:
        image: Original image
        objects: List of object dictionaries
        show_labels: Whether to show particle IDs
        
    Returns:
        RGB image with overlays
    """
    # Convert to RGB if grayscale
    if len(image.shape) == 2:
        vis = cv2.cvtColor(image, cv2.COLOR_GRAY2RGB)
    else:
        vis = image.copy()
    
    # Normalize to 8-bit for display
    if vis.dtype == np.uint16:
        vis = (vis / 256).astype(np.uint8)
    
    # Draw contours and labels
    for obj in objects:
        contour = obj['contour']
        
        # Draw contour in green
        cv2.drawContours(vis, [contour], -1, (0, 255, 0), 2)
        
        if show_labels:
            # Calculate centroid
            M = cv2.moments(contour)
            if M['m00'] != 0:
                cx = int(M['m10'] / M['m00'])
                cy = int(M['m01'] / M['m00'])
                
                # Draw particle ID
                cv2.putText(vis, str(obj['object_id']), (cx-10, cy+5),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 0), 2)
    
    return vis


def save_debug_image(image, filename, debug_dir):
    """Save debug image to debug directory"""
    if not DEBUG_ENABLED:
        return
    
    filepath = Path(debug_dir) / filename
    
    # Normalize for saving
    if image.dtype == np.uint16:
        image_save = (image / 256).astype(np.uint8)
    else:
        image_save = image
    
    cv2.imwrite(str(filepath), image_save)
    print(f"  Debug: {filename}")


def process_image(image_path, min_area=50, max_area=10000, output_dir='output'):
    """
    Process a single multi-channel image
    """
    print(f"\n{'='*60}")
    print(f"Processing: {os.path.basename(image_path)}")
    print(f"{'='*60}")
    
    # Read image
    img = cv2.imread(image_path, cv2.IMREAD_UNCHANGED)
    if img is None:
        print(f"✗ Failed to read image: {image_path}")
        return None
    
    print(f"✓ Image loaded - Shape: {img.shape}, Type: {img.dtype}")
    
    # Split channels
    if len(img.shape) == 3 and img.shape[2] == 3:
        blue_channel, green_channel, red_channel = cv2.split(img)
        print(f"✓ Split into 3 channels (BGR)")
    else:
        print(f"✗ Unexpected image format")
        return None
    
    # Create mask from RED channel only (brightfield/phase contrast)
    mask = create_mask(red_channel, method='otsu', invert=True)
    print(f"✓ Created mask using Otsu thresholding (detecting dark objects)")
    
    # Clean up mask
    kernel = np.ones((3,3), np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=2)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel, iterations=1)
    print(f"✓ Applied morphological operations")
    
    # Find contours
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    print(f"✓ Found {len(contours)} initial contours")
    
    # Filter contours by area
    filtered_contours = [cnt for cnt in contours if min_area < cv2.contourArea(cnt) < max_area]
    print(f"✓ Filtered to {len(filtered_contours)} contours (area: {min_area}-{max_area} px²)")
    
    # Measure fluorescence for each contour
    results = []
    for i, contour in enumerate(filtered_contours):
        # Create mask for this contour
        contour_mask = np.zeros(blue_channel.shape, dtype=np.uint8)
        cv2.drawContours(contour_mask, [contour], -1, 255, -1)
        
        # Measure in blue channel (fluorescence)
        blue_pixels = blue_channel[contour_mask == 255]
        mean_intensity = float(np.mean(blue_pixels.astype(np.float64))) if blue_pixels.size > 0 else 0.0
        total_intensity = float(np.sum(blue_pixels.astype(np.float64)))
        area = cv2.contourArea(contour)
        
        # Get centroid
        M = cv2.moments(contour)
        if M['m00'] != 0:
            cx = int(M['m10'] / M['m00'])
            cy = int(M['m01'] / M['m00'])
        else:
            cx, cy = 0, 0
        
        results.append({
            'object_id': i + 1,
            'area_pixels': area,
            'centroid_x': cx,
            'centroid_y': cy,
            'mean_fluorescence': mean_intensity,
            'total_fluorescence': total_intensity
        })
    
    print(f"✓ Measured fluorescence for {len(results)} objects")
    
    # Create visualization
    vis_img = cv2.cvtColor(red_channel, cv2.COLOR_GRAY2BGR)
    
    # Draw contours in green
    cv2.drawContours(vis_img, filtered_contours, -1, (0, 255, 0), 2)
    
    # Overlay blue channel fluorescence
    blue_colored = np.zeros_like(vis_img)
    blue_colored[:, :, 0] = blue_channel  # Blue channel
    vis_img = cv2.addWeighted(vis_img, 0.7, blue_colored, 0.3, 0)
    
    # Add labels
    for result in results:
        cv2.putText(vis_img, str(result['object_id']), 
                   (result['centroid_x'], result['centroid_y']),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1)
    
    # Save outputs
    base_name = os.path.splitext(os.path.basename(image_path))[0]
    
    # Save visualization
    output_path = os.path.join(output_dir, f"{base_name}_contours.png")
    cv2.imwrite(output_path, vis_img)
    print(f"✓ Saved visualization: {output_path}")
    
    # Save mask
    mask_path = os.path.join(output_dir, f"{base_name}_mask.png")
    cv2.imwrite(mask_path, mask)
    print(f"✓ Saved mask: {mask_path}")
    
    return results


def process_group(group_dir, group_name, output_dir):
    """
    Process all images in a group directory
    
    Args:
        group_dir: Directory containing images
        group_name: Name of the group
        output_dir: Directory for outputs (replaces debug_dir)
        
    Returns:
        List of result dictionaries
    """
    print(f"\n{'#'*60}")
    print(f"# Processing Group: {group_name}")
    print(f"# Directory: {group_dir}")
    print(f"{'#'*60}")
    
    group_path = Path(group_dir)
    if not group_path.exists():
        print(f"✗ Directory does not exist: {group_dir}")
        return []
    
    # Find all image files
    image_extensions = ['.tif', '.tiff', '.png', '.jpg', '.jpeg']
    image_files = []
    for ext in image_extensions:
        image_files.extend(group_path.glob(f'*{ext}'))
        image_files.extend(group_path.glob(f'*{ext.upper()}'))
    
    image_files = sorted(set(image_files))
    
    if not image_files:
        print(f"✗ No images found in {group_dir}")
        return []
    
    print(f"Found {len(image_files)} images")
    
    # Process each image with correct arguments
    results = []
    for img_path in image_files:
        result = process_image(
            str(img_path), 
            min_area=50,      # Add explicit min_area
            max_area=10000,   # Add explicit max_area
            output_dir=output_dir
        )
        if result:
            results.append(result)
    
    print(f"\n{'#'*60}")
    print(f"# Group {group_name} Complete: {len(results)}/{len(image_files)} images processed")
    print(f"{'#'*60}")
    
    return results

def style_header(ws, row_num):
    """Apply styling to header row"""
    for cell in ws[row_num]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')


def style_summary_header(cell):
    """Apply styling to summary section header"""
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.alignment = Alignment(horizontal='left')


def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def numpy_to_excel_image(img_array):
    """
    Convert numpy array to Excel-compatible image
    
    Args:
        img_array: Numpy array image
        
    Returns:
        openpyxl Image object
    """
    # Ensure 8-bit
    if img_array.dtype == np.uint16:
        img_array = (img_array / 256).astype(np.uint8)
    
    # Convert BGR to RGB if needed
    if len(img_array.shape) == 3 and img_array.shape[2] == 3:
        img_array = cv2.cvtColor(img_array, cv2.COLOR_BGR2RGB)
    
    # Convert to PIL Image
    pil_img = PILImage.fromarray(img_array)
    
    # Save to BytesIO
    img_buffer = BytesIO()
    pil_img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    # Create Excel Image
    excel_img = ExcelImage(img_buffer)
    
    # Scale down if too large
    max_width = 800
    if excel_img.width > max_width:
        scale = max_width / excel_img.width
        excel_img.width = int(excel_img.width * scale)
        excel_img.height = int(excel_img.height * scale)
    
    return excel_img


def trim_outliers(particle_list, trim_percentage=0.20):
    """
    Remove outliers by trimming top and bottom percentiles
    
    Args:
        particle_list: List of particle dictionaries
        trim_percentage: Percentage to trim from each end (default 0.20 = 20%)
        
    Returns:
        Trimmed list of particles
    """
    if not particle_list:
        return []
    
    # Extract intensity values
    intensities = [p['intensity_per_area_orig'] for p in particle_list]
    
    # Calculate percentiles
    lower_percentile = trim_percentage * 100
    upper_percentile = (1 - trim_percentage) * 100
    
    lower_threshold = np.percentile(intensities, lower_percentile)
    upper_threshold = np.percentile(intensities, upper_percentile)
    
    # Filter particles
    trimmed = [p for p in particle_list 
               if lower_threshold <= p['intensity_per_area_orig'] <= upper_threshold]
    
    print(f"  Original: {len(particle_list)} particles")
    print(f"  Trimmed: {len(trimmed)} particles (removed {len(particle_list) - len(trimmed)})")
    print(f"  Threshold range: {lower_threshold:.2f} - {upper_threshold:.2f}")
    
    return trimmed


def add_column_definitions(ws, start_row, unit_str='um', bit_depth=8):
    """
    Add column definition table to worksheet
    
    Args:
        ws: Worksheet object
        start_row: Starting row for the table
        unit_str: Unit string (um or pixels)
        bit_depth: Bit depth of the image
    """
    # Title
    title_cell = ws.cell(start_row, 1, 'COLUMN DEFINITIONS')
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='left')
    
    start_row += 1
    
    # Table headers
    headers = ['Column', 'Description', 'Calculation', 'Use Case']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    start_row += 1
    
    # Determine max value based on bit depth
    max_value = 2**bit_depth - 1
    
    # Table rows - data definitions
    definitions = [
        [
            f'Area ({unit_str}²)',
            'Physical area of the detected particle in square micrometers',
            f'Calculated from pixel count × pixel size²',
            'Shows particle size; used to filter out noise or debris'
        ],
        [
            f'Perimeter ({unit_str})',
            'Distance around the particle boundary in micrometers',
            f'Sum of boundary edge lengths × pixel size',
            'Indicates particle shape complexity; rounder particles have lower perimeter-to-area ratios'
        ],
        [
            f'Total Intensity ({bit_depth}-bit)',
            'Sum of all pixel brightness values within the particle',
            f'Σ(pixel values) ranging 0-{max_value}',
            'Raw fluorescence signal; depends on both particle size and brightness'
        ],
        [
            'Mean Intensity',
            'Average brightness of pixels within the particle',
            'Total Intensity ÷ Number of pixels',
            'Shows average fluorescence per pixel; independent of particle size'
        ],
        [
            'Std Dev',
            'Standard deviation of pixel intensities',
            '√[Σ(pixel - mean)² ÷ n]',
            'Indicates how uniform the fluorescence is within the particle; low = uniform, high = patchy'
        ],
        [
            f'Intensity per {unit_str}²',
            'Fluorescence intensity normalized by particle area',
            f'Total Intensity ÷ Area ({unit_str}²)',
            '**PRIMARY COMPARISON METRIC**: Accounts for size differences between particles'
        ]
    ]
    
    for row_data in definitions:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(start_row, col_idx, value)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # Make primary metric row bold
            if 'PRIMARY COMPARISON METRIC' in value:
                cell.font = Font(bold=True, size=9)
            else:
                cell.font = Font(size=9)
        start_row += 1
    
    # Key points section
    start_row += 1
    key_title_cell = ws.cell(start_row, 1, 'KEY POINTS:')
    key_title_cell.font = Font(bold=True, size=11, color="FFFFFF")
    key_title_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    start_row += 1
    
    key_points = [
        f'• Intensity per {unit_str}² is the primary metric for comparing Control vs Sample groups',
        '• It normalizes for particle size, so larger particles don\'t artificially inflate measurements',
        '• The trimming process (removing top/bottom 20%) removes outliers from this metric',
        '• The comparison plot shows mean ± SEM of this normalized intensity value',
        '',
        'Example:',
        f'  Particle A: 1000 total intensity, 10 {unit_str}² area → 100 intensity/{unit_str}²',
        f'  Particle B: 2000 total intensity, 20 {unit_str}² area → 100 intensity/{unit_str}²',
        '  Both have same normalized intensity despite different sizes, making them comparable.'
    ]
    
    for point in key_points:
        cell = ws.cell(start_row, 1, point)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if point.startswith('•'):
            cell.font = Font(size=9)
        elif point.startswith('Example:'):
            cell.font = Font(bold=True, size=9)
        else:
            cell.font = Font(size=9, italic=True)
        start_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 50
    
    return start_row


def add_statistical_explanations(ws, start_row, bit_depth=8, unit_str='um'):
    """
    Add statistical metric explanations to worksheet
    
    Args:
        ws: Worksheet object
        start_row: Starting row for explanations
        bit_depth: Bit depth of the image
        unit_str: Unit string (um or pixels)
    """
    # Title
    title_cell = ws.cell(start_row, 1, 'STATISTICAL METRICS EXPLAINED')
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    start_row += 2
    
    # Explanations
    explanations = [
        ['Mean:', 'Average value of all particles in the dataset'],
        ['Std Dev:', 'Standard Deviation - measure of variability/spread in the data'],
        ['SEM:', 'Standard Error of the Mean - uncertainty in the mean estimate (Std Dev ÷ √n)'],
        ['Min/Max:', 'Range of values observed in the dataset'],
        ['', ''],
        ['Trimmed Data:', 'Middle 60% of data after removing top 20% and bottom 20% outliers'],
        ['Purpose:', 'Reduces impact of extreme values and technical artifacts'],
        ['Note:', 'Use trimmed statistics for group comparisons']
    ]
    
    for label, explanation in explanations:
        if label:
            label_cell = ws.cell(start_row, 1, label)
            label_cell.font = Font(bold=True, size=10)
        exp_cell = ws.cell(start_row, 2, explanation)
        exp_cell.font = Font(size=9)
        exp_cell.alignment = Alignment(wrap_text=True)
        start_row += 1
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 60



def save_group_excel(results, group_name, output_dir, apply_trimming=True, trim_percentage=0.20, 
                    comparison_chart=None, is_sample=False):
    """
    Save Excel file for a single group with optional trimming and embedded images
    
    Args:
        results: List of result dictionaries
        group_name: Name of the group
        output_dir: Output directory path
        apply_trimming: Whether to apply outlier trimming
        trim_percentage: Percentage to trim from each end
        comparison_chart: Optional comparison chart image (numpy array) to embed
        is_sample: Whether this is the sample group (for chart placement)
    """
    if not results:
        print(f"⚠ No results to save for {group_name}")
        return None
    
    excel_file = os.path.join(output_dir, f"{group_name}_analysis.xlsx")
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    meta = results[0]['metadata']
    bit_depth = meta['bit_depth']
    unit_str = meta['unit']
    has_calibration = meta['has_calibration']
    
    # Aggregate all particles
    all_particles = []
    for result in results:
        all_particles.extend(result['object_data'])
    
    # Apply trimming if requested
    if apply_trimming and len(all_particles) > 0:
        print(f"\n📊 Trimming outliers for {group_name}:")
        trimmed_particles = trim_outliers(all_particles, trim_percentage)
    else:
        trimmed_particles = all_particles
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    ws_summary.append(['GROUP SUMMARY'])
    style_summary_header(ws_summary.cell(1, 1))
    ws_summary.append(['Group Name:', group_name])
    ws_summary.append(['Total Images:', len(results)])
    ws_summary.append([])
    
    # Statistics - Original data
    ws_summary.append(['ORIGINAL DATA STATISTICS'])
    style_summary_header(ws_summary.cell(5, 1))
    ws_summary.append(['Total Particles:', len(all_particles)])
    
    if all_particles:
        intensities = [p['intensity_per_area_orig'] for p in all_particles]
        ws_summary.append(['Mean Intensity per Area:', f"{np.mean(intensities):.2f}"])
        ws_summary.append(['Std Dev:', f"{np.std(intensities, ddof=1):.2f}"])
        ws_summary.append(['SEM:', f"{np.std(intensities, ddof=1) / np.sqrt(len(intensities)):.2f}"])
        ws_summary.append(['Min:', f"{np.min(intensities):.2f}"])
        ws_summary.append(['Max:', f"{np.max(intensities):.2f}"])
    
    ws_summary.append([])
    
    # Statistics - Trimmed data
    if apply_trimming:
        ws_summary.append(['TRIMMED DATA STATISTICS (Middle 60%)'])
        style_summary_header(ws_summary.cell(ws_summary.max_row, 1))
        ws_summary.append(['Total Particles:', len(trimmed_particles)])
        
        if trimmed_particles:
            trimmed_intensities = [p['intensity_per_area_orig'] for p in trimmed_particles]
            ws_summary.append(['Mean Intensity per Area:', f"{np.mean(trimmed_intensities):.2f}"])
            ws_summary.append(['Std Dev:', f"{np.std(trimmed_intensities, ddof=1):.2f}"])
            ws_summary.append(['SEM:', f"{np.std(trimmed_intensities, ddof=1) / np.sqrt(len(trimmed_intensities)):.2f}"])
            ws_summary.append(['Min:', f"{np.min(trimmed_intensities):.2f}"])
            ws_summary.append(['Max:', f"{np.max(trimmed_intensities):.2f}"])
    
    ws_summary.append([])
    ws_summary.append([])
    
    # Embed comparison chart in SAMPLE Excel file only
    if is_sample and comparison_chart is not None:
        chart_row = ws_summary.max_row + 1
        ws_summary.cell(chart_row, 1, 'COMPARISON CHART')
        style_summary_header(ws_summary.cell(chart_row, 1))
        
        excel_chart = numpy_to_excel_image(comparison_chart)
        ws_summary.add_image(excel_chart, f'A{chart_row + 1}')
        
        # Add extra rows to make space for chart
        for _ in range(30):
            ws_summary.append([])
    
    # Add statistical explanations
    explanation_row = ws_summary.max_row + 2
    add_statistical_explanations(ws_summary, explanation_row, bit_depth, unit_str)
    
    # Add column definitions
    column_def_row = ws_summary.max_row + 3
    add_column_definitions(ws_summary, column_def_row, unit_str, bit_depth)
    
    auto_adjust_column_width(ws_summary)
    
    # Individual image data sheets with embedded contour images
    for result in results:
        img_name = result['image_name']
        object_data = result['object_data']
        contour_img = result['contour_image']
        
        # Create sheet with truncated name
        ws = wb.create_sheet(img_name[:31])
        
        # Add data table
        headers = [
            'Particle ID',
            f'Area ({unit_str}²)',
            f'Perimeter ({unit_str})',
            f'Total Intensity ({bit_depth}-bit)',
            f'Mean Intensity',
            f'Std Dev',
            f'Intensity per {unit_str}²'
        ]
        ws.append(headers)
        style_header(ws, 1)
        
        for obj in object_data:
            ws.append([
                obj['object_id'],
                round(obj['area_physical'], 4 if has_calibration else 2),
                round(obj['perimeter_physical'], 2),
                round(obj['red_total_orig'], 1),
                round(obj['red_mean_orig'], 2),
                round(obj['red_std_orig'], 2),
                round(obj['intensity_per_area_orig'], 2)
            ])
        
        # Determine where to place image (to the right of data table, starting at column I)
        image_column = 'I'
        image_row = 1
        
        # Add image title
        ws[f'{image_column}{image_row}'] = f"Detected Particles: {len(object_data)}"
        ws[f'{image_column}{image_row}'].font = Font(bold=True, size=11)
        
        # Embed contour image
        excel_img = numpy_to_excel_image(contour_img)
        ws.add_image(excel_img, f'{image_column}{image_row + 1}')
        
        # Add column definitions below the data table
        definitions_start_row = len(object_data) + 5  # Leave some space after data
        add_column_definitions(ws, definitions_start_row, unit_str, bit_depth)
        
        auto_adjust_column_width(ws)
    
    wb.save(excel_file)
    print(f"✓ Saved: {excel_file}")
    
    # Return trimmed data for comparison
    return trimmed_particles


def save_comparison_excel(control_data, sample_data, control_name, sample_name, output_dir, unit_str='um'):
    """
    Save comparison Excel with both groups' data and statistics
    
    Args:
        control_data: List of control particle dictionaries
        sample_data: List of sample particle dictionaries
        control_name: Name of control group
        sample_name: Name of sample group
        output_dir: Output directory
        unit_str: Unit string for measurements
    """
    excel_file = os.path.join(output_dir, "Comparison_Analysis.xlsx")
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Get bit depth from first particle (assuming consistent across dataset)
    bit_depth = 8  # default
    if control_data and 'red_total_orig' in control_data[0]:
        # Estimate bit depth from max possible value
        max_val = max([p.get('red_total_orig', 0) for p in control_data + sample_data])
        if max_val > 255:
            bit_depth = 16
    
    # Summary comparison sheet
    ws = wb.create_sheet("Comparison Summary")
    
    ws.append(['CONTROL VS SAMPLE COMPARISON'])
    style_summary_header(ws.cell(1, 1))
    ws.append([])
    
    # Control statistics
    ws.append(['CONTROL GROUP'])
    style_summary_header(ws.cell(3, 1))
    ws.append(['Group Name:', control_name])
    ws.append(['Total Particles:', len(control_data)])
    
    if control_data:
        control_intensities = [p['intensity_per_area_orig'] for p in control_data]
        ws.append(['Mean Intensity per Area:', f"{np.mean(control_intensities):.2f}"])
        ws.append(['Std Dev:', f"{np.std(control_intensities, ddof=1):.2f}"])
        ws.append(['SEM:', f"{np.std(control_intensities, ddof=1) / np.sqrt(len(control_intensities)):.2f}"])
    
    ws.append([])
    
    # Sample statistics
    ws.append(['SAMPLE GROUP'])
    style_summary_header(ws.cell(ws.max_row, 1))
    ws.append(['Group Name:', sample_name])
    ws.append(['Total Particles:', len(sample_data)])
    
    if sample_data:
        sample_intensities = [p['intensity_per_area_orig'] for p in sample_data]
        ws.append(['Mean Intensity per Area:', f"{np.mean(sample_intensities):.2f}"])
        ws.append(['Std Dev:', f"{np.std(sample_intensities, ddof=1):.2f}"])
        ws.append(['SEM:', f"{np.std(sample_intensities, ddof=1) / np.sqrt(len(sample_intensities)):.2f}"])
    
    ws.append([])
    
    # Statistical comparison
    if control_data and sample_data:
        ws.append(['STATISTICAL COMPARISON'])
        style_summary_header(ws.cell(ws.max_row, 1))
        
        control_mean = np.mean(control_intensities)
        sample_mean = np.mean(sample_intensities)
        percent_change = ((sample_mean - control_mean) / control_mean) * 100
        
        ws.append(['Difference (Sample - Control):', f"{sample_mean - control_mean:.2f}"])
        ws.append(['Percent Change:', f"{percent_change:.2f}%"])
        ws.append(['Fold Change:', f"{sample_mean / control_mean:.2f}x"])
    
    # Add column definitions
    definitions_row = ws.max_row + 3
    add_column_definitions(ws, definitions_row, unit_str, bit_depth)
    
    # Add statistical explanations
    stats_row = ws.max_row + 3
    add_statistical_explanations(ws, stats_row, bit_depth, unit_str)
    
    auto_adjust_column_width(ws)
    
    # Detailed data sheets
    ws_control = wb.create_sheet(f"{control_name} Data")
    ws_sample = wb.create_sheet(f"{sample_name} Data")
    
    # Headers
    headers = [
        'Particle ID',
        f'Intensity per {unit_str}²',
        f'Area ({unit_str}²)',
        f'Mean Intensity',
        'Std Dev'
    ]
    
    ws_control.append(headers)
    ws_sample.append(headers)
    style_header(ws_control, 1)
    style_header(ws_sample, 1)
    
    # Control data
    for i, obj in enumerate(control_data, 1):
        ws_control.append([
            i,
            round(obj['intensity_per_area_orig'], 2),
            round(obj['area_physical'], 4),
            round(obj['red_mean_orig'], 2),
            round(obj['red_std_orig'], 2)
        ])
    
    # Sample data
    for i, obj in enumerate(sample_data, 1):
        ws_sample.append([
            i,
            round(obj['intensity_per_area_orig'], 2),
            round(obj['area_physical'], 4),
            round(obj['red_mean_orig'], 2),
            round(obj['red_std_orig'], 2)
        ])
    
    # Add definitions to data sheets
    for ws_data in [ws_control, ws_sample]:
        defs_row = ws_data.max_row + 3
        add_column_definitions(ws_data, defs_row, unit_str, bit_depth)
        auto_adjust_column_width(ws_data)
    
    wb.save(excel_file)
    print(f"✓ Comparison Excel saved: {excel_file}")


def create_comparison_chart(control_data, sample_data, control_name, sample_name, output_path, unit_str='um'):
    """
    Create a bar chart comparing control vs sample with error bars
    
    Args:
        control_data: List of control particle dictionaries
        sample_data: List of sample particle dictionaries
        control_name: Name of the control group
        sample_name: Name of the sample group
        output_path: Path to save the chart
        unit_str: Unit string for axis label
    """
    # Extract intensity values
    control_intensities = [p['intensity_per_area_orig'] for p in control_data] if control_data else []
    sample_intensities = [p['intensity_per_area_orig'] for p in sample_data] if sample_data else []
    
    if not control_intensities and not sample_intensities:
        print("⚠ No data to plot comparison chart")
        return None
    
    # Calculate statistics - convert to float explicitly
    control_mean = float(np.mean(control_intensities)) if control_intensities else 0.0
    control_sem = float(np.std(control_intensities, ddof=1) / np.sqrt(len(control_intensities))) if len(control_intensities) > 1 else 0.0
    
    sample_mean = float(np.mean(sample_intensities)) if sample_intensities else 0.0
    sample_sem = float(np.std(sample_intensities, ddof=1) / np.sqrt(len(sample_intensities))) if len(sample_intensities) > 1 else 0.0
    
    # Create bar chart
    fig, ax = plt.subplots(figsize=(6, 5), dpi=PLOT_DPI)
    
    # Create group labels with names and sample sizes
    control_label = f'{control_name}\n(n={len(control_intensities)})'
    sample_label = f'{sample_name}\n(n={len(sample_intensities)})'
    groups = [control_label, sample_label]
    
    means = np.array([control_mean, sample_mean], dtype=np.float64)
    errors = np.array([control_sem, sample_sem], dtype=np.float64)
    
    # Define colors matching Figure B style
    colors = ['#87CEEB', '#DDA0DD']  # Light blue and light purple
    
    # Create bars
    bars = ax.bar(groups, means, yerr=errors, capsize=10, 
                   color=colors, edgecolor='black', linewidth=1.5,
                   error_kw={'linewidth': 2, 'ecolor': 'black'})
    
    # Styling
    ax.set_ylabel(f'Fluorescence Intensity (Fluor/{unit_str}²)', fontsize=12, fontweight='bold')
    max_value = float(np.max(means + errors)) if len(means) > 0 else 1.0
    ax.set_ylim(0, max_value * 1.2)  # Set y-axis to accommodate error bars
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.tick_params(axis='both', labelsize=11)
    
    # Add grid for better readability
    ax.yaxis.grid(True, linestyle='--', alpha=0.3)
    ax.set_axisbelow(True)
    
    # Add title
    ax.set_title('Average Fluorescence Intensity Comparison', fontsize=13, fontweight='bold', pad=15)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=PLOT_DPI, bbox_inches='tight')
    plt.close()
    
    print(f"✓ Comparison chart saved: {output_path}")
    
    # Return the figure as numpy array for Excel embedding
    fig, ax = plt.subplots(figsize=(6, 5), dpi=PLOT_DPI)
    bars = ax.bar(groups, means, yerr=errors, capsize=10, 
                   color=colors, edgecolor='black', linewidth=1.5,
                   error_kw={'linewidth': 2, 'ecolor': 'black'})
    ax.set_ylabel(f'Fluorescence Intensity (Fluor/{unit_str}²)', fontsize=12, fontweight='bold')
    ax.set_ylim(0, max_value * 1.2)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.tick_params(axis='both', labelsize=11)
    ax.yaxis.grid(True, linestyle='--', alpha=0.3)
    ax.set_axisbelow(True)
    ax.set_title('Average Fluorescence Intensity Comparison', fontsize=13, fontweight='bold', pad=15)
    
    # Convert to numpy array - compatible method
    plt.tight_layout()
    
    # Save to BytesIO buffer
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=PLOT_DPI, bbox_inches='tight')
    buf.seek(0)
    
    # Read as PIL Image then convert to numpy
    pil_img = PILImage.open(buf)
    img_array = np.array(pil_img)
    
    plt.close()
    buf.close()
    
    return img_array


def prompt_for_sample_group(source_dir, control_group):
    """
    Interactively prompt user to select a sample group from available directories
    
    Args:
        source_dir: Source directory path
        control_group: Control group name to exclude from options
        
    Returns:
        Selected sample group name or None if cancelled
    """
    source_path = Path(source_dir)
    
    if not source_path.exists():
        print(f"❌ Source directory does not exist: {source_dir}")
        return None
    
    # Get all subdirectories except control group
    available_groups = [
        d.name for d in source_path.iterdir() 
        if d.is_dir() and d.name != control_group and not d.name.startswith('.')
    ]
    
    if not available_groups:
        print(f"❌ No sample groups found in {source_dir}")
        print(f"   (excluding control group: {control_group})")
        return None
    
    # Display available groups
    print("\n" + "="*60)
    print("AVAILABLE SAMPLE GROUPS:")
    print("="*60)
    for i, group in enumerate(available_groups, 1):
        # Count images in this group
        group_path = source_path / group
        image_count = len([f for f in group_path.iterdir() 
                          if f.suffix.lower() in ['.tif', '.tiff', '.png', '.jpg', '.jpeg']])
        print(f"{i}. {group} ({image_count} images)")
    print("="*60)
    
    # Get user selection
    while True:
        try:
            choice = input("\nSelect sample group number (or 'q' to quit): ").strip()
            
            if choice.lower() == 'q':
                return None
            
            choice_num = int(choice)
            if 1 <= choice_num <= len(available_groups):
                selected = available_groups[choice_num - 1]
                print(f"✓ Selected: {selected}")
                return selected
            else:
                print(f"❌ Please enter a number between 1 and {len(available_groups)}")
        except ValueError:
            print("❌ Invalid input. Please enter a number or 'q' to quit")
        except KeyboardInterrupt:
            print("\n\n❌ Selection cancelled")
            return None


def verify_group_paths(source_dir, control_group, sample_group):
    """
    Verify that group paths exist and contain valid image files
    
    Args:
        source_dir: Source directory path
        control_group: Control group name
        sample_group: Sample group name
        
    Returns:
        True if paths are valid, False otherwise
    """
    source_path = Path(source_dir)
    
    # Check source directory exists
    if not source_path.exists():
        print(f"❌ Source directory does not exist: {source_dir}")
        return False
    
    # Check control group
    control_path = source_path / control_group
    if not control_path.exists():
        print(f"❌ Control group directory does not exist: {control_path}")
        return False
    
    if not control_path.is_dir():
        print(f"❌ Control group path is not a directory: {control_path}")
        return False
    
    # Count images in control group
    control_images = [f for f in control_path.iterdir() 
                     if f.suffix.lower() in ['.tif', '.tiff', '.png', '.jpg', '.jpeg']]
    
    if not control_images:
        print(f"❌ No valid images found in control group: {control_path}")
        print(f"   Supported formats: .tif, .tiff, .png, .jpg, .jpeg")
        return False
    
    print(f"✓ Control group verified: {len(control_images)} images found")
    
    # Check sample group
    sample_path = source_path / sample_group
    if not sample_path.exists():
        print(f"❌ Sample group directory does not exist: {sample_path}")
        return False
    
    if not sample_path.is_dir():
        print(f"❌ Sample group path is not a directory: {sample_path}")
        return False
    
    # Count images in sample group
    sample_images = [f for f in sample_path.iterdir() 
                    if f.suffix.lower() in ['.tif', '.tiff', '.png', '.jpg', '.jpeg']]
    
    if not sample_images:
        print(f"❌ No valid images found in sample group: {sample_path}")
        print(f"   Supported formats: .tif, .tiff, .png, .jpg, .jpeg")
        return False
    
    print(f"✓ Sample group verified: {len(sample_images)} images found")
    
    return True


def main():
    """Main execution function"""
    parser = argparse.ArgumentParser(description='Particle Scout - Advanced Particle Analysis')
    parser.add_argument('--source-dir', '-d', default='source', help='Source directory containing group folders')
    parser.add_argument('--control-group', '-c', default='Control group', help='Control group directory name')
    parser.add_argument('--sample-group', '-s', help='Sample group directory name')
    parser.add_argument('--output-dir', '-o', help='Output directory (default: Downloads/particle_analysis_TIMESTAMP)')
    parser.add_argument('--trim', action='store_true', help='Enable outlier trimming')
    parser.add_argument('--no-trim', action='store_true', help='Disable outlier trimming (default)')
    parser.add_argument('--trim-percentage', type=float, default=0.20,
                       help='Percentage to trim from each end (default: 0.20)')
    
    args = parser.parse_args()
    
    source_dir = args.source_dir
    control_group = args.control_group
    
    # Determine trimming setting
    if args.trim:
        enable_trimming = True
    elif args.no_trim:
        enable_trimming = False
    else:
        enable_trimming = False  # Default: disabled
    
    # Interactive sample group selection if not specified
    if not args.sample_group:
        sample_group = prompt_for_sample_group(source_dir, control_group)
        if sample_group is None:
            print("\n❌ Sample group selection cancelled or failed")
            return
    else:
        sample_group = args.sample_group
    
    # Verify paths exist and contain valid data
    if not verify_group_paths(source_dir, control_group, sample_group):
        print("\n❌ Path verification failed. Please check your directory structure.")
        return
    
    # Create output directory
    if args.output_dir:
        output_dir = Path(args.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        output_dir = create_output_directory()
    
    # Create debug directory
    debug_dir = get_debug_directory()
    
    print("\n" + "="*60)
    print("PARTICLE SCOUT - ANALYSIS STARTED")
    print("="*60)
    print(f"Source directory: {source_dir}")
    print(f"Control group: {control_group}")
    print(f"Sample group: {sample_group}")
    print(f"Output directory: {output_dir}")
    print(f"Debug directory: {debug_dir}")
    print(f"Outlier trimming: {'Enabled' if enable_trimming else 'Disabled'}")
    if enable_trimming:
        print(f"Trim percentage: {args.trim_percentage*100:.0f}% each end")
    print("="*60)
    
    # Process Control group
    control_dir = Path(source_dir) / control_group
    control_results = process_group(str(control_dir), "Control", str(output_dir))  # Changed debug_dir to output_dir
    control_trimmed = None
    if control_results:
        control_trimmed = save_group_excel(control_results, "Control", str(output_dir), 
                                        enable_trimming, args.trim_percentage, 
                                        comparison_chart=None, is_sample=False)

    # Process Sample group
    sample_dir = Path(source_dir) / sample_group
    sample_results = process_group(str(sample_dir), f"Sample_{sample_group}", str(output_dir))  # Changed debug_dir to output_dir

    # Create comparison chart if both groups have data
    if control_trimmed and sample_results:
        unit_str = control_results[0]['metadata']['unit']
        
        # Aggregate sample data first
        all_sample_particles = []
        for result in sample_results:
            all_sample_particles.extend(result['object_data'])
        
        # Apply trimming if enabled
        if enable_trimming and len(all_sample_particles) > 0:
            sample_trimmed_temp = trim_outliers(all_sample_particles, args.trim_percentage)
        else:
            sample_trimmed_temp = all_sample_particles
        
        # Create comparison chart
        chart_path = os.path.join(output_dir, "comparison_chart.png")
        comparison_chart_img = create_comparison_chart(
            control_trimmed, sample_trimmed_temp, 
            control_group, sample_group,
            chart_path, unit_str
        )
    
    # Save sample Excel with embedded comparison chart
    if sample_results:
        sample_trimmed = save_group_excel(
            sample_results, f"Sample_{sample_group}", 
            str(output_dir), enable_trimming, args.trim_percentage,
            comparison_chart=comparison_chart_img, is_sample=True
        )
    
    # Also create standalone comparison Excel
    if control_trimmed and sample_trimmed:
        unit_str = control_results[0]['metadata']['unit']
        save_comparison_excel(
            control_trimmed, sample_trimmed, 
            control_group, sample_group,
            str(output_dir), unit_str
        )
    
    # Print summary
    print(f"\n{'#'*60}")
    print("# ANALYSIS COMPLETE - SUMMARY")
    print(f"{'#'*60}")
    print(f"Control images processed: {len(control_results) if control_results else 0}")
    print(f"Sample images processed: {len(sample_results) if sample_results else 0}")
    
    if control_trimmed:
        control_intensities = [p['intensity_per_area_orig'] for p in control_trimmed]
        print(f"\nControl particles: {len(control_trimmed)}")
        print(f"  Mean: {np.mean(control_intensities):.2f}")
        print(f"  SEM: {np.std(control_intensities, ddof=1) / np.sqrt(len(control_intensities)):.2f}")
    
    if sample_trimmed:
        sample_intensities = [p['intensity_per_area_orig'] for p in sample_trimmed]
        print(f"\nSample particles: {len(sample_trimmed)}")
        print(f"  Mean: {np.mean(sample_intensities):.2f}")
        print(f"  SEM: {np.std(sample_intensities, ddof=1) / np.sqrt(len(sample_intensities)):.2f}")
    
    print(f"\n📁 Excel outputs: {output_dir}")
    print(f"🔍 Debug files: {debug_dir}")
    print(f"📊 Comparison chart embedded in: Sample_{sample_group}_analysis.xlsx")
    print(f"{'#'*60}\n")


if __name__ == "__main__":
    main()