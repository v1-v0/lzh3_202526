# Standard library imports
import atexit
import csv
import json
import logging
import os
import platform
import re
import shutil
import subprocess
import sys
import textwrap
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast

# Third-party data science imports
import numpy as np
import pandas as pd
from scipy import stats as scipy_stats
from tqdm import tqdm

# Computer vision imports
import cv2
from skimage.registration import phase_cross_correlation

# Plotting imports
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

# Excel/Office imports
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.drawing.image import Image as XLImage


# ==================================================
# Unicode-Safe File I/O Functions
# ==================================================

def safe_imread(path: Path, flags: int = cv2.IMREAD_UNCHANGED) -> Optional[np.ndarray]:
    """Read image with Unicode path support on Windows
    
    Args:
        path: Path to image file
        flags: OpenCV imread flags
        
    Returns:
        numpy array of image, or None if failed
    """
    try:
        # Method: Read file as bytes, then decode with OpenCV
        with open(path, 'rb') as f:
            file_bytes = np.asarray(bytearray(f.read()), dtype=np.uint8)
        img = cv2.imdecode(file_bytes, flags)
        
        if img is None:
            print(f"[WARN] cv2.imdecode returned None for {path.name}")
            return None
            
        return img
    except Exception as e:
        print(f"[ERROR] Failed to read image {path.name}: {e}")
        return None


def safe_imwrite(path: Path, img: np.ndarray, params: Optional[list] = None) -> bool:
    """Write image with Unicode path support on Windows
    
    Args:
        path: Path where to save image
        img: Image array to save
        params: Optional OpenCV imwrite parameters
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Get file extension
        ext = path.suffix.lower()
        if not ext:
            ext = '.png'
        
        # Encode image to memory buffer
        if params is None:
            is_success, buffer = cv2.imencode(ext, img)
        else:
            is_success, buffer = cv2.imencode(ext, img, params)
        
        if not is_success:
            print(f"[WARN] cv2.imencode failed for {path.name}")
            return False
        
        # Write buffer to file
        with open(path, 'wb') as f:
            f.write(buffer.tobytes())
        
        return True
    except Exception as e:
        print(f"[ERROR] Failed to write image {path.name}: {e}")
        return False


def safe_xml_parse(xml_path: Path) -> Optional[ET.ElementTree]:
    """Parse XML file with Unicode path support
    
    Args:
        xml_path: Path to XML file
        
    Returns:
        ElementTree object, or None if failed
    """
    try:
        # Read file content first with explicit encoding
        with open(xml_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Parse from string
        root = ET.fromstring(content)
        tree = ET.ElementTree(root)
        
        return tree
    except FileNotFoundError:
        print(f"[WARN] XML file not found: {xml_path.name}")
        return None
    except ET.ParseError as e:
        print(f"[ERROR] XML parse error in {xml_path.name}: {e}")
        return None
    except Exception as e:
        print(f"[ERROR] Failed to read XML {xml_path.name}: {e}")
        return None


def validate_path_encoding(path: Path) -> bool:
    """Check if path can be properly encoded for filesystem
    
    Args:
        path: Path to validate
        
    Returns:
        True if path encoding is valid, False otherwise
    """
    try:
        path_str = str(path.resolve())
        # Try encoding to filesystem encoding
        path_str.encode(sys.getfilesystemencoding())
        return True
    except UnicodeEncodeError as e:
        print(f"[WARN] Path encoding issue: {path}")
        print(f"       {e}")
        return False
    except Exception as e:
        print(f"[WARN] Path validation error: {e}")
        return False
# ==================================================


# Basic logger setup
logger = logging.getLogger("particle_scout")
if not logger.handlers:
    logger.setLevel(logging.INFO)
    _sh = logging.StreamHandler()
    _sh.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
    logger.addHandler(_sh)

csv_paths: list[Path] = []

# ==================================================
# Logging: tee stdout/stderr to a file
# ==================================================
class Tee:
    def __init__(self, *streams):
        self.streams = streams

    def write(self, data: str) -> None:
        for s in self.streams:
            s.write(data)
            s.flush()

    def flush(self) -> None:
        for s in self.streams:
            s.flush()

def get_project_root() -> Path:
    """Get project root (works as script and as .exe)"""
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).resolve().parent
    else:
        return Path(__file__).resolve().parent

_project_root = get_project_root()
_logs_dir = _project_root / "logs"
_logs_dir.mkdir(exist_ok=True)
_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
_script_name = Path(sys.argv[0]).stem
_log_path = _logs_dir / f"run_{_timestamp}_{_script_name}.txt"

# Initialize _log_file with type annotation
_log_file: Optional[Any] = None

try:
    _log_file = open(_log_path, "w", encoding="utf-8")
    sys.stdout = Tee(sys.stdout, _log_file)
    sys.stderr = Tee(sys.stderr, _log_file)
    print(f"Saving output to: {_log_path}")
    print(f"Project root: {_project_root.resolve()}")
    print(f"Running as: {'EXECUTABLE' if getattr(sys, 'frozen', False) else 'SCRIPT'}")
except Exception as e:
    print(f"Warning: Could not set up logging: {e}")


print(f"Saving output to: {_log_path}")
print(f"Project root: {_project_root.resolve()}")
print(f"Running as: {'EXECUTABLE' if getattr(sys, 'frozen', False) else 'SCRIPT'}")

@atexit.register
def _close_log_file() -> None:
    """Close the global log file if it exists."""
    global _log_file  # Now this is OK since we initialized it explicitly
    if _log_file is not None:
        try:
            _log_file.close()
        except Exception:
            pass
        finally:
            _log_file = None

# ==================================================
# Configuration
# ==================================================
SOURCE_DIR = Path("./source")
CONTROL_DIR = None  # Will be set dynamically

# Segment only brightfield channel
IMAGE_GLOB = "*_ch00.tif"

# OUTPUT_DIR will be set dynamically in main()
OUTPUT_DIR: Optional[Path] = None


# Scale bar parameters
SCALE_BAR_LENGTH_UM = 10
SCALE_BAR_HEIGHT = 4
SCALE_BAR_MARGIN = 15
SCALE_BAR_COLOR = (255, 255, 255)
SCALE_BAR_BG_COLOR = (0, 0, 0)
SCALE_BAR_TEXT_COLOR = (255, 255, 255)
SCALE_BAR_FONT_SCALE = 0.5
SCALE_BAR_FONT_THICKNESS = 1

# --- Segmentation ---
GAUSSIAN_SIGMA = 15
MORPH_KERNEL_SIZE = 3
MORPH_ITERATIONS = 1
DILATE_ITERATIONS = 1
ERODE_ITERATIONS = 1

# Filtering (in micrometers)
MIN_AREA_UM2 = 3.0
MAX_AREA_UM2 = 2000.0
MIN_CIRCULARITY = 0.0
MAX_FRACTION_OF_IMAGE_AREA = 0.25

# --- Fluorescence segmentation (S2) ---
FLUOR_GAUSSIAN_SIGMA = 1.5
FLUOR_MORPH_KERNEL_SIZE = 3
FLUOR_MIN_AREA_UM2 = 3.0
FLUOR_MATCH_MIN_INTERSECTION_PX = 5.0

# Debug options
CLEAR_OUTPUT_DIR_EACH_RUN = True
SEPARATE_OUTPUT_BY_GROUP = True
FALLBACK_UM_PER_PX: Optional[float] = 0.109492

# ==================================================
# Bacteria-Specific Profiles
# ==================================================
BACTERIA_PROFILES = {
    'proteus_mirabilis': {
        'microgel_type': 'negative',
        'name': 'Proteus mirabilis',
        'gram_stain': 'negative',
        'description': 'Highly motile, urease-positive, swarming behavior',
        # Morphology considerations
        'typical_size_um': (0.5, 2.0),  # Width x length range
        'expected_shape': 'rod',
        # Detection parameters
        'area_adjustment_factor': 1.2,  # Swarming may create larger clusters
        'threshold_sensitivity': 0.95,  # Standard sensitivity
        'min_confidence_score': 0.7,
    },
    'klebsiella_pneumoniae': {
        'microgel_type': 'negative',
        'name': 'Klebsiella pneumoniae',
        'gram_stain': 'negative',
        'description': 'Encapsulated, non-motile, lactose-fermenting',
        # Morphology considerations
        'typical_size_um': (0.5, 2.0),
        'expected_shape': 'rod',
        # Detection parameters - capsule may shield binding
        'area_adjustment_factor': 1.3,  # Thick capsule
        'threshold_sensitivity': 0.90,  # Slightly reduced for capsule
        'min_confidence_score': 0.65,
    },
    'streptococcus_mitis': {
        'microgel_type': 'positive',
        'name': 'Streptococcus mitis',
        'gram_stain': 'positive',
        'description': 'Alpha-hemolytic, cocci in chains',
        # Morphology considerations
        'typical_size_um': (0.5, 1.0),  # Smaller cocci
        'expected_shape': 'coccus',
        # Detection parameters
        'area_adjustment_factor': 0.8,  # Smaller cells
        'threshold_sensitivity': 1.0,  # Standard sensitivity
        'min_confidence_score': 0.7,
    },
    'default': {
        'microgel_type': 'negative',
        'name': 'Unknown bacteria',
        'gram_stain': 'unknown',
        'description': 'Generic detection profile',
        'typical_size_um': (0.5, 2.0),
        'expected_shape': 'unknown',
        'area_adjustment_factor': 1.0,
        'threshold_sensitivity': 1.0,
        'min_confidence_score': 0.7,
    }
}

# Mapping for user-friendly selection
BACTERIA_DISPLAY_OPTIONS = [
    ('proteus_mirabilis', '1', 'Proteus mirabilis (G- rod, highly motile)'),
    ('klebsiella_pneumoniae', '2', 'Klebsiella pneumoniae (G- rod, encapsulated)'),
    ('streptococcus_mitis', '3', 'Streptococcus mitis (G+ coccus, chains)'),
    ('default', '4', 'Unknown/Mixed bacteria (use adaptive detection)'),
]


# ==================================================
# Helper Functions
# ==================================================
def normalize_group_folder_name(folder_name: str) -> str:
    """Standardize group folder naming"""
    name = folder_name.strip()
    
    # Control group: always "Control"
    if name.lower() == 'control':
        return "Control"
    
    # Numbered groups: keep as-is
    return name


def logged_input(prompt: str) -> str:
    """Input function that logs both prompt and user response"""
    print(prompt, end='', flush=True)
    user_input = input()
    
    if user_input.strip():
        print(user_input)
    else:
        print("(pressed Enter)")
    
    return user_input


def clear_output_dir(folder: Path) -> None:
    for p in folder.glob("*"):
        try:
            if p.is_file():
                p.unlink()
            elif p.is_dir():
                for q in p.rglob("*"):
                    try:
                        if q.is_file():
                            q.unlink()
                    except Exception:
                        pass
        except Exception:
            pass


def add_scale_bar(
    img: np.ndarray, pixel_size: float, unit: str = "um", length_um: float = 10
) -> np.ndarray:
    """Add a scale bar to the image"""
    if pixel_size is None or pixel_size <= 0:
        return img

    bar_length_px = int(round(length_um / pixel_size))
    if bar_length_px < 10:
        return img

    h, w = img.shape[:2]
    bar_x = w - bar_length_px - SCALE_BAR_MARGIN
    bar_y = h - SCALE_BAR_HEIGHT - SCALE_BAR_MARGIN

    if bar_x < 0 or bar_y < 0:
        return img

    label = (
        f"{int(length_um)} um"
        if unit in ["µm", "um"]
        else f"{int(length_um)} {unit}"
    )

    font = cv2.FONT_HERSHEY_SIMPLEX
    (text_w, text_h), baseline = cv2.getTextSize(
        label, font, SCALE_BAR_FONT_SCALE, SCALE_BAR_FONT_THICKNESS
    )

    text_x = bar_x + (bar_length_px - text_w) // 2
    text_y = bar_y - 8

    bg_padding = 5
    bg_x1 = min(bar_x, text_x) - bg_padding
    bg_y1 = text_y - text_h - bg_padding
    bg_x2 = max(bar_x + bar_length_px, text_x + text_w) + bg_padding
    bg_y2 = bar_y + SCALE_BAR_HEIGHT + bg_padding

    img = img.copy()
    overlay = img.copy()
    cv2.rectangle(overlay, (bg_x1, bg_y1), (bg_x2, bg_y2), SCALE_BAR_BG_COLOR, -1)
    cv2.addWeighted(overlay, 0.7, img, 0.3, 0, img)

    cv2.rectangle(
        img,
        (bar_x, bar_y),
        (bar_x + bar_length_px, bar_y + SCALE_BAR_HEIGHT),
        SCALE_BAR_COLOR,
        -1,
    )

    cv2.putText(
        img,
        label,
        (text_x, text_y),
        font,
        SCALE_BAR_FONT_SCALE,
        SCALE_BAR_TEXT_COLOR,
        SCALE_BAR_FONT_THICKNESS,
        cv2.LINE_AA,
    )

    return img


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def save_debug(
    folder: Path,
    name: str,
    img: np.ndarray,
    pixel_size_um: Optional[float] = None,
) -> None:
    """Save debug image with optional scale bar - memory optimized, Unicode-safe
    
    Args:
        folder: Output folder
        name: Image filename
        img: Image array to save
        pixel_size_um: Optional pixel size for scale bar
    """
    out = folder / name
    
    if pixel_size_um is not None and pixel_size_um > 0:
        img_to_save = add_scale_bar(
            img.copy(),
            float(pixel_size_um), "um", SCALE_BAR_LENGTH_UM
        )
    else:
        img_to_save = img
    
    # Use Unicode-safe write
    success = safe_imwrite(out, img_to_save)
    
    if not success:
        print(f"[ERROR] Failed to save debug image: {name}")
    
    # Clean up large images
    if img_to_save.nbytes > 10_000_000:
        del img_to_save



def list_sample_group_folders(source_dir: Path) -> list[Path]:
    groups: list[Path] = []
    if not source_dir.exists():
        raise FileNotFoundError(f"Source folder not found: {source_dir.resolve()}")

    for p in source_dir.iterdir():
        if not p.is_dir():
            continue
        if p.name.lower().startswith("control"):
            continue
        if re.fullmatch(r"\d+", p.name):
            groups.append(p)

    groups.sort(key=lambda x: int(x.name))
    return groups


def _display_group_name(name: str) -> str:
    return "Control" if name.lower().startswith("control") else name


def _group_order_key(g: str) -> tuple[int, int]:
    """Sort numeric groups ascending, Control last."""
    if g == "Control":
        return (1, 10**9)
    if g.isdigit():
        return (0, int(g))
    return (0, 10**8)


def find_metadata_paths(img_path: Path) -> tuple[Optional[Path], Optional[Path]]:
    base = img_path.stem
    if base.endswith("_ch00"):
        base = base[:-5]
    md_dir = img_path.parent / "MetaData"
    xml_main = md_dir / f"{base}.xml"
    xml_props = md_dir / f"{base}_Properties.xml"

    return (
        xml_props if xml_props.exists() else None,
        xml_main if xml_main.exists() else None,
    )


def _require_attr(elem: ET.Element, attr: str, context: str) -> str:
    v = elem.get(attr)
    if v is None:
        raise ValueError(f"Missing attribute '{attr}' in {context}")
    return v


def _parse_float(s: str) -> float:
    return float(s.strip().replace(",", "."))


def get_pixel_size_um(
    xml_props_path: Optional[Path],
    xml_main_path: Optional[Path],
) -> Tuple[float, float]:
    """Extract pixel size with detailed error reporting and Unicode support
    
    Args:
        xml_props_path: Path to Properties XML file
        xml_main_path: Path to main XML file
        
    Returns:
        Tuple of (pixel_size_x, pixel_size_y) in micrometers
        
    Raises:
        ValueError: If pixel size cannot be determined
    """
    
    errors = []
    
    # Try Properties XML first
    if xml_props_path is not None:
        try:
            tree = safe_xml_parse(xml_props_path)
            if tree is None:
                raise ValueError(f"Could not parse {xml_props_path.name}")
            
            root = tree.getroot()
            if root is None:
                raise ValueError(f"Empty XML document: {xml_props_path.name}")

            dims = root.findall(".//ImageDescription/Dimensions/DimensionDescription")
            by_id = {d.get("DimID"): d for d in dims}

            def read_dim(dim_id: str) -> Tuple[float, int, str]:
                d = by_id.get(dim_id)
                if d is None:
                    raise ValueError(
                        f"Missing DimensionDescription with DimID='{dim_id}' in {xml_props_path.name}"
                    )

                length_s = _require_attr(
                    d, "Length", f"{xml_props_path.name} DimID={dim_id}"
                )
                n_s = _require_attr(
                    d, "NumberOfElements", f"{xml_props_path.name} DimID={dim_id}"
                )
                unit = _require_attr(d, "Unit", f"{xml_props_path.name} DimID={dim_id}")

                length = _parse_float(length_s)
                n = int(n_s)
                return length, n, unit

            x_len, x_n, x_unit = read_dim("X")
            y_len, y_n, y_unit = read_dim("Y")

            if x_unit != "µm" or y_unit != "µm":
                raise ValueError(
                    f"Unexpected units in {xml_props_path.name}: X={x_unit}, Y={y_unit}"
                )

            return float(x_len / x_n), float(y_len / y_n)
        except Exception as e:
            errors.append(f"Properties XML ({xml_props_path.name}): {e}")

    # Try main XML
    if xml_main_path is not None:
        try:
            tree = safe_xml_parse(xml_main_path)
            if tree is None:
                raise ValueError(f"Could not parse {xml_main_path.name}")
            
            root = tree.getroot()
            if root is None:
                raise ValueError(f"Empty XML document: {xml_main_path.name}")

            dims = root.findall(".//ImageDescription/Dimensions/DimensionDescription")
            by_id = {d.get("DimID"): d for d in dims}

            def read_dim(dim_id: str) -> Tuple[float, int, str]:
                d = by_id.get(dim_id)
                if d is None:
                    raise ValueError(
                        f"Missing DimensionDescription with DimID='{dim_id}' in {xml_main_path.name}"
                    )

                length_s = _require_attr(
                    d, "Length", f"{xml_main_path.name} DimID={dim_id}"
                )
                n_s = _require_attr(
                    d, "NumberOfElements", f"{xml_main_path.name} DimID={dim_id}"
                )
                unit = _require_attr(d, "Unit", f"{xml_main_path.name} DimID={dim_id}")

                length = _parse_float(length_s)
                n = int(n_s)
                return length, n, unit

            x_len_m, x_n, x_unit = read_dim("1")
            y_len_m, y_n, y_unit = read_dim("2")

            if x_unit != "m" or y_unit != "m":
                raise ValueError(
                    f"Unexpected units in {xml_main_path.name}: X={x_unit}, Y={y_unit}"
                )

            return float((x_len_m * 1e6) / x_n), float((y_len_m * 1e6) / y_n)
        except Exception as e:
            errors.append(f"Main XML ({xml_main_path.name}): {e}")
    
    error_summary = "\n  - ".join(errors) if errors else "No XML files provided"
    raise ValueError(
        f"Could not determine pixel size (µm/px).\nAttempted sources:\n  - {error_summary}"
    )


def contour_perimeter_um(contour: np.ndarray, um_per_px_x: float, um_per_px_y: float) -> float:
    pts = contour.reshape(-1, 2).astype(np.float64)
    pts[:, 0] *= float(um_per_px_x)
    pts[:, 1] *= float(um_per_px_y)
    d = np.diff(np.vstack([pts, pts[0]]), axis=0)
    seg = np.sqrt((d[:, 0] ** 2) + (d[:, 1] ** 2))
    return float(seg.sum())


def equivalent_diameter_from_area(area: float) -> float:
    return float(2.0 * np.sqrt(area / np.pi)) if area > 0 else 0.0


def normalize_to_8bit(img: np.ndarray) -> np.ndarray:
    if img.dtype == np.uint8:
        return img
    if img.dtype == np.uint16:
        out = np.zeros_like(img, dtype=np.uint8)
        cv2.normalize(img, out, 0, 255, cv2.NORM_MINMAX)
        return out
    img_f = img.astype(np.float32)
    mn, mx = float(np.min(img_f)), float(np.max(img_f))
    if mx <= mn:
        return np.zeros(img.shape, dtype=np.uint8)
    return ((img_f - mn) * (255.0 / (mx - mn))).clip(0, 255).astype(np.uint8)


def _put_text_outline(
    img: np.ndarray,
    text: str,
    org: tuple[int, int],
    font_scale: float = 0.5,
    color: tuple[int, int, int] = (255, 255, 255),
    thickness: int = 1,
) -> None:
    """Draw readable text with a black outline."""
    cv2.putText(
        img,
        text,
        org,
        cv2.FONT_HERSHEY_SIMPLEX,
        font_scale,
        (0, 0, 0),
        thickness + 2,
        cv2.LINE_AA,
    )
    cv2.putText(
        img,
        text,
        org,
        cv2.FONT_HERSHEY_SIMPLEX,
        font_scale,
        color,
        thickness,
        cv2.LINE_AA,
    )


def draw_object_ids(
    img_bgr: np.ndarray, contours: list[np.ndarray], labels: Optional[list[str]] = None
) -> np.ndarray:
    """Draw object labels at contour centroids."""
    out = img_bgr.copy()
    for i, c in enumerate(contours, 1):
        M = cv2.moments(c)
        if M["m00"] == 0:
            continue
        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])
        text = labels[i - 1] if (labels is not None and i - 1 < len(labels)) else str(i)
        _put_text_outline(out, text, (cx, cy), font_scale=0.5, color=(0, 255, 0), thickness=1)
    return out


def _ids_name(original_png: str) -> str:
    """Insert '_ids' before '.png'."""
    if original_png.lower().endswith(".png"):
        return original_png[:-4] + "_ids.png"
    return original_png + "_ids"


def save_debug_ids(
    folder: Path,
    original_name: str,
    img_bgr: np.ndarray,
    accepted_contours: list[np.ndarray],
    object_ids: list[str],
    pixel_size_um: Optional[float] = None,
) -> None:
    """Save a labeled (Object_ID) version of an existing debug view."""
    labeled = draw_object_ids(img_bgr, accepted_contours, labels=object_ids)
    save_debug(folder, _ids_name(original_name), labeled, pixel_size_um)


# ==================================================
# Adaptive Thresholding System
# ==================================================
def calculate_adaptive_thresholds(
    control_data: pd.DataFrame,
    bacteria_profile: str = "default",
    confidence_level: float = 0.95
) -> Dict[str, Any]:
    """
    Calculate adaptive detection thresholds based on control statistics
    and bacteria characteristics.
    
    Args:
        control_data: DataFrame with Fluor_Density_per_BF_Area column
        bacteria_profile: Key from BACTERIA_PROFILES
        confidence_level: Statistical confidence (0.90, 0.95, or 0.99)
        
    Returns:
        Dictionary with threshold values and metadata
    """
    
    if 'Fluor_Density_per_BF_Area' not in control_data.columns:
        raise ValueError("Control data missing Fluor_Density_per_BF_Area column")
    
    # Get profile
    profile = BACTERIA_PROFILES.get(bacteria_profile, BACTERIA_PROFILES['default'])
    
    # Extract clean data
    values = pd.to_numeric(
        control_data['Fluor_Density_per_BF_Area'], 
        errors='coerce'
    ).dropna()
    
    if len(values) < 3:
        print(f"[WARN] Insufficient control data (n={len(values)}), using fallback")
        return {
            'threshold_method': 'fallback',
            'threshold_value': 0.0,
            'confidence': 'low',
            'use_default': True
        }
    
    # Calculate base statistics
    control_mean = float(values.mean())
    control_std = float(values.std(ddof=1))
    control_median = float(values.median())
    
    # Confidence interval multipliers
    z_scores = {0.90: 1.645, 0.95: 1.96, 0.99: 2.576}
    z = z_scores.get(confidence_level, 1.96)
    
    # Method 1: Standard deviation based (robust to outliers)
    threshold_std = control_mean - (z * control_std)
    
    # Method 2: Percentile based (non-parametric)
    threshold_percentile = float(values.quantile(0.05))
    
    # Method 3: Interquartile range (IQR) based (very robust)
    q1 = float(values.quantile(0.25))
    q3 = float(values.quantile(0.75))
    iqr = q3 - q1
    threshold_iqr = q1 - (1.5 * iqr)
    
    # Method 4: Bacteria-specific adjustment
    sensitivity_factor = profile.get('threshold_sensitivity', 1.0)
    threshold_bacteria = control_mean * (1 - (0.05 / sensitivity_factor))
    
    # Ensemble decision: use most conservative (lowest) threshold
    candidate_thresholds = {
        'std_based': threshold_std,
        'percentile_based': threshold_percentile,
        'iqr_based': threshold_iqr,
        'bacteria_adjusted': threshold_bacteria,
    }
    
    # Select final threshold (most conservative)
    final_threshold = min(candidate_thresholds.values())
    winning_method = min(candidate_thresholds.items(), key=lambda kv: kv[1])[0]
    
    # Calculate confidence score
    threshold_spread = max(candidate_thresholds.values()) - min(candidate_thresholds.values())
    spread_ratio = threshold_spread / control_mean if control_mean > 0 else 1.0
    
    if spread_ratio < 0.1:
        confidence = 'high'
    elif spread_ratio < 0.3:
        confidence = 'medium'
    else:
        confidence = 'low'
    
    # Compile results
    result = {
        'threshold_method': winning_method,
        'threshold_value': final_threshold,
        'control_mean': control_mean,
        'control_std': control_std,
        'control_median': control_median,
        'control_n': len(values),
        'confidence': confidence,
        'confidence_level': confidence_level,
        'bacteria_profile': bacteria_profile,
        'bacteria_name': profile['name'],
        'all_thresholds': candidate_thresholds,
        'threshold_spread': threshold_spread,
        'use_default': False,
    }
    
    return result


def apply_adaptive_classification(
    group_mean: float,
    threshold_result: Dict[str, Any],
    bacteria_profile: str = "default"
) -> Dict[str, Any]:
    """
    Classify a group using adaptive thresholds with confidence scoring.
    
    Args:
        group_mean: Mean fluorescence density for the group
        threshold_result: Output from calculate_adaptive_thresholds
        bacteria_profile: Bacteria profile key
        
    Returns:
        Dictionary with classification and confidence
    """
    
    profile = BACTERIA_PROFILES.get(bacteria_profile, BACTERIA_PROFILES['default'])
    threshold = threshold_result['threshold_value']
    control_mean = threshold_result['control_mean']
    control_std = threshold_result['control_std']
    
    # Calculate distance from control in standard deviations
    if control_std > 0:
        z_score = (group_mean - control_mean) / control_std
    else:
        z_score = 0.0
    
    # Calculate confidence score
    distance_from_threshold = abs(group_mean - threshold)
    normalized_distance = distance_from_threshold / control_mean if control_mean > 0 else 0.0
    
    # Confidence increases with distance from threshold
    if normalized_distance > 0.2:
        detection_confidence = 'high'
        confidence_score = min(0.95, 0.7 + normalized_distance)
    elif normalized_distance > 0.1:
        detection_confidence = 'medium'
        confidence_score = min(0.85, 0.6 + normalized_distance)
    else:
        detection_confidence = 'low'
        confidence_score = min(0.75, 0.5 + normalized_distance)
    
    # Classification decision
    if group_mean < threshold:
        # Below threshold = bacteria detected
        if profile['microgel_type'] == 'negative':
            classification = "NEGATIVE"
            interpretation = f"Gram-negative bacteria detected ({profile['name']})"
        else:
            classification = "POSITIVE"
            interpretation = f"Gram-positive bacteria detected ({profile['name']})"
    else:
        # Above threshold = no bacteria
        if profile['microgel_type'] == 'negative':
            classification = "POSITIVE/No obvious bacteria"
            interpretation = "No obvious bacteria"
        else:
            classification = "NEGATIVE/No obvious bacteria"
            interpretation = "No obvious bacteria"
    
    return {
        'classification': classification,
        'interpretation': interpretation,
        'confidence': detection_confidence,
        'confidence_score': confidence_score,
        'z_score': z_score,
        'distance_from_threshold': distance_from_threshold,
        'group_mean': group_mean,
        'threshold_used': threshold,
    }




# ==================================================
# Fluorescence Registration / Alignment
# ==================================================
def align_fluorescence_channel(bf_img: np.ndarray, fluor_img: np.ndarray) -> tuple[np.ndarray, tuple[float, float]]:
    """Aligns the fluorescence image to the brightfield image using Phase Correlation."""
    if bf_img.ndim == 3:
        bf_gray = cv2.cvtColor(bf_img, cv2.COLOR_BGR2GRAY)
    else:
        bf_gray = bf_img

    if fluor_img.ndim == 3:
        fluor_gray = cv2.cvtColor(fluor_img, cv2.COLOR_BGR2GRAY)
    else:
        fluor_gray = fluor_img

    bf_inverted = cv2.bitwise_not(bf_gray)

    shift, error, diffphase = phase_cross_correlation(bf_inverted, fluor_gray, upsample_factor=10)
    shift_y, shift_x = shift

    rows, cols = fluor_img.shape[:2]
    
    M = np.array([[1, 0, -shift_x], [0, 1, -shift_y]], dtype=np.float32)
    
    aligned_fluor = cv2.warpAffine(fluor_img, M, (cols, rows))

    return aligned_fluor, (shift_y, shift_x)


# ==================================================
# Fluorescence segmentation + matching
# ==================================================
def segment_fluorescence_global(fluor_img8: np.ndarray) -> np.ndarray:
    """Segment fluorescence objects globally. Returns binary mask uint8 (0/255)."""
    blur = cv2.GaussianBlur(
        fluor_img8, (0, 0), sigmaX=FLUOR_GAUSSIAN_SIGMA, sigmaY=FLUOR_GAUSSIAN_SIGMA
    )
    
    otsu_threshold = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[0]
    
    THRESHOLD_MULTIPLIER = 0.5
    adjusted_threshold = otsu_threshold * THRESHOLD_MULTIPLIER
    
    print(f"  Fluorescence threshold: Otsu={otsu_threshold:.1f}, Adjusted={adjusted_threshold:.1f}")
    
    _, bw = cv2.threshold(blur, adjusted_threshold, 255, cv2.THRESH_BINARY)
    
    k = np.ones((FLUOR_MORPH_KERNEL_SIZE, FLUOR_MORPH_KERNEL_SIZE), np.uint8)
    bw = cv2.morphologyEx(bw, cv2.MORPH_OPEN, k, iterations=1)
    bw = cv2.morphologyEx(bw, cv2.MORPH_CLOSE, k, iterations=1)
    return bw


def contour_intersection_area_px(c1: np.ndarray, c2: np.ndarray, shape_hw: tuple[int, int]) -> float:
    """Intersection area in pixels between two contours."""
    H, W = shape_hw
    m1 = np.zeros((H, W), dtype=np.uint8)
    m2 = np.zeros((H, W), dtype=np.uint8)
    cv2.drawContours(m1, [c1], -1, 255, thickness=-1)
    cv2.drawContours(m2, [c2], -1, 255, thickness=-1)
    return float(np.count_nonzero(cv2.bitwise_and(m1, m2)))


def match_fluor_to_bf_by_overlap(
    bf_contours: list[np.ndarray],
    fluor_contours: list[np.ndarray],
    img_shape_hw: tuple[int, int],
    min_intersection_px: float = FLUOR_MATCH_MIN_INTERSECTION_PX,
) -> list[Optional[int]]:
    """For each BF contour, pick the fluorescence contour that maximizes overlap."""
    matches: list[Optional[int]] = []
    fluor_boxes = [cv2.boundingRect(c) for c in fluor_contours]

    for bf in bf_contours:
        bx, by, bw, bh = cv2.boundingRect(bf)

        best_idx: Optional[int] = None
        best_inter = 0.0

        for j, (fx, fy, fw, fh) in enumerate(fluor_boxes):
            if (bx + bw < fx) or (fx + fw < bx) or (by + bh < fy) or (fy + fh < by):
                continue

            inter = contour_intersection_area_px(bf, fluor_contours[j], img_shape_hw)
            if inter > best_inter:
                best_inter = inter
                best_idx = j

        if best_idx is not None and best_inter >= float(min_intersection_px):
            matches.append(best_idx)
        else:
            matches.append(None)

    return matches


def measure_fluorescence_intensity_with_global_area(
    fluor_img: np.ndarray,
    bf_contours: list[np.ndarray],
    fluor_contours: list[np.ndarray],
    bf_to_fluor_match: list[Optional[int]],
    um_per_px_x: float,
    um_per_px_y: float,
) -> list[dict]:
    """Intensity stats: within BF contour. Fluor area: from matched fluorescence contour."""
    um2_per_px2 = float(um_per_px_x) * float(um_per_px_y)
    measurements: list[dict] = []

    for i, bf in enumerate(bf_contours, 1):
        bf_mask = np.zeros(fluor_img.shape[:2], dtype=np.uint8)
        cv2.drawContours(bf_mask, [bf], -1, 255, thickness=-1)
        fluor_values = fluor_img[bf_mask > 0]

        j = bf_to_fluor_match[i - 1]
        if j is not None:
            s2_area_px = float(cv2.contourArea(fluor_contours[j]))
        else:
            s2_area_px = 0.0
        s2_area_um2 = s2_area_px * um2_per_px2

        if fluor_values.size > 0:
            measurements.append(
                {
                    "object_id": i,
                    "fluor_area_px": s2_area_px,
                    "fluor_area_um2": s2_area_um2,
                    "fluor_mean": float(np.mean(fluor_values)),
                    "fluor_median": float(np.median(fluor_values)),
                    "fluor_std": float(np.std(fluor_values)),
                    "fluor_min": float(np.min(fluor_values)),
                    "fluor_max": float(np.max(fluor_values)),
                    "fluor_integrated_density": float(np.sum(fluor_values)),
                }
            )
        else:
            measurements.append(
                {
                    "object_id": i,
                    "fluor_area_px": s2_area_px,
                    "fluor_area_um2": s2_area_um2,
                    "fluor_mean": 0.0,
                    "fluor_median": 0.0,
                    "fluor_std": 0.0,
                    "fluor_min": 0.0,
                    "fluor_max": 0.0,
                    "fluor_integrated_density": 0.0,
                }
            )

    return measurements


def visualize_fluorescence_measurements(
    fluor_img8: np.ndarray, contours: list[np.ndarray], measurements: list[dict]
) -> np.ndarray:
    """Create visualization with contours and intensity labels"""
    vis = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(vis, contours, -1, (0, 255, 0), 1)

    for m in measurements:
        c = contours[m["object_id"] - 1]
        M = cv2.moments(c)
        if M["m00"] != 0:
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])

            label = f"{m['object_id']}: {m['fluor_mean']:.0f}"
            cv2.putText(
                vis,
                label,
                (cx, cy),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.4,
                (255, 255, 0),
                1,
                cv2.LINE_AA,
            )

    return vis


# ==================================================
# Segmentation
# ==================================================
def segment_particles_brightfield(img8: np.ndarray, pixel_size_um: float, out_dir: Path) -> np.ndarray:
    """Brightfield segmentation: dark objects on light background"""
    bg = cv2.GaussianBlur(img8, (0, 0), sigmaX=GAUSSIAN_SIGMA, sigmaY=GAUSSIAN_SIGMA)
    enhanced = cv2.subtract(bg, img8)
    enhanced_blur = cv2.GaussianBlur(enhanced, (3, 3), 0)

    save_debug(out_dir, "02_enhanced.png", enhanced, pixel_size_um)
    save_debug(out_dir, "03_enhanced_blur.png", enhanced_blur, pixel_size_um)

    _, thresh = cv2.threshold(enhanced_blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    save_debug(out_dir, "04_thresh_raw.png", thresh, pixel_size_um)

    kernel = np.ones((MORPH_KERNEL_SIZE, MORPH_KERNEL_SIZE), np.uint8)
    bw = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=MORPH_ITERATIONS)
    bw = cv2.dilate(bw, kernel, iterations=DILATE_ITERATIONS)
    bw = cv2.erode(bw, kernel, iterations=ERODE_ITERATIONS)
    save_debug(out_dir, "05_closed.png", bw, pixel_size_um)

    print(f"Mask white fraction (final): {float((bw > 0).mean()):.4f}")
    return bw


# ==================================================
# Plotting and Analysis
# ==================================================
def generate_error_bar_comparison_with_threshold(
    output_dir: Path,
    percentile: float = 0.2,
    restrict_to_groups: Optional[list[str]] = None,
    output_path: Optional[Path] = None,
    title_suffix: str = "",
    dataset_id: str = "",
    threshold_pct: float = 0.05,
    microgel_type: str = "negative",
) -> Optional[Path]:
    """Enhanced version with threshold lines and control mean."""
    
    excel_files = list(output_dir.rglob("*_master.xlsx"))

    if not excel_files:
        print(f"[INFO] No master Excel files found under {output_dir}")
        return None

    all_data_rows: list[dict[str, object]] = []
    group_stats: dict[str, dict[str, float | int]] = {}
    control_mean = None

    # Load data
    for excel_path in sorted(excel_files):
        group_name_raw = excel_path.stem.replace("_master", "")
        display_name = _display_group_name(group_name_raw)

        if restrict_to_groups is not None and display_name not in restrict_to_groups:
            continue

        try:
            typical_sheet = f"{group_name_raw}_Typical_Particles"
            df = pd.read_excel(excel_path, sheet_name=typical_sheet)

            if "Fluor_Density_per_BF_Area" not in df.columns:
                print(f"[WARN] Missing Fluor_Density_per_BF_Area column in {excel_path.name}")
                continue

            values = df["Fluor_Density_per_BF_Area"].dropna()
            
            if len(values) == 0:
                print(f"[WARN] No valid data in {group_name_raw}")
                continue

            for v in values:
                all_data_rows.append(
                    {"Group": display_name, "Fluorescence Density": float(np.asarray(v).item())}
                )

            mean_val = float(np.asarray(values.mean()).item())
            std_val = float(np.asarray(values.std()).item())
            sem_val = float(np.asarray(values.sem()).item())

            group_stats[display_name] = {
                "n": int(len(values)),
                "mean": mean_val,
                "std": std_val,
                "sem": sem_val,
            }

            if display_name == "Control":
                control_mean = mean_val

            print(f"Loaded {len(values)} points for group: {display_name}")

        except Exception as e:
            print(f"[WARN] Could not read {group_name_raw}: {e}")
            continue

    if not all_data_rows:
        print("[WARN] No valid data found for comparison")
        return None

    df_all = pd.DataFrame(all_data_rows)

    group_order: list[str] = sorted(
        df_all["Group"].dropna().astype(str).drop_duplicates().tolist(), 
        key=_group_order_key
    )

    if len(group_order) < 1:
        print("[INFO] Need at least 1 group with data to generate plot.")
        return None

    # Color palette
    palette_colors: list[Any] = ["silver", "violet"]
    if len(group_order) > 2:
        palette_colors = list(sns.color_palette("husl", len(group_order)))
    elif len(group_order) == 1:
        palette_colors = ["skyblue"]

    # Generate plot
    plt.figure(figsize=(10, 7))
    sns.set_style("ticks")

    try:
        ax = sns.barplot(
            data=df_all,
            x="Group",
            y="Fluorescence Density",
            hue="Group",
            order=group_order,
            hue_order=group_order,
            palette=palette_colors,
            legend=False,
            errorbar=None,
            edgecolor="black",
            alpha=0.7,
        )
    except TypeError:
        ax = sns.barplot(
            data=df_all,
            x="Group",
            y="Fluorescence Density",
            hue="Group",
            order=group_order,
            hue_order=group_order,
            palette=palette_colors,
            legend=False,
            ci=None,
            edgecolor="black",
            alpha=0.7,
        )

    # Add SD error bars
    means = df_all.groupby("Group")["Fluorescence Density"].mean()
    sds = df_all.groupby("Group")["Fluorescence Density"].std(ddof=1)

    for xi, g in enumerate(group_order):
        m = float(np.asarray(means.get(g, 0.0)).item())
        sd = float(np.asarray(sds.get(g, 0.0)).item())
        cap = 14 if g == "Control" else 7

        ax.errorbar(
            xi, m, yerr=sd,
            fmt="none", ecolor="black", elinewidth=1.5,
            capsize=cap, capthick=1.5, zorder=10,
        )

    # Add jitter overlay
    sns.stripplot(
        x="Group", y="Fluorescence Density",
        data=df_all, order=group_order,
        jitter=True, color="cyan",
        edgecolor="black", linewidth=0.5,
        size=6, alpha=0.6,
    )

    # ✅ Add threshold lines - BOTH types use LOWER threshold
    legend_handles = []
    
    if control_mean is not None:
        # Control mean line (dotted blue)
        control_line = ax.axhline(
            y=control_mean, 
            color='blue', 
            linestyle=':', 
            linewidth=2.5, 
            label=f'Control Mean ({control_mean:.1f})', 
            zorder=5
        )
        legend_handles.append(control_line)
        
        # Calculate LOWER threshold for BOTH microgel types
        threshold = control_mean * (1 - threshold_pct)
        threshold_label = f'Lower Threshold (-{threshold_pct*100:.0f}%: {threshold:.1f})'
        
        threshold_line = ax.axhline(
            y=threshold, 
            color='red', 
            linestyle='--', 
            linewidth=2.5,
            label=threshold_label, 
            zorder=5
        )
        legend_handles.append(threshold_line)
        
        # Add legend
        ax.legend(
            handles=legend_handles,
            loc='upper right', 
            fontsize=10, 
            framealpha=0.95,
            edgecolor='black',
            fancybox=True
        )

    # Axis labels
    plt.ylabel("Fluorescence Density (a.u./µm²)", fontsize=12, fontweight="bold")
    plt.xlabel("")
    plt.xticks(fontsize=10, fontweight="bold")
    plt.yticks(fontsize=10, fontweight="bold")
    
    # Title - Consistent messaging
    filter_pct_display = int(percentile * 100)
    threshold_pct_display = int(threshold_pct * 100)
    
    title_parts = []
    if dataset_id:
        title_parts.append(dataset_id)
    
    # Microgel type description
    microgel_desc = "G- Microgel" if microgel_type.lower() == "negative" else "G+ Microgel"
    
    title_parts.append(
        f"{microgel_desc} — Typical Particles: Middle {100 - 2*filter_pct_display}% "
        f"(Excluded top/bottom {filter_pct_display}%)"
    )
    
    if title_suffix:
        title_parts.append(title_suffix)
    
    raw_title = " — ".join(title_parts)
    wrapped_title = "\n".join(
        textwrap.wrap(raw_title, width=80, break_long_words=False, break_on_hyphens=False)
    )
    plt.title(wrapped_title, fontsize=11, pad=10)

    # Enhance spines
    for axis in ["top", "bottom", "left", "right"]:
        plt.gca().spines[axis].set_linewidth(1.5)

    plt.tight_layout()

    # Save
    if output_path is not None:
        out_path = output_path
    else:
        if restrict_to_groups is not None:
            safe = "_".join([g.replace(" ", "_") for g in group_order])
            out_path = output_dir / f"comparison_{microgel_type}_{safe}.png"
        else:
            out_path = output_dir / f"comparison_{microgel_type}_all_groups.png"

    try:
        plt.savefig(out_path, dpi=300)
        plt.close()
        print(f"Saved plot with threshold: {out_path.name}")
        return out_path
    except Exception as e:
        print(f"[ERROR] Failed to save plot to {out_path}: {e}")
        plt.close()
        return None

def generate_pairwise_group_vs_control_plots(
    output_root: Path, 
    percentile: float, 
    dataset_id: str,
    threshold_pct: float,
    microgel_type: str
) -> None:
    """Generate pairwise plots with threshold lines"""
    
    control_folder = None
    for folder in output_root.iterdir():
        if folder.is_dir() and folder.name.lower().startswith("control"):
            control_folder = folder
            break
    
    if control_folder is None:
        print("[WARN] No Control group folder found - skipping pairwise plots")
        return
    
    control_master = control_folder / f"{control_folder.name}_master.xlsx"
    if not control_master.exists():
        print(f"[WARN] Control group master file not found: {control_master}")
        return
    
    control_display_name = _display_group_name(control_folder.name)
    
    for group_dir in sorted(output_root.iterdir()):
        if not group_dir.is_dir():
            continue
        if group_dir.name.lower().startswith("control"):
            continue
        if not re.fullmatch(r"\d+", group_dir.name):
            continue

        group_master = group_dir / f"{group_dir.name}_master.xlsx"
        if not group_master.exists():
            print(f"[WARN] Missing group master: {group_master}")
            continue

        pair_plot_path = group_dir / f"Group_{group_dir.name}_vs_Control_threshold.png"
        
        result = generate_error_bar_comparison_with_threshold(
            output_dir=output_root,
            percentile=percentile,
            restrict_to_groups=[group_dir.name, control_display_name],
            output_path=pair_plot_path,
            title_suffix=f"Group {group_dir.name} vs Control",
            dataset_id=dataset_id,
            threshold_pct=threshold_pct,
            microgel_type=microgel_type,
        )
        
        if result is not None:
            print(f"  Pairwise plot: {pair_plot_path.name}")


def embed_comparison_plots_into_all_excels(
    output_root: Path,
    percentile: float = 0.2,
    plot_path: Optional[Path] = None,
) -> None:
    """Post-process Excel files and embed plots"""
    
    if plot_path is None or not plot_path.exists():
        print(f"[WARN] Plot not found, embedding skipped")
        return
    
    excel_files = sorted(output_root.rglob("*_master.xlsx"))
    if not excel_files:
        print(f"[WARN] No master Excel files found under {output_root}")
        return

    def add_png(ws, path: Path, anchor_cell: str, width_px: int = 620) -> None:
        if not path.exists():
            return
        img = XLImage(str(path))
        if getattr(img, "width", None) and getattr(img, "height", None):
            scale = width_px / float(img.width)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
        ws.add_image(img, anchor_cell)

    updated = 0
    for excel_path in excel_files:
        try:
            wb = load_workbook(excel_path)
            modified = False

            if "Summary" in wb.sheetnames:
                ws_summary = wb["Summary"]
            elif "Error_Bar_Summary" in wb.sheetnames:
                ws_summary = wb["Error_Bar_Summary"]
                ws_summary.title = "Summary"
                ws_summary = wb["Summary"]
                modified = True
            else:
                ws_summary = wb.create_sheet("Summary")
                modified = True

            if "Ratios" in wb.sheetnames:
                ws_ratios = wb["Ratios"]
                ratios_idx = wb.sheetnames.index("Ratios")
                desired_idx = 1
                if ratios_idx != desired_idx:
                    wb.move_sheet(ws_ratios, offset=desired_idx - ratios_idx)
                    modified = True

            current_idx = wb.sheetnames.index(ws_summary.title)
            if current_idx != 0:
                wb.move_sheet(ws_summary, offset=-current_idx)
                modified = True

            marker_cell = "G1"
            marker_value = "COMPARISON_PLOTS_EMBEDDED"
            if ws_summary[marker_cell].value != marker_value:
                add_png(ws_summary, plot_path, "G3")
                ws_summary[marker_cell].value = marker_value
                modified = True

            if modified:
                wb.save(excel_path)
                updated += 1

        except Exception as e:
            print(f"[WARN] Could not embed plots into {excel_path}: {e}")

    if updated > 0:
        print(f"  Embedded plots in {updated} Excel files")


def consolidate_to_excel(output_dir: Path, group_name: str, percentile: float) -> None:
    """Consolidate all CSVs in a group folder into one Excel workbook"""
    group_name = normalize_group_folder_name(group_name)
    
    csv_files = list(output_dir.glob("*/object_stats.csv"))    

    if not csv_files:
        print(f"[WARN] No CSV files found in {output_dir}")
        return

    excel_path = output_dir / f"{group_name}_master.xlsx"

    if excel_path.exists():
        try:
            excel_path.unlink()
        except PermissionError:
            print(f"[ERROR] Cannot overwrite {excel_path} - file is open")
            return

    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")

        def adjust_column_widths(ws):
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.1
                if adjusted_width > 50:
                    adjusted_width = 50
                ws.column_dimensions[column_letter].width = adjusted_width

        def format_numbers(ws):
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0000'

        all_valid_objects: list[pd.DataFrame] = []
        all_excluded_objects: list[dict] = []

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            # README
            readme_df = pd.DataFrame(
                {
                    "Column Name": [
                        "Object_ID", "BF_Area_px", "BF_Area_um2", "Perimeter_px", "Perimeter_um",
                        "EquivDiameter_px", "EquivDiameter_um", "Circularity", "AspectRatio",
                        "CentroidX_px", "CentroidY_px", "CentroidX_um", "CentroidY_um",
                        "BBoxX_px", "BBoxY_px", "BBoxW_px", "BBoxH_px", "BBoxW_um", "BBoxH_um",
                        "Fluor_Area_px", "Fluor_Area_um2", "Fluor_Mean", "Fluor_Median",
                        "Fluor_Std", "Fluor_Min", "Fluor_Max", "Fluor_IntegratedDensity",
                        "Fluor_Density_per_BF_Area", "BF_to_Fluor_Area_Ratio",
                    ],
                    "Description": [
                        "Unique particle identifier", "Brightfield particle area (pixels²)", "Brightfield particle area (µm²)",
                        "Particle perimeter (pixels)", "Particle perimeter (µm)", "Diameter of equivalent circle (pixels)",
                        "Diameter of equivalent circle (µm)", "Shape roundness (0-1)", "Bounding box width/height ratio",
                        "Particle center X (px)", "Particle center Y (px)", "Particle center X (µm)", "Particle center Y (µm)",
                        "BBox top-left X (px)", "BBox top-left Y (px)", "BBox width (px)", "BBox height (px)",
                        "BBox width (µm)", "BBox height (µm)",
                        "Fluorescent region area (pixels²)", "Fluorescent region area (µm²)",
                        "Avg fluorescence intensity", "Median fluorescence intensity", "Std Dev fluorescence",
                        "Min fluorescence", "Max fluorescence", "Total fluorescence signal",
                        "Fluor density / BF Area (Primary Metric)", "Ratio of BF area to Fluor area",
                    ],
                }
            )

            readme_df.to_excel(writer, sheet_name="README", index=False)
            ws_readme = writer.sheets["README"]
            for cell in ws_readme[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            adjust_column_widths(ws_readme)

            # Per-image sheets
            for csv_file in sorted(csv_files):
                image_name = csv_file.parent.name
                df = pd.read_csv(csv_file)

                cols_to_numeric = ["Fluor_Area_px", "Fluor_IntegratedDensity", "BF_Area_um2", "Fluor_Area_um2"]
                for col in cols_to_numeric:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

                if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                    df["Fluor_Density_per_BF_Area"] = df["Fluor_IntegratedDensity"] / df["BF_Area_um2"]
                else:
                    df["Fluor_Density_per_BF_Area"] = 0.0

                if "Fluor_Area_um2" in df.columns and "BF_Area_um2" in df.columns:
                    df["BF_to_Fluor_Area_Ratio"] = df["BF_Area_um2"] / df["Fluor_Area_um2"]
                else:
                    df["BF_to_Fluor_Area_Ratio"] = 0.0

                df = df.replace([np.inf, -np.inf], 0).fillna(0)

                for idx, row in df.iterrows():
                    reason = None
                    
                    if row["Fluor_Area_px"] == 0 and row["Fluor_IntegratedDensity"] > 0:
                        reason = "Zero fluorescence area with positive integrated density"
                    elif row["Fluor_Area_px"] == 0:
                        reason = "Zero fluorescence area"
                    elif row["Fluor_IntegratedDensity"] == 0:
                        reason = "Zero integrated density"
                    
                    if reason:
                        all_excluded_objects.append({
                            "Object_ID": row["Object_ID"],
                            "Source_Image": image_name,
                            "BF_Area_um2": row["BF_Area_um2"],
                            "Fluor_Area_px": row["Fluor_Area_px"],
                            "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"],
                            "Exclusion_Reason": reason
                        })

                df_valid = df[(df["Fluor_IntegratedDensity"] > 0) & (df["Fluor_Area_px"] > 0)].copy()
                df_valid["Source_Image"] = image_name
                
                if not df_valid.empty:
                    all_valid_objects.append(df_valid)

                sheet_name = image_name[:31]
                
                if "Fluor_Density_per_BF_Area" in df.columns:
                    df = df.sort_values("Fluor_Density_per_BF_Area", ascending=False)
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                format_numbers(ws)
                adjust_column_widths(ws)
                ws.auto_filter.ref = ws.dimensions

            # Excluded Objects
            if all_excluded_objects:
                excluded_df = pd.DataFrame(all_excluded_objects)
                excluded_df.to_excel(writer, sheet_name="Excluded_Objects", index=False)
                ws_excluded = writer.sheets["Excluded_Objects"]
                
                for cell in ws_excluded[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                for row in ws_excluded.iter_rows(min_row=2, max_row=ws_excluded.max_row):
                    for cell in row:
                        cell.fill = red_fill
                
                adjust_column_widths(ws_excluded)
                ws_excluded.auto_filter.ref = ws_excluded.dimensions

            # All Valid Objects
            if all_valid_objects:
                merged_all = pd.concat(all_valid_objects, ignore_index=True)
                merged_all = merged_all.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)
                
                all_valid_sheet_name = f"{group_name}_All_Valid_Objects"
                merged_all.to_excel(writer, sheet_name=all_valid_sheet_name, index=False)
                ws_all = writer.sheets[all_valid_sheet_name]
                
                for cell in ws_all[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                format_numbers(ws_all)
                adjust_column_widths(ws_all)
                ws_all.auto_filter.ref = ws_all.dimensions

            # Typical Particles (GROUP-LEVEL filtering)
            if all_valid_objects:
                merged_all = pd.concat(all_valid_objects, ignore_index=True)
                merged_all = merged_all.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)
                
                n_total = len(merged_all)
                n_cut = int(n_total * percentile)
                
                start_idx = n_cut
                end_idx = n_total - n_cut
                
                if start_idx < end_idx and n_total > 3:
                    typical_particles = merged_all.iloc[start_idx:end_idx].copy()
                    
                    excluded_top = merged_all.iloc[:start_idx]
                    excluded_bottom = merged_all.iloc[end_idx:]
                    
                    for idx, row in excluded_top.iterrows():
                        all_excluded_objects.append({
                            "Object_ID": row["Object_ID"],
                            "Source_Image": row["Source_Image"],
                            "BF_Area_um2": row["BF_Area_um2"],
                            "Fluor_Area_px": row["Fluor_Area_px"],
                            "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"],
                            "Exclusion_Reason": f"Outside typical particle range (top {int(percentile*100)}%)"
                        })
                    
                    for idx, row in excluded_bottom.iterrows():
                        all_excluded_objects.append({
                            "Object_ID": row["Object_ID"],
                            "Source_Image": row["Source_Image"],
                            "BF_Area_um2": row["BF_Area_um2"],
                            "Fluor_Area_px": row["Fluor_Area_px"],
                            "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"],
                            "Exclusion_Reason": f"Outside typical particle range (bottom {int(percentile*100)}%)"
                        })
                else:
                    typical_particles = merged_all.copy()

                typical_sheet_name = f"{group_name}_Typical_Particles"
                typical_particles.to_excel(writer, sheet_name=typical_sheet_name, index=False)
                ws_typ = writer.sheets[typical_sheet_name]

                for cell in ws_typ[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                for row in ws_typ.iter_rows(min_row=2, max_row=ws_typ.max_row):
                    for cell in row:
                        cell.fill = yellow_fill
                
                format_numbers(ws_typ)
                adjust_column_widths(ws_typ)
                ws_typ.auto_filter.ref = ws_typ.dimensions

            # Update Excluded Objects
            if all_excluded_objects:
                wb = writer.book
                if "Excluded_Objects" in wb.sheetnames:
                    wb.remove(wb["Excluded_Objects"])
                
                excluded_df = pd.DataFrame(all_excluded_objects)
                excluded_df.to_excel(writer, sheet_name="Excluded_Objects", index=False)
                ws_excluded = writer.sheets["Excluded_Objects"]
                
                for cell in ws_excluded[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                for row in ws_excluded.iter_rows(min_row=2, max_row=ws_excluded.max_row):
                    for cell in row:
                        cell.fill = red_fill
                
                adjust_column_widths(ws_excluded)
                ws_excluded.auto_filter.ref = ws_excluded.dimensions

            # Summary sheet
            try:
                summary_data = []
                for csv_file in sorted(csv_files):
                    image_name = csv_file.parent.name
                    df = pd.read_csv(csv_file)

                    if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                        df["Fluor_Density_per_BF_Area"] = pd.to_numeric(df["Fluor_IntegratedDensity"], errors='coerce') / pd.to_numeric(df["BF_Area_um2"], errors='coerce')
                    else:
                        df["Fluor_Density_per_BF_Area"] = 0

                    if "Fluor_Area_px" in df.columns:
                        df["Fluor_Area_px"] = pd.to_numeric(df["Fluor_Area_px"], errors='coerce').fillna(0)
                        
                    df = df.replace([np.inf, -np.inf], 0).fillna(0)
                    
                    df_stats = df[(df["Fluor_Area_px"] > 0) & (df["Fluor_IntegratedDensity"] > 0)]
                    
                    avg_fluor_density = df_stats["Fluor_Density_per_BF_Area"].mean() if not df_stats.empty else 0.0
                    
                    avg_ratio = 0.0
                    if not df_stats.empty and "BF_Area_um2" in df_stats.columns and "Fluor_Area_um2" in df_stats.columns:
                         ratios = df_stats["BF_Area_um2"] / df_stats["Fluor_Area_um2"]
                         avg_ratio = ratios.replace([np.inf, -np.inf], 0).mean()

                    summary_data.append({
                        "Image": image_name,
                        "Total_Particles_Detected": len(df),
                        "Particles_With_Fluor": len(df_stats),
                        "Avg_BF_Area_um2": df["BF_Area_um2"].mean() if "BF_Area_um2" in df.columns else 0,
                        "Avg_Fluor_Density": avg_fluor_density,
                        "Avg_BF_to_Fluor_Ratio": avg_ratio,
                    })

                summary_df = pd.DataFrame(summary_data)

                if "Avg_Fluor_Density" in summary_df.columns:
                    summary_df = summary_df.sort_values("Avg_Fluor_Density", ascending=False)

                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                ws_summary = writer.sheets["Summary"]

                for cell in ws_summary[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                format_numbers(ws_summary)
                adjust_column_widths(ws_summary)

            except Exception as e:
                print(f"[WARN] Could not create Summary sheet: {e}")

            # Ratios sheet
            try:
                wb = writer.book
                ratios_name = "Ratios"
                if ratios_name in wb.sheetnames:
                    wb.remove(wb[ratios_name])
                ws_qa = wb.create_sheet(ratios_name)

                ws_qa["A1"] = f"QA Ratios - Group: {group_name}"
                ws_qa["A1"].font = Font(bold=True, size=14)

                def add_png(ws, path: Path, anchor_cell: str, width_px: int = 360) -> None:
                    if not path.exists(): return
                    img = XLImage(str(path))
                    if getattr(img, "width", None) and getattr(img, "height", None):
                        scale = width_px / float(img.width)
                        img.width = int(img.width * scale)
                        img.height = int(img.height * scale)
                    ws.add_image(img, anchor_cell)

                row = 3
                block_h = 34

                for csv_file in sorted(csv_files):
                    image_name = csv_file.parent.name
                    df = pd.read_csv(csv_file)
                    
                    if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                        df["Fluor_Density_per_BF_Area"] = pd.to_numeric(df["Fluor_IntegratedDensity"], errors='coerce') / pd.to_numeric(df["BF_Area_um2"], errors='coerce')
                    else:
                        df["Fluor_Density_per_BF_Area"] = 0.0
                    
                    if "BF_Area_um2" in df.columns and "Fluor_Area_um2" in df.columns:
                        df["BF_to_Fluor_Area_Ratio"] = pd.to_numeric(df["BF_Area_um2"], errors='coerce') / pd.to_numeric(df["Fluor_Area_um2"], errors='coerce')
                    else:
                        df["BF_to_Fluor_Area_Ratio"] = 0.0

                    df = df.replace([np.inf, -np.inf], 0).fillna(0)

                    if "Fluor_Density_per_BF_Area" in df.columns:
                        df = df.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)

                    ws_qa[f"A{row}"] = image_name
                    ws_qa[f"A{row}"].font = Font(bold=True, size=12)
                    
                    headers = ["Object_ID", "Fluor_Density_per_BF_Area", "Rank", "Object_ID", "BF_to_Fluor_Area_Ratio"]
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws_qa.cell(row=row+1, column=col_idx, value=header)
                        cell.font = Font(bold=True)
                        cell.alignment = center_align

                    start_data_row = row + 2
                    n = len(df)
                    
                    for k, r in enumerate(df.itertuples(index=False), 0):
                        ws_qa.cell(row=start_data_row+k, column=1, value=getattr(r, "Object_ID"))
                        ws_qa.cell(row=start_data_row+k, column=2, value=float(getattr(r, "Fluor_Density_per_BF_Area", 0.0))).number_format = '0.0000'
                        ws_qa.cell(row=start_data_row+k, column=3, value=k+1)
                        ws_qa.cell(row=start_data_row+k, column=4, value=getattr(r, "Object_ID"))
                        ws_qa.cell(row=start_data_row+k, column=5, value=float(getattr(r, "BF_to_Fluor_Area_Ratio", 0.0))).number_format = '0.0000'

                    ws_qa.column_dimensions["A"].width = 20
                    ws_qa.column_dimensions["B"].width = 25
                    ws_qa.column_dimensions["C"].width = 10
                    ws_qa.column_dimensions["D"].width = 20
                    ws_qa.column_dimensions["E"].width = 25

                    if n > 0:
                        ch1 = ScatterChart()
                        ch1.title = "Fluor Density"
                        ch1.y_axis.title = "a.u./µm²"
                        ch1.x_axis.title = "Object Rank (Sorted)"
                        
                        xref = Reference(ws_qa, min_col=3, min_row=start_data_row, max_row=start_data_row + n - 1)
                        yref = Reference(ws_qa, min_col=2, min_row=start_data_row, max_row=start_data_row + n - 1)
                        
                        s1 = SeriesFactory(yref, xref, title="Density")
                        s1.marker = Marker(symbol="triangle", size=5)
                        ch1.series.append(s1)
                        ws_qa.add_chart(ch1, f"G{row+1}")

                        ch2 = ScatterChart()
                        ch2.title = "Area Ratio"
                        ch2.y_axis.title = "Ratio"
                        ch2.x_axis.title = "Object Rank (Sorted)"
                        
                        yref2 = Reference(ws_qa, min_col=5, min_row=start_data_row, max_row=start_data_row + n - 1)
                        
                        s2 = SeriesFactory(yref2, xref, title="Ratio")
                        s2.marker = Marker(symbol="circle", size=5)
                        ch2.series.append(s2)
                        ws_qa.add_chart(ch2, f"G{row+18}")

                    img_dir = csv_file.parent
                    add_png(ws_qa, img_dir / "13_mask_accepted_ids.png", f"Q{row+1}", width_px=330)
                    add_png(ws_qa, img_dir / "22_fluorescence_mask_global_ids.png", f"Q{row+18}", width_px=330)
                    add_png(ws_qa, img_dir / "23_fluorescence_contours_global_ids.png", f"Q{row+18}", width_px=330)
                    add_png(ws_qa, img_dir / "24_bf_fluor_matching_overlay_ids.png", f"Q{row+1}", width_px=330)

                    row += block_h

            except Exception as e:
                print(f"[WARN] Could not create Ratios sheet: {e}")

        # Reorder worksheets
        wb2 = load_workbook(excel_path)
        desired_order = [
            "Summary", 
            "Ratios", 
            "README", 
            f"{group_name}_Typical_Particles",
            f"{group_name}_All_Valid_Objects",
            "Excluded_Objects"
        ]
        for idx, sheet_name in enumerate(desired_order):
            if sheet_name in wb2.sheetnames:
                sheet = wb2[sheet_name]
                current_idx = wb2.sheetnames.index(sheet_name)
                if current_idx != idx:
                    wb2.move_sheet(sheet, offset=idx - current_idx)
        wb2.save(excel_path)

    except PermissionError:
        print(f"[ERROR] Cannot write to {excel_path} - file may be open")
    except Exception as e:
        print(f"[ERROR] Failed to create Excel file: {e}")
        import traceback
        traceback.print_exc()


def export_group_statistics_to_csv(output_root: Path) -> None:
    """Export statistics with enhanced console summary"""
    
    stats_list = []
    
    for excel_path in sorted(output_root.glob("*/*_master.xlsx")):
        group_name = excel_path.parent.name
        
        try:
            typical_sheet = f"{group_name}_Typical_Particles"
            df = pd.read_excel(excel_path, sheet_name=typical_sheet)
                        
            if "Fluor_Density_per_BF_Area" in df.columns:
                values = pd.to_numeric(df["Fluor_Density_per_BF_Area"], errors='coerce').dropna()
                
                if values.empty:
                    continue

                mean_val = float(values.mean())
                std_val = float(values.std(ddof=1))
                sem_val = values.sem()
                median_val = float(values.median())
                min_val = float(values.min())
                max_val = float(values.max())
                q30 = float(values.quantile(0.30))
                q70 = float(values.quantile(0.70))
                n = int(len(values))
                
                stats_list.append({
                    'Group': "Control" if group_name.lower().startswith("control") else group_name,
                    'N': n,
                    'Mean': mean_val,
                    'Std_Dev': std_val,
                    'SEM': sem_val,
                    'Median': median_val,
                    'Q30': q30,
                    'Q70': q70,
                    'Min': min_val,
                    'Max': max_val,
                    'CV_percent': (std_val / mean_val * 100) if mean_val > 0 else 0,
                })
                
        except Exception:
            pass
    
    if not stats_list:
        return
    
    stats_df = pd.DataFrame(stats_list)
    
    stats_df['sort_key'] = stats_df['Group'].apply(
        lambda x: (0, int(x)) if x.isdigit() else (1, 999)
    )
    stats_df = stats_df.sort_values('sort_key').drop('sort_key', axis=1)
    
    numeric_cols = stats_df.select_dtypes(include=[np.number]).columns
    stats_df[numeric_cols] = stats_df[numeric_cols].round(2)
    
    output_path = output_root / "group_statistics_summary.csv"
    stats_df.to_csv(output_path, index=False)


def classify_groups_clinical(
    output_root: Path, 
    microgel_type: str = "negative",
    threshold_pct: float = 0.05,
    bacteria_profile: str = "default",
    threshold_mode: str = "adaptive",
    confidence_level: float = 0.95
) -> pd.DataFrame:
    """Classify all groups - Enhanced with adaptive thresholding."""
    
    # Load control data
    control_mean: Optional[float] = None
    control_std: Optional[float] = None
    control_folder = None
    
    for folder in output_root.iterdir():
        if folder.is_dir() and folder.name.lower().startswith("control"):
            control_folder = folder
            break
    
    if control_folder is None:
        print("[WARN] No control folder found")
        return pd.DataFrame()
    
    control_master = control_folder / f"{control_folder.name}_master.xlsx"
    if not control_master.exists():
        print(f"[WARN] Control master file not found: {control_master.name}")
        return pd.DataFrame()
    
    try:
        typical_sheet = f"{control_folder.name}_Typical_Particles"
        df_control = pd.read_excel(control_master, sheet_name=typical_sheet)
        
        if "Fluor_Density_per_BF_Area" not in df_control.columns:
            print("[WARN] Missing Fluor_Density_per_BF_Area in control data")
            return pd.DataFrame()
        
        control_values = pd.to_numeric(
            df_control["Fluor_Density_per_BF_Area"], 
            errors='coerce'
        ).dropna()
        
        if len(control_values) < 3:
            print(f"[WARN] Insufficient control data (n={len(control_values)})")
            return pd.DataFrame()
        
        # Calculate thresholds based on mode
        if threshold_mode == 'adaptive':
            print(f"\n  Using ADAPTIVE thresholding:")
            print(f"    Bacteria profile: {BACTERIA_PROFILES[bacteria_profile]['name']}")
            print(f"    Confidence level: {confidence_level*100:.0f}%")
            
            threshold_result = calculate_adaptive_thresholds(
                df_control,
                bacteria_profile=bacteria_profile,
                confidence_level=confidence_level
            )
            
            threshold = threshold_result['threshold_value']
            control_mean = threshold_result['control_mean']
            control_std = threshold_result['control_std']
            
            print(f"\n    Control statistics:")
            print(f"      Mean: {control_mean:.2f}")
            print(f"      Std Dev: {control_std:.2f}")
            print(f"      Median: {threshold_result['control_median']:.2f}")
            print(f"\n    Adaptive threshold: {threshold:.2f}")
            print(f"      Method: {threshold_result['threshold_method']}")
            print(f"      Confidence: {threshold_result['confidence']}")
            print(f"\n    All calculated thresholds:")
            for method, value in threshold_result['all_thresholds'].items():
                marker = "←" if method == threshold_result['threshold_method'] else ""
                print(f"      {method}: {value:.2f} {marker}")
        
        else:
            # Fixed threshold mode
            control_mean = float(control_values.mean())
            control_std = float(control_values.std(ddof=1))
            threshold = control_mean * (1 - threshold_pct)
            
            threshold_result = {
                'threshold_value': threshold,
                'control_mean': control_mean,
                'control_std': control_std,
                'threshold_method': 'fixed_percentage',
                'confidence': 'user_defined'
            }
            
            print(f"\n  Using FIXED thresholding:")
            print(f"    Control mean: {control_mean:.2f}")
            print(f"    Threshold: {threshold:.2f} ({threshold_pct*100:.1f}% below control)")
        
    except Exception as e:
        print(f"[ERROR] Could not load control data: {e}")
        return pd.DataFrame()
    
    # Classify all groups
    results = []
    
    # Add control group
    results.append({
        'Group': 'Control',
        'N': len(control_values),
        'Mean': round(control_mean, 2) if control_mean is not None else 0.0,
        'Std_Dev': round(control_std, 2) if control_std is not None else 0.0,
        'Control_Mean': round(control_mean, 2) if control_mean is not None else 0.0,
        'Threshold': round(threshold, 2) if threshold is not None else 0.0,
        'Threshold_Method': threshold_result.get('threshold_method', 'unknown'),
        'Diff_from_Threshold': 0.0,
        'Z_Score': 0.0,
        'Classification': 'CONTROL/Reference',
        'Confidence': 'reference',
        'Confidence_Score': 1.0,
    })
    
    # Process test groups
    for excel_path in sorted(output_root.glob("*/*_master.xlsx")):
        group_name = excel_path.parent.name
        
        if group_name.lower().startswith("control"):
            continue
        
        try:
            typical_sheet = f"{group_name}_Typical_Particles"
            df = pd.read_excel(excel_path, sheet_name=typical_sheet)
            
            if "Fluor_Density_per_BF_Area" not in df.columns:
                continue
            
            values = pd.to_numeric(
                df["Fluor_Density_per_BF_Area"], 
                errors='coerce'
            ).dropna()
            
            if values.empty:
                continue
            
            mean_val = float(values.mean())
            std_val = float(values.std(ddof=1))
            n = len(values)
            
            # Apply adaptive classification
            class_result = apply_adaptive_classification(
                mean_val,
                threshold_result,
                bacteria_profile
            )
            
            results.append({
                'Group': group_name,
                'N': n,
                'Mean': round(mean_val, 2),
                'Std_Dev': round(std_val, 2),
                'Control_Mean': round(control_mean, 2) if control_mean is not None else 0.0,
                'Threshold': round(threshold, 2),
                'Threshold_Method': threshold_result.get('threshold_method', 'unknown'),
                'Diff_from_Threshold': round(class_result['distance_from_threshold'], 2),
                'Z_Score': round(class_result['z_score'], 2),
                'Classification': class_result['classification'],
                'Confidence': class_result['confidence'],
                'Confidence_Score': round(class_result['confidence_score'], 2),
            })
            
        except Exception as e:
            print(f"[WARN] Could not classify {group_name}: {e}")
            continue
    
    if not results:
        return pd.DataFrame()
    
    results_df = pd.DataFrame(results)
    
    # Sort: Control last
    results_df['sort_key'] = results_df['Group'].apply(
        lambda x: (1, 999) if x == 'Control' else (0, int(x) if x.isdigit() else 999)
    )
    results_df = results_df.sort_values('sort_key').drop('sort_key', axis=1)
    
    return results_df


def export_clinical_classification(
    output_root: Path,
    classification_df: pd.DataFrame,
    microgel_type: str = "negative",
    bacteria_profile: str = "default"
) -> Optional[Path]:
    """Export clinical classification - Enhanced with confidence scores"""
    
    if classification_df.empty:
        return None
    
    profile = BACTERIA_PROFILES.get(bacteria_profile, BACTERIA_PROFILES['default'])
    
    csv_path = output_root / f"clinical_classification_{microgel_type}_{bacteria_profile}.csv"
    classification_df.to_csv(csv_path, index=False)
    
    excel_path = output_root / f"clinical_classification_{microgel_type}_{bacteria_profile}.xlsx"
    
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            classification_df.to_excel(writer, sheet_name='Classification', index=False)
            
            ws = writer.sheets['Classification']
            
            # Color definitions
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            safe_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            low_conf_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
            
            header_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")
            
            # Format header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            # Format data rows with confidence-aware coloring
            for row_idx in range(len(classification_df)):
                excel_row = row_idx + 2
                row_data = classification_df.iloc[row_idx]
                
                # Determine fill based on classification AND confidence
                classification = row_data['Classification']
                confidence = row_data.get('Confidence', 'medium')
                
                if 'CONTROL' in classification:
                    fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                elif "No obvious bacteria" in classification:
                    if confidence == 'high':
                        fill = safe_fill
                    elif confidence == 'medium':
                        fill = medium_fill
                    else:
                        fill = low_conf_fill
                else:
                    # Bacteria detected
                    if confidence == 'high':
                        fill = warning_fill
                    elif confidence == 'medium':
                        fill = medium_fill
                    else:
                        fill = low_conf_fill
                
                for col_idx in range(1, len(classification_df.columns) + 1):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center")
            
            # Auto-width columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.1, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add metadata sheet
            metadata = pd.DataFrame({
                'Parameter': [
                    'Bacteria Profile',
                    'Bacteria Name',
                    'Gram Stain',
                    'Microgel Type',
                    'Description',
                    'Detection Method',
                    'Typical Size (μm)',
                    'Expected Shape',
                ],
                'Value': [
                    bacteria_profile,
                    profile['name'],
                    profile['gram_stain'],
                    profile['microgel_type'],
                    profile['description'],
                    'Adaptive Thresholding',
                    f"{profile['typical_size_um'][0]} - {profile['typical_size_um'][1]}",
                    profile['expected_shape'],
                ]
            })
            
            metadata.to_excel(writer, sheet_name='Detection_Parameters', index=False)
            ws_meta = writer.sheets['Detection_Parameters']
            
            for cell in ws_meta[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            ws_meta.column_dimensions['A'].width = 25
            ws_meta.column_dimensions['B'].width = 50
        
        print(f"\n  ✓ Classification saved: {excel_path.name}")
        print(f"    Bacteria: {profile['name']}")
        print(f"    Method: Adaptive thresholding")
        
        return excel_path
        
    except Exception as e:
        print(f"[ERROR] Could not create Excel: {e}")
        return None


def generate_final_clinical_matrix(
    output_root: Path,
    gplus_classification: pd.DataFrame,
    gminus_classification: pd.DataFrame,
    dataset_base_name: str
) -> Optional[Path]:
    """Generate final clinical result matrix combining G+ and G- results"""
    
    if gplus_classification.empty or gminus_classification.empty:
        print("[WARN] Missing classification data - cannot generate final matrix")
        return None
    
    # ✅ FIX: Ensure Group column is string type in both DataFrames
    gplus_classification['Group'] = gplus_classification['Group'].astype(str)
    gminus_classification['Group'] = gminus_classification['Group'].astype(str)
    
    # Clinical decision matrix (unchanged)
    decision_matrix = {
        ('POSITIVE', 'POSITIVE/No obvious bacteria'): 'POSITIVE',
        ('NEGATIVE/No obvious bacteria', 'NEGATIVE'): 'NEGATIVE',
        ('NEGATIVE/No obvious bacteria', 'POSITIVE/No obvious bacteria'): 'NO OBVIOUS BACTERIA',
        ('POSITIVE', 'NEGATIVE'): 'MIXED/CONTRADICTORY',
    }
    
    # Prepare results
    results = []
    
    # ✅ FIX: Get ALL groups including Control
    gplus_groups = set(gplus_classification['Group'])
    gminus_groups = set(gminus_classification['Group'])
    all_groups = sorted(gplus_groups | gminus_groups, key=_group_order_key)
    
    # ✅ FIX: Create lookup dictionaries with string keys
    gplus_dict = gplus_classification.set_index('Group').to_dict('index')
    gminus_dict = gminus_classification.set_index('Group').to_dict('index')
    
    # ✅ DEBUG: Print what we found
    print(f"  DEBUG: G+ groups found: {sorted(gplus_groups)}")
    print(f"  DEBUG: G- groups found: {sorted(gminus_groups)}")
    print(f"  DEBUG: All groups to process: {all_groups}")
    
    for group in all_groups:
        # Get G+ data
        gplus_data = gplus_dict.get(group)
        if gplus_data:
            gplus_mean = gplus_data['Mean']
            gplus_class = gplus_data['Classification']
            gplus_detection = 'POSITIVE' if 'POSITIVE' in gplus_class and 'No obvious' not in gplus_class else 'NEGATIVE/No obvious bacteria'
        else:
            gplus_mean = None
            gplus_class = None
            gplus_detection = '-'
        
        # Get G- data
        gminus_data = gminus_dict.get(group)
        if gminus_data:
            gminus_mean = gminus_data['Mean']
            gminus_class = gminus_data['Classification']
            gminus_detection = 'NEGATIVE' if 'NEGATIVE' in gminus_class and 'No obvious' not in gminus_class else 'POSITIVE/No obvious bacteria'
        else:
            gminus_mean = None
            gminus_class = None
            gminus_detection = '-'
        
        # ✅ FIX: Special handling for Control group
        if group == 'Control':
            final_class = 'CONTROL (Reference)'
        # Determine final classification for test groups
        elif gplus_class is None and gminus_class is None:
            final_class = 'MISSING DATA'
        elif gplus_class is None:
            final_class = 'MISSING G+'
        elif gminus_class is None:
            final_class = 'MISSING G-'
        else:
            # Look up in decision matrix
            key = (gplus_class, gminus_class)
            final_class = decision_matrix.get(key, 'UNKNOWN COMBINATION')
        
        results.append({
            'Group': group,
            'G+_Mean': gplus_mean if gplus_mean is not None else '-',
            'G+_Detection': gplus_detection,
            'G-_Mean': gminus_mean if gminus_mean is not None else '-',
            'G-_Detection': gminus_detection,
            'Final_Classification': final_class
        })
    
    # Rest of the function remains the same...
    # Create DataFrame
    final_df = pd.DataFrame(results)
    
    # Round numeric values
    for col in ['G+_Mean', 'G-_Mean']:
        final_df[col] = final_df[col].apply(
            lambda x: f"{x:.2f}" if isinstance(x, (int, float)) else x
        )
    
    # Export to CSV
    csv_path = output_root / "final_clinical_results.csv"
    final_df.to_csv(csv_path, index=False)
    print(f"  Final results CSV: {csv_path.name}")
    
    # Export to Excel with formatting (unchanged - but add Control color)
    excel_path = output_root / "final_clinical_results.xlsx"
    
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            final_df.to_excel(writer, sheet_name='Final Results', index=False)
            
            ws = writer.sheets['Final Results']
            
            # Define colors
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            positive_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            negative_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            no_bacteria_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            mixed_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
            missing_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            control_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # ✅ NEW
            
            header_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            
            # Format data rows
            for row_idx in range(len(final_df)):
                excel_row = row_idx + 2
                final_class = final_df.iloc[row_idx]['Final_Classification']
                
                # ✅ FIX: Add control group color
                if final_class == 'CONTROL (Reference)':
                    row_fill = control_fill
                elif final_class == 'POSITIVE':
                    row_fill = positive_fill
                elif final_class == 'NEGATIVE':
                    row_fill = negative_fill
                elif final_class == 'NO OBVIOUS BACTERIA':
                    row_fill = no_bacteria_fill
                elif final_class == 'MIXED/CONTRADICTORY':
                    row_fill = mixed_fill
                else:
                    row_fill = missing_fill
                
                # Apply formatting to all cells in row
                for col_idx in range(1, len(final_df.columns) + 1):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    cell.fill = row_fill
                    cell.alignment = center_align
                    cell.border = thin_border
            
            # Adjust column widths (unchanged)
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.1, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Update legend with Control
            legend_data = pd.DataFrame({
                'Classification': ['POSITIVE', 'NEGATIVE', 'NO OBVIOUS BACTERIA', 'MIXED/CONTRADICTORY', 'CONTROL (Reference)', 'MISSING DATA'],
                'Meaning': [
                    'Bacteria detected (G+ positive AND G- positive/no bacteria)',
                    'No bacteria detected (G+ negative/no bacteria AND G- negative)',
                    'No obvious bacteria (both G+ and G- show no bacteria)',
                    'Contradictory results (G+ positive AND G- negative)',
                    'Control group (reference baseline)',
                    'Insufficient data for classification'
                ],
                'Color': ['Light Red', 'Light Green', 'Light Yellow', 'Light Orange', 'Light Gray', 'Gray']
            })
            
            legend_data.to_excel(writer, sheet_name='Legend', index=False)
            ws_legend = writer.sheets['Legend']
            
            # Format legend (add control color)
            for cell in ws_legend[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            for row_idx in range(len(legend_data)):
                excel_row = row_idx + 2
                class_name = legend_data.iloc[row_idx]['Classification']
                
                if class_name == 'POSITIVE':
                    fill = positive_fill
                elif class_name == 'NEGATIVE':
                    fill = negative_fill
                elif class_name == 'NO OBVIOUS BACTERIA':
                    fill = no_bacteria_fill
                elif class_name == 'MIXED/CONTRADICTORY':
                    fill = mixed_fill
                elif class_name == 'CONTROL (Reference)':
                    fill = control_fill
                else:
                    fill = missing_fill
                
                for col_idx in range(1, 4):
                    ws_legend.cell(row=excel_row, column=col_idx).fill = fill
                    ws_legend.cell(row=excel_row, column=col_idx).alignment = center_align
            
            ws_legend.column_dimensions['A'].width = 25
            ws_legend.column_dimensions['B'].width = 60
            ws_legend.column_dimensions['C'].width = 15
        
        print(f"  Final results Excel: {excel_path.name}")
        
        # Print summary
        print("\n" + "="*80)
        print("FINAL CLINICAL RESULTS SUMMARY")
        print("="*80)
        print(final_df.to_string(index=False))
        print("="*80 + "\n")
        
        return excel_path
        
    except Exception as e:
        print(f"[ERROR] Could not create Excel file: {e}")
        return None


# ==================================================
# Source Directory Selection
# ==================================================
def select_source_directory(max_depth=2) -> Optional[Path]:
    """Lists directories that either have a Control subfolder OR contain G+/G- subdirectories"""
    root_dir = Path('source')
    
    if not root_dir.exists():
        print(f"[ERROR] Source directory not found: {root_dir.resolve()}")
        return None
    
    valid_directories = []
    
    # Check immediate subdirectories of source/
    for item in root_dir.iterdir():
        if not item.is_dir():
            continue
            
        # Check if this directory has both G+ and G- subdirectories
        has_gplus = (item / 'G+').is_dir()
        has_gminus = (item / 'G-').is_dir()
        
        if has_gplus and has_gminus:
            # This is a batch directory (like "PD sample" or "Spike sample")
            valid_directories.append(item.name)
            continue
        
        # Check if this directory has a Control subfolder (single-mode directory)
        try:
            subdirs = [d for d in item.iterdir() if d.is_dir()]
            has_control = any(d.name.lower().startswith('control') for d in subdirs)
            if has_control:
                valid_directories.append(item.name)
        except OSError:
            continue
    
    if not valid_directories:
        print("[ERROR] No valid directories found.")
        print("Valid directories must either:")
        print("  1. Contain both 'G+' and 'G-' subfolders (for batch processing)")
        print("  2. Contain a 'Control' subfolder (for single processing)")
        return None
    
    valid_directories.sort()
    
    print("\n" + "="*80)
    print("SELECT SOURCE DIRECTORY")
    print("="*80)
    print("\nAvailable directories:")
    for i, dir_name in enumerate(valid_directories, 1):
        dir_path = root_dir / dir_name
        has_gplus = (dir_path / 'G+').is_dir()
        has_gminus = (dir_path / 'G-').is_dir()
        
        if has_gplus and has_gminus:
            mode_label = "[BATCH: G+ and G-]"
        else:
            mode_label = "[SINGLE]"
        
        print(f"  [{i}] {dir_name} {mode_label}")
    
    while True:
        selected = logged_input("\nEnter the number or folder name (or 'q' to quit): ").strip()
        
        if selected.lower() in {'q', 'quit', 'exit'}:
            raise SystemExit(0)
        
        if selected.isdigit():
            num = int(selected)
            if 1 <= num <= len(valid_directories):
                selected_name = valid_directories[num - 1]
                full_selected = root_dir / selected_name
                return full_selected
            else:
                print(f"Invalid number. Please enter between 1 and {len(valid_directories)}.")
        elif selected in valid_directories:
            full_selected = root_dir / selected
            return full_selected
        else:
            print("Invalid selection. Please enter a valid number or folder name.")

def collect_configuration() -> dict:
    """Collect all user configuration upfront - Enhanced with bacteria selection
    
    Returns:
        dict: Configuration dictionary with all user settings
        
    Raises:
        SystemExit: If no source directory is selected
    """
    
    print("\n" + "="*80)
    print("PHASE 1: CONFIGURATION")
    print("="*80)
    print("\nPlease answer the following questions:\n")
    
    config = {}
    
    # Step 1: Source Directory
    print("━" * 80)
    print("STEP 1/5: Select Source Directory")
    print("━" * 80)
    config['source_dir'] = select_source_directory()
    if config['source_dir'] is None:
        raise SystemExit("No source directory selected.")
    
    # Auto-detect batch mode
    gplus_path = config['source_dir'] / 'G+'
    gminus_path = config['source_dir'] / 'G-'
    
    has_gplus = gplus_path.is_dir()
    has_gminus = gminus_path.is_dir()
    
    if has_gplus and has_gminus:
        config['batch_mode'] = True
        config['dataset_base_name'] = config['source_dir'].name
        config['subdirs'] = []
        
        print(f"\nDetected BATCH PROCESSING mode")
        print(f"  Parent folder: {config['dataset_base_name']}")
        print(f"  Will process:")
        print(f"    → {config['dataset_base_name']}/G+ (Gram-positive)")
        print(f"    → {config['dataset_base_name']}/G- (Gram-negative)")
        
        config['subdirs'].append({
            'path': gplus_path,
            'microgel_type': 'positive',
            'label': 'G+',
            'safe_label': 'Positive'
        })
        config['subdirs'].append({
            'path': gminus_path,
            'microgel_type': 'negative',
            'label': 'G-',
            'safe_label': 'Negative'
        })
        
    else:
        config['batch_mode'] = False
        config['dataset_base_name'] = None
        
        # Detect microgel type from folder structure
        has_control = any(
            d.name.lower().startswith('control') 
            for d in config['source_dir'].iterdir() 
            if d.is_dir()
        )
        
        if not has_control:
            print("\n⚠ No Control folder found in source directory")
            print("  A Control folder is required for processing")
            raise SystemExit("Missing Control folder")
    
    # Step 2: Bacteria Type Selection
    print("\n" + "━" * 80)
    print("STEP 2/5: Expected Bacteria Type")
    print("━" * 80)
    print("\nSelect expected bacteria for optimized detection:")
    print("(This will auto-tune thresholds and validation)")
    print()
    
    for key, num, desc in BACTERIA_DISPLAY_OPTIONS:
        print(f"  [{num}] {desc}")
    
    print("\n  → Press Enter to use adaptive detection (recommended for unknown samples)")
    
    while True:
        bacteria_choice = logged_input("\nEnter number: ").strip()
        
        if bacteria_choice == "":
            config['bacteria_profile'] = 'default'
            print("  Using adaptive detection mode")
            break
        
        # Find matching profile
        selected_profile = None
        for key, num, desc in BACTERIA_DISPLAY_OPTIONS:
            if bacteria_choice == num:
                selected_profile = key
                break
        
        if selected_profile:
            config['bacteria_profile'] = selected_profile
            profile = BACTERIA_PROFILES[selected_profile]
            print(f"\n  ✓ Selected: {profile['name']}")
            print(f"    Gram stain: {profile['gram_stain']}")
            print(f"    Description: {profile['description']}")
            
            # Auto-set microgel type if not batch mode
            if not config.get('batch_mode', False):
                config['microgel_type'] = profile['microgel_type']
                print(f"    Auto-selected: {profile['microgel_type']} microgel")
            
            break
        else:
            print("  Invalid selection. Please enter 1-4 or press Enter.")
    
    # Step 3: Dataset ID
    print("\n" + "━" * 80)
    print("STEP 3/5: Dataset Identifier")
    print("━" * 80)
    
    if config.get('batch_mode', False):
        # Batch mode: ask for base ID
        print("\nEnter base dataset identifier:")
        print(f"  Example: 'v48_PD' will create:")
        print(f"    - v48_PD Positive")
        print(f"    - v48_PD Negative")
        
        while True:
            dataset_id_base = logged_input("\nDataset base ID: ").strip()
            if dataset_id_base:
                config['dataset_id_base'] = dataset_id_base
                print(f"  ✓ Base ID: {dataset_id_base}")
                break
            else:
                print("  Dataset ID cannot be empty")
    else:
        # Single mode: ask for full ID
        print("\nEnter dataset identifier:")
        print(f"  Example: 'v48_PD_Sample1'")
        
        while True:
            dataset_id = logged_input("\nDataset ID: ").strip()
            if dataset_id:
                config['dataset_id'] = dataset_id
                print(f"  ✓ Dataset ID: {dataset_id}")
                break
            else:
                print("  Dataset ID cannot be empty")
    
    # Step 4: Percentile
    print("\n" + "━" * 80)
    print("STEP 4/5: Percentile for Top/Bottom Filtering")
    print("━" * 80)
    print("\nEnter percentile to exclude from top and bottom:")
    print("  → Default: 20% (excludes top 20% and bottom 20%)")
    print("  → Range: 0-40%")
    print("  → Enter 0 to include all particles")
    
    while True:
        percentile_input = logged_input("\nPercentile (%): ").strip()
        
        if percentile_input == "":
            config['percentile'] = 0.20
            print("  Using default: 20%")
            break
        
        try:
            value = float(percentile_input)
            if value > 1:
                value = value / 100
            
            if 0 <= value <= 0.40:
                config['percentile'] = value
                print(f"  Percentile: {value*100:.0f}%")
                break
            else:
                print("  Must be between 0-40%")
        except ValueError:
            print("  Invalid number")
    
    # Step 5: Clinical Threshold
    print("\n" + "━" * 80)
    print("STEP 5/5: Clinical Classification Method")
    print("━" * 80)
    print("\nChoose threshold calculation method:")
    print("  [1] Adaptive (recommended) - Auto-calculates from control data")
    print("  [2] Fixed percentage - Manual threshold setting")
    print()
    
    method_choice = logged_input("Enter number (Enter=Adaptive): ").strip()
    
    if method_choice == "2":
        # Fixed threshold mode
        config['threshold_mode'] = 'fixed'
        print("\nEnter threshold percentage:")
        print("  → Default: 5%")
        print("  → Range: 1-20%")
        
        while True:
            choice = logged_input("Threshold (%): ").strip()
            if choice == "":
                config['threshold_pct'] = 0.05
                print("  Using default: 5%")
                break
            try:
                value = float(choice)
                if value > 1:
                    value = value / 100
                if 0.01 <= value <= 0.20:
                    config['threshold_pct'] = value
                    print(f"  Threshold: {value*100:.1f}%")
                    break
                else:
                    print("  Must be 1-20%")
            except ValueError:
                print("  Invalid number")
    else:
        # Adaptive mode1
        config['threshold_mode'] = 'adaptive'
        config['confidence_level'] = 0.95
        config['threshold_pct'] = 0.05  # Fallback value
        print("  ✓ Using adaptive thresholding")
        print("    - Calculates optimal threshold from control statistics")
        print("    - Adjusts for bacteria-specific characteristics")
        print("    - Confidence level: 95%")
    
    return config



def display_configuration_summary(config: dict) -> None:
    """Display configuration summary - Enhanced"""
    print("\n" + "="*80)
    print("CONFIGURATION SUMMARY")
    print("="*80 + "\n")
    
    if config.get('batch_mode', False):
        print(f"Mode: BATCH PROCESSING")
        print(f"1. Base Directory: {config['source_dir']}")
        print(f"2. Dataset Base ID: {config.get('dataset_id_base', 'N/A')}")
        print(f"3. Bacteria Profile: {BACTERIA_PROFILES[config['bacteria_profile']]['name']}")  # NEW
        print(f"   → Processing:")
        for subdir in config['subdirs']:
            dataset_name = f"{config.get('dataset_id_base', '')} {subdir['safe_label']}"
            print(f"     • {subdir['label']}: {dataset_name}")
        print(f"4. Percentile: {config['percentile']*100:.0f}%")
        print(f"5. Threshold Mode: {config.get('threshold_mode', 'adaptive').upper()}")  # NEW
        if config.get('threshold_mode') == 'fixed':
            print(f"   Fixed threshold: {config['threshold_pct']*100:.1f}%")
        else:
            print(f"   Confidence level: {config.get('confidence_level', 0.95)*100:.0f}%")
    else:
        print(f"Mode: SINGLE PROCESSING")
        print(f"1. Source Directory: {config['source_dir']}")
        print(f"2. Dataset ID: {config['dataset_id']}")
        print(f"3. Bacteria Profile: {BACTERIA_PROFILES[config['bacteria_profile']]['name']}")  # NEW
        print(f"   Gram stain: {BACTERIA_PROFILES[config['bacteria_profile']]['gram_stain']}")  # NEW
        print(f"4. Percentile: {config['percentile']*100:.0f}%")
        print(f"5. Microgel Type: {config['microgel_type'].capitalize()}")
        print(f"6. Threshold Mode: {config.get('threshold_mode', 'adaptive').upper()}")  # NEW
        if config.get('threshold_mode') == 'fixed':
            print(f"   Fixed threshold: {config['threshold_pct']*100:.1f}%")
        else:
            print(f"   Adaptive with {config.get('confidence_level', 0.95)*100:.0f}% confidence")
    
    print("\n" + "="*80 + "\n")
    
    confirm = logged_input("Proceed with this configuration? (y/n, Enter=yes): ").strip().lower()
    
    if confirm not in ["", "y", "yes"]:
        raise SystemExit("Configuration cancelled by user.")
    
    print("\nConfiguration confirmed - starting processing...\n")


# ==================================================
# Image Processing
# ==================================================
def process_image(img_path: Path, output_root: Path) -> None:
    """Process a single image - Unicode-safe version
    
    Args:
        img_path: Path to input image
        output_root: Root output directory
    """
    
    # Validate path encoding
    if not validate_path_encoding(img_path):
        print(f"[ERROR] Cannot process image with problematic path: {img_path}")
        return
    
    xml_props, xml_main = find_metadata_paths(img_path)
    
    try:
        um_per_px_x, um_per_px_y = get_pixel_size_um(xml_props, xml_main)
        um_per_px_x = float(um_per_px_x)
        um_per_px_y = float(um_per_px_y)
    except Exception as e:
        if FALLBACK_UM_PER_PX is None:
            raise
        print(f"[WARN] Using fallback pixel size for {img_path.name}: {e}")
        um_per_px_x = um_per_px_y = float(FALLBACK_UM_PER_PX)

    um_per_px_avg = (um_per_px_x + um_per_px_y) / 2.0

    original_folder_name = img_path.parent.name
    if original_folder_name.lower().strip() == 'control':
        normalized_folder_name = "Control"  # Standardize to "Control"
    else:
        normalized_folder_name = original_folder_name
    
    img_out = output_root / normalized_folder_name

    ensure_dir(img_out)

    # Use Unicode-safe imread
    img = safe_imread(img_path, cv2.IMREAD_UNCHANGED)
    if img is None:
        raise FileNotFoundError(f"Could not read image: {img_path}")
    if img.ndim == 3:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    img8 = normalize_to_8bit(img)
    save_debug(img_out, "01_gray_8bit.png", img8, um_per_px_avg)

    mask = segment_particles_brightfield(img8, float(um_per_px_avg), img_out)

    _fc = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = cast(list[np.ndarray], _fc[-2])

    vis_all = cv2.cvtColor(img8, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(vis_all, contours, -1, (0, 0, 255), 1)
    save_debug(img_out, "10_contours_all.png", vis_all, um_per_px_avg)

    um2_per_px2 = float(um_per_px_x) * float(um_per_px_y)
    min_area_px = MIN_AREA_UM2 / um2_per_px2
    max_area_px = MAX_AREA_UM2 / um2_per_px2

    H, W = img8.shape[:2]
    img_area_px = float(H * W)
    max_big_area_px = MAX_FRACTION_OF_IMAGE_AREA * img_area_px

    accepted: list[np.ndarray] = []
    rejected: list[np.ndarray] = []

    for c in contours:
        area_px = float(cv2.contourArea(c))
        if area_px <= 0:
            rejected.append(c)
            continue

        if area_px >= max_big_area_px:
            rejected.append(c)
            continue

        perim_px = float(cv2.arcLength(c, True))
        circ = (4 * np.pi * area_px / (perim_px**2)) if perim_px > 0 else 0.0

        ok = (min_area_px <= area_px <= max_area_px) and (circ >= MIN_CIRCULARITY)
        (accepted if ok else rejected).append(c)

    vis_acc = cv2.cvtColor(img8, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(vis_acc, rejected, -1, (0, 165, 255), 1)
    cv2.drawContours(vis_acc, accepted, -1, (0, 255, 255), 1)
    vis_acc = draw_object_ids(vis_acc, accepted)
    save_debug(img_out, "11_contours_rejected_orange_accepted_yellow_ids_green.png", vis_acc, um_per_px_avg)

    mask_all = np.zeros_like(mask)
    cv2.drawContours(mask_all, contours, -1, 255, thickness=-1)
    save_debug(img_out, "12_mask_all.png", mask_all)

    mask_acc = np.zeros_like(mask)
    cv2.drawContours(mask_acc, accepted, -1, 255, thickness=-1)
    save_debug(img_out, "13_mask_accepted.png", mask_acc)

    fluor_path = img_path.parent / img_path.name.replace("_ch00", "_ch01")
    fluor_measurements: Optional[list[dict]] = None

    fluor_bw: Optional[np.ndarray] = None
    vis_fluor: Optional[np.ndarray] = None
    vis_match: Optional[np.ndarray] = None

    if fluor_path.exists():
        # Use Unicode-safe imread for fluorescence
        fluor_img = safe_imread(fluor_path, cv2.IMREAD_UNCHANGED)
        if fluor_img is not None:
            fluor_img, (sy, sx) = align_fluorescence_channel(img, fluor_img)
            
            save_debug(img_out, "20_fluorescence_aligned_raw.png", normalize_to_8bit(fluor_img), um_per_px_avg)

            if fluor_img.ndim == 3:
                fluor_img = cv2.cvtColor(fluor_img, cv2.COLOR_BGR2GRAY)

            fluor_img8 = normalize_to_8bit(fluor_img)
            save_debug(img_out, "20_fluorescence_8bit.png", fluor_img8, um_per_px_avg)

            fluor_bw = segment_fluorescence_global(fluor_img8)
            save_debug(img_out, "22_fluorescence_mask_global.png", fluor_bw, um_per_px_avg)

            _fc2 = cv2.findContours(fluor_bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            fluor_contours_all = cast(list[np.ndarray], _fc2[-2])

            min_fluor_area_px = FLUOR_MIN_AREA_UM2 / um2_per_px2 if um2_per_px2 > 0 else 0.0
            fluor_contours = [c for c in fluor_contours_all if float(cv2.contourArea(c)) >= float(min_fluor_area_px)]

            vis_fluor = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
            cv2.drawContours(vis_fluor, fluor_contours, -1, (0, 255, 0), 1)
            save_debug(img_out, "23_fluorescence_contours_global.png", vis_fluor, um_per_px_avg)

            matches = match_fluor_to_bf_by_overlap(accepted, fluor_contours, fluor_img8.shape[:2])

            fluor_measurements = measure_fluorescence_intensity_with_global_area(
                fluor_img, accepted, fluor_contours, matches, float(um_per_px_x), float(um_per_px_y)
            )

            vis_match = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
            for idx, bf_c in enumerate(accepted):
                j = matches[idx]
                cv2.drawContours(vis_match, [bf_c], -1, (0, 0, 255), 1)
                if j is not None:
                    cv2.drawContours(vis_match, [fluor_contours[j]], -1, (0, 255, 0), 2)
            save_debug(img_out, "24_bf_fluor_matching_overlay.png", vis_match, um_per_px_avg)

            fluor_overlay = visualize_fluorescence_measurements(fluor_img8, accepted, fluor_measurements)
            save_debug(img_out, "21_fluorescence_overlay.png", fluor_overlay, um_per_px_avg)

    parts = img_path.stem.split()
    group_id = parts[0] if parts else "unk"
    sequence_num = parts[-1].split("_")[0] if parts else "0"

    object_ids = [f"{group_id}_{sequence_num}_{i}" for i in range(1, len(accepted) + 1)]

    mask_acc_bgr = cv2.cvtColor(mask_acc, cv2.COLOR_GRAY2BGR)
    save_debug_ids(img_out, "13_mask_accepted.png", mask_acc_bgr, accepted, object_ids, um_per_px_avg)

    if fluor_bw is not None:
        fluor_bw_bgr = cv2.cvtColor(fluor_bw, cv2.COLOR_GRAY2BGR)
        save_debug_ids(img_out, "22_fluorescence_mask_global.png", fluor_bw_bgr, accepted, object_ids, um_per_px_avg)

    if vis_fluor is not None:
        save_debug_ids(img_out, "23_fluorescence_contours_global.png", vis_fluor, accepted, object_ids, um_per_px_avg)

    if vis_match is not None:
        save_debug_ids(img_out, "24_bf_fluor_matching_overlay.png", vis_match, accepted, object_ids, um_per_px_avg)

    csv_path = img_out / "object_stats.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "Object_ID",
                "BF_Area_px",
                "BF_Area_um2",
                "Perimeter_px",
                "Perimeter_um",
                "EquivDiameter_px",
                "EquivDiameter_um",
                "Circularity",
                "AspectRatio",
                "CentroidX_px",
                "CentroidY_px",
                "CentroidX_um",
                "CentroidY_um",
                "BBoxX_px",
                "BBoxY_px",
                "BBoxW_px",
                "BBoxH_px",
                "BBoxW_um",
                "BBoxH_um",
                "Fluor_Area_px",
                "Fluor_Area_um2",
                "Fluor_Mean",
                "Fluor_Median",
                "Fluor_Std",
                "Fluor_Min",
                "Fluor_Max",
                "Fluor_IntegratedDensity",
            ]
        )

        for i, c in enumerate(accepted, 1):
            compound_id = f"{group_id}_{sequence_num}_{i}"

            area_px = float(cv2.contourArea(c))
            area_um2 = area_px * (float(um_per_px_x) * float(um_per_px_y))

            perim_px = float(cv2.arcLength(c, True))
            perim_um = contour_perimeter_um(c, float(um_per_px_x), float(um_per_px_y))

            eqd_px = equivalent_diameter_from_area(area_px)
            eqd_um = equivalent_diameter_from_area(area_um2)

            circ = (4 * np.pi * area_px / (perim_px**2)) if perim_px > 0 else 0.0

            x, y, bw, bh = cv2.boundingRect(c)
            aspect = (float(bw) / float(bh)) if bh > 0 else 0.0

            M = cv2.moments(c)
            if M["m00"] != 0:
                cx = float(M["m10"] / M["m00"])
                cy = float(M["m01"] / M["m00"])
            else:
                cx, cy = 0.0, 0.0

            cx_um = cx * float(um_per_px_x)
            cy_um = cy * float(um_per_px_y)
            bw_um = float(bw) * float(um_per_px_x)
            bh_um = float(bh) * float(um_per_px_y)

            if fluor_measurements is not None:
                fm = fluor_measurements[i - 1]
            else:
                fm = {
                    "fluor_area_px": 0.0,
                    "fluor_area_um2": 0.0,
                    "fluor_mean": 0.0,
                    "fluor_median": 0.0,
                    "fluor_std": 0.0,
                    "fluor_min": 0.0,
                    "fluor_max": 0.0,
                    "fluor_integrated_density": 0.0,
                }

            w.writerow(
                [
                    compound_id,
                    f"{area_px:.2f}",
                    f"{area_um2:.4f}",
                    f"{perim_px:.2f}",
                    f"{perim_um:.4f}",
                    f"{eqd_px:.2f}",
                    f"{eqd_um:.4f}",
                    f"{circ:.4f}",
                    f"{aspect:.4f}",
                    f"{cx:.2f}",
                    f"{cy:.2f}",
                    f"{cx_um:.4f}",
                    f"{cy_um:.4f}",
                    x,
                    y,
                    bw,
                    bh,
                    f"{bw_um:.4f}",
                    f"{bh_um:.4f}",
                    f"{float(fm['fluor_area_px']):.2f}",
                    f"{float(fm['fluor_area_um2']):.4f}",
                    f"{float(fm['fluor_mean']):.2f}",
                    f"{float(fm['fluor_median']):.2f}",
                    f"{float(fm['fluor_std']):.2f}",
                    f"{float(fm['fluor_min']):.2f}",
                    f"{float(fm['fluor_max']):.2f}",
                    f"{float(fm['fluor_integrated_density']):.2f}",
                ]
            )

def open_folder(folder_path: Path) -> None:
    """Open folder in file explorer (cross-platform, Unicode-safe)
    
    Args:
        folder_path: Path to folder to open
    """
    try:
        folder_str = str(folder_path.resolve())
        
        if sys.platform == 'win32':
            # Use os.startfile for better Unicode support on Windows
            os.startfile(folder_str)
        elif sys.platform == 'darwin':  # macOS
            subprocess.run(['open', folder_str])
        else:  # Linux
            subprocess.run(['xdg-open', folder_str])
        
        print(f"  ✓ Opened folder: {folder_path.name}")
    except Exception as e:
        print(f"  ⚠ Could not open folder automatically: {e}")
        print(f"  Please open manually: {folder_path.resolve()}")




def setup_output_directory(config: dict) -> Path:
    """Create and setup output directory structure with timestamp naming.
    
    Args:
        config: Configuration dictionary
        
    Returns:
        Path to the main output directory
    """
    # Resolve the repository root from this file
    project_root = Path(__file__).resolve().parent
    output_root = project_root / 'outputs'
    output_root.mkdir(exist_ok=True)
    
    # Create timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if config.get('batch_mode', False):
        # Batch mode: base_name_timestamp_test
        # Example: v48_PD_20260119_012703_test
        base_name = config.get('dataset_id_base', 'dataset')
        output_dir_name = f"{base_name}_{timestamp}_test"
        
    else:
        # Single mode: base_name_microgel_timestamp_test
        # Example: v48_PD_Negative_20260119_012703_test
        dataset_id = config.get('dataset_id', 'dataset')
        microgel_type = config.get('microgel_type', '')
        
        if microgel_type:
            # Capitalize first letter: negative -> Negative, positive -> Positive
            microgel_label = microgel_type.capitalize()
            output_dir_name = f"{dataset_id}_{microgel_label}_{timestamp}_test"
        else:
            output_dir_name = f"{dataset_id}_{timestamp}_test"
    
    # Create the output directory
    output_dir = output_root / output_dir_name
    output_dir.mkdir(parents=True, exist_ok=True)
    
    return output_dir


def copy_log_to_output(log_path: Path, output_dir: Path) -> Optional[Path]:
    """Copy log file to output directory
    
    Args:
        log_path: Path to the original log file
        output_dir: Output directory to copy to
        
    Returns:
        Path to the copied log file, or None if copy failed
    """
    try:
        if log_path and log_path.exists():
            dest_path = output_dir / log_path.name
            shutil.copy2(log_path, dest_path)
            return dest_path
        return None
    except Exception as e:
        print(f"Warning: Could not copy log file: {e}")
        return None


def check_log_for_errors(log_path: Path) -> Dict[str, Any]:
    """Check log file for errors and warnings
    
    Args:
        log_path: Path to log file
        
    Returns:
        Dictionary with error/warning counts and details
    """
    result = {
        'errors': [],
        'warnings': [],
        'error_count': 0,
        'warning_count': 0
    }
    
    try:
        if not log_path or not log_path.exists():
            return result
            
        with open(log_path, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                line_lower = line.lower()
                
                # Check for errors
                if any(keyword in line_lower for keyword in ['error', 'exception', 'traceback', 'failed']):
                    if not any(skip in line_lower for skip in ['0 failed', 'no error', 'error_count']):
                        result['errors'].append((line_num, line.strip()))
                        result['error_count'] += 1
                
                # Check for warnings
                elif 'warning' in line_lower or '⚠' in line:
                    result['warnings'].append((line_num, line.strip()))
                    result['warning_count'] += 1
                    
    except Exception as e:
        print(f"Could not analyze log file: {e}")
    
    return result


def display_log_analysis(log_analysis: Dict, log_path: Path):
    """Display log analysis results
    
    Args:
        log_analysis: Results from check_log_for_errors
        log_path: Path to the log file
    """
    print("\n" + "="*80)
    print("LOG FILE ANALYSIS")
    print("="*80)
    print(f"Log file: {log_path.name}\n")
    
    if log_analysis['error_count'] == 0 and log_analysis['warning_count'] == 0:
        print("No errors or warnings found")
    else:
        if log_analysis['error_count'] > 0:
            print(f"⚠ Found {log_analysis['error_count']} error(s):")
            for line_num, line in log_analysis['errors'][:5]:  # Show first 5
                print(f"  Line {line_num}: {line[:100]}")
            if len(log_analysis['errors']) > 5:
                print(f"  ... and {len(log_analysis['errors']) - 5} more")
            print()
        
        if log_analysis['warning_count'] > 0:
            print(f"⚠ Found {log_analysis['warning_count']} warning(s):")
            for line_num, line in log_analysis['warnings'][:5]:  # Show first 5
                print(f"  Line {line_num}: {line[:100]}")
            if len(log_analysis['warnings']) > 5:
                print(f"  ... and {len(log_analysis['warnings']) - 5} more")
    
    print("="*80)


def process_single_dataset(config: dict) -> dict:
    """Process a single dataset with given configuration
    
    Args:
        config: Configuration dictionary containing all settings
        
    Returns:
        dict: Processing results including success status
    """
    
    try:
        # Extract configuration
        source_dir = Path(config['source_dir'])
        dataset_id = config['dataset_id']
        percentile = config['percentile']
        microgel_type = config['microgel_type']
        bacteria_profile = config.get('bacteria_profile', 'default')  # NEW
        threshold_mode = config.get('threshold_mode', 'adaptive')  # NEW
        confidence_level = config.get('confidence_level', 0.95)  # NEW
        threshold_pct = config['threshold_pct']
        
        # Setup output directory
        output_dir = config.get('output_dir')
        if output_dir is None:
            project_root = Path(__file__).resolve().parent
            output_root = project_root / 'outputs'
            output_root.mkdir(exist_ok=True)
            output_dir = output_root / dataset_id
            output_dir.mkdir(exist_ok=True)
        else:
            output_dir = Path(output_dir)
            output_dir.mkdir(exist_ok=True)
        
        # Step 1: Collect images
        print("━" * 80)
        print("STEP 1/7: Collecting Images")
        print("━" * 80)
        
        group_folders = list_sample_group_folders(source_dir)
        
        # Find control folder
        control_folder = None
        for folder in source_dir.iterdir():
            if folder.is_dir() and folder.name.lower().startswith("control"):
                control_folder = folder
                break
        
        all_groups = group_folders + ([control_folder] if control_folder else [])
        
        total_images = sum(len(list(g.glob(IMAGE_GLOB))) for g in all_groups if g)
        print(f"  Found {total_images} images across {len(all_groups)} groups")
        
        # Step 2: Process images
        print("\n" + "━" * 80)
        print("STEP 2/7: Processing Images")
        print("━" * 80)
        
        all_image_paths = []
        for group in all_groups:
            if group:
                all_image_paths.extend(sorted(group.glob(IMAGE_GLOB)))
        
        success_count = 0
        fail_count = 0
        
        with tqdm(total=len(all_image_paths), desc="  Processing", unit="img") as pbar:
            for img_path in all_image_paths:
                try:
                    # Process each image into group-specific output folder
                    group_output_dir = output_dir / img_path.parent.name
                    ensure_dir(group_output_dir)
                    process_image(img_path, group_output_dir)
                    success_count += 1
                except Exception as e:
                    print(f"\n[ERROR] Failed to process {img_path.name}: {e}")
                    fail_count += 1
                finally:
                    pbar.update(1)
        
        print(f"\n  Processed: {success_count} succeeded, {fail_count} failed")
        
        # Step 3: Consolidate to Excel
        print("\n" + "━" * 80)
        print("STEP 3/7: Consolidating to Excel")
        print("━" * 80)
        
        # Consolidate each group folder
        for group_dir in sorted(output_dir.iterdir()):
            if group_dir.is_dir() and len(list(group_dir.glob("*/object_stats.csv"))) > 0:
                group_name = group_dir.name
                display_name = _display_group_name(group_name)
                print(f"  → Consolidating {display_name}...")
                consolidate_to_excel(group_dir, group_name, percentile)
        
        print("  Excel consolidation complete")
        
        # Step 4: Generate comparison plots
        print("\n" + "━" * 80)
        print("STEP 4/7: Generating Comparison Plots")
        print("━" * 80)
        
        print("  → Generating pairwise plots...")
        generate_pairwise_group_vs_control_plots(
            output_dir, 
            percentile, 
            dataset_id,
            threshold_pct,
            microgel_type
        )
        
        print("  → Generating all-groups plot...")
        all_groups_plot = generate_error_bar_comparison_with_threshold(
            output_dir=output_dir,
            percentile=percentile,
            restrict_to_groups=None,
            output_path=None,
            title_suffix="",
            dataset_id=dataset_id,
            threshold_pct=threshold_pct,
            microgel_type=microgel_type,
        )
        print("  Plots generated")
        
        # Step 5: Embed plots into Excel
        print("\n" + "━" * 80)
        print("STEP 5/7: Embedding Plots into Excel")
        print("━" * 80)
        
        # Embed plots into each group's Excel file
        for group_dir in sorted(output_dir.iterdir()):
            if not group_dir.is_dir():
                continue

            if group_dir.name.lower().startswith("control"):
                plot = all_groups_plot
            elif re.fullmatch(r"\d+", group_dir.name):
                plot = group_dir / f"Group_{group_dir.name}_vs_Control_threshold.png"
            else:
                continue

            if plot and plot.exists():
                embed_comparison_plots_into_all_excels(group_dir, percentile, plot_path=plot)
        
        print("  Plots embedded")
        
        # Step 6: Export statistics
        print("\n" + "━" * 80)
        print("STEP 6/7: Exporting Statistics")
        print("━" * 80)
        
        export_group_statistics_to_csv(output_dir)
        print("  Statistics exported")
        
        # Step 7: Clinical classification
        print("\n" + "━" * 80)
        print("STEP 7/7: Clinical Classification")
        print("━" * 80)
        
        classification_df = classify_groups_clinical(
            output_dir,
            microgel_type=microgel_type,
            threshold_pct=threshold_pct,
            bacteria_profile=bacteria_profile,
            threshold_mode=threshold_mode,
            confidence_level=confidence_level
        )
        
        if not classification_df.empty:
            export_clinical_classification(
                output_dir,
                classification_df,
                microgel_type=microgel_type,
                bacteria_profile=bacteria_profile
            )
            print("  Classification complete")
        else:
            print("  ⚠ No classification data available")
        
        print()
        
        return {
            'success': True,
            'output_dir': output_dir,
            'dataset_id': dataset_id,
            'images_processed': success_count,
            'images_failed': fail_count
        }
        
    except Exception as e:
        print(f"\nProcessing failed: {e}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e),
            'dataset_id': config.get('dataset_id', 'Unknown')
        }

def launch_results_viewer(output_dir: Optional[Path] = None):
    """Launch GUI viewer for results
    
    Args:
        output_dir: Optional path to output directory to load automatically
    """
    try:
        # Check if gui_viewer.py exists
        viewer_path = Path(__file__).parent / "gui_viewer.py"
        
        if not viewer_path.exists():
            print("  ⚠ GUI viewer not found (gui_viewer.py)")
            print(f"    Expected location: {viewer_path}")
            return False
        
        print("  🚀 Launching GUI viewer...")
        
        # Launch as separate process
        if getattr(sys, 'frozen', False):
            # Running as compiled executable
            # GUI should be bundled
            try:
                from image_viewer import launch_viewer
                launch_viewer(output_dir)
                return True
            except ImportError:
                print("  ⚠ GUI viewer not available in executable")
                return False
        else:
            # Running as script - launch in new process
            cmd = [sys.executable, str(viewer_path)]
            
            # Pass output directory as argument if provided
            if output_dir:
                cmd.append(str(output_dir))
            
            subprocess.Popen(cmd)
            print("  ✓ GUI viewer launched in separate window")
            return True
            
    except Exception as e:
        print(f"  ⚠ Could not launch GUI viewer: {e}")
        return False

def get_user_inputs():
    """Get all configuration inputs from user
    
    Returns:
        dict: Configuration dictionary with all settings
        None: If user cancels
    """
    try:
        print("\n" + "─" * 80)
        print("DATASET CONFIGURATION")
        print("─" * 80)
        
        # Dataset ID
        dataset_id = logged_input("\nEnter dataset name/ID: ").strip()
        if not dataset_id:
            print("  ✗ Dataset ID cannot be empty")
            return None
        
        # Image directory
        print("\n" + "─" * 40)
        print("Image Directory")
        print("─" * 40)
        image_dir_input = logged_input("Enter image directory path (or press Enter for current directory): ").strip()
        
        if image_dir_input:
            image_dir = Path(image_dir_input)
        else:
            image_dir = Path.cwd()
        
        if not image_dir.exists():
            print(f"  ✗ Directory not found: {image_dir}")
            return None
        
        print(f"  ✓ Using: {image_dir}")
        
        # Excel file
        print("\n" + "─" * 40)
        print("Excel Data File")
        print("─" * 40)
        excel_path_input = logged_input("Enter Excel file path: ").strip()
        
        if not excel_path_input:
            print("  ✗ Excel path cannot be empty")
            return None
        
        excel_path = Path(excel_path_input)
        
        if not excel_path.exists():
            print(f"  ✗ File not found: {excel_path}")
            return None
        
        print(f"  ✓ Using: {excel_path}")
        
        # Control groups
        print("\n" + "─" * 40)
        print("Control Groups")
        print("─" * 40)
        control_input = logged_input("Enter control group IDs (comma-separated, e.g., 1,2,3): ").strip()
        
        if not control_input:
            print("  ✗ Control groups cannot be empty")
            return None
        
        try:
            control_groups = [int(x.strip()) for x in control_input.split(',')]
            print(f"  ✓ Control groups: {control_groups}")
        except ValueError:
            print("  ✗ Invalid control group format")
            return None
        
        # Threshold
        print("\n" + "─" * 40)
        print("Detection Threshold")
        print("─" * 40)
        threshold_input = logged_input("Enter number of standard deviations for threshold (Enter=2.0): ").strip()
        
        if threshold_input:
            try:
                num_std_threshold = float(threshold_input)
            except ValueError:
                print("  ✗ Invalid threshold value, using default 2.0")
                num_std_threshold = 2.0
        else:
            num_std_threshold = 2.0
        
        print(f"  ✓ Threshold: {num_std_threshold} std deviations")
        
        # Auto-open folder
        auto_open_input = logged_input("\nAuto-open output folder when complete? (y/n, Enter=yes): ").strip().lower()
        auto_open = auto_open_input in ["", "y", "yes"]
        
        # Return configuration
        config = {
            'dataset_id': dataset_id,
            'image_dir': image_dir,
            'excel_path': excel_path,
            'control_groups': control_groups,
            'num_std_threshold': num_std_threshold,
            'auto_open': auto_open
        }
        
        return config
        
    except KeyboardInterrupt:
        print("\n\n  ✗ Configuration cancelled by user")
        return None
    except Exception as e:
        print(f"\n  ✗ Error during configuration: {e}")
        return None
    
def process_pipeline(config):
    """Execute the full processing pipeline
    
    Args:
        config: Configuration dictionary
        
    Returns:
        dict: Results dictionary with success status and output paths
    """
    try:
        # Extract configuration
        dataset_id = config['dataset_id']
        image_dir = config['image_dir']
        excel_path = config['excel_path']
        control_groups = config['control_groups']
        num_std_threshold = config['num_std_threshold']
        output_dir = config.get('output_dir', Path('outputs') / f"{dataset_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        microgel_type = config.get('microgel_type', 'positive')
        
        # Ensure output directory exists
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        result = {
            'success': False,
            'output_dir': output_dir,
            'dataset_id': dataset_id,
            'errors': []
        }
        
        # Step 1: Load and process images
        print("\n" + "─" * 80)
        print("STEP 1: IMAGE PROCESSING")
        print("─" * 80)
        
        # Load Excel data
        print(f"\nLoading Excel data from: {excel_path.name}")
        excel_data = pd.read_excel(excel_path)
        print(f"  ✓ Loaded {len(excel_data)} rows")
        
        # Process images and extract fluorescence
        print("\nProcessing images...")
        fluorescence_results = []
        
        for idx, row in excel_data.iterrows():
            group = row.get('Group', idx)
            image_name = row.get('Image', f"image_{idx}.png")
            
            image_path = image_dir / image_name
            
            if not image_path.exists():
                print(f"  ⚠ Image not found: {image_name}")
                continue
            
            # Extract fluorescence (simplified - you'd use your actual extraction logic)
            fluorescence = extract_fluorescence(image_path)
            
            fluorescence_results.append({
                'Group': group,
                'Image': image_name,
                'Fluorescence': fluorescence
            })
        
        print(f"  ✓ Processed {len(fluorescence_results)} images")
        
        # Step 2: Statistical analysis
        print("\n" + "─" * 80)
        print("STEP 2: STATISTICAL ANALYSIS")
        print("─" * 80)
        
        df = pd.DataFrame(fluorescence_results)
        
        # Calculate statistics
        stats = df.groupby('Group')['Fluorescence'].agg(['mean', 'std', 'count']).reset_index()
        
        # Control statistics
        control_data = df[df['Group'].isin(control_groups)]
        control_mean = control_data['Fluorescence'].mean()
        control_std = control_data['Fluorescence'].std()
        threshold = control_mean + (num_std_threshold * control_std)
        
        print(f"\nControl statistics:")
        print(f"  Mean: {control_mean:.2f}")
        print(f"  Std Dev: {control_std:.2f}")
        print(f"  Threshold: {threshold:.2f}")
        
        # Step 3: Clinical classification
        print("\n" + "─" * 80)
        print("STEP 3: CLINICAL CLASSIFICATION")
        print("─" * 80)
        
        def classify(mean_val):
            if mean_val > threshold:
                return "POSITIVE"
            elif mean_val < control_mean - control_std:
                return "NEGATIVE"
            else:
                return "NO OBVIOUS BACTERIA"
        
        stats['Classification'] = stats['mean'].apply(classify)
        stats['Threshold'] = threshold
        stats['Control_Mean'] = control_mean
        
        # Save results
        classification_file = output_dir / f"clinical_classification_{microgel_type}.csv"
        stats.to_csv(classification_file, index=False)
        print(f"\n  ✓ Classification saved: {classification_file.name}")
        
        # Save raw data
        raw_data_file = output_dir / f"fluorescence_data_{microgel_type}.csv"
        df.to_csv(raw_data_file, index=False)
        print(f"  ✓ Raw data saved: {raw_data_file.name}")
        
        # Step 4: Generate plots
        print("\n" + "─" * 80)
        print("STEP 4: GENERATING PLOTS")
        print("─" * 80)
        
        # Create comparison plot
        fig, ax = plt.subplots(figsize=(12, 6))
        
        x = range(len(stats))
        colors_map = {'POSITIVE': 'red', 'NEGATIVE': 'green', 'NO OBVIOUS BACTERIA': 'yellow'}
        colors = [colors_map[c] for c in stats['Classification']]
        
        ax.bar(x, stats['mean'], color=colors, alpha=0.6, edgecolor='black')
        ax.axhline(threshold, color='red', linestyle='--', label='Threshold')
        ax.axhline(control_mean, color='blue', linestyle='--', label='Control Mean')
        
        ax.set_xlabel('Group')
        ax.set_ylabel('Mean Fluorescence')
        ax.set_title(f'{microgel_type.upper()} Microgel - Clinical Classification')
        ax.legend()
        
        plot_file = output_dir / f"classification_plot_{microgel_type}.png"
        plt.savefig(plot_file, dpi=300, bbox_inches='tight')
        plt.close()
        
        print(f"  ✓ Plot saved: {plot_file.name}")
        
        # Success
        result['success'] = True
        result['classification_file'] = classification_file
        result['plot_file'] = plot_file
        
        return result
        
    except Exception as e:
        print(f"\n✗ Pipeline error: {e}")
        import traceback
        traceback.print_exc()
        result['errors'].append(str(e))
        return result


def extract_fluorescence(image_path):
    """Extract fluorescence from image (simplified placeholder)
    
    Args:
        image_path: Path to image file
        
    Returns:
        float: Fluorescence intensity value
    """
    # This is a simplified placeholder
    # Replace with your actual fluorescence extraction logic
    try:
        img = cv2.imread(str(image_path))
        if img is None:
            return 0.0
        
        # Simple mean intensity as placeholder
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        return float(gray.mean())  # Use ndarray's mean() method instead
        
    except Exception:
        return 0.0

# ==================================================
# Main Function
# ==================================================
def main():
    """Main execution function"""
    
    # ==================== INITIALIZATION ====================
    print("\n" + "="*80)
    print("MICROGEL FLUORESCENCE ANALYSIS PIPELINE")
    print("="*80 + "\n")
    
    try:
        # Collect configuration (handles both single and batch mode detection)
        config = collect_configuration()
        
        # Display summary and get confirmation
        display_configuration_summary(config)
        
        # Setup main output directory
        output_dir = setup_output_directory(config)
        print(f"\n📁 Output directory: {output_dir.relative_to(Path.cwd())}")
        
        # ==================== BATCH MODE ====================
        if config.get('batch_mode', False):
            print("\n" + "="*80)
            print("BATCH PROCESSING MODE")
            print("="*80)
            
            # Process each subdirectory (G+ and G-)
            all_results = []
            
            for subdir_config in config['subdirs']:
                print("\n" + "━" * 80)
                print(f"PROCESSING {subdir_config['label']} MICROGEL")
                print("━" * 80)
                
                # Create dataset-specific configuration
                dataset_id = f"{config['dataset_id_base']} {subdir_config['safe_label']}"
                
                single_config = {
                    'source_dir': subdir_config['path'],
                    'dataset_id': dataset_id,
                    'percentile': config['percentile'],
                    'microgel_type': subdir_config['microgel_type'],
                    'threshold_pct': config['threshold_pct'],
                    'output_dir': output_dir / subdir_config['safe_label']
                }
                
                # Process this dataset
                result = process_single_dataset(single_config)
                
                all_results.append({
                    'label': subdir_config['label'],
                    'result': result
                })
                
                if result['success']:
                    print(f"\n✓ {subdir_config['label']} processing completed")
                else:
                    print(f"\n✗ {subdir_config['label']} processing failed")
            
            # ==================== GENERATE FINAL MATRIX ====================
            print("\n" + "="*80)
            print("GENERATING FINAL CLINICAL MATRIX")
            print("="*80)
            
            # Check if both succeeded
            if all(item['result']['success'] for item in all_results):
                # Load classification results
                gplus_output = output_dir / "Positive"
                gminus_output = output_dir / "Negative"
                
                gplus_csv_files = list(gplus_output.glob("clinical_classification_positive*.csv"))
                gminus_csv_files = list(gminus_output.glob("clinical_classification_negative*.csv"))

                gplus_csv_path = gplus_csv_files[0] if gplus_csv_files else None
                gminus_csv_path = gminus_csv_files[0] if gminus_csv_files else None

                if gplus_csv_path and gminus_csv_path:
                    print("\n  Loading classification results...")
                    gplus_df = pd.read_csv(gplus_csv_path)
                    gminus_df = pd.read_csv(gminus_csv_path)
                    
                    print("  Generating final clinical matrix...")
                    # Generate final matrix
                    final_matrix = generate_final_clinical_matrix(
                        output_root=output_dir,
                        gplus_classification=gplus_df,
                        gminus_classification=gminus_df,
                        dataset_base_name=config['dataset_id_base']
                    )
                    
                    if final_matrix:
                        print(f"\n✓ Final clinical matrix: {final_matrix.name}")
                    else:
                        print("\n⚠ Could not generate final matrix")
                else:
                    print("\n⚠ Missing classification files - cannot generate final matrix")
                    if not gplus_csv_path:
                        print(f"    Missing: G+ classification CSV in {gplus_output}")
                    if not gminus_csv_path:
                        print(f"    Missing: G- classification CSV in {gminus_output}")
            else:
                print("\n⚠ Some processing failed - skipping final matrix")
            
            # ==================== COMPLETION ====================
            print("\n" + "="*80)
            print("BATCH PROCESSING COMPLETE")
            print("="*80)
            
            for item in all_results:
                status = "✓" if item['result']['success'] else "✗"
                dataset_name = item['result'].get('dataset_id', item['label'])
                print(f"  {status} {item['label']}: {dataset_name}")
            
            print()
        
        # ==================== SINGLE MODE ====================
        else:
            print("\n" + "="*80)
            print("SINGLE PROCESSING MODE")
            print("="*80)
            
            single_config = {
                'source_dir': config['source_dir'],
                'dataset_id': config['dataset_id'],
                'percentile': config['percentile'],
                'microgel_type': config['microgel_type'],
                'threshold_pct': config['threshold_pct'],
                'output_dir': output_dir
            }
            
            result = process_single_dataset(single_config)
            
            if result['success']:
                print("\n✓ Processing completed successfully")
                print(f"   Dataset: {result.get('dataset_id', 'Unknown')}")
                print(f"   Images processed: {result.get('images_processed', 0)}")
                if result.get('images_failed', 0) > 0:
                    print(f"   Images failed: {result.get('images_failed', 0)}")
            else:
                print("\n✗ Processing failed")
                if 'error' in result:
                    print(f"   Error: {result['error']}")
        
        # ==================== OPEN RESULTS ====================
        print("\n" + "━" * 80)
        print("RESULTS")
        print("━" * 80)
        
        # Copy log file to output
        if _log_path and _log_path.exists():
            copied_log = copy_log_to_output(_log_path, output_dir)
            if copied_log:
                print(f"\n✓ Log file copied to output directory")
        
        # Analyze log for errors
        if _log_path and _log_path.exists():
            log_analysis = check_log_for_errors(_log_path)
            
            if log_analysis['error_count'] > 0 or log_analysis['warning_count'] > 0:
                print(f"\n⚠ Log analysis:")
                if log_analysis['error_count'] > 0:
                    print(f"   Errors found: {log_analysis['error_count']}")
                if log_analysis['warning_count'] > 0:
                    print(f"   Warnings found: {log_analysis['warning_count']}")
                print(f"   See log file for details: {_log_path.name}")
        
        # Offer to open folder or launch viewer
        print()
        '''
        view_choice = logged_input("Open results? (1=Folder, 2=GUI Viewer, Enter=Folder): ").strip()
        
        if view_choice == "2":
            if not launch_results_viewer(output_dir):
                print("  Falling back to folder view...")
                open_folder(output_dir)
        else:
            open_folder(output_dir)
        '''
        open_folder(output_dir)

        print("\n" + "="*80)
        print("PROCESSING COMPLETE")
        print("="*80 + "\n")
        
    except SystemExit:
        print("\n\nProgram terminated by user.\n")
    except KeyboardInterrupt:
        print("\n\nProgram interrupted by user.\n")
    except Exception as e:
        print(f"\n\nFatal error: {e}")
        import traceback
        traceback.print_exc()
        print("\nPlease check the log file for details.\n")


if __name__ == "__main__":
    main()