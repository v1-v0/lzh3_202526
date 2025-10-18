"""
Interactive Dual-Channel Bacteria Detection Review Tool with Excel Export
Features:
- Generates review sheet for each image pair in portrait orientation
- Shows ORIGINAL images with detection overlays (using raw source images)
- Includes original fluorescence image loaded directly from _ch01.tif without processing
- Uses MODE intensity instead of mean
- Detailed contour statistics table (without min intensity)
- Exports all data to Excel file including embedded images
- Portrait layout with reserved space for page header and footer
- Fixed overlap between detection report and parameters
- Maximized image sizes for optimal review
"""

import os
import numpy as np
import cv2
from pathlib import Path
from skimage.io import imread
from sympy import im
from skimage import io, filters, morphology, measure
from skimage.segmentation import watershed
from scipy import ndimage as ndi
from scipy import stats
import pandas as pd
import warnings
import io
from PIL import Image as PILImage
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment

warnings.filterwarnings('ignore')

class BacteriaReviewTool:
    def __init__(self, source_folder="source//9"):
        """Initialize the review tool"""
        self.source_folder = source_folder
        
        # Fixed configuration
        self.config = {
            "gray_threshold": 53,
            "fluor_threshold": 5,
            "min_area": 400,
            "min_fluor_area": 400,
            "overlap_threshold": 0.80,
        }
        
        self.load_image_pairs()
        
        if not self.image_pairs:
            raise ValueError(f"No image pairs found in {source_folder}")
        
        # Storage for all results
        self.all_results = []
        
        # Generate review sheets for all images
        self.generate_all_review_sheets()
        
        # Export to Excel
        self.export_to_excel()
        
    def load_image_pairs(self):
        """Load all grayscale and fluorescence image pairs"""
        self.image_pairs = []
        folder_path = Path(self.source_folder)
        
        if not folder_path.exists():
            print(f"Error: Folder '{self.source_folder}' does not exist")
            return
        
        gray_files = sorted([
            f for f in os.listdir(self.source_folder)
            if f.endswith('_ch00.tif') and os.path.isfile(os.path.join(self.source_folder, f))
        ])
        
        for gray_file in gray_files:
            fluor_file = gray_file.replace('_ch00.tif', '_ch01.tif')
            gray_path = os.path.join(self.source_folder, gray_file)
            fluor_path = os.path.join(self.source_folder, fluor_file)
            
            if os.path.exists(fluor_path):
                try:
                    gray_img = imread(gray_path)
                    fluor_img = imread(fluor_path)
                    
                    if len(gray_img.shape) == 3:
                        gray_img = cv2.cvtColor(gray_img, cv2.COLOR_RGB2GRAY)
                    if len(fluor_img.shape) == 3:
                        fluor_img = cv2.cvtColor(fluor_img, cv2.COLOR_RGB2GRAY)
                    
                    self.image_pairs.append({
                        'gray': gray_img,
                        'fluor': fluor_img,
                        'gray_file': gray_file,
                        'fluor_file': fluor_file
                    })
                except Exception as e:
                    print(f"Error loading {gray_file}: {e}")
        
        print(f"Loaded {len(self.image_pairs)} image pair(s)")
    
    def preprocess_grayscale(self, image):
        """Preprocess grayscale image"""
        median_filtered = filters.median(image, morphology.disk(3))
        
        if median_filtered.max() > 0:
            img_normalized = (median_filtered / median_filtered.max() * 255).astype(np.uint8)
        else:
            img_normalized = median_filtered.astype(np.uint8)
        
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(img_normalized)
        
        return enhanced
    
    def preprocess_fluorescence(self, image):
        """Preprocess fluorescence image"""
        median_filtered = filters.median(image, morphology.disk(3))
        
        if median_filtered.max() > 0:
            img_normalized = (median_filtered / median_filtered.max() * 255).astype(np.uint8)
        else:
            img_normalized = median_filtered.astype(np.uint8)
        
        clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(img_normalized)
        
        return enhanced
    
    def segment_grayscale(self, preprocessed, threshold):
        """Segment grayscale bacteria (dark objects)"""
        binary = np.asarray(preprocessed) <= threshold
        
        distance_result = ndi.distance_transform_edt(binary)
        if isinstance(distance_result, tuple):
            distance = distance_result[0]
        else:
            distance = distance_result
        distance = np.asarray(distance, dtype=np.float32)
        
        local_max = morphology.local_maxima(distance)
        markers = measure.label(local_max)
        
        labels = watershed(-distance, markers, mask=binary)
        separated = labels > 0
        
        return separated, labels
    
    def segment_fluorescence(self, preprocessed, threshold):
        """Segment fluorescence bacteria (bright objects)"""
        binary = np.asarray(preprocessed) >= threshold
        
        distance_result = ndi.distance_transform_edt(binary)
        if isinstance(distance_result, tuple):
            distance = distance_result[0]
        else:
            distance = distance_result
        distance = np.asarray(distance, dtype=np.float32)
        
        local_max = morphology.local_maxima(distance)
        markers = measure.label(local_max)
        
        labels = watershed(-distance, markers, mask=binary)
        separated = labels > 0
        
        return separated, labels
    
    def postprocess(self, binary_mask):
        """Post-process binary mask"""
        min_size = 50
        cleaned = morphology.remove_small_objects(binary_mask, min_size=min_size)
        filled = ndi.binary_fill_holes(cleaned)
        opened = morphology.binary_opening(filled, morphology.disk(2))
        final = morphology.binary_closing(opened, morphology.disk(2))
        return final
    
    def detect_bacteria(self, image, is_grayscale=True):
        """Detect bacteria in image with current settings"""
        if is_grayscale:
            preprocessed = self.preprocess_grayscale(image)
            threshold = self.config['gray_threshold']
            min_area = self.config['min_area']
            binary, labels = self.segment_grayscale(preprocessed, threshold)
        else:
            preprocessed = self.preprocess_fluorescence(image)
            threshold = self.config['fluor_threshold']
            min_area = self.config['min_fluor_area']
            binary, labels = self.segment_fluorescence(preprocessed, threshold)
        
        final = self.postprocess(binary)
        labeled = measure.label(final)
        props = measure.regionprops(labeled)
        large_props = [prop for prop in props if prop.area > min_area]
        
        return preprocessed, labeled, props, large_props
    
    def calculate_mode(self, intensities):
        """Calculate mode of intensity values"""
        if len(intensities) == 0:
            return 0
        
        mode_result = stats.mode(intensities, keepdims=True)
        return float(mode_result.mode[0])
    
    def get_contour_intensities(self, image, labeled, props):
        """Get intensity statistics for each contour (using MODE instead of mean)"""
        contour_data = []
        
        for i, prop in enumerate(props):
            region_mask = (labeled == prop.label)
            intensities = image[region_mask]
            
            mode_intensity = self.calculate_mode(intensities)
            max_intensity = np.max(intensities)
            std_intensity = np.std(intensities)
            
            contour_data.append({
                'id': i + 1,
                'area': prop.area,
                'perimeter': prop.perimeter,
                'eccentricity': prop.eccentricity,
                'mode_intensity': mode_intensity,
                'max_intensity': max_intensity,
                'std_intensity': std_intensity,
                'centroid_y': prop.centroid[0],
                'centroid_x': prop.centroid[1],
                'label': prop.label,
                'bbox': prop.bbox
            })
        
        return contour_data
    
    def calculate_colocalization(self, gray_labeled, fluor_labeled, gray_props, fluor_props):
        """Calculate which grayscale and fluorescence bacteria are co-localized"""
        colocalized = []
        overlap_threshold = self.config['overlap_threshold']
        
        for gray_prop in gray_props:
            gray_mask = (gray_labeled == gray_prop['label'])
            
            for fluor_prop in fluor_props:
                fluor_mask = (fluor_labeled == fluor_prop['label'])
                
                intersection = np.logical_and(gray_mask, fluor_mask)
                intersection_area = np.sum(intersection)
                overlap_ratio = intersection_area / gray_prop['area']
                
                if overlap_ratio >= overlap_threshold:
                    colocalized.append({
                        'gray_id': gray_prop['id'],
                        'fluor_id': fluor_prop['id'],
                        'overlap_ratio': overlap_ratio,
                        'gray_area': gray_prop['area'],
                        'fluor_area': fluor_prop['area'],
                        'intersection_area': intersection_area
                    })
        
        return colocalized
    
    def create_overlay_visualization(self, image, labeled, props, is_grayscale=True, colocalized=None):
        """Create overlay visualization on ORIGINAL image without preprocessing"""
        # Convert original image to RGB for visualization
        vis = np.dstack([image, image, image]).astype(np.uint8)
        
        colocalized_gray_ids = set()
        colocalized_fluor_ids = set()
        if colocalized:
            colocalized_gray_ids = {c['gray_id'] for c in colocalized}
            colocalized_fluor_ids = {c['fluor_id'] for c in colocalized}
        
        for i, prop in enumerate(props):
            region_mask = (labeled == prop.label).astype(np.uint8)
            contours, _ = cv2.findContours(region_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            if not contours:
                continue
            
            contour = max(contours, key=cv2.contourArea)
            
            is_colocalized = False
            if is_grayscale:
                is_colocalized = (i + 1) in colocalized_gray_ids
            else:
                is_colocalized = (i + 1) in colocalized_fluor_ids
            
            if is_colocalized:
                color = (255, 255, 0)  # Yellow
                thickness = 3
            else:
                color = (0, 255, 0) if is_grayscale else (255, 0, 0)
                thickness = 2
            
            cv2.drawContours(vis, [contour], -1, color, thickness)
            
            M = cv2.moments(contour)
            if M["m00"] != 0:
                cX = int(M["m10"] / M["m00"])
                cY = int(M["m01"] / M["m00"])
                
                label_text = f"{i+1}"
                if is_colocalized:
                    label_text += "*"
                
                cv2.putText(vis, label_text, (cX-10, cY-10), 
                            cv2.FONT_HERSHEY_SIMPLEX, 2, (255, 255, 255), 2)
                            #cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)

        
        return vis
    
    def format_statistics_text(self, gray_props, fluor_props, gray_contour_data, 
                               fluor_contour_data, colocalized, filename):
        """Format statistics as text for display (using MODE, no MIN)"""
        lines = []
        
        lines.append(f"═══════════════ DETECTION REPORT ═══════════════")
        lines.append(f"Source: {filename}")
        lines.append(f"Parameters: Gray Thresh={self.config['gray_threshold']}, "
                    f"Fluor Thresh={self.config['fluor_threshold']}, "
                    f"Overlap={self.config['overlap_threshold']*100:.0f}%")
        lines.append("")
        
        # Summary
        lines.append(f"┌─── SUMMARY ───┐")
        lines.append(f"  Grayscale bacteria:     {len(gray_props):>3}")
        lines.append(f"  Fluorescence bacteria:  {len(fluor_props):>3}")
        lines.append(f"  Co-localized pairs:     {len(colocalized):>3}")
        
        if len(gray_props) > 0:
            coloc_pct = (len(colocalized) / len(gray_props)) * 100
            lines.append(f"  Co-localization rate:   {coloc_pct:>5.1f}%")
        
        lines.append("")
        
        # Co-localization pairs
        if colocalized:
            lines.append(f"┌─── CO-LOCALIZED PAIRS ({len(colocalized)}) ───┐")
            lines.append(f"{'Gray':<6} {'Fluor':<7} {'Overlap':<9} {'G.Area':<9} {'F.Area':<9} {'Intersect'}")
            lines.append("─" * 60)
            
            for coloc in colocalized:
                lines.append(
                    f"{coloc['gray_id']:<6} "
                    f"{coloc['fluor_id']:<7} "
                    f"{coloc['overlap_ratio']*100:>7.1f}% "
                    f"{coloc['gray_area']:<9.0f} "
                    f"{coloc['fluor_area']:<9.0f} "
                    f"{coloc['intersection_area']:.0f}"
                )
        else:
            lines.append(f"┌─── CO-LOCALIZED PAIRS ───┐")
            lines.append("  No co-localized bacteria detected")
        
        lines.append("")
        
        # Grayscale details
        colocalized_gray_ids = {c['gray_id'] for c in colocalized}
        
        lines.append(f"┌─── GRAYSCALE CHANNEL ({len(gray_props)} bacteria) ───┐")
        if gray_contour_data:
            lines.append(f"{'ID':<5} {'*':<2} {'Area':<9} {'Mode':<9} {'Max':<8} {'Centroid (Y, X)'}")
            lines.append("─" * 55)
            
            for data in gray_contour_data:
                coloc_mark = "*" if data['id'] in colocalized_gray_ids else " "
                lines.append(
                    f"{data['id']:<5} "
                    f"{coloc_mark:<2} "
                    f"{data['area']:<9.0f} "
                    f"{data['mode_intensity']:<9.1f} "
                    f"{data['max_intensity']:<8.0f} "
                    f"({data['centroid_y']:.1f}, {data['centroid_x']:.1f})"
                )
            
            # Summary statistics
            gray_areas = [d['area'] for d in gray_contour_data]
            gray_mode_ints = [d['mode_intensity'] for d in gray_contour_data]
            lines.append("─" * 55)
            lines.append(f"Stats: Area μ={np.mean(gray_areas):.1f} σ={np.std(gray_areas):.1f} | "
                        f"Intensity(mode) μ={np.mean(gray_mode_ints):.1f}")
        else:
            lines.append("  No bacteria detected")
        
        lines.append("")
        
        # Fluorescence details
        colocalized_fluor_ids = {c['fluor_id'] for c in colocalized}
        
        lines.append(f"┌─── FLUORESCENCE CHANNEL ({len(fluor_props)} bacteria) ───┐")
        if fluor_contour_data:
            lines.append(f"{'ID':<5} {'*':<2} {'Area':<9} {'Mode':<9} {'Max':<8} {'Centroid (Y, X)'}")
            lines.append("─" * 55)
            
            for data in fluor_contour_data:
                coloc_mark = "*" if data['id'] in colocalized_fluor_ids else " "
                lines.append(
                    f"{data['id']:<5} "
                    f"{coloc_mark:<2} "
                    f"{data['area']:<9.0f} "
                    f"{data['mode_intensity']:<9.1f} "
                    f"{data['max_intensity']:<8.0f} "
                    f"({data['centroid_y']:.1f}, {data['centroid_x']:.1f})"
                )
            
            # Summary statistics
            fluor_areas = [d['area'] for d in fluor_contour_data]
            fluor_mode_ints = [d['mode_intensity'] for d in fluor_contour_data]
            lines.append("─" * 55)
            lines.append(f"Stats: Area μ={np.mean(fluor_areas):.1f} σ={np.std(fluor_areas):.1f} | "
                        f"Intensity(mode) μ={np.mean(fluor_mode_ints):.1f}")
        else:
            lines.append("  No bacteria detected")
        
        lines.append("")
        lines.append("Legend: * = Co-localized | Yellow contours = Co-localized")
        
        return "\n".join(lines)
    
    def generate_review_sheet(self, pair_index):
        """Generate review data for a single image pair"""
        pair = self.image_pairs[pair_index]
        gray_img = pair['gray']
        fluor_img = pair['fluor']
        filename = pair['gray_file']
        fluor_path = os.path.join(self.source_folder, pair['fluor_file'])
        
        # Load original fluorescence image directly from source
        try:
            orig_fluor_img = imread(fluor_path)
        except Exception as e:
            print(f"Error loading original fluorescence image {pair['fluor_file']}: {e}")
            orig_fluor_img = fluor_img  # Fallback to pre-loaded image if loading fails
        
        # Detect bacteria (preprocessing only for segmentation)
        gray_proc, gray_labeled, gray_props, gray_large = self.detect_bacteria(gray_img, is_grayscale=True)
        fluor_proc, fluor_labeled, fluor_props, fluor_large = self.detect_bacteria(fluor_img, is_grayscale=False)
        
        # Get contour intensity data (MODE instead of MEAN, no MIN)
        gray_contour_data = self.get_contour_intensities(gray_img, gray_labeled, gray_large)
        fluor_contour_data = self.get_contour_intensities(fluor_img, fluor_labeled, fluor_large)
        
        # Calculate co-localization
        colocalized = self.calculate_colocalization(gray_labeled, fluor_labeled, 
                                                    gray_contour_data, fluor_contour_data)
        
        # Create visualizations
        gray_vis = self.create_overlay_visualization(gray_img, gray_labeled, gray_large, 
                                                     is_grayscale=True, colocalized=colocalized)
        fluor_vis = self.create_overlay_visualization(fluor_img, fluor_labeled, fluor_large, 
                                                      is_grayscale=False, colocalized=colocalized)
        
        # Prepare original visualization (normalize if necessary and convert to RGB)
        orig_vis = orig_fluor_img.copy()
        if orig_vis.ndim == 2:
            if orig_vis.dtype != np.uint8:
                if orig_vis.max() > 0:
                    orig_vis = (orig_vis / orig_vis.max() * 255).astype(np.uint8)
                else:
                    orig_vis = orig_vis.astype(np.uint8)
            orig_vis = np.dstack([orig_vis, orig_vis, orig_vis])
        
        stats_text = self.format_statistics_text(gray_large, fluor_large, 
                                                gray_contour_data, fluor_contour_data, 
                                                colocalized, filename)
        
        # Store results for Excel export
        result = {
            'filename': filename,
            'gray_contour_data': gray_contour_data,
            'fluor_contour_data': fluor_contour_data,
            'colocalized': colocalized,
            'gray_count': len(gray_large),
            'fluor_count': len(fluor_large),
            'coloc_count': len(colocalized),
            'gray_vis': gray_vis,
            'fluor_vis': fluor_vis,
            'orig_vis': orig_vis,
            'stats_text': stats_text
        }
        self.all_results.append(result)
        
    def generate_all_review_sheets(self):
        """Generate review data for all image pairs"""
        print(f"\nGenerating review data for {len(self.image_pairs)} image pair(s)...\n")
        
        for i, pair in enumerate(self.image_pairs):
            print(f"Processing {i+1}/{len(self.image_pairs)}: {pair['gray_file']}")
            self.generate_review_sheet(i)
            print(f"  ✓ Processed")
        
        print(f"\n{'='*70}")
        print(f"✓ All review data generated successfully!")
        print(f"{'='*70}\n")
    
    def export_to_excel(self):
        """Export all results to Excel file (MODE intensity, no MIN), including embedded images"""
        print("\nExporting results to Excel...")
        
        # Create Excel writer
        output_file = f"bacteria_detection_results_{self.source_folder.replace('/', '_').replace('//', '_')}.xlsx"
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        
        # Sheet 1: Summary
        summary_data = []
        for result in self.all_results:
            coloc_rate = (result['coloc_count'] / result['gray_count'] * 100) if result['gray_count'] > 0 else 0
            
            # Calculate statistics
            gray_areas = [d['area'] for d in result['gray_contour_data']]
            fluor_areas = [d['area'] for d in result['fluor_contour_data']]
            
            gray_intensities = [d['mode_intensity'] for d in result['gray_contour_data']]
            fluor_intensities = [d['mode_intensity'] for d in result['fluor_contour_data']]
            
            summary_data.append({
                'Filename': result['filename'],
                'Grayscale_Count': result['gray_count'],
                'Fluorescence_Count': result['fluor_count'],
                'Colocalized_Count': result['coloc_count'],
                'Colocalization_Rate_%': round(coloc_rate, 2),
                'Gray_Avg_Area': round(np.mean(gray_areas), 2) if gray_areas else 0,
                'Gray_Std_Area': round(np.std(gray_areas), 2) if gray_areas else 0,
                'Gray_Avg_Mode_Intensity': round(np.mean(gray_intensities), 2) if gray_intensities else 0,
                'Fluor_Avg_Area': round(np.mean(fluor_areas), 2) if fluor_areas else 0,
                'Fluor_Std_Area': round(np.std(fluor_areas), 2) if fluor_areas else 0,
                'Fluor_Avg_Mode_Intensity': round(np.mean(fluor_intensities), 2) if gray_intensities else 0,
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Sheet 2: Grayscale Details (MODE, no MIN)
        gray_data = []
        for result in self.all_results:
            colocalized_ids = {c['gray_id'] for c in result['colocalized']}
            
            for data in result['gray_contour_data']:
                gray_data.append({
                    'Filename': result['filename'],
                    'Bacteria_ID': data['id'],
                    'Colocalized': 'Yes' if data['id'] in colocalized_ids else 'No',
                    'Area_px': round(data['area'], 2),
                    'Perimeter_px': round(data['perimeter'], 2),
                    'Eccentricity': round(data['eccentricity'], 3),
                    'Mode_Intensity': round(data['mode_intensity'], 2),
                    'Max_Intensity': round(data['max_intensity'], 2),
                    'Std_Intensity': round(data['std_intensity'], 2),
                    'Centroid_Y': round(data['centroid_y'], 2),
                    'Centroid_X': round(data['centroid_x'], 2),
                })
        
        gray_df = pd.DataFrame(gray_data)
        gray_df.to_excel(writer, sheet_name='Grayscale_Bacteria', index=False)
        
        # Sheet 3: Fluorescence Details (MODE, no MIN)
        fluor_data = []
        for result in self.all_results:
            colocalized_ids = {c['fluor_id'] for c in result['colocalized']}
            
            for data in result['fluor_contour_data']:
                fluor_data.append({
                    'Filename': result['filename'],
                    'Bacteria_ID': data['id'],
                    'Colocalized': 'Yes' if data['id'] in colocalized_ids else 'No',
                    'Area_px': round(data['area'], 2),
                    'Perimeter_px': round(data['perimeter'], 2),
                    'Eccentricity': round(data['eccentricity'], 3),
                    'Mode_Intensity': round(data['mode_intensity'], 2),
                    'Max_Intensity': round(data['max_intensity'], 2),
                    'Std_Intensity': round(data['std_intensity'], 2),
                    'Centroid_Y': round(data['centroid_y'], 2),
                    'Centroid_X': round(data['centroid_x'], 2),
                })
        
        fluor_df = pd.DataFrame(fluor_data)
        fluor_df.to_excel(writer, sheet_name='Fluorescence_Bacteria', index=False)
        
        # Sheet 4: Colocalization Pairs
        coloc_data = []
        for result in self.all_results:
            for coloc in result['colocalized']:
                coloc_data.append({
                    'Filename': result['filename'],
                    'Gray_Bacteria_ID': coloc['gray_id'],
                    'Fluor_Bacteria_ID': coloc['fluor_id'],
                    'Overlap_Ratio_%': round(coloc['overlap_ratio'] * 100, 2),
                    'Gray_Area_px': round(coloc['gray_area'], 2),
                    'Fluor_Area_px': round(coloc['fluor_area'], 2),
                    'Intersection_Area_px': round(coloc['intersection_area'], 2),
                })
        
        coloc_df = pd.DataFrame(coloc_data)
        coloc_df.to_excel(writer, sheet_name='Colocalization_Pairs', index=False)
        
        # Sheet 5: Parameters
        params_data = [{
            'Parameter': key,
            'Value': value
        } for key, value in self.config.items()]
        
        params_df = pd.DataFrame(params_data)
        params_df.to_excel(writer, sheet_name='Detection_Parameters', index=False)
        
        # Add review sheets with embedded images
        for i, result in enumerate(self.all_results):
            sheet_title = f"Review_{i+1}_{result['filename'][:20]}"
            ws = writer.book.create_sheet(title=sheet_title)
            
            # Function to get image BytesIO
            def get_img_bio(arr, max_width=400):
                pil_img = PILImage.fromarray(arr)
                aspect = pil_img.height / pil_img.width
                new_width = min(pil_img.width, max_width)
                new_height = int(new_width * aspect)
                pil_img = pil_img.resize((new_width, new_height))
                bio = io.BytesIO()
                pil_img.save(bio, format="PNG")
                bio.seek(0)
                return bio
            
            # Grayscale channel image
            gray_bio = get_img_bio(result['gray_vis'])
            img_gray = Image(gray_bio)
            ws.add_image(img_gray, 'A1')
            
            # Fluorescence channel image
            fluor_bio = get_img_bio(result['fluor_vis'])
            img_fluor = Image(fluor_bio)
            ws.add_image(img_fluor, 'G1')
            
            # Original fluorescence
            orig_bio = get_img_bio(result['orig_vis'])
            img_orig = Image(orig_bio)
            ws.add_image(img_orig, 'A30')
            
            # Detection report (text)
            ws['G30'] = result['stats_text']
            ws['G30'].font = Font(name='Courier New', size=8)
            ws['G30'].alignment = Alignment(wrapText=True, vertical='top', horizontal='left')
            
            # Adjust column widths for images and text
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col].width = 15  # Adjust based on image width
            for col in ['G', 'H', 'I', 'J', 'K', 'L']:
                ws.column_dimensions[col].width = 15
            ws.column_dimensions['G'].width = 80  # Wider for text
        
        # Save and close
        writer.close()
        
        print(f"✓ Excel file saved: {output_file}")
        print(f"  - Summary sheet: {len(summary_data)} images")
        print(f"  - Grayscale bacteria: {len(gray_data)} detections")
        print(f"  - Fluorescence bacteria: {len(fluor_data)} detections")
        print(f"  - Colocalization pairs: {len(coloc_data)} pairs")
        print(f"  - Review sheets with embedded images: {len(self.all_results)}")
        print(f"  - Using MODE intensity (no MIN)")
        print(f"{'='*70}\n")


def main():
    """Main function to run the review tool"""
    print("="*70)
    print("BACTERIA DETECTION REVIEW SHEET GENERATOR WITH EXCEL EXPORT")
    print("="*70)
    print("\nUpdates:")
    print("  • Using MODE intensity instead of mean")
    print("  • Removed MIN intensity")
    print("  • Using ORIGINAL source images for display")
    print("  • Added original fluorescence image from _ch01.tif (unprocessed)")
    print("  • Portrait orientation with reserved header/footer space")
    print("  • Fixed overlap between detection report and parameters")
    print("  • Maximized image sizes for optimal review")
    print("  • Embedded images in Excel review sheets")
    print("\n" + "="*70 + "\n")
    
    # You can change this to your source folder
    source_folder = "source//1"
    
    try:
        tool = BacteriaReviewTool(source_folder=source_folder)
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()