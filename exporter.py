"""
Report Generation Module for Particle Scout
Handles Excel report creation, chart generation, and data visualization
"""

import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path
from datetime import datetime
import cv2


# Configuration
PLOT_DPI = 150


def numpy_to_excel_image(img_array, format='PNG'):
    """Convert numpy array to Excel-compatible image"""
    if img_array is None:
        return None
    
    # Convert to PIL Image
    if len(img_array.shape) == 2:  # Grayscale
        pil_img = PILImage.fromarray(img_array)
    else:  # Color (BGR to RGB)
        pil_img = PILImage.fromarray(cv2.cvtColor(img_array, cv2.COLOR_BGR2RGB))
    
    # Save to bytes buffer
    img_buffer = BytesIO()
    pil_img.save(img_buffer, format=format)
    img_buffer.seek(0)
    
    # Create Excel image
    excel_img = XLImage(img_buffer)
    # Scale down for Excel (max width ~600 pixels)
    scale_factor = min(1.0, 600 / pil_img.width)
    excel_img.width = int(pil_img.width * scale_factor)
    excel_img.height = int(pil_img.height * scale_factor)
    
    return excel_img


def create_comparison_chart(control_data, sample_data, sample_name, output_path, unit_str='um'):
    """
    Create a bar chart comparing control vs sample with error bars
    Similar to Figure B in the reference image
    
    Args:
        control_data: List of control particle dictionaries
        sample_data: List of sample particle dictionaries
        sample_name: Name of the sample group
        output_path: Path to save the chart
        unit_str: Unit string for axis label
    
    Returns:
        numpy.ndarray: Image array for Excel embedding, or None if no data
    """
    # Extract intensity values (updated key name)
    control_intensities = [p['intensity_per_area'] for p in control_data] if control_data else []
    sample_intensities = [p['intensity_per_area'] for p in sample_data] if sample_data else []
    
    if not control_intensities and not sample_intensities:
        print("⚠ No data to plot comparison chart")
        return None
    
    # Calculate statistics - convert to float explicitly
    control_mean = float(np.mean(control_intensities)) if control_intensities else 0.0
    control_sem = float(np.std(control_intensities, ddof=1) / np.sqrt(len(control_intensities))) if len(control_intensities) > 1 else 0.0
    
    sample_mean = float(np.mean(sample_intensities)) if sample_intensities else 0.0
    sample_sem = float(np.std(sample_intensities, ddof=1) / np.sqrt(len(sample_intensities))) if len(sample_intensities) > 1 else 0.0
    
    # Create bar chart
    fig, ax = plt.subplots(figsize=(6, 5), dpi=PLOT_DPI)
    
    groups = ['Control', 'Sample']
    means = np.array([control_mean, sample_mean], dtype=np.float64)
    errors = np.array([control_sem, sample_sem], dtype=np.float64)
    
    # Define colors matching Figure B style
    colors = ['#87CEEB', '#DDA0DD']  # Light blue and light purple
    
    # Create bars
    bars = ax.bar(groups, means, yerr=errors, capsize=10, 
                   color=colors, edgecolor='black', linewidth=1.5,
                   error_kw={'linewidth': 2, 'ecolor': 'black'})
    
    # Styling
    ax.set_ylabel(f'Fluorescence Intensity (Fluor/{unit_str}²)', fontsize=12, fontweight='bold')
    max_value = float(np.max(means)) if len(means) > 0 else 1.0
    ax.set_ylim(0, max_value * 1.3)  # Set y-axis to accommodate error bars
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.tick_params(axis='both', labelsize=11)
    
    # Add grid for better readability
    ax.yaxis.grid(True, linestyle='--', alpha=0.3)
    ax.set_axisbelow(True)
    
    # Add title
    ax.set_title('Average Fluorescence Intensity Comparison', fontsize=13, fontweight='bold', pad=15)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=PLOT_DPI, bbox_inches='tight')
    plt.close()
    
    print(f"✓ Comparison chart saved: {output_path}")
    
    # Return the figure as numpy array for Excel embedding
    fig, ax = plt.subplots(figsize=(6, 5), dpi=PLOT_DPI)
    bars = ax.bar(groups, means, yerr=errors, capsize=10, 
                   color=colors, edgecolor='black', linewidth=1.5,
                   error_kw={'linewidth': 2, 'ecolor': 'black'})
    ax.set_ylabel(f'Fluorescence Intensity (Fluor/{unit_str}²)', fontsize=12, fontweight='bold')
    ax.set_ylim(0, max_value * 1.3)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.tick_params(axis='both', labelsize=11)
    ax.yaxis.grid(True, linestyle='--', alpha=0.3)
    ax.set_axisbelow(True)
    ax.set_title('Average Fluorescence Intensity Comparison', fontsize=13, fontweight='bold', pad=15)
    
    # Convert to numpy array - compatible method
    plt.tight_layout()
    
    # Save to BytesIO buffer
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=PLOT_DPI, bbox_inches='tight')
    buf.seek(0)
    
    # Read as PIL Image then convert to numpy
    pil_img = PILImage.open(buf)
    img_array = np.array(pil_img)
    
    plt.close()
    buf.close()
    
    return img_array


def create_intensity_histogram(data, group_name, output_path, unit_str='um'):
    """
    Create histogram of particle intensities
    
    Args:
        data: List of particle dictionaries
        group_name: Name of the group (e.g., 'Control', 'Sample')
        output_path: Path to save the histogram
        unit_str: Unit string for axis label
    
    Returns:
        numpy.ndarray: Image array for Excel embedding, or None if no data
    """
    if not data:
        print(f"⚠ No data to plot histogram for {group_name}")
        return None
    
    intensities = [p['intensity_per_area'] for p in data]
    
    fig, ax = plt.subplots(figsize=(6, 4), dpi=PLOT_DPI)
    
    ax.hist(intensities, bins=20, color='#87CEEB', edgecolor='black', alpha=0.7)
    ax.set_xlabel(f'Fluorescence Intensity (Fluor/{unit_str}²)', fontsize=11, fontweight='bold')
    ax.set_ylabel('Frequency', fontsize=11, fontweight='bold')
    ax.set_title(f'{group_name} - Intensity Distribution', fontsize=12, fontweight='bold')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(True, linestyle='--', alpha=0.3, axis='y')
    ax.set_axisbelow(True)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=PLOT_DPI, bbox_inches='tight')
    plt.close()
    
    print(f"✓ Histogram saved: {output_path}")
    
    # Convert to numpy array for Excel
    fig, ax = plt.subplots(figsize=(6, 4), dpi=PLOT_DPI)
    ax.hist(intensities, bins=20, color='#87CEEB', edgecolor='black', alpha=0.7)
    ax.set_xlabel(f'Fluorescence Intensity (Fluor/{unit_str}²)', fontsize=11, fontweight='bold')
    ax.set_ylabel('Frequency', fontsize=11, fontweight='bold')
    ax.set_title(f'{group_name} - Intensity Distribution', fontsize=12, fontweight='bold')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(True, linestyle='--', alpha=0.3, axis='y')
    ax.set_axisbelow(True)
    plt.tight_layout()
    
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=PLOT_DPI, bbox_inches='tight')
    buf.seek(0)
    pil_img = PILImage.open(buf)
    img_array = np.array(pil_img)
    plt.close()
    buf.close()
    
    return img_array


def style_header(ws, row_num, fill_color='4472C4'):
    """Apply header styling to a row"""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(bold=True, color='FFFFFF', size=11)
    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row_num]:
        if cell.value:  # Only style non-empty cells
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            cell.border = border


def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_excel_report(control_data, sample_data, sample_name, output_excel_path, 
                       comparison_chart_img=None, control_hist_img=None, sample_hist_img=None,
                       unit_str='um', pixels_per_um=1.0):
    """
    Create comprehensive Excel report with data and visualizations
    
    Args:
        control_data: List of control particle dictionaries
        sample_data: List of sample particle dictionaries
        sample_name: Name of the sample group
        output_excel_path: Path to save Excel file
        comparison_chart_img: numpy array of comparison chart
        control_hist_img: numpy array of control histogram
        sample_hist_img: numpy array of sample histogram
        unit_str: Unit string for measurements
        pixels_per_um: Conversion factor for pixel measurements
    """
    wb = Workbook()
    
    # Style definitions
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== Summary Sheet ==========
    ws_summary = wb.active
    if ws_summary is None:
        ws_summary = wb.create_sheet("Summary")
    else:
        ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = 'Particle Analysis Report'
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws_summary['A2'].font = Font(italic=True, size=10)
    
    # Statistics
    row = 5
    ws_summary[f'A{row}'] = 'Group'
    ws_summary[f'B{row}'] = 'N (particles)'
    ws_summary[f'C{row}'] = f'Mean Intensity (Fluor/{unit_str}²)'
    ws_summary[f'D{row}'] = 'SEM'
    ws_summary[f'E{row}'] = 'SD'
    ws_summary[f'F{row}'] = f'Mean Area ({unit_str}²)'
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_summary[f'{col}{row}'].fill = header_fill
        ws_summary[f'{col}{row}'].font = header_font
        ws_summary[f'{col}{row}'].border = border
        ws_summary[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    # Control statistics
    row += 1
    control_intensities = [p['intensity_per_area'] for p in control_data] if control_data else []
    control_areas = [p['area_um2'] for p in control_data] if control_data else []
    
    ws_summary[f'A{row}'] = 'Control'
    ws_summary[f'B{row}'] = len(control_data)
    ws_summary[f'C{row}'] = float(np.mean(control_intensities)) if control_intensities else 0
    ws_summary[f'D{row}'] = float(np.std(control_intensities, ddof=1) / np.sqrt(len(control_intensities))) if len(control_intensities) > 1 else 0
    ws_summary[f'E{row}'] = float(np.std(control_intensities, ddof=1)) if len(control_intensities) > 1 else 0
    ws_summary[f'F{row}'] = float(np.mean(control_areas)) if control_areas else 0
    
    # Sample statistics
    row += 1
    sample_intensities = [p['intensity_per_area'] for p in sample_data] if sample_data else []
    sample_areas = [p['area_um2'] for p in sample_data] if sample_data else []
    
    ws_summary[f'A{row}'] = sample_name
    ws_summary[f'B{row}'] = len(sample_data)
    ws_summary[f'C{row}'] = float(np.mean(sample_intensities)) if sample_intensities else 0
    ws_summary[f'D{row}'] = float(np.std(sample_intensities, ddof=1) / np.sqrt(len(sample_intensities))) if len(sample_intensities) > 1 else 0
    ws_summary[f'E{row}'] = float(np.std(sample_intensities, ddof=1)) if len(sample_intensities) > 1 else 0
    ws_summary[f'F{row}'] = float(np.mean(sample_areas)) if sample_areas else 0
    
    # Format numbers
    for r in range(row-1, row+1):
        for col in ['C', 'D', 'E', 'F']:
            ws_summary[f'{col}{r}'].number_format = '0.00'
            ws_summary[f'{col}{r}'].border = border
            ws_summary[f'{col}{r}'].alignment = Alignment(horizontal='center')
        ws_summary[f'A{r}'].border = border
        ws_summary[f'B{r}'].border = border
        ws_summary[f'B{r}'].alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_summary.column_dimensions[col].width = 20
    
    # Insert comparison chart
    if comparison_chart_img is not None:
        xl_img = numpy_to_excel_image(comparison_chart_img)
        if xl_img:
            ws_summary.add_image(xl_img, f'A{row+3}')
    
    # ========== Control Data Sheet ==========
    ws_control = wb.create_sheet("Control Data")
    if ws_control is not None:
        _write_particle_data_sheet(ws_control, control_data, "Control", unit_str, 
                                    header_fill, header_font, border)
        
        # Insert histogram
        if control_hist_img is not None:
            xl_img = numpy_to_excel_image(control_hist_img)
            if xl_img:
                last_row = len(control_data) + 5
                ws_control.add_image(xl_img, f'H{last_row}')
        
        # Insert overlay images (if available)
        _add_overlay_images(ws_control, control_data, 'H', len(control_data) + 20)
    
    # ========== Sample Data Sheet ==========
    ws_sample = wb.create_sheet(f"{sample_name} Data")
    if ws_sample is not None:
        _write_particle_data_sheet(ws_sample, sample_data, sample_name, unit_str,
                                   header_fill, header_font, border)
        
        # Insert histogram
        if sample_hist_img is not None:
            xl_img = numpy_to_excel_image(sample_hist_img)
            if xl_img:
                last_row = len(sample_data) + 5
                ws_sample.add_image(xl_img, f'H{last_row}')
        
        # Insert overlay images (if available)
        _add_overlay_images(ws_sample, sample_data, 'H', len(sample_data) + 20)
    
    # Save workbook
    wb.save(output_excel_path)
    print(f"✓ Excel report saved: {output_excel_path}")


def _write_particle_data_sheet(ws, data, group_name, unit_str, header_fill, header_font, border):
    """Helper function to write particle data to a worksheet"""
    # Title
    ws['A1'] = f'{group_name} - Particle Data'
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    headers = ['Particle ID', f'Area ({unit_str}²)', f'Perimeter ({unit_str})', 
               'Circularity', f'Intensity/Area (Fluor/{unit_str}²)', 
               'Mean Intensity', 'Total Intensity', 'Std Dev']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Data rows
    for row_idx, particle in enumerate(data, start=4):
        ws.cell(row=row_idx, column=1, value=particle['id'])
        ws.cell(row=row_idx, column=2, value=particle['area_um2'])
        ws.cell(row=row_idx, column=3, value=particle['perimeter_um'])
        ws.cell(row=row_idx, column=4, value=particle['circularity'])
        ws.cell(row=row_idx, column=5, value=particle['intensity_per_area'])
        ws.cell(row=row_idx, column=6, value=particle['mean_intensity'])
        ws.cell(row=row_idx, column=7, value=particle['total_intensity'])
        ws.cell(row=row_idx, column=8, value=particle.get('std_intensity', 0))
        
        # Apply borders and number formatting
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            if col > 1:  # Format numbers
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _add_overlay_images(ws, data, start_column, start_row):
    """Helper function to add overlay images to worksheet"""
    if not data:
        return
    
    # Group particles by image name
    images_dict = {}
    for particle in data:
        img_name = particle.get('image_name', 'unknown')
        if img_name not in images_dict and 'overlay_image' in particle:
            images_dict[img_name] = particle['overlay_image']
    
    # Add title
    ws[f'{start_column}{start_row}'] = "Detected Particles - Overlay Images"
    ws[f'{start_column}{start_row}'].font = Font(bold=True, size=11)
    
    current_row = start_row + 2
    
    # Add each unique overlay image
    for img_name, overlay_img in images_dict.items():
        if overlay_img is not None:
            # Add image label
            ws[f'{start_column}{current_row}'] = img_name
            ws[f'{start_column}{current_row}'].font = Font(bold=True, size=10)
            
            # Add image
            xl_img = numpy_to_excel_image(overlay_img)
            if xl_img:
                ws.add_image(xl_img, f'{start_column}{current_row + 1}')
                current_row += 25  # Space for image (approximate)