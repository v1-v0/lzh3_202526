from __future__ import annotations

from pathlib import Path
import io
import datetime
from typing import Any, cast
from tkinter import messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image

from register_dataset import ImagePairRecord


# ---------------------------------------------------------------------------
# Internal helpers to read data from the viewer
# ---------------------------------------------------------------------------

def _get_stats_as_dataframe(viewer: Any) -> pd.DataFrame:
    """
    Convert viewer.stats_tree content into a pandas DataFrame.

    Columns:
      idx, area_px, eq_diam_px, total_intensity, intensity_per_area
    """
    cols = ["idx", "area_px", "eq_diam_px", "total_intensity", "intensity_per_area"]
    rows: list[dict] = []

    tree = viewer.stats_tree
    for iid in tree.get_children():
        vals = tree.item(iid, "values")
        if not vals:
            continue
        row = dict(zip(cols, vals))
        for k in cols:
            try:
                row[k] = float(row[k])
            except Exception:
                pass
        rows.append(row)

    if not rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows)


def _get_histogram_png_bytes(viewer: Any) -> bytes | None:
    """
    Render viewer.histo_fig into PNG bytes, or None if figure is missing.

    This captures the 3-panel histogram (Intensity/Area, Total Fluo,
    Eq. diameter) exactly as shown in the GUI.
    """
    fig = getattr(viewer, "histo_fig", None)
    if fig is None:
        return None
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    return buf.getvalue()


def _get_display_images(viewer: Any) -> dict[str, Image.Image]:
    """
    Get the four main display PIL images from viewer for export.

    These are exactly the images currently displayed in the GUI panes,
    including contours, rank labels, arrows, and scale bar
    (if enabled when update_images was last called).

    Keys:
      - 'BF_raw_top'
      - 'FLUO_raw_top'
      - 'BF_enh_bottom'
      - 'FLUO_enh_bottom'
    """
    imgs: dict[str, Image.Image] = {}
    if getattr(viewer, "_bf_top_pil", None) is not None:
        imgs["BF_raw_top"] = viewer._bf_top_pil
    if getattr(viewer, "_fluo_top_pil", None) is not None:
        imgs["FLUO_raw_top"] = viewer._fluo_top_pil
    if getattr(viewer, "_bf_bottom_pil", None) is not None:
        imgs["BF_enh_bottom"] = viewer._bf_bottom_pil
    if getattr(viewer, "_fluo_bottom_pil", None) is not None:
        imgs["FLUO_enh_bottom"] = viewer._fluo_bottom_pil
    return imgs


def _write_spec_sheet(wb: Workbook, rec: ImagePairRecord) -> None:
    """
    Create 'Spec' sheet in Excel workbook with microscope/acquisition spec,
    modeled on the Leica LAS X metadata and your Word summary.
    """
    ws: Worksheet = wb.create_sheet(title="Spec")

    row = 1
    ws.cell(row=row, column=1, value="Leica LAS X metadata summary")
    row += 2

    ws.cell(row=row, column=1, value="Image Dimensions & Resolution")
    row += 1

    if rec.field_length_um_x is not None and rec.field_length_um_y is not None:
        ws.cell(row=row, column=1, value="Field size X (µm)")
        ws.cell(row=row, column=2, value=float(rec.field_length_um_x))
        row += 1
        ws.cell(row=row, column=1, value="Field size Y (µm)")
        ws.cell(row=row, column=2, value=float(rec.field_length_um_y))
        row += 1

    ws.cell(row=row, column=1, value="Pixel size (µm/pixel)")
    ws.cell(
        row=row,
        column=2,
        value=float(rec.pixel_size_um) if rec.pixel_size_um is not None else "n/a",
    )
    row += 1

    ws.cell(row=row, column=1, value="Bit depth")
    ws.cell(row=row, column=2, value=rec.bit_depth if rec.bit_depth is not None else "n/a")
    row += 2

    ws.cell(row=row, column=1, value="Optical Configuration")
    row += 1
    ws.cell(row=row, column=1, value="Objective")
    ws.cell(row=row, column=2, value=rec.objective_name or "N PLAN 100x/1.25 Oil")
    row += 1
    ws.cell(row=row, column=1, value="Numerical Aperture (NA)")
    ws.cell(
        row=row,
        column=2,
        value=float(rec.numerical_aperture) if rec.numerical_aperture is not None else 1.25,
    )
    row += 1
    ws.cell(row=row, column=1, value="Microscope")
    ws.cell(row=row, column=2, value="Leica DMI8 (inverted)")
    row += 1
    ws.cell(row=row, column=1, value="Total magnification")
    ws.cell(row=row, column=2, value="100x")
    row += 2

    ws.cell(row=row, column=1, value="Channel Information")
    row += 1
    ws.cell(row=row, column=1, value="Channel 1 (Brightfield - TL-BF)")
    row += 1
    ws.cell(row=row, column=1, value="Exposure (s)")
    ws.cell(
        row=row,
        column=2,
        value=float(rec.exposure_bf_s) if rec.exposure_bf_s is not None else "0.138",
    )
    row += 1
    ws.cell(row=row, column=1, value="Channel 2 (Fluorescence - FLUO)")
    row += 1
    ws.cell(row=row, column=1, value="Exposure (s)")
    ws.cell(
        row=row,
        column=2,
        value=float(rec.exposure_fluo_s) if rec.exposure_fluo_s is not None else "0.138",
    )
    row += 1
    ws.cell(row=row, column=1, value="Emission wavelength (nm)")
    ws.cell(row=row, column=2, value="594")
    row += 1
    ws.cell(row=row, column=1, value="Excitation LED (nm)")
    ws.cell(row=row, column=2, value="555")
    row += 2

    ws.cell(row=row, column=1, value="Display/Viewer Channel Scaling")
    row += 1
    if rec.channel_scaling:
        for ch_idx, ch in enumerate(rec.channel_scaling):
            ws.cell(row=row, column=1, value=f"Channel {ch_idx}")
            ws.cell(
                row=row,
                column=2,
                value=f"Black={ch.get('black_norm')}, "
                      f"White={ch.get('white_norm')}, "
                      f"Gamma={ch.get('gamma')}",
            )
            row += 1
    else:
        ws.cell(row=row, column=1, value="No viewer scaling metadata available.")
        row += 1

    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 40


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def _format_stats_sheet(ws: Worksheet, df_stats: pd.DataFrame) -> None:
    """
    Apply simple formatting to 'Statistics' sheet:
    - header style, alignment, borders, column widths
    - highlight 'middle range' rows based on intensity_per_area (33–66 percentile)
    """
    if df_stats.empty:
        ws["A1"].font = Font(bold=True)
        return

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    max_col = len(df_stats.columns)
    max_row = len(df_stats) + 1  # header + data

    # Header row
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Body cells: basic formatting
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border

    # Highlight middle range based on intensity_per_area if present
    if "intensity_per_area" in df_stats.columns and not df_stats["intensity_per_area"].empty:
        try:
            intensity_series = pd.to_numeric(df_stats["intensity_per_area"], errors="coerce")
            intensity_non_null = intensity_series.dropna()
            if not intensity_non_null.empty:
                low = intensity_non_null.quantile(1 / 3)
                high = intensity_non_null.quantile(2 / 3)
                middle_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow

                for r_idx in range(len(df_stats)):
                    val = intensity_series.iloc[r_idx]
                    excel_row = r_idx + 2  # +1 for header, +1 for 1-based indexing
                    if pd.notna(val) and low <= val <= high:
                        for c_idx in range(1, max_col + 1):
                            ws.cell(row=excel_row, column=c_idx).fill = middle_fill
        except Exception:
            # If anything goes wrong, skip highlighting
            pass

    # Column widths
    for col_idx, col_name in enumerate(df_stats.columns, start=1):
        max_len = len(str(col_name))
        for val in df_stats[col_name]:
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, max_len + 2)


def _format_spec_sheet(ws: Worksheet) -> None:
    """
    Apply simple formatting to 'Spec' sheet:
    - Bold major section headers
    """
    bold_font = Font(bold=True)

    section_keywords = [
        "Leica LAS X metadata summary",
        "Image Dimensions & Resolution",
        "Optical Configuration",
        "Channel Information",
        "Display/Viewer Channel Scaling",
    ]

    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and any(key in val for key in section_keywords):
            ws.cell(row=row, column=1).font = bold_font

    # Left-align first column
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")


# ---------------------------------------------------------------------------
# Main entry point called from BFFluoViewer
# ---------------------------------------------------------------------------

def export_current_view(viewer: Any) -> None:
    """
    Export the current viewer state to an Excel workbook.

    Sheets:
      - 'Statistics': stats table, histogram to the right
      - 'BF Images' : brightfield images (A1, Q1; at most two images)
      - 'FLUO Images': fluorescence images (A1, Q1; at most two images)
      - 'Spec'      : metadata/spec summary

    The images exported are exactly those shown in the GUI (including
    scale bar and contour labels). The histogram is the same 3-panel
    figure (Fluo/Area, Total Fluo, Eq. diameter) displayed in the GUI.
    """
    if viewer.current_record is None:
        messagebox.showwarning("Export", "No image pair selected.")
        return

    rec: ImagePairRecord = viewer.current_record

    df_stats = _get_stats_as_dataframe(viewer)
    display_imgs = _get_display_images(viewer)
    hist_png_bytes = _get_histogram_png_bytes(viewer)

    wb = Workbook()
    _active = wb.active
    ws_stats = cast(Worksheet, _active)
    ws_stats.title = "Statistics"

    from openpyxl.drawing.image import Image as XLImage  # import here for clarity

    # ----------------- Statistics sheet -----------------
    if not df_stats.empty:
        # Write headers
        for col_idx, col_name in enumerate(df_stats.columns, start=1):
            ws_stats.cell(row=1, column=col_idx, value=col_name)
        # Write data
        for r_idx, (_, row) in enumerate(df_stats.iterrows(), start=2):
            for c_idx, val in enumerate(row, start=1):
                try:
                    ws_stats.cell(row=r_idx, column=c_idx, value=float(val))
                except Exception:
                    ws_stats.cell(row=r_idx, column=c_idx, value=str(val))

        # Histogram to the right (3-panel, including eq_diam_px distribution)
        if hist_png_bytes is not None:
            last_col = len(df_stats.columns)
            hist_col = last_col + 2  # one empty col, then histogram
            anchor_cell = f"{get_column_letter(hist_col)}1"
            buf_hist = io.BytesIO(hist_png_bytes)
            xl_hist = XLImage(buf_hist)
            xl_hist.anchor = anchor_cell
            ws_stats.add_image(xl_hist)
    else:
        ws_stats.cell(row=1, column=1, value="No particle statistics available.")
        if hist_png_bytes is not None:
            anchor_cell = "B1"
            buf_hist = io.BytesIO(hist_png_bytes)
            xl_hist = XLImage(buf_hist)
            xl_hist.anchor = anchor_cell
            ws_stats.add_image(xl_hist)

    _format_stats_sheet(ws_stats, df_stats)

    # ----------------- BF / FLUO images in separate sheets -----------------
    bf_keys = [k for k in display_imgs if k.startswith("BF")]
    fluo_keys = [k for k in display_imgs if k.startswith("FLUO")]

    # Warn if too many images
    if len(bf_keys) > 2:
        messagebox.showwarning(
            "Export - BF Images",
            f"More than 2 BF images detected ({len(bf_keys)}). "
            "Only the first 2 will be exported (A1, Q1).",
        )
        bf_keys = bf_keys[:2]

    if len(fluo_keys) > 2:
        messagebox.showwarning(
            "Export - FLUO Images",
            f"More than 2 FLUO images detected ({len(fluo_keys)}). "
            "Only the first 2 will be exported (A1, Q1).",
        )
        fluo_keys = fluo_keys[:2]

    def _add_images_to_sheet(sheet: Worksheet, keys: list[str]) -> None:
        """
        Place up to two images on the sheet:
          - first at A1 (col 1)
          - second at Q1 (col 17)

        The images are taken directly from the viewer's cached PIL images,
        which already include contours, labels and scale bar.
        """
        for i in range(1, 50):
            sheet.row_dimensions[i].height = 20.0

        scale_factor = 0.80  # change to 0.70 if you prefer 70%

        positions = [
            ("A1", 1),   # first image
            ("Q1", 17),  # second image
        ]

        from openpyxl.drawing.image import Image as XLImage  # local import

        for idx, name in enumerate(keys):
            if idx >= len(positions):
                break

            img = display_imgs[name]

            new_w = int(img.width * scale_factor)
            new_h = int(img.height * scale_factor)
            img_resized = img.resize((new_w, new_h), Image.Resampling.LANCZOS)

            anchor_cell, col_index = positions[idx]

            buf = io.BytesIO()
            if img_resized.mode not in ("RGB", "RGBA"):
                img_resized = img_resized.convert("RGB")
            img_resized.save(buf, format="PNG")
            buf.seek(0)

            xl_img = XLImage(buf)
            xl_img.anchor = anchor_cell
            sheet.add_image(xl_img)

            label_col = col_index + 4
            sheet.cell(
                row=1,
                column=label_col,
                value=name,
            )

    ws_bf = wb.create_sheet(title="BF Images")
    if bf_keys:
        _add_images_to_sheet(ws_bf, bf_keys)
    else:
        ws_bf["A1"] = "No BF images available."

    ws_fluo = wb.create_sheet(title="FLUO Images")
    if fluo_keys:
        _add_images_to_sheet(ws_fluo, fluo_keys)
    else:
        ws_fluo["A1"] = "No FLUO images available."

    # ----------------- Spec sheet -----------------
    _write_spec_sheet(wb, rec)
    ws_spec = wb["Spec"]
    _format_spec_sheet(ws_spec)

    # ----------------- Save workbook -----------------
    try:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_pair = rec.pair_name.replace(" ", "_").replace("/", "_")
        filename = f"Export_{safe_pair}_{ts}.xlsx"

        # Save to current user's Downloads folder
        downloads_dir = Path.home() / "Downloads"
        downloads_dir.mkdir(parents=True, exist_ok=True)
        out_path = downloads_dir / filename

        wb.save(out_path)
        messagebox.showinfo("Export", f"Exported to:\n{out_path}")
    except Exception as e:
        messagebox.showerror("Export error", f"Could not save Excel file:\n{e}")