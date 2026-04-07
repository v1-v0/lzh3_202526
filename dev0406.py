# Standard library imports
import atexit
import csv
import json
import logging

import os

import re
import shutil
import stat
import subprocess
import sys
import textwrap
import time as pytime
# from tracemalloc import start
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional, Tuple, cast

# Third-party data science imports


import numpy as np
import pandas as pd

from scipy import stats as scipy_stats

from tqdm import tqdm

# Computer vision imports
import cv2
from skimage.registration import phase_cross_correlation

# Plotting imports
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.axes import Axes as MplAxes
import seaborn as sns

# Excel/Office imports
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.drawing.image import Image as XLImage

from bacteria_configs import (
    SegmentationConfig, 
)

# ==================================================
# Helper Functions
# ==================================================

def _compute_group_vs_control_stats(
    group_values: pd.Series,
    control_values: pd.Series,
    threshold: float
) -> dict:
    """Compute CI, effect size, p-value, and confidence level for one group vs control.
    
    Args:
        group_values: Fluor_Density values for the test group
        control_values: Fluor_Density values for the control group
        threshold: Clinical decision threshold
        
    Returns:
        dict with CI_Lower, CI_Upper, Cohens_d, P_Value, Classification_Confidence
    """
    n = len(group_values)
    mean_val = float(group_values.mean())
    std_val = float(group_values.std(ddof=1)) if n > 1 else 0.0

    # --- 95% Confidence Interval ---
    if n >= 2 and std_val > 0:
        t_crit = float(scipy_stats.t.ppf(0.975, df=n - 1))
        sem = std_val / np.sqrt(n)
        ci_lower = mean_val - t_crit * sem
        ci_upper = mean_val + t_crit * sem
    else:
        ci_lower = mean_val
        ci_upper = mean_val

    # --- Cohen's d (vs control) ---
    ctrl_mean = float(control_values.mean())
    ctrl_std = float(control_values.std(ddof=1)) if len(control_values) > 1 else 0.0
    pooled_std = np.sqrt(
        ((n - 1) * std_val ** 2 + (len(control_values) - 1) * ctrl_std ** 2)
        / (n + len(control_values) - 2)
    ) if (n + len(control_values) - 2) > 0 else 1.0

    cohens_d = (mean_val - ctrl_mean) / pooled_std if pooled_std > 0 else 0.0

    # --- Welch's t-test ---
    if n >= 2 and len(control_values) >= 2:
        try:
            t_raw, p_raw = scipy_stats.ttest_ind(
                group_values.values,
                control_values.values,
                equal_var=False,
            )
            p_value = float(np.asarray(p_raw).item())
        except Exception:
            p_value = np.nan
    else:
        p_value = np.nan

    # --- Significance label ---
    if pd.isna(p_value):
        sig_label = "N/A (n<2)"
    elif p_value < 0.001:
        sig_label = "***"
    elif p_value < 0.01:
        sig_label = "**"
    elif p_value < 0.05:
        sig_label = "*"
    else:
        sig_label = "ns"

    # --- Classification confidence ---
    # Based on whether the 95% CI overlaps the threshold
    if n < 3:
        confidence = "Low"
    elif (ci_upper < threshold) or (ci_lower > threshold):
        # CI entirely on one side of threshold
        if abs(cohens_d) >= 0.8:
            confidence = "High"
        else:
            confidence = "Moderate"
    else:
        # CI straddles the threshold — ambiguous
        confidence = "Low"

    return {
        'CI_Lower': round(ci_lower, 2),
        'CI_Upper': round(ci_upper, 2),
        'Cohens_d': round(cohens_d, 3),
        'P_Value': round(p_value, 6) if not pd.isna(p_value) else np.nan,
        'Significance': sig_label,
        'Classification_Confidence': confidence,
    }


def select_bacteria_config() -> dict:
    """Wrapper for select_bacteria_configuration with consistent return format"""
    return select_bacteria_configuration()


def configure_dataset() -> dict:
    """Collect complete dataset configuration from the user.

    Consolidates the old configure_dataset() (which auto-used the folder name
    as dataset ID with no prompt) and the dead collect_configuration() (which
    was never called from main() but contained the interactive ID prompt).

    Returns
    -------
    dict with keys:
        source_dir, batch_mode,
        dataset_id, dataset_id_base,
        source_dir_positive, source_dir_negative  (batch mode only),
        percentile, threshold_pct
    """
    config: dict = {}

    # ── Source directory ───────────────────────────────────────────────────
    source_dir = select_source_directory()
    if source_dir is None:
        raise SystemExit("No source directory selected")

    config['source_dir'] = source_dir

    # ── Batch mode detection ───────────────────────────────────────────────
    gplus_path  = source_dir / 'G+'
    gminus_path = source_dir / 'G-'
    has_gplus   = gplus_path.is_dir()
    has_gminus  = gminus_path.is_dir()

    if has_gplus and has_gminus:
        config['batch_mode']          = True
        config['source_dir_positive'] = gplus_path
        config['source_dir_negative'] = gminus_path
        print(f"\n  Detected BATCH mode — will process G+ and G-")
        print(f"    G+: {gplus_path}")
        print(f"    G-: {gminus_path}")
    else:
        config['batch_mode'] = False
        print(f"\n  Detected SINGLE mode")

    # ── Dataset identifier ─────────────────────────────────────────────────
    print("\n" + "─" * 80)
    print("Dataset Identifier")
    print("─" * 80)

    default_id = source_dir.name

    if config['batch_mode']:
        print(f"  Will create: '<id> Positive/' and '<id> Negative/' output folders")
    print(f"  Press Enter to use folder name: '{default_id}'")

    while True:
        dataset_id = logged_input("Dataset label: ").strip()

        if dataset_id == "":
            dataset_id = default_id
            print(f"  Using folder name: {default_id}")
            break

        if len(dataset_id) > 50:
            print("  Too long (max 50 characters)")
            continue

        invalid_chars = set('<>:"|?*\\/')
        found_invalid = sorted({c for c in dataset_id if c in invalid_chars})
        if found_invalid:
            print(f"  Invalid characters: {', '.join(repr(c) for c in found_invalid)}")
            continue

        confirm = logged_input(
            f"  Confirm '{dataset_id}'? (y/n, Enter=yes): "
        ).strip().lower()
        if confirm in ("", "y", "yes"):
            print(f"  Confirmed: {dataset_id}")
            break

    config['dataset_id']      = dataset_id
    config['dataset_id_base'] = dataset_id

    # ── Percentile ─────────────────────────────────────────────────────────
    print("\n" + "─" * 80)
    print("Percentile Configuration  (top / bottom filtering)")
    print("─" * 80)
    print("  Default: 30%  |  Valid range: 1–40%")

    while True:
        choice = logged_input("Enter percentile (% or Enter for 30%): ").strip()

        if choice == "":
            config['percentile'] = 0.30
            print("  Using default: 30%")
            break

        try:
            value = float(choice)
            if 1.0 <= value <= 40.0:
                config['percentile'] = value / 100.0
                print(f"  Selected: {value:.0f}%")
                break
            else:
                print("  Must be between 1 and 40")
        except ValueError:
            print("  Please enter a valid number")

    # ── Clinical threshold ─────────────────────────────────────────────────
    print("\n" + "─" * 80)
    print("Clinical Classification Threshold")
    print("─" * 80)
    print("  Default: 5%  |  Valid range: 1–20%")

    while True:
        choice = logged_input("Enter threshold (% or Enter for 5%): ").strip()

        if choice == "":
            config['threshold_pct'] = 0.05
            print("  Using default: 5%")
            break

        try:
            value = float(choice)
            if value > 1.0:          # user typed "5", not "0.05"
                value = value / 100.0

            if 0.01 <= value <= 0.20:
                config['threshold_pct'] = value
                print(f"  Selected: {value * 100:.1f}%")
                break
            else:
                print("  Must be between 1% and 20%")
        except ValueError:
            print("  Please enter a valid number")

    return config


def validate_config(config: dict) -> bool:
    """Validate configuration dictionary.

    Checks every key that main() unconditionally accesses, including
    batch-mode-specific source paths and numeric range guards.
    Previously only checked 4 keys, missing dataset_id and batch paths.

    Returns True if config is valid, False otherwise.
    """

    # ── Keys required in every mode ────────────────────────────────────────
    always_required = [
        'source_dir',
        'batch_mode',
        'dataset_id',
        'dataset_id_base',
        'percentile',
        'threshold_pct',
    ]

    for key in always_required:
        if key not in config:
            print(f"[ERROR] Missing required config key: '{key}'")
            return False

    # ── source_dir must exist ──────────────────────────────────────────────
    source_dir = config['source_dir']
    if source_dir is None:
        print("[ERROR] source_dir is None")
        return False
    if not source_dir.exists():
        print(f"[ERROR] Source directory does not exist: {source_dir}")
        return False

    # ── Batch-mode-specific keys ───────────────────────────────────────────
    if config['batch_mode']:
        batch_required = ['source_dir_positive', 'source_dir_negative']
        for key in batch_required:
            if key not in config:
                print(f"[ERROR] Batch mode requires config key: '{key}'")
                return False

        for key, label in [
            ('source_dir_positive', 'G+'),
            ('source_dir_negative', 'G-'),
        ]:
            path = config[key]
            if path is None:
                print(f"[ERROR] {key} is None")
                return False
            if not path.exists():
                print(f"[ERROR] {label} source directory does not exist: {path}")
                return False

    # ── Numeric range guards ───────────────────────────────────────────────
    percentile = config['percentile']
    if not (0.0 < percentile <= 0.5):
        print(
            f"[ERROR] percentile out of range: {percentile:.4f} "
            f"(expected 0 < p ≤ 0.5)"
        )
        return False

    threshold_pct = config['threshold_pct']
    if not (0.0 < threshold_pct <= 0.5):
        print(
            f"[ERROR] threshold_pct out of range: {threshold_pct:.4f} "
            f"(expected 0 < t ≤ 0.5)"
        )
        return False

    return True


def setup_output_directory(config: Dict) -> Path:
    """Create the output directory structure and store subdirectory paths in config.

    Previously dead — main() contained 15 lines of inline mkdir logic that
    duplicated this function. main() now calls this instead.

    In batch mode creates:
        outputs/<timestamp>_<dataset_id>/
            Positive/
            Negative/
    and stores config['positive_output'] and config['negative_output'].

    In single mode creates:
        outputs/<timestamp>_<dataset_id>/

    The timestamp is read from config['timestamp'] so the directory name is
    consistent with what has already been logged at the point of creation.

    Returns
    -------
    Path
        The root output directory (parent in batch mode, the single dir otherwise).
    """
    # Use the timestamp already set in config so directory name is consistent
    # with log messages.  Fall back to a fresh timestamp only as a safety net.
    timestamp  = config.get('timestamp', datetime.now().strftime("%Y%m%d_%H%M%S"))
    dataset_id = config['dataset_id']
    folder_name = f"{timestamp}_{dataset_id}"

    if config.get('batch_mode', False):
        output_root = OUTPUTS_DIR / folder_name
        output_root.mkdir(parents=True, exist_ok=True)

        positive_dir = output_root / "Positive"
        negative_dir = output_root / "Negative"
        positive_dir.mkdir(exist_ok=True)
        negative_dir.mkdir(exist_ok=True)

        # Store in config so downstream functions can locate them
        config['positive_output'] = positive_dir
        config['negative_output'] = negative_dir

        print(f"📁 Output directory: {output_root}")
        print(f"   ├── Positive/")
        print(f"   └── Negative/\n")

        return output_root

    else:
        output_dir = OUTPUTS_DIR / folder_name
        output_dir.mkdir(parents=True, exist_ok=True)

        print(f"📁 Output directory: {output_dir}\n")

        return output_dir


def display_configuration_summary(config: dict) -> None:
    """Display a full configuration summary and ask for user confirmation.

    Previously broken because it read config['bacteria_mode'] but main()
    stores the mode under config['processing_mode'].  Also previously called
    load_bacteria_config_from_json() a second time; now reads the already-
    loaded object from bacteria_config_info.

    Raises SystemExit if the user cancels.
    """
    print("\n" + "=" * 80)
    print("CONFIGURATION SUMMARY")
    print("=" * 80 + "\n")

    # ── Processing mode ────────────────────────────────────────────────────
    # main() stores the mode as config['processing_mode'], not 'bacteria_mode'
    mode                = config.get('processing_mode', 'single')
    bacteria_config_info = config.get('bacteria_config_info', {})

    if mode == 'multi_scan':
        configs_to_scan = bacteria_config_info.get('configs_to_scan', [])
        config_names    = bacteria_config_info.get('config_names', {})
        print(f"Processing Mode   : MULTI-CONFIGURATION SCAN")
        print(f"Configurations    : {len(configs_to_scan)} to test")
        for key in configs_to_scan[:5]:
            print(f"                    • {config_names.get(key, key)}")
        if len(configs_to_scan) > 5:
            print(f"                    … and {len(configs_to_scan) - 5} more")
    else:
        # Use the already-loaded config object — no second JSON read needed
        bacteria_cfg = bacteria_config_info.get('selected_config')
        bacteria_key = bacteria_config_info.get('bacteria_type', 'unknown')
        print(f"Processing Mode   : SINGLE CONFIGURATION")
        if bacteria_cfg is not None:
            print(f"Bacteria Config   : {bacteria_cfg.name}")
            print(f"Description       : {bacteria_cfg.description}")
            print(f"Gaussian σ        : {bacteria_cfg.gaussian_sigma}")
            print(
                f"Area range        : "
                f"{bacteria_cfg.min_area_um2} – {bacteria_cfg.max_area_um2} µm²"
            )
        else:
            print(f"Bacteria Config   : {bacteria_key}  (config object not loaded)")

    print()

    # ── Source / output ────────────────────────────────────────────────────
    if config.get('batch_mode', False):
        print(f"Dataset Mode      : BATCH  (G+ and G-)")
        print(f"Source directory  : {config['source_dir']}")
        print(f"Dataset ID        : {config.get('dataset_id', '—')}")
        print(f"  G+ source       : {config.get('source_dir_positive', '—')}")
        print(f"  G- source       : {config.get('source_dir_negative', '—')}")
    else:
        print(f"Dataset Mode      : SINGLE")
        print(f"Source directory  : {config['source_dir']}")
        print(f"Dataset ID        : {config.get('dataset_id', '—')}")

    print()

    # ── Parameters ─────────────────────────────────────────────────────────
    print(f"Percentile filter : {config.get('percentile', 0.3) * 100:.0f}%")
    print(f"Clinical threshold: {config.get('threshold_pct', 0.05) * 100:.1f}%")

    print("\n" + "=" * 80 + "\n")

    confirm = logged_input(
        "Proceed with this configuration? (y/n, Enter=yes): "
    ).strip().lower()
    if confirm not in ("", "y", "yes"):
        raise SystemExit("Configuration cancelled by user.")

    print("\nConfiguration confirmed — starting processing…\n")


def run_multi_config_scan(config: dict, bacteria_config_info: dict) -> dict:
    """Run multi-configuration scan - ensures correct source directory is used
    
    Args:
        config: Configuration dictionary - MUST have 'current_source' set
        bacteria_config_info: Bacteria configuration info
    
    Returns:
        dict: Results from process_multi_configuration
    """
    
    # ✅ VALIDATION: Ensure current_source is set
    if 'current_source' not in config:
        print("\n⚠️  WARNING: 'current_source' not set in config")
        print("   Falling back to 'source_dir' - results may be incorrect")
        config['current_source'] = config.get('source_dir')
    
    # ✅ DEBUG: Log which directory will be scanned
    source_to_scan = config['current_source']
    print(f"\n🔍 Multi-scan will process: {source_to_scan}")
    
    if source_to_scan is None:
        print("\n❌ ERROR: source_to_scan is None")
        return {
            'ranked_results': [],
            'all_results': {},
            'comparison_df': pd.DataFrame(),
            'report_path': None,
            'configs': {}
        }

    # Verify directory exists and has images
    if not source_to_scan.exists():
        print(f"\n❌ ERROR: Source directory does not exist: {source_to_scan}")
        return {
            'ranked_results': [],
            'all_results': {},
            'comparison_df': pd.DataFrame(),
            'report_path': None,
            'configs': {}
        }
    
    # Quick check for images
    test_images = list(source_to_scan.rglob("*_ch00.tif"))
    print(f"   Found {len(test_images)} images matching pattern\n")
    
    if len(test_images) == 0:
        print(f"⚠️  WARNING: No images found matching '*_ch00.tif' in {source_to_scan}")
        print(f"   Multi-scan may return 0 results\n")
    
    # Call the existing process_multi_configuration function
    results = process_multi_configuration(config)
    
    # Load all configs for return
    configs_dict = {}
    for cfg_key in bacteria_config_info.get('configs_to_scan', []):
        loaded_config = load_bacteria_config_from_json(cfg_key)
        if loaded_config:
            configs_dict[cfg_key] = loaded_config
        
    # Format results for compatibility
    ranked_results = []
    if not results['comparison_df'].empty:
        for idx, row in results['comparison_df'].iterrows():
            ranked_results.append({
                'rank': int(row['Rank']),
                'config_key': str(row['Config_Key']),
                'bacteria_name': str(row['Bacteria_Type']),
                'confidence': float(row['Confidence_Score']),
                'particles': int(row['Particles_Detected']),
                'mean_fluorescence': float(row['Mean_Fluorescence'])
            })
    
    return {
        'ranked_results': ranked_results,
        'all_results': results['all_results'],
        'comparison_df': results['comparison_df'],
        'report_path': results.get('report_path'),
        'configs': configs_dict
    }

def run_single_config_analysis(config: dict) -> None:
    """Run single configuration analysis"""
    
    # ========== STEP 1: Extract and validate bacteria_config ==========
    bacteria_config_raw = config.get('bacteria_config')
    
    # Handle string key
    if isinstance(bacteria_config_raw, str):
        bacteria_config = load_bacteria_config_from_json(bacteria_config_raw)
        if bacteria_config is None:
            print(f"[ERROR] Failed to load bacteria configuration: {bacteria_config_raw}")
            return  # ✅ FIX: Add return statement
    elif bacteria_config_raw is None:
        print("[ERROR] bacteria configuration not found in config dictionary")
        return
    else:
        # Assume it's already a SegmentationConfig object
        bacteria_config = bacteria_config_raw
    
    # ✅ Type assertion for Pylance
    from bacteria_configs import SegmentationConfig
    if not isinstance(bacteria_config, SegmentationConfig):
        print("[ERROR] bacteria_config is not a valid SegmentationConfig object")
        return
    
    # ✅ NOW GUARANTEED: bacteria_config is a valid SegmentationConfig object
    
    source_dir = config['current_source']
    output_dir = config['output_dir']
    percentile = config['percentile']
    threshold_pct = config['threshold_pct']
    
    # Determine microgel type
    if 'Positive' in str(output_dir):
        microgel_type = 'positive'
    elif 'Negative' in str(output_dir):
        microgel_type = 'negative'
    else:
        microgel_type = 'positive'
    
    # Save configuration
    config_file = output_dir / "segmentation_config.txt"
    with open(config_file, 'w', encoding='utf-8') as f:
        f.write(f"Configuration: {bacteria_config.name}\n")  # ✅ Now safe
        f.write(f"Description: {bacteria_config.description}\n")
        f.write(f"Gaussian Sigma: {bacteria_config.gaussian_sigma}\n")
        f.write(f"Area Range: {bacteria_config.min_area_um2} - {bacteria_config.max_area_um2} µm²\n")
    
    # Collect and process images
    print(f"\n📂 Collecting images from: {source_dir.name}")
    image_groups = collect_images_from_directory(source_dir)
    total_images = sum(len(group['images']) for group in image_groups.values())
    print(f"   Found {total_images} images in {len(image_groups)} groups")
    
    print(f"\n⚙️  Processing images...")
    for group_name, group_data in image_groups.items():
        group_output = output_dir / group_name
        
        for img_path in tqdm(group_data['images'], desc=f"  {group_name}"):
            try:
                process_image(img_path, group_output, bacteria_config)
            except Exception as e:
                print(f"\n  ✗ Failed: {img_path.name} - {e}")
    
    # Consolidate to Excel
    print(f"\n📊 Consolidating to Excel...")
    for group_name in image_groups.keys():
        group_dir = output_dir / group_name
        if group_dir.exists():
            try:
                consolidate_to_excel(group_dir, group_name, percentile)
                print(f"   ✓ {group_name}")
            except Exception as e:
                print(f"   ✗ {group_name}: {e}")
    
    # Generate statistics
    print(f"\n📈 Generating statistics...")
    export_group_statistics_to_csv(output_dir)
    
    # Clinical classification
    print(f"\n🔬 Clinical classification...")
    classification_df = classify_groups_clinical(
        output_root=output_dir,
        microgel_type=microgel_type,
        threshold_pct=threshold_pct
    )
    
    if not classification_df.empty:
        csv_path = export_clinical_classification(
            output_root=output_dir,
            classification_df=classification_df,
            microgel_type=microgel_type
        )
        if csv_path:  # ✅ Add None check
            print(f"   ✓ Classification saved: {csv_path.name}")
    
    # Generate plots
    print(f"\n📊 Generating plots...")
    plot_path = generate_error_bar_comparison_with_threshold(
        output_dir=output_dir,
        percentile=percentile,
        threshold_pct=threshold_pct,
        microgel_type=microgel_type,
        dataset_id=config.get('dataset_id_current', '')
    )
    
    if plot_path:
        print(f"   ✓ Comparison plot: {plot_path.name}")
    
    # Pairwise plots
    generate_pairwise_group_vs_control_plots(
        output_root=output_dir,
        percentile=percentile,
        dataset_id=config.get('dataset_id_current', ''),
        threshold_pct=threshold_pct,
        microgel_type=microgel_type
    )
    
    # Embed in Excel
    if plot_path and plot_path.exists():
        embed_comparison_plots_into_all_excels(
            output_root=output_dir,
            percentile=percentile,
            plot_path=plot_path
        )


def generate_final_clinical_matrix_wrapper(output_root: Path, config: dict) -> None:
    """Generate final clinical matrix + PDF laboratory report."""

    print("\n" + "=" * 80)
    print("GENERATING FINAL CLINICAL MATRIX")
    print("=" * 80)

    gplus_csv  = config['positive_output'] / "clinical_classification_positive.csv"
    gminus_csv = config['negative_output'] / "clinical_classification_negative.csv"

    if not gplus_csv.exists() or not gminus_csv.exists():
        print("  ⚠ Missing classification files")
        return

    gplus_df  = pd.read_csv(gplus_csv)
    gminus_df = pd.read_csv(gminus_csv)

    dataset_base = config.get('dataset_id_base', config.get('dataset_id', 'Dataset'))

    matrix_path = generate_final_clinical_matrix(
        output_root=output_root,
        gplus_classification=gplus_df,
        gminus_classification=gminus_df,
        dataset_base_name=dataset_base,
    )
    if matrix_path:
        print(f"  ✓ Final matrix: {matrix_path.name}")

    # ── PDF ──
    print("\n" + "=" * 80)
    print("GENERATING PDF LABORATORY REPORT")
    print("=" * 80)

    final_csv = output_root / "final_clinical_results.csv"
    final_df = pd.read_csv(final_csv) if final_csv.exists() else pd.DataFrame()

    pdf_path = generate_laboratory_report_pdf(
        output_root=output_root,
        config=config,
        gplus_classification=gplus_df,
        gminus_classification=gminus_df,
        final_df=final_df,
    )
    if pdf_path:
        print(f"  ✓ PDF report: {pdf_path.name}")


# ==================================================
# Constants
# ==================================================


def cleanup_and_reorganize_output(output_root: Path, config: dict) -> None:
    """Clean up multi-scan artifacts and reorganize essential files
    
    Keeps only:
    - Positive/ and Negative/ folders (with their contents)
    - confidence_report.txt (explains multi-scan decision) ✅ KEPT
    - clinical_classification_positive.xlsx
    - clinical_classification_negative.xlsx
    - comparison_positive_all_groups.png
    - comparison_negative_all_groups.png
    - final_clinical_results.xlsx (combined G+/G- matrix)
    
    Removes (inside Positive/ and Negative/):
    - Individual bacteria configuration folders (default/, klebsiella_pneumoniae/, etc.)
    - configuration_comparison.csv (redundant with confidence_report.txt)
    - multi_config_comparison_with_statistics.png (intermediate)
    - .cache/ directory
    
    Args:
        output_root: Root output directory
        config: Configuration dictionary
    """
    
    print("\n" + "="*80)
    print("CLEANING UP OUTPUT FOLDER")
    print("="*80)
    
    try:
        # ========== STEP 1: Copy Essential Files to Root ==========
        print("\n📋 Copying essential files to root...")
        
        files_copied = 0
        
        # Copy from Positive/
        positive_dir = config.get('positive_output')
        if positive_dir and positive_dir.exists():
            # Copy classification Excel
            pos_class_excel = positive_dir / "clinical_classification_positive.xlsx"
            if pos_class_excel.exists():
                dest = output_root / "clinical_classification_positive.xlsx"
                shutil.copy2(pos_class_excel, dest)
                print(f"  ✓ Copied: clinical_classification_positive.xlsx")
                files_copied += 1
            
            # Copy comparison plot
            pos_plot = positive_dir / "comparison_positive_all_groups.png"
            if pos_plot.exists():
                dest = output_root / "comparison_positive_all_groups.png"
                shutil.copy2(pos_plot, dest)
                print(f"  ✓ Copied: comparison_positive_all_groups.png")
                files_copied += 1
            
            # Copy confidence report from Positive/ (if not already in root)
            pos_report = positive_dir / "confidence_report.txt"
            if pos_report.exists() and not (output_root / "confidence_report.txt").exists():
                dest = output_root / "confidence_report.txt"
                shutil.copy2(pos_report, dest)
                print(f"  ✓ Copied: confidence_report.txt")
                files_copied += 1
        
        # Copy from Negative/
        negative_dir = config.get('negative_output')
        if negative_dir and negative_dir.exists():
            # Copy classification Excel
            neg_class_excel = negative_dir / "clinical_classification_negative.xlsx"
            if neg_class_excel.exists():
                dest = output_root / "clinical_classification_negative.xlsx"
                shutil.copy2(neg_class_excel, dest)
                print(f"  ✓ Copied: clinical_classification_negative.xlsx")
                files_copied += 1
            
            # Copy comparison plot
            neg_plot = negative_dir / "comparison_negative_all_groups.png"
            if neg_plot.exists():
                dest = output_root / "comparison_negative_all_groups.png"
                shutil.copy2(neg_plot, dest)
                print(f"  ✓ Copied: comparison_negative_all_groups.png")
                files_copied += 1
            
            # Copy confidence report from Negative/ (if not already copied)
            neg_report = negative_dir / "confidence_report.txt"
            if neg_report.exists() and not (output_root / "confidence_report.txt").exists():
                dest = output_root / "confidence_report.txt"
                shutil.copy2(neg_report, dest)
                print(f"  ✓ Copied: confidence_report.txt")
                files_copied += 1
        
        # Check if confidence_report.txt already exists in root (from multi-scan)
        root_report = output_root / "confidence_report.txt"
        if root_report.exists():
            print(f"  ✓ Kept: confidence_report.txt (already in root)")
            files_copied += 1
        
        print(f"  Total files copied/kept: {files_copied}")
        
        # ========== STEP 2 + 3: Remove Artifacts Inside Positive/ and Negative/ ==========
        # 
        # IMPORTANT: In batch mode, process_multi_configuration creates config subfolders
        # *inside* Positive/ and Negative/ (e.g. Positive/default/, Positive/klebsiella/),
        # NOT at output_root level. The old code searched output_root/config_key, which
        # never existed, so nothing was ever removed. This version searches the correct
        # subdirectories.
        
        # Determine which subdirectories to clean
        sub_output_dirs: list[Path] = []
        for key in ('positive_output', 'negative_output'):
            d = config.get(key)
            if isinstance(d, Path) and d.exists():
                sub_output_dirs.append(d)
        
        # Single-directory mode fallback (non-batch)
        if not sub_output_dirs:
            sub_output_dirs = [output_root]
        
        bacteria_config_info = config.get('bacteria_config_info', {})
        configs_scanned = bacteria_config_info.get('configs_to_scan', [])
        
        intermediate_files = [
            "configuration_comparison.csv",          # Redundant with confidence_report.txt
            "multi_config_comparison_with_statistics.png",  # Intermediate plot
        ]
        
        removed_files = 0
        removed_dirs  = 0
        
        for sub_dir in sub_output_dirs:
            print(f"\n🗑️  Cleaning inside: {sub_dir.name}/")
            
            # --- Remove intermediate flat files ---
            for filename in intermediate_files:
                fp = sub_dir / filename
                if fp.exists():
                    try:
                        fp.unlink()
                        print(f"    ✓ Removed: {filename}")
                        removed_files += 1
                    except Exception as e:
                        print(f"    ⚠ Could not remove {filename}: {e}")
            
            # --- Remove .cache directory ---
            cache_dir = sub_dir / ".cache"
            if cache_dir.exists():
                try:
                    shutil.rmtree(cache_dir)
                    print(f"    ✓ Removed: .cache/")
                    removed_dirs += 1
                except Exception as e:
                    print(f"    ⚠ Could not remove .cache/: {e}")
            
            # --- Remove per-configuration scan folders ---
            if not configs_scanned:
                print(f"    (No configuration keys found — skipping folder removal)")
            else:
                for config_key in configs_scanned:
                    config_dir = sub_dir / config_key
                    if config_dir.exists() and config_dir.is_dir():
                        try:
                            shutil.rmtree(config_dir)
                            print(f"    ✓ Removed: {config_key}/")
                            removed_dirs += 1
                        except Exception as e:
                            print(f"    ⚠ Could not remove {config_key}/: {e}")
        
        if removed_files == 0 and removed_dirs == 0:
            print(f"\n  (No intermediate files or configuration folders found to remove)")
        else:
            print(f"\n  Files removed : {removed_files}")
            print(f"  Folders removed: {removed_dirs}")
        
        # ========== STEP 4: Summary ==========
        print("\n" + "─"*80)
        print("FINAL FOLDER STRUCTURE:")
        print("─"*80)
        
        # List what remains at the root level
        remaining_items = []
        for item in sorted(output_root.iterdir()):
            if item.is_dir():
                remaining_items.append(f"📁 {item.name}/")
            else:
                remaining_items.append(f"📄 {item.name}")
        
        for item in remaining_items:
            print(f"  {item}")
        
        # Also list what remains inside Positive/ and Negative/
        for sub_dir in sub_output_dirs:
            if sub_dir.exists():
                print(f"\n  Inside {sub_dir.name}/:")
                for item in sorted(sub_dir.iterdir()):
                    if item.is_dir():
                        print(f"    📁 {item.name}/")
                    else:
                        print(f"    📄 {item.name}")
        
        print("─"*80)
        print(f"✓ Cleanup complete — confidence report preserved")
        print("="*80 + "\n")
        
    except Exception as e:
        print(f"\n⚠ Error during cleanup: {e}")
        import traceback
        traceback.print_exc()


def get_cache_key(image_path: Path, bacteria_config: 'SegmentationConfig') -> str:
    """Generate unique cache key for image + config combination
    
    Args:
        image_path: Path to image
        bacteria_config: Configuration object
        
    Returns:
        str: MD5 hash of image path + config parameters
    """
    import hashlib
    
    # Combine image path and key config parameters
    cache_string = f"{image_path}_{bacteria_config.name}_{bacteria_config.gaussian_sigma}_{bacteria_config.min_area_um2}_{bacteria_config.max_area_um2}"
    
    return hashlib.md5(cache_string.encode()).hexdigest()


def check_cache(cache_dir: Path, cache_key: str) -> Optional[dict]:
    """Check if cached results exist
    
    Args:
        cache_dir: Cache directory
        cache_key: Unique cache key
        
    Returns:
        dict: Cached results or None
    """
    cache_file = cache_dir / f"{cache_key}.json"
    
    if not cache_file.exists():
        return None
    
    try:
        with open(cache_file, 'r', encoding='utf-8') as f:
            cached_data = json.load(f)
        
        # Validate cache
        if 'timestamp' in cached_data and 'results' in cached_data:
            # Check if cache is recent (optional: add expiration)
            return cached_data['results']
    except Exception as e:
        print(f"  ⚠ Cache read error: {e}")
    
    return None


def save_cache(cache_dir: Path, cache_key: str, results: dict) -> None:
    """Save results to cache
    
    Args:
        cache_dir: Cache directory
        cache_key: Unique cache key
        results: Results dictionary to cache
    """
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"{cache_key}.json"
    
    try:
        cache_data = {
            'timestamp': datetime.now().isoformat(),
            'results': results
        }
        
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, indent=2, default=str)
    except Exception as e:
        print(f"  ⚠ Cache write error: {e}")


def clear_old_cache(cache_dir: Path, max_age_days: int = 7) -> None:
    """Clear cache files older than specified days
    
    Args:
        cache_dir: Cache directory
        max_age_days: Maximum age in days
    """
    if not cache_dir.exists():
        return
    
    from datetime import timedelta
    
    cutoff_time = datetime.now() - timedelta(days=max_age_days)
    cleared_count = 0
    
    for cache_file in cache_dir.glob("*.json"):
        try:
            # Check file modification time
            mtime = datetime.fromtimestamp(cache_file.stat().st_mtime)
            
            if mtime < cutoff_time:
                cache_file.unlink()
                cleared_count += 1
        except Exception:
            pass
    
    if cleared_count > 0:
        print(f"  ✓ Cleared {cleared_count} old cache files")


# ==================================================
# process_multi_configuration  (patched)
# Key change: stores intensity_threshold and
# use_intensity_threshold in scan_results so that
# calculate_confidence_score can normalise fluorescence
# against each config's own detection cutoff.
# ==================================================


def process_multi_configuration(config: dict) -> dict:
    """Process images with all available bacteria configurations.

    v2 changes vs original:
        • Separates per-image fluorescence into test vs control lists
          so calculate_confidence_score can measure biological separation.
        • Stores 'test_fluorescences', 'control_fluorescences',
          'test_particles', 'control_particles' in each scan_results dict.
    """
    bacteria_config_info = config.get('bacteria_config_info', {})
    configs_to_scan      = bacteria_config_info.get('configs_to_scan', [])
    config_names         = bacteria_config_info.get('config_names', {})

    _empty = {
        'mode': 'multi_scan',
        'all_results': {},
        'comparison_df': pd.DataFrame(),
        'best_match': None,
        'confidence_report': "No configurations available",
        'comparison_path': None,
        'report_path': None,
        'plot_path': None,
    }

    if not configs_to_scan:
        print("[ERROR] No configurations to scan")
        return _empty

    source_to_scan = config.get('current_source', config.get('source_dir'))
    if source_to_scan is None:
        print("[ERROR] No source directory specified")
        return _empty

    print("\n" + "=" * 80)
    print("MULTI-CONFIGURATION SCANNING")
    print("=" * 80)
    print(f"\n📁 Scanning directory: {source_to_scan.name}")
    print(f"📊 Testing {len(configs_to_scan)} bacteria configurations:")
    for cfg_key in configs_to_scan:
        print(f"   • {config_names[cfg_key]}")
    print()

    image_groups = collect_images_from_directory(source_to_scan)
    total_images = sum(len(g['images']) for g in image_groups.values())

    if total_images == 0:
        print(f"\n⚠️  WARNING: No images found in {source_to_scan}")
        print(f"   Check that:")
        print(f"     1. Images exist in this directory")
        print(f"     2. Images match pattern '*_ch00.tif'")
        print(f"     3. Directory structure is correct")
        print(f"\n   Directory contents:")
        try:
            for item in source_to_scan.iterdir():
                print(f"     - {item.name} {'(dir)' if item.is_dir() else '(file)'}")
        except Exception as e:
            print(f"     Could not list directory: {e}")
        _empty['confidence_report'] = "No images found to process"
        return _empty

    print(f"📂 Found {total_images} images across {len(image_groups)} groups\n")

    cache_dir = config['output_dir'] / ".cache"
    cache_dir.mkdir(exist_ok=True)
    clear_old_cache(cache_dir, max_age_days=7)

    all_results  = {}
    cache_hits   = 0
    cache_misses = 0

    for bacteria_type in configs_to_scan:
        print("─" * 80)
        print(f"Testing: {config_names[bacteria_type]}")
        print("─" * 80)

        bacteria_config = load_bacteria_config_from_json(bacteria_type)
        if bacteria_config is None:
            print(f"  ✗ Failed to load configuration")
            continue

        config_output = config['output_dir'] / bacteria_type
        config_output.mkdir(exist_ok=True)

        scan_results: dict[str, Any] = {
            'particles_detected':       0,
            'fluorescence_sum':         0.0,
            'mean_fluorescence':        0.0,
            'images_processed':         0,
            'images_failed':            0,
            'cached':                   False,
            'per_image_particles':      [],
            'per_image_fluorescence':   [],
            # NEW — test vs control split
            'test_fluorescences':       [],
            'control_fluorescences':    [],
            'test_particles':           [],
            'control_particles':        [],
            # Threshold metadata for report display
            'use_intensity_threshold':  bacteria_config.use_intensity_threshold,
            'intensity_threshold':      (bacteria_config.intensity_threshold
                                         if bacteria_config.use_intensity_threshold
                                         else 80.0),
        }

        for group_name, group_data in image_groups.items():
            is_control = group_name.lower().startswith('control')

            for img_path in group_data['images']:
                cache_key    = get_cache_key(img_path, bacteria_config)
                cached_result = check_cache(cache_dir, cache_key)

                if cached_result is not None:
                    cache_hits     += 1
                    particle_count  = cached_result.get('particle_count', 0)
                    mean_fluor      = cached_result.get('mean_fluorescence', 0.0)
                else:
                    cache_misses += 1
                    try:
                        group_output = config_output / group_name
                        process_image(img_path, group_output, bacteria_config)

                        csv_path = group_output / img_path.stem / "object_stats.csv"
                        if csv_path.exists():
                            df             = pd.read_csv(csv_path)
                            particle_count = len(df)

                            if 'Fluor_Mean' in df.columns and len(df) > 0:
                                fluor_values = pd.to_numeric(df['Fluor_Mean'],
                                                             errors='coerce')
                                valid_fluor  = fluor_values.dropna()
                                mean_fluor   = (float(valid_fluor.mean())
                                                if len(valid_fluor) > 0 else 0.0)
                            else:
                                mean_fluor = 0.0

                            save_cache(cache_dir, cache_key, {
                                'particle_count':    particle_count,
                                'mean_fluorescence': mean_fluor,
                            })
                        else:
                            continue
                    except Exception as e:
                        print(f"  ✗ Failed: {img_path.name} - {e}")
                        scan_results['images_failed'] += 1
                        continue

                # ── Accumulate into both pooled AND split lists ──
                scan_results['particles_detected']     += particle_count
                scan_results['fluorescence_sum']       += mean_fluor
                scan_results['images_processed']       += 1
                scan_results['per_image_particles'].append(particle_count)
                scan_results['per_image_fluorescence'].append(mean_fluor)

                if cached_result is not None:
                    scan_results['cached'] = True

                # NEW: test vs control split
                if is_control:
                    scan_results['control_fluorescences'].append(mean_fluor)
                    scan_results['control_particles'].append(particle_count)
                else:
                    scan_results['test_fluorescences'].append(mean_fluor)
                    scan_results['test_particles'].append(particle_count)

        # ── Aggregate statistics (pooled — for display) ──
        n_proc = scan_results['images_processed']
        if n_proc > 0:
            scan_results['mean_fluorescence'] = (
                scan_results['fluorescence_sum'] / n_proc)

            fa = np.array(scan_results['per_image_fluorescence'])
            if len(fa) > 1:
                scan_results['std_fluorescence'] = float(np.std(fa, ddof=1))
                scan_results['sem_fluorescence'] = float(scipy_stats.sem(fa))
            else:
                scan_results['std_fluorescence'] = 0.0
                scan_results['sem_fluorescence'] = 0.0

            pa = np.array(scan_results['per_image_particles'])
            if len(pa) > 1:
                scan_results['mean_particles_per_image'] = float(np.mean(pa))
                scan_results['std_particles']            = float(np.std(pa, ddof=1))
            else:
                scan_results['mean_particles_per_image'] = float(
                    scan_results['particles_detected'])
                scan_results['std_particles'] = 0.0
        else:
            for k in ('mean_fluorescence', 'std_fluorescence',
                      'sem_fluorescence', 'mean_particles_per_image',
                      'std_particles'):
                scan_results[k] = 0.0

        # Excel consolidation
        if n_proc > 0:
            for group_name in image_groups.keys():
                group_dir = config_output / group_name
                if group_dir.exists():
                    try:
                        consolidate_to_excel(
                            group_dir, group_name, config['percentile'])
                    except Exception as e:
                        print(f"  ⚠ Excel consolidation failed: {e}")

        all_results[bacteria_type] = scan_results

        # ── Console summary ──
        n_test = len(scan_results['test_fluorescences'])
        n_ctrl = len(scan_results['control_fluorescences'])
        print(f"\n  Results:")
        print(f"    Particles detected: "
              f"{scan_results['particles_detected']} total")
        print(f"    Particles per image: "
              f"{scan_results['mean_particles_per_image']:.1f} "
              f"± {scan_results['std_particles']:.1f}")
        print(f"    Mean fluorescence: "
              f"{scan_results['mean_fluorescence']:.2f} "
              f"± {scan_results['std_fluorescence']:.2f} a.u./µm²")
        print(f"    SEM fluorescence: "
              f"±{scan_results['sem_fluorescence']:.2f}")
        print(f"    Images processed: {n_proc} "
              f"(test: {n_test}, control: {n_ctrl})")
        if scan_results['cached']:
            print(f"    Cache hits: ✓")
        print()

    print(f"\n📊 Cache Statistics:")
    print(f"   Cache hits: {cache_hits}")
    print(f"   Cache misses: {cache_misses}")
    total_cache = cache_hits + cache_misses
    if total_cache > 0:
        print(f"   Hit rate: {cache_hits / total_cache * 100:.1f}%")
    else:
        print(f"   Hit rate: N/A")

    comparison_df, confidence_report = generate_confidence_report(
        all_results, config_names, config['output_dir'])

    plot_path = generate_multi_config_comparison_plot(
        comparison_df, all_results, config['output_dir'])

    comparison_path = config['output_dir'] / "configuration_comparison.csv"
    comparison_df.to_csv(comparison_path, index=False)

    report_path = config['output_dir'] / "confidence_report.txt"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(confidence_report)

    print("\n" + "=" * 80)
    print("MULTI-SCAN RESULTS")
    print("=" * 80)
    print("\n" + confidence_report)

    best_match = comparison_df.iloc[0] if not comparison_df.empty else None

    return {
        'mode':              'multi_scan',
        'all_results':       all_results,
        'comparison_df':     comparison_df,
        'best_match':        best_match,
        'confidence_report': confidence_report,
        'comparison_path':   comparison_path,
        'report_path':       report_path,
        'plot_path':         plot_path,
    }


# ==================================================
# Scores based on how well
# a config separates test images from control images,
# NOT on absolute fluorescence magnitude.
#
# Scoring breakdown (max 100):
#   Factor 1 — Effect size (Cohen's d)           0-35
#   Factor 2 — Statistical significance           0-20
#   Factor 3 — Particle count appropriateness     0-15  ← Approach 1 applied
#   Factor 4 — Within-group consistency           0-15
#   Factor 5 — Processing completeness            0-10  (reserved, not yet scored)
#   Factor 6 — Direction & magnitude bonus         0-5
#
# Post-scoring modifiers:
#   Approach 2 — Non-specific detection disqualifier
#                (overcount + low fluorescence hard cap / multiplier)
#   Approach 3 — Stores '_overcount_disqualified' flag back into
#                the results dict for generate_confidence_report labeling
#
# Tunable constants (adjust to match your assay biology):
#   OVERCOUNT_HARD_LIMIT    : particles/image above which hard cap triggers
#   OVERCOUNT_MID_LIMIT     : particles/image for moderate penalty
#   OVERCOUNT_SOFT_LIMIT    : particles/image for borderline penalty
#   NONSPECIFIC_FLUOR_CAP   : a.u./µm²  — pooled mean below this is suspicious
#   NONSPECIFIC_FLUOR_STRICT: a.u./µm²  — stricter version for borderline band
# ==================================================

def calculate_confidence_score(
        results: dict,
        default_fluor: Optional[float] = None
    ) -> float:
    """Score a bacteria configuration based on test-vs-control separation (0-100).

    Approach 1 — Particle-count ceiling tightened to 30/image with
                 a steep degradation ladder above it.  Configs detecting
                 hundreds of particles/image at low fluorescence (noise,
                 debris, crystal artefacts) can no longer score well on
                 Factor 3 alone.

    Approach 2 — Non-specific detection disqualifier applied AFTER all
                 factor scores are summed.  Targets the specific artefact
                 signature: high particle count AND low mean fluorescence.
                 Three severity bands with progressively stronger penalties:
                   • Extreme  (>100/img, fluor<2.0) → hard cap at 10 pts
                   • Moderate ( >50/img, fluor<2.0) → score × 0.35
                   • Borderline(>30/img, fluor<1.0) → score × 0.55

    Approach 3 — Stores three diagnostic keys back into the results dict
                 so that generate_confidence_report can label and section
                 off disqualified configs without re-computing anything:
                   results['_overcount_disqualified'] : bool
                   results['_mean_per_image_used']    : float  (rounded)
                   results['_pooled_fluor_mean']      : float  (rounded)

    If the results dict does NOT contain test/control split keys
    (backward compat), falls back to a minimal heuristic so the
    function never crashes.

    Args:
        results     : per-config dict produced by process_multi_configuration
        default_fluor: mean_fluorescence of the 'default' config run, used
                       as a signal-to-background reference (optional)

    Returns:
        float in [0, 100]
    """

    # ── Tunable disqualification thresholds ────────────────────────────────────
    OVERCOUNT_HARD_LIMIT     = 100.0   # particles/image
    OVERCOUNT_MID_LIMIT      =  50.0   # particles/image
    OVERCOUNT_SOFT_LIMIT     =  30.0   # particles/image
    NONSPECIFIC_FLUOR_CAP    =   2.0   # a.u./µm²  (hard / moderate bands)
    NONSPECIFIC_FLUOR_STRICT =   1.0   # a.u./µm²  (borderline band)

    # ── Approach 3: reset flags at the top so every call is clean ──────────────
    results['_overcount_disqualified'] = False
    results['_mean_per_image_used']    = 0.0
    results['_pooled_fluor_mean']      = 0.0

    # ── Pull arrays from results ────────────────────────────────────────────────
    test_fluor = np.array(
        results.get('test_fluorescences', []), dtype=float)
    ctrl_fluor = np.array(
        results.get('control_fluorescences', []), dtype=float)
    test_particles = np.array(
        results.get('test_particles', []), dtype=float)
    ctrl_particles = np.array(
        results.get('control_particles', []), dtype=float)

    n_test = len(test_fluor)
    n_ctrl = len(ctrl_fluor)

    # ── Compute mean_per_image early — needed by Approaches 1-3 ────────────────
    # Priority: per_image_particles list (most accurate) → test/ctrl split
    # totals → stored aggregate value.
    per_image_particles = results.get('per_image_particles', [])
    if per_image_particles:
        mean_per_image = float(np.mean(per_image_particles))
    elif (n_test + n_ctrl) > 0:
        total_particles = (
            float(np.sum(test_particles)) + float(np.sum(ctrl_particles))
        )
        mean_per_image = total_particles / max(n_test + n_ctrl, 1)
    else:
        mean_per_image = float(results.get('mean_particles_per_image', 0))

    # ── Guard: need ≥2 images in BOTH groups for separation scoring ─────────────
    if n_test < 2 or n_ctrl < 2:
        # Fallback: minimal heuristic from particle count alone.
        # Still apply Approach 2/3 flags using aggregate fluorescence.
        mean_fluor_fallback = float(results.get('mean_fluorescence', 0.0))
        results['_mean_per_image_used'] = round(mean_per_image, 2)
        results['_pooled_fluor_mean']   = round(mean_fluor_fallback, 3)

        is_disqualified_fb = (
            mean_per_image > OVERCOUNT_SOFT_LIMIT
            and mean_fluor_fallback < NONSPECIFIC_FLUOR_CAP
        )
        results['_overcount_disqualified'] = is_disqualified_fb

        n_particles = results.get('particles_detected', 0)
        if is_disqualified_fb:
            # Even in fallback, overcount disqualification applies
            return 3.0
        if 20 <= n_particles <= 200:
            return 15.0
        elif 5 <= n_particles < 20 or 200 < n_particles <= 500:
            return 8.0
        elif n_particles > 0:
            return 3.0
        return 0.0

    # ── Descriptive stats ───────────────────────────────────────────────────────
    test_mean = float(np.mean(test_fluor))
    ctrl_mean = float(np.mean(ctrl_fluor))
    test_std  = float(np.std(test_fluor,  ddof=1))
    ctrl_std  = float(np.std(ctrl_fluor,  ddof=1))

    score = 0.0

    # ── Factor 1: Effect size — |Cohen's d| (0-35 pts) ─────────────────────────
    denom = n_test + n_ctrl - 2
    if denom > 0:
        pooled_std = np.sqrt(
            ((n_test - 1) * test_std ** 2 + (n_ctrl - 1) * ctrl_std ** 2)
            / denom
        )
    else:
        pooled_std = 1.0

    cohens_d = (
        abs(test_mean - ctrl_mean) / pooled_std
        if pooled_std > 0 else 0.0
    )

    if   cohens_d >= 2.0: effect_score = 35
    elif cohens_d >= 1.5: effect_score = 30
    elif cohens_d >= 1.0: effect_score = 25
    elif cohens_d >= 0.8: effect_score = 20
    elif cohens_d >= 0.5: effect_score = 12
    elif cohens_d >= 0.2: effect_score = 5
    else:                 effect_score = 0
    score += effect_score

    # ── Factor 2: Statistical significance (0-20 pts) ──────────────────────────
    try:
        _, p_raw = scipy_stats.ttest_ind(
            test_fluor, ctrl_fluor, equal_var=False)
        p_value = float(np.asarray(p_raw).item())
    except Exception:
        p_value = 1.0

    if   p_value < 0.001: sig_score = 20
    elif p_value < 0.01:  sig_score = 15
    elif p_value < 0.05:  sig_score = 10
    elif p_value < 0.10:  sig_score = 3
    else:                 sig_score = 0
    score += sig_score

    # ── Factor 3: Particle count appropriateness (0-15 pts) ────────────────────
    # APPROACH 1 — Ceiling tightened from 50 → 30 particles/image.
    # Steep degradation bands replace the single upper check so that
    # configs detecting 100+ particles/image (likely noise) cannot earn
    # a respectable Factor 3 score at all.
    #
    #   3 – 30  /img  → 15 pts  (optimal biological range)
    #   2 – 3   /img  → 10 pts  (slightly sparse but plausible)
    #  30 – 60  /img  →  5 pts  (borderline overcount)
    #  60 – 150 /img  →  2 pts  (high overcount, probably artefact)
    #  > 150    /img  →  0 pts  (extreme overcount — certain artefact)
    #   1 – 2   /img  →  5 pts  (very sparse detection)
    #   0 – 1   /img  →  0 pts  (nothing detected)
    if   3  <= mean_per_image <= 30:   particle_score = 15
    elif 2  <= mean_per_image <   3:   particle_score = 10
    elif 30 <  mean_per_image <= 60:   particle_score =  5
    elif 60 <  mean_per_image <= 150:  particle_score =  2
    elif mean_per_image > 150:         particle_score =  0
    elif mean_per_image >= 1:          particle_score =  5
    else:                              particle_score =  0
    score += particle_score

    # ── Factor 4: Within-group consistency (0-15 pts) ──────────────────────────
    test_cv = (test_std / test_mean) if test_mean > 0 else 999.0
    ctrl_cv = (ctrl_std / ctrl_mean) if ctrl_mean > 0 else 999.0
    avg_cv  = (test_cv + ctrl_cv) / 2.0

    if   avg_cv < 0.30: consistency_score = 15
    elif avg_cv < 0.50: consistency_score = 10
    elif avg_cv < 0.80: consistency_score = 5
    else:               consistency_score = 0
    score += consistency_score

    # ── Factor 5: Processing completeness (0-10 pts) — reserved ───────────────
    # Not yet scored; kept as a placeholder to preserve the max-100 design.

    # ── Factor 6: Direction & magnitude bonus (0-5 pts) ────────────────────────
    # Bacteria presence quenches fluorescence → test < control is the
    # expected direction.  Reward configs that show that.
    direction_bonus = 0.0
    if test_mean < ctrl_mean and cohens_d >= 0.5:
        direction_bonus = 3.0
    elif test_mean < ctrl_mean:
        direction_bonus = 1.0

    sb_bonus = 0.0
    if default_fluor is not None and default_fluor > 0:
        pooled_mean_sb = float(
            np.mean(np.concatenate([test_fluor, ctrl_fluor]))
        )
        sb_ratio = pooled_mean_sb / default_fluor
        if sb_ratio >= 1.5:
            sb_bonus = 2.0
        elif sb_ratio >= 1.1:
            sb_bonus = 1.0

    score += min(direction_bonus + sb_bonus, 5.0)

    # ══════════════════════════════════════════════════════════════════════════════
    # APPROACH 2 — Non-specific Detection Disqualifier
    # ══════════════════════════════════════════════════════════════════════════════
    # Applied AFTER all factor scores are summed.
    #
    # Rationale: a config can score well on Cohen's d even when both test
    # and control groups are flooded with noise detections, because the
    # relative difference between two noisy distributions can still be
    # statistically significant.  Biological plausibility requires BOTH a
    # reasonable particle count AND a meaningful fluorescence signal.
    #
    # Three severity bands (use OVERCOUNT_* constants at top to re-tune):
    #
    #  Extreme   >100/img AND fluor < 2.0  → hard cap at 10
    #             (Streptococcus 410/img, fluor=0.51 → hits this band)
    #             (Default       855/img, fluor=0.64 → hits this band)
    #
    #  Moderate   >50/img AND fluor < 2.0  → score × 0.35
    #
    #  Borderline >30/img AND fluor < 1.0  → score × 0.55
    #
    #  Proteus  4.7/img, fluor=38.4  → no band triggered, score unchanged
    #  Klebsiella 3.9/img, fluor=23  → no band triggered, score unchanged
    # ══════════════════════════════════════════════════════════════════════════════

    pooled_fluor_mean = float(
        np.mean(np.concatenate([test_fluor, ctrl_fluor]))
    )
    is_disqualified = False

    if mean_per_image > OVERCOUNT_HARD_LIMIT and pooled_fluor_mean < NONSPECIFIC_FLUOR_CAP:
        # ── Extreme band ──────────────────────────────────────────────────────
        original_score = score
        score = min(score, 10.0)
        is_disqualified = True
        print(
            f"  ⛔ OVERCOUNT DISQUALIFIED [extreme]: "
            f"{mean_per_image:.1f} particles/img, "
            f"fluor={pooled_fluor_mean:.3f} a.u./µm² "
            f"→ score capped: {original_score:.1f} → {score:.1f}"
        )

    elif mean_per_image > OVERCOUNT_MID_LIMIT and pooled_fluor_mean < NONSPECIFIC_FLUOR_CAP:
        # ── Moderate band ─────────────────────────────────────────────────────
        original_score = score
        score = score * 0.35
        is_disqualified = True
        print(
            f"  ⚠ OVERCOUNT PENALTY [moderate]: "
            f"{mean_per_image:.1f} particles/img, "
            f"fluor={pooled_fluor_mean:.3f} a.u./µm² "
            f"→ {original_score:.1f} × 0.35 = {score:.1f}"
        )

    elif mean_per_image > OVERCOUNT_SOFT_LIMIT and pooled_fluor_mean < NONSPECIFIC_FLUOR_STRICT:
        # ── Borderline band ───────────────────────────────────────────────────
        original_score = score
        score = score * 0.55
        is_disqualified = True
        print(
            f"  ⚠ OVERCOUNT PENALTY [borderline]: "
            f"{mean_per_image:.1f} particles/img, "
            f"fluor={pooled_fluor_mean:.3f} a.u./µm² "
            f"→ {original_score:.1f} × 0.55 = {score:.1f}"
        )

    # ══════════════════════════════════════════════════════════════════════════════
    # APPROACH 3 — Store diagnostic flags back into results dict
    # ══════════════════════════════════════════════════════════════════════════════
    # Consumed by generate_confidence_report to:
    #   • Print a dedicated "DISQUALIFIED CONFIGURATIONS" section above Top-3
    #   • Add an "Overcount_Disqualified" column to comparison_df
    #   • Adjust clinical recommendations text accordingly
    #
    # Keys written:
    #   _overcount_disqualified : bool  — True if any band above was triggered
    #   _mean_per_image_used    : float — the mean_per_image value that was tested
    #   _pooled_fluor_mean      : float — the pooled fluorescence mean that was tested
    # ══════════════════════════════════════════════════════════════════════════════
    results['_overcount_disqualified'] = is_disqualified
    results['_mean_per_image_used']    = round(mean_per_image, 2)
    results['_pooled_fluor_mean']      = round(pooled_fluor_mean, 3)

    return min(score, 100.0)




def _apply_pairwise_discrimination_penalties(
    all_results: dict,
    raw_scores: dict[str, float],
) -> tuple[dict[str, float], dict[str, dict]]:
    """Adjust confidence scores based on pairwise statistical separation.

    For each adjacent pair (by descending score), if their per-image
    fluorescence distributions are NOT significantly different (p >= 0.05),
    the higher-scoring config's score is pulled toward the lower one.

    Returns
    -------
    adjusted_scores : dict  config_key -> adjusted score
    penalty_log     : dict  config_key -> {peer, p_value, factor, old_score, new_score}
    """
    adjusted = dict(raw_scores)
    penalty_log: dict[str, dict] = {}

    sorted_keys = sorted(raw_scores, key=lambda k: raw_scores[k], reverse=True)

    for i in range(len(sorted_keys) - 1):
        key_hi = sorted_keys[i]
        key_lo = sorted_keys[i + 1]

        fluor_hi = [
            float(x) for x in
            all_results[key_hi].get('per_image_fluorescence', [])
            if x is not None and not (isinstance(x, float) and np.isnan(x))
        ]
        fluor_lo = [
            float(x) for x in
            all_results[key_lo].get('per_image_fluorescence', [])
            if x is not None and not (isinstance(x, float) and np.isnan(x))
        ]

        if len(fluor_hi) < 3 or len(fluor_lo) < 3:
            continue

        try:
            _, p_raw = scipy_stats.ttest_ind(
                fluor_hi, fluor_lo, equal_var=False)
            p_val = float(np.asarray(p_raw).item())
        except Exception:
            continue

        if p_val >= 0.05:
            score_hi = adjusted[key_hi]
            score_lo = adjusted[key_lo]

            convergence_factor = 0.30 + 0.60 * min(
                (p_val - 0.05) / 0.95, 1.0)

            new_hi = score_hi - (score_hi - score_lo) * convergence_factor
            new_hi = round(max(new_hi, score_lo), 1)

            adjusted[key_hi] = new_hi

            penalty_log[key_hi] = {
                'peer':      key_lo,
                'p_value':   round(p_val, 6),
                'factor':    round(convergence_factor, 3),
                'old_score': round(score_hi, 1),
                'new_score': new_hi,
            }

            print(f"  ⚠ Pairwise penalty: {key_hi} "
                  f"{score_hi:.1f}% → {new_hi:.1f}% "
                  f"(vs {key_lo}, p={p_val:.4f}, "
                  f"factor={convergence_factor:.2f})")

    return adjusted, penalty_log


# ==================================================
# generate_confidence_report  (patched)
# Changes vs original:
#   • Computes signal-to-background (S/B) ratio using
#     the 'default' config run as a background reference.
#   • Stores S/B in comparison_df and displays it.
#   • After the top-2 t-test, applies a convergence
#     penalty to the rank-1 score when p ≥ 0.05,
#     preventing a statistically indistinguishable
#     runner-up from being confidently dismissed.
#   • Updates report text to surface ambiguity clearly.
# ==================================================


def generate_confidence_report(
    all_results: dict,
    config_names: dict,
    output_dir: Path,
) -> tuple[pd.DataFrame, str]:
    """Generate confidence report with test-vs-control separation metrics."""

    # ── Background baseline from 'default' config ─────────────
    _default_fluor = None
    if 'default' in all_results:
        _raw = all_results['default'].get('mean_fluorescence', 0.0)
        _default_fluor = max(float(_raw), 0.1)

    # ── PASS 1: Compute raw per-config scores ─────────────────
    raw_scores: dict[str, float] = {}
    for bacteria_type, results in all_results.items():
        raw_scores[bacteria_type] = calculate_confidence_score(
            results, default_fluor=_default_fluor)

    # ── PASS 2: Apply pairwise discrimination penalties ───────
    adjusted_scores, penalty_log = _apply_pairwise_discrimination_penalties(
        all_results, raw_scores)

    # ── FIX: Compute and store per-config Cohen's d and p-value ──
    # calculate_confidence_score computes these internally but never
    # stores them back, so the display table always showed d=0.00,
    # p=1.0000. We compute them here and write them into each
    # results dict so comparison_data can read them correctly.
    for bacteria_type, results in all_results.items():
        test_fl  = np.array(results.get('test_fluorescences',    []), dtype=float)
        ctrl_fl  = np.array(results.get('control_fluorescences', []), dtype=float)

        if len(test_fl) >= 2 and len(ctrl_fl) >= 2:
            n_t   = len(test_fl)
            n_c   = len(ctrl_fl)
            denom = n_t + n_c - 2

            pooled_std = (
                np.sqrt(
                    ((n_t - 1) * np.std(test_fl,  ddof=1) ** 2
                   + (n_c - 1) * np.std(ctrl_fl, ddof=1) ** 2)
                    / denom
                )
                if denom > 0 else 1.0
            )

            d_display = (
                float((np.mean(test_fl) - np.mean(ctrl_fl)) / pooled_std)
                if pooled_std > 0 else 0.0
            )

            try:
                _, p_raw   = scipy_stats.ttest_ind(test_fl, ctrl_fl, equal_var=False)
                p_display  = float(np.asarray(p_raw).item())
            except Exception:
                p_display  = np.nan
        else:
            d_display = 0.0
            p_display = np.nan

        results['_score_cohens_d'] = round(d_display, 3)
        results['_score_p_value']  = (
            round(p_display, 6) if not np.isnan(p_display) else np.nan
        )

    # ── PASS 3: Build comparison table using ADJUSTED scores ──
    comparison_data = []

    for bacteria_type, results in all_results.items():
        confidence_score = adjusted_scores[bacteria_type]

        # S/B ratio
        if (_default_fluor is not None
                and _default_fluor > 0
                and bacteria_type != 'default'):
            sb_ratio = results['mean_fluorescence'] / _default_fluor
        else:
            sb_ratio = 1.0

        # Test vs control means (for display)
        test_fl = results.get('test_fluorescences', [])
        ctrl_fl = results.get('control_fluorescences', [])
        test_mean = float(np.mean(test_fl)) if test_fl else 0.0
        ctrl_mean = float(np.mean(ctrl_fl)) if ctrl_fl else 0.0

        # Low-fluorescence non-specific detection penalty (display flag only)
        pooled_mean_fluor = float(
            np.mean(np.array(test_fl + ctrl_fl, dtype=float))
        ) if (test_fl or ctrl_fl) else 0.0
        low_fluor_penalty = pooled_mean_fluor < 1.0

        comparison_data.append({
            'Rank':                   0,
            'Bacteria_Type':          config_names.get(bacteria_type,
                                                       bacteria_type),
            'Config_Key':             bacteria_type,
            'Particles_Detected':     int(results['particles_detected']),
            'Particles_Per_Image':    float(results.get(
                'mean_particles_per_image', 0)),
            'Particles_Std':          float(results.get('std_particles', 0)),
            'Mean_Fluorescence':      float(results['mean_fluorescence']),
            'Fluor_Std':              float(results.get('std_fluorescence', 0)),
            'Fluor_SEM':              float(results.get('sem_fluorescence', 0)),
            'Signal_to_Background':   round(sb_ratio, 2),
            'Test_Mean':              round(test_mean, 2),
            'Control_Mean':           round(ctrl_mean, 2),
            'Cohens_d':               float(results.get('_score_cohens_d', 0)),
            'P_Value_TvC':            (
                float(results['_score_p_value'])
                if not pd.isna(results.get('_score_p_value', np.nan))
                else np.nan
            ),
            'Images_Processed':       int(results['images_processed']),
            'Images_Failed':          int(results['images_failed']),
            'Raw_Score':              round(raw_scores[bacteria_type], 1),
            'Confidence_Score':       float(confidence_score),
            'Confidence_Percent':     float(confidence_score),
            'Penalty_Applied':        bacteria_type in penalty_log,
            'Low_Fluor_Flag':         low_fluor_penalty,

            'Overcount_Disqualified': bool(results.get('_overcount_disqualified', False)),
        })

    comparison_df = pd.DataFrame(comparison_data)

    if comparison_df.empty:
        return comparison_df, "No results to report"

    comparison_df = comparison_df.sort_values(
        'Confidence_Percent', ascending=False).reset_index(drop=True)
    comparison_df['Rank'] = range(1, len(comparison_df) + 1)

    # ── t-test between rank-1 and rank-2 ──────────────────────
    config_keys = list(comparison_df['Config_Key'].values)
    top1_key = str(config_keys[0])
    top2_key = str(config_keys[1]) if len(config_keys) > 1 else None

    stat_ambiguous      = False
    p_value_top2        = np.nan
    t_stat_top2         = np.nan

    if top2_key is not None:
        top1_fluor = [
            float(x) for x in
            all_results[top1_key].get('per_image_fluorescence', [])
            if x is not None and not (isinstance(x, float) and np.isnan(x))
        ]
        top2_fluor = [
            float(x) for x in
            all_results[top2_key].get('per_image_fluorescence', [])
            if x is not None and not (isinstance(x, float) and np.isnan(x))
        ]

        if len(top1_fluor) >= 3 and len(top2_fluor) >= 3:
            try:
                t_raw, p_raw = scipy_stats.ttest_ind(
                    top1_fluor, top2_fluor, equal_var=False)
                t_stat_top2  = float(np.asarray(t_raw).item())
                p_value_top2 = float(np.asarray(p_raw).item())

                if p_value_top2 >= 0.05:
                    stat_ambiguous = True

            except Exception as e:
                print(f"  Could not perform t-test: {e}")

    # ── Build report text ─────────────────────────────────────
    lines: list[str] = []
    lines.append("=" * 80)
    lines.append("BACTERIA IDENTIFICATION CONFIDENCE REPORT")
    lines.append("=" * 80)
    lines.append("")
    lines.append(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"Output Directory: {output_dir.name}")
    lines.append("")

    top_result     = comparison_df.iloc[0]
    top_confidence = float(top_result['Confidence_Percent'])
    top_particles  = int(top_result['Particles_Detected'])
    top_name       = str(top_result['Bacteria_Type'])

    lines.append("DETECTION SUMMARY:")
    lines.append("─" * 80)

    if top_particles == 0:
        lines.append("Status: NO BACTERIA DETECTED")
        lines.append("Confidence: High")
        lines.append("")
        lines.append("⚪ All configurations detected 0 particles")
        lines.append("   → Sample appears clean")
    elif stat_ambiguous:
        runner = str(comparison_df.iloc[1]['Bacteria_Type'])
        lines.append(
            "Status: BACTERIA DETECTED — AMBIGUOUS IDENTIFICATION")
        lines.append(
            f"Confidence: Reduced ({top_confidence:.1f}% after "
            f"convergence penalty, p={p_value_top2:.4f})")
        lines.append("")
        lines.append(
            "⚠ Top two configurations are NOT statistically "
            "distinguishable")
        lines.append(f"  Rank 1: {top_name}")
        lines.append(f"  Rank 2: {runner}")
        lines.append(
            "  → Manual review or culture confirmation "
            "STRONGLY recommended")
    elif top_confidence >= 70:
        lines.append("Status: BACTERIA DETECTED")
        lines.append(f"Confidence: High ({top_confidence:.1f}%)")
        lines.append("")
        lines.append(f"✓ Strong match to: {top_name}")
    elif top_confidence >= 50:
        lines.append("Status: BACTERIA DETECTED")
        lines.append(f"Confidence: Moderate ({top_confidence:.1f}%)")
        lines.append("")
        lines.append(f"⚠ Possible match to: {top_name}")
        lines.append("   → Manual review recommended")
    else:
        lines.append("Status: AMBIGUOUS")
        lines.append(f"Confidence: Low ({top_confidence:.1f}%)")
        lines.append("")
        lines.append("⚠ MANUAL REVIEW REQUIRED")
        lines.append(
            "   → Multiple configurations show similar results")
        lines.append(
            "   → Consider culture-based confirmation")
    lines.append("")
    lines.append("")




    disqualified_keys = [
        k for k, v in all_results.items()
        if v.get('_overcount_disqualified', False)
    ]
    if disqualified_keys:
        lines.append("⛔ DISQUALIFIED CONFIGURATIONS (overcount + low fluorescence):")
        lines.append("─" * 80)
        for key in disqualified_keys:
            ppi  = all_results[key].get('_mean_per_image_used', 0)
            mf   = all_results[key].get('_pooled_fluor_mean',   0)
            lines.append(
                f"  ✗ {config_names.get(key, key)}: "
                f"{ppi:.0f} particles/image, "
                f"mean fluor = {mf:.3f} a.u./µm²  "
                f"→ Non-specific detections (overcount artefact)"
            )
        lines.append("")

    # ── Top-3 matches ─────────────────────────────────────────
    lines.append("TOP 3 CONFIGURATION MATCHES:")
    lines.append("─" * 80)

    for idx in range(min(3, len(comparison_df))):
        row  = comparison_df.iloc[idx]
        rank = int(row['Rank'])
        name = str(row['Bacteria_Type'])

        prefix = ("🥇" if rank == 1 else "🥈" if rank == 2 else "🥉")
        conf   = float(row['Confidence_Percent'])
        d_val  = float(row['Cohens_d'])
        p_tvc  = row['P_Value_TvC']
        p_tvc_display = (
            f"{float(p_tvc):.4f}"
            if not pd.isna(p_tvc) else "N/A"
        )

        fluor_mean = float(row['Mean_Fluorescence'])
        fluor_std  = float(row['Fluor_Std'])
        fluor_sem  = float(row['Fluor_SEM'])
        sb         = float(row['Signal_to_Background'])
        t_mean     = float(row['Test_Mean'])
        c_mean     = float(row['Control_Mean'])
        low_flag   = bool(row.get('Low_Fluor_Flag', False))

        fluor_display = (
            f"{fluor_mean:.2f} ± {fluor_std:.2f} a.u./µm² "
            f"(SEM: ±{fluor_sem:.2f})"
            if fluor_mean > 0 else "N/A")

        lines.append(f"{prefix} Rank {rank}: {name}")
        lines.append(
            f"   Confidence: {conf:.1f}%"
            + (f" (raw: {float(row.get('Raw_Score', conf)):.1f}%,"
               f" pairwise penalty applied)"
               if row.get('Penalty_Applied', False) else ""))
        if low_flag:
            lines.append(
                f"   ⚠ Low fluorescence flag: mean < 1.0 a.u./µm² "
                f"— detections may be non-specific")
        lines.append(
            f"   Total particles: {int(row['Particles_Detected'])}")
        lines.append(
            f"   Particles/image: "
            f"{float(row['Particles_Per_Image']):.1f} "
            f"± {float(row['Particles_Std']):.1f}")
        lines.append(f"   Mean fluorescence: {fluor_display}")
        lines.append(
            f"   Signal-to-background: {sb:.1f}×")
        lines.append(
            f"   Test vs Control: {t_mean:.2f} vs {c_mean:.2f} "
            f"(d={d_val:.3f}, p={p_tvc_display})")
        lines.append("")

    # ── Full comparison table ─────────────────────────────────
    lines.append("")
    lines.append("FULL CONFIGURATION COMPARISON WITH STATISTICS:")
    lines.append("─" * 80)

    display_df = comparison_df.copy()

    for col, fmt in [
        ('Particles_Per_Image',  '{:.1f}'),
        ('Particles_Std',        '±{:.1f}'),
        ('Mean_Fluorescence',    '{:.2f}'),
        ('Fluor_Std',            '±{:.2f}'),
        ('Fluor_SEM',            '±{:.2f}'),
        ('Signal_to_Background', '{:.1f}×'),
        ('Confidence_Score',     '{:.1f}'),
        ('Cohens_d',             '{:.3f}'),
    ]:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(
                lambda x, f=fmt: f.format(float(x)))  # type: ignore[arg-type]

    # Format P_Value_TvC separately to handle NaN
    if 'P_Value_TvC' in display_df.columns:
        display_df['P_Value_TvC'] = display_df['P_Value_TvC'].apply(
            lambda x: f"{float(x):.4f}" if not pd.isna(x) else "N/A"
        )

    display_cols = [
        'Rank', 'Bacteria_Type', 'Particles_Detected',
        'Particles_Per_Image', 'Particles_Std',
        'Mean_Fluorescence', 'Fluor_Std', 'Fluor_SEM',
        'Signal_to_Background', 'Cohens_d', 'P_Value_TvC',
        'Confidence_Score', 'Images_Processed', 'Low_Fluor_Flag',
    ]
    display_cols = [c for c in display_cols if c in display_df.columns]
    lines.append(display_df[display_cols].to_string(index=False))
    lines.append("")

    # ── Statistical significance analysis ─────────────────────
    if len(all_results) >= 2 and top2_key is not None:
        lines.append("")
        lines.append("STATISTICAL SIGNIFICANCE ANALYSIS:")
        lines.append("─" * 80)
        lines.append("")

        top1_name_s = str(config_names[top1_key])
        top2_name_s = str(config_names[top2_key])

        if not np.isnan(t_stat_top2):
            lines.append(
                f"Comparison: {top1_name_s} vs {top2_name_s}")
            lines.append("  Independent t-test (Welch's) on per-image fluorescence:")
            lines.append(f"    t-statistic: {t_stat_top2:.4f}")
            lines.append(f"    p-value: {p_value_top2:.6f}")

            if   p_value_top2 < 0.001:
                sig_label = "*** Highly significant (p < 0.001)"
            elif p_value_top2 < 0.01:
                sig_label = "** Very significant (p < 0.01)"
            elif p_value_top2 < 0.05:
                sig_label = "* Significant (p < 0.05)"
            else:
                sig_label = "Not significant (p ≥ 0.05)"

            lines.append(f"    Significance: {sig_label}")
            lines.append("")

            # Per-config test-vs-control breakdown
            lines.append("  Per-configuration test-vs-control statistics:")
            for _, row in comparison_df.iterrows():
                d   = float(row['Cohens_d'])
                p_v = row['P_Value_TvC']
                p_s = f"{float(p_v):.4f}" if not pd.isna(p_v) else "N/A"
                lf  = "⚠ low-fluor" if row.get('Low_Fluor_Flag', False) else ""
                lines.append(
                    f"    {str(row['Bacteria_Type']):<35}"
                    f"  d={d:+.3f}  p={p_s}  {lf}"
                )
            lines.append("")

            if p_value_top2 < 0.05:
                lines.append(
                    f"  ✓ Top configuration ({top1_name_s}) is "
                    f"statistically distinguishable from rank 2")
            else:
                lines.append(
                    "  ⚠ Top configurations are NOT statistically "
                    "different")
                lines.append(
                    "    → Convergence penalty applied to rank-1 "
                    "score")
                lines.append(
                    "    → Consider both as viable candidates")
        else:
            lines.append(
                "  Insufficient data for statistical testing")
        lines.append("")

    # ── Clinical recommendations ──────────────────────────────
    lines.append("")
    lines.append("CLINICAL RECOMMENDATIONS:")
    lines.append("─" * 80)

    if top_particles == 0:
        lines.append(
            "✓ No bacteria detected across all configurations")
        lines.append("  → Sample appears clean")
        lines.append("  → No immediate action required")
    elif stat_ambiguous:
        runner_r = str(comparison_df.iloc[1]['Bacteria_Type'])

        # Extra guidance when ambiguity is driven by low-fluor configs
        top1_low  = bool(comparison_df.iloc[0].get('Low_Fluor_Flag', False))
        top2_low  = bool(comparison_df.iloc[1].get('Low_Fluor_Flag', False))
        both_low  = top1_low and top2_low
        one_low   = top1_low or top2_low

        lines.append(
            f"⚠ AMBIGUOUS — {top_name} vs {runner_r} "
            f"not statistically separable")
        if both_low:
            lines.append(
                "  ⚠ Both top configs have low mean fluorescence "
                "(<1.0 a.u./µm²) — likely non-specific detections")
            lines.append(
                "  → Re-tune the bacteria configuration parameters "
                "or check segmentation debug images")
        elif one_low:
            lines.append(
                "  ⚠ One of the top configs has low mean fluorescence "
                "(<1.0 a.u./µm²) — it may be detecting background")
        lines.append(
            "  → Culture-based identification REQUIRED "
            "before targeted therapy")
        lines.append(
            "  → Consider empirical broad-spectrum coverage")
        lines.append(
            "  → Infectious disease consultation recommended")
    elif top_confidence >= 70:
        lines.append(f"✓ High confidence match: {top_name}")
        lines.append(
            "  → Proceed with targeted antimicrobial therapy")
        lines.append(
            "  → Consider culture confirmation if treatment fails")
    elif top_confidence >= 50:
        lines.append(f"⚠ Moderate confidence match: {top_name}")
        lines.append(
            "  → Recommend culture-based confirmation")
        lines.append(
            "  → Consider broad-spectrum coverage initially")
    else:
        lines.append("⚠ LOW CONFIDENCE - MANUAL REVIEW REQUIRED")
        lines.append(
            "  → Multiple configurations show similar results")
        lines.append(
            "  → STRONGLY recommend culture-based identification")
        lines.append(
            "  → Consider infectious disease consultation")

    lines.append("")
    lines.append("=" * 80)

    return comparison_df, "\n".join(lines)



# ==================================================
# generate_multi_config_comparison_plot
# ==================================================
def generate_multi_config_comparison_plot(
    comparison_df: pd.DataFrame,
    all_results: dict,
    output_dir: Path
) -> Optional[Path]:
    """Generate bar plot comparing configurations with error bars"""
    
    if comparison_df.empty:
        return None
    
    try:
        plt.figure(figsize=(14, 8))
        
        # Extract data
        config_keys: list[str] = comparison_df["Config_Key"].astype(str).to_list()
        bacteria_names: list[str] = comparison_df["Bacteria_Type"].astype(str).to_list()

        mean_fluor = comparison_df["Mean_Fluorescence"].to_numpy(dtype=float, na_value=0.0)
        confidence = comparison_df["Confidence_Percent"].to_numpy(dtype=float, na_value=0.0)
        
        # Get error bars (standard deviation)
        std_fluor = []
        for key in config_keys:
            std_val = all_results[key].get('std_fluorescence', 0.0)
            std_fluor.append(float(std_val))  # ✅ Ensure float
        
        std_fluor = np.array(
            [float(all_results[k].get("std_fluorescence", 0.0)) for k in config_keys],
            dtype=float,
            )
        
        # Create bar plot
        x = np.arange(len(bacteria_names), dtype=float)  # Ensure x is float array
        bars = plt.bar(x, mean_fluor,  # Already float now
                       width=0.6,
                       alpha=0.8, 
                       edgecolor='black',
                       linewidth=2)
        
        # Color bars by confidence
        colors = []
        for conf in confidence:
            if conf >= 70:
                colors.append('darkgreen')
            elif conf >= 50:
                colors.append('gold')
            else:
                colors.append('orangered')
        
        for bar, color in zip(bars, colors):
            bar.set_color(color)
        
        # Add error bars
        plt.errorbar(x, mean_fluor, yerr=std_fluor,  # Already float arrays
                    fmt='none',
                    ecolor='black',
                    elinewidth=2.5,
                    capsize=10,
                    capthick=2.5,
                    zorder=10)
        
        # Labels and formatting
        plt.xlabel('Bacteria Configuration', fontsize=14, fontweight='bold')
        plt.ylabel('Mean Fluorescence (a.u./µm²)', fontsize=14, fontweight='bold')
        plt.title('Multi-Configuration Scan Results\n(Error bars: ±1 Standard Deviation)', 
                 fontsize=16, fontweight='bold', pad=20)
        
        # ✅ FIXED: Convert bacteria_names to strings explicitly
        plt.xticks(x, [str(name) for name in bacteria_names], rotation=45, ha='right', fontsize=11)
        plt.yticks(fontsize=11)
        plt.grid(axis='y', alpha=0.3, linestyle='--', linewidth=1)
        
        # Add confidence scores as text labels
        max_y = float(np.max(mean_fluor + std_fluor)) if len(mean_fluor) > 0 else 1.0
        
        for i, (xi, yi, std_i, conf) in enumerate(zip(x, mean_fluor, std_fluor, confidence)):
            # Position text above error bar
            text_y = float(yi + std_i + max_y * 0.03)
            
            plt.text(float(xi), text_y,  # ✅ Convert to float
                    f'{float(conf):.0f}%',  # ✅ Convert conf to float
                    ha='center', va='bottom',
                    fontsize=11, fontweight='bold',
                    bbox=dict(boxstyle='round,pad=0.4', 
                             facecolor='white', 
                             edgecolor='black',
                             linewidth=1.5,
                             alpha=0.9))
        
        # Add legend
        from matplotlib.patches import Patch
        legend_elements = [
            Patch(facecolor='darkgreen', edgecolor='black', label='High Confidence (≥70%)'),
            Patch(facecolor='gold', edgecolor='black', label='Moderate Confidence (50-69%)'),
            Patch(facecolor='orangered', edgecolor='black', label='Low Confidence (<50%)')
        ]
        plt.legend(handles=legend_elements, loc='upper right', fontsize=10, framealpha=0.95)
        
        plt.tight_layout()
        
        # Save
        plot_path = output_dir / "multi_config_comparison_with_statistics.png"
        plt.savefig(plot_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        print(f"  ✓ Comparison plot with error bars: {plot_path.name}")
        return plot_path
        
    except Exception as e:
        print(f"  ✗ Failed to generate comparison plot: {e}")
        import traceback
        traceback.print_exc()
        plt.close()
        return None

def load_bacteria_config_from_json(bacteria_key: str) -> Optional['SegmentationConfig']:
    """Load bacteria configuration directly from JSON file with robust error handling
    
    Args:
        bacteria_key: Configuration key (filename without .json)
        
    Returns:
        SegmentationConfig object or None if loading fails
    """
    from bacteria_configs import SegmentationConfig
    
    config_file = Path("bacteria_configs") / f"{bacteria_key}.json"
    
    if not config_file.exists():
        print(f"[WARN] Config file not found: {config_file}")
        return None
    
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # Extract config data (support both wrapped and direct formats)
        if "config" in json_data:
            config_data = json_data["config"]
        else:
            config_data = json_data
        
        # ✅ Helper function to safely convert to float
        def safe_float(value, default):
            """Convert to float, handling None and invalid values"""
            if value is None:
                return default
            try:
                return float(value)
            except (TypeError, ValueError):
                return default
        
        # ✅ Helper function to safely convert to int
        def safe_int(value, default):
            """Convert to int, handling None and invalid values"""
            if value is None:
                return default
            try:
                return int(value)
            except (TypeError, ValueError):
                return default
        
        # ✅ Helper function to safely convert to bool
        def safe_bool(value, default):
            """Convert to bool, handling None"""
            if value is None:
                return default
            return bool(value)
        
        # Create config with safe conversions
        config = SegmentationConfig(
            name=config_data.get('name', 'Unknown'),
            description=config_data.get('description', ''),
            
            # Core parameters with safe defaults
            gaussian_sigma=safe_float(config_data.get('gaussian_sigma'), 15.0),
            min_area_um2=safe_float(config_data.get('min_area_um2'), 3.0),
            max_area_um2=safe_float(config_data.get('max_area_um2'), 2000.0),
            
            # Morphological parameters
            dilate_iterations=safe_int(config_data.get('dilate_iterations'), 0),
            erode_iterations=safe_int(config_data.get('erode_iterations'), 0),
            morph_kernel_size=safe_int(config_data.get('morph_kernel_size'), 3),
            morph_iterations=safe_int(config_data.get('morph_iterations'), 1),
            
            # Shape filters
            min_circularity=safe_float(config_data.get('min_circularity'), 0.0),
            max_circularity=safe_float(config_data.get('max_circularity'), 1.0),
            min_aspect_ratio=safe_float(config_data.get('min_aspect_ratio'), 0.2),
            max_aspect_ratio=safe_float(config_data.get('max_aspect_ratio'), 10.0),
            
            # Intensity filters
            min_mean_intensity=safe_int(config_data.get('min_mean_intensity'), 0),
            max_mean_intensity=safe_int(config_data.get('max_mean_intensity'), 255),
            max_edge_gradient=safe_int(config_data.get('max_edge_gradient'), 200),
            
            # Advanced filters
            min_solidity=safe_float(config_data.get('min_solidity'), 0.3),
            max_fraction_of_image=safe_float(config_data.get('max_fraction_of_image'), 0.25),
            
            # Fluorescence parameters
            fluor_min_area_um2=safe_float(config_data.get('fluor_min_area_um2'), 3.0),
            fluor_max_area_um2=safe_float(config_data.get('fluor_max_area_um2'), 2000.0),
            fluor_match_min_intersection_px=safe_float(config_data.get('fluor_match_min_intersection_px'), 5.0),
            
            # Image processing flags
            invert_image=safe_bool(config_data.get('invert_image'), False),
            
            # Intensity threshold parameters
            use_intensity_threshold=safe_bool(config_data.get('use_intensity_threshold'), False),
            intensity_threshold=safe_float(config_data.get('intensity_threshold'), 80.0),
            
            # Metadata
            pixel_size_um=safe_float(config_data.get('pixel_size_um'), 0.109492),
            last_modified=config_data.get('last_modified'),
            tuned_by=config_data.get('tuned_by')
        )
        
        print(f"✅ Loaded config from JSON: {config_file.name}")
        print(f"   Name: {config.name}")
        print(f"   Gaussian σ: {config.gaussian_sigma:.2f}")
        
        if config.use_intensity_threshold:
            print(f"   Intensity threshold: {config.intensity_threshold:.1f}")
        else:
            print(f"   Intensity threshold: Disabled")
        
        print(f"   Area range: {config.min_area_um2:.1f} - {config.max_area_um2:.1f} µm²")
        print(f"   Invert image: {'ON' if config.invert_image else 'OFF'}")
        
        return config
        
    except json.JSONDecodeError as e:
        print(f"[ERROR] Invalid JSON in {config_file}: {e}")
        import traceback
        traceback.print_exc()
        return None
    except Exception as e:
        print(f"[ERROR] Failed to load config from {config_file}: {e}")
        import traceback
        traceback.print_exc()
        return None


# ==================================================
# Unicode-Safe File I/O Functions
# ==================================================

def safe_imread(path: Path, flags: int = cv2.IMREAD_UNCHANGED) -> Optional[np.ndarray]:
    """Read image with Unicode path support on Windows
    
    Args:
        path: Path to image file
        flags: OpenCV imread flags
        
    Returns:
        numpy array of image, or None if failed
    """
    try:
        # Method: Read file as bytes, then decode with OpenCV
        with open(path, 'rb') as f:
            file_bytes = np.asarray(bytearray(f.read()), dtype=np.uint8)
        img = cv2.imdecode(file_bytes, flags)
        
        if img is None:
            print(f"[WARN] cv2.imdecode returned None for {path.name}")
            return None
            
        return img
    except Exception as e:
        print(f"[ERROR] Failed to read image {path.name}: {e}")
        return None


def safe_imwrite(path: Path, img: np.ndarray, params: Optional[list] = None) -> bool:
    """Write image with Unicode path support on Windows
    
    Args:
        path: Path where to save image
        img: Image array to save
        params: Optional OpenCV imwrite parameters
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Get file extension
        ext = path.suffix.lower()
        if not ext:
            ext = '.png'
        
        # Encode image to memory buffer
        if params is None:
            is_success, buffer = cv2.imencode(ext, img)
        else:
            is_success, buffer = cv2.imencode(ext, img, params)
        
        if not is_success:
            print(f"[WARN] cv2.imencode failed for {path.name}")
            return False
        
        # Write buffer to file
        with open(path, 'wb') as f:
            f.write(buffer.tobytes())
        
        return True
    except Exception as e:
        print(f"[ERROR] Failed to write image {path.name}: {e}")
        return False


def safe_xml_parse(xml_path: Path) -> Optional[ET.ElementTree]:
    """Parse XML file with Unicode path support
    
    Args:
        xml_path: Path to XML file
        
    Returns:
        ElementTree object, or None if failed
    """
    try:
        # Read file content first with explicit encoding
        with open(xml_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Parse from string
        root = ET.fromstring(content)
        tree = ET.ElementTree(root)
        
        return tree
    except FileNotFoundError:
        print(f"[WARN] XML file not found: {xml_path.name}")
        return None
    except ET.ParseError as e:
        print(f"[ERROR] XML parse error in {xml_path.name}: {e}")
        return None
    except Exception as e:
        print(f"[ERROR] Failed to read XML {xml_path.name}: {e}")
        return None


def validate_path_encoding(path: Path) -> bool:
    """Check if path can be properly encoded for filesystem
    
    Args:
        path: Path to validate
        
    Returns:
        True if path encoding is valid, False otherwise
    """
    try:
        path_str = str(path.resolve())
        # Try encoding to filesystem encoding
        path_str.encode(sys.getfilesystemencoding())
        return True
    except UnicodeEncodeError as e:
        print(f"[WARN] Path encoding issue: {path}")
        print(f"       {e}")
        return False
    except Exception as e:
        print(f"[WARN] Path validation error: {e}")
        return False
# ==================================================


# Basic logger setup
logger = logging.getLogger("particle_scout")
if not logger.handlers:
    logger.setLevel(logging.INFO)
    _sh = logging.StreamHandler()
    _sh.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
    logger.addHandler(_sh)


# ==================================================
# Logging: tee stdout/stderr to a file
# ==================================================
class Tee:
    def __init__(self, *streams):
        self.streams = streams

    def write(self, data: str) -> None:
        for s in self.streams:
            s.write(data)
            s.flush()

    def flush(self) -> None:
        for s in self.streams:
            s.flush()

def get_project_root() -> Path:
    """Get project root (works as script and as .exe)"""
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).resolve().parent
    else:
        return Path(__file__).resolve().parent

PROJECT_ROOT = get_project_root()
OUTPUTS_DIR = PROJECT_ROOT / "outputs"
OUTPUTS_DIR.mkdir(exist_ok=True)
_project_root = get_project_root()
_logs_dir = _project_root / "logs"
_logs_dir.mkdir(exist_ok=True)
_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
_script_name = Path(sys.argv[0]).stem
_log_path = _logs_dir / f"run_{_timestamp}_{_script_name}.txt"

# Initialize _log_file with type annotation
_log_file: Optional[Any] = None

try:
    _log_file = open(_log_path, "w", encoding="utf-8")
    sys.stdout = Tee(sys.stdout, _log_file)
    sys.stderr = Tee(sys.stderr, _log_file)

except Exception as e:
    print(f"Warning: Could not set up logging: {e}")


print(f"Saving output to: {_log_path}")
print(f"Project root: {_project_root.resolve()}")
print(f"Running as: {'EXECUTABLE' if getattr(sys, 'frozen', False) else 'SCRIPT'}")

@atexit.register
def _close_log_file() -> None:
    """Close the global log file if it exists."""
    global _log_file  # Now this is OK since we initialized it explicitly
    if _log_file is not None:
        try:
            _log_file.close()
        except Exception:
            pass
        finally:
            _log_file = None

# ==================================================
# Configuration
# ==================================================
SOURCE_DIR = Path("./source")
CONTROL_DIR = None  # Will be set dynamically

# Segment only brightfield channel
IMAGE_GLOB = "*_ch00.tif"

# OUTPUT_DIR will be set dynamically in main()
OUTPUT_DIR: Optional[Path] = None


# Scale bar parameters
SCALE_BAR_LENGTH_UM = 10
SCALE_BAR_HEIGHT = 4
SCALE_BAR_MARGIN = 15
SCALE_BAR_COLOR = (255, 255, 255)
SCALE_BAR_BG_COLOR = (0, 0, 0)
SCALE_BAR_TEXT_COLOR = (255, 255, 255)
SCALE_BAR_FONT_SCALE = 0.5
SCALE_BAR_FONT_THICKNESS = 1

# --- Segmentation ---
GAUSSIAN_SIGMA = 15
MORPH_KERNEL_SIZE = 3
MORPH_ITERATIONS = 1
DILATE_ITERATIONS = 1
ERODE_ITERATIONS = 1

# Filtering (in micrometers)
MIN_AREA_UM2 = 3.0
MAX_AREA_UM2 = 2000.0
MIN_CIRCULARITY = 0.0
MAX_FRACTION_OF_IMAGE_AREA = 0.25

# --- Fluorescence segmentation (S2) ---
FLUOR_GAUSSIAN_SIGMA = 1.5
FLUOR_MORPH_KERNEL_SIZE = 3
FLUOR_MIN_AREA_UM2 = 3.0
FLUOR_MATCH_MIN_INTERSECTION_PX = 5.0

# Debug options
CLEAR_OUTPUT_DIR_EACH_RUN = True
SEPARATE_OUTPUT_BY_GROUP = True
FALLBACK_UM_PER_PX: Optional[float] = 0.109492


# ==================================================
# Helper Functions
# ==================================================
def logged_input(prompt: str) -> str:
    """Input function that logs both prompt and user response"""
    print(prompt, end='', flush=True)
    user_input = input()
    
    if user_input.strip():
        print(user_input)
    else:
        print("(pressed Enter)")
    
    return user_input


def add_scale_bar(
    img: np.ndarray, pixel_size: float, unit: str = "um", length_um: float = 10
) -> np.ndarray:
    """Add a scale bar to the image"""
    if pixel_size is None or pixel_size <= 0:
        return img

    bar_length_px = int(round(length_um / pixel_size))
    if bar_length_px < 10:
        return img

    h, w = img.shape[:2]
    bar_x = w - bar_length_px - SCALE_BAR_MARGIN
    bar_y = h - SCALE_BAR_HEIGHT - SCALE_BAR_MARGIN

    if bar_x < 0 or bar_y < 0:
        return img

    label = (
        f"{int(length_um)} um"
        if unit in ["µm", "um"]
        else f"{int(length_um)} {unit}"
    )

    font = cv2.FONT_HERSHEY_SIMPLEX
    (text_w, text_h), baseline = cv2.getTextSize(
        label, font, SCALE_BAR_FONT_SCALE, SCALE_BAR_FONT_THICKNESS
    )

    text_x = bar_x + (bar_length_px - text_w) // 2
    text_y = bar_y - 8

    bg_padding = 5
    bg_x1 = min(bar_x, text_x) - bg_padding
    bg_y1 = text_y - text_h - bg_padding
    bg_x2 = max(bar_x + bar_length_px, text_x + text_w) + bg_padding
    bg_y2 = bar_y + SCALE_BAR_HEIGHT + bg_padding

    img = img.copy()
    overlay = img.copy()
    cv2.rectangle(overlay, (bg_x1, bg_y1), (bg_x2, bg_y2), SCALE_BAR_BG_COLOR, -1)
    cv2.addWeighted(overlay, 0.7, img, 0.3, 0, img)

    cv2.rectangle(
        img,
        (bar_x, bar_y),
        (bar_x + bar_length_px, bar_y + SCALE_BAR_HEIGHT),
        SCALE_BAR_COLOR,
        -1,
    )

    cv2.putText(
        img,
        label,
        (text_x, text_y),
        font,
        SCALE_BAR_FONT_SCALE,
        SCALE_BAR_TEXT_COLOR,
        SCALE_BAR_FONT_THICKNESS,
        cv2.LINE_AA,
    )

    return img


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def save_debug(
    folder: Path,
    name: str,
    img: np.ndarray,
    pixel_size_um: Optional[float] = None,
) -> None:
    """Save debug image with optional scale bar - memory optimized, Unicode-safe
    
    Args:
        folder: Output folder
        name: Image filename
        img: Image array to save
        pixel_size_um: Optional pixel size for scale bar
    """
    out = folder / name
    
    if pixel_size_um is not None and pixel_size_um > 0:
        img_to_save = add_scale_bar(
            img.copy(),
            float(pixel_size_um), "um", SCALE_BAR_LENGTH_UM
        )
    else:
        img_to_save = img
    
    # Use Unicode-safe write
    success = safe_imwrite(out, img_to_save)
    
    if not success:
        print(f"[ERROR] Failed to save debug image: {name}")
    
    # Clean up large images
    if img_to_save.nbytes > 10_000_000:
        del img_to_save


def _display_group_name(name: str) -> str:
    return "Control" if name.lower().startswith("control") else name


def _group_order_key(g: str) -> tuple[int, int]:
    """Sort numeric groups ascending, Control last."""
    if g == "Control":
        return (1, 10**9)
    if g.isdigit():
        return (0, int(g))
    return (0, 10**8)


def find_metadata_paths(img_path: Path) -> tuple[Optional[Path], Optional[Path]]:
    base = img_path.stem
    if base.endswith("_ch00"):
        base = base[:-5]
    md_dir = img_path.parent / "MetaData"
    xml_main = md_dir / f"{base}.xml"
    xml_props = md_dir / f"{base}_Properties.xml"

    return (
        xml_props if xml_props.exists() else None,
        xml_main if xml_main.exists() else None,
    )


def _require_attr(elem: ET.Element, attr: str, context: str) -> str:
    v = elem.get(attr)
    if v is None:
        raise ValueError(f"Missing attribute '{attr}' in {context}")
    return v


def _parse_float(s: str) -> float:
    return float(s.strip().replace(",", "."))


def get_pixel_size_um(
    xml_props_path: Optional[Path],
    xml_main_path: Optional[Path],
) -> Tuple[float, float]:
    """Extract pixel size with detailed error reporting and Unicode support
    
    Args:
        xml_props_path: Path to Properties XML file
        xml_main_path: Path to main XML file
        
    Returns:
        Tuple of (pixel_size_x, pixel_size_y) in micrometers
        
    Raises:
        ValueError: If pixel size cannot be determined
    """
    
    errors = []
    
    # Try Properties XML first
    if xml_props_path is not None:
        try:
            tree = safe_xml_parse(xml_props_path)
            if tree is None:
                raise ValueError(f"Could not parse {xml_props_path.name}")
            
            root = tree.getroot()
            if root is None:
                raise ValueError(f"Empty XML document: {xml_props_path.name}")

            dims = root.findall(".//ImageDescription/Dimensions/DimensionDescription")
            by_id = {d.get("DimID"): d for d in dims}

            def read_dim(dim_id: str) -> Tuple[float, int, str]:
                d = by_id.get(dim_id)
                if d is None:
                    raise ValueError(
                        f"Missing DimensionDescription with DimID='{dim_id}' in {xml_props_path.name}"
                    )

                length_s = _require_attr(
                    d, "Length", f"{xml_props_path.name} DimID={dim_id}"
                )
                n_s = _require_attr(
                    d, "NumberOfElements", f"{xml_props_path.name} DimID={dim_id}"
                )
                unit = _require_attr(d, "Unit", f"{xml_props_path.name} DimID={dim_id}")

                length = _parse_float(length_s)
                n = int(n_s)
                return length, n, unit

            x_len, x_n, x_unit = read_dim("X")
            y_len, y_n, y_unit = read_dim("Y")

            if x_unit != "µm" or y_unit != "µm":
                raise ValueError(
                    f"Unexpected units in {xml_props_path.name}: X={x_unit}, Y={y_unit}"
                )

            return float(x_len / x_n), float(y_len / y_n)
        except Exception as e:
            errors.append(f"Properties XML ({xml_props_path.name}): {e}")

    # Try main XML
    if xml_main_path is not None:
        try:
            tree = safe_xml_parse(xml_main_path)
            if tree is None:
                raise ValueError(f"Could not parse {xml_main_path.name}")
            
            root = tree.getroot()
            if root is None:
                raise ValueError(f"Empty XML document: {xml_main_path.name}")

            dims = root.findall(".//ImageDescription/Dimensions/DimensionDescription")
            by_id = {d.get("DimID"): d for d in dims}

            def read_dim(dim_id: str) -> Tuple[float, int, str]:
                d = by_id.get(dim_id)
                if d is None:
                    raise ValueError(
                        f"Missing DimensionDescription with DimID='{dim_id}' in {xml_main_path.name}"
                    )

                length_s = _require_attr(
                    d, "Length", f"{xml_main_path.name} DimID={dim_id}"
                )
                n_s = _require_attr(
                    d, "NumberOfElements", f"{xml_main_path.name} DimID={dim_id}"
                )
                unit = _require_attr(d, "Unit", f"{xml_main_path.name} DimID={dim_id}")

                length = _parse_float(length_s)
                n = int(n_s)
                return length, n, unit

            x_len_m, x_n, x_unit = read_dim("1")
            y_len_m, y_n, y_unit = read_dim("2")

            if x_unit != "m" or y_unit != "m":
                raise ValueError(
                    f"Unexpected units in {xml_main_path.name}: X={x_unit}, Y={y_unit}"
                )

            return float((x_len_m * 1e6) / x_n), float((y_len_m * 1e6) / y_n)
        except Exception as e:
            errors.append(f"Main XML ({xml_main_path.name}): {e}")
    
    error_summary = "\n  - ".join(errors) if errors else "No XML files provided"
    raise ValueError(
        f"Could not determine pixel size (µm/px).\nAttempted sources:\n  - {error_summary}"
    )


def contour_perimeter_um(contour: np.ndarray, um_per_px_x: float, um_per_px_y: float) -> float:
    pts = contour.reshape(-1, 2).astype(np.float64)
    pts[:, 0] *= float(um_per_px_x)
    pts[:, 1] *= float(um_per_px_y)
    d = np.diff(np.vstack([pts, pts[0]]), axis=0)
    seg = np.sqrt((d[:, 0] ** 2) + (d[:, 1] ** 2))
    return float(seg.sum())


def equivalent_diameter_from_area(area: float) -> float:
    return float(2.0 * np.sqrt(area / np.pi)) if area > 0 else 0.0


def normalize_to_8bit(img: np.ndarray) -> np.ndarray:
    if img.dtype == np.uint8:
        return img
    if img.dtype == np.uint16:
        out = np.zeros_like(img, dtype=np.uint8)
        cv2.normalize(img, out, 0, 255, cv2.NORM_MINMAX)
        return out
    img_f = img.astype(np.float32)
    mn, mx = float(np.min(img_f)), float(np.max(img_f))
    if mx <= mn:
        return np.zeros(img.shape, dtype=np.uint8)
    return ((img_f - mn) * (255.0 / (mx - mn))).clip(0, 255).astype(np.uint8)


def _put_text_outline(
    img: np.ndarray,
    text: str,
    org: tuple[int, int],
    font_scale: float = 0.5,
    color: tuple[int, int, int] = (255, 255, 255),
    thickness: int = 1,
) -> None:
    """Draw readable text with a black outline."""
    cv2.putText(
        img,
        text,
        org,
        cv2.FONT_HERSHEY_SIMPLEX,
        font_scale,
        (0, 0, 0),
        thickness + 2,
        cv2.LINE_AA,
    )
    cv2.putText(
        img,
        text,
        org,
        cv2.FONT_HERSHEY_SIMPLEX,
        font_scale,
        color,
        thickness,
        cv2.LINE_AA,
    )


def draw_object_ids(
    img_bgr: np.ndarray, contours: list[np.ndarray], labels: Optional[list[str]] = None
) -> np.ndarray:
    """Draw object labels at contour centroids."""
    out = img_bgr.copy()
    for i, c in enumerate(contours, 1):
        M = cv2.moments(c)
        if M["m00"] == 0:
            continue
        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])
        text = labels[i - 1] if (labels is not None and i - 1 < len(labels)) else str(i)
        _put_text_outline(out, text, (cx, cy), font_scale=0.5, color=(0, 255, 0), thickness=1)
    return out


def _ids_name(original_png: str) -> str:
    """Insert '_ids' before '.png'."""
    if original_png.lower().endswith(".png"):
        return original_png[:-4] + "_ids.png"
    return original_png + "_ids"


def save_debug_ids(
    folder: Path,
    original_name: str,
    img_bgr: np.ndarray,
    accepted_contours: list[np.ndarray],
    object_ids: list[str],
    pixel_size_um: Optional[float] = None,
) -> None:
    """Save a labeled (Object_ID) version of an existing debug view."""
    labeled = draw_object_ids(img_bgr, accepted_contours, labels=object_ids)
    save_debug(folder, _ids_name(original_name), labeled, pixel_size_um)







# ==================================================
# Fluorescence Registration / Alignment
# ==================================================

def align_fluorescence_channel(bf_img: np.ndarray, fluor_img: np.ndarray) -> tuple[np.ndarray, tuple[float, float], dict]:
    """Aligns fluorescence to brightfield using phase correlation with validation
    
    Returns:
        (aligned_image, (shift_y, shift_x), diagnostics)
    """
    
    # Convert to grayscale
    if bf_img.ndim == 3:
        bf_gray = cv2.cvtColor(bf_img, cv2.COLOR_BGR2GRAY)
    else:
        bf_gray = bf_img.copy()

    if fluor_img.ndim == 3:
        fluor_gray = cv2.cvtColor(fluor_img, cv2.COLOR_BGR2GRAY)
    else:
        fluor_gray = fluor_img.copy()

    # Normalize for better correlation
    bf_norm = np.zeros_like(bf_gray, dtype=np.uint8)
    cv2.normalize(bf_gray, bf_norm, 0, 255, cv2.NORM_MINMAX)
    
    fluor_norm = np.zeros_like(fluor_gray, dtype=np.uint8)
    cv2.normalize(fluor_gray, fluor_norm, 0, 255, cv2.NORM_MINMAX)
    
    # Prepare for correlation
    bf_for_corr = bf_norm.astype(np.float32) / 255.0
    fluor_for_corr = fluor_norm.astype(np.float32) / 255.0
    
    # Test 1: Direct correlation
    try:
        shift1, error1, _ = phase_cross_correlation(
            bf_for_corr, 
            fluor_for_corr, 
            upsample_factor=10
        )
    except Exception as e:
        print(f"  ⚠ Direct correlation failed: {e}")
        shift1, error1 = (0.0, 0.0), 1.0
    
    # Test 2: Inverted BF correlation
    bf_inverted = 255 - bf_norm
    bf_inv_for_corr = bf_inverted.astype(np.float32) / 255.0
    
    try:
        shift2, error2, _ = phase_cross_correlation(
            bf_inv_for_corr, 
            fluor_for_corr, 
            upsample_factor=10
        )
    except Exception as e:
        print(f"  ⚠ Inverted correlation failed: {e}")
        shift2, error2 = (0.0, 0.0), 1.0
    
    shift_y1, shift_x1 = shift1
    shift_y2, shift_x2 = shift2
    
    print(f"  Alignment test:")
    print(f"    Direct BF:   shift=({shift_x1:.2f}, {shift_y1:.2f}), error={error1:.6f}")
    print(f"    Inverted BF: shift=({shift_x2:.2f}, {shift_y2:.2f}), error={error2:.6f}")
    
    # ========== DECISION LOGIC ==========
    # If both errors are very high (> 0.5), correlation failed
    CORRELATION_FAILURE_THRESHOLD = 0.5
    
    if error1 > CORRELATION_FAILURE_THRESHOLD and error2 > CORRELATION_FAILURE_THRESHOLD:
        print(f"  ⚠️ Correlation failed (both errors > {CORRELATION_FAILURE_THRESHOLD})")
        print(f"     Images may already be aligned or have no overlap")
        print(f"     Using NO alignment (original fluorescence)")
        
        # Create diagnostics with no shift
        diagnostics = create_alignment_diagnostics(
            bf_norm, fluor_img, fluor_img, fluor_img,
            (0.0, 0.0), (0.0, 0.0), error1, error2
        )
        
        return fluor_img.copy(), (0.0, 0.0), diagnostics
    
    # If shifts are unreasonably large (> 50px), suspect failure
    MAX_REASONABLE_SHIFT_PX = 50
    shift1_magnitude = np.sqrt(shift_x1**2 + shift_y1**2)
    shift2_magnitude = np.sqrt(shift_x2**2 + shift_y2**2)
    
    if shift1_magnitude > MAX_REASONABLE_SHIFT_PX and shift2_magnitude > MAX_REASONABLE_SHIFT_PX:
        print(f"  ⚠️ Both shifts unreasonably large (>{MAX_REASONABLE_SHIFT_PX}px)")
        print(f"     Using NO alignment")
        
        diagnostics = create_alignment_diagnostics(
            bf_norm, fluor_img, fluor_img, fluor_img,
            (0.0, 0.0), (0.0, 0.0), error1, error2
        )
        
        return fluor_img.copy(), (0.0, 0.0), diagnostics
    
    # Choose the better correlation
    if error1 < error2:
        chosen_shift = shift1
        chosen_error = error1
        method = "Direct BF"
        shift_y, shift_x = shift_y1, shift_x1
    else:
        chosen_shift = shift2
        chosen_error = error2
        method = "Inverted BF"
        shift_y, shift_x = shift_y2, shift_x2
    
    print(f"    Using: {method}, shift=({shift_x:.2f}, {shift_y:.2f})px, error={chosen_error:.6f}")
    
    # Apply shift if it's reasonable
    if np.sqrt(shift_x**2 + shift_y**2) > MAX_REASONABLE_SHIFT_PX:
        print(f"  ⚠️ Chosen shift too large (>{MAX_REASONABLE_SHIFT_PX}px) - using NO alignment")
        
        diagnostics = create_alignment_diagnostics(
            bf_norm, fluor_img, fluor_img, fluor_img,
            (0.0, 0.0), (0.0, 0.0), error1, error2
        )
        
        return fluor_img.copy(), (0.0, 0.0), diagnostics
    
    # Apply the shift
    rows, cols = fluor_img.shape[:2]
    
    # Create shifted versions
    no_shift = fluor_img.copy()
    
    M_pos = np.array([[1, 0, shift_x], [0, 1, shift_y]], dtype=np.float32)
    aligned_pos = cv2.warpAffine(
        fluor_img, M_pos, (cols, rows),
        flags=cv2.INTER_LINEAR,
        borderMode=cv2.BORDER_CONSTANT,
        borderValue=0
    )
    
    M_neg = np.array([[1, 0, -shift_x], [0, 1, -shift_y]], dtype=np.float32)
    aligned_neg = cv2.warpAffine(
        fluor_img, M_neg, (cols, rows),
        flags=cv2.INTER_LINEAR,
        borderMode=cv2.BORDER_CONSTANT,
        borderValue=0
    )
    
    # Create diagnostics
    diagnostics = create_alignment_diagnostics(
        bf_norm, no_shift, aligned_pos, aligned_neg,
        (shift_y, shift_x), (-shift_y, -shift_x),
        error1, error2
    )
    
    # Return the version that matches the chosen shift direction
    if method == "Direct BF":
        return aligned_pos, (shift_y, shift_x), diagnostics
    else:
        return aligned_neg, (-shift_y, -shift_x), diagnostics


def create_alignment_diagnostics(
    bf_norm: np.ndarray,
    fluor_none: np.ndarray,
    fluor_pos: np.ndarray,
    fluor_neg: np.ndarray,
    shift_pos: tuple[float, float],
    shift_neg: tuple[float, float],
    error_direct: float,
    error_inverted: float
) -> dict:
    """Create diagnostic overlay images for alignment verification"""
    
    def make_overlay(fluor_version: np.ndarray) -> np.ndarray:
        overlay = np.zeros((bf_norm.shape[0], bf_norm.shape[1], 3), dtype=np.uint8)
        overlay[:, :, 1] = bf_norm  # Green = BF
        
        if fluor_version.ndim == 3:
            fluor_disp = cv2.cvtColor(fluor_version, cv2.COLOR_BGR2GRAY)
        else:
            fluor_disp = fluor_version.copy()
        
        fluor_disp_norm = np.zeros_like(fluor_disp, dtype=np.uint8)
        cv2.normalize(fluor_disp, fluor_disp_norm, 0, 255, cv2.NORM_MINMAX)
        overlay[:, :, 0] = fluor_disp_norm  # Red = Fluorescence
        
        return overlay
    
    return {
        'overlay_none': make_overlay(fluor_none),
        'overlay_pos': make_overlay(fluor_pos),
        'overlay_neg': make_overlay(fluor_neg),
        'shift_pos': shift_pos,
        'shift_neg': shift_neg,
        'error_direct': error_direct,
        'error_inverted': error_inverted,
    }


# ==================================================
# Fluorescence segmentation + matching
# ==================================================
def segment_fluorescence_global(
    fluor_img8: np.ndarray,
    bacteria_config: 'SegmentationConfig'  # NEW parameter
) -> np.ndarray:
    """Segment fluorescence objects globally with bacteria-specific parameters"""
    
    # Use config parameter
    blur = cv2.GaussianBlur(
        fluor_img8, (0, 0), 
        sigmaX=bacteria_config.gaussian_sigma * 0.1,  # Scale for fluorescence
        sigmaY=bacteria_config.gaussian_sigma * 0.1
    )
    
    otsu_threshold = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[0]
    
    THRESHOLD_MULTIPLIER = 0.5
    adjusted_threshold = otsu_threshold * THRESHOLD_MULTIPLIER
    
    print(f"  Fluorescence threshold: Otsu={otsu_threshold:.1f}, Adjusted={adjusted_threshold:.1f}")
    
    _, bw = cv2.threshold(blur, adjusted_threshold, 255, cv2.THRESH_BINARY)
    
    # Use config kernel size
    k = np.ones(
        (bacteria_config.morph_kernel_size, bacteria_config.morph_kernel_size), 
        np.uint8
    )
    bw = cv2.morphologyEx(bw, cv2.MORPH_OPEN, k, iterations=1)
    bw = cv2.morphologyEx(bw, cv2.MORPH_CLOSE, k, iterations=1)
    return bw



def contour_intersection_area_px(c1: np.ndarray, c2: np.ndarray, shape_hw: tuple[int, int]) -> float:
    """Intersection area in pixels between two contours."""
    H, W = shape_hw
    m1 = np.zeros((H, W), dtype=np.uint8)
    m2 = np.zeros((H, W), dtype=np.uint8)
    cv2.drawContours(m1, [c1], -1, 255, thickness=-1)
    cv2.drawContours(m2, [c2], -1, 255, thickness=-1)
    return float(np.count_nonzero(cv2.bitwise_and(m1, m2)))


def match_fluor_to_bf_by_overlap(
    bf_contours: list[np.ndarray],
    fluor_contours: list[np.ndarray],
    img_shape_hw: tuple[int, int],
    min_intersection_px: float = FLUOR_MATCH_MIN_INTERSECTION_PX,
) -> list[Optional[int]]:
    """For each BF contour, pick the fluorescence contour that maximizes overlap."""
    matches: list[Optional[int]] = []
    fluor_boxes = [cv2.boundingRect(c) for c in fluor_contours]

    for bf in bf_contours:
        bx, by, bw, bh = cv2.boundingRect(bf)

        best_idx: Optional[int] = None
        best_inter = 0.0

        for j, (fx, fy, fw, fh) in enumerate(fluor_boxes):
            if (bx + bw < fx) or (fx + fw < bx) or (by + bh < fy) or (fy + fh < by):
                continue

            inter = contour_intersection_area_px(bf, fluor_contours[j], img_shape_hw)
            if inter > best_inter:
                best_inter = inter
                best_idx = j

        if best_idx is not None and best_inter >= float(min_intersection_px):
            matches.append(best_idx)
        else:
            matches.append(None)

    return matches


def measure_fluorescence_intensity_with_global_area(
    fluor_img: np.ndarray,
    bf_contours: list[np.ndarray],
    fluor_contours: list[np.ndarray],
    bf_to_fluor_match: list[Optional[int]],
    um_per_px_x: float,
    um_per_px_y: float,
) -> list[dict]:
    """Intensity stats: within BF contour. Fluor area: from matched fluorescence contour."""
    um2_per_px2 = float(um_per_px_x) * float(um_per_px_y)
    measurements: list[dict] = []

    for i, bf in enumerate(bf_contours, 1):
        bf_mask = np.zeros(fluor_img.shape[:2], dtype=np.uint8)
        cv2.drawContours(bf_mask, [bf], -1, 255, thickness=-1)
        fluor_values = fluor_img[bf_mask > 0]

        j = bf_to_fluor_match[i - 1]
        if j is not None:
            s2_area_px = float(cv2.contourArea(fluor_contours[j]))
        else:
            s2_area_px = 0.0
        s2_area_um2 = s2_area_px * um2_per_px2

        if fluor_values.size > 0:
            measurements.append(
                {
                    "object_id": i,
                    "fluor_area_px": s2_area_px,
                    "fluor_area_um2": s2_area_um2,
                    "fluor_mean": float(np.mean(fluor_values)),
                    "fluor_median": float(np.median(fluor_values)),
                    "fluor_std": float(np.std(fluor_values)),
                    "fluor_min": float(np.min(fluor_values)),
                    "fluor_max": float(np.max(fluor_values)),
                    "fluor_integrated_density": float(np.sum(fluor_values)),
                }
            )
        else:
            measurements.append(
                {
                    "object_id": i,
                    "fluor_area_px": s2_area_px,
                    "fluor_area_um2": s2_area_um2,
                    "fluor_mean": 0.0,
                    "fluor_median": 0.0,
                    "fluor_std": 0.0,
                    "fluor_min": 0.0,
                    "fluor_max": 0.0,
                    "fluor_integrated_density": 0.0,
                }
            )

    return measurements


def visualize_fluorescence_measurements(
    fluor_img8: np.ndarray, contours: list[np.ndarray], measurements: list[dict]
) -> np.ndarray:
    """Create visualization with contours and intensity labels"""
    vis = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(vis, contours, -1, (0, 255, 0), 1)

    for m in measurements:
        c = contours[m["object_id"] - 1]
        M = cv2.moments(c)
        if M["m00"] != 0:
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])

            label = f"{m['object_id']}: {m['fluor_mean']:.0f}"
            cv2.putText(
                vis,
                label,
                (cx, cy),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.4,
                (255, 255, 0),
                1,
                cv2.LINE_AA,
            )

    return vis


# ==================================================
# Segmentation
# ==================================================


def segment_particles_brightfield(
    img8: np.ndarray, 
    pixel_size_um: float, 
    out_dir: Path,
    bacteria_config: 'SegmentationConfig'
) -> np.ndarray:
    """Brightfield segmentation - MAXIMUM intensity threshold for dark bacteria
    
    Args:
        img8: Input image (8-bit grayscale)
        pixel_size_um: Pixel size in micrometers
        out_dir: Output directory for debug images
        bacteria_config: SegmentationConfig object
    
    Returns:
        Binary mask (uint8, 0/255)
    """
    
    # Step 1: Apply inversion if needed (for DARK particles)
    if bacteria_config.invert_image:
        # DON'T invert the image - work directly on dark pixels
        save_debug(out_dir, "01a_original.png", img8, pixel_size_um)
        print("  Using DARK particle mode (no inversion)")
    
    # Step 2: STRONG blur to remove crystal texture
    blur = cv2.GaussianBlur(
        img8, (0, 0), 
        sigmaX=bacteria_config.gaussian_sigma, 
        sigmaY=bacteria_config.gaussian_sigma
    )
    save_debug(out_dir, "02_blurred.png", blur, pixel_size_um)
    
    # Step 3: Threshold - Keep only DARK pixels (bacteria)
    if bacteria_config.use_intensity_threshold:
        # METHOD A: Maximum intensity threshold (for very dark bacteria)
        # Keep pixels DARKER than threshold (invert logic)
        print(f"  Using MAXIMUM intensity threshold: {bacteria_config.intensity_threshold}")
        
        # Threshold: pixels < threshold become 255 (white/foreground)
        _, thresh = cv2.threshold(
            blur, 
            bacteria_config.intensity_threshold,  # e.g., 30
            255, 
            cv2.THRESH_BINARY_INV  # ✅ INVERTED - dark pixels become foreground
        )
        save_debug(out_dir, "03_thresh_dark_only.png", thresh, pixel_size_um)
        
    else:
        # METHOD B: Traditional enhancement + Otsu (original method)
        bg = cv2.GaussianBlur(
            img8, (0, 0), 
            sigmaX=bacteria_config.gaussian_sigma, 
            sigmaY=bacteria_config.gaussian_sigma
        )
        enhanced = cv2.subtract(bg, img8)
        enhanced_blur = cv2.GaussianBlur(enhanced, (3, 3), 0)
        
        save_debug(out_dir, "02_enhanced.png", enhanced, pixel_size_um)
        save_debug(out_dir, "03_enhanced_blur.png", enhanced_blur, pixel_size_um)
        
        _, thresh = cv2.threshold(
            enhanced_blur, 0, 255, 
            cv2.THRESH_BINARY + cv2.THRESH_OTSU
        )
        save_debug(out_dir, "04_thresh_otsu.png", thresh, pixel_size_um)
    
    # Step 4: Morphological cleanup (unchanged)
    kernel = np.ones(
        (bacteria_config.morph_kernel_size, bacteria_config.morph_kernel_size), 
        np.uint8
    )
    
    # Remove small noise
    bw = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, 
                         iterations=bacteria_config.morph_iterations)
    save_debug(out_dir, "05_opened.png", bw, pixel_size_um)
    
    # Fill holes and expand slightly
    bw = cv2.morphologyEx(bw, cv2.MORPH_CLOSE, kernel, 
                         iterations=bacteria_config.morph_iterations + 1)
    save_debug(out_dir, "06_closed.png", bw, pixel_size_um)
    
    # Optional dilate/erode
    if bacteria_config.dilate_iterations > 0:
        bw = cv2.dilate(bw, kernel, iterations=bacteria_config.dilate_iterations)
        save_debug(out_dir, "07_dilated.png", bw, pixel_size_um)
    
    if bacteria_config.erode_iterations > 0:
        bw = cv2.erode(bw, kernel, iterations=bacteria_config.erode_iterations)
        save_debug(out_dir, "08_eroded.png", bw, pixel_size_um)
    
    print(f"  Mask white fraction (final): {float((bw > 0).mean()):.4f}")
    return bw



def generate_rejection_analysis(output_root: Path) -> Optional[Path]:
    """Generate comprehensive rejection analysis across all groups
    
    Args:
        output_root: Root output directory
        
    Returns:
        Path to analysis CSV, or None if no data
    """
    
    print("\n" + "="*80)
    print("REJECTION ANALYSIS")
    print("="*80 + "\n")
    
    all_rejections = []
    
    # ✅ FIX: Use rglob with depth filter instead of fixed-depth glob.
    #
    # Old (broken in batch mode):
    #   output_root.glob("*/*_master.xlsx")
    #   → only matches depth-2 paths like  group/group_master.xlsx
    #   → misses batch-mode paths like     Positive/group/group_master.xlsx
    #
    # New:
    #   output_root.rglob("*_master.xlsx")  filtered to depth ≤ 3
    #   → matches single mode:  group/group_master.xlsx            (depth 2)
    #   → matches batch mode:   Positive/group/group_master.xlsx   (depth 3)
    all_master_excels = sorted(
        p for p in output_root.rglob("*_master.xlsx")
        if len(p.relative_to(output_root).parts) <= 3
    )
    
    if not all_master_excels:
        print("  No master Excel files found\n")
        return None
    
    # Collect rejection data from all groups
    for excel_path in all_master_excels:
        group_name = excel_path.parent.name
        
        # ✅ FIX: Build display-friendly group label that includes polarity
        # in batch mode so groups from Positive/ and Negative/ don't collide.
        rel_parts = excel_path.relative_to(output_root).parts
        if len(rel_parts) >= 3:
            # Batch mode: rel_parts = ("Positive", "1", "1_master.xlsx")
            polarity = rel_parts[0]  # "Positive" or "Negative"
            display_group = f"{polarity}/{group_name}"
        else:
            display_group = group_name
        
        try:
            rejected_sheet = f"{group_name}_Rejected_Objects"
            df = pd.read_excel(excel_path, sheet_name=rejected_sheet)
            
            if not df.empty:
                df['Group'] = display_group
                all_rejections.append(df)
                
        except Exception:
            pass  # Sheet doesn't exist or couldn't be read
    
    if not all_rejections:
        print("  No rejection data found\n")
        return None
    
    merged = pd.concat(all_rejections, ignore_index=True)
    
    # ========== ANALYSIS 1: Rejection reasons by frequency ==========
    print("📊 REJECTION REASONS BY FREQUENCY:\n")
    
    # Parse all reasons (semicolon-separated)
    all_reasons = []
    for reasons_str in merged['Rejection_Reasons']:
        if pd.notna(reasons_str):
            all_reasons.extend([r.strip() for r in str(reasons_str).split(';')])
    
    reason_counts = pd.Series(all_reasons).value_counts()
    
    for reason, count in reason_counts.head(10).items():
        pct = count / len(merged) * 100
        print(f"  • {reason}: {count} ({pct:.1f}%)")
    
    # ========== ANALYSIS 2: Rejection by group ==========
    print(f"\n📊 REJECTIONS BY GROUP:\n")
    
    group_summary = merged.groupby('Group').agg({
        'Object_ID': 'count',
        'BF_Area_px': ['mean', 'std'],
        'Circularity': ['mean', 'std'],
    }).round(2)
    
    print(group_summary.to_string())
    
    # ========== ANALYSIS 3: Size distribution of rejected vs accepted ==========
    print(f"\n📊 SIZE COMPARISON (Rejected vs Accepted):\n")
    
    # ✅ FIX: Re-use the same all_master_excels list (already depth-aware)
    # instead of a second broken glob.
    all_accepted = []
    for excel_path in all_master_excels:
        group_name = excel_path.parent.name
        
        try:
            typical_sheet = f"{group_name}_Typical_Particles"
            df = pd.read_excel(excel_path, sheet_name=typical_sheet)
            
            if 'BF_Area_um2' in df.columns:
                rel_parts = excel_path.relative_to(output_root).parts
                if len(rel_parts) >= 3:
                    display_group = f"{rel_parts[0]}/{group_name}"
                else:
                    display_group = group_name
                
                df['Group'] = display_group
                df['Status'] = 'Accepted'
                all_accepted.append(df[['Group', 'Status', 'BF_Area_um2', 'Circularity', 'AspectRatio']])
        except:
            pass
    
    if all_accepted:
        accepted_df = pd.concat(all_accepted, ignore_index=True)
        rejected_df = merged.copy()
        rejected_df['Status'] = 'Rejected'
        
        # Use BF_Area_um2 directly — rejected_objects.csv already has it
        # (renaming BF_Area_px would create a duplicate column name)
        rej_cols = ['Status']
        if 'BF_Area_um2' in rejected_df.columns:
            rej_cols.append('BF_Area_um2')
        elif 'BF_Area_px' in rejected_df.columns:
            rejected_df = rejected_df.rename(columns={'BF_Area_px': 'BF_Area_um2'})
            rej_cols.append('BF_Area_um2')
        if 'Circularity' in rejected_df.columns:
            rej_cols.append('Circularity')
        
        comparison = pd.concat([
            accepted_df[['Status', 'BF_Area_um2', 'Circularity']].reset_index(drop=True),
            rejected_df[rej_cols].reset_index(drop=True)
        ], ignore_index=True)
    
    # ========== Export full analysis ==========
    analysis_path = output_root / "rejection_analysis_summary.csv"
    
    # Create comprehensive analysis dataframe
    analysis_data = []
    
    for group in merged['Group'].unique():
        group_data = merged[merged['Group'] == group]
        
        # Parse reasons for this group
        group_reasons: list[str] = []

        if 'Rejection_Reasons' in group_data.columns:
            rejection_series = cast(pd.Series, group_data['Rejection_Reasons'])
            rejection_values = rejection_series.dropna().astype(str).tolist()

            for reasons_str in rejection_values:
                group_reasons.extend(
                    [r.strip() for r in reasons_str.split(';') if r.strip()]
                )
        
        reason_counts_group = pd.Series(group_reasons).value_counts()
        
        analysis_data.append({
            'Group': group,
            'Total_Rejected': len(group_data),
            'Mean_Area_px': group_data['BF_Area_px'].mean(),
            'Mean_Circularity': group_data['Circularity'].mean(),
            'Mean_AspectRatio': group_data['AspectRatio'].mean(),
            'Mean_Solidity': group_data['Solidity'].mean(),
            'Top_Rejection_Reason': reason_counts_group.index[0] if len(reason_counts_group) > 0 else 'N/A',
            'Top_Reason_Count': reason_counts_group.iloc[0] if len(reason_counts_group) > 0 else 0,
        })
    
    analysis_df = pd.DataFrame(analysis_data)
    analysis_df.to_csv(analysis_path, index=False)
    
    print(f"\n✓ Rejection analysis saved: {analysis_path.name}\n")
    print("="*80 + "\n")
    
    return analysis_path



# ==================================================
# Plotting and Analysis
# ==================================================
def generate_error_bar_comparison_with_threshold(
    output_dir: Path,
    percentile: float = 0.3,
    restrict_to_groups: Optional[list[str]] = None,
    output_path: Optional[Path] = None,
    title_suffix: str = "",
    dataset_id: str = "",
    threshold_pct: float = 0.05,
    microgel_type: str = "negative",
) -> Optional[Path]:
    """Enhanced version with threshold lines and control mean."""
    
    excel_files = [
        p for p in output_dir.rglob("*_master.xlsx")
        if len(p.relative_to(output_dir).parts) == 2
    ]   

    if not excel_files:
        print(f"[INFO] No master Excel files found under {output_dir}")
        return None

    all_data_rows: list[dict[str, object]] = []
    group_stats: dict[str, dict[str, float | int]] = {}
    control_mean = None

    # Load data
    for excel_path in sorted(excel_files):
        group_name_raw = excel_path.stem.replace("_master", "")
        display_name = _display_group_name(group_name_raw)

        if restrict_to_groups is not None and display_name not in restrict_to_groups:
            continue

        try:
            typical_sheet = f"{group_name_raw}_Typical_Particles"
            df = pd.read_excel(excel_path, sheet_name=typical_sheet)

            if "Fluor_Density_per_BF_Area" not in df.columns:
                print(f"[WARN] Missing Fluor_Density_per_BF_Area column in {excel_path.name}")
                continue

            values = df["Fluor_Density_per_BF_Area"].dropna()
            
            if len(values) == 0:
                print(f"[WARN] No valid data in {group_name_raw}")
                continue

            for v in values:
                all_data_rows.append(
                    {"Group": display_name, "Fluorescence Density": float(np.asarray(v).item())}
                )

            mean_val = float(np.asarray(values.mean()).item())
            std_val = float(np.asarray(values.std()).item())
            sem_val = float(np.asarray(values.sem()).item())

            group_stats[display_name] = {
                "n": int(len(values)),
                "mean": mean_val,
                "std": std_val,
                "sem": sem_val,
            }

            if display_name == "Control":
                control_mean = mean_val

            print(f"Loaded {len(values)} points for group: {display_name}")

        except Exception as e:
            print(f"[WARN] Could not read {group_name_raw}: {e}")
            continue

    if not all_data_rows:
        print("[WARN] No valid data found for comparison")
        return None

    df_all = pd.DataFrame(all_data_rows)

    group_order: list[str] = sorted(
        df_all["Group"].dropna().astype(str).drop_duplicates().tolist(), 
        key=_group_order_key
    )

    if len(group_order) < 1:
        print("[INFO] Need at least 1 group with data to generate plot.")
        return None

    # Color palette
    palette_colors: list[Any] = ["silver", "violet"]
    if len(group_order) > 2:
        palette_colors = list(sns.color_palette("husl", len(group_order)))
    elif len(group_order) == 1:
        palette_colors = ["skyblue"]

    # Generate plot
    plt.figure(figsize=(10, 7))
    sns.set_style("ticks")

    try:
        ax = sns.barplot(
            data=df_all,
            x="Group",
            y="Fluorescence Density",
            hue="Group",
            order=group_order,
            hue_order=group_order,
            palette=palette_colors,
            legend=False,
            errorbar=None,
            edgecolor="black",
            alpha=0.7,
        )
    except TypeError:
        ax = sns.barplot(
            data=df_all,
            x="Group",
            y="Fluorescence Density",
            hue="Group",
            order=group_order,
            hue_order=group_order,
            palette=palette_colors,
            legend=False,
            ci=None,
            edgecolor="black",
            alpha=0.7,
        )

    # Add SD error bars
    means = df_all.groupby("Group")["Fluorescence Density"].mean()
    sds = df_all.groupby("Group")["Fluorescence Density"].std(ddof=1)

    for xi, g in enumerate(group_order):
        m = float(np.asarray(means.get(g, 0.0)).item())
        sd = float(np.asarray(sds.get(g, 0.0)).item())
        cap = 14 if g == "Control" else 7

        ax.errorbar(
            xi, m, yerr=sd,
            fmt="none", ecolor="black", elinewidth=1.5,
            capsize=cap, capthick=1.5, zorder=10,
        )

    # Add jitter overlay
    sns.stripplot(
        x="Group", y="Fluorescence Density",
        data=df_all, order=group_order,
        jitter=True, color="cyan",
        edgecolor="black", linewidth=0.5,
        size=6, alpha=0.6,
    )

    # ✅ Add threshold lines - BOTH types use LOWER threshold
    legend_handles = []
    
    if control_mean is not None:
        # Control mean line (dotted blue)
        control_line = ax.axhline(
            y=control_mean, 
            color='blue', 
            linestyle=':', 
            linewidth=2.5, 
            label=f'Control Mean ({control_mean:.1f})', 
            zorder=5
        )
        legend_handles.append(control_line)
        
        # Calculate LOWER threshold for BOTH microgel types
        threshold = control_mean * (1 - threshold_pct)
        threshold_label = f'Lower Threshold (-{threshold_pct*100:.0f}%: {threshold:.1f})'
        
        threshold_line = ax.axhline(
            y=threshold, 
            color='red', 
            linestyle='--', 
            linewidth=2.5,
            label=threshold_label, 
            zorder=5
        )
        legend_handles.append(threshold_line)
        
        # Add legend
        ax.legend(
            handles=legend_handles,
            loc='upper right', 
            fontsize=10, 
            framealpha=0.95,
            edgecolor='black',
            fancybox=True
        )

    # Axis labels
    plt.ylabel("Fluorescence Density (a.u./µm²)", fontsize=12, fontweight="bold")
    plt.xlabel("")
    plt.xticks(fontsize=10, fontweight="bold")
    plt.yticks(fontsize=10, fontweight="bold")
    
    # Title - Consistent messaging
    filter_pct_display = int(percentile * 100)
    threshold_pct_display = int(threshold_pct * 100)
    
    title_parts = []
    if dataset_id:
        title_parts.append(dataset_id)
    
    # Microgel type description
    microgel_desc = "G- Microgel" if microgel_type.lower() == "negative" else "G+ Microgel"
    
    title_parts.append(
        f"{microgel_desc} — Typical Particles: Middle {100 - 2*filter_pct_display}% "
        f"(Excluded top/bottom {filter_pct_display}%)"
    )
    
    if title_suffix:
        title_parts.append(title_suffix)
    
    raw_title = " — ".join(title_parts)
    wrapped_title = "\n".join(
        textwrap.wrap(raw_title, width=80, break_long_words=False, break_on_hyphens=False)
    )
    plt.title(wrapped_title, fontsize=11, pad=10)

    # Enhance spines
    for axis in ["top", "bottom", "left", "right"]:
        plt.gca().spines[axis].set_linewidth(1.5)

    plt.tight_layout()

    # Save
    if output_path is not None:
        out_path = output_path
    else:
        if restrict_to_groups is not None:
            safe = "_".join([g.replace(" ", "_") for g in group_order])
            out_path = output_dir / f"comparison_{microgel_type}_{safe}.png"
        else:
            out_path = output_dir / f"comparison_{microgel_type}_all_groups.png"

    try:
        plt.savefig(out_path, dpi=300)
        plt.close()
        print(f"Saved plot with threshold: {out_path.name}")
        return out_path
    except Exception as e:
        print(f"[ERROR] Failed to save plot to {out_path}: {e}")
        plt.close()
        return None

def generate_pairwise_group_vs_control_plots(
    output_root: Path, 
    percentile: float, 
    dataset_id: str,
    threshold_pct: float,
    microgel_type: str
) -> None:
    """Generate pairwise plots with threshold lines"""
    
    control_folder = None
    for folder in output_root.iterdir():
        if folder.is_dir() and folder.name.lower().startswith("control"):
            control_folder = folder
            break
    
    if control_folder is None:
        print("[WARN] No Control group folder found - skipping pairwise plots")
        return
    
    control_master = control_folder / f"{control_folder.name}_master.xlsx"
    if not control_master.exists():
        print(f"[WARN] Control group master file not found: {control_master}")
        return
    
    control_display_name = _display_group_name(control_folder.name)
    
    for group_dir in sorted(output_root.iterdir()):
        if not group_dir.is_dir():
            continue
        if group_dir.name.lower().startswith("control"):
            continue
        if not re.fullmatch(r"\d+", group_dir.name):
            continue

        group_master = group_dir / f"{group_dir.name}_master.xlsx"
        if not group_master.exists():
            print(f"[WARN] Missing group master: {group_master}")
            continue

        pair_plot_path = group_dir / f"Group_{group_dir.name}_vs_Control_threshold.png"
        
        result = generate_error_bar_comparison_with_threshold(
            output_dir=output_root,
            percentile=percentile,
            restrict_to_groups=[group_dir.name, control_display_name],
            output_path=pair_plot_path,
            title_suffix=f"Group {group_dir.name} vs Control",
            dataset_id=dataset_id,
            threshold_pct=threshold_pct,
            microgel_type=microgel_type,
        )
        
        if result is not None:
            print(f"  Pairwise plot: {pair_plot_path.name}")


def embed_comparison_plots_into_all_excels(
    output_root: Path,
    percentile: float = 0.2,
    plot_path: Optional[Path] = None,
) -> None:
    """Post-process Excel files and embed plots"""
    
    if plot_path is None or not plot_path.exists():
        print(f"[WARN] Plot not found, embedding skipped")
        return
    
    excel_files = sorted(
        p for p in output_root.rglob("*_master.xlsx")
        if len(p.relative_to(output_root).parts) == 2
    )


    if not excel_files:
        print(f"[WARN] No master Excel files found under {output_root}")
        return

    def add_png(ws, path: Path, anchor_cell: str, width_px: int = 620) -> None:
        if not path.exists():
            return
        img = XLImage(str(path))
        if getattr(img, "width", None) and getattr(img, "height", None):
            scale = width_px / float(img.width)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
        ws.add_image(img, anchor_cell)

    updated = 0
    for excel_path in excel_files:
        try:
            wb = load_workbook(excel_path)
            modified = False

            if "Summary" in wb.sheetnames:
                ws_summary = wb["Summary"]
            elif "Error_Bar_Summary" in wb.sheetnames:
                ws_summary = wb["Error_Bar_Summary"]
                ws_summary.title = "Summary"
                ws_summary = wb["Summary"]
                modified = True
            else:
                ws_summary = wb.create_sheet("Summary")
                modified = True

            if "Ratios" in wb.sheetnames:
                ws_ratios = wb["Ratios"]
                ratios_idx = wb.sheetnames.index("Ratios")
                desired_idx = 1
                if ratios_idx != desired_idx:
                    wb.move_sheet(ws_ratios, offset=desired_idx - ratios_idx)
                    modified = True

            current_idx = wb.sheetnames.index(ws_summary.title)
            if current_idx != 0:
                wb.move_sheet(ws_summary, offset=-current_idx)
                modified = True

            marker_cell = "G1"
            marker_value = "COMPARISON_PLOTS_EMBEDDED"
            if ws_summary[marker_cell].value != marker_value:
                add_png(ws_summary, plot_path, "G3")
                ws_summary[marker_cell].value = marker_value
                modified = True

            if modified:
                wb.save(excel_path)
                updated += 1

        except Exception as e:
            print(f"[WARN] Could not embed plots into {excel_path}: {e}")

    if updated > 0:
        print(f"  Embedded plots in {updated} Excel files")


def consolidate_to_excel(output_dir: Path, group_name: str, percentile: float) -> None:
    """Consolidate all CSVs in a group folder into one Excel workbook"""
    csv_files = list(output_dir.glob("*/object_stats.csv"))

    if not csv_files:
        print(f"[WARN] No CSV files found in {output_dir}")
        return

    excel_path = output_dir / f"{group_name}_master.xlsx"

    if excel_path.exists():
        try:
            excel_path.unlink()
        except PermissionError:
            print(f"[ERROR] Cannot overwrite {excel_path} - file is open")
            return

    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")

        def adjust_column_widths(ws):
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.1
                if adjusted_width > 50:
                    adjusted_width = 50
                ws.column_dimensions[column_letter].width = adjusted_width

        def format_numbers(ws):
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0000'

        all_valid_objects: list[pd.DataFrame] = []
        all_excluded_objects: list[dict] = []

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            # README
            readme_df = pd.DataFrame(
                {
                    "Column Name": [
                        "Object_ID", "BF_Area_px", "BF_Area_um2", "Perimeter_px", "Perimeter_um",
                        "EquivDiameter_px", "EquivDiameter_um", "Circularity", "AspectRatio",
                        "CentroidX_px", "CentroidY_px", "CentroidX_um", "CentroidY_um",
                        "BBoxX_px", "BBoxY_px", "BBoxW_px", "BBoxH_px", "BBoxW_um", "BBoxH_um",
                        "Fluor_Area_px", "Fluor_Area_um2", "Fluor_Mean", "Fluor_Median",
                        "Fluor_Std", "Fluor_Min", "Fluor_Max", "Fluor_IntegratedDensity",
                        "Fluor_Density_per_BF_Area", "BF_to_Fluor_Area_Ratio",
                    ],
                    "Description": [
                        "Unique particle identifier", "Brightfield particle area (pixels²)", "Brightfield particle area (µm²)",
                        "Particle perimeter (pixels)", "Particle perimeter (µm)", "Diameter of equivalent circle (pixels)",
                        "Diameter of equivalent circle (µm)", "Shape roundness (0-1)", "Bounding box width/height ratio",
                        "Particle center X (px)", "Particle center Y (px)", "Particle center X (µm)", "Particle center Y (µm)",
                        "BBox top-left X (px)", "BBox top-left Y (px)", "BBox width (px)", "BBox height (px)",
                        "BBox width (µm)", "BBox height (µm)",
                        "Fluorescent region area (pixels²)", "Fluorescent region area (µm²)",
                        "Avg fluorescence intensity", "Median fluorescence intensity", "Std Dev fluorescence",
                        "Min fluorescence", "Max fluorescence", "Total fluorescence signal",
                        "Fluor density / BF Area (Primary Metric)", "Ratio of BF area to Fluor area",
                    ],
                }
            )

            readme_df.to_excel(writer, sheet_name="README", index=False)
            ws_readme = writer.sheets["README"]
            for cell in ws_readme[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            adjust_column_widths(ws_readme)

            # Per-image sheets
            for csv_file in sorted(csv_files):
                image_name = csv_file.parent.name
                df = pd.read_csv(csv_file)

                cols_to_numeric = ["Fluor_Area_px", "Fluor_IntegratedDensity", "BF_Area_um2", "Fluor_Area_um2"]
                for col in cols_to_numeric:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

                if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                    df["Fluor_Density_per_BF_Area"] = df["Fluor_IntegratedDensity"] / df["BF_Area_um2"]
                else:
                    df["Fluor_Density_per_BF_Area"] = 0.0

                if "Fluor_Area_um2" in df.columns and "BF_Area_um2" in df.columns:
                    df["BF_to_Fluor_Area_Ratio"] = np.where(df["BF_Area_um2"] > 0, df["BF_Area_um2"] / df["Fluor_Area_um2"], 0.0)
                else:
                    df["BF_to_Fluor_Area_Ratio"] = 0.0

                df = df.replace([np.inf, -np.inf], 0).fillna(0)

                for idx, row in df.iterrows():
                    reason = None
                    
                    if row["Fluor_Area_px"] == 0 and row["Fluor_IntegratedDensity"] > 0:
                        reason = "Zero fluorescence area with positive integrated density"
                    elif row["Fluor_Area_px"] == 0:
                        reason = "Zero fluorescence area"
                    elif row["Fluor_IntegratedDensity"] == 0:
                        reason = "Zero integrated density"
                    
                    if reason:
                        all_excluded_objects.append({
                            "Object_ID": row["Object_ID"],
                            "Source_Image": image_name,
                            "BF_Area_um2": row["BF_Area_um2"],
                            "Fluor_Area_px": row["Fluor_Area_px"],
                            "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"],
                            "Exclusion_Reason": reason
                        })

                df_valid = df[(df["Fluor_IntegratedDensity"] > 0) & (df["Fluor_Area_px"] > 0)].copy()
                df_valid["Source_Image"] = image_name
                
                if not df_valid.empty:
                    all_valid_objects.append(df_valid)

                sheet_name = image_name[:31]
                
                if "Fluor_Density_per_BF_Area" in df.columns:
                    df = df.sort_values("Fluor_Density_per_BF_Area", ascending=False)
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                format_numbers(ws)
                adjust_column_widths(ws)
                ws.auto_filter.ref = ws.dimensions

            # Excluded Objects
            if all_excluded_objects:
                excluded_df = pd.DataFrame(all_excluded_objects)
                excluded_df.to_excel(writer, sheet_name="Excluded_Objects", index=False)
                ws_excluded = writer.sheets["Excluded_Objects"]
                
                for cell in ws_excluded[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                for row in ws_excluded.iter_rows(min_row=2, max_row=ws_excluded.max_row):
                    for cell in row:
                        cell.fill = red_fill
                
                adjust_column_widths(ws_excluded)
                ws_excluded.auto_filter.ref = ws_excluded.dimensions

            # All Valid Objects
            if all_valid_objects:
                merged_all = pd.concat(all_valid_objects, ignore_index=True)
                merged_all = merged_all.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)
                
                all_valid_sheet_name = f"{group_name}_All_Valid_Objects"
                merged_all.to_excel(writer, sheet_name=all_valid_sheet_name, index=False)
                ws_all = writer.sheets[all_valid_sheet_name]
                
                for cell in ws_all[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                format_numbers(ws_all)
                adjust_column_widths(ws_all)
                ws_all.auto_filter.ref = ws_all.dimensions

            # Typical Particles (GROUP-LEVEL filtering)
            if all_valid_objects:
                merged_all = pd.concat(all_valid_objects, ignore_index=True)
                merged_all = merged_all.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)
                
                n_total = len(merged_all)
                n_cut = int(n_total * percentile)
                
                start_idx = n_cut
                end_idx = n_total - n_cut
                
                if start_idx < end_idx and n_total > 3:
                    typical_particles = merged_all.iloc[start_idx:end_idx].copy()
                    
                    excluded_top = merged_all.iloc[:start_idx]
                    excluded_bottom = merged_all.iloc[end_idx:]
                    
                    for idx, row in excluded_top.iterrows():
                        all_excluded_objects.append({
                            "Object_ID": row["Object_ID"],
                            "Source_Image": row["Source_Image"],
                            "BF_Area_um2": row["BF_Area_um2"],
                            "Fluor_Area_px": row["Fluor_Area_px"],
                            "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"],
                            "Exclusion_Reason": f"Outside typical particle range (top {int(percentile*100)}%)"
                        })
                    
                    for idx, row in excluded_bottom.iterrows():
                        all_excluded_objects.append({
                            "Object_ID": row["Object_ID"],
                            "Source_Image": row["Source_Image"],
                            "BF_Area_um2": row["BF_Area_um2"],
                            "Fluor_Area_px": row["Fluor_Area_px"],
                            "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"],
                            "Exclusion_Reason": f"Outside typical particle range (bottom {int(percentile*100)}%)"
                        })
                else:
                    typical_particles = merged_all.copy()

                typical_sheet_name = f"{group_name}_Typical_Particles"
                typical_particles.to_excel(writer, sheet_name=typical_sheet_name, index=False)
                ws_typ = writer.sheets[typical_sheet_name]

                for cell in ws_typ[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                for row in ws_typ.iter_rows(min_row=2, max_row=ws_typ.max_row):
                    for cell in row:
                        cell.fill = yellow_fill
                
                format_numbers(ws_typ)
                adjust_column_widths(ws_typ)
                ws_typ.auto_filter.ref = ws_typ.dimensions


            # ========== REJECTED OBJECTS SUMMARY (FIXED) ==========
            # ✅ Explicitly initialize list
            all_rejected_objects: list[pd.DataFrame] = []
            
            for csv_file in sorted(csv_files):
                image_name = csv_file.parent.name
                rejected_csv = csv_file.parent / "rejected_objects.csv"
                
                # ✅ Check if file exists before trying to read
                if rejected_csv.exists():
                    try:
                        rej_df = pd.read_csv(rejected_csv)
                        
                        # ✅ Check if DataFrame is not empty
                        if not rej_df.empty:
                            rej_df['Source_Image'] = image_name
                            all_rejected_objects.append(rej_df)
                    except Exception as e:
                        print(f"  ⚠ Could not read rejected objects from {image_name}: {e}")
            
            # ✅ Only create sheet if we actually have rejected data
            if all_rejected_objects:  # Check the list, not undefined variables
                try:
                    merged_rejected = pd.concat(all_rejected_objects, ignore_index=True)
                    
                    # Sort by area (largest first)
                    if 'BF_Area_px' in merged_rejected.columns:
                        merged_rejected = merged_rejected.sort_values('BF_Area_px', ascending=False).reset_index(drop=True)
                    
                    rejected_sheet_name = f"{group_name}_Rejected_Objects"
                    merged_rejected.to_excel(writer, sheet_name=rejected_sheet_name, index=False)
                    ws_rej = writer.sheets[rejected_sheet_name]
                    
                    # Format header
                    for cell in ws_rej[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    
                    # Highlight rows with red
                    for row in ws_rej.iter_rows(min_row=2, max_row=ws_rej.max_row):
                        for cell in row:
                            cell.fill = red_fill
                    
                    format_numbers(ws_rej)
                    adjust_column_widths(ws_rej)
                    ws_rej.auto_filter.ref = ws_rej.dimensions
                    
                except Exception as e:
                    print(f"  ⚠ Could not create Rejected_Objects sheet: {e}")


            # Update Excluded Objects
            if all_excluded_objects:
                wb = writer.book
                if "Excluded_Objects" in wb.sheetnames:
                    wb.remove(wb["Excluded_Objects"])
                
                excluded_df = pd.DataFrame(all_excluded_objects)
                excluded_df.to_excel(writer, sheet_name="Excluded_Objects", index=False)
                ws_excluded = writer.sheets["Excluded_Objects"]
                
                for cell in ws_excluded[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                for row in ws_excluded.iter_rows(min_row=2, max_row=ws_excluded.max_row):
                    for cell in row:
                        cell.fill = red_fill
                
                adjust_column_widths(ws_excluded)
                ws_excluded.auto_filter.ref = ws_excluded.dimensions

            # Summary sheet
            try:
                summary_data = []
                for csv_file in sorted(csv_files):
                    image_name = csv_file.parent.name
                    df = pd.read_csv(csv_file)

                    if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                        df["Fluor_Density_per_BF_Area"] = pd.to_numeric(df["Fluor_IntegratedDensity"], errors='coerce') / pd.to_numeric(df["BF_Area_um2"], errors='coerce')
                    else:
                        df["Fluor_Density_per_BF_Area"] = 0

                    if "Fluor_Area_px" in df.columns:
                        df["Fluor_Area_px"] = pd.to_numeric(df["Fluor_Area_px"], errors='coerce').fillna(0)
                        
                    df = df.replace([np.inf, -np.inf], 0).fillna(0)
                    
                    df_stats = df[(df["Fluor_Area_px"] > 0) & (df["Fluor_IntegratedDensity"] > 0)]
                    
                    avg_fluor_density = df_stats["Fluor_Density_per_BF_Area"].mean() if not df_stats.empty else 0.0
                    
                    avg_ratio = 0.0
                    if not df_stats.empty and "BF_Area_um2" in df_stats.columns and "Fluor_Area_um2" in df_stats.columns:
                         ratios = df_stats["BF_Area_um2"] / df_stats["Fluor_Area_um2"]
                         avg_ratio = ratios.replace([np.inf, -np.inf], 0).mean()

                    summary_data.append({
                        "Image": image_name,
                        "Total_Particles_Detected": len(df),
                        "Particles_With_Fluor": len(df_stats),
                        "Avg_BF_Area_um2": df["BF_Area_um2"].mean() if "BF_Area_um2" in df.columns else 0,
                        "Avg_Fluor_Density": avg_fluor_density,
                        "Avg_BF_to_Fluor_Ratio": avg_ratio,
                    })

                summary_df = pd.DataFrame(summary_data)

                if "Avg_Fluor_Density" in summary_df.columns:
                    summary_df = summary_df.sort_values("Avg_Fluor_Density", ascending=False)

                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                ws_summary = writer.sheets["Summary"]

                for cell in ws_summary[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                format_numbers(ws_summary)
                adjust_column_widths(ws_summary)

            except Exception as e:
                print(f"[WARN] Could not create Summary sheet: {e}")

            # Ratios sheet (unchanged - keeping existing code)
            try:
                wb = writer.book
                ratios_name = "Ratios"
                if ratios_name in wb.sheetnames:
                    wb.remove(wb[ratios_name])
                ws_qa = wb.create_sheet(ratios_name)

                ws_qa["A1"] = f"QA Ratios - Group: {group_name}"
                ws_qa["A1"].font = Font(bold=True, size=14)

                def add_png(ws, path: Path, anchor_cell: str, width_px: int = 360) -> None:
                    if not path.exists(): return
                    img = XLImage(str(path))
                    if getattr(img, "width", None) and getattr(img, "height", None):
                        scale = width_px / float(img.width)
                        img.width = int(img.width * scale)
                        img.height = int(img.height * scale)
                    ws.add_image(img, anchor_cell)

                row = 3
                block_h = 34

                for csv_file in sorted(csv_files):
                    image_name = csv_file.parent.name
                    df = pd.read_csv(csv_file)
                    
                    if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                        df["Fluor_Density_per_BF_Area"] = pd.to_numeric(df["Fluor_IntegratedDensity"], errors='coerce') / pd.to_numeric(df["BF_Area_um2"], errors='coerce')
                    else:
                        df["Fluor_Density_per_BF_Area"] = 0.0
                    
                    if "BF_Area_um2" in df.columns and "Fluor_Area_um2" in df.columns:
                        df["BF_to_Fluor_Area_Ratio"] = pd.to_numeric(df["BF_Area_um2"], errors='coerce') / pd.to_numeric(df["Fluor_Area_um2"], errors='coerce')
                    else:
                        df["BF_to_Fluor_Area_Ratio"] = 0.0

                    df = df.replace([np.inf, -np.inf], 0).fillna(0)

                    if "Fluor_Density_per_BF_Area" in df.columns:
                        df = df.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)

                    ws_qa[f"A{row}"] = image_name
                    ws_qa[f"A{row}"].font = Font(bold=True, size=12)
                    
                    headers = ["Object_ID", "Fluor_Density_per_BF_Area", "Rank", "Object_ID", "BF_to_Fluor_Area_Ratio"]
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws_qa.cell(row=row+1, column=col_idx, value=header)
                        cell.font = Font(bold=True)
                        cell.alignment = center_align

                    start_data_row = row + 2
                    n = len(df)
                    
                    for k, r in enumerate(df.itertuples(index=False), 0):
                        ws_qa.cell(row=start_data_row+k, column=1, value=getattr(r, "Object_ID"))
                        ws_qa.cell(row=start_data_row+k, column=2, value=float(getattr(r, "Fluor_Density_per_BF_Area", 0.0))).number_format = '0.0000'
                        ws_qa.cell(row=start_data_row+k, column=3, value=k+1)
                        ws_qa.cell(row=start_data_row+k, column=4, value=getattr(r, "Object_ID"))
                        ws_qa.cell(row=start_data_row+k, column=5, value=float(getattr(r, "BF_to_Fluor_Area_Ratio", 0.0))).number_format = '0.0000'

                    ws_qa.column_dimensions["A"].width = 20
                    ws_qa.column_dimensions["B"].width = 25
                    ws_qa.column_dimensions["C"].width = 10
                    ws_qa.column_dimensions["D"].width = 20
                    ws_qa.column_dimensions["E"].width = 25

                    if n > 0:
                        ch1 = ScatterChart()
                        ch1.title = "Fluor Density"
                        ch1.y_axis.title = "a.u./µm²"
                        ch1.x_axis.title = "Object Rank (Sorted)"
                        
                        xref = Reference(ws_qa, min_col=3, min_row=start_data_row, max_row=start_data_row + n - 1)
                        yref = Reference(ws_qa, min_col=2, min_row=start_data_row, max_row=start_data_row + n - 1)
                        
                        s1 = SeriesFactory(yref, xref, title="Density")
                        s1.marker = Marker(symbol="triangle", size=5)
                        ch1.series.append(s1)
                        ws_qa.add_chart(ch1, f"G{row+1}")

                        ch2 = ScatterChart()
                        ch2.title = "Area Ratio"
                        ch2.y_axis.title = "Ratio"
                        ch2.x_axis.title = "Object Rank (Sorted)"
                        
                        yref2 = Reference(ws_qa, min_col=5, min_row=start_data_row, max_row=start_data_row + n - 1)
                        
                        s2 = SeriesFactory(yref2, xref, title="Ratio")
                        s2.marker = Marker(symbol="circle", size=5)
                        ch2.series.append(s2)
                        ws_qa.add_chart(ch2, f"G{row+18}")

                    img_dir = csv_file.parent
                    add_png(ws_qa, img_dir / "13_mask_accepted_ids.png", f"Q{row+1}", width_px=330)
                    add_png(ws_qa, img_dir / "22_fluorescence_mask_global_ids.png", f"Q{row+18}", width_px=330)
                    add_png(ws_qa, img_dir / "23_fluorescence_contours_global_ids.png", f"Q{row+18}", width_px=330)
                    add_png(ws_qa, img_dir / "24_bf_fluor_matching_overlay_ids.png", f"Q{row+1}", width_px=330)

                    row += block_h

            except Exception as e:
                print(f"[WARN] Could not create Ratios sheet: {e}")

        # Reorder worksheets
        wb2 = load_workbook(excel_path)
        desired_order = [
            "Summary", 
            "Ratios", 
            "README", 
            f"{group_name}_Typical_Particles",
            f"{group_name}_All_Valid_Objects",
            "Excluded_Objects"
        ]
        
        # ✅ Add Rejected_Objects to desired order if it exists
        rejected_sheet_name = f"{group_name}_Rejected_Objects"
        if rejected_sheet_name in wb2.sheetnames:
            desired_order.insert(5, rejected_sheet_name)
        
        for idx, sheet_name in enumerate(desired_order):
            if sheet_name in wb2.sheetnames:
                sheet = wb2[sheet_name]
                current_idx = wb2.sheetnames.index(sheet_name)
                if current_idx != idx:
                    wb2.move_sheet(sheet, offset=idx - current_idx)
        wb2.save(excel_path)

    except PermissionError:
        print(f"[ERROR] Cannot write to {excel_path} - file may be open")
    except Exception as e:
        print(f"[ERROR] Failed to create Excel file: {e}")
        import traceback
        traceback.print_exc()



def export_group_statistics_to_csv(output_root: Path) -> None:
    """Export statistics with enhanced console summary"""
    
    stats_list = []
    
    for excel_path in sorted(output_root.glob("*/*_master.xlsx")):
        group_name = excel_path.parent.name
        
        try:
            typical_sheet = f"{group_name}_Typical_Particles"
            df = pd.read_excel(excel_path, sheet_name=typical_sheet)
                        
            if "Fluor_Density_per_BF_Area" in df.columns:
                values = pd.to_numeric(df["Fluor_Density_per_BF_Area"], errors='coerce').dropna()
                
                if values.empty:
                    continue

                mean_val = float(values.mean())
                std_val = float(values.std(ddof=1))
                sem_val = values.sem()
                median_val = float(values.median())
                min_val = float(values.min())
                max_val = float(values.max())
                q30 = float(values.quantile(0.30))
                q70 = float(values.quantile(0.70))
                n = int(len(values))
                
                stats_list.append({
                    'Group': "Control" if group_name.lower().startswith("control") else group_name,
                    'N': n,
                    'Mean': mean_val,
                    'Std_Dev': std_val,
                    'SEM': sem_val,
                    'Median': median_val,
                    'Q30': q30,
                    'Q70': q70,
                    'Min': min_val,
                    'Max': max_val,
                    'CV_percent': (std_val / mean_val * 100) if mean_val > 0 else 0,
                })
                
        except Exception:
            pass
    
    if not stats_list:
        return
    
    stats_df = pd.DataFrame(stats_list)
    
    stats_df['sort_key'] = stats_df['Group'].apply(
        lambda x: (0, int(x)) if x.isdigit() else (1, 999)
    )
    stats_df = stats_df.sort_values('sort_key').drop('sort_key', axis=1)
    
    numeric_cols = stats_df.select_dtypes(include=[np.number]).columns
    stats_df[numeric_cols] = stats_df[numeric_cols].round(2)
    
    output_path = output_root / "group_statistics_summary.csv"
    stats_df.to_csv(output_path, index=False)

def _as_float(x: object) -> float:
    v = np.asarray(x).item()
    if isinstance(v, complex):
        raise TypeError("Expected real number, got complex")
    return float(v)



def classify_groups_clinical(
    output_root: Path,
    microgel_type: str = "negative",
    threshold_pct: float = 0.05
) -> pd.DataFrame:
    """Classify groups with comprehensive statistical annotations.

    Columns produced (per group):
        N, Mean, Std_Dev, SEM, Median,
        CI_Lower, CI_Upper,
        Control_Mean, Threshold,
        Diff_from_Threshold, Diff_from_Control, Pct_Diff_from_Control,
        Cohens_d, d_CI_Lower, d_CI_Upper, Effect_Size_Label,
        P_Value, Significance,
        Classification, Classification_Confidence, Strength_of_Evidence
    """

    if not (0.01 <= threshold_pct <= 0.50):
        print(f"  ⚠ Unusual threshold: {threshold_pct*100:.1f}%")

    # ========== Locate control ==========
    control_folder = None
    for folder in output_root.iterdir():
        if folder.is_dir() and folder.name.lower().startswith("control"):
            control_folder = folder
            break

    if control_folder is None:
        return pd.DataFrame()

    control_master = control_folder / f"{control_folder.name}_master.xlsx"
    if not control_master.exists():
        return pd.DataFrame()

    try:
        typical_sheet = f"{control_folder.name}_Typical_Particles"
        df_control = pd.read_excel(control_master, sheet_name=typical_sheet)
        if "Fluor_Density_per_BF_Area" not in df_control.columns:
            return pd.DataFrame()
        control_values = pd.to_numeric(
            df_control["Fluor_Density_per_BF_Area"], errors='coerce'
        ).dropna()
        control_mean = float(control_values.mean())
        control_std = float(control_values.std(ddof=1))
    except Exception:
        return pd.DataFrame()

    threshold = control_mean * (1 - threshold_pct)

    # ── helpers ──
    def _effect_label(d: float) -> str:
        ad = abs(d)
        if ad < 0.2:
            return "Negligible"
        elif ad < 0.5:
            return "Small"
        elif ad < 0.8:
            return "Medium"
        else:
            return "Large"

    def _d_ci(d: float, n1: int, n2: int) -> tuple[float, float]:
        """Approximate 95 % CI for Cohen's d (Hedges & Olkin)."""
        if n1 + n2 < 4:
            return (d, d)
        d_f = float(d)
        n_sum = float(n1 + n2)
        n_prod = float(n1 * n2)
        se = float(np.sqrt(n_sum / n_prod + d_f * d_f / (2.0 * n_sum)))
        return (round(float(d_f - 1.96 * se), 3), round(float(d_f + 1.96 * se), 3))


    def _strength(p: float, d: float, n: int) -> str:
        if pd.isna(p):
            return "Insufficient (n<2)"
        if p < 0.01 and abs(d) >= 0.8 and n >= 5:
            return "Strong"
        if p < 0.05 and abs(d) >= 0.5 and n >= 3:
            return "Moderate"
        if p < 0.05 or abs(d) >= 0.5:
            return "Weak"
        return "Insufficient"

    # ========== Control row ==========
    results = []
    ctrl_n = len(control_values)
    ctrl_sem = (control_std / np.sqrt(ctrl_n)) if ctrl_n >= 2 else 0.0
    ctrl_median = float(control_values.median()) if ctrl_n > 0 else 0.0

    if ctrl_n >= 2:
        t_crit = float(scipy_stats.t.ppf(0.975, df=ctrl_n - 1))
        ctrl_ci_lo = control_mean - t_crit * ctrl_sem
        ctrl_ci_hi = control_mean + t_crit * ctrl_sem
    else:
        ctrl_ci_lo = control_mean
        ctrl_ci_hi = control_mean

    results.append({
        'Group': 'Control',
        'N': ctrl_n,
        'Mean': round(control_mean, 2),
        'Std_Dev': round(control_std, 2),
        'SEM': round(ctrl_sem, 2),
        'Median': round(ctrl_median, 2),
        'CI_Lower': round(ctrl_ci_lo, 2),
        'CI_Upper': round(ctrl_ci_hi, 2),
        'Control_Mean': round(control_mean, 2),
        'Threshold': round(threshold, 2),
        'Diff_from_Threshold': 0.0,
        'Diff_from_Control': 0.0,
        'Pct_Diff_from_Control': 0.0,
        'Cohens_d': 0.0,
        'd_CI_Lower': 0.0,
        'd_CI_Upper': 0.0,
        'Effect_Size_Label': '—',
        'P_Value': np.nan,
        'Significance': '—',
        'Classification': 'CONTROL/Reference',
        'Classification_Confidence': '—',
        'Strength_of_Evidence': '—',
    })

    # ========== Test groups ==========
    for excel_path in sorted(output_root.glob("*/*_master.xlsx")):
        group_name = excel_path.parent.name
        if group_name.lower().startswith("control"):
            continue

        try:
            typical_sheet = f"{group_name}_Typical_Particles"
            df = pd.read_excel(excel_path, sheet_name=typical_sheet)
            if "Fluor_Density_per_BF_Area" not in df.columns:
                continue
            values = pd.to_numeric(
                df["Fluor_Density_per_BF_Area"], errors='coerce'
            ).dropna()
            if values.empty:
                continue
        except Exception:
            continue

        mean_val = _as_float(values.mean())
        std_val = _as_float(values.std(ddof=1))
        n = len(values)
        sem_val = _as_float(values.sem()) if n >= 2 else 0.0
        median_val = _as_float(values.median())

        # Classification logic
        if mean_val < threshold:
            if microgel_type.lower() == "negative":
                classification = "NEGATIVE / Bacteria Detected"
            else:
                classification = "POSITIVE / Bacteria Detected"
        else:
            if microgel_type.lower() == "negative":
                classification = "POSITIVE / No obvious bacteria"
            else:
                classification = "NEGATIVE / No obvious bacteria"

        # Statistical comparison vs control
        stats = _compute_group_vs_control_stats(values, control_values, threshold)

        d_val = stats['Cohens_d']
        d_lo, d_hi = _d_ci(d_val, n, ctrl_n)

        diff_from_threshold = mean_val - threshold
        diff_from_control = mean_val - control_mean
        pct_diff = (diff_from_control / control_mean) * 100 if control_mean != 0 else 0

        results.append({
            'Group': group_name,
            'N': n,
            'Mean': round(mean_val, 2),
            'Std_Dev': round(std_val, 2),
            'SEM': round(sem_val, 2),
            'Median': round(median_val, 2),
            'CI_Lower': stats['CI_Lower'],
            'CI_Upper': stats['CI_Upper'],
            'Control_Mean': round(control_mean, 2),
            'Threshold': round(threshold, 2),
            'Diff_from_Threshold': round(diff_from_threshold, 2),
            'Diff_from_Control': round(diff_from_control, 2),
            'Pct_Diff_from_Control': round(pct_diff, 1),
            'Cohens_d': d_val,
            'd_CI_Lower': d_lo,
            'd_CI_Upper': d_hi,
            'Effect_Size_Label': _effect_label(d_val),
            'P_Value': stats['P_Value'],
            'Significance': stats['Significance'],
            'Classification': classification,
            'Classification_Confidence': stats['Classification_Confidence'],
            'Strength_of_Evidence': _strength(
                stats['P_Value'], d_val, n
            ),
        })

    if not results:
        return pd.DataFrame()

    results_df = pd.DataFrame(results)
    results_df['sort_key'] = results_df['Group'].apply(
        lambda x: (1, 999) if x == 'Control' else (0, int(x) if x.isdigit() else 999)
    )
    results_df = results_df.sort_values('sort_key').drop('sort_key', axis=1)
    return results_df


def export_clinical_classification(
    output_root: Path,
    classification_df: pd.DataFrame,
    microgel_type: str = "negative"
) -> Optional[Path]:
    """Export with VIEWER-COMPATIBLE naming"""
    
    if classification_df.empty:
        return None
    
    # VIEWER-COMPATIBLE naming: clinical_classification_positive.csv
    csv_path = output_root / f"clinical_classification_{microgel_type}.csv"
    classification_df.to_csv(csv_path, index=False)
    
    excel_path = output_root / f"clinical_classification_{microgel_type}.xlsx"
    
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            classification_df.to_excel(writer, sheet_name='Classification', index=False)
            
            ws = writer.sheets['Classification']
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            safe_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            for row_idx in range(len(classification_df)):
                excel_row = row_idx + 2
                row_data = classification_df.iloc[row_idx]
                
                if "NEGATIVE" in row_data['Classification'] or "POSITIVE" in row_data['Classification']:
                    if "No obvious bacteria" in row_data['Classification']:
                        fill = safe_fill
                    else:
                        fill = warning_fill
                else:
                    fill = safe_fill
                
                for col_idx in range(1, len(classification_df.columns) + 1):
                    ws.cell(row=excel_row, column=col_idx).fill = fill
                    ws.cell(row=excel_row, column=col_idx).alignment = Alignment(horizontal="center")
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.1, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
    except Exception:
        pass
    
    return csv_path


def generate_final_clinical_matrix(
    output_root: Path,
    gplus_classification: pd.DataFrame,
    gminus_classification: pd.DataFrame,
    dataset_base_name: str
) -> Optional[Path]:
    """Generate comprehensive final matrix combining G+/G- with all metrics."""

    if gplus_classification.empty or gminus_classification.empty:
        print("[WARN] Missing classification data")
        return None

    gplus_classification['Group'] = gplus_classification['Group'].astype(str)
    gminus_classification['Group'] = gminus_classification['Group'].astype(str)

    def _norm(classification):
        if classification is None:
            return 'unknown'
        upper = str(classification).upper()
        if 'CONTROL' in upper:
            return 'control'
        if 'NO OBVIOUS' in upper:
            return 'not_detected'
        if 'BACTERIA DETECTED' in upper or 'DETECTED' in upper:
            return 'detected'
        return 'unknown'

    decision_matrix = {
        ('detected',     'not_detected'): 'POSITIVE',
        ('not_detected', 'detected'):     'NEGATIVE',
        ('not_detected', 'not_detected'): 'NO OBVIOUS BACTERIA',
        ('detected',     'detected'):     'MIXED/CONTRADICTORY',
    }
    detection_labels = {
        'detected': 'Detected', 'not_detected': 'Not Detected',
        'control': 'Control', 'unknown': 'Unknown',
    }

    gplus_dict = gplus_classification.set_index('Group').to_dict('index')
    gminus_dict = gminus_classification.set_index('Group').to_dict('index')

    all_groups = sorted(
        set(gplus_classification['Group']) | set(gminus_classification['Group']),
        key=_group_order_key,
    )

    def _safe(d, key, fmt="{:.2f}"):
        v = d.get(key)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return '—'
        try:
            return fmt.format(float(v))
        except (TypeError, ValueError):
            return str(v)

    def _safe_str(d, key):
        v = d.get(key)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return '—'
        return str(v)

    results = []
    for group in all_groups:
        gp = gplus_dict.get(group, {})
        gm = gminus_dict.get(group, {})

        gp_class = str(gp.get('Classification', '')) if gp else None
        gm_class = str(gm.get('Classification', '')) if gm else None
        gp_norm = _norm(gp_class)
        gm_norm = _norm(gm_class)

        if group == 'Control':
            final_class = 'CONTROL (Reference)'
        elif gp_class is None and gm_class is None:
            final_class = 'MISSING DATA'
        elif gp_class is None:
            final_class = 'MISSING G+'
        elif gm_class is None:
            final_class = 'MISSING G-'
        else:
            final_class = decision_matrix.get((gp_norm, gm_norm), 'UNKNOWN COMBINATION')

        # Interpretation
        if group == 'Control':
            interpretation = 'CONTROL (Reference)'
        else:
            parts = []
            if final_class == 'NEGATIVE':
                parts.append("Gram-negative bacteria DETECTED")
            elif final_class == 'POSITIVE':
                parts.append("Gram-positive bacteria DETECTED")
            elif final_class == 'NO OBVIOUS BACTERIA':
                parts.append("No obvious bacteria detected")
            elif final_class == 'MIXED/CONTRADICTORY':
                parts.append("Contradictory — manual review required")
            else:
                parts.append(final_class)
            for label, dd in [("G+", gp), ("G-", gm)]:
                sig = _safe_str(dd, 'Significance')
                pv = dd.get('P_Value')
                if pv is not None and not (isinstance(pv, float) and pd.isna(pv)):
                    try:
                        pf = float(pv)
                        pstr = f"p<0.001" if pf < 0.001 else f"p={pf:.3f}"
                        parts.append(f"[{label} {pstr} {sig}]")
                    except (TypeError, ValueError):
                        pass
            interpretation = " ".join(parts)

        row = {
            'Group': group,
            # G+
            'G+_N': _safe(gp, 'N', "{:.0f}"),
            'G+_Mean': _safe(gp, 'Mean'),
            'G+_Std': _safe(gp, 'Std_Dev'),
            'G+_SEM': _safe(gp, 'SEM'),
            'G+_Median': _safe(gp, 'Median'),
            'G+_CI': f"{_safe(gp, 'CI_Lower')} – {_safe(gp, 'CI_Upper')}",
            'G+_Cohens_d': _safe(gp, 'Cohens_d', "{:.3f}"),
            'G+_d_CI': f"{_safe(gp, 'd_CI_Lower', '{:.3f}')} – {_safe(gp, 'd_CI_Upper', '{:.3f}')}",
            'G+_Effect_Size': _safe_str(gp, 'Effect_Size_Label'),
            'G+_P_Value': _safe(gp, 'P_Value', "{:.4f}"),
            'G+_Significance': _safe_str(gp, 'Significance'),
            'G+_Detection': detection_labels.get(gp_norm, 'Unknown'),
            'G+_Confidence': _safe_str(gp, 'Classification_Confidence'),
            'G+_Strength': _safe_str(gp, 'Strength_of_Evidence'),
            'G+_Pct_Diff': _safe(gp, 'Pct_Diff_from_Control', "{:.1f}"),
            'G+_Classification': _safe_str(gp, 'Classification'),
            'G+_Control_Mean': _safe(gp, 'Control_Mean'),
            'G+_Threshold': _safe(gp, 'Threshold'),
            # G-
            'G-_N': _safe(gm, 'N', "{:.0f}"),
            'G-_Mean': _safe(gm, 'Mean'),
            'G-_Std': _safe(gm, 'Std_Dev'),
            'G-_SEM': _safe(gm, 'SEM'),
            'G-_Median': _safe(gm, 'Median'),
            'G-_CI': f"{_safe(gm, 'CI_Lower')} – {_safe(gm, 'CI_Upper')}",
            'G-_Cohens_d': _safe(gm, 'Cohens_d', "{:.3f}"),
            'G-_d_CI': f"{_safe(gm, 'd_CI_Lower', '{:.3f}')} – {_safe(gm, 'd_CI_Upper', '{:.3f}')}",
            'G-_Effect_Size': _safe_str(gm, 'Effect_Size_Label'),
            'G-_P_Value': _safe(gm, 'P_Value', "{:.4f}"),
            'G-_Significance': _safe_str(gm, 'Significance'),
            'G-_Detection': detection_labels.get(gm_norm, 'Unknown'),
            'G-_Confidence': _safe_str(gm, 'Classification_Confidence'),
            'G-_Strength': _safe_str(gm, 'Strength_of_Evidence'),
            'G-_Pct_Diff': _safe(gm, 'Pct_Diff_from_Control', "{:.1f}"),
            'G-_Classification': _safe_str(gm, 'Classification'),
            'G-_Control_Mean': _safe(gm, 'Control_Mean'),
            'G-_Threshold': _safe(gm, 'Threshold'),
            # Combined
            'Final_Classification': final_class,
            'Interpretation': interpretation,
        }
        results.append(row)

    final_df = pd.DataFrame(results)
    csv_path = output_root / "final_clinical_results.csv"
    final_df.to_csv(csv_path, index=False)

    # ── Excel ──
    excel_path = output_root / "final_clinical_results.xlsx"
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        fills = {
            'POSITIVE':             PatternFill("solid", fgColor="FFC7CE"),
            'NEGATIVE':             PatternFill("solid", fgColor="C6EFCE"),
            'NO OBVIOUS BACTERIA':  PatternFill("solid", fgColor="FFEB9C"),
            'MIXED/CONTRADICTORY':  PatternFill("solid", fgColor="FCD5B4"),
            'CONTROL (Reference)':  PatternFill("solid", fgColor="E7E6E6"),
        }
        default_fill = PatternFill("solid", fgColor="D9D9D9")
        header_fill  = PatternFill("solid", fgColor="4472C4")
        header_font  = Font(bold=True, color="FFFFFF", size=10)
        center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side    = Side(style='thin')
        thin_border  = Border(left=thin_side, right=thin_side,
                              top=thin_side, bottom=thin_side)

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: Final Results
            final_df.to_excel(writer, sheet_name='Final Results', index=False)
            ws = writer.sheets['Final Results']
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = thin_border
            for row_idx in range(len(final_df)):
                fc = final_df.iloc[row_idx]['Final_Classification']
                fill = fills.get(fc, default_fill)
                for col_idx in range(1, len(final_df.columns) + 1):
                    cell = ws.cell(row=row_idx + 2, column=col_idx)
                    cell.fill = fill
                    cell.alignment = center
                    cell.border = thin_border
            for col in ws.columns:
                letter = get_column_letter(col[0].column)
                max_len = max(len(str(c.value or '')) for c in col)
                ws.column_dimensions[letter].width = min(max_len + 3, 28)

            # Sheet 2: Group Detail (one row per group × channel)
            detail_rows = []
            for group in all_groups:
                if group == 'Control':
                    continue
                for ch_label, ch_dict in [('G+', gplus_dict.get(group, {})),
                                           ('G-', gminus_dict.get(group, {}))]:
                    if not ch_dict:
                        continue
                    detail_rows.append({
                        'Group': group,
                        'Channel': ch_label,
                        'N': ch_dict.get('N', '—'),
                        'Mean': ch_dict.get('Mean', '—'),
                        'SD': ch_dict.get('Std_Dev', '—'),
                        'SEM': ch_dict.get('SEM', '—'),
                        'Median': ch_dict.get('Median', '—'),
                        '95% CI': f"{_safe(ch_dict, 'CI_Lower')} – {_safe(ch_dict, 'CI_Upper')}",
                        "Cohen's d": ch_dict.get('Cohens_d', '—'),
                        'd 95% CI': f"{_safe(ch_dict, 'd_CI_Lower', '{:.3f}')} – {_safe(ch_dict, 'd_CI_Upper', '{:.3f}')}",
                        'Effect Size': ch_dict.get('Effect_Size_Label', '—'),
                        'p-value': ch_dict.get('P_Value', '—'),
                        'Significance': ch_dict.get('Significance', '—'),
                        '% Diff from Control': ch_dict.get('Pct_Diff_from_Control', '—'),
                        'Detection': ch_dict.get('Classification', '—'),
                        'Confidence': ch_dict.get('Classification_Confidence', '—'),
                        'Strength': ch_dict.get('Strength_of_Evidence', '—'),
                    })
            if detail_rows:
                detail_df = pd.DataFrame(detail_rows)
                detail_df.to_excel(writer, sheet_name='Group Detail', index=False)
                ws_d = writer.sheets['Group Detail']
                for cell in ws_d[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                for col in ws_d.columns:
                    letter = get_column_letter(col[0].column)
                    max_len = max(len(str(c.value or '')) for c in col)
                    ws_d.column_dimensions[letter].width = min(max_len + 3, 28)

            # Sheet 3: Parameters
            params_data = {
                'Parameter': [
                    'Dataset', 'Date Generated', 'Processing Mode',
                    'Percentile Filter', 'Clinical Threshold Formula',
                    'G+ Control Mean', 'G+ Control SD', 'G+ Threshold',
                    'G- Control Mean', 'G- Control SD', 'G- Threshold',
                    'Number of Test Groups',
                ],
                'Value': [
                    dataset_base_name,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'Batch (G+ and G-)',
                    f"{gplus_classification.iloc[0].get('N', '?')} (see individual files)",
                    'Control Mean × (1 − threshold%)',
                    _safe(gplus_dict.get('Control', {}), 'Mean'),
                    _safe(gplus_dict.get('Control', {}), 'Std_Dev'),
                    _safe(gplus_dict.get('Control', {}), 'Threshold'),
                    _safe(gminus_dict.get('Control', {}), 'Mean'),
                    _safe(gminus_dict.get('Control', {}), 'Std_Dev'),
                    _safe(gminus_dict.get('Control', {}), 'Threshold'),
                    str(len([g for g in all_groups if g != 'Control'])),
                ],
            }
            pd.DataFrame(params_data).to_excel(writer, sheet_name='Parameters', index=False)

            # Sheet 4: Legend
            legend_data = pd.DataFrame({
                'Classification': ['POSITIVE', 'NEGATIVE', 'NO OBVIOUS BACTERIA',
                                   'MIXED/CONTRADICTORY', 'CONTROL (Reference)', 'MISSING DATA'],
                'Meaning': [
                    'Gram-positive bacteria detected (G+ below threshold, G- above)',
                    'Gram-negative bacteria detected (G- below threshold, G+ above)',
                    'No obvious bacteria in either channel',
                    'Contradictory G+ and G- results — manual review required',
                    'Control group (reference baseline)',
                    'Insufficient data for classification'],
            })
            legend_data.to_excel(writer, sheet_name='Legend', index=False)

            # Sheet 5: Significance Key
            sig_key = pd.DataFrame({
                'Symbol': ['***', '**', '*', 'ns'],
                'Criterion': ['p < 0.001', 'p < 0.01', 'p < 0.05', 'p ≥ 0.05'],
                'Interpretation': ['Highly significant', 'Very significant',
                                   'Significant', 'Not significant'],
            })
            effect_key = pd.DataFrame({
                "Cohen's d Range": ['|d| < 0.2', '0.2–0.5', '0.5–0.8', '|d| ≥ 0.8'],
                'Label': ['Negligible', 'Small', 'Medium', 'Large'],
            })
            strength_key = pd.DataFrame({
                'Level': ['Strong', 'Moderate', 'Weak', 'Insufficient'],
                'Criteria': [
                    'p < 0.01 AND |d| ≥ 0.8 AND n ≥ 5',
                    'p < 0.05 AND |d| ≥ 0.5 AND n ≥ 3',
                    'p < 0.05 OR |d| ≥ 0.5',
                    'Neither criterion met or n < 2',
                ],
            })
            sig_key.to_excel(writer, sheet_name='Significance Key', index=False, startrow=0)
            effect_key.to_excel(writer, sheet_name='Significance Key', index=False,
                                startrow=len(sig_key) + 3)
            strength_key.to_excel(writer, sheet_name='Significance Key', index=False,
                                  startrow=len(sig_key) + len(effect_key) + 7)

        print(f"  ✓ Final results Excel: {excel_path.name}")

    except Exception as e:
        print(f"[ERROR] Could not create final Excel: {e}")
        import traceback; traceback.print_exc()

    # Console summary
    print("\n" + "=" * 100)
    print("FINAL CLINICAL RESULTS SUMMARY")
    print("=" * 100)
    display_cols = [
        'Group', 'G+_N', 'G+_Mean', 'G+_Std', 'G+_CI', 'G+_Detection',
        'G+_P_Value', 'G+_Cohens_d',
        'G-_N', 'G-_Mean', 'G-_Std', 'G-_CI', 'G-_Detection',
        'G-_P_Value', 'G-_Cohens_d',
        'Final_Classification',
    ]
    available = [c for c in display_cols if c in final_df.columns]
    print(final_df[available].to_string(index=False))
    print("=" * 100 + "\n")

    return excel_path


# ==================================================
# PDF Laboratory Report Generator
# ==================================================

from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.patches as mpatches


def generate_laboratory_report_pdf(
    output_root: Path,
    config: dict,
    gplus_classification: pd.DataFrame,
    gminus_classification: pd.DataFrame,
    final_df: pd.DataFrame,
) -> Optional[Path]:
    """Generate 5-page A4 PDF with forest plot and decision heatmap."""

    pdf_path = output_root / "laboratory_report.pdf"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    dataset_id = config.get('dataset_id_base', config.get('dataset_id', 'Unknown'))

    HEADER_BG  = "#1B3A5C"
    ACCENT     = "#2E75B6"
    LIGHT_GRAY = "#F2F2F2"
    A4_W, A4_H = 8.27, 11.69

    RESULT_COLOURS = {
        'POSITIVE':             ("#FFC7CE", "#9C0006"),
        'NEGATIVE':             ("#C6EFCE", "#006100"),
        'NO OBVIOUS BACTERIA':  ("#FFEB9C", "#9C6500"),
        'MIXED/CONTRADICTORY':  ("#FCD5B4", "#974706"),
        'CONTROL (Reference)':  ("#E7E6E6", "#333333"),
        'MISSING DATA':         ("#D9D9D9", "#666666"),
    }

    def _make_ax(fig):
        ax = fig.add_axes([0, 0, 1, 1]); ax.set_xlim(0, 1); ax.set_ylim(0, 1); ax.axis('off')
        return ax

    def _draw_header(ax, subtitle=""):
        ax.fill_between([0, 1], 0.92, 1.0, color=HEADER_BG)
        ax.text(0.05, 0.965, "MICROGEL FLUORESCENCE ANALYSIS",
                fontsize=16, fontweight='bold', color='white', va='center', family='sans-serif')
        ax.text(0.05, 0.935, "Laboratory Report",
                fontsize=11, color='#A0C4E8', va='center', family='sans-serif')
        if subtitle:
            ax.text(0.95, 0.935, subtitle, fontsize=9, color='#A0C4E8',
                    va='center', ha='right', family='sans-serif')
        ax.fill_between([0, 1], 0.915, 0.92, color=ACCENT)

    def _draw_footer(ax, page_num, total_pages):
        ax.plot([0.05, 0.95], [0.04, 0.04], color='#CCCCCC', linewidth=0.5)
        ax.text(0.05, 0.025, f"Generated: {timestamp}", fontsize=7, color='#999999', va='center')
        ax.text(0.5, 0.025, f"Page {page_num} of {total_pages}",
                fontsize=7, color='#999999', va='center', ha='center')
        ax.text(0.95, 0.025, "CONFIDENTIAL", fontsize=7, color='#CC0000',
                va='center', ha='right', fontweight='bold')

    try:
        with PdfPages(str(pdf_path)) as pdf:
            total_pages = 5

            # ============ PAGE 1 — Summary ============
            fig1 = plt.figure(figsize=(A4_W, A4_H))
            ax = _make_ax(fig1); _draw_header(ax, "Summary"); _draw_footer(ax, 1, total_pages)

            ax.fill_between([0.05, 0.95], 0.82, 0.90, color=LIGHT_GRAY)
            ax.plot([0.05, 0.95, 0.95, 0.05, 0.05],
                    [0.90, 0.90, 0.82, 0.82, 0.90], color='#999999', lw=0.5)
            info_items = [
                ("Sample ID:", dataset_id),
                ("Date:", timestamp),
                ("Bacteria Config:", config.get('bacteria_config_info', {}).get(
                    'config_names', {}).get(config.get('bacteria_type', ''), 'Multi-scan / Auto')),
                ("Percentile Filter:", f"{config.get('percentile', 0.3)*100:.0f}%"),
                ("Clinical Threshold:", f"{config.get('threshold_pct', 0.05)*100:.1f}%"),
            ]
            for i, (label, value) in enumerate(info_items):
                col = 0.07 if i < 3 else 0.52
                row_y = 0.885 - (i % 3) * 0.022
                ax.text(col, row_y, label, fontsize=8, fontweight='bold', va='center', color='#333333')
                ax.text(col + 0.12, row_y, str(value), fontsize=8, va='center', color='#333333')

            ax.text(0.05, 0.79, "FINAL RESULTS", fontsize=13, fontweight='bold', color=HEADER_BG, va='center')
            ax.plot([0.05, 0.95], [0.78, 0.78], color=ACCENT, lw=1.5)

            col_headers = ['Group', 'G+ Mean±SD', 'G+ Detection', 'G− Mean±SD', 'G− Detection', 'Final']
            col_widths = [0.10, 0.18, 0.14, 0.18, 0.14, 0.18]
            col_x = [0.05]
            for w in col_widths[:-1]:
                col_x.append(col_x[-1] + w)
            y_top = 0.76; row_h = 0.032

            for j, header in enumerate(col_headers):
                ax.fill_between([col_x[j], col_x[j]+col_widths[j]], y_top-row_h, y_top, color=HEADER_BG)
                ax.text(col_x[j]+col_widths[j]/2, y_top-row_h/2, header,
                        fontsize=7.5, fontweight='bold', color='white', va='center', ha='center')

            for i, (_, row) in enumerate(final_df.iterrows()):
                y = y_top - (i+1)*row_h
                fc = row.get('Final_Classification', '')
                bg, tc = RESULT_COLOURS.get(fc, ('#FFFFFF', '#333333'))
                gp_str = f"{row.get('G+_Mean','—')}±{row.get('G+_Std','—')}" if row.get('G+_Mean','—') != '—' else '—'
                gm_str = f"{row.get('G-_Mean','—')}±{row.get('G-_Std','—')}" if row.get('G-_Mean','—') != '—' else '—'
                cells = [str(row.get('Group','')), gp_str, str(row.get('G+_Detection','')),
                         gm_str, str(row.get('G-_Detection','')), fc]
                for j, ct in enumerate(cells):
                    ax.fill_between([col_x[j], col_x[j]+col_widths[j]], y-row_h, y,
                                    color=bg if j==len(cells)-1 else LIGHT_GRAY)
                    ax.text(col_x[j]+col_widths[j]/2, y-row_h/2, ct, fontsize=7,
                            va='center', ha='center', color=tc if j==len(cells)-1 else '#333333')

            # Interpretation
            box_y = y_top - (len(final_df)+2)*row_h
            ax.text(0.05, box_y, "INTERPRETATION", fontsize=11, fontweight='bold', color=HEADER_BG)
            ax.plot([0.05, 0.95], [box_y-0.01, box_y-0.01], color=ACCENT, lw=1)
            interp_y = box_y - 0.04
            for _, row in final_df.iterrows():
                grp = row.get('Group', ''); fc = row.get('Final_Classification', '')
                if grp == 'Control':
                    continue
                symbol_map = {'NEGATIVE': ("●","#006100"), 'POSITIVE': ("●","#9C0006"),
                              'NO OBVIOUS BACTERIA': ("○","#9C6500"), 'MIXED/CONTRADICTORY': ("◆","#974706")}
                sym, clr = symbol_map.get(fc, ("?","#666666"))
                interp_text = str(row.get('Interpretation', fc))
                # Split interpretation into main text and p-value annotations
                main_part = interp_text.split('[')[0].strip()
                pval_parts = re.findall(r'\[.*?\]', interp_text)
                pval_note = "  ".join(pval_parts)

                ax.text(0.07, interp_y, sym, fontsize=10, color=clr, va='center', ha='center')
                ax.text(0.09, interp_y, f"Group {grp}: {main_part}", fontsize=8.5,
                        color='#333333', va='center', fontweight='bold')
                if pval_note:
                    ax.text(0.09, interp_y-0.018, pval_note, fontsize=7, color='#666666', va='center')
                    interp_y -= 0.045
                else:
                    interp_y -= 0.030

            pdf.savefig(fig1); plt.close(fig1)

            # ============ PAGE 2 — Charts ============
            fig2 = plt.figure(figsize=(A4_W, A4_H))
            tpct = config.get('threshold_pct', 0.05)
            ax_gp = fig2.add_axes((0.10, 0.53, 0.82, 0.35))
            _draw_chart_on_axis(ax_gp, gplus_classification, "G+ Microgel (Positive)", ACCENT,
                                output_dir=config.get('positive_output'), threshold_pct=tpct)
            ax_gm = fig2.add_axes((0.10, 0.08, 0.82, 0.35))
            _draw_chart_on_axis(ax_gm, gminus_classification, "G− Microgel (Negative)", "#C0504D",
                                output_dir=config.get('negative_output'), threshold_pct=tpct)
            ax_o = _make_ax(fig2); _draw_header(ax_o, "Comparison Charts"); _draw_footer(ax_o, 2, total_pages)
            pdf.savefig(fig2); plt.close(fig2)

            # ============ PAGE 3 — Forest Plot + Decision Heatmap ============
            fig3 = plt.figure(figsize=(A4_W, A4_H))
            ax_forest = fig3.add_axes((0.12, 0.52, 0.78, 0.36))
            _draw_forest_plot_on_axis(ax_forest, gplus_classification, gminus_classification)
            ax_heat = fig3.add_axes((0.12, 0.10, 0.78, 0.30))
            _draw_decision_heatmap_on_axis(ax_heat, final_df, gplus_classification, gminus_classification)
            ax_o3 = _make_ax(fig3); _draw_header(ax_o3, "Effect Size & Decision Matrix"); _draw_footer(ax_o3, 3, total_pages)
            pdf.savefig(fig3); plt.close(fig3)

            # ============ PAGE 4 — Statistical Tables ============
            fig4 = plt.figure(figsize=(A4_W, A4_H))
            ax = _make_ax(fig4)
            _draw_header(ax, "Statistical Analysis")
            _draw_footer(ax, 4, total_pages)

            # ── Fixed layout constants ─────────────────────────────────────────
            # Page content area: y = 0.06 (footer top) to y = 0.88 (header bottom)
            # Divide into three non-overlapping bands:
            #   G+  table  : 0.88 → TABLE_MID   (top half)
            #   G-  table  : TABLE_MID → SIG_BOT (bottom half)
            #   Sig Key     : SIG_KEY_Y  (reserved band, always fixed)
            _TABLE_MID  = 0.54      # G+ stops here; G- starts here
            _SIG_KEY_Y  = 0.22      # Significance Key header – always at this y
            _SIG_GUARD  = _SIG_KEY_Y + 0.04   # G- hard stop: 4 % gap above sig key

            # ── G+ table – occupies the TOP half of the content area ──
            _draw_stats_table(
                ax, gplus_classification, "G+ Microgel",
                y_top=0.87,
                y_bottom_limit=_TABLE_MID + 0.01,   # must not cross into G- zone
            )

            # ── G- table – occupies the BOTTOM half, above the Significance Key ──
            _draw_stats_table(
                ax, gminus_classification, "G− Microgel",
                y_top=_TABLE_MID - 0.01,
                y_bottom_limit=_SIG_GUARD,          # hard stop before sig key
            )

            # ── Significance Key – FIXED position, never moved by table content ──
            ax.plot([0.05, 0.95], [_SIG_KEY_Y + 0.008, _SIG_KEY_Y + 0.008],
                    color='#CCCCCC', lw=0.5)
            ax.text(0.05, _SIG_KEY_Y, "Significance Key",
                    fontsize=9, fontweight='bold', color=HEADER_BG)

            key_lines = [
                "*** p < 0.001 (highly significant)",
                "**  p < 0.01  (very significant)",
                "*   p < 0.05  (significant)",
                "ns  p ≥ 0.05  (not significant)",
                "Effect size: |d| < 0.2 negligible, 0.2–0.5 small, "
                "0.5–0.8 medium, > 0.8 large (Cohen's d)",
            ]

            for i, line in enumerate(key_lines):
                ax.text(0.07, _SIG_KEY_Y - 0.022 - i * 0.018,
                        line, fontsize=7, color='#555555', family='monospace')

            pdf.savefig(fig4)
            plt.close(fig4)

            # ============ PAGE 5 — Methodology ============
            fig5 = plt.figure(figsize=(A4_W, A4_H))
            ax = _make_ax(fig5); _draw_header(ax, "Methodology & Quality Control"); _draw_footer(ax, 5, total_pages)
            sections = [
                ("METHODOLOGY", 0.88, [
                    "Brightfield and fluorescence images were acquired using a Leica microscope.",
                    "Particles were segmented from brightfield (ch00) using Gaussian blur and intensity thresholding.",
                    "Fluorescence intensity (ch01) was measured within brightfield contours.",
                    "The primary metric is Fluorescence Integrated Density / BF Area (a.u./µm²).",
                    f"Typical particles were selected as the middle {100-2*int(config.get('percentile',0.3)*100)}% "
                    f"(top/bottom {int(config.get('percentile',0.3)*100)}% excluded).",
                    f"Clinical threshold = Control Mean × (1 − {config.get('threshold_pct',0.05)*100:.0f}%).",
                    "Groups whose typical-particle mean falls below the threshold are classified as 'Bacteria Detected'.",
                    "Cohen's d effect size computed with pooled SD; 95% CI via Hedges–Olkin approximation.",
                ]),
                ("QUALITY CONTROL", 0.56, [
                    f"Pixel size derived from Leica XML metadata (fallback: {FALLBACK_UM_PER_PX} µm/px).",
                    "Alignment: phase cross-correlation; fallback to no-shift if correlation error > 0.5.",
                    "Excluded objects logged in each group's master Excel (Excluded_Objects sheet).",
                ]),
                ("LIMITATIONS", 0.43, [
                    "This assay is for research use only.",
                    "Results should be confirmed by culture-based methods.",
                    "Classification depends on threshold chosen and control baseline.",
                    "Low particle counts (N < 5) may yield unreliable statistics.",
                ]),
            ]
            for title_t, ys, lines in sections:
                ax.text(0.05, ys, title_t, fontsize=11, fontweight='bold', color=HEADER_BG)
                ax.plot([0.05, 0.95], [ys-0.01, ys-0.01], color=ACCENT, lw=1)
                for i, line in enumerate(lines):
                    ax.text(0.07, ys-0.03-i*0.02, f"• {line}", fontsize=7.5, color='#333333', wrap=True)
            sig_y = 0.15
            ax.plot([0.05, 0.95], [sig_y+0.02, sig_y+0.02], color='#CCCCCC', lw=0.5)
            ax.text(0.05, sig_y, "APPROVAL", fontsize=10, fontweight='bold', color=HEADER_BG)
            for label, xp in [("Performed by:", 0.05), ("Reviewed by:", 0.38), ("Date:", 0.71)]:
                ax.text(xp, sig_y-0.03, label, fontsize=8, fontweight='bold', color='#555555')
                ax.plot([xp, xp+0.25], [sig_y-0.07, sig_y-0.07], color='#333333', lw=0.8)
            pdf.savefig(fig5); plt.close(fig5)

        print(f"  ✓ Laboratory report PDF: {pdf_path.name}")
        return pdf_path

    except Exception as e:
        print(f"  ✗ PDF generation failed: {e}")
        import traceback; traceback.print_exc()
        return None


# ── Chart drawing helper ──

def _draw_chart_on_axis(
    ax: MplAxes,
    classification_df: pd.DataFrame,
    title: str,
    bar_colour: str,
    output_dir: Optional[Path] = None,
    threshold_pct: float = 0.05,
) -> None:
    """Draw a bar chart matching the standalone comparison plots:
    palette bars, SD error bars, jitter dots, threshold lines, significance."""

    if classification_df.empty:
        ax.text(0.5, 0.5, "No data", ha='center', va='center')
        return

    df = classification_df.copy()
    df['sort_key'] = df['Group'].apply(
        lambda x: (1, 999) if x == 'Control' else (0, int(x) if x.isdigit() else 999)
    )
    df = df.sort_values('sort_key').reset_index(drop=True)

    groups = df['Group'].tolist()
    
    # ✅ FIX: Create consistent numpy arrays with explicit dtype
    means = np.array(df['Mean'].astype(float).tolist(), dtype=float)
    stds = np.array(df['Std_Dev'].astype(float).tolist(), dtype=float)
    x = np.arange(len(groups), dtype=float)

    # ── Load individual data points from master Excel files ──
    individual_data: dict[str, list[float]] = {}
    if output_dir is not None and output_dir.exists():
        for group_name in groups:
            master_path: Optional[Path] = None
            typical_sheet: Optional[str] = None

            if group_name == 'Control':
                for folder in output_dir.iterdir():
                    if folder.is_dir() and folder.name.lower().startswith('control'):
                        master_path = folder / f"{folder.name}_master.xlsx"
                        typical_sheet = f"{folder.name}_Typical_Particles"
                        break
            else:
                folder = output_dir / group_name
                master_path = folder / f"{group_name}_master.xlsx"
                typical_sheet = f"{group_name}_Typical_Particles"

            if master_path is not None and master_path.exists() and typical_sheet is not None:
                try:
                    tp_df = pd.read_excel(master_path, sheet_name=typical_sheet)
                    if 'Fluor_Density_per_BF_Area' in tp_df.columns:
                        vals = pd.to_numeric(
                            tp_df['Fluor_Density_per_BF_Area'], errors='coerce'
                        ).dropna().tolist()
                        if vals:
                            individual_data[group_name] = vals
                except Exception:
                    pass

    # ── Colour palette (match standalone) ──
    n_test = sum(1 for g in groups if g != 'Control')
    palette = list(sns.color_palette("husl", max(n_test, 1)))
    colours: list[Any] = []
    palette_idx = 0
    for g in groups:
        if g == 'Control':
            colours.append('#A0A0A0')
        else:
            colours.append(palette[palette_idx % len(palette)])
            palette_idx += 1

    # ── Bars ──
    bars = ax.bar(x, means, color=colours, edgecolor='black',
                  linewidth=0.8, alpha=0.7, width=0.6)

    # ── SD error bars (thick cap for control) ──
    # ✅ FIX: Explicit float conversion in loop
    for xi, mean_val, sd_val, group_name in zip(x, means, stds, groups):
        cap = 14 if group_name == 'Control' else 7
        ax.errorbar(
            float(xi),      # ✅ Convert to float
            float(mean_val), # ✅ Convert to float
            yerr=float(sd_val),  # ✅ Convert to float
            fmt='none', 
            ecolor='black',
            elinewidth=1.5, 
            capsize=cap, 
            capthick=1.5, 
            zorder=10
        )

    # ── Jitter dots ──
    if individual_data:
        rng = np.random.RandomState(42)
        for xi, g in zip(x, groups):
            if g in individual_data:
                vals = individual_data[g]
                jitter = rng.uniform(-0.15, 0.15, size=len(vals))
                ax.scatter(
                    float(xi) + jitter,  # ✅ Convert to float
                    vals,
                    color='cyan', 
                    edgecolor='black', 
                    linewidth=0.5,
                    s=30, 
                    alpha=0.6, 
                    zorder=11,
                )

    # ── Threshold & control-mean lines ──
    ctrl_row = df[df['Group'] == 'Control']
    if not ctrl_row.empty:
        ctrl_mean = float(ctrl_row.iloc[0]['Mean'])
        threshold = float(
            ctrl_row.iloc[0].get('Threshold', ctrl_mean * (1 - threshold_pct))
        )
        ax.axhline(
            ctrl_mean, 
            color='blue', 
            ls=':', 
            lw=2.0,
            label=f'Control Mean ({ctrl_mean:.0f})'
        )
        ax.axhline(
            threshold, 
            color='red', 
            ls='--', 
            lw=2.0,
            label=f'Threshold ({threshold:.0f})'
        )
        ax.legend(fontsize=7, loc='upper right', framealpha=0.9)

    # ── Significance annotations ──
    y_max = float(np.max(means + stds)) if len(means) > 0 else 1.0
    for i, (_, row_data) in enumerate(df.iterrows()):
        if row_data['Group'] == 'Control':
            continue
        sig = str(row_data.get('Significance', ''))
        if sig and sig not in ('—', 'N/A (n<2)'):
            y_pos = float(means[i]) + float(stds[i]) + y_max * 0.03
            ax.text(
                float(x[i]),  # ✅ Convert to float
                y_pos, 
                sig, 
                ha='center', 
                va='bottom',
                fontsize=9, 
                fontweight='bold', 
                color='#333333'
            )

    # ── N labels beneath bars ──
    for i, (_, row_data) in enumerate(df.iterrows()):
        n_val = row_data.get('N', '')
        ax.text(
            float(x[i]),  # ✅ Convert to float
            -y_max * 0.04, 
            f"n={n_val}", 
            ha='center',
            va='top', 
            fontsize=7, 
            color='#666666'
        )

    ax.set_xticks(x)
    ax.set_xticklabels(groups, fontsize=8, fontweight='bold')
    ax.set_ylabel("Fluor Density (a.u./µm²)", fontsize=8)
    ax.set_title(title, fontsize=10, fontweight='bold', pad=8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='y', alpha=0.2, ls='--')


def _draw_forest_plot_on_axis(
    ax: MplAxes,
    gplus_df: pd.DataFrame,
    gminus_df: pd.DataFrame,
) -> None:
    """Draw a Cohen's d forest plot with 95% CI for every group × channel."""

    rows: list[dict] = []
    for channel_label, cdf, colour in [
        ("G+", gplus_df, "#2E75B6"),
        ("G−", gminus_df, "#C0504D"),
    ]:
        df = cdf.copy()
        df['sort_key'] = df['Group'].apply(
            lambda x: (1, 999) if x == 'Control' else (0, int(x) if x.isdigit() else 999)
        )
        df = df.sort_values('sort_key').reset_index(drop=True)

        for _, r in df.iterrows():
            if r['Group'] == 'Control':
                continue
            d_val = r.get('Cohens_d', 0.0)
            d_lo  = r.get('d_CI_Lower', d_val)
            d_hi  = r.get('d_CI_Upper', d_val)
            if pd.isna(d_val):
                continue
            rows.append({
                'label': f"Grp {r['Group']} {channel_label}",
                'd': float(d_val),
                'lo': float(d_lo) if not pd.isna(d_lo) else float(d_val),
                'hi': float(d_hi) if not pd.isna(d_hi) else float(d_val),
                'colour': colour,
                'sig': str(r.get('Significance', '')),
            })

    if not rows:
        ax.text(0.5, 0.5, "No effect-size data", ha='center', va='center', fontsize=10)
        return

    rows = rows[::-1]  # bottom-to-top
    y_pos = np.arange(len(rows))

    for i, r in enumerate(rows):
        ax.plot([r['lo'], r['hi']], [i, i], color=r['colour'], lw=2.5, solid_capstyle='round')
        marker = 'D' if abs(r['d']) >= 0.8 else 'o'
        ax.plot(r['d'], i, marker, color=r['colour'], markersize=7, markeredgecolor='black',
                markeredgewidth=0.8, zorder=10)
        # significance annotation
        sig = r['sig']
        if sig and sig not in ('—', 'ns', 'N/A (n<2)'):
            ax.text(r['hi'] + 0.15, i, sig, fontsize=8, fontweight='bold',
                    va='center', color='#006100')
        elif sig == 'ns':
            ax.text(r['hi'] + 0.15, i, 'ns', fontsize=7, va='center', color='#999999')

    ax.axvline(0, color='black', ls='--', lw=1, zorder=0)

    # Shade negligible zone
    ax.axvspan(-0.2, 0.2, color='#E8E8E8', alpha=0.5, zorder=0, label='Negligible (|d|<0.2)')

    ax.set_yticks(y_pos)
    ax.set_yticklabels([r['label'] for r in rows], fontsize=7.5)
    ax.set_xlabel("Cohen's d (effect size vs Control)", fontsize=8, fontweight='bold')
    ax.set_title("Effect Size Forest Plot (95% CI)", fontsize=10, fontweight='bold', pad=8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='x', alpha=0.2, ls='--')

    # Legend
    from matplotlib.lines import Line2D
    handles = [
        Line2D([0], [0], color='#2E75B6', lw=2.5, label='G+ channel'),
        Line2D([0], [0], color='#C0504D', lw=2.5, label='G− channel'),
        Line2D([0], [0], color='black', ls='--', lw=1, label='No effect (d=0)'),
    ]
    ax.legend(handles=handles, fontsize=7, loc='lower right', framealpha=0.9)


def _draw_decision_heatmap_on_axis(
    ax: MplAxes,
    final_df: pd.DataFrame,
    gplus_df: pd.DataFrame,
    gminus_df: pd.DataFrame,
) -> None:
    """Draw a coloured decision-matrix heatmap: rows = groups, cols = G+ / G- / Final."""
    
    test_groups = [str(g) for g in final_df['Group'] if str(g) != 'Control']

    if not test_groups:
        ax.text(0.5, 0.5, "No test groups", ha='center', va='center')
        return

    col_labels = ['G+ Detection', 'G− Detection', 'Final Classification',
                  'G+ Strength', 'G− Strength']
    n_rows = len(test_groups)
    n_cols = len(col_labels)

    colour_map = {
        'Detected':             '#FFC7CE',
        'Not Detected':         '#C6EFCE',
        'Control':              '#E7E6E6',
        'POSITIVE':             '#FFC7CE',
        'NEGATIVE':             '#C6EFCE',
        'NO OBVIOUS BACTERIA':  '#FFEB9C',
        'MIXED/CONTRADICTORY':  '#FCD5B4',
        'CONTROL (Reference)':  '#E7E6E6',
        'Strong':               '#C6EFCE',
        'Moderate':             '#FFEB9C',
        'Weak':                 '#FCD5B4',
        'Insufficient':         '#E7E6E6',
        'Insufficient (n<2)':   '#E7E6E6',
    }

    # Build data grid
    grid_text = []
    grid_colour = []
    for grp in test_groups:
        row_match = final_df[final_df['Group'] == grp]
        if row_match.empty:
            grid_text.append(['—'] * n_cols)
            grid_colour.append(['#FFFFFF'] * n_cols)
            continue
        r = row_match.iloc[0]
        vals = [
            str(r.get('G+_Detection', '—')),
            str(r.get('G-_Detection', '—')),
            str(r.get('Final_Classification', '—')),
            str(r.get('G+_Strength', '—')),
            str(r.get('G-_Strength', '—')),
        ]
        grid_text.append(vals)
        grid_colour.append([colour_map.get(v, '#FFFFFF') for v in vals])

    # Draw grid
    cell_h = 0.8 / max(n_rows, 1)
    cell_w = 0.85 / n_cols
    x0, y0 = 0.05, 0.15

    # Column headers
    for j, label in enumerate(col_labels):
        cx = x0 + j * cell_w + cell_w / 2
        cy = y0 + n_rows * cell_h + cell_h * 0.4
        ax.text(cx, cy, label, fontsize=7.5, fontweight='bold', ha='center', va='center',
                color='white',
                bbox=dict(boxstyle='round,pad=0.3', facecolor='#1B3A5C', edgecolor='none'))

    # Row headers
    for i, grp in enumerate(test_groups):
        ry = y0 + (n_rows - 1 - i) * cell_h + cell_h / 2
        ax.text(x0 - 0.02, ry, f"Grp {grp}", fontsize=8, fontweight='bold',
                ha='right', va='center', color='#333333')

    # Cells
    for i in range(n_rows):
        for j in range(n_cols):
            cx = x0 + j * cell_w
            cy = y0 + (n_rows - 1 - i) * cell_h
            rect = mpatches.Rectangle((cx, cy), cell_w, cell_h,
                                  facecolor=grid_colour[i][j],
                                  edgecolor='#CCCCCC', linewidth=0.5)
            ax.add_patch(rect)
            ax.text(cx + cell_w / 2, cy + cell_h / 2,
                    grid_text[i][j], fontsize=7, ha='center', va='center',
                    color='#333333', fontweight='bold' if j == 2 else 'normal')

    ax.set_xlim(-0.05, 1.05)
    ax.set_ylim(0, y0 + (n_rows + 1) * cell_h + 0.05)
    ax.axis('off')
    ax.set_title("Decision Matrix", fontsize=10, fontweight='bold', pad=10, color='#1B3A5C')




# ── Statistics table helper ──

def _draw_stats_table(
    ax: MplAxes,
    classification_df: pd.DataFrame,
    title: str,
    y_top: float,
    y_bottom_limit: float = 0.05,
) -> float:
    """Draw a formatted statistics table on the given Axes.

    Rows whose centre would fall below *y_bottom_limit* are silently omitted
    so the table never overflows into whatever content lives below it.

    Returns:
        float: y-coordinate just below the last drawn row (for chaining).
    """
    ax.text(0.05, y_top, title, fontsize=10, fontweight='bold', color="#1B3A5C")
    ax.plot([0.05, 0.95], [y_top - 0.01, y_top - 0.01], color="#2E75B6", lw=1)

    headers = ['Group', 'N', 'Mean', 'SD', '95% CI', "Cohen's d",
               'p-value', 'Sig', 'Confidence']
    col_x = [0.05, 0.13, 0.19, 0.28, 0.36, 0.52, 0.62, 0.73, 0.80]
    row_h = 0.022
    y = y_top - 0.035

    # Header row — only draw if it fits above the limit
    if y - row_h / 2 >= y_bottom_limit:
        for j, h in enumerate(headers):
            ax.text(col_x[j], y, h, fontsize=6.5, fontweight='bold',
                    color='white', va='center')
        ax.fill_between([0.05, 0.95],
                        y - row_h / 2, y + row_h / 2, color="#1B3A5C")

    # Sort rows
    df = classification_df.copy()
    df['sort_key'] = df['Group'].apply(
        lambda x: (1, 999) if x == 'Control'
                  else (0, int(x) if x.isdigit() else 999)
    )
    df = df.sort_values('sort_key').reset_index(drop=True)

    last_row_bottom: float = y - row_h / 2     # header bottom as initial value

    for i, (_, row) in enumerate(df.iterrows()):
        y_row = y - (i + 1) * row_h

        # Stop before crossing the hard limit
        if y_row - row_h / 2 < y_bottom_limit:
            break

        bg = "#F2F2F2" if i % 2 == 0 else "#FFFFFF"
        ax.fill_between([0.05, 0.95],
                        y_row - row_h / 2, y_row + row_h / 2, color=bg)

        ci_lo  = row.get('CI_Lower', '—')
        ci_hi  = row.get('CI_Upper', '—')
        ci_str = f"{ci_lo}–{ci_hi}" if ci_lo != '—' else '—'

        p_val = row.get('P_Value', np.nan)
        if pd.isna(p_val):
            p_str = '—'
        elif p_val < 0.001:
            p_str = '<0.001'
        else:
            p_str = f"{p_val:.4f}"

        cells = [
            str(row.get('Group', '')),
            str(row.get('N', '')),
            f"{float(row.get('Mean', 0)):.1f}",
            f"{float(row.get('Std_Dev', 0)):.1f}",
            ci_str,
            str(row.get('Cohens_d', '—')),
            p_str,
            str(row.get('Significance', '—')),
            str(row.get('Classification_Confidence', '—')),
        ]

        for j, cell in enumerate(cells):
            colour = '#333333'
            weight = 'normal'
            sig_val = str(row.get('Significance', ''))
            if j == 7 and sig_val.startswith('*'):
                colour = '#006100'
                weight = 'bold'
            elif j == 7 and sig_val == 'ns':
                colour = '#999999'

            ax.text(col_x[j], y_row, cell, fontsize=6, va='center',
                    color=colour, fontweight=weight)

        last_row_bottom = y_row - row_h / 2

    return float(last_row_bottom - 0.015)   # small gap below the last drawn row


# ==================================================
# Source Directory Selection
# ==================================================

def select_source_directory(max_depth=2) -> Optional[Path]:
    """Lists directories that either have a Control subfolder OR contain G+/G- subdirectories"""
    root_dir = Path('source')
    
    if not root_dir.exists():
        print(f"[ERROR] Source directory not found: {root_dir.resolve()}")
        return None
    
    valid_directories = []
    
    # Check immediate subdirectories of source/
    for item in root_dir.iterdir():
        if not item.is_dir():
            continue
            
        # Check if this directory has both G+ and G- subdirectories
        has_gplus = (item / 'G+').is_dir()
        has_gminus = (item / 'G-').is_dir()
        
        if has_gplus and has_gminus:
            # This is a batch directory (like "PD sample" or "Spike sample")
            valid_directories.append(item.name)
            continue
        
        # Check if this directory has a Control subfolder (single-mode directory)
        try:
            subdirs = [d for d in item.iterdir() if d.is_dir()]
            has_control = any(d.name.lower().startswith('control') for d in subdirs)
            if has_control:
                valid_directories.append(item.name)
        except OSError:
            continue
    
    if not valid_directories:
        print("[ERROR] No valid directories found.")
        print("Valid directories must either:")
        print("  1. Contain both 'G+' and 'G-' subfolders (for batch processing)")
        print("  2. Contain a 'Control' subfolder (for single processing)")
        return None
    
    valid_directories.sort()
    
    print("\n" + "="*80)
    print("SELECT SOURCE DIRECTORY")
    print("="*80)
    print("\nAvailable directories:")
    for i, dir_name in enumerate(valid_directories, 1):
        dir_path = root_dir / dir_name
        has_gplus = (dir_path / 'G+').is_dir()
        has_gminus = (dir_path / 'G-').is_dir()
        
        if has_gplus and has_gminus:
            mode_label = "[BATCH: G+ and G-]"
        else:
            mode_label = "[SINGLE]"
        
        print(f"  [{i}] {dir_name} {mode_label}")
    
    while True:
        selected = logged_input("\nEnter the number or folder name (or 'q' to quit): ").strip()
        
        if selected.lower() in {'q', 'quit', 'exit'}:
            raise SystemExit(0)
        
        if selected.isdigit():
            num = int(selected)
            if 1 <= num <= len(valid_directories):
                selected_name = valid_directories[num - 1]
                full_selected = root_dir / selected_name
                return full_selected
            else:
                print(f"Invalid number. Please enter between 1 and {len(valid_directories)}.")
        elif selected in valid_directories:
            full_selected = root_dir / selected
            return full_selected
        else:
            print("Invalid selection. Please enter a valid number or folder name.")


# ==================================================
# Image Processing
# ==================================================


def process_image(
    img_path: Path, 
    output_root: Path,
    bacteria_config: 'SegmentationConfig'
) -> None:
    """Process a single image - Unicode-safe version with bacteria-specific parameters
    
    Args:
        img_path: Path to input image
        output_root: Root output directory
        bacteria_config: Bacteria-specific segmentation configuration
    
    Raises:
        FileNotFoundError: If image cannot be read
        ValueError: If pixel size cannot be determined
    """
    
    # ========== VARIABLE INITIALIZATION ==========
    # Initialize all variables that need cleanup
    img = None
    img8 = None
    fluor_img = None
    fluor_img_aligned = None
    mask = None
    vis_all = None
    vis_acc = None
    vis_rej = None  # ✅ NEW: Visualization for rejected objects
    mask_all = None
    mask_acc = None
    mask_rej = None  # ✅ NEW: Mask for rejected objects
    fluor_img8 = None
    fluor_bw = None
    vis_fluor = None
    vis_match = None
    
    try:
        # ========== VALIDATION ==========
        # Validate path encoding
        if not validate_path_encoding(img_path):
            print(f"[ERROR] Cannot process image with problematic path: {img_path}")
            return
        
        # ✅ FIX: MOVE OBJECT ID GENERATION TO THE TOP
        # Generate object IDs early so they're available throughout
        parts = img_path.stem.split()
        group_id = parts[0] if parts else "unk"
        sequence_num = parts[-1].split("_")[0] if parts else "0"
        
        # ========== METADATA EXTRACTION ==========
        xml_props, xml_main = find_metadata_paths(img_path)
        
        try:
            um_per_px_x, um_per_px_y = get_pixel_size_um(xml_props, xml_main)
            um_per_px_x = float(um_per_px_x)
            um_per_px_y = float(um_per_px_y)
        except Exception as e:
            if FALLBACK_UM_PER_PX is None:
                raise
            print(f"[WARN] Using fallback pixel size for {img_path.name}: {e}")
            um_per_px_x = um_per_px_y = float(FALLBACK_UM_PER_PX)

        um_per_px_avg = (um_per_px_x + um_per_px_y) / 2.0
        um2_per_px2 = float(um_per_px_x) * float(um_per_px_y)

        # ========== OUTPUT DIRECTORY SETUP ==========
        img_out = output_root / img_path.stem
        ensure_dir(img_out)

        # ========== LOAD BRIGHTFIELD IMAGE ==========
        img = safe_imread(img_path, cv2.IMREAD_UNCHANGED)
        if img is None:
            raise FileNotFoundError(f"Could not read image: {img_path}")
        
        # Convert to grayscale if needed
        if img.ndim == 3:
            img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        img8 = normalize_to_8bit(img)
        save_debug(img_out, "01_gray_8bit.png", img8, um_per_px_avg)

        # ========== SEGMENT PARTICLES ==========
        mask = segment_particles_brightfield(img8, float(um_per_px_avg), img_out, bacteria_config)

        # Find contours
        _fc = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contours = cast(list[np.ndarray], _fc[-2])

        # Visualize all contours
        vis_all = cv2.cvtColor(img8, cv2.COLOR_GRAY2BGR)
        cv2.drawContours(vis_all, contours, -1, (0, 0, 255), 1)
        save_debug(img_out, "10_contours_all.png", vis_all, um_per_px_avg)

        # ========== FILTER PARTICLES (ENHANCED WITH REJECTION TRACKING) ==========
        min_area_px = bacteria_config.min_area_px
        max_area_px = bacteria_config.max_area_px

        H, W = img8.shape[:2]
        img_area_px = float(H * W)
        max_big_area_px = bacteria_config.max_fraction_of_image * img_area_px

        # ✅ FIX: Explicitly initialize lists with type hints
        accepted: list[np.ndarray] = []
        rejected: list[np.ndarray] = []
        rejection_reasons: list[dict] = []  # Track why each object was rejected
        
        # ✅ FIX: Handle empty contours case
        if len(contours) == 0:
            print(f"  [WARN] No contours found in {img_path.name}")
            # Continue with empty lists - the code will handle this gracefully

        for contour_idx, c in enumerate(contours, 1):
            area_px = float(cv2.contourArea(c))
            
            # ✅ NEW: Initialize rejection tracking for this contour
            rejection_info: dict = {
                'contour_index': contour_idx,
                'area_px': area_px,
                'reasons': []
            }
            
            if area_px <= 0:
                rejection_info['reasons'].append('Zero area')
                rejected.append(c)
                rejection_reasons.append(rejection_info)
                continue

            # Filter by absolute size
            if area_px >= max_big_area_px:
                rejection_info['reasons'].append(f'Too large (>{max_big_area_px:.1f} px²)')
                rejected.append(c)
                rejection_reasons.append(rejection_info)
                continue

            perim_px = float(cv2.arcLength(c, True))
            circ = (4 * np.pi * area_px / (perim_px ** 2)) if perim_px > 0 else 0.0

            # Get bounding box for aspect ratio
            x, y, w, h = cv2.boundingRect(c)
            aspect_ratio = float(w) / h if h > 0 else 0.0

            # Calculate solidity (convex hull ratio)
            hull = cv2.convexHull(c)
            hull_area = float(cv2.contourArea(hull))
            solidity = area_px / hull_area if hull_area > 0 else 0.0

            # ✅ NEW: Track each filter criterion
            rejection_info['perim_px'] = perim_px
            rejection_info['circularity'] = circ
            rejection_info['aspect_ratio'] = aspect_ratio
            rejection_info['solidity'] = solidity
            rejection_info['bbox'] = (x, y, w, h)
            
            # Apply bacteria-specific filters with detailed rejection tracking
            passed = True
            
            if area_px < min_area_px:
                rejection_info['reasons'].append(f'Area too small ({area_px:.1f} < {min_area_px:.1f} px²)')
                passed = False
            elif area_px > max_area_px:
                rejection_info['reasons'].append(f'Area too large ({area_px:.1f} > {max_area_px:.1f} px²)')
                passed = False
            
            if circ < bacteria_config.min_circularity:
                rejection_info['reasons'].append(f'Circularity too low ({circ:.3f} < {bacteria_config.min_circularity:.3f})')
                passed = False
            elif circ > bacteria_config.max_circularity:
                rejection_info['reasons'].append(f'Circularity too high ({circ:.3f} > {bacteria_config.max_circularity:.3f})')
                passed = False
            
            if aspect_ratio < bacteria_config.min_aspect_ratio:
                rejection_info['reasons'].append(f'Aspect ratio too low ({aspect_ratio:.3f} < {bacteria_config.min_aspect_ratio:.3f})')
                passed = False
            elif aspect_ratio > bacteria_config.max_aspect_ratio:
                rejection_info['reasons'].append(f'Aspect ratio too high ({aspect_ratio:.3f} > {bacteria_config.max_aspect_ratio:.3f})')
                passed = False
            
            if solidity < bacteria_config.min_solidity:
                rejection_info['reasons'].append(f'Solidity too low ({solidity:.3f} < {bacteria_config.min_solidity:.3f})')
                passed = False

            if passed:
                accepted.append(c)
            else:
                rejected.append(c)
                rejection_reasons.append(rejection_info)

        # ========== VISUALIZE FILTERED PARTICLES (ENHANCED) ==========
        # Generate object IDs for both accepted and rejected
        accepted_ids = [f"{group_id}_{sequence_num}_{i}" for i in range(1, len(accepted) + 1)]
        rejected_ids = [f"{group_id}_{sequence_num}_REJ{i}" for i in range(1, len(rejected) + 1)]
        
        # Visualization with rejected=orange, accepted=yellow
        vis_acc = cv2.cvtColor(img8, cv2.COLOR_GRAY2BGR)
        cv2.drawContours(vis_acc, rejected, -1, (0, 165, 255), 1)  # Orange
        cv2.drawContours(vis_acc, accepted, -1, (0, 255, 255), 1)  # Yellow
        vis_acc = draw_object_ids(vis_acc, accepted, labels=accepted_ids)
        save_debug(
            img_out, 
            "11_contours_rejected_orange_accepted_yellow_ids_green.png", 
            vis_acc, 
            um_per_px_avg
        )

        # ✅ NEW: Separate visualization for rejected objects with IDs and reasons
        if len(rejected) > 0:
            vis_rej = cv2.cvtColor(img8, cv2.COLOR_GRAY2BGR)
            cv2.drawContours(vis_rej, rejected, -1, (0, 0, 255), 2)  # Bright red for rejected
            
            # Add labels showing rejection reasons
            for i, (c, rej_id) in enumerate(zip(rejected, rejected_ids)):
                M = cv2.moments(c)
                if M["m00"] != 0:
                    cx = int(M["m10"] / M["m00"])
                    cy = int(M["m01"] / M["m00"])
                    
                    # Draw ID
                    _put_text_outline(vis_rej, rej_id, (cx, cy - 10), 
                                    font_scale=0.4, color=(0, 255, 255), thickness=1)
                    
                    # Draw first rejection reason (short version)
                    if i < len(rejection_reasons) and rejection_reasons[i]['reasons']:
                        first_reason = rejection_reasons[i]['reasons'][0]
                        # Shorten reason for display
                        if len(first_reason) > 30:
                            first_reason = first_reason[:27] + "..."
                        _put_text_outline(vis_rej, first_reason, (cx, cy + 10),
                                        font_scale=0.3, color=(255, 255, 255), thickness=1)
            
            save_debug(img_out, "11b_rejected_objects_detailed.png", vis_rej, um_per_px_avg)

        # Save masks
        mask_all = np.zeros_like(mask)
        cv2.drawContours(mask_all, contours, -1, 255, thickness=-1)
        save_debug(img_out, "12_mask_all.png", mask_all)

        mask_acc = np.zeros_like(mask)
        cv2.drawContours(mask_acc, accepted, -1, 255, thickness=-1)
        save_debug(img_out, "13_mask_accepted.png", mask_acc)

        # ✅ NEW: Save rejected mask
        if len(rejected) > 0:
            mask_rej = np.zeros_like(mask)
            cv2.drawContours(mask_rej, rejected, -1, 255, thickness=-1)
            save_debug(img_out, "13b_mask_rejected.png", mask_rej)

        # ========== FLUORESCENCE PROCESSING ==========
        fluor_path = img_path.parent / img_path.name.replace("_ch00", "_ch01")
        fluor_measurements: Optional[list[dict]] = None

        if fluor_path.exists():
            # Load fluorescence image
            fluor_img = safe_imread(fluor_path, cv2.IMREAD_UNCHANGED)
            
            if fluor_img is not None:
                # ========== ALIGN FLUORESCENCE CHANNEL ==========
                fluor_img_aligned, (sy, sx), diagnostics = align_fluorescence_channel(img, fluor_img)
                
                # Save alignment diagnostics
                safe_imwrite(img_out / "DIAG_A_no_shift.png", diagnostics['overlay_none'])
                safe_imwrite(img_out / "DIAG_B_positive_shift.png", diagnostics['overlay_pos'])
                safe_imwrite(img_out / "DIAG_C_negative_shift.png", diagnostics['overlay_neg'])
                
                print(f"  📊 Alignment shift: ({sx:.2f}, {sy:.2f}) px")
                print(f"     Check DIAG_*.png files to verify alignment quality")

                save_debug(
                    img_out, 
                    "20_fluorescence_aligned_raw.png", 
                    normalize_to_8bit(fluor_img_aligned), 
                    um_per_px_avg
                )

                # Convert to grayscale if needed
                if fluor_img_aligned.ndim == 3:
                    fluor_img_aligned = cv2.cvtColor(fluor_img_aligned, cv2.COLOR_BGR2GRAY)

                fluor_img8 = normalize_to_8bit(fluor_img_aligned)
                save_debug(img_out, "20_fluorescence_8bit.png", fluor_img8, um_per_px_avg)

                # ========== SEGMENT FLUORESCENCE ==========
                fluor_bw = segment_fluorescence_global(fluor_img8, bacteria_config)
                save_debug(img_out, "22_fluorescence_mask_global.png", fluor_bw, um_per_px_avg)

                # Find fluorescence contours
                _fc2 = cv2.findContours(fluor_bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                fluor_contours_all = cast(list[np.ndarray], _fc2[-2])

                # Filter fluorescence contours by size
                min_fluor_area_px = bacteria_config.fluor_min_area_um2 / um2_per_px2 if um2_per_px2 > 0 else 0.0
                fluor_contours = [
                    c for c in fluor_contours_all 
                    if float(cv2.contourArea(c)) >= float(min_fluor_area_px)
                ]

                # Visualize fluorescence contours
                vis_fluor = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
                cv2.drawContours(vis_fluor, fluor_contours, -1, (0, 255, 0), 1)
                save_debug(img_out, "23_fluorescence_contours_global.png", vis_fluor, um_per_px_avg)

                # ========== MATCH FLUORESCENCE TO BF PARTICLES ==========
                matches = match_fluor_to_bf_by_overlap(
                    accepted, 
                    fluor_contours, 
                    fluor_img8.shape[:2],
                    min_intersection_px=bacteria_config.fluor_match_min_intersection_px
                )

                # Measure fluorescence intensity
                fluor_measurements = measure_fluorescence_intensity_with_global_area(
                    fluor_img_aligned, 
                    accepted, 
                    fluor_contours, 
                    matches, 
                    float(um_per_px_x), 
                    float(um_per_px_y)
                )

                # Visualize matching
                vis_match = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
                for idx, bf_c in enumerate(accepted):
                    j = matches[idx]
                    cv2.drawContours(vis_match, [bf_c], -1, (0, 0, 255), 1)  # BF in red
                    if j is not None:
                        cv2.drawContours(vis_match, [fluor_contours[j]], -1, (0, 255, 0), 2)  # Fluor in green
                save_debug(img_out, "24_bf_fluor_matching_overlay.png", vis_match, um_per_px_avg)

                # Overlay with measurements
                fluor_overlay = visualize_fluorescence_measurements(fluor_img8, accepted, fluor_measurements)
                save_debug(img_out, "21_fluorescence_overlay.png", fluor_overlay, um_per_px_avg)

        # ========== SAVE LABELED VERSIONS ==========
        # BF mask with IDs
        mask_acc_bgr = cv2.cvtColor(mask_acc, cv2.COLOR_GRAY2BGR)
        save_debug_ids(
            img_out, 
            "13_mask_accepted.png", 
            mask_acc_bgr, 
            accepted, 
            accepted_ids, 
            um_per_px_avg
        )

        # Fluorescence visualizations with IDs
        if fluor_bw is not None:
            fluor_bw_bgr = cv2.cvtColor(fluor_bw, cv2.COLOR_GRAY2BGR)
            save_debug_ids(
                img_out, 
                "22_fluorescence_mask_global.png", 
                fluor_bw_bgr, 
                accepted, 
                accepted_ids, 
                um_per_px_avg
            )

        if vis_fluor is not None:
            save_debug_ids(
                img_out, 
                "23_fluorescence_contours_global.png", 
                vis_fluor, 
                accepted, 
                accepted_ids, 
                um_per_px_avg
            )

        if vis_match is not None:
            save_debug_ids(
                img_out, 
                "24_bf_fluor_matching_overlay.png", 
                vis_match, 
                accepted, 
                accepted_ids, 
                um_per_px_avg
            )

        # ========== EXPORT REJECTED OBJECTS TO CSV (NEW) ==========
        csv_rejected_path = img_out / "rejected_objects.csv"
        
        # ✅ FIX: Ensure rejection_reasons exists and matches rejected length
        if len(rejected) > 0 and len(rejection_reasons) == len(rejected):
            with open(csv_rejected_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                
                # Write header
                w.writerow([
                    "Object_ID",
                    "Rejection_Reasons",
                    "BF_Area_px",
                    "BF_Area_um2",
                    "Perimeter_px",
                    "Circularity",
                    "AspectRatio",
                    "Solidity",
                    "BBoxX_px",
                    "BBoxY_px",
                    "BBoxW_px",
                    "BBoxH_px",
                    "CentroidX_px",
                    "CentroidY_px",
                ])
                
                # Write rejected objects
                for i, (c, rej_info) in enumerate(zip(rejected, rejection_reasons)):
                    rej_id = rejected_ids[i] if i < len(rejected_ids) else f"{group_id}_{sequence_num}_REJ{i+1}"
                    reasons_str = "; ".join(rej_info['reasons'])
                    
                    area_px = rej_info['area_px']
                    area_um2 = area_px * um2_per_px2
                    perim_px = rej_info.get('perim_px', 0.0)
                    circ = rej_info.get('circularity', 0.0)
                    aspect = rej_info.get('aspect_ratio', 0.0)
                    solidity = rej_info.get('solidity', 0.0)
                    x, y, bw, bh = rej_info.get('bbox', (0, 0, 0, 0))
                    
                    # Calculate centroid
                    M = cv2.moments(c)
                    if M["m00"] != 0:
                        cx = float(M["m10"] / M["m00"])
                        cy = float(M["m01"] / M["m00"])
                    else:
                        cx, cy = 0.0, 0.0
                    
                    w.writerow([
                        rej_id,
                        reasons_str,
                        f"{area_px:.2f}",
                        f"{area_um2:.4f}",
                        f"{perim_px:.2f}",
                        f"{circ:.4f}",
                        f"{aspect:.4f}",
                        f"{solidity:.4f}",
                        x,
                        y,
                        bw,
                        bh,
                        f"{cx:.2f}",
                        f"{cy:.2f}",
                    ])
        else:
            # No rejected objects or mismatch in data - write empty CSV with header only
            with open(csv_rejected_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow([
                    "Object_ID",
                    "Rejection_Reasons",
                    "BF_Area_px",
                    "BF_Area_um2",
                    "Perimeter_px",
                    "Circularity",
                    "AspectRatio",
                    "Solidity",
                    "BBoxX_px",
                    "BBoxY_px",
                    "BBoxW_px",
                    "BBoxH_px",
                    "CentroidX_px",
                    "CentroidY_px",
                ])

        # ========== EXPORT ACCEPTED OBJECTS TO CSV ==========
        csv_path = img_out / "object_stats.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            
            # Write header
            w.writerow([
                "Object_ID",
                "Status",  # ✅ NEW: "Accepted" for all rows
                "BF_Area_px",
                "BF_Area_um2",
                "Perimeter_px",
                "Perimeter_um",
                "EquivDiameter_px",
                "EquivDiameter_um",
                "Circularity",
                "AspectRatio",
                "Solidity",
                "CentroidX_px",
                "CentroidY_px",
                "CentroidX_um",
                "CentroidY_um",
                "BBoxX_px",
                "BBoxY_px",
                "BBoxW_px",
                "BBoxH_px",
                "BBoxW_um",
                "BBoxH_um",
                "Fluor_Area_px",
                "Fluor_Area_um2",
                "Fluor_Mean",
                "Fluor_Median",
                "Fluor_Std",
                "Fluor_Min",
                "Fluor_Max",
                "Fluor_IntegratedDensity",
            ])

            # Write data for each accepted particle
            for i, c in enumerate(accepted, 1):
                compound_id = accepted_ids[i-1]

                # Area measurements
                area_px = float(cv2.contourArea(c))
                area_um2 = area_px * um2_per_px2

                # Perimeter measurements
                perim_px = float(cv2.arcLength(c, True))
                perim_um = contour_perimeter_um(c, float(um_per_px_x), float(um_per_px_y))

                # Equivalent diameter
                eqd_px = equivalent_diameter_from_area(area_px)
                eqd_um = equivalent_diameter_from_area(area_um2)

                # Shape metrics
                circ = (4 * np.pi * area_px / (perim_px**2)) if perim_px > 0 else 0.0

                x, y, bw, bh = cv2.boundingRect(c)
                aspect = (float(bw) / float(bh)) if bh > 0 else 0.0

                # Solidity
                hull = cv2.convexHull(c)
                hull_area = float(cv2.contourArea(hull))
                solidity = area_px / hull_area if hull_area > 0 else 0.0

                # Centroid
                M = cv2.moments(c)
                if M["m00"] != 0:
                    cx = float(M["m10"] / M["m00"])
                    cy = float(M["m01"] / M["m00"])
                else:
                    cx, cy = 0.0, 0.0

                cx_um = cx * float(um_per_px_x)
                cy_um = cy * float(um_per_px_y)
                bw_um = float(bw) * float(um_per_px_x)
                bh_um = float(bh) * float(um_per_px_y)

                # Fluorescence measurements
                if fluor_measurements is not None:
                    fm = fluor_measurements[i - 1]
                else:
                    fm = {
                        "fluor_area_px": 0.0,
                        "fluor_area_um2": 0.0,
                        "fluor_mean": 0.0,
                        "fluor_median": 0.0,
                        "fluor_std": 0.0,
                        "fluor_min": 0.0,
                        "fluor_max": 0.0,
                        "fluor_integrated_density": 0.0,
                    }

                # Write row
                w.writerow([
                    compound_id,
                    "Accepted",  # ✅ NEW: Status column
                    f"{area_px:.2f}",
                    f"{area_um2:.4f}",
                    f"{perim_px:.2f}",
                    f"{perim_um:.4f}",
                    f"{eqd_px:.2f}",
                    f"{eqd_um:.4f}",
                    f"{circ:.4f}",
                    f"{aspect:.4f}",
                    f"{solidity:.4f}",
                    f"{cx:.2f}",
                    f"{cy:.2f}",
                    f"{cx_um:.4f}",
                    f"{cy_um:.4f}",
                    x,
                    y,
                    bw,
                    bh,
                    f"{bw_um:.4f}",
                    f"{bh_um:.4f}",
                    f"{float(fm['fluor_area_px']):.2f}",
                    f"{float(fm['fluor_area_um2']):.4f}",
                    f"{float(fm['fluor_mean']):.2f}",
                    f"{float(fm['fluor_median']):.2f}",
                    f"{float(fm['fluor_std']):.2f}",
                    f"{float(fm['fluor_min']):.2f}",
                    f"{float(fm['fluor_max']):.2f}",
                    f"{float(fm['fluor_integrated_density']):.2f}",
                ])

        print(f"  ✓ Processed: {img_path.name} ({len(accepted)} accepted, {len(rejected)} rejected)")

    except Exception as e:
        print(f"[ERROR] Failed to process {img_path.name}: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    finally:
        # ========== EXPLICIT MEMORY CLEANUP ==========
        # This ensures cleanup happens even if an exception occurs
        import gc
        
        # Delete all large image arrays
        for var_name in ['img', 'img8', 'fluor_img', 'fluor_img_aligned', 
                         'mask', 'vis_all', 'vis_acc', 'vis_rej', 'mask_all', 'mask_acc', 'mask_rej',
                         'fluor_img8', 'fluor_bw', 'vis_fluor', 'vis_match']:
            var = locals().get(var_name)
            if var is not None:
                try:
                    del var
                except:
                    pass
        
        # Force garbage collection for large images (>10MB)
        gc.collect()

def open_folder(folder_path: Path) -> None:
    """Open folder in file explorer (cross-platform, Unicode-safe)
    
    Args:
        folder_path: Path to folder to open
    """
    try:
        folder_str = str(folder_path.resolve())
        
        if sys.platform == 'win32':
            # Use os.startfile for better Unicode support on Windows
            os.startfile(folder_str)
        elif sys.platform == 'darwin':  # macOS
            subprocess.run(['open', folder_str])
        else:  # Linux
            subprocess.run(['xdg-open', folder_str])
        
        print(f"  ✓ Opened folder: {folder_path.name}")
    except Exception as e:
        print(f"  ⚠ Could not open folder automatically: {e}")
        print(f"  Please open manually: {folder_path.resolve()}")


def copy_log_to_output(log_path: Path, output_dir: Path) -> Optional[Path]:
    """Copy log file to output directory
    
    Args:
        log_path: Path to the original log file
        output_dir: Output directory to copy to
        
    Returns:
        Path to the copied log file, or None if copy failed
    """
    try:
        if log_path and log_path.exists():
            dest_path = output_dir / log_path.name
            shutil.copy2(log_path, dest_path)
            return dest_path
        return None
    except Exception as e:
        print(f"Warning: Could not copy log file: {e}")
        return None


# ==================================================
# Missing Helper Functions
# ==================================================

def collect_images_from_directory(source_dir: Path) -> Dict[str, Dict]:
    """Collect images organized by group folders
    
    Args:
        source_dir: Source directory containing group folders
        
    Returns:
        Dictionary mapping group names to image lists
    """
    image_groups = {}
    
    # Find all group folders
    for group_dir in sorted(source_dir.iterdir()):
        if not group_dir.is_dir():
            continue
        
        # Collect images in this group
        images = list(group_dir.glob("*_ch00.tif"))
        
        if images:
            image_groups[group_dir.name] = {
                'path': group_dir,
                'images': images
            }
    
    return image_groups


# Configurations included in the "Unknown Sample" multi-scan.
# Streptococcus mitis is excluded (insufficient validation samples).
# Default (General Purpose) is excluded (uninformative for identification).


def _load_multi_scan_whitelist() -> list[str]:
    """Return validated bacteria keys from the registry.

    Falls back to the original hardcoded list if the registry module is
    unavailable (e.g. frozen executable without the file).
    """
    try:
        from bacteria_registry import registry as _reg
        wl = _reg.get_whitelist()
        if wl:
            print(f"[Whitelist] {len(wl)} validated bacteria: {wl}")
            return wl
        print("[Whitelist] Registry returned empty whitelist; using fallback.")
    except Exception as exc:
        print(f"[Whitelist] Could not load from registry ({exc}); using fallback.")

    return ['klebsiella_pneumoniae', 'proteus_mirabilis']   # fallback


MULTI_SCAN_WHITELIST: list[str] = _load_multi_scan_whitelist()


def select_bacteria_configuration() -> dict:
    """Interactive bacteria configuration selector with multi-scan support
    
    Returns:
        dict: Configuration info with keys:
            - mode: 'single' or 'multi_scan'
            - bacteria_type: specific type (for single mode) or None
            - selected_config: SegmentationConfig object (for single mode)
            - configs_to_scan: list of config keys (for multi_scan)
            - config_names: dict mapping keys to display names
            - configs: dict mapping keys to SegmentationConfig objects
    """
    print("\n" + "="*80)
    print("BACTERIA CONFIGURATION SELECTION")
    print("="*80)
    
    # Find all JSON config files
    config_dir = Path("bacteria_configs")
    
    if not config_dir.exists():
        print("\n⚠ bacteria_configs/ directory not found!")
        print("  Run tuner.py first to create configurations")
        return {
            'mode': 'single',
            'bacteria_type': 'default',
            'selected_config': None,
            'configs_to_scan': [],
            'config_names': {},
            'configs': {}
        }
    
    json_files = list(config_dir.glob("*.json"))
    
    if not json_files:
        print("\n⚠ No JSON configuration files found!")
        print("  Run tuner.py first to create configurations")
        return {
            'mode': 'single',
            'bacteria_type': 'default',
            'selected_config': None,
            'configs_to_scan': [],
            'config_names': {},
            'configs': {}
        }
    
    # Extract bacteria keys from filenames and load names
    available_configs = []
    config_names = {}
    
    for json_file in sorted(json_files):
        bacteria_key = json_file.stem
        
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Extract display name
            if "config" in data:
                name = data["config"].get("name", bacteria_key)
            else:
                name = data.get("name", bacteria_key)
            
            available_configs.append(bacteria_key)
            config_names[bacteria_key] = name
            
        except Exception as e:
            print(f"  ⚠ Could not read {json_file.name}: {e}")
            available_configs.append(bacteria_key)
            config_names[bacteria_key] = bacteria_key
    
    if not available_configs:
        print("\n⚠ No valid configuration files found!")
        return {
            'mode': 'single',
            'bacteria_type': 'default',
            'selected_config': None,
            'configs_to_scan': [],
            'config_names': {},
            'configs': {}
        }
    
    # Display mode selection
    print("\n" + "─" * 80)
    print("PROCESSING MODE SELECTION")
    print("─" * 80)
    print("\nSelect processing mode:")
    print("  [1] UNKNOWN SAMPLE - Scan validated configurations (Recommended)")
    print("      → Tests clinically validated bacteria types and provides confidence report")
    print("      → Best for clinical samples with unknown bacteria")
    print()
    print("  [2] KNOWN BACTERIA - Select specific configuration")
    print("      → Faster processing for known bacteria type")
    print("      → Best for research/validation")
    print()
    
    # Try to load last selection
    last_selection_file = Path(".last_processing_mode")
    last_mode = None
    if last_selection_file.exists():
        try:
            last_mode = last_selection_file.read_text().strip()
            if last_mode in ['1', '2']:
                print(f"💡 Last used: Mode {last_mode}")
                print()
        except:
            pass
    
    # Mode selection loop
    while True:
        if last_mode:
            choice = logged_input("Select mode (or press Enter for last used): ").strip()
            if choice == "":
                choice = last_mode
        else:
            choice = logged_input("Select mode [1-2]: ").strip()
        
        if choice in ['q', 'quit', 'exit']:
            raise SystemExit(0)
        
        # ========================================
        # MODE 1: MULTI-SCAN (Unknown Sample)
        # ========================================
        if choice == "1":
            print("\n✓ Selected: UNKNOWN SAMPLE (Multi-Configuration Scan)")

            # ── Filter to whitelisted configs only ──────────────────────────
            configs_to_scan = [k for k in available_configs if k in MULTI_SCAN_WHITELIST]
            skipped = [k for k in available_configs if k not in MULTI_SCAN_WHITELIST]

            if not configs_to_scan:
                print("\n  ❌ No whitelisted configurations found in bacteria_configs/")
                print(f"     Required files: "
                      f"{', '.join(f'{k}.json' for k in MULTI_SCAN_WHITELIST)}")
                print("     Please run tuner.py to create these configurations first.")
                continue  # Back to mode selection

            # Warn about excluded configs so the user knows this is intentional
            print(f"\n  Will test {len(configs_to_scan)} validated configurations:")
            for bacteria_key in configs_to_scan:
                print(f"    ✓ {config_names[bacteria_key]}")

            if skipped:
                print(f"\n  Excluded from scan ({len(skipped)} config(s)):")
                for k in skipped:
                    if 'streptococcus' in k:
                        reason = "under testing"
                    elif k == 'default':
                        reason = "baseline configuration only"
                    else:
                        reason = "not in multi-scan whitelist"
                    print(f"    ✗ {config_names.get(k, k)}  ({reason})")

            confirm = logged_input("\n  Proceed with multi-scan? (y/n, Enter=yes): ").strip().lower()
            if confirm not in ['', 'y', 'yes']:
                continue  # Go back to mode selection

            # Save selection for next time
            try:
                last_selection_file.write_text('1')
            except:
                pass

            # Load whitelisted configs
            print("\n  Loading configurations...")
            configs_dict = {}
            loaded_count = 0

            for bacteria_key in configs_to_scan:
                loaded_config = load_bacteria_config_from_json(bacteria_key)
                if loaded_config:
                    configs_dict[bacteria_key] = loaded_config
                    loaded_count += 1
                else:
                    print(f"  ⚠ Failed to load: {bacteria_key}")

            if loaded_count == 0:
                print("  ❌ Could not load any configurations. Aborting.")
                continue

            print(f"  ✓ Loaded {loaded_count}/{len(configs_to_scan)} configurations")

            return {
                'mode': 'multi_scan',
                'bacteria_type': None,
                'selected_config': None,
                'configs_to_scan': configs_to_scan,   # ← whitelisted subset only
                'config_names': config_names,
                'configs': configs_dict
            }
        
        # ========================================
        # MODE 2: SINGLE CONFIG (Known Bacteria)
        # ========================================
        elif choice == "2":
            print("\n✓ Selected: KNOWN BACTERIA (Single Configuration)")
            print("\nAvailable bacteria configurations:")
            
            for i, bacteria_key in enumerate(available_configs, 1):
                name = config_names[bacteria_key]
                print(f"  [{i}] {name}")
            
            print()
            
            # Try to load last bacteria selection
            last_bacteria_file = Path(".last_bacteria_selection")
            last_bacteria = None
            if last_bacteria_file.exists():
                try:
                    last_bacteria = last_bacteria_file.read_text().strip()
                    if last_bacteria in available_configs:
                        last_name = config_names[last_bacteria]
                        last_idx = available_configs.index(last_bacteria) + 1
                        print(f"💡 Last used: [{last_idx}] {last_name}")
                        print()
                except:
                    pass
            
            # Bacteria selection loop
            while True:
                bacteria_choice = logged_input("Select bacteria configuration: ").strip()
                
                # Handle default (last used)
                if bacteria_choice == "" and last_bacteria:
                    selected_key = last_bacteria
                    print(f"  ✓ Using: {config_names[selected_key]}")
                    break
                
                # Handle numeric input
                if bacteria_choice.isdigit():
                    idx = int(bacteria_choice)
                    if 1 <= idx <= len(available_configs):
                        selected_key = available_configs[idx - 1]
                        print(f"  ✓ Selected: {config_names[selected_key]}")
                        break
                    else:
                        print(f"  ✗ Invalid number. Enter 1-{len(available_configs)}")
                        continue
                
                # Handle string key input
                if bacteria_choice in available_configs:
                    selected_key = bacteria_choice
                    print(f"  ✓ Selected: {config_names[selected_key]}")
                    break
                
                print(f"  ✗ Invalid selection. Try again.")
            
            # Load the selected configuration
            print(f"\n  Loading configuration...")
            selected_config = load_bacteria_config_from_json(selected_key)
            
            if selected_config is None:
                print(f"  ✗ Failed to load configuration: {selected_key}")
                print(f"  Please try a different configuration.")
                continue  # Go back to mode selection
            
            print(f"  ✓ Configuration loaded successfully")
            
            # Save selections for next time
            try:
                last_selection_file.write_text('2')
                last_bacteria_file.write_text(selected_key)
            except:
                pass
            
            return {
                'mode': 'single',
                'bacteria_type': selected_key,
                'selected_config': selected_config,
                'configs_to_scan': [selected_key],
                'config_names': config_names,
                'configs': {selected_key: selected_config}
            }
        
        # Invalid choice
        else:
            print("  ✗ Invalid choice. Enter 1 or 2 (or 'q' to quit)")




def validate_batch_structure(config: dict) -> bool:
    """Validate that batch mode directories contain images before processing
    
    Args:
        config: Configuration dictionary
        
    Returns:
        bool: True if structure is valid, False otherwise
    """
    if not config.get('batch_mode'):
        return True
    
    print("\n" + "="*80)
    print("VALIDATING BATCH STRUCTURE")
    print("="*80 + "\n")
    
    issues = []
    warnings = []
    
    # Check G+ directory
    if 'source_dir_positive' in config:
        g_plus = config['source_dir_positive']
        
        if not g_plus.exists():
            issues.append(f"G+ directory does not exist: {g_plus}")
        else:
            images = list(g_plus.rglob("*_ch00.tif"))
            print(f"✓ G+ directory: {g_plus.name}")
            print(f"  Found {len(images)} images")
            
            if len(images) == 0:
                issues.append(f"No images (*_ch00.tif) found in {g_plus}")
            else:
                # Check for group folders
                group_folders = [d for d in g_plus.iterdir() if d.is_dir() and not d.name.startswith('.')]
                print(f"  Found {len(group_folders)} group folders")
                
                # Check for control
                has_control = any('control' in d.name.lower() for d in group_folders)
                if not has_control:
                    warnings.append("No 'Control' folder found in G+")
    else:
        issues.append("'source_dir_positive' not set in config")
    
    print()
    
    # Check G- directory
    if 'source_dir_negative' in config:
        g_minus = config['source_dir_negative']
        
        if not g_minus.exists():
            issues.append(f"G- directory does not exist: {g_minus}")
        else:
            images = list(g_minus.rglob("*_ch00.tif"))
            print(f"✓ G- directory: {g_minus.name}")
            print(f"  Found {len(images)} images")
            
            if len(images) == 0:
                issues.append(f"No images (*_ch00.tif) found in {g_minus}")
            else:
                # Check for group folders
                group_folders = [d for d in g_minus.iterdir() if d.is_dir() and not d.name.startswith('.')]
                print(f"  Found {len(group_folders)} group folders")
                
                # Check for control
                has_control = any('control' in d.name.lower() for d in group_folders)
                if not has_control:
                    warnings.append("No 'Control' folder found in G-")
    else:
        issues.append("'source_dir_negative' not set in config")
    
    print()
    
    # Report issues
    if issues:
        print("❌ VALIDATION FAILED:")
        for issue in issues:
            print(f"  • {issue}")
        print()
        return False
    
    if warnings:
        print("⚠️  WARNINGS:")
        for warning in warnings:
            print(f"  • {warning}")
        print()
    
    print("✓ Structure validation passed\n")
    print("="*80 + "\n")
    return True



# ==================================================
# Main Function
# ==================================================


def main():
    """Main execution function with enhanced error handling and cleanup"""

    # ==================== INITIALIZATION ====================
    print("\n" + "=" * 80)
    print("MICROGEL FLUORESCENCE ANALYSIS PIPELINE")
    print("=" * 80 + "\n")

    output_dir: Optional[Path] = None
    config: dict = {}
    output_root: Optional[Path] = None

    # Results containers
    positive_results: Optional[dict[str, Any]] = None
    negative_results: Optional[dict[str, Any]] = None
    results: Optional[dict[str, Any]] = None

    global _log_path, _log_file

    try:
        # ==================== STEP 1: MODE & CONFIGURATION SELECTION ====================
        print("STEP 1: Select Processing Mode & Configuration")
        print("─" * 80 + "\n")

        bacteria_config_info = select_bacteria_config()
        mode = bacteria_config_info['mode']

        config['bacteria_config_info'] = bacteria_config_info
        config['processing_mode'] = mode

        # ==================== STEP 2: DATASET CONFIGURATION ====================
        print("\nSTEP 2: Dataset Configuration")
        print("─" * 80 + "\n")

        config.update(configure_dataset())

        if not validate_config(config):
            print("\n❌ Configuration validation failed")
            return

        if config.get('batch_mode') and not validate_batch_structure(config):
            print("❌ Batch structure validation failed - cannot proceed")
            return

        # ==================== STEP 3: CREATE OUTPUT DIRECTORY ====================
        display_configuration_summary(config)
        config['timestamp'] = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_root = setup_output_directory(config)
        output_dir = output_root

        # ==================== STEP 4: PROCESSING ====================
        print("=" * 80)
        print("PHASE 2: PROCESSING")
        print("=" * 80 + "\n")

        if mode == "multi_scan":
            if config['batch_mode']:
                print("Processing in BATCH mode...\n")

                # ✅ G+ multi-scan
                print("─" * 80)
                print("Processing: G+ (positive)")
                print("─" * 80 + "\n")

                config['current_source'] = config['source_dir_positive']
                config['output_dir'] = config['positive_output']
                config['dataset_id_current'] = f"{config['dataset_id']} Positive"

                positive_results = run_multi_config_scan(config, bacteria_config_info)

                print("\n✓ G+ processing completed\n")

                # ✅ G- multi-scan
                print("─" * 80)
                print("Processing: G- (negative)")
                print("─" * 80 + "\n")

                config['current_source'] = config['source_dir_negative']
                config['output_dir'] = config['negative_output']
                config['dataset_id_current'] = f"{config['dataset_id']} Negative"

                negative_results = run_multi_config_scan(config, bacteria_config_info)

                print("\n✓ G- processing completed\n")

                # Auto clinical follow-up
                print("=" * 80)
                print("AUTO CLINICAL FOLLOW-UP (POST MULTI-SCAN)")
                print("=" * 80)

                positive_ranked = positive_results.get('ranked_results', []) if positive_results is not None else []
                negative_ranked = negative_results.get('ranked_results', []) if negative_results is not None else []

                best_positive = positive_ranked[0] if positive_ranked else None
                best_negative = negative_ranked[0] if negative_ranked else None

                chosen_config_key = None
                selection_rule = None

                if best_positive and best_negative:
                    if best_positive['config_key'] == best_negative['config_key']:
                        chosen_config_key = best_positive['config_key']
                        selection_rule = "G+ and G- best matches agree"
                    else:
                        if best_positive['confidence'] > best_negative['confidence']:
                            chosen_config_key = best_positive['config_key']
                            selection_rule = (
                                f"G+ has higher confidence "
                                f"({best_positive['confidence']:.1f}% vs {best_negative['confidence']:.1f}%)"
                            )
                        else:
                            chosen_config_key = best_negative['config_key']
                            selection_rule = (
                                f"G- has higher confidence "
                                f"({best_negative['confidence']:.1f}% vs {best_positive['confidence']:.1f}%)"
                            )
                elif best_positive:
                    chosen_config_key = best_positive['config_key']
                    selection_rule = "Only G+ results available"
                elif best_negative:
                    chosen_config_key = best_negative['config_key']
                    selection_rule = "Only G- results available"

                if chosen_config_key:
                    print(f"Chosen bacteria configuration: {chosen_config_key}")
                    print(f"Selection rule: {selection_rule}\n")

                    chosen_bacteria_config = None
                    if positive_results is not None:
                        chosen_bacteria_config = positive_results.get('configs', {}).get(chosen_config_key)
                    if chosen_bacteria_config is None and negative_results is not None:
                        chosen_bacteria_config = negative_results.get('configs', {}).get(chosen_config_key)
                    if chosen_bacteria_config is None:
                        chosen_bacteria_config = load_bacteria_config_from_json(chosen_config_key)

                    # ✅ G+ clinical run
                    print("─" * 80)
                    print(f"CLINICAL RUN: G+ using config '{chosen_config_key}'")
                    print("─" * 80 + "\n")

                    config['bacteria_config'] = chosen_bacteria_config
                    config['current_source'] = config['source_dir_positive']
                    config['output_dir'] = config['positive_output']
                    config['dataset_id_current'] = f"{config['dataset_id']} Positive"

                    run_single_config_analysis(config)

                    # ✅ G- clinical run
                    print("\n─" * 80)
                    print(f"CLINICAL RUN: G- using config '{chosen_config_key}'")
                    print("─" * 80 + "\n")

                    config['current_source'] = config['source_dir_negative']
                    config['output_dir'] = config['negative_output']
                    config['dataset_id_current'] = f"{config['dataset_id']} Negative"

                    run_single_config_analysis(config)

                    # Final combined outputs
                    if output_root is not None:
                        generate_final_clinical_matrix_wrapper(output_root, config)

                        if config.get('batch_mode', False):
                            generate_rejection_analysis(output_root)
                        elif output_dir is not None:
                            generate_rejection_analysis(output_dir)

                else:
                    print("⚠ Could not determine best bacteria configuration")
                    print("  Manual review of multi-scan results recommended\n")

            else:
                # Single directory multi-scan
                print("Processing single directory...\n")

                config['current_source'] = config['source_dir']
                config['output_dir'] = output_dir
                config['dataset_id_current'] = config['dataset_id']

                results = run_multi_config_scan(config, bacteria_config_info)

                ranked_results = results.get('ranked_results', []) if results is not None else []

                if ranked_results:
                    best_match = ranked_results[0]
                    chosen_config_key = best_match['config_key']

                    print("\n" + "=" * 80)
                    print("AUTO CLINICAL FOLLOW-UP")
                    print("=" * 80)
                    print(f"Using best match: {chosen_config_key} ({best_match['confidence']:.1f}% confidence)\n")

                    config['bacteria_config'] = bacteria_config_info['configs'][chosen_config_key]
                    run_single_config_analysis(config)

        else:
            # Single config mode
            config['bacteria_config'] = bacteria_config_info['selected_config']

            if config['batch_mode']:
                print("Processing in BATCH mode...\n")

                # ✅ G+
                print("─" * 80)
                print("Processing: G+ (positive)")
                print("─" * 80 + "\n")

                config['current_source'] = config['source_dir_positive']
                config['output_dir'] = config['positive_output']
                config['dataset_id_current'] = f"{config['dataset_id']} Positive"

                run_single_config_analysis(config)

                print("\n✓ G+ processing completed\n")

                # ✅ G-
                print("─" * 80)
                print("Processing: G- (negative)")
                print("─" * 80 + "\n")

                config['current_source'] = config['source_dir_negative']
                config['output_dir'] = config['negative_output']
                config['dataset_id_current'] = f"{config['dataset_id']} Negative"

                run_single_config_analysis(config)

                print("\n✓ G- processing completed\n")

                if output_root is not None:
                    generate_final_clinical_matrix_wrapper(output_root, config)

            else:
                print("Processing single directory...\n")

                config['current_source'] = config['source_dir']
                config['output_dir'] = output_dir
                config['dataset_id_current'] = config['dataset_id']

                run_single_config_analysis(config)

        # Cleanup multi-scan artifacts
        if mode == "multi_scan" and config.get('batch_mode', False) and output_root is not None:
            cleanup_and_reorganize_output(output_root, config)

        # ==================== FINAL SUMMARY ====================
        print("\n" + "=" * 80)
        if config.get('batch_mode', False):
            print("BATCH PROCESSING COMPLETE")
            print("=" * 80)
            print(f"  Output folder: {output_root}")

            if mode == "multi_scan":
                if positive_results is not None and negative_results is not None:
                    positive_ranked = positive_results.get('ranked_results', [])
                    negative_ranked = negative_results.get('ranked_results', [])

                    pos_best = positive_ranked[0] if positive_ranked else None
                    neg_best = negative_ranked[0] if negative_ranked else None

                    if pos_best:
                        print(f"  ✓ G+: Best match = {pos_best['bacteria_name']} ({pos_best['confidence']:.1f}%)")
                    if neg_best:
                        print(f"  ✓ G-: Best match = {neg_best['bacteria_name']} ({neg_best['confidence']:.1f}%)")
        else:
            print("PROCESSING COMPLETE")
            print("=" * 80)
            print(f"  Output folder: {output_dir}")

            if mode == "multi_scan" and results is not None:
                ranked_results = results.get('ranked_results', [])
                best = ranked_results[0] if ranked_results else None
                if best:
                    print(f"  ✓ Best match: {best['bacteria_name']} ({best['confidence']:.1f}%)")

        folder_to_open = output_root if output_root is not None else output_dir
        if folder_to_open is not None:
            open_folder(folder_to_open)

        print("\n")

    except KeyboardInterrupt:
        print("\n\n⚠ Processing interrupted by user")
        print("  Partial results may be available in the output directory\n")

    except Exception as e:
        print("\n" + "=" * 80)
        print("❌ ERROR OCCURRED")
        print("=" * 80)
        print(f"Error: {e}\n")

        import traceback
        traceback.print_exc()
        print()

    finally:
        # Log file handling
        try:
            if _log_file is not None:
                _log_file.flush()
                os.fsync(_log_file.fileno())
        except Exception:
            pass

        # Copy log file to output directory
        try:
            if output_root is not None and _log_path and _log_path.exists():
                print("\n📄 Saving log file...")

                log_destinations = []

                if config.get('batch_mode', False):
                    log_destinations.append(output_root)

                    if 'positive_output' in config:
                        pos_out = config['positive_output']
                        if isinstance(pos_out, Path) and pos_out.exists():
                            log_destinations.append(pos_out)

                    if 'negative_output' in config:
                        neg_out = config['negative_output']
                        if isinstance(neg_out, Path) and neg_out.exists():
                            log_destinations.append(neg_out)
                else:
                    if output_dir is not None and output_dir.exists():
                        log_destinations.append(output_dir)

                copied_count = 0
                for dest in log_destinations:
                    try:
                        log_copy = copy_log_to_output(_log_path, dest)
                        if log_copy:
                            try:
                                rel_path = log_copy.relative_to(PROJECT_ROOT)
                                print(f"  ✓ Log saved: {rel_path}")
                            except ValueError:
                                print(f"  ✓ Log saved: {log_copy.name}")
                            copied_count += 1
                    except Exception as e:
                        print(f"  ✗ Failed to copy to {dest.name}: {e}")

                if copied_count == 0:
                    print(f"  ⚠ Log not copied (see original): {_log_path}")
                else:
                    print(f"  ✓ Total copies: {copied_count}")

            elif _log_path and _log_path.exists():
                print(f"\n📄 Log file: {_log_path}")
                print("   (Not copied - no output directory)")

        except NameError as e:
            print(f"\n⚠ Could not access log variables: {e}")
        except Exception as e:
            print(f"\n⚠ Error in log copy section: {e}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    start_time = pytime.time()
    main()
    print("="*80)
    print(f"  • Total runtime: {pytime.time() - start_time:.1f} seconds")
