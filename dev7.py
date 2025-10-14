import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from skimage import io, filters, morphology, measure, segmentation
from skimage.feature import peak_local_max
from scipy import ndimage
from pathlib import Path
import seaborn as sns
from io import BytesIO

def analyze_bacteria_images(image_path_ch00, image_path_ch01, output_folder='./outputs'):
    """
    Analyzes two channels of bacteria screening images to identify, quantify, label, and measure bacteria.

    Args:
        image_path_ch00 (str): Path to the first channel image (brightfield/phase contrast).
        image_path_ch01 (str): Path to the second channel image (fluorescence).
        output_folder (str): Path to save output files.

    Returns:
        dict: A dictionary containing analysis results, measurements DataFrame, and output paths.
    """
    try:
        # Create output folder if it doesn't exist
        output_path = Path(output_folder)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # Load images
        image_ch00 = io.imread(image_path_ch00)
        image_ch01 = io.imread(image_path_ch01)

        print(f"Loaded {image_path_ch00} with shape {image_ch00.shape} and dtype {image_ch00.dtype}")
        print(f"Loaded {image_path_ch01} with shape {image_ch01.shape} and dtype {image_ch01.dtype}")

        # --- Preprocessing for ch00 (brightfield channel) ---
        if image_ch00.ndim == 3:
            image_ch00_gray = image_ch00[:, :, 0]
        else:
            image_ch00_gray = image_ch00

        # --- Preprocessing for ch01 (fluorescence channel) ---
        if image_ch01.ndim == 3:
            image_ch01_gray = image_ch01[:, :, 0]  # Red channel
        else:
            image_ch01_gray = image_ch01

        # Normalize images
        image_ch00_norm = image_ch00_gray / image_ch00_gray.max() if image_ch00_gray.max() > 0 else image_ch00_gray
        image_ch01_norm = image_ch01_gray / image_ch01_gray.max() if image_ch01_gray.max() > 0 else image_ch01_gray

        # --- IMPROVED Segmentation with Quick Fixes ---
        # Apply MORE Gaussian blur for noise reduction
        blurred_image = filters.gaussian(image_ch01_norm, sigma=2)

        # Apply Otsu's threshold
        thresh = filters.threshold_otsu(blurred_image)
        binary_image = blurred_image > thresh

        # Clean up segmentation - LARGER min size to remove noise
        cleaned_image = morphology.remove_small_objects(binary_image, min_size=50)
        cleaned_image = morphology.remove_small_holes(cleaned_image, area_threshold=100)

        # ADDED: Morphological closing to merge close parts of same bacterium
        cleaned_image = morphology.binary_closing(cleaned_image, morphology.disk(4))

        # Watershed segmentation to separate touching bacteria
        distance = ndimage.distance_transform_edt(cleaned_image)
        
        # IMPROVED: More smoothing to prevent over-segmentation
        distance_smooth = filters.gaussian(distance, sigma=2.5)

        # IMPROVED: Use peak_local_max with better control
        coordinates = peak_local_max(
            distance_smooth, 
            min_distance=15,
            threshold_abs=0.3 * distance_smooth.max(),
            exclude_border=True
        )
        
        # Create marker image from peaks
        mask_peaks = np.zeros(distance_smooth.shape, dtype=bool)
        mask_peaks[tuple(coordinates.T)] = True
        markers = measure.label(mask_peaks)
        
        # Apply watershed
        labeled_bacteria = segmentation.watershed(-distance_smooth, markers, mask=cleaned_image)

        num_bacteria = int(labeled_bacteria.max())
        
        print(f"Detected {num_bacteria} bacteria (after improved segmentation)")

        # --- Extract Comprehensive Measurements ---
        measurements_df = extract_measurements(
            labeled_bacteria, 
            image_ch00_gray, 
            image_ch01_gray,
            image_path_ch00
        )

        # Classify bacteria by fluorescence intensity
        measurements_df = classify_bacteria(measurements_df)

        # --- Export Measurements to Tables ---
        # Get base filename
        base_name = Path(image_path_ch00).stem
        
        # Export to Excel with visualizations
        excel_path = output_path / f'{base_name}_measurements.xlsx'
        export_to_excel_with_charts(
            measurements_df, 
            excel_path,
            image_ch00_gray,
            image_ch01_gray,
            labeled_bacteria
        )
        
        # Export to CSV
        csv_path = output_path / f'{base_name}_measurements.csv'
        measurements_df.to_csv(csv_path, index=False)
        print(f"✓ Exported measurements to CSV: {csv_path}")

        # --- Simple Tagged Image Only ---
        result_image_path = output_path / f'{base_name}_tagged.png'
        create_simple_tagged_image(
            image_ch00_gray,
            measurements_df,
            num_bacteria,
            result_image_path
        )

        # --- Generate Summary Report ---
        summary_path = output_path / f'{base_name}_summary.txt'
        generate_summary_report(measurements_df, summary_path)

        print(f"\n{'='*60}")
        print(f"Analysis complete for: {Path(image_path_ch00).name}")
        print(f"Total bacteria detected: {num_bacteria}")
        print(f"High fluorescence: {len(measurements_df[measurements_df['fluorescence_class']=='High'])}")
        print(f"Low fluorescence: {len(measurements_df[measurements_df['fluorescence_class']=='Low'])}")
        print(f"Results saved to: {output_folder}")
        print(f"{'='*60}\n")

        return {
            'num_bacteria': num_bacteria,
            'measurements': measurements_df,
            'result_image_path': str(result_image_path),
            'excel_path': str(excel_path),
            'csv_path': str(csv_path),
            'summary_path': str(summary_path)
        }

    except Exception as e:
        print(f"An error occurred during analysis: {e}")
        import traceback
        traceback.print_exc()
        return {'error': str(e)}


def extract_measurements(labeled_bacteria, image_ch00, image_ch01, image_name):
    """Extract comprehensive measurements for each bacterium"""
    
    # Measurements from brightfield channel
    props_ch00 = measure.regionprops_table(
        labeled_bacteria,
        intensity_image=image_ch00,
        properties=['label', 'area', 'perimeter', 'eccentricity',
                   'solidity', 'orientation', 'major_axis_length',
                   'minor_axis_length', 'centroid', 'bbox']
    )
    
    # Measurements from fluorescence channel
    props_ch01 = measure.regionprops_table(
        labeled_bacteria,
        intensity_image=image_ch01,
        properties=['label', 'mean_intensity', 'max_intensity',
                   'min_intensity']
    )
    
    # Create DataFrames
    df_ch00 = pd.DataFrame(props_ch00)
    df_ch01 = pd.DataFrame(props_ch01)
    
    # Merge measurements
    measurements = pd.merge(df_ch00, df_ch01, on='label')
    
    # Add calculated features
    measurements['aspect_ratio'] = (
        measurements['major_axis_length'] / measurements['minor_axis_length']
    )
    
    measurements['circularity'] = (
        4 * np.pi * measurements['area'] / (measurements['perimeter'] ** 2)
    )
    
    measurements['integrated_intensity'] = (
        measurements['mean_intensity'] * measurements['area']
    )
    
    measurements['orientation_degrees'] = np.degrees(measurements['orientation'])
    
    # Calculate bounding box dimensions
    measurements['bbox_width'] = measurements['bbox-3'] - measurements['bbox-1']
    measurements['bbox_height'] = measurements['bbox-2'] - measurements['bbox-0']
    
    # Rename centroid columns for clarity
    measurements.rename(columns={
        'centroid-0': 'centroid_y',
        'centroid-1': 'centroid_x'
    }, inplace=True)
    
    # Add image identifier
    measurements['image_name'] = Path(image_name).stem
    
    # Round numeric columns
    numeric_cols = measurements.select_dtypes(include=[np.number]).columns
    measurements[numeric_cols] = measurements[numeric_cols].round(3)
    
    # Reorder columns for better readability
    column_order = ['label', 'image_name', 'centroid_x', 'centroid_y', 
                   'area', 'perimeter', 'circularity',
                   'major_axis_length', 'minor_axis_length', 'aspect_ratio',
                   'eccentricity', 'solidity', 'orientation_degrees',
                   'mean_intensity', 'max_intensity', 'min_intensity', 
                   'integrated_intensity']
    
    # Add remaining columns
    remaining_cols = [col for col in measurements.columns if col not in column_order]
    measurements = measurements[column_order + remaining_cols]
    
    return measurements


def classify_bacteria(measurements_df, intensity_threshold=None):
    """Classify bacteria based on fluorescence intensity"""
    
    if intensity_threshold is None:
        # Use median as threshold
        intensity_threshold = measurements_df['mean_intensity'].median()
    
    measurements_df['fluorescence_class'] = np.where(
        measurements_df['mean_intensity'] > intensity_threshold,
        'High',
        'Low'
    )
    
    measurements_df['intensity_threshold'] = intensity_threshold
    
    return measurements_df


def create_simple_tagged_image(image_ch00, measurements_df, num_bacteria, save_path):
    """Create a simple tagged and labeled image only"""
    
    fig, ax = plt.subplots(1, 1, figsize=(12, 12))
    
    # Display brightfield image
    ax.imshow(image_ch00, cmap='gray')
    
    # Add labels and color coding
    for _, row in measurements_df.iterrows():
        color = 'lime' if row['fluorescence_class'] == 'High' else 'cyan'
        marker_size = 14 if row['fluorescence_class'] == 'High' else 10
        
        # Draw circle
        ax.plot(row['centroid_x'], row['centroid_y'], 'o',
                color=color, markersize=marker_size, 
                markeredgewidth=2.5, markerfacecolor='none')
        
        # Add label number
        ax.text(row['centroid_x'] + 15, row['centroid_y'] + 15,
                str(int(row['label'])),
                color='yellow', fontsize=11, fontweight='bold',
                bbox=dict(boxstyle='round,pad=0.4', facecolor='black', alpha=0.7))
    
    # Add legend
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor='lime', edgecolor='lime', label=f'High Fluorescence (n={len(measurements_df[measurements_df["fluorescence_class"]=="High"])})'),
        Patch(facecolor='cyan', edgecolor='cyan', label=f'Low Fluorescence (n={len(measurements_df[measurements_df["fluorescence_class"]=="Low"])})')
    ]
    ax.legend(handles=legend_elements, loc='upper right', fontsize=10, framealpha=0.9)
    
    ax.set_title(f'Screening (Total: {num_bacteria})', 
                fontsize=14, fontweight='bold', pad=15)
    ax.axis('off')
    
    plt.tight_layout()
    plt.savefig(save_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    
    print(f"✓ Tagged image saved: {save_path}")


def create_chart_images(measurements_df, image_ch00, image_ch01, labeled_bacteria):
    """Create chart images for Excel embedding"""
    
    chart_images = {}
    
    # Chart 1: Original images comparison
    fig, axes = plt.subplots(1, 3, figsize=(15, 5))

    axes[0].imshow(image_ch00, cmap='gray')
    axes[0].imshow(labeled_bacteria, cmap='nipy_spectral', alpha=0.5)
    axes[0].set_title('Detected Bacteria', fontweight='bold')
    axes[0].axis('off')
    
    axes[1].imshow(image_ch01, cmap='hot')
    axes[1].set_title('Channel 01 (Fluorescence)', fontweight='bold')
    axes[1].axis('off')
    
    axes[2].imshow(image_ch00, cmap='gray')
    axes[2].set_title('Channel 00 (Brightfield)', fontweight='bold')
    axes[2].axis('off')
    
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    chart_images['original_images'] = buf
    plt.close(fig)
    
    # Chart 2: Distribution plots
    fig, axes = plt.subplots(2, 2, figsize=(12, 10))
    
    # Area distribution
    axes[0, 0].hist(measurements_df['area'], bins=20, color='steelblue', edgecolor='black', alpha=0.7)
    axes[0, 0].axvline(measurements_df['area'].mean(), color='red', linestyle='--', 
                       linewidth=2, label=f'Mean: {measurements_df["area"].mean():.1f}')
    axes[0, 0].set_xlabel('Area (pixels²)', fontsize=10)
    axes[0, 0].set_ylabel('Count', fontsize=10)
    axes[0, 0].set_title('Size Distribution', fontsize=11, fontweight='bold')
    axes[0, 0].legend()
    axes[0, 0].grid(alpha=0.3)
    
    # Intensity distribution
    axes[0, 1].hist(measurements_df['mean_intensity'], bins=20, 
                    color='orangered', edgecolor='black', alpha=0.7)
    axes[0, 1].axvline(measurements_df['mean_intensity'].median(), color='blue', 
                       linestyle='--', linewidth=2, 
                       label=f'Median: {measurements_df["mean_intensity"].median():.1f}')
    axes[0, 1].set_xlabel('Mean Fluorescence Intensity', fontsize=10)
    axes[0, 1].set_ylabel('Count', fontsize=10)
    axes[0, 1].set_title('Intensity Distribution', fontsize=11, fontweight='bold')
    axes[0, 1].legend()
    axes[0, 1].grid(alpha=0.3)
    
    # Size vs Intensity scatter
    scatter = axes[1, 0].scatter(measurements_df['area'],
                                 measurements_df['mean_intensity'],
                                 c=measurements_df['mean_intensity'],
                                 cmap='hot', s=60, edgecolors='black', 
                                 linewidth=0.5, alpha=0.7)
    axes[1, 0].set_xlabel('Area (pixels²)', fontsize=10)
    axes[1, 0].set_ylabel('Mean Fluorescence Intensity', fontsize=10)
    axes[1, 0].set_title('Size vs Intensity', fontsize=11, fontweight='bold')
    axes[1, 0].grid(alpha=0.3)
    plt.colorbar(scatter, ax=axes[1, 0], label='Intensity')
    
    # Classification pie chart
    class_counts = measurements_df['fluorescence_class'].value_counts()
    colors_pie = ['#2ecc71', '#3498db']
    axes[1, 1].pie(class_counts.values, labels=class_counts.index, autopct='%1.1f%%',
                   colors=colors_pie, startangle=90, 
                   textprops={'fontsize': 10, 'fontweight': 'bold'})
    axes[1, 1].set_title('Fluorescence Classification', fontsize=11, fontweight='bold')
    
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    chart_images['distributions'] = buf
    plt.close(fig)
    
    # Chart 3: Top bacteria
    fig, ax = plt.subplots(1, 1, figsize=(10, 6))
    top10 = measurements_df.nlargest(10, 'mean_intensity')[['label', 'mean_intensity']]
    ax.barh(top10['label'].astype(str), top10['mean_intensity'], 
            color='orangered', edgecolor='black', alpha=0.7)
    ax.set_xlabel('Mean Intensity', fontsize=11)
    ax.set_ylabel('Bacteria Label', fontsize=11)
    ax.set_title('Top 10 Brightest Bacteria', fontsize=12, fontweight='bold')
    ax.grid(axis='x', alpha=0.3)
    ax.invert_yaxis()
    
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    chart_images['top_bacteria'] = buf
    plt.close(fig)
    
    return chart_images


def export_to_excel_with_charts(measurements_df, excel_path, image_ch00, image_ch01, labeled_bacteria):
    """Export measurements to formatted Excel file with embedded charts"""
    
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Create chart images
    print("Generating charts for Excel...")
    chart_images = create_chart_images(measurements_df, image_ch00, image_ch01, labeled_bacteria)
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        workbook = writer.book
        
        # Sheet 1: Summary Dashboard (text only, no images)
        summary_sheet = workbook.create_sheet('Summary Dashboard', 0)
        
        # Add title
        summary_sheet['A1'] = 'BACTERIA SCREENING ANALYSIS REPORT'
        summary_sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        summary_sheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        summary_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        summary_sheet.merge_cells('A1:F1')
        summary_sheet.row_dimensions[1].height = 30
        
        # Add summary statistics
        summary_data = [
            ['Metric', 'Value'],
            ['Total Bacteria Detected', len(measurements_df)],
            ['High Fluorescence Count', len(measurements_df[measurements_df['fluorescence_class'] == 'High'])],
            ['Low Fluorescence Count', len(measurements_df[measurements_df['fluorescence_class'] == 'Low'])],
            ['Mean Area (pixels²)', f"{measurements_df['area'].mean():.2f}"],
            ['Std Area (pixels²)', f"{measurements_df['area'].std():.2f}"],
            ['Mean Intensity', f"{measurements_df['mean_intensity'].mean():.2f}"],
            ['Std Intensity', f"{measurements_df['mean_intensity'].std():.2f}"],
            ['Median Intensity', f"{measurements_df['mean_intensity'].median():.2f}"],
            ['Mean Aspect Ratio', f"{measurements_df['aspect_ratio'].mean():.2f}"],
            ['Mean Circularity', f"{measurements_df['circularity'].mean():.2f}"],
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 1:  # Metric column
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Value column
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths for summary sheet
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 25
        
        # Add note about visualizations
        note_row = len(summary_data) + 5
        summary_sheet[f'A{note_row}'] = 'Note: Visualizations are available in separate worksheets →'
        summary_sheet[f'A{note_row}'].font = Font(size=11, italic=True, color='0066CC')
        summary_sheet.merge_cells(f'A{note_row}:B{note_row}')
        
        # Sheet 2: Visual - Original Images
        images_sheet = workbook.create_sheet('Visuals - Original Images', 1)
        images_sheet['A1'] = 'Original Images Comparison'
        images_sheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        images_sheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        images_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        images_sheet.merge_cells('A1:F1')
        images_sheet.row_dimensions[1].height = 25
        
        img = XLImage(chart_images['original_images'])
        img.width = int(img.width * 1.10)
        img.height = int(img.height * 1.10)
        images_sheet.add_image(img, 'A3')
        
        # Sheet 3: Visual - Distributions
        dist_sheet = workbook.create_sheet('Visuals - Distributions', 2)
        dist_sheet['A1'] = 'Statistical Distributions & Analysis'
        dist_sheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        dist_sheet['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        dist_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        dist_sheet.merge_cells('A1:F1')
        dist_sheet.row_dimensions[1].height = 25
        
        img = XLImage(chart_images['distributions'])
        img.width = 900
        img.height = int(img.height * (900 / img.width))
        dist_sheet.add_image(img, 'A3')
        
        # Sheet 4: Visual - Top Bacteria
        top_sheet = workbook.create_sheet('Visuals - Top Bacteria', 3)
        top_sheet['A1'] = 'Top 10 Brightest Bacteria'
        top_sheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        top_sheet['A1'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        top_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        top_sheet.merge_cells('A1:F1')
        top_sheet.row_dimensions[1].height = 25
        
        img = XLImage(chart_images['top_bacteria'])
        img.width = 750
        img.height = int(img.height * (750 / img.width))
        top_sheet.add_image(img, 'A3')
        
        # Sheet 5: All measurements
        measurements_df.to_excel(writer, sheet_name='Data - All Measurements', index=False)
        
        # Format measurements sheet
        ws_measurements = writer.sheets['Data - All Measurements']
        for cell in ws_measurements[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths for measurements
        for idx in range(1, ws_measurements.max_column + 1):
            column_letter = get_column_letter(idx)
            max_length = 0
            for row in range(1, min(ws_measurements.max_row + 1, 100)):
                cell = ws_measurements.cell(row=row, column=idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_measurements.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 6: Summary statistics
        summary_stats = measurements_df.describe().T
        summary_stats.to_excel(writer, sheet_name='Data - Summary Statistics')
        
        ws_stats = writer.sheets['Data - Summary Statistics']
        for cell in ws_stats[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths for stats
        for idx in range(1, ws_stats.max_column + 1):
            column_letter = get_column_letter(idx)
            max_length = 0
            for row in range(1, ws_stats.max_row + 1):
                cell = ws_stats.cell(row=row, column=idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_stats.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 7: Classification breakdown
        classification = measurements_df.groupby('fluorescence_class').agg({
            'label': 'count',
            'area': ['mean', 'std'],
            'mean_intensity': ['mean', 'std'],
            'aspect_ratio': ['mean', 'std']
        }).round(3)
        classification.to_excel(writer, sheet_name='Data - Classification')
        
        ws_class = writer.sheets['Data - Classification']
        for idx in range(1, ws_class.max_column + 1):
            column_letter = get_column_letter(idx)
            max_length = 0
            for row in range(1, ws_class.max_row + 1):
                cell = ws_class.cell(row=row, column=idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_class.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 8: Top performers
        top_bacteria = measurements_df.nlargest(20, 'mean_intensity')
        top_bacteria.to_excel(writer, sheet_name='Data - Top 20 Intensity', index=False)
        
        ws_top = writer.sheets['Data - Top 20 Intensity']
        for cell in ws_top[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths for top bacteria
        for idx in range(1, ws_top.max_column + 1):
            column_letter = get_column_letter(idx)
            max_length = 0
            for row in range(1, ws_top.max_row + 1):
                cell = ws_top.cell(row=row, column=idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_top.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✓ Exported measurements with charts to Excel: {excel_path}")


def generate_summary_report(measurements_df, summary_path):
    """Generate text summary report"""
    
    with open(summary_path, 'w') as f:
        f.write("=" * 70 + "\n")
        f.write("BACTERIA SCREENING ANALYSIS SUMMARY REPORT\n")
        f.write("=" * 70 + "\n\n")
        
        f.write(f"Total Bacteria Detected: {len(measurements_df)}\n\n")
        
        f.write("CLASSIFICATION:\n")
        f.write("-" * 70 + "\n")
        class_counts = measurements_df['fluorescence_class'].value_counts()
        for class_name, count in class_counts.items():
            percentage = (count / len(measurements_df)) * 100
            f.write(f"  {class_name} Fluorescence: {count} ({percentage:.1f}%)\n")
        
        f.write("\nMORPHOLOGY STATISTICS:\n")
        f.write("-" * 70 + "\n")
        f.write(f"  Mean Area: {measurements_df['area'].mean():.2f} ± {measurements_df['area'].std():.2f} pixels²\n")
        f.write(f"  Mean Aspect Ratio: {measurements_df['aspect_ratio'].mean():.2f} ± {measurements_df['aspect_ratio'].std():.2f}\n")
        f.write(f"  Mean Circularity: {measurements_df['circularity'].mean():.2f} ± {measurements_df['circularity'].std():.2f}\n")
        
        f.write("\nFLUORESCENCE STATISTICS:\n")
        f.write("-" * 70 + "\n")
        f.write(f"  Mean Intensity: {measurements_df['mean_intensity'].mean():.2f} ± {measurements_df['mean_intensity'].std():.2f}\n")
        f.write(f"  Median Intensity: {measurements_df['mean_intensity'].median():.2f}\n")
        f.write(f"  Max Intensity: {measurements_df['max_intensity'].max():.2f}\n")
        
        f.write("\nTOP 5 BRIGHTEST BACTERIA:\n")
        f.write("-" * 70 + "\n")
        top5 = measurements_df.nlargest(5, 'mean_intensity')
        for idx, row in top5.iterrows():
            f.write(f"  #{int(row['label'])}: Intensity = {row['mean_intensity']:.2f}, Area = {row['area']:.1f} px²\n")
        
        f.write("\n" + "=" * 70 + "\n")
    
    print(f"✓ Summary report saved: {summary_path}")


def create_consolidated_summary_plots(all_measurements, output_folder):
    """Create consolidated plots across all images"""
    
    output_path = Path(output_folder)
    
    # Create comprehensive consolidated figure
    fig = plt.figure(figsize=(20, 12))
    gs = fig.add_gridspec(3, 3, hspace=0.3, wspace=0.3)
    
    # Plot 1: Bacteria count per image
    ax1 = fig.add_subplot(gs[0, 0])
    count_per_image = all_measurements.groupby('image_name').size()
    ax1.bar(range(len(count_per_image)), count_per_image.values, 
            color='steelblue', edgecolor='black', alpha=0.7)
    ax1.set_xlabel('Image Number', fontsize=11, fontweight='bold')
    ax1.set_ylabel('Bacteria Count', fontsize=11, fontweight='bold')
    ax1.set_title('Bacteria Count per Image', fontsize=12, fontweight='bold')
    ax1.set_xticks(range(len(count_per_image)))
    ax1.set_xticklabels([f"Img {i+1}" for i in range(len(count_per_image))], rotation=45)
    ax1.grid(axis='y', alpha=0.3)
    
    # Plot 2: Overall area distribution
    ax2 = fig.add_subplot(gs[0, 1])
    ax2.hist(all_measurements['area'], bins=30, color='forestgreen', 
             edgecolor='black', alpha=0.7)
    ax2.axvline(all_measurements['area'].mean(), color='red', 
                linestyle='--', linewidth=2, label=f"Mean: {all_measurements['area'].mean():.1f}")
    ax2.set_xlabel('Area (pixels²)', fontsize=11, fontweight='bold')
    ax2.set_ylabel('Frequency', fontsize=11, fontweight='bold')
    ax2.set_title('Overall Size Distribution', fontsize=12, fontweight='bold')
    ax2.legend()
    ax2.grid(axis='y', alpha=0.3)
    
    # Plot 3: Overall intensity distribution
    ax3 = fig.add_subplot(gs[0, 2])
    ax3.hist(all_measurements['mean_intensity'], bins=30, 
             color='orangered', edgecolor='black', alpha=0.7)
    ax3.axvline(all_measurements['mean_intensity'].median(), color='blue', 
                linestyle='--', linewidth=2, 
                label=f"Median: {all_measurements['mean_intensity'].median():.1f}")
    ax3.set_xlabel('Mean Intensity', fontsize=11, fontweight='bold')
    ax3.set_ylabel('Frequency', fontsize=11, fontweight='bold')
    ax3.set_title('Overall Intensity Distribution', fontsize=12, fontweight='bold')
    ax3.legend()
    ax3.grid(axis='y', alpha=0.3)
    
    # Plot 4: Fluorescence classification per image
    ax4 = fig.add_subplot(gs[1, 0])
    class_pivot = all_measurements.groupby(['image_name', 'fluorescence_class']).size().unstack(fill_value=0)
    class_pivot.plot(kind='bar', stacked=True, ax=ax4, 
                     color=['#3498db', '#2ecc71'], alpha=0.8, edgecolor='black')
    ax4.set_xlabel('Image', fontsize=11, fontweight='bold')
    ax4.set_ylabel('Bacteria Count', fontsize=11, fontweight='bold')
    ax4.set_title('High vs Low Fluorescence by Image', fontsize=12, fontweight='bold')
    ax4.legend(title='Classification', loc='upper right')
    ax4.set_xticklabels(ax4.get_xticklabels(), rotation=45, ha='right')
    ax4.grid(axis='y', alpha=0.3)
    
    # Plot 5: Size vs Intensity scatter (all data)
    ax5 = fig.add_subplot(gs[1, 1])
    scatter = ax5.scatter(all_measurements['area'], 
                         all_measurements['mean_intensity'],
                         c=all_measurements['mean_intensity'],
                         cmap='hot', s=40, alpha=0.6, edgecolors='black', linewidth=0.5)
    ax5.set_xlabel('Area (pixels²)', fontsize=11, fontweight='bold')
    ax5.set_ylabel('Mean Intensity', fontsize=11, fontweight='bold')
    ax5.set_title('Size vs Intensity (All Images)', fontsize=12, fontweight='bold')
    ax5.grid(alpha=0.3)
    plt.colorbar(scatter, ax=ax5, label='Intensity')
    
    # Plot 6: Overall classification pie chart
    ax6 = fig.add_subplot(gs[1, 2])
    overall_class = all_measurements['fluorescence_class'].value_counts()
    colors = ['#2ecc71', '#3498db']
    ax6.pie(overall_class.values, labels=overall_class.index, autopct='%1.1f%%',
            colors=colors, startangle=90, 
            textprops={'fontsize': 11, 'fontweight': 'bold'},
            explode=(0.05, 0.05))
    ax6.set_title('Overall Fluorescence Classification', fontsize=12, fontweight='bold')
    
    # Plot 7: Average metrics per image
    ax7 = fig.add_subplot(gs[2, 0])
    avg_metrics = all_measurements.groupby('image_name').agg({
        'area': 'mean',
        'mean_intensity': 'mean'
    })
    x = np.arange(len(avg_metrics))
    width = 0.35
    ax7_twin = ax7.twinx()
    
    bars1 = ax7.bar(x - width/2, avg_metrics['area'], width, 
                    label='Avg Area', color='steelblue', alpha=0.7, edgecolor='black')
    bars2 = ax7_twin.bar(x + width/2, avg_metrics['mean_intensity'], width,
                         label='Avg Intensity', color='orangered', alpha=0.7, edgecolor='black')
    
    ax7.set_xlabel('Image', fontsize=11, fontweight='bold')
    ax7.set_ylabel('Average Area (pixels²)', fontsize=10, fontweight='bold', color='steelblue')
    ax7_twin.set_ylabel('Average Intensity', fontsize=10, fontweight='bold', color='orangered')
    ax7.set_title('Average Metrics per Image', fontsize=12, fontweight='bold')
    ax7.set_xticks(x)
    ax7.set_xticklabels([f"Img {i+1}" for i in range(len(avg_metrics))], rotation=45)
    ax7.tick_params(axis='y', labelcolor='steelblue')
    ax7_twin.tick_params(axis='y', labelcolor='orangered')
    ax7.grid(axis='y', alpha=0.3)
    
    # Combine legends
    lines1, labels1 = ax7.get_legend_handles_labels()
    lines2, labels2 = ax7_twin.get_legend_handles_labels()
    ax7.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
    
    # Plot 8: Top 20 brightest bacteria across all images
    ax8 = fig.add_subplot(gs[2, 1:])
    top20 = all_measurements.nlargest(20, 'mean_intensity')
    colors_top = ['orangered' if c == 'High' else 'dodgerblue' 
                  for c in top20['fluorescence_class']]
    bars = ax8.barh(range(len(top20)), top20['mean_intensity'].values,
                    color=colors_top, edgecolor='black', alpha=0.8)
    ax8.set_yticks(range(len(top20)))
    ax8.set_yticklabels([f"{row['image_name'][:10]}..#{int(row['label'])}" 
                         for _, row in top20.iterrows()], fontsize=9)
    ax8.set_xlabel('Mean Intensity', fontsize=11, fontweight='bold')
    ax8.set_title('Top 20 Brightest Bacteria (All Images)', fontsize=12, fontweight='bold')
    ax8.invert_yaxis()
    ax8.grid(axis='x', alpha=0.3)
    
    # Add legend for top 20
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor='orangered', label='High Fluorescence'),
        Patch(facecolor='dodgerblue', label='Low Fluorescence')
    ]
    ax8.legend(handles=legend_elements, loc='lower right')
    
    plt.suptitle('CONSOLIDATED BACTERIA SCREENING ANALYSIS - ALL IMAGES', 
                 fontsize=16, fontweight='bold', y=0.995)
    
    consolidated_plot_path = output_path / 'consolidated_analysis_plots.png'
    plt.savefig(consolidated_plot_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    
    print(f"✓ Consolidated plots saved: {consolidated_plot_path}")
    
    return str(consolidated_plot_path)


def generate_consolidated_report(all_measurements, all_results, output_folder):
    """Generate comprehensive consolidated text report"""
    
    output_path = Path(output_folder)
    report_path = output_path / 'CONSOLIDATED_SUMMARY_REPORT.txt'
    
    with open(report_path, 'w') as f:
        f.write("=" * 80 + "\n")
        f.write("CONSOLIDATED BACTERIA SCREENING ANALYSIS REPORT\n")
        f.write("ALL IMAGE PAIRS (1 N NO 1 - 1 N NO 5)\n")
        f.write("=" * 80 + "\n\n")
        
        # Overall summary
        f.write("OVERALL SUMMARY:\n")
        f.write("-" * 80 + "\n")
        f.write(f"Total Images Analyzed: {len(all_results)}\n")
        f.write(f"Total Bacteria Detected: {len(all_measurements)}\n")
        f.write(f"Average Bacteria per Image: {len(all_measurements) / len(all_results):.1f}\n\n")
        
        # Per-image breakdown
        f.write("PER-IMAGE BREAKDOWN:\n")
        f.write("-" * 80 + "\n")
        count_per_image = all_measurements.groupby('image_name').size()
        for idx, (img_name, count) in enumerate(count_per_image.items(), 1):
            high_count = len(all_measurements[(all_measurements['image_name'] == img_name) & 
                                             (all_measurements['fluorescence_class'] == 'High')])
            low_count = len(all_measurements[(all_measurements['image_name'] == img_name) & 
                                            (all_measurements['fluorescence_class'] == 'Low')])
            f.write(f"  Image {idx} ({img_name}):\n")
            f.write(f"    Total: {count} bacteria\n")
            f.write(f"    High Fluorescence: {high_count} ({high_count/count*100:.1f}%)\n")
            f.write(f"    Low Fluorescence: {low_count} ({low_count/count*100:.1f}%)\n\n")
        
        # Overall classification
        f.write("OVERALL CLASSIFICATION:\n")
        f.write("-" * 80 + "\n")
        overall_class = all_measurements['fluorescence_class'].value_counts()
        for class_name, count in overall_class.items():
            percentage = (count / len(all_measurements)) * 100
            f.write(f"  {class_name} Fluorescence: {count} ({percentage:.1f}%)\n")
        f.write(f"  Intensity Threshold (median): {all_measurements['mean_intensity'].median():.2f}\n\n")
        
        # Overall morphology statistics
        f.write("OVERALL MORPHOLOGY STATISTICS:\n")
        f.write("-" * 80 + "\n")
        f.write(f"  Mean Area: {all_measurements['area'].mean():.2f} ± {all_measurements['area'].std():.2f} pixels²\n")
        f.write(f"  Median Area: {all_measurements['area'].median():.2f} pixels²\n")
        f.write(f"  Area Range: {all_measurements['area'].min():.2f} - {all_measurements['area'].max():.2f} pixels²\n")
        f.write(f"  Mean Aspect Ratio: {all_measurements['aspect_ratio'].mean():.2f} ± {all_measurements['aspect_ratio'].std():.2f}\n")
        f.write(f"  Mean Circularity: {all_measurements['circularity'].mean():.2f} ± {all_measurements['circularity'].std():.2f}\n")
        f.write(f"  Mean Eccentricity: {all_measurements['eccentricity'].mean():.2f} ± {all_measurements['eccentricity'].std():.2f}\n\n")
        
        # Overall fluorescence statistics
        f.write("OVERALL FLUORESCENCE STATISTICS:\n")
        f.write("-" * 80 + "\n")
        f.write(f"  Mean Intensity: {all_measurements['mean_intensity'].mean():.2f} ± {all_measurements['mean_intensity'].std():.2f}\n")
        f.write(f"  Median Intensity: {all_measurements['mean_intensity'].median():.2f}\n")
        f.write(f"  Intensity Range: {all_measurements['mean_intensity'].min():.2f} - {all_measurements['mean_intensity'].max():.2f}\n")
        f.write(f"  Mean Integrated Intensity: {all_measurements['integrated_intensity'].mean():.2f}\n\n")
        
        # Top performers
        f.write("TOP 10 BRIGHTEST BACTERIA (ALL IMAGES):\n")
        f.write("-" * 80 + "\n")
        top10 = all_measurements.nlargest(10, 'mean_intensity')
        for rank, (idx, row) in enumerate(top10.iterrows(), 1):
            f.write(f"  #{rank}: {row['image_name']} - Bacteria #{int(row['label'])}\n")
            f.write(f"       Intensity: {row['mean_intensity']:.2f}, Area: {row['area']:.1f} px²\n")
        
        # Statistical comparison
        f.write("\n" + "STATISTICAL COMPARISON BY IMAGE:\n")
        f.write("-" * 80 + "\n")
        stats_by_image = all_measurements.groupby('image_name').agg({
            'area': ['mean', 'std'],
            'mean_intensity': ['mean', 'std'],
            'circularity': 'mean'
        }).round(2)
        f.write(stats_by_image.to_string())
        f.write("\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("END OF CONSOLIDATED REPORT\n")
        f.write("=" * 80 + "\n")
    
    print(f"✓ Consolidated report saved: {report_path}")
    
    return str(report_path)


def batch_analyze_specific_images(base_folder, image_numbers, output_folder='./outputs'):
    """
    Batch process specific numbered image pairs
    
    Args:
        base_folder (str): Base folder containing the images
        image_numbers (list): List of image numbers to process (e.g., [1, 2, 3, 4, 5])
        output_folder (str): Folder to save all results
    
    Returns:
        dict: Consolidated results
    """
    base_path = Path(base_folder)
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)
    
    all_results = []
    all_measurements = []
    
    print("\n" + "="*80)
    print("STARTING BATCH ANALYSIS FOR MULTIPLE IMAGE PAIRS")
    print("="*80 + "\n")
    
    for img_num in image_numbers:
        # Construct file paths
        ch00_path = base_path / f'1 N NO {img_num}_ch00.tif'
        ch01_path = base_path / f'1 N NO {img_num}_ch01.tif'
        
        # Check if both files exist
        if not ch00_path.exists():
            print(f"⚠ Warning: {ch00_path.name} not found, skipping...")
            continue
        
        if not ch01_path.exists():
            print(f"⚠ Warning: {ch01_path.name} not found, skipping...")
            continue
        
        print(f"\n{'='*80}")
        print(f"Processing Image Pair #{img_num}: {ch00_path.name}")
        print(f"{'='*80}")
        
        # Analyze
        results = analyze_bacteria_images(str(ch00_path), str(ch01_path), output_folder)
        
        if 'error' not in results:
            all_results.append(results)
            all_measurements.append(results['measurements'])
            print(f"✓ Successfully processed image pair #{img_num}")
        else:
            print(f"✗ Error processing image pair #{img_num}: {results['error']}")
    
    # Combine all measurements
    if all_measurements:
        print(f"\n{'='*80}")
        print("CREATING CONSOLIDATED OUTPUTS")
        print(f"{'='*80}\n")
        
        combined_measurements = pd.concat(all_measurements, ignore_index=True)
        
        # Save consolidated CSV
        combined_csv = output_path / 'CONSOLIDATED_all_measurements.csv'
        combined_measurements.to_csv(combined_csv, index=False)
        print(f"✓ Consolidated CSV saved: {combined_csv}")
        
        # Save consolidated Excel with multiple sheets
        combined_excel = output_path / 'CONSOLIDATED_all_measurements.xlsx'
        
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(combined_excel, engine='openpyxl') as writer:
            # Sheet 1: All measurements
            combined_measurements.to_excel(writer, sheet_name='All Measurements', index=False)
            
            # Format header
            ws = writer.sheets['All Measurements']
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust columns
            for idx in range(1, ws.max_column + 1):
                column_letter = get_column_letter(idx)
                max_length = 0
                for row in range(1, min(ws.max_row + 1, 100)):
                    cell = ws.cell(row=row, column=idx)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Sheet 2: Summary statistics
            summary_stats = combined_measurements.describe().T
            summary_stats.to_excel(writer, sheet_name='Summary Statistics')
            
            ws_stats = writer.sheets['Summary Statistics']
            for cell in ws_stats[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Sheet 3: Classification by image
            classification = combined_measurements.groupby(['image_name', 'fluorescence_class']).size().unstack(fill_value=0)
            classification.to_excel(writer, sheet_name='Classification by Image')
            
            # Sheet 4: Average metrics by image
            avg_metrics = combined_measurements.groupby('image_name').agg({
                'label': 'count',
                'area': ['mean', 'std'],
                'mean_intensity': ['mean', 'std'],
                'circularity': ['mean', 'std'],
                'aspect_ratio': ['mean', 'std']
            }).round(3)
            avg_metrics.columns = ['_'.join(col).strip() for col in avg_metrics.columns.values]
            avg_metrics.to_excel(writer, sheet_name='Metrics by Image')
            
            # Sheet 5: Top 50 brightest bacteria
            top50 = combined_measurements.nlargest(50, 'mean_intensity')
            top50.to_excel(writer, sheet_name='Top 50 Brightest', index=False)
            
            ws_top = writer.sheets['Top 50 Brightest']
            for cell in ws_top[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        print(f"✓ Consolidated Excel saved: {combined_excel}")
        
        # Create consolidated plots
        plot_path = create_consolidated_summary_plots(combined_measurements, output_folder)
        
        # Generate consolidated report
        report_path = generate_consolidated_report(combined_measurements, all_results, output_folder)
        
        # Final summary
        print(f"\n{'='*80}")
        print("BATCH PROCESSING COMPLETE - FINAL SUMMARY")
        print(f"{'='*80}")
        print(f"Total images processed: {len(all_results)}")
        print(f"Total bacteria detected: {len(combined_measurements)}")
        print(f"Average bacteria per image: {len(combined_measurements) / len(all_results):.1f}")
        print(f"\nHigh fluorescence: {len(combined_measurements[combined_measurements['fluorescence_class']=='High'])} ({len(combined_measurements[combined_measurements['fluorescence_class']=='High'])/len(combined_measurements)*100:.1f}%)")
        print(f"Low fluorescence: {len(combined_measurements[combined_measurements['fluorescence_class']=='Low'])} ({len(combined_measurements[combined_measurements['fluorescence_class']=='Low'])/len(combined_measurements)*100:.1f}%)")
        print(f"\nAll results saved to: {output_folder}")
        print(f"{'='*80}\n")
        
        return {
            'total_images': len(all_results),
            'total_bacteria': len(combined_measurements),
            'combined_measurements': combined_measurements,
            'individual_results': all_results,
            'consolidated_csv': str(combined_csv),
            'consolidated_excel': str(combined_excel),
            'consolidated_plots': plot_path,
            'consolidated_report': report_path
        }
    else:
        print("\n✗ No images were successfully processed!")
        return {'error': 'No images processed successfully'}


if __name__ == '__main__':
    # Analyze all 5 image pairs (1 N NO 1 through 1 N NO 5)
    base_folder = './PD image/1'
    image_numbers = [1, 2, 3, 4, 5]  # Process images 1 through 5
    output_folder = './outputs'
    
    consolidated_results = batch_analyze_specific_images(
        base_folder=base_folder,
        image_numbers=image_numbers,
        output_folder=output_folder
    )
    
    if 'error' not in consolidated_results:
        print("\n" + "="*80)
        print("ALL OUTPUTS GENERATED:")
        print("="*80)
        print(f"1. Individual tagged images: {output_folder}/*_tagged.png")
        print(f"2. Individual Excel reports: {output_folder}/*_measurements.xlsx")
        print(f"3. Individual CSV files: {output_folder}/*_measurements.csv")
        print(f"4. Individual summaries: {output_folder}/*_summary.txt")
        print(f"5. Consolidated CSV: {consolidated_results['consolidated_csv']}")
        print(f"6. Consolidated Excel: {consolidated_results['consolidated_excel']}")
        print(f"7. Consolidated plots: {consolidated_results['consolidated_plots']}")
        print(f"8. Consolidated report: {consolidated_results['consolidated_report']}")
        print("="*80 + "\n")