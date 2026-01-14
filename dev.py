import os
import cv2
import numpy as np
import sys
import csv
import textwrap
import time

import atexit
import re
from datetime import datetime
from pathlib import Path
import xml.etree.ElementTree as ET
from typing import Optional, Tuple, Any, cast

from tqdm import tqdm
import pandas as pd
from scipy import stats as scipy_stats

# --- NEW IMPORT FOR REGISTRATION ---
from skimage.registration import phase_cross_correlation

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.series_factory import SeriesFactory

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


# ==================================================
# Logging: tee stdout/stderr to a file
# ==================================================
class Tee:
    def __init__(self, *streams):
        self.streams = streams

    def write(self, data: str) -> None:
        for s in self.streams:
            s.write(data)
            s.flush()

    def flush(self) -> None:
        for s in self.streams:
            s.flush()

def get_project_root() -> Path:
    """Get project root (works as script and as .exe)"""
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).resolve().parent
    else:
        return Path(__file__).resolve().parent

_project_root = get_project_root()
_logs_dir = _project_root / "logs"
_logs_dir.mkdir(exist_ok=True)
_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
_script_name = Path(sys.argv[0]).stem
_log_path = _logs_dir / f"run_{_timestamp}_{_script_name}.txt"
_log_file = open(_log_path, "w", encoding="utf-8")

sys.stdout = Tee(sys.stdout, _log_file)
sys.stderr = Tee(sys.stderr, _log_file)
print(f"Saving output to: {_log_path}")
print(f"Project root: {_project_root.resolve()}")
print(f"Running as: {'EXECUTABLE' if getattr(sys, 'frozen', False) else 'SCRIPT'}")


@atexit.register
def _close_log_file() -> None:
    try:
        _log_file.close()
    except Exception:
        pass

# ==================================================
# Configuration
# ==================================================
SOURCE_DIR = Path("./source")
CONTROL_DIR = None  # Will be set dynamically

# Segment only brightfield channel
IMAGE_GLOB = "*_ch00.tif"

# OUTPUT_DIR will be set dynamically in main()
OUTPUT_DIR: Optional[Path] = None


# Scale bar parameters
SCALE_BAR_LENGTH_UM = 10
SCALE_BAR_HEIGHT = 4
SCALE_BAR_MARGIN = 15
SCALE_BAR_COLOR = (255, 255, 255)
SCALE_BAR_BG_COLOR = (0, 0, 0)
SCALE_BAR_TEXT_COLOR = (255, 255, 255)
SCALE_BAR_FONT_SCALE = 0.5
SCALE_BAR_FONT_THICKNESS = 1

# --- Segmentation ---
GAUSSIAN_SIGMA = 15
MORPH_KERNEL_SIZE = 3
MORPH_ITERATIONS = 1
DILATE_ITERATIONS = 1
ERODE_ITERATIONS = 1

# Filtering (in micrometers)
MIN_AREA_UM2 = 3.0
MAX_AREA_UM2 = 2000.0
MIN_CIRCULARITY = 0.0
MAX_FRACTION_OF_IMAGE_AREA = 0.25

# --- Fluorescence segmentation (S2) ---
FLUOR_GAUSSIAN_SIGMA = 1.5
FLUOR_MORPH_KERNEL_SIZE = 3
FLUOR_MIN_AREA_UM2 = 3.0
FLUOR_MATCH_MIN_INTERSECTION_PX = 5.0

# Debug options
CLEAR_OUTPUT_DIR_EACH_RUN = True
SEPARATE_OUTPUT_BY_GROUP = True
FALLBACK_UM_PER_PX: Optional[float] = 0.109492


# ==================================================
# Helper Functions
# ==================================================
def logged_input(prompt: str) -> str:
    """Input function that logs both prompt and user response"""
    print(prompt, end='', flush=True)
    user_input = input()
    
    if user_input.strip():
        print(user_input)
    else:
        print("(pressed Enter)")
    
    return user_input


def clear_output_dir(folder: Path) -> None:
    for p in folder.glob("*"):
        try:
            if p.is_file():
                p.unlink()
            elif p.is_dir():
                for q in p.rglob("*"):
                    try:
                        if q.is_file():
                            q.unlink()
                    except Exception:
                        pass
        except Exception:
            pass


def add_scale_bar(
    img: np.ndarray, pixel_size: float, unit: str = "um", length_um: float = 10
) -> np.ndarray:
    """Add a scale bar to the image"""
    if pixel_size is None or pixel_size <= 0:
        return img

    bar_length_px = int(round(length_um / pixel_size))
    if bar_length_px < 10:
        return img

    h, w = img.shape[:2]
    bar_x = w - bar_length_px - SCALE_BAR_MARGIN
    bar_y = h - SCALE_BAR_HEIGHT - SCALE_BAR_MARGIN

    if bar_x < 0 or bar_y < 0:
        return img

    label = (
        f"{int(length_um)} um"
        if unit in ["µm", "um"]
        else f"{int(length_um)} {unit}"
    )

    font = cv2.FONT_HERSHEY_SIMPLEX
    (text_w, text_h), baseline = cv2.getTextSize(
        label, font, SCALE_BAR_FONT_SCALE, SCALE_BAR_FONT_THICKNESS
    )

    text_x = bar_x + (bar_length_px - text_w) // 2
    text_y = bar_y - 8

    bg_padding = 5
    bg_x1 = min(bar_x, text_x) - bg_padding
    bg_y1 = text_y - text_h - bg_padding
    bg_x2 = max(bar_x + bar_length_px, text_x + text_w) + bg_padding
    bg_y2 = bar_y + SCALE_BAR_HEIGHT + bg_padding

    img = img.copy()
    overlay = img.copy()
    cv2.rectangle(overlay, (bg_x1, bg_y1), (bg_x2, bg_y2), SCALE_BAR_BG_COLOR, -1)
    cv2.addWeighted(overlay, 0.7, img, 0.3, 0, img)

    cv2.rectangle(
        img,
        (bar_x, bar_y),
        (bar_x + bar_length_px, bar_y + SCALE_BAR_HEIGHT),
        SCALE_BAR_COLOR,
        -1,
    )

    cv2.putText(
        img,
        label,
        (text_x, text_y),
        font,
        SCALE_BAR_FONT_SCALE,
        SCALE_BAR_TEXT_COLOR,
        SCALE_BAR_FONT_THICKNESS,
        cv2.LINE_AA,
    )

    return img


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def save_debug(
    folder: Path,
    name: str,
    img: np.ndarray,
    pixel_size_um: Optional[float] = None,
) -> None:
    """Save debug image with optional scale bar - memory optimized"""
    out = folder / name
    
    if pixel_size_um is not None and pixel_size_um > 0:
        img_to_save = add_scale_bar(
            img.copy(),
            float(pixel_size_um), "um", SCALE_BAR_LENGTH_UM
        )
    else:
        img_to_save = img
    
    cv2.imwrite(str(out), img_to_save)
    
    if img_to_save.nbytes > 10_000_000:
        del img_to_save


def list_sample_group_folders(source_dir: Path) -> list[Path]:
    groups: list[Path] = []
    if not source_dir.exists():
        raise FileNotFoundError(f"Source folder not found: {source_dir.resolve()}")

    for p in source_dir.iterdir():
        if not p.is_dir():
            continue
        if p.name.lower().startswith("control"):
            continue
        if re.fullmatch(r"\d+", p.name):
            groups.append(p)

    groups.sort(key=lambda x: int(x.name))
    return groups


def _display_group_name(name: str) -> str:
    return "Control" if name.lower().startswith("control") else name


def _group_order_key(g: str) -> tuple[int, int]:
    """Sort numeric groups ascending, Control last."""
    if g == "Control":
        return (1, 10**9)
    if g.isdigit():
        return (0, int(g))
    return (0, 10**8)


def find_metadata_paths(img_path: Path) -> tuple[Optional[Path], Optional[Path]]:
    base = img_path.stem
    if base.endswith("_ch00"):
        base = base[:-5]
    md_dir = img_path.parent / "MetaData"
    xml_main = md_dir / f"{base}.xml"
    xml_props = md_dir / f"{base}_Properties.xml"

    return (
        xml_props if xml_props.exists() else None,
        xml_main if xml_main.exists() else None,
    )


def _require_attr(elem: ET.Element, attr: str, context: str) -> str:
    v = elem.get(attr)
    if v is None:
        raise ValueError(f"Missing attribute '{attr}' in {context}")
    return v


def _parse_float(s: str) -> float:
    return float(s.strip().replace(",", "."))


def get_pixel_size_um(
    xml_props_path: Optional[Path],
    xml_main_path: Optional[Path],
) -> Tuple[float, float]:
    """Extract pixel size with detailed error reporting"""
    
    errors = []
    
    if xml_props_path is not None:
        try:
            tree = ET.parse(xml_props_path)
            root = tree.getroot()

            dims = root.findall(".//ImageDescription/Dimensions/DimensionDescription")
            by_id = {d.get("DimID"): d for d in dims}

            def read_dim(dim_id: str) -> Tuple[float, int, str]:
                d = by_id.get(dim_id)
                if d is None:
                    raise ValueError(
                        f"Missing DimensionDescription with DimID='{dim_id}' in {xml_props_path.name}"
                    )

                length_s = _require_attr(
                    d, "Length", f"{xml_props_path.name} DimID={dim_id}"
                )
                n_s = _require_attr(
                    d, "NumberOfElements", f"{xml_props_path.name} DimID={dim_id}"
                )
                unit = _require_attr(d, "Unit", f"{xml_props_path.name} DimID={dim_id}")

                length = _parse_float(length_s)
                n = int(n_s)
                return length, n, unit

            x_len, x_n, x_unit = read_dim("X")
            y_len, y_n, y_unit = read_dim("Y")

            if x_unit != "µm" or y_unit != "µm":
                raise ValueError(
                    f"Unexpected units in {xml_props_path.name}: X={x_unit}, Y={y_unit}"
                )

            return float(x_len / x_n), float(y_len / y_n)
        except Exception as e:
            errors.append(f"Properties XML ({xml_props_path.name}): {e}")

    if xml_main_path is not None:
        try:
            tree = ET.parse(xml_main_path)
            root = tree.getroot()

            dims = root.findall(".//ImageDescription/Dimensions/DimensionDescription")
            by_id = {d.get("DimID"): d for d in dims}

            def read_dim(dim_id: str) -> Tuple[float, int, str]:
                d = by_id.get(dim_id)
                if d is None:
                    raise ValueError(
                        f"Missing DimensionDescription with DimID='{dim_id}' in {xml_main_path.name}"
                    )

                length_s = _require_attr(
                    d, "Length", f"{xml_main_path.name} DimID={dim_id}"
                )
                n_s = _require_attr(
                    d, "NumberOfElements", f"{xml_main_path.name} DimID={dim_id}"
                )
                unit = _require_attr(d, "Unit", f"{xml_main_path.name} DimID={dim_id}")

                length = _parse_float(length_s)
                n = int(n_s)
                return length, n, unit

            x_len_m, x_n, x_unit = read_dim("1")
            y_len_m, y_n, y_unit = read_dim("2")

            if x_unit != "m" or y_unit != "m":
                raise ValueError(
                    f"Unexpected units in {xml_main_path.name}: X={x_unit}, Y={y_unit}"
                )

            return float((x_len_m * 1e6) / x_n), float((y_len_m * 1e6) / y_n)
        except Exception as e:
            errors.append(f"Main XML ({xml_main_path.name}): {e}")
    
    error_summary = "\n  - ".join(errors) if errors else "No XML files provided"
    raise ValueError(
        f"Could not determine pixel size (µm/px).\nAttempted sources:\n  - {error_summary}"
    )


def contour_perimeter_um(contour: np.ndarray, um_per_px_x: float, um_per_px_y: float) -> float:
    pts = contour.reshape(-1, 2).astype(np.float64)
    pts[:, 0] *= float(um_per_px_x)
    pts[:, 1] *= float(um_per_px_y)
    d = np.diff(np.vstack([pts, pts[0]]), axis=0)
    seg = np.sqrt((d[:, 0] ** 2) + (d[:, 1] ** 2))
    return float(seg.sum())


def equivalent_diameter_from_area(area: float) -> float:
    return float(2.0 * np.sqrt(area / np.pi)) if area > 0 else 0.0


def normalize_to_8bit(img: np.ndarray) -> np.ndarray:
    if img.dtype == np.uint8:
        return img
    if img.dtype == np.uint16:
        out = np.zeros_like(img, dtype=np.uint8)
        cv2.normalize(img, out, 0, 255, cv2.NORM_MINMAX)
        return out
    img_f = img.astype(np.float32)
    mn, mx = float(np.min(img_f)), float(np.max(img_f))
    if mx <= mn:
        return np.zeros(img.shape, dtype=np.uint8)
    return ((img_f - mn) * (255.0 / (mx - mn))).clip(0, 255).astype(np.uint8)


def _put_text_outline(
    img: np.ndarray,
    text: str,
    org: tuple[int, int],
    font_scale: float = 0.5,
    color: tuple[int, int, int] = (255, 255, 255),
    thickness: int = 1,
) -> None:
    """Draw readable text with a black outline."""
    cv2.putText(
        img,
        text,
        org,
        cv2.FONT_HERSHEY_SIMPLEX,
        font_scale,
        (0, 0, 0),
        thickness + 2,
        cv2.LINE_AA,
    )
    cv2.putText(
        img,
        text,
        org,
        cv2.FONT_HERSHEY_SIMPLEX,
        font_scale,
        color,
        thickness,
        cv2.LINE_AA,
    )


def draw_object_ids(
    img_bgr: np.ndarray, contours: list[np.ndarray], labels: Optional[list[str]] = None
) -> np.ndarray:
    """Draw object labels at contour centroids."""
    out = img_bgr.copy()
    for i, c in enumerate(contours, 1):
        M = cv2.moments(c)
        if M["m00"] == 0:
            continue
        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])
        text = labels[i - 1] if (labels is not None and i - 1 < len(labels)) else str(i)
        _put_text_outline(out, text, (cx, cy), font_scale=0.5, color=(0, 255, 0), thickness=1)
    return out


def _ids_name(original_png: str) -> str:
    """Insert '_ids' before '.png'."""
    if original_png.lower().endswith(".png"):
        return original_png[:-4] + "_ids.png"
    return original_png + "_ids"


def save_debug_ids(
    folder: Path,
    original_name: str,
    img_bgr: np.ndarray,
    accepted_contours: list[np.ndarray],
    object_ids: list[str],
    pixel_size_um: Optional[float] = None,
) -> None:
    """Save a labeled (Object_ID) version of an existing debug view."""
    labeled = draw_object_ids(img_bgr, accepted_contours, labels=object_ids)
    save_debug(folder, _ids_name(original_name), labeled, pixel_size_um)


# ==================================================
# Fluorescence Registration / Alignment
# ==================================================
def align_fluorescence_channel(bf_img: np.ndarray, fluor_img: np.ndarray) -> tuple[np.ndarray, tuple[float, float]]:
    """Aligns the fluorescence image to the brightfield image using Phase Correlation."""
    if bf_img.ndim == 3:
        bf_gray = cv2.cvtColor(bf_img, cv2.COLOR_BGR2GRAY)
    else:
        bf_gray = bf_img

    if fluor_img.ndim == 3:
        fluor_gray = cv2.cvtColor(fluor_img, cv2.COLOR_BGR2GRAY)
    else:
        fluor_gray = fluor_img

    bf_inverted = cv2.bitwise_not(bf_gray)

    shift, error, diffphase = phase_cross_correlation(bf_inverted, fluor_gray, upsample_factor=10)
    shift_y, shift_x = shift

    rows, cols = fluor_img.shape[:2]
    
    M = np.array([[1, 0, -shift_x], [0, 1, -shift_y]], dtype=np.float32)
    
    aligned_fluor = cv2.warpAffine(fluor_img, M, (cols, rows))

    return aligned_fluor, (shift_y, shift_x)


# ==================================================
# Fluorescence segmentation + matching
# ==================================================
def segment_fluorescence_global(fluor_img8: np.ndarray) -> np.ndarray:
    """Segment fluorescence objects globally. Returns binary mask uint8 (0/255)."""
    blur = cv2.GaussianBlur(
        fluor_img8, (0, 0), sigmaX=FLUOR_GAUSSIAN_SIGMA, sigmaY=FLUOR_GAUSSIAN_SIGMA
    )
    
    otsu_threshold = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[0]
    
    THRESHOLD_MULTIPLIER = 0.5
    adjusted_threshold = otsu_threshold * THRESHOLD_MULTIPLIER
    
    print(f"  Fluorescence threshold: Otsu={otsu_threshold:.1f}, Adjusted={adjusted_threshold:.1f}")
    
    _, bw = cv2.threshold(blur, adjusted_threshold, 255, cv2.THRESH_BINARY)
    
    k = np.ones((FLUOR_MORPH_KERNEL_SIZE, FLUOR_MORPH_KERNEL_SIZE), np.uint8)
    bw = cv2.morphologyEx(bw, cv2.MORPH_OPEN, k, iterations=1)
    bw = cv2.morphologyEx(bw, cv2.MORPH_CLOSE, k, iterations=1)
    return bw


def contour_intersection_area_px(c1: np.ndarray, c2: np.ndarray, shape_hw: tuple[int, int]) -> float:
    """Intersection area in pixels between two contours."""
    H, W = shape_hw
    m1 = np.zeros((H, W), dtype=np.uint8)
    m2 = np.zeros((H, W), dtype=np.uint8)
    cv2.drawContours(m1, [c1], -1, 255, thickness=-1)
    cv2.drawContours(m2, [c2], -1, 255, thickness=-1)
    return float(np.count_nonzero(cv2.bitwise_and(m1, m2)))


def match_fluor_to_bf_by_overlap(
    bf_contours: list[np.ndarray],
    fluor_contours: list[np.ndarray],
    img_shape_hw: tuple[int, int],
    min_intersection_px: float = FLUOR_MATCH_MIN_INTERSECTION_PX,
) -> list[Optional[int]]:
    """For each BF contour, pick the fluorescence contour that maximizes overlap."""
    matches: list[Optional[int]] = []
    fluor_boxes = [cv2.boundingRect(c) for c in fluor_contours]

    for bf in bf_contours:
        bx, by, bw, bh = cv2.boundingRect(bf)

        best_idx: Optional[int] = None
        best_inter = 0.0

        for j, (fx, fy, fw, fh) in enumerate(fluor_boxes):
            if (bx + bw < fx) or (fx + fw < bx) or (by + bh < fy) or (fy + fh < by):
                continue

            inter = contour_intersection_area_px(bf, fluor_contours[j], img_shape_hw)
            if inter > best_inter:
                best_inter = inter
                best_idx = j

        if best_idx is not None and best_inter >= float(min_intersection_px):
            matches.append(best_idx)
        else:
            matches.append(None)

    return matches


def measure_fluorescence_intensity_with_global_area(
    fluor_img: np.ndarray,
    bf_contours: list[np.ndarray],
    fluor_contours: list[np.ndarray],
    bf_to_fluor_match: list[Optional[int]],
    um_per_px_x: float,
    um_per_px_y: float,
) -> list[dict]:
    """Intensity stats: within BF contour. Fluor area: from matched fluorescence contour."""
    um2_per_px2 = float(um_per_px_x) * float(um_per_px_y)
    measurements: list[dict] = []

    for i, bf in enumerate(bf_contours, 1):
        bf_mask = np.zeros(fluor_img.shape[:2], dtype=np.uint8)
        cv2.drawContours(bf_mask, [bf], -1, 255, thickness=-1)
        fluor_values = fluor_img[bf_mask > 0]

        j = bf_to_fluor_match[i - 1]
        if j is not None:
            s2_area_px = float(cv2.contourArea(fluor_contours[j]))
        else:
            s2_area_px = 0.0
        s2_area_um2 = s2_area_px * um2_per_px2

        if fluor_values.size > 0:
            measurements.append(
                {
                    "object_id": i,
                    "fluor_area_px": s2_area_px,
                    "fluor_area_um2": s2_area_um2,
                    "fluor_mean": float(np.mean(fluor_values)),
                    "fluor_median": float(np.median(fluor_values)),
                    "fluor_std": float(np.std(fluor_values)),
                    "fluor_min": float(np.min(fluor_values)),
                    "fluor_max": float(np.max(fluor_values)),
                    "fluor_integrated_density": float(np.sum(fluor_values)),
                }
            )
        else:
            measurements.append(
                {
                    "object_id": i,
                    "fluor_area_px": s2_area_px,
                    "fluor_area_um2": s2_area_um2,
                    "fluor_mean": 0.0,
                    "fluor_median": 0.0,
                    "fluor_std": 0.0,
                    "fluor_min": 0.0,
                    "fluor_max": 0.0,
                    "fluor_integrated_density": 0.0,
                }
            )

    return measurements


def visualize_fluorescence_measurements(
    fluor_img8: np.ndarray, contours: list[np.ndarray], measurements: list[dict]
) -> np.ndarray:
    """Create visualization with contours and intensity labels"""
    vis = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(vis, contours, -1, (0, 255, 0), 1)

    for m in measurements:
        c = contours[m["object_id"] - 1]
        M = cv2.moments(c)
        if M["m00"] != 0:
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])

            label = f"{m['object_id']}: {m['fluor_mean']:.0f}"
            cv2.putText(
                vis,
                label,
                (cx, cy),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.4,
                (255, 255, 0),
                1,
                cv2.LINE_AA,
            )

    return vis


# ==================================================
# Segmentation
# ==================================================
def segment_particles_brightfield(img8: np.ndarray, pixel_size_um: float, out_dir: Path) -> np.ndarray:
    """Brightfield segmentation: dark objects on light background"""
    bg = cv2.GaussianBlur(img8, (0, 0), sigmaX=GAUSSIAN_SIGMA, sigmaY=GAUSSIAN_SIGMA)
    enhanced = cv2.subtract(bg, img8)
    enhanced_blur = cv2.GaussianBlur(enhanced, (3, 3), 0)

    save_debug(out_dir, "02_enhanced.png", enhanced, pixel_size_um)
    save_debug(out_dir, "03_enhanced_blur.png", enhanced_blur, pixel_size_um)

    _, thresh = cv2.threshold(enhanced_blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    save_debug(out_dir, "04_thresh_raw.png", thresh, pixel_size_um)

    kernel = np.ones((MORPH_KERNEL_SIZE, MORPH_KERNEL_SIZE), np.uint8)
    bw = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=MORPH_ITERATIONS)
    bw = cv2.dilate(bw, kernel, iterations=DILATE_ITERATIONS)
    bw = cv2.erode(bw, kernel, iterations=ERODE_ITERATIONS)
    save_debug(out_dir, "05_closed.png", bw, pixel_size_um)

    print(f"Mask white fraction (final): {float((bw > 0).mean()):.4f}")
    return bw


# ==================================================
# Plotting and Analysis
# ==================================================
def generate_error_bar_comparison_with_threshold(
    output_dir: Path,
    percentile: float = 0.2,
    restrict_to_groups: Optional[list[str]] = None,
    output_path: Optional[Path] = None,
    title_suffix: str = "",
    dataset_id: str = "",
    threshold_pct: float = 0.05,
    microgel_type: str = "negative",
) -> Optional[Path]:
    """Enhanced version with threshold lines and control mean."""
    
    excel_files = list(output_dir.rglob("*_master.xlsx"))

    if not excel_files:
        print(f"[INFO] No master Excel files found under {output_dir}")
        return None

    all_data_rows: list[dict[str, object]] = []
    group_stats: dict[str, dict[str, float | int]] = {}
    control_mean = None

    # Load data
    for excel_path in sorted(excel_files):
        group_name_raw = excel_path.stem.replace("_master", "")
        display_name = _display_group_name(group_name_raw)

        if restrict_to_groups is not None and display_name not in restrict_to_groups:
            continue

        try:
            typical_sheet = f"{group_name_raw}_Typical_Particles"
            df = pd.read_excel(excel_path, sheet_name=typical_sheet)

            if "Fluor_Density_per_BF_Area" not in df.columns:
                print(f"[WARN] Missing Fluor_Density_per_BF_Area column in {excel_path.name}")
                continue

            values = df["Fluor_Density_per_BF_Area"].dropna()
            
            if len(values) == 0:
                print(f"[WARN] No valid data in {group_name_raw}")
                continue

            for v in values:
                all_data_rows.append(
                    {"Group": display_name, "Fluorescence Density": float(np.asarray(v).item())}
                )

            mean_val = float(np.asarray(values.mean()).item())
            std_val = float(np.asarray(values.std()).item())
            sem_val = float(np.asarray(values.sem()).item())

            group_stats[display_name] = {
                "n": int(len(values)),
                "mean": mean_val,
                "std": std_val,
                "sem": sem_val,
            }

            if display_name == "Control":
                control_mean = mean_val

            print(f"Loaded {len(values)} points for group: {display_name}")

        except Exception as e:
            print(f"[WARN] Could not read {group_name_raw}: {e}")
            continue

    if not all_data_rows:
        print("[WARN] No valid data found for comparison")
        return None

    df_all = pd.DataFrame(all_data_rows)

    group_order: list[str] = sorted(
        df_all["Group"].dropna().astype(str).drop_duplicates().tolist(), 
        key=_group_order_key
    )

    if len(group_order) < 1:
        print("[INFO] Need at least 1 group with data to generate plot.")
        return None

    # Color palette
    palette_colors: list[Any] = ["silver", "violet"]
    if len(group_order) > 2:
        palette_colors = list(sns.color_palette("husl", len(group_order)))
    elif len(group_order) == 1:
        palette_colors = ["skyblue"]

    # Generate plot
    plt.figure(figsize=(10, 7))
    sns.set_style("ticks")

    try:
        ax = sns.barplot(
            data=df_all,
            x="Group",
            y="Fluorescence Density",
            hue="Group",
            order=group_order,
            hue_order=group_order,
            palette=palette_colors,
            legend=False,
            errorbar=None,
            edgecolor="black",
            alpha=0.7,
        )
    except TypeError:
        ax = sns.barplot(
            data=df_all,
            x="Group",
            y="Fluorescence Density",
            hue="Group",
            order=group_order,
            hue_order=group_order,
            palette=palette_colors,
            legend=False,
            ci=None,
            edgecolor="black",
            alpha=0.7,
        )

    # Add SD error bars
    means = df_all.groupby("Group")["Fluorescence Density"].mean()
    sds = df_all.groupby("Group")["Fluorescence Density"].std(ddof=1)

    for xi, g in enumerate(group_order):
        m = float(np.asarray(means.get(g, 0.0)).item())
        sd = float(np.asarray(sds.get(g, 0.0)).item())
        cap = 14 if g == "Control" else 7

        ax.errorbar(
            xi, m, yerr=sd,
            fmt="none", ecolor="black", elinewidth=1.5,
            capsize=cap, capthick=1.5, zorder=10,
        )

    # Add jitter overlay
    sns.stripplot(
        x="Group", y="Fluorescence Density",
        data=df_all, order=group_order,
        jitter=True, color="cyan",
        edgecolor="black", linewidth=0.5,
        size=6, alpha=0.6,
    )

    # ✅ Add threshold lines - BOTH types use LOWER threshold
    legend_handles = []
    
    if control_mean is not None:
        # Control mean line (dotted blue)
        control_line = ax.axhline(
            y=control_mean, 
            color='blue', 
            linestyle=':', 
            linewidth=2.5, 
            label=f'Control Mean ({control_mean:.1f})', 
            zorder=5
        )
        legend_handles.append(control_line)
        
        # Calculate LOWER threshold for BOTH microgel types
        threshold = control_mean * (1 - threshold_pct)
        threshold_label = f'Lower Threshold (-{threshold_pct*100:.0f}%: {threshold:.1f})'
        
        threshold_line = ax.axhline(
            y=threshold, 
            color='red', 
            linestyle='--', 
            linewidth=2.5,
            label=threshold_label, 
            zorder=5
        )
        legend_handles.append(threshold_line)
        
        # Add legend
        ax.legend(
            handles=legend_handles,
            loc='upper right', 
            fontsize=10, 
            framealpha=0.95,
            edgecolor='black',
            fancybox=True
        )

    # Axis labels
    plt.ylabel("Fluorescence Density (a.u./µm²)", fontsize=12, fontweight="bold")
    plt.xlabel("")
    plt.xticks(fontsize=10, fontweight="bold")
    plt.yticks(fontsize=10, fontweight="bold")
    
    # Title - Consistent messaging
    filter_pct_display = int(percentile * 100)
    threshold_pct_display = int(threshold_pct * 100)
    
    title_parts = []
    if dataset_id:
        title_parts.append(dataset_id)
    
    # Microgel type description
    microgel_desc = "G- Microgel" if microgel_type.lower() == "negative" else "G+ Microgel"
    
    title_parts.append(
        f"{microgel_desc} — Typical Particles: Middle {100 - 2*filter_pct_display}% "
        f"(Excluded top/bottom {filter_pct_display}%)"
    )
    
    if title_suffix:
        title_parts.append(title_suffix)
    
    raw_title = " — ".join(title_parts)
    wrapped_title = "\n".join(
        textwrap.wrap(raw_title, width=80, break_long_words=False, break_on_hyphens=False)
    )
    plt.title(wrapped_title, fontsize=11, pad=10)

    # Enhance spines
    for axis in ["top", "bottom", "left", "right"]:
        plt.gca().spines[axis].set_linewidth(1.5)

    plt.tight_layout()

    # Save
    if output_path is not None:
        out_path = output_path
    else:
        if restrict_to_groups is not None:
            safe = "_".join([g.replace(" ", "_") for g in group_order])
            out_path = output_dir / f"comparison_{microgel_type}_{safe}.png"
        else:
            out_path = output_dir / f"comparison_{microgel_type}_all_groups.png"

    try:
        plt.savefig(out_path, dpi=300)
        plt.close()
        print(f"✓ Saved plot with threshold: {out_path.name}")
        return out_path
    except Exception as e:
        print(f"[ERROR] Failed to save plot to {out_path}: {e}")
        plt.close()
        return None

def generate_pairwise_group_vs_control_plots(
    output_root: Path, 
    percentile: float, 
    dataset_id: str,
    threshold_pct: float,
    microgel_type: str
) -> None:
    """Generate pairwise plots with threshold lines"""
    
    control_folder = None
    for folder in output_root.iterdir():
        if folder.is_dir() and folder.name.lower().startswith("control"):
            control_folder = folder
            break
    
    if control_folder is None:
        print("[WARN] No Control group folder found - skipping pairwise plots")
        return
    
    control_master = control_folder / f"{control_folder.name}_master.xlsx"
    if not control_master.exists():
        print(f"[WARN] Control group master file not found: {control_master}")
        return
    
    control_display_name = _display_group_name(control_folder.name)
    
    for group_dir in sorted(output_root.iterdir()):
        if not group_dir.is_dir():
            continue
        if group_dir.name.lower().startswith("control"):
            continue
        if not re.fullmatch(r"\d+", group_dir.name):
            continue

        group_master = group_dir / f"{group_dir.name}_master.xlsx"
        if not group_master.exists():
            print(f"[WARN] Missing group master: {group_master}")
            continue

        pair_plot_path = group_dir / f"Group_{group_dir.name}_vs_Control_threshold.png"
        
        result = generate_error_bar_comparison_with_threshold(
            output_dir=output_root,
            percentile=percentile,
            restrict_to_groups=[group_dir.name, control_display_name],
            output_path=pair_plot_path,
            title_suffix=f"Group {group_dir.name} vs Control",
            dataset_id=dataset_id,
            threshold_pct=threshold_pct,
            microgel_type=microgel_type,
        )
        
        if result is not None:
            print(f"  ✓ Pairwise plot: {pair_plot_path.name}")


def embed_comparison_plots_into_all_excels(
    output_root: Path,
    percentile: float = 0.2,
    plot_path: Optional[Path] = None,
) -> None:
    """Post-process Excel files and embed plots"""
    
    if plot_path is None or not plot_path.exists():
        print(f"[WARN] Plot not found, embedding skipped")
        return
    
    excel_files = sorted(output_root.rglob("*_master.xlsx"))
    if not excel_files:
        print(f"[WARN] No master Excel files found under {output_root}")
        return

    def add_png(ws, path: Path, anchor_cell: str, width_px: int = 620) -> None:
        if not path.exists():
            return
        img = XLImage(str(path))
        if getattr(img, "width", None) and getattr(img, "height", None):
            scale = width_px / float(img.width)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
        ws.add_image(img, anchor_cell)

    updated = 0
    for excel_path in excel_files:
        try:
            wb = load_workbook(excel_path)
            modified = False

            if "Summary" in wb.sheetnames:
                ws_summary = wb["Summary"]
            elif "Error_Bar_Summary" in wb.sheetnames:
                ws_summary = wb["Error_Bar_Summary"]
                ws_summary.title = "Summary"
                ws_summary = wb["Summary"]
                modified = True
            else:
                ws_summary = wb.create_sheet("Summary")
                modified = True

            if "Ratios" in wb.sheetnames:
                ws_ratios = wb["Ratios"]
                ratios_idx = wb.sheetnames.index("Ratios")
                desired_idx = 1
                if ratios_idx != desired_idx:
                    wb.move_sheet(ws_ratios, offset=desired_idx - ratios_idx)
                    modified = True

            current_idx = wb.sheetnames.index(ws_summary.title)
            if current_idx != 0:
                wb.move_sheet(ws_summary, offset=-current_idx)
                modified = True

            marker_cell = "G1"
            marker_value = "COMPARISON_PLOTS_EMBEDDED"
            if ws_summary[marker_cell].value != marker_value:
                add_png(ws_summary, plot_path, "G3")
                ws_summary[marker_cell].value = marker_value
                modified = True

            if modified:
                wb.save(excel_path)
                updated += 1

        except Exception as e:
            print(f"[WARN] Could not embed plots into {excel_path}: {e}")

    if updated > 0:
        print(f"  ✓ Embedded plots in {updated} Excel files")


def consolidate_to_excel(output_dir: Path, group_name: str, percentile: float) -> None:
    """Consolidate all CSVs in a group folder into one Excel workbook"""
    csv_files = list(output_dir.glob("*/object_stats.csv"))

    if not csv_files:
        print(f"[WARN] No CSV files found in {output_dir}")
        return

    excel_path = output_dir / f"{group_name}_master.xlsx"

    if excel_path.exists():
        try:
            excel_path.unlink()
        except PermissionError:
            print(f"[ERROR] Cannot overwrite {excel_path} - file is open")
            return

    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")

        def adjust_column_widths(ws):
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.1
                if adjusted_width > 50:
                    adjusted_width = 50
                ws.column_dimensions[column_letter].width = adjusted_width

        def format_numbers(ws):
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0000'

        all_valid_objects: list[pd.DataFrame] = []
        all_excluded_objects: list[dict] = []

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            # README
            readme_df = pd.DataFrame(
                {
                    "Column Name": [
                        "Object_ID", "BF_Area_px", "BF_Area_um2", "Perimeter_px", "Perimeter_um",
                        "EquivDiameter_px", "EquivDiameter_um", "Circularity", "AspectRatio",
                        "CentroidX_px", "CentroidY_px", "CentroidX_um", "CentroidY_um",
                        "BBoxX_px", "BBoxY_px", "BBoxW_px", "BBoxH_px", "BBoxW_um", "BBoxH_um",
                        "Fluor_Area_px", "Fluor_Area_um2", "Fluor_Mean", "Fluor_Median",
                        "Fluor_Std", "Fluor_Min", "Fluor_Max", "Fluor_IntegratedDensity",
                        "Fluor_Density_per_BF_Area", "BF_to_Fluor_Area_Ratio",
                    ],
                    "Description": [
                        "Unique particle identifier", "Brightfield particle area (pixels²)", "Brightfield particle area (µm²)",
                        "Particle perimeter (pixels)", "Particle perimeter (µm)", "Diameter of equivalent circle (pixels)",
                        "Diameter of equivalent circle (µm)", "Shape roundness (0-1)", "Bounding box width/height ratio",
                        "Particle center X (px)", "Particle center Y (px)", "Particle center X (µm)", "Particle center Y (µm)",
                        "BBox top-left X (px)", "BBox top-left Y (px)", "BBox width (px)", "BBox height (px)",
                        "BBox width (µm)", "BBox height (µm)",
                        "Fluorescent region area (pixels²)", "Fluorescent region area (µm²)",
                        "Avg fluorescence intensity", "Median fluorescence intensity", "Std Dev fluorescence",
                        "Min fluorescence", "Max fluorescence", "Total fluorescence signal",
                        "Fluor density / BF Area (Primary Metric)", "Ratio of BF area to Fluor area",
                    ],
                }
            )

            readme_df.to_excel(writer, sheet_name="README", index=False)
            ws_readme = writer.sheets["README"]
            for cell in ws_readme[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            adjust_column_widths(ws_readme)

            # Per-image sheets
            for csv_file in sorted(csv_files):
                image_name = csv_file.parent.name
                df = pd.read_csv(csv_file)

                cols_to_numeric = ["Fluor_Area_px", "Fluor_IntegratedDensity", "BF_Area_um2", "Fluor_Area_um2"]
                for col in cols_to_numeric:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

                if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                    df["Fluor_Density_per_BF_Area"] = df["Fluor_IntegratedDensity"] / df["BF_Area_um2"]
                else:
                    df["Fluor_Density_per_BF_Area"] = 0.0

                if "Fluor_Area_um2" in df.columns and "BF_Area_um2" in df.columns:
                    df["BF_to_Fluor_Area_Ratio"] = df["BF_Area_um2"] / df["Fluor_Area_um2"]
                else:
                    df["BF_to_Fluor_Area_Ratio"] = 0.0

                df = df.replace([np.inf, -np.inf], 0).fillna(0)

                for idx, row in df.iterrows():
                    reason = None
                    
                    if row["Fluor_Area_px"] == 0 and row["Fluor_IntegratedDensity"] > 0:
                        reason = "Zero fluorescence area with positive integrated density"
                    elif row["Fluor_Area_px"] == 0:
                        reason = "Zero fluorescence area"
                    elif row["Fluor_IntegratedDensity"] == 0:
                        reason = "Zero integrated density"
                    
                    if reason:
                        all_excluded_objects.append({
                            "Object_ID": row["Object_ID"],
                            "Source_Image": image_name,
                            "BF_Area_um2": row["BF_Area_um2"],
                            "Fluor_Area_px": row["Fluor_Area_px"],
                            "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"],
                            "Exclusion_Reason": reason
                        })

                df_valid = df[(df["Fluor_IntegratedDensity"] > 0) & (df["Fluor_Area_px"] > 0)].copy()
                df_valid["Source_Image"] = image_name
                
                if not df_valid.empty:
                    all_valid_objects.append(df_valid)

                sheet_name = image_name[:31]
                
                if "Fluor_Density_per_BF_Area" in df.columns:
                    df = df.sort_values("Fluor_Density_per_BF_Area", ascending=False)
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                format_numbers(ws)
                adjust_column_widths(ws)
                ws.auto_filter.ref = ws.dimensions

            # Excluded Objects
            if all_excluded_objects:
                excluded_df = pd.DataFrame(all_excluded_objects)
                excluded_df.to_excel(writer, sheet_name="Excluded_Objects", index=False)
                ws_excluded = writer.sheets["Excluded_Objects"]
                
                for cell in ws_excluded[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                for row in ws_excluded.iter_rows(min_row=2, max_row=ws_excluded.max_row):
                    for cell in row:
                        cell.fill = red_fill
                
                adjust_column_widths(ws_excluded)
                ws_excluded.auto_filter.ref = ws_excluded.dimensions

            # All Valid Objects
            if all_valid_objects:
                merged_all = pd.concat(all_valid_objects, ignore_index=True)
                merged_all = merged_all.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)
                
                all_valid_sheet_name = f"{group_name}_All_Valid_Objects"
                merged_all.to_excel(writer, sheet_name=all_valid_sheet_name, index=False)
                ws_all = writer.sheets[all_valid_sheet_name]
                
                for cell in ws_all[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                format_numbers(ws_all)
                adjust_column_widths(ws_all)
                ws_all.auto_filter.ref = ws_all.dimensions

            # Typical Particles (GROUP-LEVEL filtering)
            if all_valid_objects:
                merged_all = pd.concat(all_valid_objects, ignore_index=True)
                merged_all = merged_all.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)
                
                n_total = len(merged_all)
                n_cut = int(n_total * percentile)
                
                start_idx = n_cut
                end_idx = n_total - n_cut
                
                if start_idx < end_idx and n_total > 3:
                    typical_particles = merged_all.iloc[start_idx:end_idx].copy()
                    
                    excluded_top = merged_all.iloc[:start_idx]
                    excluded_bottom = merged_all.iloc[end_idx:]
                    
                    for idx, row in excluded_top.iterrows():
                        all_excluded_objects.append({
                            "Object_ID": row["Object_ID"],
                            "Source_Image": row["Source_Image"],
                            "BF_Area_um2": row["BF_Area_um2"],
                            "Fluor_Area_px": row["Fluor_Area_px"],
                            "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"],
                            "Exclusion_Reason": f"Outside typical particle range (top {int(percentile*100)}%)"
                        })
                    
                    for idx, row in excluded_bottom.iterrows():
                        all_excluded_objects.append({
                            "Object_ID": row["Object_ID"],
                            "Source_Image": row["Source_Image"],
                            "BF_Area_um2": row["BF_Area_um2"],
                            "Fluor_Area_px": row["Fluor_Area_px"],
                            "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"],
                            "Exclusion_Reason": f"Outside typical particle range (bottom {int(percentile*100)}%)"
                        })
                else:
                    typical_particles = merged_all.copy()

                typical_sheet_name = f"{group_name}_Typical_Particles"
                typical_particles.to_excel(writer, sheet_name=typical_sheet_name, index=False)
                ws_typ = writer.sheets[typical_sheet_name]

                for cell in ws_typ[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                for row in ws_typ.iter_rows(min_row=2, max_row=ws_typ.max_row):
                    for cell in row:
                        cell.fill = yellow_fill
                
                format_numbers(ws_typ)
                adjust_column_widths(ws_typ)
                ws_typ.auto_filter.ref = ws_typ.dimensions

            # Update Excluded Objects
            if all_excluded_objects:
                wb = writer.book
                if "Excluded_Objects" in wb.sheetnames:
                    wb.remove(wb["Excluded_Objects"])
                
                excluded_df = pd.DataFrame(all_excluded_objects)
                excluded_df.to_excel(writer, sheet_name="Excluded_Objects", index=False)
                ws_excluded = writer.sheets["Excluded_Objects"]
                
                for cell in ws_excluded[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                for row in ws_excluded.iter_rows(min_row=2, max_row=ws_excluded.max_row):
                    for cell in row:
                        cell.fill = red_fill
                
                adjust_column_widths(ws_excluded)
                ws_excluded.auto_filter.ref = ws_excluded.dimensions

            # Summary sheet
            try:
                summary_data = []
                for csv_file in sorted(csv_files):
                    image_name = csv_file.parent.name
                    df = pd.read_csv(csv_file)

                    if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                        df["Fluor_Density_per_BF_Area"] = pd.to_numeric(df["Fluor_IntegratedDensity"], errors='coerce') / pd.to_numeric(df["BF_Area_um2"], errors='coerce')
                    else:
                        df["Fluor_Density_per_BF_Area"] = 0

                    if "Fluor_Area_px" in df.columns:
                        df["Fluor_Area_px"] = pd.to_numeric(df["Fluor_Area_px"], errors='coerce').fillna(0)
                        
                    df = df.replace([np.inf, -np.inf], 0).fillna(0)
                    
                    df_stats = df[(df["Fluor_Area_px"] > 0) & (df["Fluor_IntegratedDensity"] > 0)]
                    
                    avg_fluor_density = df_stats["Fluor_Density_per_BF_Area"].mean() if not df_stats.empty else 0.0
                    
                    avg_ratio = 0.0
                    if not df_stats.empty and "BF_Area_um2" in df_stats.columns and "Fluor_Area_um2" in df_stats.columns:
                         ratios = df_stats["BF_Area_um2"] / df_stats["Fluor_Area_um2"]
                         avg_ratio = ratios.replace([np.inf, -np.inf], 0).mean()

                    summary_data.append({
                        "Image": image_name,
                        "Total_Particles_Detected": len(df),
                        "Particles_With_Fluor": len(df_stats),
                        "Avg_BF_Area_um2": df["BF_Area_um2"].mean() if "BF_Area_um2" in df.columns else 0,
                        "Avg_Fluor_Density": avg_fluor_density,
                        "Avg_BF_to_Fluor_Ratio": avg_ratio,
                    })

                summary_df = pd.DataFrame(summary_data)

                if "Avg_Fluor_Density" in summary_df.columns:
                    summary_df = summary_df.sort_values("Avg_Fluor_Density", ascending=False)

                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                ws_summary = writer.sheets["Summary"]

                for cell in ws_summary[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                format_numbers(ws_summary)
                adjust_column_widths(ws_summary)

            except Exception as e:
                print(f"[WARN] Could not create Summary sheet: {e}")

            # Ratios sheet
            try:
                wb = writer.book
                ratios_name = "Ratios"
                if ratios_name in wb.sheetnames:
                    wb.remove(wb[ratios_name])
                ws_qa = wb.create_sheet(ratios_name)

                ws_qa["A1"] = f"QA Ratios - Group: {group_name}"
                ws_qa["A1"].font = Font(bold=True, size=14)

                def add_png(ws, path: Path, anchor_cell: str, width_px: int = 360) -> None:
                    if not path.exists(): return
                    img = XLImage(str(path))
                    if getattr(img, "width", None) and getattr(img, "height", None):
                        scale = width_px / float(img.width)
                        img.width = int(img.width * scale)
                        img.height = int(img.height * scale)
                    ws.add_image(img, anchor_cell)

                row = 3
                block_h = 34

                for csv_file in sorted(csv_files):
                    image_name = csv_file.parent.name
                    df = pd.read_csv(csv_file)
                    
                    if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                        df["Fluor_Density_per_BF_Area"] = pd.to_numeric(df["Fluor_IntegratedDensity"], errors='coerce') / pd.to_numeric(df["BF_Area_um2"], errors='coerce')
                    else:
                        df["Fluor_Density_per_BF_Area"] = 0.0
                    
                    if "BF_Area_um2" in df.columns and "Fluor_Area_um2" in df.columns:
                        df["BF_to_Fluor_Area_Ratio"] = pd.to_numeric(df["BF_Area_um2"], errors='coerce') / pd.to_numeric(df["Fluor_Area_um2"], errors='coerce')
                    else:
                        df["BF_to_Fluor_Area_Ratio"] = 0.0

                    df = df.replace([np.inf, -np.inf], 0).fillna(0)

                    if "Fluor_Density_per_BF_Area" in df.columns:
                        df = df.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)

                    ws_qa[f"A{row}"] = image_name
                    ws_qa[f"A{row}"].font = Font(bold=True, size=12)
                    
                    headers = ["Object_ID", "Fluor_Density_per_BF_Area", "Rank", "Object_ID", "BF_to_Fluor_Area_Ratio"]
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws_qa.cell(row=row+1, column=col_idx, value=header)
                        cell.font = Font(bold=True)
                        cell.alignment = center_align

                    start_data_row = row + 2
                    n = len(df)
                    
                    for k, r in enumerate(df.itertuples(index=False), 0):
                        ws_qa.cell(row=start_data_row+k, column=1, value=getattr(r, "Object_ID"))
                        ws_qa.cell(row=start_data_row+k, column=2, value=float(getattr(r, "Fluor_Density_per_BF_Area", 0.0))).number_format = '0.0000'
                        ws_qa.cell(row=start_data_row+k, column=3, value=k+1)
                        ws_qa.cell(row=start_data_row+k, column=4, value=getattr(r, "Object_ID"))
                        ws_qa.cell(row=start_data_row+k, column=5, value=float(getattr(r, "BF_to_Fluor_Area_Ratio", 0.0))).number_format = '0.0000'

                    ws_qa.column_dimensions["A"].width = 20
                    ws_qa.column_dimensions["B"].width = 25
                    ws_qa.column_dimensions["C"].width = 10
                    ws_qa.column_dimensions["D"].width = 20
                    ws_qa.column_dimensions["E"].width = 25

                    if n > 0:
                        ch1 = ScatterChart()
                        ch1.title = "Fluor Density"
                        ch1.y_axis.title = "a.u./µm²"
                        ch1.x_axis.title = "Object Rank (Sorted)"
                        
                        xref = Reference(ws_qa, min_col=3, min_row=start_data_row, max_row=start_data_row + n - 1)
                        yref = Reference(ws_qa, min_col=2, min_row=start_data_row, max_row=start_data_row + n - 1)
                        
                        s1 = SeriesFactory(yref, xref, title="Density")
                        s1.marker = Marker(symbol="triangle", size=5)
                        ch1.series.append(s1)
                        ws_qa.add_chart(ch1, f"G{row+1}")

                        ch2 = ScatterChart()
                        ch2.title = "Area Ratio"
                        ch2.y_axis.title = "Ratio"
                        ch2.x_axis.title = "Object Rank (Sorted)"
                        
                        yref2 = Reference(ws_qa, min_col=5, min_row=start_data_row, max_row=start_data_row + n - 1)
                        
                        s2 = SeriesFactory(yref2, xref, title="Ratio")
                        s2.marker = Marker(symbol="circle", size=5)
                        ch2.series.append(s2)
                        ws_qa.add_chart(ch2, f"G{row+18}")

                    img_dir = csv_file.parent
                    add_png(ws_qa, img_dir / "13_mask_accepted_ids.png", f"Q{row+1}", width_px=330)
                    add_png(ws_qa, img_dir / "22_fluorescence_mask_global_ids.png", f"Q{row+18}", width_px=330)
                    add_png(ws_qa, img_dir / "23_fluorescence_contours_global_ids.png", f"Q{row+18}", width_px=330)
                    add_png(ws_qa, img_dir / "24_bf_fluor_matching_overlay_ids.png", f"Q{row+1}", width_px=330)

                    row += block_h

            except Exception as e:
                print(f"[WARN] Could not create Ratios sheet: {e}")

        # Reorder worksheets
        wb2 = load_workbook(excel_path)
        desired_order = [
            "Summary", 
            "Ratios", 
            "README", 
            f"{group_name}_Typical_Particles",
            f"{group_name}_All_Valid_Objects",
            "Excluded_Objects"
        ]
        for idx, sheet_name in enumerate(desired_order):
            if sheet_name in wb2.sheetnames:
                sheet = wb2[sheet_name]
                current_idx = wb2.sheetnames.index(sheet_name)
                if current_idx != idx:
                    wb2.move_sheet(sheet, offset=idx - current_idx)
        wb2.save(excel_path)

    except PermissionError:
        print(f"[ERROR] Cannot write to {excel_path} - file may be open")
    except Exception as e:
        print(f"[ERROR] Failed to create Excel file: {e}")
        import traceback
        traceback.print_exc()


def export_group_statistics_to_csv(output_root: Path) -> None:
    """Export statistics with enhanced console summary"""
    
    stats_list = []
    
    for excel_path in sorted(output_root.glob("*/*_master.xlsx")):
        group_name = excel_path.parent.name
        
        try:
            typical_sheet = f"{group_name}_Typical_Particles"
            df = pd.read_excel(excel_path, sheet_name=typical_sheet)
                        
            if "Fluor_Density_per_BF_Area" in df.columns:
                values = pd.to_numeric(df["Fluor_Density_per_BF_Area"], errors='coerce').dropna()
                
                if values.empty:
                    continue

                mean_val = float(values.mean())
                std_val = float(values.std(ddof=1))
                sem_val = values.sem()
                median_val = float(values.median())
                min_val = float(values.min())
                max_val = float(values.max())
                q30 = float(values.quantile(0.30))
                q70 = float(values.quantile(0.70))
                n = int(len(values))
                
                stats_list.append({
                    'Group': "Control" if group_name.lower().startswith("control") else group_name,
                    'N': n,
                    'Mean': mean_val,
                    'Std_Dev': std_val,
                    'SEM': sem_val,
                    'Median': median_val,
                    'Q30': q30,
                    'Q70': q70,
                    'Min': min_val,
                    'Max': max_val,
                    'CV_percent': (std_val / mean_val * 100) if mean_val > 0 else 0,
                })
                
        except Exception:
            pass
    
    if not stats_list:
        return
    
    stats_df = pd.DataFrame(stats_list)
    
    stats_df['sort_key'] = stats_df['Group'].apply(
        lambda x: (0, int(x)) if x.isdigit() else (1, 999)
    )
    stats_df = stats_df.sort_values('sort_key').drop('sort_key', axis=1)
    
    numeric_cols = stats_df.select_dtypes(include=[np.number]).columns
    stats_df[numeric_cols] = stats_df[numeric_cols].round(2)
    
    output_path = output_root / "group_statistics_summary.csv"
    stats_df.to_csv(output_path, index=False)


def classify_groups_clinical(
    output_root: Path, 
    microgel_type: str = "negative",
    threshold_pct: float = 0.05
) -> pd.DataFrame:
    """Classify all groups based on clinical thresholds."""
    
    control_mean: Optional[float] = None
    control_std: Optional[float] = None
    control_folder = None
    
    for folder in output_root.iterdir():
        if folder.is_dir() and folder.name.lower().startswith("control"):
            control_folder = folder
            break
    
    if control_folder is None:
        return pd.DataFrame()
    
    control_master = control_folder / f"{control_folder.name}_master.xlsx"
    if not control_master.exists():
        return pd.DataFrame()
    
    try:
        typical_sheet = f"{control_folder.name}_Typical_Particles"
        df_control = pd.read_excel(control_master, sheet_name=typical_sheet)
        
        if "Fluor_Density_per_BF_Area" not in df_control.columns:
            return pd.DataFrame()
        
        control_values = pd.to_numeric(df_control["Fluor_Density_per_BF_Area"], errors='coerce').dropna()
        control_mean = float(control_values.mean())
        control_std = float(control_values.std(ddof=1))
        
    except Exception:
        return pd.DataFrame()
    
    if control_mean is None:
        return pd.DataFrame()
    
    # Calculate threshold based on microgel type
    if microgel_type.lower() == "negative":
        # Negative microgel: threshold is ABOVE control mean
        threshold = control_mean * (1 + threshold_pct)
    else:
        # Positive microgel: threshold is BELOW control mean
        threshold = control_mean * (1 - threshold_pct)
    
    results = []
    
    for excel_path in sorted(output_root.glob("*/*_master.xlsx")):
        group_name = excel_path.parent.name
        
        if group_name.lower().startswith("control"):
            continue
        
        try:
            typical_sheet = f"{group_name}_Typical_Particles"
            df = pd.read_excel(excel_path, sheet_name=typical_sheet)
            
            if "Fluor_Density_per_BF_Area" not in df.columns:
                continue
            
            values = pd.to_numeric(df["Fluor_Density_per_BF_Area"], errors='coerce').dropna()
            
            if values.empty:
                continue
            
            mean_val = float(values.mean())
            std_val = float(values.std(ddof=1))
            n = len(values)
            
            # Classification logic based on microgel type
            if microgel_type.lower() == "negative":
                # Negative microgel logic:
                # Mean < Threshold (control mean + threshold_pct%) → NEGATIVE
                # Mean ≥ Threshold → POSITIVE/No obvious bacteria
                if mean_val < threshold:
                    classification = "NEGATIVE"
                    bacteria_status = "Gram-negative bacteria detected"
                else:
                    classification = "POSITIVE/No obvious bacteria"
                    bacteria_status = "No obvious bacteria"
                    
            else:
                # Positive microgel logic:
                # Mean < Threshold (control mean - threshold_pct%) → POSITIVE
                # Mean ≥ Threshold → NEGATIVE/No obvious bacteria
                if mean_val < threshold:
                    classification = "POSITIVE"
                    bacteria_status = "Gram-positive bacteria detected"
                else:
                    classification = "NEGATIVE/No obvious bacteria"
                    bacteria_status = "No obvious bacteria"
            
            diff_from_threshold = mean_val - threshold
            diff_from_control = mean_val - control_mean
            pct_diff_from_control = (diff_from_control / control_mean) * 100
            
            results.append({
                'Group': group_name,
                'N': n,
                'Mean': round(mean_val, 2),
                'Std_Dev': round(std_val, 2),
                'Control_Mean': round(control_mean, 2),
                'Threshold': round(threshold, 2),
                'Diff_from_Threshold': round(diff_from_threshold, 2),
                'Diff_from_Control': round(diff_from_control, 2),
                'Pct_Diff_from_Control': round(pct_diff_from_control, 1),
                'Classification': classification,
            })
            
        except Exception:
            pass
    
    if not results:
        return pd.DataFrame()
    
    results_df = pd.DataFrame(results)
    
    results_df['sort_key'] = results_df['Group'].apply(
        lambda x: int(x) if x.isdigit() else 999
    )
    results_df = results_df.sort_values('sort_key').drop('sort_key', axis=1)
    
    return results_df


def export_clinical_classification(
    output_root: Path,
    classification_df: pd.DataFrame,
    microgel_type: str = "negative"
) -> Optional[Path]:
    """Export clinical classification results to CSV with color coding"""
    
    if classification_df.empty:
        return None
    
    csv_path = output_root / f"clinical_classification_{microgel_type}.csv"
    classification_df.to_csv(csv_path, index=False)
    
    excel_path = output_root / f"clinical_classification_{microgel_type}.xlsx"
    
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            classification_df.to_excel(writer, sheet_name='Classification', index=False)
            
            ws = writer.sheets['Classification']
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            safe_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            for row_idx in range(len(classification_df)):
                excel_row = row_idx + 2
                row_data = classification_df.iloc[row_idx]
                
                if "NEGATIVE" in row_data['Classification'] or "POSITIVE" in row_data['Classification']:
                    if "No obvious bacteria" in row_data['Classification']:
                        fill = safe_fill
                    else:
                        fill = warning_fill
                else:
                    fill = safe_fill
                
                for col_idx in range(1, len(classification_df.columns) + 1):
                    ws.cell(row=excel_row, column=col_idx).fill = fill
                    ws.cell(row=excel_row, column=col_idx).alignment = Alignment(horizontal="center")
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.1, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
    except Exception:
        pass
    
    return csv_path


# ==================================================
# Source Directory Selection
# ==================================================
def select_source_directory(max_depth=2) -> Optional[Path]:
    """Lists directories up to 2 levels deep that have a Control subfolder"""
    root_dir = Path('source')
    
    if not root_dir.exists():
        print(f"[ERROR] Source directory not found: {root_dir.resolve()}")
        return None
    
    directories = []
    
    for root, dirs, files in os.walk(root_dir):
        rel_path = os.path.relpath(root, root_dir)
        if rel_path == '.':
            depth = 0
        else:
            depth = rel_path.count(os.sep) + 1
        
        if depth > max_depth + 1:
            dirs[:] = []
            continue
        
        if depth > 0:
            directories.append(rel_path.replace(os.sep, '\\'))
    
    valid_directories = []
    for dir_path in directories:
        if len(dir_path.split('\\')) > max_depth:
            continue
            
        full_path = root_dir / dir_path.replace('\\', os.sep)
        if full_path.is_dir():
            try:
                subdirs = [d for d in os.listdir(full_path) 
                          if (full_path / d).is_dir()]
                has_control = any(d.lower().startswith('control') for d in subdirs)
                if has_control:
                    valid_directories.append(dir_path)
            except OSError:
                continue
    
    directories = sorted(set(valid_directories))
    
    if not directories:
        print("[ERROR] No valid directories found with Control subfolders.")
        return None
    
    print("\n" + "="*80)
    print("SELECT SOURCE DIRECTORY")
    print("="*80)
    print("\nAvailable directories (up to 2 levels deep):")
    for i, dir_path in enumerate(directories, 1):
        print(f"  [{i}] {dir_path}")
    
    while True:
        selected = logged_input("\nEnter the number or full path (or 'q' to quit): ").strip()
        
        if selected.lower() in {'q', 'quit', 'exit'}:
            raise SystemExit(0)
        
        if selected.isdigit():
            num = int(selected)
            if 1 <= num <= len(directories):
                selected_path = directories[num - 1]
                full_selected = root_dir / selected_path.replace('\\', os.sep)
                return full_selected
            else:
                print(f"Invalid number. Please enter between 1 and {len(directories)}.")
        elif selected in directories:
            full_selected = root_dir / selected.replace('\\', os.sep)
            return full_selected
        else:
            print("Invalid selection. Please enter a valid number or path.")


# ==================================================
# Configuration Collection
# ==================================================
def collect_configuration() -> dict:
    """Collect all user configuration upfront"""
    
    print("\n" + "="*80)
    print("PHASE 1: CONFIGURATION")
    print("="*80)
    print("\nPlease answer the following 4 questions:\n")
    
    config = {}
    
    # Step 1/4: Source Directory
    print("━" * 80)
    print("STEP 1/4: Select Source Directory")
    print("━" * 80)
    config['source_dir'] = select_source_directory()
    if config['source_dir'] is None:
        raise SystemExit("No source directory selected.")
    
    # Step 2/4: Dataset ID
    print("\n" + "━" * 80)
    print("STEP 2/4: Dataset Identifier")
    print("━" * 80)
    print("\nEnter dataset identifier (e.g., 'PD G-', 'Spike G+'):")
    print("  → Press Enter to use timestamp as label")
    
    while True:
        dataset_id = logged_input("Dataset label: ").strip()
        
        if dataset_id == "":
            timestamp_label = datetime.now().strftime("%Y%m%d_%H%M%S")
            config['dataset_id'] = timestamp_label
            print(f"  ✓ Using timestamp: {timestamp_label}")
            break
        
        if len(dataset_id) > 50:
            print("  ✗ Too long (max 50 characters)")
            continue
        
        invalid_chars = set('<>:"|?*\\/')
        found_invalid = [c for c in dataset_id if c in invalid_chars]
        if found_invalid:
            print(f"  ✗ Invalid characters: {', '.join(repr(c) for c in set(found_invalid))}")
            continue
        
        confirm = logged_input(f"  → Confirm '{dataset_id}'? (y/n, Enter=yes): ").strip().lower()
        
        if confirm in ["", "y", "yes"]:
            config['dataset_id'] = dataset_id
            print(f"  ✓ Confirmed: {dataset_id}")
            break
    
    # Step 3/4: Percentile
    print("\n" + "━" * 80)
    print("STEP 3/4: Percentile for Top/Bottom Filtering")
    print("━" * 80)
    print("\nEnter threshold percentage:")
    print("  → Default: 30% (recommended)")
    print("  → Range: 1-100%")

    while True:
        choice = logged_input("Threshold (% or Enter for 30%): ").strip()
        
        if choice == "":
            config['percentile'] = 0.3
            print("  ✓ Using default: 30%")
            break
        else:
            try:
                value = float(choice)
                if 1 <= value <= 30:
                    config['percentile'] = value / 40
                    print(f"  ✓ Selected: {value}%")
                    break
                else:
                    print("Invalid input. Enter a number between 1 and 40, or press Enter.")
            except ValueError:
                print("Invalid input. Enter a number between 1 and 40, or press Enter.")
    



    # Step 4/4: Microgel Type and Threshold
    print("\n" + "━" * 80)
    print("STEP 4/4: Clinical Classification Settings")
    print("━" * 80)
    
    # Microgel type
    print("\nSelect microgel type:")
    print("  [1] Positive microgel (G+)")
    print("  [2] Negative microgel (G-)")

    while True:
        mg_choice = logged_input("Enter number: ").strip()
        if mg_choice == "1":
            config['microgel_type'] = "positive"
            print("  ✓ Selected: Positive microgel (G+)")
            break
        elif mg_choice == "2":
            config['microgel_type'] = "negative"
            print("  ✓ Selected: Negative microgel (G-)")
            break
        else:
            print("Invalid choice. Enter 1 or 2.")


    # Threshold
    print("\nEnter threshold percentage:")
    print("  → Default: 5% (recommended)")
    print("  → Range: 1-20%")
    
    while True:
        choice = logged_input("Threshold (% or Enter for 5%): ").strip()
        
        if choice == "":
            config['threshold_pct'] = 0.05
            print("  ✓ Using default: 5%")
            break
        
        try:
            value = float(choice)
            if value > 1:
                value = value / 100
            
            if 0.01 <= value <= 0.20:
                config['threshold_pct'] = value
                print(f"  ✓ Threshold set: {value*100:.1f}%")
                break
            else:
                print("  ✗ Must be between 1% and 20%")
        except ValueError:
            print("  ✗ Please enter a valid number")
    
    return config


def display_configuration_summary(config: dict) -> None:
    """Display configuration summary for user confirmation"""
    print("\n" + "="*80)
    print("CONFIGURATION SUMMARY")
    print("="*80)
    print(f"\n1. Source Directory: {config['source_dir']}")
    print(f"2. Dataset ID: {config['dataset_id']}")
    print(f"3. Percentile: {int(config['percentile']*100)}%")
    print(f"4. Microgel Type: {config['microgel_type'].capitalize()}")
    print(f"5. Threshold: {int(config['threshold_pct']*100)}%")
    print("\n" + "="*80)
    
    confirm = logged_input("\nProceed with this configuration? (y/n, Enter=yes): ").strip().lower()
    
    if confirm not in ["", "y", "yes"]:
        print("\n✗ Configuration cancelled by user")
        raise SystemExit(0)
    
    print("\n✓ Configuration confirmed - starting processing...")


# ==================================================
# Image Processing
# ==================================================
def process_image(img_path: Path, output_root: Path) -> None:
    """Process a single image"""
    
    xml_props, xml_main = find_metadata_paths(img_path)
    
    try:
        um_per_px_x, um_per_px_y = get_pixel_size_um(xml_props, xml_main)
        um_per_px_x = float(um_per_px_x)
        um_per_px_y = float(um_per_px_y)
    except Exception as e:
        if FALLBACK_UM_PER_PX is None:
            raise
        um_per_px_x = um_per_px_y = float(FALLBACK_UM_PER_PX)

    um_per_px_avg = (um_per_px_x + um_per_px_y) / 2.0

    img_out = output_root / img_path.stem
    ensure_dir(img_out)

    img = cv2.imread(str(img_path), cv2.IMREAD_UNCHANGED)
    if img is None:
        raise FileNotFoundError(str(img_path))
    if img.ndim == 3:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    img8 = normalize_to_8bit(img)
    save_debug(img_out, "01_gray_8bit.png", img8, um_per_px_avg)

    mask = segment_particles_brightfield(img8, float(um_per_px_avg), img_out)

    _fc = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = cast(list[np.ndarray], _fc[-2])

    vis_all = cv2.cvtColor(img8, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(vis_all, contours, -1, (0, 0, 255), 1)
    save_debug(img_out, "10_contours_all.png", vis_all, um_per_px_avg)

    um2_per_px2 = float(um_per_px_x) * float(um_per_px_y)
    min_area_px = MIN_AREA_UM2 / um2_per_px2
    max_area_px = MAX_AREA_UM2 / um2_per_px2

    H, W = img8.shape[:2]
    img_area_px = float(H * W)
    max_big_area_px = MAX_FRACTION_OF_IMAGE_AREA * img_area_px

    accepted: list[np.ndarray] = []
    rejected: list[np.ndarray] = []

    for c in contours:
        area_px = float(cv2.contourArea(c))
        if area_px <= 0:
            rejected.append(c)
            continue

        if area_px >= max_big_area_px:
            rejected.append(c)
            continue

        perim_px = float(cv2.arcLength(c, True))
        circ = (4 * np.pi * area_px / (perim_px**2)) if perim_px > 0 else 0.0

        ok = (min_area_px <= area_px <= max_area_px) and (circ >= MIN_CIRCULARITY)
        (accepted if ok else rejected).append(c)

    vis_acc = cv2.cvtColor(img8, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(vis_acc, rejected, -1, (0, 165, 255), 1)
    cv2.drawContours(vis_acc, accepted, -1, (0, 255, 255), 1)
    vis_acc = draw_object_ids(vis_acc, accepted)
    save_debug(img_out, "11_contours_rejected_orange_accepted_yellow_ids_green.png", vis_acc, um_per_px_avg)

    mask_all = np.zeros_like(mask)
    cv2.drawContours(mask_all, contours, -1, 255, thickness=-1)
    save_debug(img_out, "12_mask_all.png", mask_all)

    mask_acc = np.zeros_like(mask)
    cv2.drawContours(mask_acc, accepted, -1, 255, thickness=-1)
    save_debug(img_out, "13_mask_accepted.png", mask_acc)

    fluor_path = img_path.parent / img_path.name.replace("_ch00", "_ch01")
    fluor_measurements: Optional[list[dict]] = None

    fluor_bw: Optional[np.ndarray] = None
    vis_fluor: Optional[np.ndarray] = None
    vis_match: Optional[np.ndarray] = None

    if fluor_path.exists():
        fluor_img = cv2.imread(str(fluor_path), cv2.IMREAD_UNCHANGED)
        if fluor_img is not None:
            fluor_img, (sy, sx) = align_fluorescence_channel(img, fluor_img)
            
            save_debug(img_out, "20_fluorescence_aligned_raw.png", normalize_to_8bit(fluor_img), um_per_px_avg)

            if fluor_img.ndim == 3:
                fluor_img = cv2.cvtColor(fluor_img, cv2.COLOR_BGR2GRAY)

            fluor_img8 = normalize_to_8bit(fluor_img)
            save_debug(img_out, "20_fluorescence_8bit.png", fluor_img8, um_per_px_avg)

            fluor_bw = segment_fluorescence_global(fluor_img8)
            save_debug(img_out, "22_fluorescence_mask_global.png", fluor_bw, um_per_px_avg)

            _fc2 = cv2.findContours(fluor_bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            fluor_contours_all = cast(list[np.ndarray], _fc2[-2])

            min_fluor_area_px = FLUOR_MIN_AREA_UM2 / um2_per_px2 if um2_per_px2 > 0 else 0.0
            fluor_contours = [c for c in fluor_contours_all if float(cv2.contourArea(c)) >= float(min_fluor_area_px)]

            vis_fluor = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
            cv2.drawContours(vis_fluor, fluor_contours, -1, (0, 255, 0), 1)
            save_debug(img_out, "23_fluorescence_contours_global.png", vis_fluor, um_per_px_avg)

            matches = match_fluor_to_bf_by_overlap(accepted, fluor_contours, fluor_img8.shape[:2])

            fluor_measurements = measure_fluorescence_intensity_with_global_area(
                fluor_img, accepted, fluor_contours, matches, float(um_per_px_x), float(um_per_px_y)
            )

            vis_match = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
            for idx, bf_c in enumerate(accepted):
                j = matches[idx]
                cv2.drawContours(vis_match, [bf_c], -1, (0, 0, 255), 1)
                if j is not None:
                    cv2.drawContours(vis_match, [fluor_contours[j]], -1, (0, 255, 0), 2)
            save_debug(img_out, "24_bf_fluor_matching_overlay.png", vis_match, um_per_px_avg)

            fluor_overlay = visualize_fluorescence_measurements(fluor_img8, accepted, fluor_measurements)
            save_debug(img_out, "21_fluorescence_overlay.png", fluor_overlay, um_per_px_avg)

    parts = img_path.stem.split()
    group_id = parts[0] if parts else "unk"
    sequence_num = parts[-1].split("_")[0] if parts else "0"

    object_ids = [f"{group_id}_{sequence_num}_{i}" for i in range(1, len(accepted) + 1)]

    mask_acc_bgr = cv2.cvtColor(mask_acc, cv2.COLOR_GRAY2BGR)
    save_debug_ids(img_out, "13_mask_accepted.png", mask_acc_bgr, accepted, object_ids, um_per_px_avg)

    if fluor_bw is not None:
        fluor_bw_bgr = cv2.cvtColor(fluor_bw, cv2.COLOR_GRAY2BGR)
        save_debug_ids(img_out, "22_fluorescence_mask_global.png", fluor_bw_bgr, accepted, object_ids, um_per_px_avg)

    if vis_fluor is not None:
        save_debug_ids(img_out, "23_fluorescence_contours_global.png", vis_fluor, accepted, object_ids, um_per_px_avg)

    if vis_match is not None:
        save_debug_ids(img_out, "24_bf_fluor_matching_overlay.png", vis_match, accepted, object_ids, um_per_px_avg)

    csv_path = img_out / "object_stats.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "Object_ID",
                "BF_Area_px",
                "BF_Area_um2",
                "Perimeter_px",
                "Perimeter_um",
                "EquivDiameter_px",
                "EquivDiameter_um",
                "Circularity",
                "AspectRatio",
                "CentroidX_px",
                "CentroidY_px",
                "CentroidX_um",
                "CentroidY_um",
                "BBoxX_px",
                "BBoxY_px",
                "BBoxW_px",
                "BBoxH_px",
                "BBoxW_um",
                "BBoxH_um",
                "Fluor_Area_px",
                "Fluor_Area_um2",
                "Fluor_Mean",
                "Fluor_Median",
                "Fluor_Std",
                "Fluor_Min",
                "Fluor_Max",
                "Fluor_IntegratedDensity",
            ]
        )

        for i, c in enumerate(accepted, 1):
            compound_id = f"{group_id}_{sequence_num}_{i}"

            area_px = float(cv2.contourArea(c))
            area_um2 = area_px * (float(um_per_px_x) * float(um_per_px_y))

            perim_px = float(cv2.arcLength(c, True))
            perim_um = contour_perimeter_um(c, float(um_per_px_x), float(um_per_px_y))

            eqd_px = equivalent_diameter_from_area(area_px)
            eqd_um = equivalent_diameter_from_area(area_um2)

            circ = (4 * np.pi * area_px / (perim_px**2)) if perim_px > 0 else 0.0

            x, y, bw, bh = cv2.boundingRect(c)
            aspect = (float(bw) / float(bh)) if bh > 0 else 0.0

            M = cv2.moments(c)
            if M["m00"] != 0:
                cx = float(M["m10"] / M["m00"])
                cy = float(M["m01"] / M["m00"])
            else:
                cx, cy = 0.0, 0.0

            cx_um = cx * float(um_per_px_x)
            cy_um = cy * float(um_per_px_y)
            bw_um = float(bw) * float(um_per_px_x)
            bh_um = float(bh) * float(um_per_px_y)

            if fluor_measurements is not None:
                fm = fluor_measurements[i - 1]
            else:
                fm = {
                    "fluor_area_px": 0.0,
                    "fluor_area_um2": 0.0,
                    "fluor_mean": 0.0,
                    "fluor_median": 0.0,
                    "fluor_std": 0.0,
                    "fluor_min": 0.0,
                    "fluor_max": 0.0,
                    "fluor_integrated_density": 0.0,
                }

            w.writerow(
                [
                    compound_id,
                    f"{area_px:.2f}",
                    f"{area_um2:.4f}",
                    f"{perim_px:.2f}",
                    f"{perim_um:.4f}",
                    f"{eqd_px:.2f}",
                    f"{eqd_um:.4f}",
                    f"{circ:.4f}",
                    f"{aspect:.4f}",
                    f"{cx:.2f}",
                    f"{cy:.2f}",
                    f"{cx_um:.4f}",
                    f"{cy_um:.4f}",
                    x,
                    y,
                    bw,
                    bh,
                    f"{bw_um:.4f}",
                    f"{bh_um:.4f}",
                    f"{float(fm['fluor_area_px']):.2f}",
                    f"{float(fm['fluor_area_um2']):.4f}",
                    f"{float(fm['fluor_mean']):.2f}",
                    f"{float(fm['fluor_median']):.2f}",
                    f"{float(fm['fluor_std']):.2f}",
                    f"{float(fm['fluor_min']):.2f}",
                    f"{float(fm['fluor_max']):.2f}",
                    f"{float(fm['fluor_integrated_density']):.2f}",
                ]
            )


# ==================================================
# Main Function
# ==================================================
def main() -> None:
    start_time = time.time()
    
    # PHASE 1: Configuration
    config = collect_configuration()
    display_configuration_summary(config)
    
    # Setup paths
    global SOURCE_DIR, CONTROL_DIR, OUTPUT_DIR
    SOURCE_DIR = config['source_dir']
    
    control_candidates = [
        d for d in SOURCE_DIR.iterdir() 
        if d.is_dir() and d.name.lower().startswith('control')
    ]
    
    if control_candidates:
        CONTROL_DIR = control_candidates[0]
    else:
        CONTROL_DIR = SOURCE_DIR / "Control group"
    
    safe_dataset_id = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in config['dataset_id'])
    safe_dataset_id = safe_dataset_id.strip().replace(' ', '_')
    output_folder_name = f"{safe_dataset_id}_{_timestamp}_{_script_name}"
    
    OUTPUT_DIR = _project_root / "outputs" / output_folder_name
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    if CLEAR_OUTPUT_DIR_EACH_RUN:
        clear_output_dir(OUTPUT_DIR)
    
    # PHASE 2: Processing
    print("\n" + "="*80)
    print("PHASE 2: PROCESSING")
    print("="*80)
    
    # Step 1/7: Collect images
    print("\n━" * 80)
    print("STEP 1/7: Collecting Images")
    print("━" * 80)
    
    groups = list_sample_group_folders(SOURCE_DIR)
    dirs_to_process = groups.copy()
    
    if CONTROL_DIR.exists():
        dirs_to_process.append(CONTROL_DIR)
    
    img_paths: list[Path] = []
    for d in dirs_to_process:
        img_paths.extend(sorted(d.rglob(IMAGE_GLOB)))
    
    print(f"  ✓ Found {len(img_paths)} images across {len(dirs_to_process)} groups")
    
    if not img_paths:
        raise FileNotFoundError(f"No images found matching {IMAGE_GLOB}")
    
    # Step 2/7: Process images
    print("\n━" * 80)
    print("STEP 2/7: Processing Images")
    print("━" * 80)
    
    total_processed = 0
    total_failed = 0

    for p in tqdm(img_paths, desc="  Processing", unit="img"):
        out_root = (OUTPUT_DIR / p.parent.name) if SEPARATE_OUTPUT_BY_GROUP else OUTPUT_DIR
        ensure_dir(out_root)

        try:
            process_image(p, out_root)
            total_processed += 1
        except Exception as e:
            tqdm.write(f"  [ERROR] {p.name}: {e}")
            total_failed += 1
    
    print(f"\n  ✓ Processed: {total_processed} succeeded, {total_failed} failed")
    
    # Step 3/7: Consolidate to Excel
    print("\n━" * 80)
    print("STEP 3/7: Consolidating to Excel")
    print("━" * 80)
    
    for group_dir in OUTPUT_DIR.iterdir():
        if group_dir.is_dir() and len(list(group_dir.glob("*/object_stats.csv"))) > 0:
            print(f"  → Consolidating {group_dir.name}...")
            consolidate_to_excel(group_dir, group_dir.name, config['percentile'])
    
    print("  ✓ Excel consolidation complete")
    
    # Step 4/7: Generate plots
    print("\n━" * 80)
    print("STEP 4/7: Generating Comparison Plots")
    print("━" * 80)
    
    # Pairwise plots
    print("  → Generating pairwise plots...")
    generate_pairwise_group_vs_control_plots(
        OUTPUT_DIR, 
        config['percentile'], 
        config['dataset_id'],
        config['threshold_pct'],
        config['microgel_type']
    )
    
    # All-groups plot
    print("  → Generating all-groups plot...")
    all_groups_plot_path = generate_error_bar_comparison_with_threshold(
        output_dir=OUTPUT_DIR,
        percentile=config['percentile'],
        restrict_to_groups=None,
        output_path=OUTPUT_DIR / f"comparison_{config['microgel_type']}_all_groups.png",
        title_suffix="All Groups",
        dataset_id=config['dataset_id'],
        threshold_pct=config['threshold_pct'],
        microgel_type=config['microgel_type'],
    )
    
    print("  ✓ Plots generated")
    
    # Step 5/7: Embed plots
    print("\n━" * 80)
    print("STEP 5/7: Embedding Plots into Excel")
    print("━" * 80)
    
    for group_dir in sorted(OUTPUT_DIR.iterdir()):
        if not group_dir.is_dir():
            continue

        if group_dir.name.lower().startswith("control"):
            plot = all_groups_plot_path
        elif re.fullmatch(r"\d+", group_dir.name):
            plot = group_dir / f"Group_{group_dir.name}_vs_Control_threshold.png"
        else:
            continue

        if plot and plot.exists():
            embed_comparison_plots_into_all_excels(group_dir, config['percentile'], plot_path=plot)
    
    print("  ✓ Plots embedded")
    
    # Step 6/7: Export statistics
    print("\n━" * 80)
    print("STEP 6/7: Exporting Statistics")
    print("━" * 80)
    
    export_group_statistics_to_csv(OUTPUT_DIR)
    print("  ✓ Statistics exported")
    
    # Step 7/7: Clinical classification
    print("\n━" * 80)
    print("STEP 7/7: Clinical Classification")
    print("━" * 80)
    
    classification_df = classify_groups_clinical(
        OUTPUT_DIR, 
        config['microgel_type'], 
        config['threshold_pct']
    )
    
    if not classification_df.empty:
        export_clinical_classification(OUTPUT_DIR, classification_df, config['microgel_type'])
        print("  ✓ Classification complete")
    else:
        print("  ✗ Classification failed")
    
    # Copy log file
    try:
        import shutil
        log_copy_path = OUTPUT_DIR / _log_path.name
        shutil.copy2(_log_path, log_copy_path)
    except Exception:
        pass
    
    # PHASE 3: Summary
    elapsed_time = time.time() - start_time
    
    print("\n" + "="*80)
    print("PHASE 3: SUMMARY")
    print("="*80)
    print(f"\n✓ Processing complete in {elapsed_time:.1f} seconds")
    print(f"\n📁 Output directory: {OUTPUT_DIR}")
    print(f"\nKey files generated:")
    print(f"  • Master Excel files (per group)")
    print(f"  • Comparison plots with threshold lines")
    print(f"  • Clinical classification results")
    print(f"  • Group statistics summary")
    
    # Auto-open output folder
    try:
        import subprocess
        import platform
        
        system = platform.system()
        if system == "Windows":
            subprocess.run(["explorer", str(OUTPUT_DIR)])
        elif system == "Darwin":  # macOS
            subprocess.run(["open", str(OUTPUT_DIR)])
        else:  # Linux
            subprocess.run(["xdg-open", str(OUTPUT_DIR)])
        
        print(f"\n✓ Output folder opened")
    except Exception:
        print(f"\n→ Please manually open: {OUTPUT_DIR}")
    
    print("\n" + "="*80)
    
    sys.stdout = sys.__stdout__
    sys.stderr = sys.__stderr__


if __name__ == "__main__":
    main()