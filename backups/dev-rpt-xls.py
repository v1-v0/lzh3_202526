import os
import numpy as np
from pathlib import Path
from skimage.io import imread
from skimage import filters, morphology, measure
from skimage.segmentation import watershed
from scipy import ndimage as ndi
from scipy import stats
import pandas as pd
import cv2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from typing import cast
from openpyxl.cell.cell import Cell
import warnings

warnings.filterwarnings('ignore')

def set_cell_value_safe(ws: Worksheet, row: int, col: int, value: str) -> None:
    """
    Safely set a cell value. If (row,col) is inside a merged range, write to the
    top-left cell of that merged range. Cast to Cell to satisfy type checkers.
    """
    coord = f"{get_column_letter(col)}{row}"
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            min_col, min_row, _, _ = rng.bounds
            target = cast(Cell, ws.cell(row=min_row, column=min_col))
            target.value = value
            return
    target = cast(Cell, ws.cell(row=row, column=col))
    target.value = value

class BacteriaReviewTool:
    def __init__(self, root_folder="source"):
        """Initialize the review tool with root folder"""
        self.root_folder = root_folder
        self.config = {
            "gray_threshold": 53,
            "fluor_threshold": 5,
            "min_area": 400,
            "min_fluor_area": 400,
            "overlap_threshold": 0.80,
        }
        self.all_results = []
        self.process_all_subfolders()
        self.export_to_excel()

    def process_all_subfolders(self):
        """Process all subfolders under root_folder"""
        root_path = Path(self.root_folder)
        if not root_path.exists():
            print(f"Error: Root folder '{self.root_folder}' does not exist")
            return

        subfolders = [f for f in root_path.iterdir() if f.is_dir()]
        print(f"Found {len(subfolders)} subfolder(s) to process")

        for subfolder in subfolders:
            print(f"\nProcessing subfolder: {subfolder.name}")
            self.image_pairs = []
            self.load_image_pairs(subfolder)
            if self.image_pairs:
                self.generate_detection_reports(subfolder.name)
            else:
                print(f"  No image pairs found in {subfolder.name}")

    def load_image_pairs(self, folder_path):
        """Load all grayscale and fluorescence image pairs from a folder"""
        self.image_pairs = []
        folder_path = Path(folder_path)

        gray_files = sorted([
            f for f in os.listdir(folder_path)
            if f.endswith('_ch00.tif') and os.path.isfile(folder_path / f)
        ])

        for gray_file in gray_files:
            fluor_file = gray_file.replace('_ch00.tif', '_ch01.tif')
            gray_path = folder_path / gray_file
            fluor_path = folder_path / fluor_file

            if fluor_path.exists():
                try:
                    gray_img = imread(gray_path)
                    fluor_img = imread(fluor_path)

                    if len(gray_img.shape) == 3:
                        gray_img = cv2.cvtColor(gray_img, cv2.COLOR_RGB2GRAY)
                    if len(fluor_img.shape) == 3:
                        fluor_img = cv2.cvtColor(fluor_img, cv2.COLOR_RGB2GRAY)

                    self.image_pairs.append({
                        'gray': gray_img,
                        'fluor': fluor_img,
                        'gray_file': gray_file,
                        'fluor_file': fluor_file
                    })
                except Exception as e:
                    print(f"  Error loading {gray_file}: {e}")

        print(f"  Loaded {len(self.image_pairs)} image pair(s)")

    def preprocess_grayscale(self, image):
        """Preprocess grayscale image"""
        median_filtered = filters.median(image, morphology.disk(3))
        if median_filtered.max() > 0:
            img_normalized = (median_filtered / median_filtered.max() * 255).astype(np.uint8)
        else:
            img_normalized = median_filtered.astype(np.uint8)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(img_normalized)
        return enhanced

    def preprocess_fluorescence(self, image):
        """Preprocess fluorescence image"""
        median_filtered = filters.median(image, morphology.disk(3))
        if median_filtered.max() > 0:
            img_normalized = (median_filtered / median_filtered.max() * 255).astype(np.uint8)
        else:
            img_normalized = median_filtered.astype(np.uint8)
        clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(img_normalized)
        return enhanced

    def segment_grayscale(self, preprocessed, threshold):
        """Segment grayscale bacteria (dark objects)"""
        binary = np.asarray(preprocessed) <= threshold
        distance = ndi.distance_transform_edt(binary)
        distance = np.asarray(distance, dtype=np.float32)
        local_max = morphology.local_maxima(distance)
        markers = measure.label(local_max)
        labels = watershed(-distance, markers, mask=binary)
        return labels > 0, labels

    def segment_fluorescence(self, preprocessed, threshold):
        """Segment fluorescence bacteria (bright objects)"""
        binary = np.asarray(preprocessed) >= threshold
        distance = ndi.distance_transform_edt(binary)
        distance = np.asarray(distance, dtype=np.float32)
        local_max = morphology.local_maxima(distance)
        markers = measure.label(local_max)
        labels = watershed(-distance, markers, mask=binary)
        return labels > 0, labels

    def postprocess(self, binary_mask):
        """Post-process binary mask"""
        min_size = 50
        cleaned = morphology.remove_small_objects(binary_mask, min_size=min_size)
        filled = ndi.binary_fill_holes(cleaned)
        opened = morphology.binary_opening(filled, morphology.disk(2))
        final = morphology.binary_closing(opened, morphology.disk(2))
        return final

    def detect_bacteria(self, image, is_grayscale=True):
        """Detect bacteria in image with current settings"""
        if is_grayscale:
            preprocessed = self.preprocess_grayscale(image)
            threshold = self.config['gray_threshold']
            min_area = self.config['min_area']
            binary, labels = self.segment_grayscale(preprocessed, threshold)
        else:
            preprocessed = self.preprocess_fluorescence(image)
            threshold = self.config['fluor_threshold']
            min_area = self.config['min_fluor_area']
            binary, labels = self.segment_fluorescence(preprocessed, threshold)

        final = self.postprocess(binary)
        labeled = measure.label(final)
        props = measure.regionprops(labeled)
        large_props = [prop for prop in props if prop.area > min_area]
        return labeled, large_props

    def calculate_mode(self, intensities):
        """Calculate mode of intensity values"""
        if len(intensities) == 0:
            return 0
        mode_result = stats.mode(intensities, keepdims=True)
        return float(mode_result.mode[0])

    def get_contour_intensities(self, image, labeled, props):
        """Get intensity statistics for each contour (using MODE)"""
        contour_data = []
        for i, prop in enumerate(props):
            region_mask = (labeled == prop.label)
            intensities = image[region_mask]
            contour_data.append({
                'id': i + 1,
                'area': prop.area,
                'mode_intensity': self.calculate_mode(intensities),
                'max_intensity': np.max(intensities) if len(intensities) > 0 else 0,
                'centroid_y': prop.centroid[0],
                'centroid_x': prop.centroid[1],
                'label': prop.label
            })
        return contour_data

    def calculate_colocalization(self, gray_labeled, fluor_labeled, gray_props, fluor_props):
        """Calculate which grayscale and fluorescence bacteria are co-localized"""
        colocalized = []
        overlap_threshold = self.config['overlap_threshold']

        for gray_prop in gray_props:
            gray_mask = (gray_labeled == gray_prop['label'])
            for fluor_prop in fluor_props:
                fluor_mask = (fluor_labeled == fluor_prop['label'])
                intersection = np.logical_and(gray_mask, fluor_mask)
                intersection_area = np.sum(intersection)
                overlap_ratio = intersection_area / gray_prop['area'] if gray_prop['area'] > 0 else 0
                if overlap_ratio >= overlap_threshold:
                    colocalized.append({
                        'gray_id': gray_prop['id'],
                        'fluor_id': fluor_prop['id'],
                        'overlap_ratio': overlap_ratio,
                        'gray_area': gray_prop['area'],
                        'fluor_area': fluor_prop['area'],
                        'intersection_area': intersection_area
                    })
        return colocalized

    def generate_detection_reports(self, subfolder_name):
        """Generate detection reports for all image pairs in a subfolder"""
        for pair in self.image_pairs:
            gray_img = pair['gray']
            fluor_img = pair['fluor']
            filename = pair['gray_file']

            print(f"  Processing {filename}")

            # Detect bacteria
            gray_labeled, gray_large = self.detect_bacteria(gray_img, is_grayscale=True)
            fluor_labeled, fluor_large = self.detect_bacteria(fluor_img, is_grayscale=False)

            # Get contour intensity data
            gray_contour_data = self.get_contour_intensities(gray_img, gray_labeled, gray_large)
            fluor_contour_data = self.get_contour_intensities(fluor_img, fluor_labeled, fluor_large)

            # Calculate co-localization
            colocalized = self.calculate_colocalization(gray_labeled, fluor_labeled, 
                                                       gray_contour_data, fluor_contour_data)

            # Store results
            coloc_rate = (len(colocalized) / len(gray_large) * 100) if len(gray_large) > 0 else 0
            self.all_results.append({
                'subfolder': subfolder_name,
                'filename': filename,
                'gray_count': len(gray_large),
                'fluor_count': len(fluor_large),
                'coloc_count': len(colocalized),
                'coloc_rate': coloc_rate,
                'gray_contour_data': gray_contour_data,
                'fluor_contour_data': fluor_contour_data,
                'colocalized': colocalized
            })

    def get_merged_cell_top_left(self, ws: Worksheet, row: int, column: int) -> tuple[int, int]:
        """Get the top-left cell coordinates for a given cell if it's part of a merged range"""
        cell = ws.cell(row=row, column=column)
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                min_col, min_row, _, _ = merged_range.bounds
                return min_row, min_col
        return row, column

    def export_to_excel(self):
        """Export detection reports to a single Excel sheet with formatted layout"""
        print("\nExporting detection reports to Excel...")
        output_file = "bacteria_detection_reports.xlsx"
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Detection_Reports")
        else:
            ws.title = "Detection_Reports"

        # Define styles
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1
        for result in self.all_results:
            # Detection Report Header
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, f"═══════════════ DETECTION REPORT ═══════════════")
            ws.cell(row=top_left_row, column=top_left_col).font = header_font
            row += 1
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, f"Source: {result['subfolder']} {result['filename']}")
            row += 1
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, (
                f"Parameters: Gray Thresh={self.config['gray_threshold']}, "
                f"Fluor Thresh={self.config['fluor_threshold']}, "
                f"Overlap={self.config['overlap_threshold']*100:.0f}%"
            ))
            row += 2

            # Summary Section
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, "┌─── SUMMARY ───┐")
            ws.cell(row=top_left_row, column=top_left_col).font = header_font
            row += 1
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, f"  Grayscale bacteria:     {result['gray_count']:>3}")
            row += 1
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, f"  Fluorescence bacteria:  {result['fluor_count']:>3}")
            row += 1
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, f"  Co-localized pairs:     {result['coloc_count']:>3}")
            row += 1
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, f"  Co-localization rate:   {result['coloc_rate']:>5.1f}%")
            row += 2

            # Co-localized Pairs
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, f"┌─── CO-LOCALIZED PAIRS ({result['coloc_count']}) ───┐")
            cast(Cell, ws.cell(row=top_left_row, column=top_left_col)).font = header_font
            row += 1
            headers = ["Gray", "Fluor", "Overlap", "G.Area", "F.Area", "Intersect"]
            for col, header in enumerate(headers, start=1):
                top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, col)
                cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                cell.value = header
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            row += 1
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, "─" * 60)
            row += 1

            if result['colocalized']:
                for coloc in result['colocalized']:
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{coloc['gray_id']:<6}"
                    cell.alignment = center_align
                    cell.border = border
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 2)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{coloc['fluor_id']:<7}"
                    cell.alignment = center_align
                    cell.border = border
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 3)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{coloc['overlap_ratio']*100:>7.1f}%"
                    cell.alignment = center_align
                    cell.border = border
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 4)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{coloc['gray_area']:<9.0f}"
                    cell.alignment = center_align
                    cell.border = border
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 5)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{coloc['fluor_area']:<9.0f}"
                    cell.alignment = center_align
                    cell.border = border
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 6)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{coloc['intersection_area']:.0f}"
                    cell.alignment = center_align
                    cell.border = border
                    row += 1
            else:
                top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
                set_cell_value_safe(ws, top_left_row, top_left_col, "  No co-localized bacteria detected")
                row += 1
            row += 1

            # Grayscale Channel
            colocalized_gray_ids = {c['gray_id'] for c in result['colocalized']}
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, f"┌─── GRAYSCALE CHANNEL ({result['gray_count']} bacteria) ───┐")
            cast(Cell, ws.cell(row=top_left_row, column=top_left_col)).font = header_font
            row += 1
            headers = ["ID", "*", "Area", "Mode", "Max", "Centroid (Y, X)"]
            for col, header in enumerate(headers, start=1):
                top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, col)
                cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                cell.value = header
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            row += 1
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, "─" * 55)
            row += 1

            if result['gray_contour_data']:
                for data in result['gray_contour_data']:
                    coloc_mark = "*" if data['id'] in colocalized_gray_ids else " "
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{data['id']:<5}"
                    cell.alignment = center_align
                    cell.border = border

                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 2)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{coloc_mark:<2}"
                    cell.alignment = center_align
                    cell.border = border

                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 3)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{data['area']:<9.0f}"
                    cell.alignment = center_align
                    cell.border = border

                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 4)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{data['mode_intensity']:<9.1f}"
                    cell.alignment = center_align
                    cell.border = border

                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 5)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{data['max_intensity']:<8.0f}"
                    cell.alignment = center_align
                    cell.border = border

                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 6)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"({data['centroid_y']:.1f}, {data['centroid_x']:.1f})"
                    cell.alignment = center_align
                    cell.border = border
                    row += 1
                gray_areas = [d['area'] for d in result['gray_contour_data']]
                gray_mode_ints = [d['mode_intensity'] for d in result['gray_contour_data']]
                top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
                set_cell_value_safe(ws, top_left_row, top_left_col, "─" * 55)
                row += 1
                top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
                set_cell_value_safe(ws, top_left_row, top_left_col, (
                    f"Stats: Area μ={np.mean(gray_areas):.1f} "
                    f"σ={np.std(gray_areas):.1f} | "
                    f"Intensity(mode) μ={np.mean(gray_mode_ints):.1f}"
                ))
                row += 1
            else:
                top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
                set_cell_value_safe(ws, top_left_row, top_left_col, "  No bacteria detected")
                row += 1
            row += 1

            # Fluorescence Channel
            colocalized_fluor_ids = {c['fluor_id'] for c in result['colocalized']}
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, f"┌─── FLUORESCENCE CHANNEL ({result['fluor_count']} bacteria) ───┐")
            cast(Cell, ws.cell(row=top_left_row, column=top_left_col)).font = header_font
            row += 1
            for col, header in enumerate(headers, start=1):
                top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, col)
                cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                cell.value = header
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            row += 1
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, "─" * 55)
            row += 1

            if result['fluor_contour_data']:
                for data in result['fluor_contour_data']:
                    coloc_mark = "*" if data['id'] in colocalized_fluor_ids else " "
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{data['id']:<5}"
                    cell.alignment = center_align
                    cell.border = border
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 2)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{coloc_mark:<2}"
                    cell.alignment = center_align
                    cell.border = border
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 3)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{data['area']:<9.0f}"
                    cell.alignment = center_align
                    cell.border = border
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 4)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{data['mode_intensity']:<9.1f}"
                    cell.alignment = center_align
                    cell.border = border
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 5)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"{data['max_intensity']:<8.0f}"
                    cell.alignment = center_align
                    cell.border = border
                    top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 6)
                    cell = cast(Cell, ws.cell(row=top_left_row, column=top_left_col))
                    cell.value = f"({data['centroid_y']:.1f}, {data['centroid_x']:.1f})"
                    cell.alignment = center_align
                    cell.border = border
                    row += 1
                fluor_areas = [d['area'] for d in result['fluor_contour_data']]
                fluor_mode_ints = [d['mode_intensity'] for d in result['fluor_contour_data']]
                top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
                set_cell_value_safe(ws, top_left_row, top_left_col, "─" * 55)
                row += 1
                top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
                set_cell_value_safe(ws, top_left_row, top_left_col, (
                    f"Stats: Area μ={np.mean(fluor_areas):.1f} "
                    f"σ={np.std(fluor_areas):.1f} | "
                    f"Intensity(mode) μ={np.mean(fluor_mode_ints):.1f}"
                ))
                row += 1
            else:
                top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
                set_cell_value_safe(ws, top_left_row, top_left_col, "  No bacteria detected")
                row += 1
            row += 1

            # Legend
            top_left_row, top_left_col = self.get_merged_cell_top_left(ws, row, 1)
            set_cell_value_safe(ws, top_left_row, top_left_col, "Legend: * = Co-localized | Yellow contours = Co-localized")
            row += 2

        # Adjust column widths
        for col in range(1, 7):
            column_letter = get_column_letter(col)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_file)
        print(f"✓ Excel file saved: {output_file}")
        print(f"  - Detection reports: {len(self.all_results)} entries")
        print(f"{'='*70}\n")

def main():
    """Main function to run the review tool"""
    print("="*70)
    print("BACTERIA DETECTION REPORT GENERATOR")
    print("="*70)
    print("\nFeatures:")
    print("  • Processes all subfolders under 'source'")
    print("  • Generates formatted Detection Reports sheet in Excel")
    print("  • Includes summary, co-localized pairs, grayscale and fluorescence tables")
    print("  • Uses MODE intensity (no MIN)")
    print("\n" + "="*70 + "\n")

    try:
        tool = BacteriaReviewTool(root_folder="source")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
