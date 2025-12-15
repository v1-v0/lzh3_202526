import cv2
import numpy as np
import sys
import csv

import atexit
import re
from datetime import datetime
from pathlib import Path
import xml.etree.ElementTree as ET
from typing import Optional, Tuple, Any, cast

from tqdm import tqdm
import pandas as pd
from scipy import stats as scipy_stats

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.series_factory import SeriesFactory

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage




# ==================================================
# Logging: tee stdout/stderr to a file
class Tee:
    def __init__(self, *streams):
        self.streams = streams

    def write(self, data: str) -> None:
        for s in self.streams:
            s.write(data)
            s.flush()

    def flush(self) -> None:
        for s in self.streams:
            s.flush()


_project_root = Path(__file__).resolve().parent
_logs_dir = _project_root / "logs"
_logs_dir.mkdir(exist_ok=True)
_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
_script_name = Path(__file__).stem
_log_path = _logs_dir / f"run_{_timestamp}_{_script_name}.txt"
_log_file = open(_log_path, "w", encoding="utf-8")

sys.stdout = Tee(sys.stdout, _log_file)
sys.stderr = Tee(sys.stderr, _log_file)
print(f"Saving output to: {_log_path}")


@atexit.register
def _close_log_file() -> None:
    try:
        _log_file.close()
    except Exception:
        pass


# ==================================================
# Configuration
# ==================================================
SOURCE_DIR = Path("./source")
CONTROL_DIR = SOURCE_DIR / "Control group"

# Segment only brightfield channel
IMAGE_GLOB = "*_ch00.tif"

OUTPUT_DIR = _project_root / "debug"
OUTPUT_DIR.mkdir(exist_ok=True)

# Scale bar parameters
SCALE_BAR_LENGTH_UM = 10
SCALE_BAR_HEIGHT = 4
SCALE_BAR_MARGIN = 15
SCALE_BAR_COLOR = (255, 255, 255)
SCALE_BAR_BG_COLOR = (0, 0, 0)
SCALE_BAR_TEXT_COLOR = (255, 255, 255)
SCALE_BAR_FONT_SCALE = 0.5
SCALE_BAR_FONT_THICKNESS = 1

# --- Segmentation ---
GAUSSIAN_SIGMA = 15
MORPH_KERNEL_SIZE = 3
MORPH_ITERATIONS = 1
DILATE_ITERATIONS = 1
ERODE_ITERATIONS = 1

# Filtering (in micrometers)
MIN_AREA_UM2 = 3.0
MAX_AREA_UM2 = 2000.0
MIN_CIRCULARITY = 0.0
MAX_FRACTION_OF_IMAGE_AREA = 0.25

# --- Fluorescence segmentation (S2) ---
FLUOR_GAUSSIAN_SIGMA = 1.5
FLUOR_MORPH_KERNEL_SIZE = 3
FLUOR_MIN_AREA_UM2 = 3.0
FLUOR_MATCH_MIN_INTERSECTION_PX = 5.0

# Debug options
CLEAR_OUTPUT_DIR_EACH_RUN = True
SEPARATE_OUTPUT_BY_GROUP = True
FALLBACK_UM_PER_PX: Optional[float] = 0.109492


# ==================================================
# Helpers
# ==================================================
def clear_output_dir(folder: Path) -> None:
    for p in folder.glob("*"):
        try:
            if p.is_file():
                p.unlink()
            elif p.is_dir():
                for q in p.rglob("*"):
                    try:
                        if q.is_file():
                            q.unlink()
                    except Exception:
                        pass
        except Exception:
            pass


def add_scale_bar(
    img: np.ndarray, pixel_size: float, unit: str = "um", length_um: float = 10
) -> np.ndarray:
    """Add a scale bar to the image"""
    if pixel_size is None or pixel_size <= 0:
        return img

    bar_length_px = int(round(length_um / pixel_size))
    if bar_length_px < 10:
        return img

    h, w = img.shape[:2]
    bar_x = w - bar_length_px - SCALE_BAR_MARGIN
    bar_y = h - SCALE_BAR_HEIGHT - SCALE_BAR_MARGIN

    # Check for overflow
    if bar_x < 0 or bar_y < 0:
        return img

    label = (
        f"{int(length_um)} um"
        if unit in ["µm", "um"]
        else f"{int(length_um)} {unit}"
    )

    font = cv2.FONT_HERSHEY_SIMPLEX
    (text_w, text_h), baseline = cv2.getTextSize(
        label, font, SCALE_BAR_FONT_SCALE, SCALE_BAR_FONT_THICKNESS
    )

    text_x = bar_x + (bar_length_px - text_w) // 2
    text_y = bar_y - 8

    bg_padding = 5
    bg_x1 = min(bar_x, text_x) - bg_padding
    bg_y1 = text_y - text_h - bg_padding
    bg_x2 = max(bar_x + bar_length_px, text_x + text_w) + bg_padding
    bg_y2 = bar_y + SCALE_BAR_HEIGHT + bg_padding

    img = img.copy()
    overlay = img.copy()
    cv2.rectangle(overlay, (bg_x1, bg_y1), (bg_x2, bg_y2), SCALE_BAR_BG_COLOR, -1)
    cv2.addWeighted(overlay, 0.7, img, 0.3, 0, img)

    cv2.rectangle(
        img,
        (bar_x, bar_y),
        (bar_x + bar_length_px, bar_y + SCALE_BAR_HEIGHT),
        SCALE_BAR_COLOR,
        -1,
    )

    cv2.putText(
        img,
        label,
        (text_x, text_y),
        font,
        SCALE_BAR_FONT_SCALE,
        SCALE_BAR_TEXT_COLOR,
        SCALE_BAR_FONT_THICKNESS,
        cv2.LINE_AA,
    )

    return img


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def save_debug(
    folder: Path,
    name: str,
    img: np.ndarray,
    pixel_size_um: Optional[float] = None,
) -> None:
    """Save debug image with optional scale bar"""
    out = folder / name
    img_to_save = img.copy()

    if pixel_size_um is not None and pixel_size_um > 0:
        img_to_save = add_scale_bar(
            img_to_save, float(pixel_size_um), "um", SCALE_BAR_LENGTH_UM
        )

    cv2.imwrite(str(out), img_to_save)


def list_sample_group_folders(source_dir: Path) -> list[Path]:
    groups: list[Path] = []
    if not source_dir.exists():
        raise FileNotFoundError(f"Source folder not found: {source_dir.resolve()}")

    for p in source_dir.iterdir():
        if not p.is_dir():
            continue
        if p.name == "Control group":
            continue
        if re.fullmatch(r"\d+", p.name):
            groups.append(p)

    groups.sort(key=lambda x: int(x.name))
    return groups


def prompt_user_select_group(groups: list[Path]) -> Optional[Path]:
    """Return a selected group folder, or None if user chooses 'all'."""
    if not groups:
        return None

    print("\nSelect sample group folder to process:")
    print("  [0] ALL numeric groups")
    for i, g in enumerate(groups, 1):
        print(f"  [{i}] {g.name}")

    while True:
        s = input("Enter number (or 'q' to quit): ").strip().lower()
        if s in {"q", "quit", "exit"}:
            raise SystemExit(0)

        if not s.isdigit():
            print("Please enter a valid number.")
            continue

        idx = int(s)
        if idx == 0:
            return None
        if 1 <= idx <= len(groups):
            return groups[idx - 1]

        print("Out of range. Try again.")


def get_percentile_option() -> float:
    """Prompt user to select percentile for top/bottom group analysis."""
    print("\nSelect percentile for top/bottom group analysis:")
    print("  [1] 20% (default)")
    print("  [2] 25%")
    print("  [3] 30%")
    
    while True:
        choice = input("Enter number (or press Enter for default): ").strip()
        if choice == '' or choice == '1':
            return 0.2
        elif choice == '2':
            return 0.25
        elif choice == '3':
            return 0.3
        else:
            print("Invalid choice. Please enter 1, 2, or 3, or press Enter.")


def find_metadata_paths(img_path: Path) -> tuple[Optional[Path], Optional[Path]]:
    base = img_path.stem
    if base.endswith("_ch00"):
        base = base[:-5]
    md_dir = img_path.parent / "MetaData"
    xml_main = md_dir / f"{base}.xml"
    xml_props = md_dir / f"{base}_Properties.xml"

    return (
        xml_props if xml_props.exists() else None,
        xml_main if xml_main.exists() else None,
    )


def _require_attr(elem: ET.Element, attr: str, context: str) -> str:
    v = elem.get(attr)
    if v is None:
        raise ValueError(f"Missing attribute '{attr}' in {context}")
    return v


def _parse_float(s: str) -> float:
    return float(s.strip().replace(",", "."))


def get_pixel_size_um(
    xml_props_path: Optional[Path],
    xml_main_path: Optional[Path],
) -> Tuple[float, float]:
    if xml_props_path is not None:
        try:
            tree = ET.parse(xml_props_path)
            root = tree.getroot()

            dims = root.findall(".//ImageDescription/Dimensions/DimensionDescription")
            by_id = {d.get("DimID"): d for d in dims}

            def read_dim(dim_id: str) -> Tuple[float, int, str]:
                d = by_id.get(dim_id)
                if d is None:
                    raise ValueError(
                        f"Missing DimensionDescription with DimID='{dim_id}' in {xml_props_path.name}"
                    )

                length_s = _require_attr(
                    d, "Length", f"{xml_props_path.name} DimID={dim_id}"
                )
                n_s = _require_attr(
                    d, "NumberOfElements", f"{xml_props_path.name} DimID={dim_id}"
                )
                unit = _require_attr(d, "Unit", f"{xml_props_path.name} DimID={dim_id}")

                length = _parse_float(length_s)
                n = int(n_s)
                return length, n, unit

            x_len, x_n, x_unit = read_dim("X")
            y_len, y_n, y_unit = read_dim("Y")

            if x_unit != "µm" or y_unit != "µm":
                raise ValueError(
                    f"Unexpected units in {xml_props_path.name}: X={x_unit}, Y={y_unit}"
                )

            return float(x_len / x_n), float(y_len / y_n)

        except Exception as e:
            print(f"[WARN] Failed to read pixel size from {xml_props_path}: {e}")

    if xml_main_path is not None:
        try:
            tree = ET.parse(xml_main_path)
            root = tree.getroot()

            dims = root.findall(".//ImageDescription/Dimensions/DimensionDescription")
            by_id = {d.get("DimID"): d for d in dims}

            def read_dim(dim_id: str) -> Tuple[float, int, str]:
                d = by_id.get(dim_id)
                if d is None:
                    raise ValueError(
                        f"Missing DimensionDescription with DimID='{dim_id}' in {xml_main_path.name}"
                    )

                length_s = _require_attr(
                    d, "Length", f"{xml_main_path.name} DimID={dim_id}"
                )
                n_s = _require_attr(
                    d, "NumberOfElements", f"{xml_main_path.name} DimID={dim_id}"
                )
                unit = _require_attr(d, "Unit", f"{xml_main_path.name} DimID={dim_id}")

                length = _parse_float(length_s)
                n = int(n_s)
                return length, n, unit

            x_len_m, x_n, x_unit = read_dim("1")
            y_len_m, y_n, y_unit = read_dim("2")

            if x_unit != "m" or y_unit != "m":
                raise ValueError(
                    f"Unexpected units in {xml_main_path.name}: X={x_unit}, Y={y_unit}"
                )

            return float((x_len_m * 1e6) / x_n), float((y_len_m * 1e6) / y_n)

        except Exception as e:
            print(f"[WARN] Failed to read pixel size from {xml_main_path}: {e}")

    raise ValueError(
        "Could not determine pixel size (µm/px). Missing/invalid metadata XML."
    )


def contour_perimeter_um(contour: np.ndarray, um_per_px_x: float, um_per_px_y: float) -> float:
    pts = contour.reshape(-1, 2).astype(np.float64)
    pts[:, 0] *= float(um_per_px_x)
    pts[:, 1] *= float(um_per_px_y)
    d = np.diff(np.vstack([pts, pts[0]]), axis=0)
    seg = np.sqrt((d[:, 0] ** 2) + (d[:, 1] ** 2))
    return float(seg.sum())


def equivalent_diameter_from_area(area: float) -> float:
    return float(2.0 * np.sqrt(area / np.pi)) if area > 0 else 0.0


def normalize_to_8bit(img: np.ndarray) -> np.ndarray:
    if img.dtype == np.uint8:
        return img
    if img.dtype == np.uint16:
        out = np.zeros_like(img, dtype=np.uint8)
        cv2.normalize(img, out, 0, 255, cv2.NORM_MINMAX)
        return out
    img_f = img.astype(np.float32)
    mn, mx = float(np.min(img_f)), float(np.max(img_f))
    if mx <= mn:
        return np.zeros(img.shape, dtype=np.uint8)
    return ((img_f - mn) * (255.0 / (mx - mn))).clip(0, 255).astype(np.uint8)


def _put_text_outline(
    img: np.ndarray,
    text: str,
    org: tuple[int, int],
    font_scale: float = 0.5,
    color: tuple[int, int, int] = (255, 255, 255),
    thickness: int = 1,
) -> None:
    """Draw readable text with a black outline."""
    cv2.putText(
        img,
        text,
        org,
        cv2.FONT_HERSHEY_SIMPLEX,
        font_scale,
        (0, 0, 0),
        thickness + 2,
        cv2.LINE_AA,
    )
    cv2.putText(
        img,
        text,
        org,
        cv2.FONT_HERSHEY_SIMPLEX,
        font_scale,
        color,
        thickness,
        cv2.LINE_AA,
    )


def draw_object_ids(
    img_bgr: np.ndarray, contours: list[np.ndarray], labels: Optional[list[str]] = None
) -> np.ndarray:
    """
    Draw object labels at contour centroids.
    If labels is None, uses 1..N.
    """
    out = img_bgr.copy()
    for i, c in enumerate(contours, 1):
        M = cv2.moments(c)
        if M["m00"] == 0:
            continue
        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])
        text = labels[i - 1] if (labels is not None and i - 1 < len(labels)) else str(i)
        _put_text_outline(
            out, text, (cx, cy), font_scale=0.5, color=(0, 255, 0), thickness=1
        )
    return out


def _ids_name(original_png: str) -> str:
    """Insert '_ids' before '.png' (rule requested)."""
    if original_png.lower().endswith(".png"):
        return original_png[:-4] + "_ids.png"
    return original_png + "_ids"


def save_debug_ids(
    folder: Path,
    original_name: str,
    img_bgr: np.ndarray,
    accepted_contours: list[np.ndarray],
    object_ids: list[str],
    pixel_size_um: Optional[float] = None,
) -> None:
    """
    Save a labeled (Object_ID) version of an existing debug view.
    Output name follows rule: insert '_ids' before '.png'.
    """
    labeled = draw_object_ids(img_bgr, accepted_contours, labels=object_ids)
    save_debug(folder, _ids_name(original_name), labeled, pixel_size_um)


# ==================================================
# Fluorescence segmentation + matching (many-to-one allowed)
# ==================================================
def segment_fluorescence_global(fluor_img8: np.ndarray) -> np.ndarray:
    """Segment fluorescence objects globally. Returns binary mask uint8 (0/255)."""
    blur = cv2.GaussianBlur(
        fluor_img8,
        (0, 0),
        sigmaX=FLUOR_GAUSSIAN_SIGMA,
        sigmaY=FLUOR_GAUSSIAN_SIGMA,
    )
    _, bw = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    k = np.ones((FLUOR_MORPH_KERNEL_SIZE, FLUOR_MORPH_KERNEL_SIZE), np.uint8)
    bw = cv2.morphologyEx(bw, cv2.MORPH_OPEN, k, iterations=1)
    bw = cv2.morphologyEx(bw, cv2.MORPH_CLOSE, k, iterations=1)
    return bw


def contour_intersection_area_px(
    c1: np.ndarray, c2: np.ndarray, shape_hw: tuple[int, int]
) -> float:
    """Intersection area in pixels between two contours (matching heuristic only)."""
    H, W = shape_hw
    m1 = np.zeros((H, W), dtype=np.uint8)
    m2 = np.zeros((H, W), dtype=np.uint8)
    cv2.drawContours(m1, [c1], -1, 255, thickness=-1)
    cv2.drawContours(m2, [c2], -1, 255, thickness=-1)
    return float(np.count_nonzero(cv2.bitwise_and(m1, m2)))


def match_fluor_to_bf_by_overlap(
    bf_contours: list[np.ndarray],
    fluor_contours: list[np.ndarray],
    img_shape_hw: tuple[int, int],
    min_intersection_px: float = FLUOR_MATCH_MIN_INTERSECTION_PX,
) -> list[Optional[int]]:
    """
    For each BF contour, pick the fluorescence contour that maximizes overlap.
    Many-to-one allowed. Overlap used ONLY for assignment.
    """
    matches: list[Optional[int]] = []
    fluor_boxes = [cv2.boundingRect(c) for c in fluor_contours]

    for bf in bf_contours:
        bx, by, bw, bh = cv2.boundingRect(bf)

        best_idx: Optional[int] = None
        best_inter = 0.0

        for j, (fx, fy, fw, fh) in enumerate(fluor_boxes):
            if (bx + bw < fx) or (fx + fw < bx) or (by + bh < fy) or (fy + fh < by):
                continue

            inter = contour_intersection_area_px(bf, fluor_contours[j], img_shape_hw)
            if inter > best_inter:
                best_inter = inter
                best_idx = j

        if best_idx is not None and best_inter >= float(min_intersection_px):
            matches.append(best_idx)
        else:
            matches.append(None)

    return matches


def measure_fluorescence_intensity_with_global_area(
    fluor_img: np.ndarray,
    bf_contours: list[np.ndarray],
    fluor_contours: list[np.ndarray],
    bf_to_fluor_match: list[Optional[int]],
    um_per_px_x: float,
    um_per_px_y: float,
) -> list[dict]:
    """
    Intensity stats: within BF contour.
    Fluor area: from matched fluorescence contour (global), can extend outside BF.
    """
    um2_per_px2 = float(um_per_px_x) * float(um_per_px_y)
    measurements: list[dict] = []

    for i, bf in enumerate(bf_contours, 1):
        bf_mask = np.zeros(fluor_img.shape[:2], dtype=np.uint8)
        cv2.drawContours(bf_mask, [bf], -1, 255, thickness=-1)
        fluor_values = fluor_img[bf_mask > 0]

        j = bf_to_fluor_match[i - 1]
        if j is not None:
            s2_area_px = float(cv2.contourArea(fluor_contours[j]))
        else:
            s2_area_px = 0.0
        s2_area_um2 = s2_area_px * um2_per_px2

        if fluor_values.size > 0:
            measurements.append(
                {
                    "object_id": i,
                    "fluor_area_px": s2_area_px,
                    "fluor_area_um2": s2_area_um2,
                    "fluor_mean": float(np.mean(fluor_values)),
                    "fluor_median": float(np.median(fluor_values)),
                    "fluor_std": float(np.std(fluor_values)),
                    "fluor_min": float(np.min(fluor_values)),
                    "fluor_max": float(np.max(fluor_values)),
                    "fluor_integrated_density": float(np.sum(fluor_values)),
                }
            )
        else:
            measurements.append(
                {
                    "object_id": i,
                    "fluor_area_px": s2_area_px,
                    "fluor_area_um2": s2_area_um2,
                    "fluor_mean": 0.0,
                    "fluor_median": 0.0,
                    "fluor_std": 0.0,
                    "fluor_min": 0.0,
                    "fluor_max": 0.0,
                    "fluor_integrated_density": 0.0,
                }
            )

    return measurements


def visualize_fluorescence_measurements(
    fluor_img8: np.ndarray, contours: list[np.ndarray], measurements: list[dict]
) -> np.ndarray:
    """Create visualization with contours and intensity labels"""
    vis = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(vis, contours, -1, (0, 255, 0), 1)

    for m in measurements:
        c = contours[m["object_id"] - 1]
        M = cv2.moments(c)
        if M["m00"] != 0:
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])

            label = f"{m['object_id']}: {m['fluor_mean']:.0f}"
            cv2.putText(
                vis,
                label,
                (cx, cy),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.4,
                (255, 255, 0),
                1,
                cv2.LINE_AA,
            )

    return vis


# ==================================================
# Segmentation
# ==================================================
def segment_particles_brightfield(img8: np.ndarray, pixel_size_um: float, out_dir: Path) -> np.ndarray:
    """Brightfield segmentation: dark objects on light background"""
    bg = cv2.GaussianBlur(img8, (0, 0), sigmaX=GAUSSIAN_SIGMA, sigmaY=GAUSSIAN_SIGMA)
    enhanced = cv2.subtract(bg, img8)
    enhanced_blur = cv2.GaussianBlur(enhanced, (3, 3), 0)

    save_debug(out_dir, "02_enhanced.png", enhanced, pixel_size_um)
    save_debug(out_dir, "03_enhanced_blur.png", enhanced_blur, pixel_size_um)

    _, thresh = cv2.threshold(enhanced_blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    save_debug(out_dir, "04_thresh_raw.png", thresh, pixel_size_um)

    kernel = np.ones((MORPH_KERNEL_SIZE, MORPH_KERNEL_SIZE), np.uint8)
    bw = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=MORPH_ITERATIONS)
    bw = cv2.dilate(bw, kernel, iterations=DILATE_ITERATIONS)
    bw = cv2.erode(bw, kernel, iterations=ERODE_ITERATIONS)
    save_debug(out_dir, "05_closed.png", bw, pixel_size_um)

    print(f"Mask white fraction (final): {float((bw > 0).mean()):.4f}")
    return bw


# ==================================================
# Excel consolidation with enhanced statistics
# ==================================================
def generate_error_bar_comparison(output_dir: Path, percentile: float = 0.2) -> None:
    """Generate error bar comparison plot with overlaid jitter (strip plot) - SD only"""

    excel_files = list(output_dir.rglob("*_master.xlsx"))

    if len(excel_files) < 2:
        print(f"[INFO] Need at least 2 groups for comparison. Found {len(excel_files)}")
        return

    all_data_rows = []
    group_stats = {}

    for excel_path in sorted(excel_files):
        group_name = excel_path.stem.replace("_master", "")

        try:
            typical_sheet = f"{group_name}_Typical_Particles"
            df = pd.read_excel(excel_path, sheet_name=typical_sheet)

            if "Fluor_Density_per_BF_Area" in df.columns:
                values = df["Fluor_Density_per_BF_Area"].dropna()

                for v in values:
                    all_data_rows.append(
                        {"Group": group_name, "Fluorescence Density": float(np.asarray(v).item())}
                    )

                mean_val = float(np.asarray(values.mean()).item())
                std_val = float(np.asarray(values.std()).item())
                sem_val = float(np.asarray(values.sem()).item())

                group_stats[group_name] = {
                    "n": len(values),
                    "mean": mean_val,
                    "std": std_val,
                    "sem": sem_val,
                    "ci_95": 1.96 * sem_val,
                }

                print(f"Loaded {len(values)} points for group: {group_name}")

        except Exception as e:
            print(f"[WARN] Could not read {group_name}: {e}")
            continue

    if not all_data_rows:
        print("[WARN] No valid data found for comparison")
        return

    df_all = pd.DataFrame(all_data_rows)

    unique_groups = df_all["Group"].unique()
    palette_colors = ["silver", "violet"]
    if len(unique_groups) > 2:
        palette_colors = sns.color_palette("husl", len(unique_groups))

    # Generate SD plot
    plt.figure(figsize=(6, 5))
    sns.set_style("ticks")

    try:
        sns.barplot(
            data=df_all,
            x="Group",
            y="Fluorescence Density",
            hue="Group",
            palette=palette_colors,
            legend=False,
            errorbar="sd",
            capsize=0.1,
            edgecolor="black",
            alpha=0.7,
            err_kws={"color": "black", "linewidth": 1.5},
        )
    except TypeError:
        # Fallback for older seaborn versions
        sns.barplot(
            x="Group",
            y="Fluorescence Density",
            data=df_all,
            ci="sd",
            capsize=0.1,
            palette=palette_colors,
            edgecolor="black",
            errcolor="black",
            errwidth=1.5,
            alpha=0.7,
        )

    sns.stripplot(
        x="Group",
        y="Fluorescence Density",
        data=df_all,
        jitter=True,
        color="cyan",
        edgecolor="black",
        linewidth=0.5,
        size=6,
        alpha=0.6,
    )

    plt.ylabel("Fluorescence Density (a.u./µm²)", fontsize=12, fontweight="bold")
    plt.xlabel("")
    plt.xticks(fontsize=10, fontweight="bold")
    plt.yticks(fontsize=10, fontweight="bold")
    
    # Updated title with percentile
    percentile_pct = int(percentile * 100)
    plt.title(f"Top {percentile_pct}% vs Bottom {percentile_pct}% Comparison\n(Error Bars: Standard Deviation)", fontsize=11)

    for axis in ["top", "bottom", "left", "right"]:
        plt.gca().spines[axis].set_linewidth(1.5)

    plt.tight_layout()

    out_path = output_dir / f"error_bar_jitter_comparison_SD_{percentile_pct}pct.png"
    plt.savefig(out_path, dpi=300)
    plt.close()
    print(f"Saved plot: {out_path}")

    # Save statistics CSV
    pd.DataFrame(
        [
            {
                "Group": g_name,
                "n": stats["n"],
                "Mean": f"{stats['mean']:.2f}",
                "SD": f"{stats['std']:.2f}",
                "SEM": f"{stats['sem']:.2f}",
                "95% CI": f"±{stats['ci_95']:.2f}",
            }
            for g_name, stats in group_stats.items()
        ]
    ).to_csv(output_dir / "comparison_statistics.csv", index=False)

    # Perform t-test if exactly 2 groups
    if len(unique_groups) == 2:
        g1, g2 = unique_groups[0], unique_groups[1]
        data1 = df_all[df_all["Group"] == g1]["Fluorescence Density"]
        data2 = df_all[df_all["Group"] == g2]["Fluorescence Density"]

        t_stat_any, p_val_any = cast(Any, scipy_stats.ttest_ind(data1, data2))

        t_stat = float(np.asarray(t_stat_any).item())
        p_val_float = float(np.asarray(p_val_any).item())

        with open(output_dir / "statistical_test.txt", "w", encoding="utf-8") as f:
            f.write("Two-Sample t-Test Results\n")
            f.write(f"{'=' * 60}\n")
            f.write(f"Group 1: {g1} (n={len(data1)})\n")
            f.write(f"Group 2: {g2} (n={len(data2)})\n\n")
            f.write(f"t-statistic: {t_stat:.4f}\n")
            f.write(f"p-value: {p_val_float:.4f}\n")
            f.write(f"Significance (alpha=0.05): {'YES' if p_val_float < 0.05 else 'NO'}\n")


def embed_comparison_plots_into_all_excels(
    output_root: Path,
    percentile: float = 0.2,
) -> None:
    """
    Post-process all master Excel files under output_root:
      - rename 'Error_Bar_Summary' to 'Summary' (or keep 'Summary' if already present)
      - move 'Summary' to the first worksheet
      - embed SD comparison plot into 'Summary'
    """
    percentile_pct = int(percentile * 100)
    sd_plot = f"error_bar_jitter_comparison_SD_{percentile_pct}pct.png"
    sd_img_path = output_root / sd_plot

    if not sd_img_path.exists():
        print(f"[WARN] SD plot not found, embedding skipped: {sd_img_path}")

    excel_files = sorted(output_root.rglob("*_master.xlsx"))
    if not excel_files:
        print(f"[WARN] No master Excel files found under {output_root}")
        return

    def add_png(ws, path: Path, anchor_cell: str, width_px: int = 620) -> None:
        if not path.exists():
            return
        img = XLImage(str(path))
        if getattr(img, "width", None) and getattr(img, "height", None):
            scale = width_px / float(img.width)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
        ws.add_image(img, anchor_cell)
    
    updated = 0
    for excel_path in excel_files:
        try:
            wb = load_workbook(excel_path)

            modified = False

            if "Summary" in wb.sheetnames:
                ws_summary = wb["Summary"]
            elif "Error_Bar_Summary" in wb.sheetnames:
                ws_summary = wb["Error_Bar_Summary"]
                ws_summary.title = "Summary"
                ws_summary = wb["Summary"]
                modified = True
            else:
                ws_summary = wb.create_sheet("Summary")
                modified = True

            # After moving Summary to index 0:
            if "Ratios" in wb.sheetnames:
                ws_ratios = wb["Ratios"]
                ratios_idx = wb.sheetnames.index("Ratios")
                desired_idx = 1
                if ratios_idx != desired_idx:
                    wb.move_sheet(ws_ratios, offset=desired_idx - ratios_idx)
                    modified = True

            current_idx = wb.sheetnames.index(ws_summary.title)
            if current_idx != 0:
                wb.move_sheet(ws_summary, offset=-current_idx)
                modified = True

            marker_cell = "G1"
            marker_value = "COMPARISON_PLOTS_EMBEDDED"
            if ws_summary[marker_cell].value != marker_value:
                add_png(ws_summary, sd_img_path, "G3")
                
                ws_summary[marker_cell].value = marker_value
                modified = True
            else:
                print(f"Plots already embedded in {excel_path.name}; skipping embedding.")

            if modified:
                wb.save(excel_path)
                updated += 1
                print(f"Updated workbook: {excel_path}")
            else:
                print(f"No changes needed: {excel_path}")

        except Exception as e:
            print(f"[WARN] Could not embed plots into {excel_path}: {e}")

    print(f"Embedding complete. Updated {updated}/{len(excel_files)} workbooks.")


def consolidate_to_excel(output_dir: Path, group_name: str, percentile: float) -> None:
    """Consolidate all CSVs in a group folder into one Excel workbook with statistics and color coding"""
    csv_files = list(output_dir.glob("*/object_stats.csv"))

    if not csv_files:
        print(f"[WARN] No CSV files found in {output_dir}")
        return

    excel_path = output_dir / f"{group_name}_master.xlsx"

    if excel_path.exists():
        try:
            excel_path.unlink()
        except PermissionError:
            print(f"[ERROR] Cannot overwrite {excel_path} - file is open in another program")
            print("        Please close the file and run again, or delete it manually")
            return

    try:
        from openpyxl.styles import PatternFill, Font, Alignment

        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")

        all_typical_particles: list[pd.DataFrame] = []

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            # --- README sheet ---
            readme_df = pd.DataFrame(
                {
                    "Column Name": [
                        "Object_ID",
                        "BF_Area_px",
                        "BF_Area_um2",
                        "Perimeter_px",
                        "Perimeter_um",
                        "EquivDiameter_px",
                        "EquivDiameter_um",
                        "Circularity",
                        "AspectRatio",
                        "CentroidX_px",
                        "CentroidY_px",
                        "CentroidX_um",
                        "CentroidY_um",
                        "BBoxX_px",
                        "BBoxY_px",
                        "BBoxW_px",
                        "BBoxH_px",
                        "BBoxW_um",
                        "BBoxH_um",
                        "Fluor_Area_px",
                        "Fluor_Area_um2",
                        "Fluor_Mean",
                        "Fluor_Median",
                        "Fluor_Std",
                        "Fluor_Min",
                        "Fluor_Max",
                        "Fluor_IntegratedDensity",
                        "Fluor_Density_per_BF_Area",
                        "BF_to_Fluor_Area_Ratio",
                    ],
                    "Description": [
                        "Unique particle identifier (format: GroupID_ImageSequence_ParticleID, e.g., 12_2_5)",
                        "Brightfield particle area (pixels²)",
                        "Brightfield particle area (µm²)",
                        "Particle perimeter (pixels)",
                        "Particle perimeter (µm)",
                        "Diameter of equivalent circle (pixels)",
                        "Diameter of equivalent circle (µm)",
                        "Shape roundness: 4π×Area/Perimeter² (0-1, 1=perfect circle)",
                        "Bounding box width/height ratio",
                        "Particle center X-coordinate (pixels)",
                        "Particle center Y-coordinate (pixels)",
                        "Particle center X-coordinate (µm)",
                        "Particle center Y-coordinate (µm)",
                        "Bounding box top-left X (pixels)",
                        "Bounding box top-left Y (pixels)",
                        "Bounding box width (pixels)",
                        "Bounding box height (pixels)",
                        "Bounding box width (µm)",
                        "Bounding box height (µm)",
                        "Fluorescent region area (pixels²) (global fluorescence contour, can exceed BF)",
                        "Fluorescent region area (µm²) (global fluorescence contour, can exceed BF)",
                        "Average fluorescence intensity (measured within BF region)",
                        "Median fluorescence intensity (measured within BF region)",
                        "Standard deviation of fluorescence (measured within BF region)",
                        "Minimum fluorescence value (measured within BF region)",
                        "Maximum fluorescence value (measured within BF region)",
                        "Total fluorescence signal (sum of BF-mask pixel values)",
                        "Fluorescence density normalized by brightfield area (IntegratedDensity/BF_Area_um2) - PRIMARY METRIC",
                        "Ratio of brightfield area to fluorescence area (BF_Area_um2/Fluor_Area_um2)",
                    ],
                    "Unit": [
                        "String",
                        "pixels²",
                        "µm²",
                        "pixels",
                        "µm",
                        "pixels",
                        "µm",
                        "dimensionless",
                        "dimensionless",
                        "pixels",
                        "pixels",
                        "µm",
                        "µm",
                        "pixels",
                        "pixels",
                        "pixels",
                        "pixels",
                        "µm",
                        "µm",
                        "pixels²",
                        "µm²",
                        "a.u.",
                        "a.u.",
                        "a.u.",
                        "a.u.",
                        "a.u.",
                        "a.u.",
                        "a.u./µm²",
                        "dimensionless",
                    ],
                }
            )

            readme_df.to_excel(writer, sheet_name="README", index=False)
            ws_readme = writer.sheets["README"]
            for cell in ws_readme[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            ws_readme.column_dimensions["A"].width = 30
            ws_readme.column_dimensions["B"].width = 80
            ws_readme.column_dimensions["C"].width = 15

            # --- Per-image sheets ---
            for csv_file in sorted(csv_files):
                image_name = csv_file.parent.name
                df = pd.read_csv(csv_file)

                if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                    df["Fluor_Density_per_BF_Area"] = (
                        pd.to_numeric(df["Fluor_IntegratedDensity"], errors="coerce")
                        / pd.to_numeric(df["BF_Area_um2"], errors="coerce")
                    )
                else:
                    df["Fluor_Density_per_BF_Area"] = 0.0

                if "Fluor_Area_um2" in df.columns and "BF_Area_um2" in df.columns:
                    df["BF_to_Fluor_Area_Ratio"] = (
                        df["BF_Area_um2"].astype(float) / df["Fluor_Area_um2"].astype(float)
                    )
                else:
                    df["BF_to_Fluor_Area_Ratio"] = 0.0

                df["Fluor_Density_per_BF_Area"] = df["Fluor_Density_per_BF_Area"].replace(
                    [np.inf, -np.inf], 0
                )
                df["BF_to_Fluor_Area_Ratio"] = df["BF_to_Fluor_Area_Ratio"].replace(
                    [np.inf, -np.inf], 0
                )
                df = df.fillna(0)

                # Sorting for color coding + typical selection
                df_sorted = df.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)

                # Filter out zero-fluorescence objects before percentile calculation
                df_sorted = df_sorted[df_sorted["Fluor_Density_per_BF_Area"] > 0].reset_index(drop=True)

                n_rows = len(df_sorted)
                # Use user-selected percentile
                top_threshold = int(np.ceil(n_rows * percentile))
                bottom_threshold = int(np.floor(n_rows * (1 - percentile)))

                typical_particles = df_sorted.iloc[top_threshold:bottom_threshold].copy()
                typical_particles["Source_Image"] = image_name
                all_typical_particles.append(typical_particles)

                sheet_name = image_name[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]

                # Header formatting
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align

                ws.auto_filter.ref = ws.dimensions

            # --- Restore {group}_Typical_Particles merged sheet (for group comparison) ---
            try:
                if all_typical_particles:
                    merged_typical = pd.concat(all_typical_particles, ignore_index=True)
                    if "Fluor_Density_per_BF_Area" in merged_typical.columns:
                        merged_typical = merged_typical.sort_values(
                            "Fluor_Density_per_BF_Area", ascending=False
                        )

                    typical_sheet_name = f"{group_name}_Typical_Particles"
                    merged_typical.to_excel(writer, sheet_name=typical_sheet_name, index=False)
                    ws_typ = writer.sheets[typical_sheet_name]

                    for cell in ws_typ[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    ws_typ.auto_filter.ref = ws_typ.dimensions
            except Exception as e:
                print(f"[WARN] Could not create {group_name}_Typical_Particles sheet: {e}")

            # --- Summary sheet ---
            try:
                summary_data = []
                for csv_file in sorted(csv_files):
                    image_name = csv_file.parent.name
                    df = pd.read_csv(csv_file)
                    
                    # Calculate Fluor_Density_per_BF_Area if not present
                    if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                        df["Fluor_Density_per_BF_Area"] = (
                            pd.to_numeric(df["Fluor_IntegratedDensity"], errors="coerce")
                            / pd.to_numeric(df["BF_Area_um2"], errors="coerce")
                        )
                        df["Fluor_Density_per_BF_Area"] = df["Fluor_Density_per_BF_Area"].replace([np.inf, -np.inf], 0)
                    
                    if "Fluor_Area_um2" in df.columns and "BF_Area_um2" in df.columns:
                        df["BF_to_Fluor_Area_Ratio"] = (
                            pd.to_numeric(df["BF_Area_um2"], errors="coerce")
                            / pd.to_numeric(df["Fluor_Area_um2"], errors="coerce")
                        )
                        df["BF_to_Fluor_Area_Ratio"] = df["BF_to_Fluor_Area_Ratio"].replace([np.inf, -np.inf], 0)
                    
                    # Calculate statistics for this image
                    summary_data.append({
                        'Image': image_name,
                        'Total_Particles': len(df),
                        'Avg_BF_Area_um2': df['BF_Area_um2'].mean() if 'BF_Area_um2' in df.columns else 0,
                        'Avg_Fluor_Density': df['Fluor_Density_per_BF_Area'].mean() if 'Fluor_Density_per_BF_Area' in df.columns else 0,
                        'Avg_BF_to_Fluor_Ratio': df['BF_to_Fluor_Area_Ratio'].mean() if 'BF_to_Fluor_Area_Ratio' in df.columns else 0
                    })
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                ws_summary = writer.sheets["Summary"]
                
                # Format header
                for cell in ws_summary[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    
            except Exception as e:
                print(f"[WARN] Could not create Summary sheet: {e}")

            # --- Ratios sheet (plots + support PNGs) ---
            try:
                wb = writer.book

                ratios_name = "Ratios"

                # Remove existing "Ratios" (and legacy "QA_Ratios") if present
                for legacy in (ratios_name, "QA_Ratios"):
                    if legacy in wb.sheetnames:
                        wb.remove(wb[legacy])

                ws_qa = wb.create_sheet(ratios_name)

                def add_png(ws, path: Path, anchor_cell: str, width_px: int = 360) -> None:
                    if not path.exists():
                        return
                    img = XLImage(str(path))
                    if getattr(img, "width", None) and getattr(img, "height", None):
                        scale = width_px / float(img.width)
                        img.width = int(img.width * scale)
                        img.height = int(img.height * scale)
                    ws.add_image(img, anchor_cell)

                ws_qa["A1"] = f"QA Ratios - Group: {group_name}"
                ws_qa["A1"].font = Font(bold=True, size=14)

                row = 3
                block_h = 34

                for csv_file in sorted(csv_files):
                    image_name = csv_file.parent.name
                    df = pd.read_csv(csv_file)

                    if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns:
                        df["Fluor_Density_per_BF_Area"] = (
                            pd.to_numeric(df["Fluor_IntegratedDensity"], errors="coerce")
                            / pd.to_numeric(df["BF_Area_um2"], errors="coerce")
                        )
                    else:
                        df["Fluor_Density_per_BF_Area"] = 0.0

                    if "Fluor_Area_um2" in df.columns and "BF_Area_um2" in df.columns:
                        df["BF_to_Fluor_Area_Ratio"] = (
                            pd.to_numeric(df["BF_Area_um2"], errors="coerce")
                            / pd.to_numeric(df["Fluor_Area_um2"], errors="coerce")
                        )
                    else:
                        df["BF_to_Fluor_Area_Ratio"] = 0.0

                    df["Fluor_Density_per_BF_Area"] = (
                        df["Fluor_Density_per_BF_Area"]
                        .replace([np.inf, -np.inf], 0)
                        .fillna(0)
                    )
                    df["BF_to_Fluor_Area_Ratio"] = (
                        df["BF_to_Fluor_Area_Ratio"]
                        .replace([np.inf, -np.inf], 0)
                        .fillna(0)
                    )

                    # SORT HERE - BEFORE writing rows
                    df = df.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)

                    ws_qa[f"A{row}"] = image_name
                    ws_qa[f"A{row}"].font = Font(bold=True, size=12)
                    row0 = row

                    # SWAPPED: Now Fluor_Density on left (A,B), BF_to_Fluor on right (D,E)
                    ws_qa[f"A{row+1}"] = "Object_ID"
                    ws_qa[f"B{row+1}"] = "Fluor_Density_per_BF_Area"
                    ws_qa[f"D{row+1}"] = "Object_ID"
                    ws_qa[f"E{row+1}"] = "BF_to_Fluor_Area_Ratio"
                    for c in ["A", "B", "D", "E"]:
                        ws_qa[f"{c}{row+1}"].font = Font(bold=True)
                        ws_qa[f"{c}{row+1}"].alignment = Alignment(horizontal="center")

                    start_data_row = row + 2

                    # Column widths
                    ws_qa.column_dimensions["A"].width = 16  # Object_ID
                    ws_qa.column_dimensions["B"].width = 26  # Fluor_Density_per_BF_Area
                    ws_qa.column_dimensions["C"].width = 3   # spacer
                    ws_qa.column_dimensions["D"].width = 16  # Object_ID
                    ws_qa.column_dimensions["E"].width = 22  # BF_to_Fluor_Area_Ratio

                    for k, r in enumerate(df.itertuples(index=False), 0):
                        ws_qa[f"A{start_data_row+k}"] = getattr(r, "Object_ID")
                        ws_qa[f"B{start_data_row+k}"] = float(
                            getattr(r, "Fluor_Density_per_BF_Area", 0.0)
                        )
                        ws_qa[f"D{start_data_row+k}"] = getattr(r, "Object_ID")
                        ws_qa[f"E{start_data_row+k}"] = float(
                            getattr(r, "BF_to_Fluor_Area_Ratio", 0.0)
                        )

                    n = len(df)
                    if n > 0:
                        # SWAPPED Chart 1: Now Fluor_Density on left
                        ch1 = ScatterChart()
                        ch1.title = "Fluor intensity / BF area (Fluor_Density_per_BF_Area)"
                        ch1.y_axis.title = "a.u./µm²"
                        ch1.x_axis.title = "Object_ID"
                        xref = Reference(
                            ws_qa,
                            min_col=1,
                            min_row=start_data_row,
                            max_row=start_data_row + n - 1,
                        )
                        yref = Reference(
                            ws_qa,
                            min_col=2,
                            min_row=start_data_row,
                            max_row=start_data_row + n - 1,
                        )
                        s1 = cast(Any, SeriesFactory(yref, xref))
                        try:
                            s1.title = "Fluor Density"
                        except Exception:
                            pass
                        try:
                            s1.marker = Marker(symbol="triangle", size=5)
                        except Exception:
                            pass
                        ch1.series.append(s1)
                        ws_qa.add_chart(ch1, f"G{row+1}")

                        # SWAPPED Chart 2: Now BF_to_Fluor on right
                        ch2 = ScatterChart()
                        ch2.title = "BF area / FL area (BF_to_Fluor_Area_Ratio)"
                        ch2.y_axis.title = "Ratio"
                        ch2.x_axis.title = "Object_ID"
                        xref2 = Reference(
                            ws_qa,
                            min_col=4,
                            min_row=start_data_row,
                            max_row=start_data_row + n - 1,
                        )
                        yref2 = Reference(
                            ws_qa,
                            min_col=5,
                            min_row=start_data_row,
                            max_row=start_data_row + n - 1,
                        )
                        s2 = cast(Any, SeriesFactory(yref2, xref2))
                        try:
                            s2.title = "BF/FL Area Ratio"
                        except Exception:
                            pass
                        try:
                            s2.marker = Marker(symbol="circle", size=5)
                        except Exception:
                            pass
                        ch2.series.append(s2)
                        ws_qa.add_chart(ch2, f"G{row+18}")

                    img_dir = csv_file.parent
                    add_png(ws_qa, img_dir / "13_mask_accepted_ids.png", f"Q{row+1}", width_px=330)
                    add_png(ws_qa, img_dir / "22_fluorescence_mask_global_ids.png", f"Q{row+18}", width_px=330)
                    add_png(ws_qa, img_dir / "23_fluorescence_contours_global_ids.png", f"Q{row+18}", width_px=330)
                    add_png(ws_qa, img_dir / "24_bf_fluor_matching_overlay_ids.png", f"Q{row+1}", width_px=330)

                    row = row0 + block_h

            except Exception as e:
                print(f"[WARN] Could not create Ratios sheet: {e}")

        # Reorder worksheets AFTER ExcelWriter closes
        wb = load_workbook(excel_path)
        desired_order = ["Summary", "Ratios", "README", f"{group_name}_Typical_Particles"]

        # Move sheets to desired positions
        for idx, sheet_name in enumerate(desired_order):
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                current_idx = wb.sheetnames.index(sheet_name)
                if current_idx != idx:
                    wb.move_sheet(sheet, offset=idx - current_idx)

        wb.save(excel_path)

        print(f"Excel consolidation saved: {excel_path}")
        print("  - Ratios sheet with ratio plots + support PNGs")

    except PermissionError:
        print(f"[ERROR] Cannot write to {excel_path} - file may be open")
        print("        Close Excel and try again")
    except Exception as e:
        print(f"[ERROR] Failed to create Excel file: {e}")
        import traceback
        traceback.print_exc()
        
# ==================================================
# Main processing
# ==================================================
def process_image(img_path: Path, output_root: Path) -> None:
    print("\n" + "=" * 80)
    print(f"Processing: {img_path}")

    xml_props, xml_main = find_metadata_paths(img_path)
    print(f"Metadata (Properties): {xml_props}")
    print(f"Metadata (Main):       {xml_main}")

    try:
        um_per_px_x, um_per_px_y = get_pixel_size_um(xml_props, xml_main)
        um_per_px_x = float(um_per_px_x)
        um_per_px_y = float(um_per_px_y)
    except Exception as e:
        if FALLBACK_UM_PER_PX is None:
            raise
        print(f"[WARN] {e} -> using fallback pixel size {FALLBACK_UM_PER_PX} µm/px")
        um_per_px_x = um_per_px_y = float(FALLBACK_UM_PER_PX)

    um_per_px_avg = (um_per_px_x + um_per_px_y) / 2.0
    print(
        f"Pixel size: X={um_per_px_x:.6f} µm/px, Y={um_per_px_y:.6f} µm/px (avg={um_per_px_avg:.6f})"
    )

    img_out = output_root / img_path.stem
    ensure_dir(img_out)

    img = cv2.imread(str(img_path), cv2.IMREAD_UNCHANGED)
    if img is None:
        raise FileNotFoundError(str(img_path))
    if img.ndim == 3:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    print(f"Loaded: dtype={img.dtype}, shape={img.shape}, range=[{img.min()}-{img.max()}]")

    img8 = normalize_to_8bit(img)
    save_debug(img_out, "01_gray_8bit.png", img8, um_per_px_avg)

    mask = segment_particles_brightfield(img8, float(um_per_px_avg), img_out)

    _fc = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = cast(list[np.ndarray], _fc[-2])

    print(f"Contours found (pre-filter): {len(contours)}")

    vis_all = cv2.cvtColor(img8, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(vis_all, contours, -1, (0, 0, 255), 1)
    save_debug(img_out, "10_contours_all.png", vis_all, um_per_px_avg)

    um2_per_px2 = float(um_per_px_x) * float(um_per_px_y)
    min_area_px = MIN_AREA_UM2 / um2_per_px2
    max_area_px = MAX_AREA_UM2 / um2_per_px2

    H, W = img8.shape[:2]
    img_area_px = float(H * W)
    max_big_area_px = MAX_FRACTION_OF_IMAGE_AREA * img_area_px

    accepted: list[np.ndarray] = []
    rejected: list[np.ndarray] = []

    for c in contours:
        area_px = float(cv2.contourArea(c))
        if area_px <= 0:
            rejected.append(c)
            continue

        if area_px >= max_big_area_px:
            rejected.append(c)
            continue

        perim_px = float(cv2.arcLength(c, True))
        circ = (4 * np.pi * area_px / (perim_px**2)) if perim_px > 0 else 0.0

        ok = (min_area_px <= area_px <= max_area_px) and (circ >= MIN_CIRCULARITY)
        (accepted if ok else rejected).append(c)

    print(f"Accepted: {len(accepted)} | Rejected: {len(rejected)}")
    print(
        f"Filter thresholds: area [{MIN_AREA_UM2}-{MAX_AREA_UM2}] µm² "
        f"(~[{min_area_px:.1f}-{max_area_px:.1f}] px²), circularity >= {MIN_CIRCULARITY}, "
        f"max_single_contour_area <= {MAX_FRACTION_OF_IMAGE_AREA:.0%} of image"
    )

    vis_acc = cv2.cvtColor(img8, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(vis_acc, rejected, -1, (0, 165, 255), 1)
    cv2.drawContours(vis_acc, accepted, -1, (0, 0, 255), 1)
    vis_acc = draw_object_ids(vis_acc, accepted)
    save_debug(img_out, "11_contours_rejected_orange_accepted_red_ids_green.png", vis_acc, um_per_px_avg)

    mask_all = np.zeros_like(mask)
    cv2.drawContours(mask_all, contours, -1, 255, thickness=-1)
    save_debug(img_out, "12_mask_all.png", mask_all)

    mask_acc = np.zeros_like(mask)
    cv2.drawContours(mask_acc, accepted, -1, 255, thickness=-1)
    save_debug(img_out, "13_mask_accepted.png", mask_acc)

    fluor_path = img_path.parent / img_path.name.replace("_ch00", "_ch01")
    fluor_measurements: Optional[list[dict]] = None

    fluor_bw: Optional[np.ndarray] = None
    vis_fluor: Optional[np.ndarray] = None
    vis_match: Optional[np.ndarray] = None

    if fluor_path.exists():
        fluor_img = cv2.imread(str(fluor_path), cv2.IMREAD_UNCHANGED)
        if fluor_img is not None:
            if fluor_img.ndim == 3:
                fluor_img = cv2.cvtColor(fluor_img, cv2.COLOR_BGR2GRAY)

            print(
                f"Fluorescence loaded: dtype={fluor_img.dtype}, range=[{fluor_img.min()}-{fluor_img.max()}]"
            )

            fluor_img8 = normalize_to_8bit(fluor_img)
            save_debug(img_out, "20_fluorescence_8bit.png", fluor_img8, um_per_px_avg)

            fluor_bw = segment_fluorescence_global(fluor_img8)
            save_debug(img_out, "22_fluorescence_mask_global.png", fluor_bw, um_per_px_avg)

            _fc2 = cv2.findContours(fluor_bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            fluor_contours_all = cast(list[np.ndarray], _fc2[-2])

            min_fluor_area_px = FLUOR_MIN_AREA_UM2 / um2_per_px2 if um2_per_px2 > 0 else 0.0
            fluor_contours = [
                c for c in fluor_contours_all if float(cv2.contourArea(c)) >= float(min_fluor_area_px)
            ]

            vis_fluor = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
            cv2.drawContours(vis_fluor, fluor_contours, -1, (0, 255, 0), 1)
            save_debug(img_out, "23_fluorescence_contours_global.png", vis_fluor, um_per_px_avg)

            matches = match_fluor_to_bf_by_overlap(accepted, fluor_contours, fluor_img8.shape[:2])

            # DIAGNOSTIC LOGGING
            unmatched = sum(1 for m in matches if m is None)
            print(f"Fluorescence matching: {len(accepted) - unmatched}/{len(accepted)} BF objects matched")
            print(f"Total fluorescence contours available: {len(fluor_contours)}")
            if unmatched > 0:
                print(f"  → {unmatched} BF objects have NO fluorescence match (will have Fluor_Area_px=0)")

            fluor_measurements = measure_fluorescence_intensity_with_global_area(
                fluor_img, accepted, fluor_contours, matches, float(um_per_px_x), float(um_per_px_y)
            )

            vis_match = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
            for idx, bf_c in enumerate(accepted):
                j = matches[idx]
                cv2.drawContours(vis_match, [bf_c], -1, (0, 0, 255), 1)
                if j is not None:
                    cv2.drawContours(vis_match, [fluor_contours[j]], -1, (0, 255, 0), 2)
            save_debug(img_out, "24_bf_fluor_matching_overlay.png", vis_match, um_per_px_avg)

            fluor_overlay = visualize_fluorescence_measurements(
                fluor_img8, accepted, fluor_measurements
            )
            save_debug(img_out, "21_fluorescence_overlay.png", fluor_overlay, um_per_px_avg)
        else:
            print(f"[WARN] Could not load fluorescence image: {fluor_path}")
    else:
        print(f"[WARN] Fluorescence channel not found: {fluor_path}")

    parts = img_path.stem.split()
    group_id = parts[0] if parts else "unk"
    sequence_num = parts[-1].split("_")[0] if parts else "0"

    object_ids = [f"{group_id}_{sequence_num}_{i}" for i in range(1, len(accepted) + 1)]

    mask_acc_bgr = cv2.cvtColor(mask_acc, cv2.COLOR_GRAY2BGR)
    save_debug_ids(img_out, "13_mask_accepted.png", mask_acc_bgr, accepted, object_ids, um_per_px_avg)

    if fluor_bw is not None:
        fluor_bw_bgr = cv2.cvtColor(fluor_bw, cv2.COLOR_GRAY2BGR)
        save_debug_ids(
            img_out,
            "22_fluorescence_mask_global.png",
            fluor_bw_bgr,
            accepted,
            object_ids,
            um_per_px_avg,
        )
    if vis_fluor is not None:
        save_debug_ids(
            img_out,
            "23_fluorescence_contours_global.png",
            vis_fluor,
            accepted,
            object_ids,
            um_per_px_avg,
        )
    if vis_match is not None:
        save_debug_ids(
            img_out,
            "24_bf_fluor_matching_overlay.png",
            vis_match,
            accepted,
            object_ids,
            um_per_px_avg,
        )

    csv_path = img_out / "object_stats.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "Object_ID",
                "BF_Area_px",
                "BF_Area_um2",
                "Perimeter_px",
                "Perimeter_um",
                "EquivDiameter_px",
                "EquivDiameter_um",
                "Circularity",
                "AspectRatio",
                "CentroidX_px",
                "CentroidY_px",
                "CentroidX_um",
                "CentroidY_um",
                "BBoxX_px",
                "BBoxY_px",
                "BBoxW_px",
                "BBoxH_px",
                "BBoxW_um",
                "BBoxH_um",
                "Fluor_Area_px",
                "Fluor_Area_um2",
                "Fluor_Mean",
                "Fluor_Median",
                "Fluor_Std",
                "Fluor_Min",
                "Fluor_Max",
                "Fluor_IntegratedDensity",
            ]
        )

        for i, c in enumerate(accepted, 1):
            compound_id = f"{group_id}_{sequence_num}_{i}"

            area_px = float(cv2.contourArea(c))
            area_um2 = area_px * (float(um_per_px_x) * float(um_per_px_y))

            perim_px = float(cv2.arcLength(c, True))
            perim_um = contour_perimeter_um(c, float(um_per_px_x), float(um_per_px_y))

            eqd_px = equivalent_diameter_from_area(area_px)
            eqd_um = equivalent_diameter_from_area(area_um2)

            circ = (4 * np.pi * area_px / (perim_px**2)) if perim_px > 0 else 0.0

            x, y, bw, bh = cv2.boundingRect(c)
            aspect = (float(bw) / float(bh)) if bh > 0 else 0.0

            M = cv2.moments(c)
            if M["m00"] != 0:
                cx = float(M["m10"] / M["m00"])
                cy = float(M["m01"] / M["m00"])
            else:
                cx, cy = 0.0, 0.0

            cx_um = cx * float(um_per_px_x)
            cy_um = cy * float(um_per_px_y)
            bw_um = float(bw) * float(um_per_px_x)
            bh_um = float(bh) * float(um_per_px_y)

            if fluor_measurements is not None:
                fm = fluor_measurements[i - 1]
            else:
                fm = {
                    "fluor_area_px": 0.0,
                    "fluor_area_um2": 0.0,
                    "fluor_mean": 0.0,
                    "fluor_median": 0.0,
                    "fluor_std": 0.0,
                    "fluor_min": 0.0,
                    "fluor_max": 0.0,
                    "fluor_integrated_density": 0.0,
                }

            w.writerow(
                [
                    compound_id,
                    f"{area_px:.2f}",
                    f"{area_um2:.4f}",
                    f"{perim_px:.2f}",
                    f"{perim_um:.4f}",
                    f"{eqd_px:.2f}",
                    f"{eqd_um:.4f}",
                    f"{circ:.4f}",
                    f"{aspect:.4f}",
                    f"{cx:.2f}",
                    f"{cy:.2f}",
                    f"{cx_um:.4f}",
                    f"{cy_um:.4f}",
                    x,
                    y,
                    bw,
                    bh,
                    f"{bw_um:.4f}",
                    f"{bh_um:.4f}",
                    f"{float(fm['fluor_area_px']):.2f}",
                    f"{float(fm['fluor_area_um2']):.4f}",
                    f"{float(fm['fluor_mean']):.2f}",
                    f"{float(fm['fluor_median']):.2f}",
                    f"{float(fm['fluor_std']):.2f}",
                    f"{float(fm['fluor_min']):.2f}",
                    f"{float(fm['fluor_max']):.2f}",
                    f"{float(fm['fluor_integrated_density']):.2f}",
                ]
            )

    print(f"CSV saved: {csv_path} ({len(accepted)} objects)")
    print("✓ Done")




def main() -> None:
    if CLEAR_OUTPUT_DIR_EACH_RUN:
        clear_output_dir(OUTPUT_DIR)

    print(f"Input dir: {_project_root.resolve()}")
    
    # Get percentile choice early
    percentile = get_percentile_option()
    
    groups = list_sample_group_folders(SOURCE_DIR)
    selected_group_dir = prompt_user_select_group(groups)

    if selected_group_dir is None:
        dirs_to_process = groups[:]
    else:
        dirs_to_process = [selected_group_dir]

    if CONTROL_DIR.exists():
        dirs_to_process.append(CONTROL_DIR)

    img_paths: list[Path] = []
    for d in dirs_to_process:
        img_paths.extend(sorted(d.rglob(IMAGE_GLOB)))

    print(f"Found {len(img_paths)} brightfield images matching '{IMAGE_GLOB}'")
    if not img_paths:
        raise FileNotFoundError(f"No images found under {SOURCE_DIR} matching {IMAGE_GLOB}")

    total_processed = 0
    total_failed = 0

    for p in tqdm(img_paths, desc="Processing images", unit="img"):
        out_root = (OUTPUT_DIR / p.parent.name) if SEPARATE_OUTPUT_BY_GROUP else OUTPUT_DIR
        ensure_dir(out_root)

        try:
            process_image(p, out_root)
            total_processed += 1
        except Exception as e:
            tqdm.write(f"[ERROR] Failed processing {p}: {e}")
            total_failed += 1

    print(f"\n{'=' * 80}")
    print(f"SUMMARY: {total_processed} succeeded, {total_failed} failed")

    if SEPARATE_OUTPUT_BY_GROUP:
        for group_dir in OUTPUT_DIR.iterdir():
            if group_dir.is_dir():
                consolidate_to_excel(group_dir, group_dir.name, percentile)

        print(f"\n{'=' * 80}")
        print("Generating error bar comparison plots...")
        generate_error_bar_comparison(OUTPUT_DIR, percentile)

        print("Embedding comparison plots into master Excel files...")
        embed_comparison_plots_into_all_excels(OUTPUT_DIR, percentile)


if __name__ == "__main__":
    main()