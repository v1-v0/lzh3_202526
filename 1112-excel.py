#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Bacteria Segmentation Batch Processor (Enhanced)
- Processes BF + Fluor image pairs from source/ directory (with subdirectory support)
- Filters out AppleDouble files (._*) and hidden files
- Exports one worksheet per pair to a single Excel file
- Each worksheet contains:
    • 3 embedded images (BF, Fluor, Overlay with labels)
    • Detection statistics table
- LANDSCAPE A4 format for larger images
- Compact layout with maximum image size
- Smart label positioning matching interactive viewer
- Output: outputs/bacterial_analysis_results.xlsx

File Naming Convention:
    - Brightfield: *_ch00.tif
    - Fluorescence: *_ch01.tif

Configuration matches the interactive viewer (dev.py)
"""

import cv2
import numpy as np
from pathlib import Path
from scipy import ndimage
from typing import List, Tuple, Dict, Optional, cast
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.page import PrintPageSetup
import tempfile
import os
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont

# --------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------- #
INPUT_DIR = Path("source")
OUTPUT_DIR = Path("outputs")
OUTPUT_FILE = OUTPUT_DIR / "bacterial_analysis_results.xlsx"
MAX_IMAGES = 50
TARGET_IMAGE_SIZE = (480, 480)  # Width, Height for embedded images (LARGER)

OUTPUT_DIR.mkdir(exist_ok=True)

# Match parameters from dev.py
PARAMS = {
    'use_otsu': False,
    'manual_threshold': 110,
    'enable_clahe': True,
    'clahe_clip': 5.0,
    'clahe_tile': 32,
    'open_kernel': 3,
    'close_kernel': 5,
    'open_iter': 3,
    'close_iter': 2,
    'min_area': 50,
    'watershed_dilate': 10,
    'fluor_brightness': 2.0,
    'fluor_gamma': 0.5,
    # Label settings (matching interactive viewer)
    'show_labels': True,
    'label_font_size': 20,
    'arrow_length': 60,
    'label_offset': 15,
    'min_fluor_per_area': 0.1,
}

# --------------------------------------------------------------------- #
# Helper: Validate image files (filter AppleDouble and hidden files)
# --------------------------------------------------------------------- #
def is_valid_image_file(filepath: Path) -> bool:
    """
    Check if a file is a valid image file (not an AppleDouble or hidden file).
    
    Returns True if:
    - Filename doesn't start with '._' (AppleDouble files)
    - Filename doesn't start with '.' (hidden files)
    - File ends with _ch00.tif (brightfield)
    """
    filename = filepath.name
    
    # Filter out AppleDouble files (._filename)
    if filename.startswith('._'):
        return False
    
    # Filter out hidden files (.filename)
    if filename.startswith('.'):
        return False
    
    # Must be a brightfield file
    if not filename.endswith('_ch00.tif'):
        return False
    
    return True

def get_fluorescence_path(brightfield_path: Path) -> Optional[Path]:
    """Get the corresponding fluorescence image path for a brightfield image."""
    if not brightfield_path.name.endswith('_ch00.tif'):
        return None
    
    fluor_path = brightfield_path.parent / brightfield_path.name.replace('_ch00.tif', '_ch01.tif')
    
    # Check if it's a valid file (not AppleDouble)
    if fluor_path.exists() and not fluor_path.name.startswith('._'):
        return fluor_path
    return None

# --------------------------------------------------------------------- #
# Helper: Find image pairs
# --------------------------------------------------------------------- #
def find_image_pairs(input_dir: Path) -> List[Tuple[Path, Path, str, str]]:
    """
    Find all valid BF + Fluor image pairs in the input directory recursively.
    
    Returns:
        List of tuples: (bf_path, fluor_path, index, relative_path)
    """
    if not input_dir.exists():
        print(f"❌ Error: Input directory '{input_dir}' not found.")
        return []
    
    # Find all _ch00.tif files recursively (brightfield only)
    all_bf_files = list(input_dir.glob('**/*_ch00.tif'))
    
    # Filter out AppleDouble and hidden files
    valid_bf_files = [f for f in all_bf_files if is_valid_image_file(f)]
    
    # Sort by string representation of path for consistent ordering
    bf_files = sorted(valid_bf_files, key=lambda p: str(p))
    
    pairs = []
    missing_pairs = []
    
    print(f"\n{'='*70}")
    print(f"Scanning for image pairs in: {input_dir.absolute()}")
    print(f"{'='*70}")
    
    for bf_path in bf_files:
        if len(pairs) >= MAX_IMAGES:
            print(f"\n⚠ Reached maximum of {MAX_IMAGES} images. Stopping scan.")
            break
        
        # Get corresponding fluorescence image
        fluor_path = get_fluorescence_path(bf_path)
        
        if fluor_path:
            index = f"{len(pairs)+1:02d}"
            # Get relative path for better readability
            try:
                rel_path = bf_path.relative_to(input_dir)
            except ValueError:
                rel_path = bf_path.name
            
            pairs.append((bf_path, fluor_path, index, str(rel_path)))
            print(f"  ✓ Pair {index}: {rel_path}")
        else:
            missing_pairs.append(bf_path)
    
    if missing_pairs:
        print(f"\n{'='*70}")
        print(f"⚠ Warning: {len(missing_pairs)} BF files missing Fluor pairs:")
        print(f"{'='*70}")
        for bf_path in missing_pairs[:10]:  # Show first 10
            try:
                rel_path = bf_path.relative_to(input_dir)
            except ValueError:
                rel_path = bf_path.name
            print(f"  - {rel_path}")
        if len(missing_pairs) > 10:
            print(f"  ... and {len(missing_pairs) - 10} more")
    
    print(f"\n{'='*70}")
    print(f"✓ Total valid pairs found: {len(pairs)}")
    print(f"{'='*70}\n")
    
    return pairs

# --------------------------------------------------------------------- #
# Segmentation function
# --------------------------------------------------------------------- #
def segment_bacteria(gray_bf: np.ndarray) -> Tuple[np.ndarray, np.ndarray, np.ndarray, List[np.ndarray]]:
    """
    Perform bacteria segmentation using CLAHE, thresholding, morphology, and watershed.
    
    Returns:
        Tuple of (enhanced, threshold, cleaned, bacteria_contours)
    """
    # CLAHE Enhancement
    if PARAMS['enable_clahe']:
        bf8 = cv2.convertScaleAbs(gray_bf)
        clahe = cv2.createCLAHE(clipLimit=PARAMS['clahe_clip'],
                                tileGridSize=(PARAMS['clahe_tile'],)*2)
        enhanced = clahe.apply(bf8)
    else:
        enhanced = cv2.convertScaleAbs(gray_bf)

    # Thresholding
    if PARAMS['use_otsu']:
        _, thresh = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    else:
        _, thresh = cv2.threshold(enhanced, PARAMS['manual_threshold'], 255, cv2.THRESH_BINARY_INV)

    # Morphological Operations
    open_k_size = max(1, PARAMS['open_kernel'] - (PARAMS['open_kernel'] % 2 == 0))
    close_k_size = max(1, PARAMS['close_kernel'] - (PARAMS['close_kernel'] % 2 == 0))
    open_k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (open_k_size, open_k_size))
    close_k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (close_k_size, close_k_size))

    opened = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, open_k, iterations=PARAMS['open_iter'])
    cleaned = cv2.morphologyEx(opened, cv2.MORPH_CLOSE, close_k, iterations=PARAMS['close_iter'])

    # Watershed Segmentation
    distance = cv2.distanceTransform(cleaned, cv2.DIST_L2, 5)
    _, sure_fg = cv2.threshold(distance, PARAMS['watershed_dilate'] * distance.max() / 100.0, 255, 0)
    sure_fg = sure_fg.astype(np.uint8)
    sure_fg = cv2.morphologyEx(sure_fg, cv2.MORPH_OPEN, open_k, iterations=1)

    markers, _ = cast(Tuple[np.ndarray, int], ndimage.label(sure_fg))
    markers = markers.astype(np.int32)
    markers += 1
    markers[cleaned == 0] = 0

    watershed_input = cv2.cvtColor(gray_bf, cv2.COLOR_GRAY2BGR)
    markers = cv2.watershed(watershed_input, markers)

    contour_mask = (markers > 1).astype(np.uint8) * 255

    # Find contours (compatible with OpenCV 3.x and 4.x)
    contour_result = cv2.findContours(contour_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = cast(List[np.ndarray], contour_result[-2])

    # Filter by minimum area
    bacteria = [c for c in contours if cv2.contourArea(c) >= PARAMS['min_area']]
    
    return enhanced, thresh, cleaned, bacteria

# --------------------------------------------------------------------- #
# Statistics (with filter and sort matching dev.py)
# --------------------------------------------------------------------- #
def get_processed_stats_and_contours(contours: List[np.ndarray], bf_img: np.ndarray, fluor_img: Optional[np.ndarray], min_fpa: float) -> Tuple[pd.DataFrame, Dict, List[np.ndarray]]:
    """
    Calculate per-bacterium statistics, apply filter/sort matching dev.py, and return df, summary, filtered/sorted contours.
    """
    all_stats: List[Dict] = []
    
    for contour in contours:
        area = cv2.contourArea(contour)
        
        # Create mask for this bacterium
        mask = np.zeros(bf_img.shape, dtype=np.uint8)
        cv2.drawContours(mask, [contour], -1, 255, -1)
        
        # Calculate fluorescence statistics
        fluor_mean = 0.0
        fluor_total = 0.0
        fluor_per_area = 0.0
        if fluor_img is not None:
            values = fluor_img[mask == 255]
            if len(values) > 0:
                fluor_mean = float(np.mean(values))
                fluor_total = float(np.sum(values))
                fluor_per_area = fluor_total / area if area > 0 else 0.0
        
        all_stats.append({
            'contour': contour,
            'area': area,
            'fluor_mean': fluor_mean,
            'fluor_total': fluor_total,
            'fluor_per_area': fluor_per_area
        })
    
    # Filter by min_fluor_per_area (matching dev.py)
    if fluor_img is not None and min_fpa > 0:
        filtered_stats = [s for s in all_stats if s['fluor_per_area'] >= min_fpa]
    else:
        filtered_stats = all_stats
    
    # Sort by fluor_per_area descending (matching default in dev.py)
    filtered_stats.sort(key=lambda s: s['fluor_per_area'], reverse=True)
    
    # Extract sorted contours
    filtered_contours = [s['contour'] for s in filtered_stats]
    
    # Create rounded stats for df
    stats_list = []
    for idx, s in enumerate(filtered_stats, 1):
        stats_list.append({
            'ID': idx,
            'Area (px²)': round(s['area'], 1),
            'Fluor Mean': round(s['fluor_mean'], 2),
            'Fluor Total': round(s['fluor_total'], 1),
            'Fluor/Area': round(s['fluor_per_area'], 3)
        })
    
    df = pd.DataFrame(stats_list)

    # Calculate summary statistics
    total_bacteria = len(filtered_stats)
    total_area = df['Area (px²)'].sum() if not df.empty else 0
    image_area = bf_img.shape[0] * bf_img.shape[1]
    coverage = (total_area / image_area) * 100 if image_area > 0 else 0
    avg_area = df['Area (px²)'].mean() if not df.empty else 0
    avg_fluor_area = df['Fluor/Area'].mean() if not df.empty else 0

    summary = {
        'Total Bacteria': total_bacteria,
        'Total Area (px²)': round(total_area, 1),
        'Coverage (%)': round(coverage, 3),
        'Avg Area (px²)': round(avg_area, 1),
        'Avg Fluor/Area': round(avg_fluor_area, 3)
    }
    
    return df, summary, filtered_contours

# --------------------------------------------------------------------- #
# Smart Label Positioning (matching interactive viewer)
# --------------------------------------------------------------------- #
def create_occupancy_map(img_shape, contours, margin=20):
    """
    Create a binary occupancy map showing where objects and borders are.
    """
    h, w = img_shape
    occupancy = np.zeros((h, w), dtype=np.uint8)
    
    # Mark all contours as occupied (with margin)
    for contour in contours:
        cv2.drawContours(occupancy, [contour], -1, 255, -1)
        if margin > 0:
            kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (margin*2, margin*2))
            occupancy = cv2.dilate(occupancy, kernel, iterations=1)
    
    # Mark borders as occupied
    border_margin = margin
    occupancy[:border_margin, :] = 255  # Top
    occupancy[-border_margin:, :] = 255  # Bottom
    occupancy[:, :border_margin] = 255  # Left
    occupancy[:, -border_margin:] = 255  # Right
    
    return occupancy

def calculate_direction_score(centroid, angle, arrow_len, label_size, occupancy_map, offset):
    """
    Calculate a score for placing a label in a given direction.
    Lower score = better placement.
    """
    cx, cy = centroid
    h, w = occupancy_map.shape
    label_w, label_h = label_size
    rad = np.deg2rad(angle)
    arrow_x = int(cx + arrow_len * np.cos(rad))
    arrow_y = int(cy - arrow_len * np.sin(rad))
    label_x = int(arrow_x + offset * np.cos(rad) - label_w / 2)
    label_y = int(arrow_y - offset * np.sin(rad) - label_h / 2)
    if (label_x < 0 or label_x + label_w >= w or
        label_y < 0 or label_y + label_h >= h):
        return float('inf')
    label_region = occupancy_map[label_y:label_y+label_h, label_x:label_x+label_w]
    occupied_pixels = np.sum(label_region > 0)
    total_pixels = label_region.size
    if total_pixels == 0:
        return float('inf')
    arrow_score = 0
    num_samples = 10
    for i in range(num_samples):
        t = i / num_samples
        sx = int(cx + t * arrow_len * np.cos(rad))
        sy = int(cy - t * arrow_len * np.sin(rad))
        if 0 <= sx < w and 0 <= sy < h and occupancy_map[sy, sx] > 0:
            arrow_score += 10
    score = (occupied_pixels / total_pixels) * 100 + arrow_score
    return score

def find_best_label_position(centroid, img_shape, arrow_len, label_size, offset, occupancy_map):
    """Find best position for label by testing multiple angles."""
    cx, cy = centroid
    h, w = img_shape
    label_w, label_h = label_size
    angles = [i * 22.5 for i in range(16)]
    best_score = float('inf')
    best_pos = None
    for angle in angles:
        score = calculate_direction_score(
            centroid, angle, arrow_len, label_size, occupancy_map, offset
        )
        if score < best_score:
            best_score = score
            rad = np.deg2rad(angle)
            arrow_x = int(cx + arrow_len * np.cos(rad))
            arrow_y = int(cy - arrow_len * np.sin(rad))
            label_x = int(arrow_x + offset * np.cos(rad) - label_w / 2)
            label_y = int(arrow_y - offset * np.sin(rad) - label_h / 2)
            label_x = max(0, min(label_x, w - label_w))
            label_y = max(0, min(label_y, h - label_h))
            best_pos = (arrow_x, arrow_y, label_x, label_y, angle)
    return best_pos

def get_label_font():
    """Get font for labels, trying multiple system font paths."""
    font_size = PARAMS['label_font_size']
    font_paths = [
        "arial.ttf",
        "/System/Library/Fonts/Helvetica.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "C:\\Windows\\Fonts\\arial.ttf",
    ]
    for fp in font_paths:
        try:
            return ImageFont.truetype(fp, font_size)
        except (OSError, IOError):
            continue
    return ImageFont.load_default()

def draw_labels_on_contours(img_bgr, contours):
    """Draw numbered labels on bacteria contours with smart positioning."""
    if not PARAMS['show_labels'] or not contours:
        return img_bgr
    img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    pil_img = Image.fromarray(img_rgb)
    draw = ImageDraw.Draw(pil_img)
    arrow_len = PARAMS['arrow_length']
    label_offset = PARAMS['label_offset']
    font = get_label_font()
    h, w = img_bgr.shape[:2]
    occupancy_map = create_occupancy_map((h, w), contours, margin=20)
    for idx, contour in enumerate(contours, 1):
        M = cv2.moments(contour)
        if M["m00"] == 0:
            continue
        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])
        label_text = str(idx)
        bbox = draw.textbbox((0, 0), label_text, font=font)
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
        result = find_best_label_position(
            (cx, cy), (h, w), arrow_len, (text_w, text_h), label_offset, occupancy_map
        )
        if result is None:
            continue
        arrow_x, arrow_y, label_x, label_y, angle = result
        arrow_color = (255, 255, 0)
        arrow_width = 2
        draw.line([(cx, cy), (arrow_x, arrow_y)], fill=arrow_color, width=arrow_width)
        head_len = 8
        head_angle = 25
        angle_rad = np.deg2rad(angle)
        left_angle = angle_rad + np.deg2rad(180 - head_angle)
        left_x = int(arrow_x + head_len * np.cos(left_angle))
        left_y = int(arrow_y - head_len * np.sin(left_angle))
        draw.line([(arrow_x, arrow_y), (left_x, left_y)], fill=arrow_color, width=arrow_width)
        right_angle = angle_rad + np.deg2rad(180 + head_angle)
        right_x = int(arrow_x + head_len * np.cos(right_angle))
        right_y = int(arrow_y - head_len * np.sin(right_angle))
        draw.line([(arrow_x, arrow_y), (right_x, right_y)], fill=arrow_color, width=arrow_width)
        padding = 4
        bg_rect = [
            label_x - padding, label_y - padding,
            label_x + text_w + padding, label_y + text_h + padding
        ]
        draw.rectangle(bg_rect, fill=(0, 0, 0, 200))
        draw.text((label_x, label_y), label_text, font=font, fill=arrow_color)
        occupancy_map[label_y:label_y+text_h, label_x:label_x+text_w] = 255
    img_rgb_array = np.array(pil_img)
    return cv2.cvtColor(img_rgb_array, cv2.COLOR_RGB2BGR)

# --------------------------------------------------------------------- #
# Image creation
# --------------------------------------------------------------------- #
def create_fluorescence_image(fluor_img: np.ndarray) -> np.ndarray:
    """
    Create enhanced RED fluorescence image with gamma correction.
    """
    img_float = fluor_img.astype(np.float32)
    
    if fluor_img.dtype == np.uint16:
        img_max = img_float.max()
        if img_max > 0:
            img_float /= img_max
    else:
        img_float /= 255.0
    
    img_gamma = np.power(img_float, PARAMS['fluor_gamma'])
    img_enhanced = img_gamma * PARAMS['fluor_brightness']
    img_enhanced = np.clip(img_enhanced, 0, 1)
    
    img_8bit = (img_enhanced * 255).astype(np.uint8)
    
    # RED channel (RGB format)
    red_img = np.zeros((img_8bit.shape[0], img_8bit.shape[1], 3), dtype=np.uint8)
    red_img[:, :, 0] = img_8bit  # Red channel
    
    return red_img

def create_overlay_image(bf_img: np.ndarray, fluor_img: Optional[np.ndarray], contours: List[np.ndarray]) -> np.ndarray:
    """
    Create overlay image with labels, matching interactive viewer.
    """
    # Start with original brightfield
    overlay = cv2.cvtColor(bf_img, cv2.COLOR_GRAY2RGB)
    
    if fluor_img is not None:
        fluor_float = fluor_img.astype(np.float32)
        
        if fluor_img.dtype == np.uint16:
            fluor_max = fluor_float.max()
            if fluor_max > 0:
                fluor_float = fluor_float / fluor_max
        else:
            fluor_float = fluor_float / 255.0
        
        fluor_processed = np.power(fluor_float, PARAMS['fluor_gamma']) * PARAMS['fluor_brightness']
        fluor_processed = np.clip(fluor_processed, 0, 1)
        fluor_uint8 = (fluor_processed * 255).astype(np.uint8)
        
        # Add to RED channel (RGB format)
        red_channel = overlay[:, :, 0].astype(np.float32)
        red_channel = np.clip(red_channel + fluor_uint8.astype(np.float32), 0, 255)
        overlay[:, :, 0] = red_channel.astype(np.uint8)
    
    # Draw contours in yellow
    cv2.drawContours(overlay, contours, -1, (255, 255, 0), 2)
    
    # Convert to BGR for label drawing
    overlay_bgr = cv2.cvtColor(overlay, cv2.COLOR_RGB2BGR)
    
    # Draw labels with smart positioning (matching interactive viewer)
    overlay_bgr = draw_labels_on_contours(overlay_bgr, contours)
    
    # Convert back to RGB
    overlay = cv2.cvtColor(overlay_bgr, cv2.COLOR_BGR2RGB)
    
    return overlay

def save_temp_image(img: np.ndarray, prefix: str) -> Path:
    """
    Save image to temporary file for Excel embedding.
    """
    temp_path = Path(tempfile.gettempdir()) / f"{prefix}_{os.urandom(4).hex()}.png"
    img_bgr = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
    cv2.imwrite(str(temp_path), img_bgr)
    return temp_path

# --------------------------------------------------------------------- #
# Excel export - LANDSCAPE FORMAT WITH COMPACT LAYOUT
# --------------------------------------------------------------------- #
def export_to_excel(pairs: List[Tuple[Path, Path, str, str]]):
    """
    Export all image pairs to a single Excel workbook.
    LANDSCAPE format with larger, more compact images.
    """
    wb = Workbook()
    
    if wb.active:
        wb.remove(wb.active)

    temp_files = []

    print(f"\n{'='*70}")
    print(f"Starting Excel export (LANDSCAPE format)...")
    print(f"{'='*70}")

    for bf_path, fluor_path, idx, rel_path in pairs:
        print(f"\n[{idx}/{len(pairs):02d}] Processing: {rel_path}")

        # Load images
        bf_img = cv2.imread(str(bf_path), cv2.IMREAD_UNCHANGED)
        fluor_img = cv2.imread(str(fluor_path), cv2.IMREAD_UNCHANGED)
        
        if bf_img is None:
            print(f"  ✗ Error: Cannot read {bf_path}")
            continue
        
        if len(bf_img.shape) == 3:
            bf_img = cv2.cvtColor(bf_img, cv2.COLOR_BGR2GRAY)
        if fluor_img is not None and len(fluor_img.shape) == 3:
            fluor_img = cv2.cvtColor(fluor_img, cv2.COLOR_BGR2GRAY)

        # Segment bacteria
        print(f"  → Segmenting bacteria...")
        _, _, _, all_contours = segment_bacteria(bf_img)
        print(f"  → Found {len(all_contours)} initial bacteria")

        # Calculate statistics with filter and sort
        print(f"  → Calculating statistics with filter/sort...")
        stats_df, summary, contours = get_processed_stats_and_contours(all_contours, bf_img, fluor_img, PARAMS['min_fluor_per_area'])
        print(f"  → After filter/sort: {len(contours)} bacteria")

        # Create visualization images
        print(f"  → Creating visualizations with smart labels...")
        bf_rgb = cv2.cvtColor(bf_img, cv2.COLOR_GRAY2RGB)
        fluor_rgb = create_fluorescence_image(fluor_img) if fluor_img is not None else np.zeros_like(bf_rgb)
        overlay_rgb = create_overlay_image(bf_img, fluor_img, contours)

        # Save to temporary files
        bf_temp = save_temp_image(bf_rgb, f"bf_{idx}")
        fluor_temp = save_temp_image(fluor_rgb, f"fluor_{idx}")
        overlay_temp = save_temp_image(overlay_rgb, f"overlay_{idx}")
        
        temp_files.extend([bf_temp, fluor_temp, overlay_temp])

        # Create worksheet
        print(f"  → Creating Excel worksheet (landscape layout)...")
        source_folder_name = bf_path.parent.name
        sheet_title = f"{source_folder_name}-{idx}"
        ws = wb.create_sheet(title=sheet_title)

        # === COMPACT LANDSCAPE LAYOUT ===
        # Images side by side at top, statistics below
        
        # Image row
        img_row = 2
        img_paths = [bf_temp, fluor_temp, overlay_temp]
        titles = ["Brightfield (BF)", "Fluorescence (Red)", "BF + Fluor Overlay (Labeled)"]
        
        for col_idx, (path, title) in enumerate(zip(img_paths, titles), 0):
            # Column positions
            if col_idx == 0:
                col = 1
            elif col_idx == 1:
                col = 7
            elif col_idx == 2:
                col = 15

            # Title
            title_cell = ws.cell(row=img_row, column=col, value=title)
            title_cell.font = Font(bold=True, size=12)
            title_cell.alignment = Alignment(horizontal='left')
            #ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row, end_column=col+8)
            
            # Image (LARGER SIZE)
            img = XLImage(str(path))
            img.width, img.height = TARGET_IMAGE_SIZE
            ws.add_image(img, anchor=f"{get_column_letter(int(col))}{img_row + 1}")

        # === Statistics Section (below images) ===
        stats_row = img_row + 28  # After larger images
        
        # File path header
        header_cell = ws.cell(row=stats_row, column=1, value=f"📊 Detection Statistics - {rel_path}")
        header_cell.font = Font(bold=True, size=13, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=stats_row, start_column=1, end_row=stats_row, end_column=5)
        ws.row_dimensions[stats_row].height = 25

        # Summary statistics (compact, side by side)
        summary_row = stats_row + 2
        
        # Left column summaries
        for i, (k, v) in enumerate(list(summary.items())[:3]):
            key_cell = ws.cell(row=summary_row + i, column=1, value=k)
            key_cell.font = Font(bold=True, size=10)
            val_cell = ws.cell(row=summary_row + i, column=2, value=v)
            val_cell.alignment = Alignment(horizontal='right')
            val_cell.font = Font(size=10)
        
        # Right column summaries
        remaining = list(summary.items())[3:]
        for i, (k, v) in enumerate(remaining):
            key_cell = ws.cell(row=summary_row + i, column=4, value=k)
            key_cell.font = Font(bold=True, size=10)
            val_cell = ws.cell(row=summary_row + i, column=5, value=v)
            val_cell.alignment = Alignment(horizontal='right')
            val_cell.font = Font(size=10)

        # Detailed bacteria table
        detail_row = summary_row + 5
        
        if not stats_df.empty:
            # Column headers
            for c_idx, col in enumerate(stats_df.columns, 1):
                header_cell = ws.cell(row=detail_row, column=c_idx, value=col)
                header_cell.font = Font(bold=True, size=10)
                header_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                header_cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for r_idx, row in enumerate(stats_df.itertuples(), detail_row + 1):
                for c_idx, val in enumerate(row[1:], 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.alignment = Alignment(horizontal='center' if c_idx == 1 else 'right')
                    cell.font = Font(size=9)
            
            # Borders
            thin = Side(border_style="thin", color="000000")
            for row in ws[f"A{detail_row}:E{detail_row + len(stats_df)}"]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        else:
            ws.cell(row=detail_row + 1, column=1, value="No bacteria detected")

        # === Column Formatting ===
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 14

        # === LANDSCAPE Page Setup ===
        ws.page_setup.orientation = 'landscape'  # LANDSCAPE
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = 1
        ws.page_setup.fitToWidth = 1
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.3
        ws.page_margins.bottom = 0.3

        # Headers and footers
        ws.oddHeader.center.text = f"{Path(rel_path).stem} - Bacteria Analysis"
        ws.oddHeader.center.size = 12
        ws.oddFooter.left.text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws.oddFooter.right.text = "&P / &N"

        print(f"  ✓ Worksheet '{sheet_title}' created (LANDSCAPE, large images)")

    # Save workbook
    print(f"\n{'='*70}")
    print(f"Saving Excel file...")
    wb.save(OUTPUT_FILE)
    
    # Clean up temporary files
    print(f"Cleaning up temporary files...")
    for temp_file in temp_files:
        try:
            temp_file.unlink()
        except Exception as e:
            print(f"  Warning: Could not delete {temp_file}: {e}")
    
    file_size_kb = OUTPUT_FILE.stat().st_size / 1024
    print(f"✓ Export complete: {OUTPUT_FILE}")
    print(f"  → File size: {file_size_kb:.1f} KB")
    print(f"  → Total worksheets: {len(wb.sheetnames)}")
    print(f"  → Format: LANDSCAPE A4 with large images")
    print(f"{'='*70}\n")

# --------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------- #
if __name__ == "__main__":
    print(f"\n{'='*70}")
    print(f"Bacteria Segmentation Batch Processor (Enhanced)")
    print(f"{'='*70}\n")
    
    pairs = find_image_pairs(INPUT_DIR)
    
    if pairs:
        export_to_excel(pairs)
    else:
        print(f"\n{'='*70}")
        print(f"✗ No valid image pairs found!")
        print(f"{'='*70}")
        print(f"\nPlease ensure:")
        print(f"  1. Images are in: {INPUT_DIR.absolute()}")
        print(f"  2. Files follow naming convention: *_ch00.tif + *_ch01.tif")
        print(f"  3. No AppleDouble files (._*) or hidden files (.*)") 
        print(f"\nDirectory structure example:")
        print(f"  source/")
        print(f"    ├── experiment1/")
        print(f"    │   ├── sample_01_ch00.tif")
        print(f"    │   └── sample_01_ch01.tif")
        print(f"    └── experiment2/")
        print(f"        ├── sample_02_ch00.tif")
        print(f"        └── sample_02_ch01.tif")
        print(f"{'='*70}\n")