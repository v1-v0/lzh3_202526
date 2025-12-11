# modules/reporter.py
"""
Excel report generation module
"""

import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

class ExcelReporter:
    """Handles Excel report generation with plots and statistics"""
    
    def __init__(self, config):
        self.config = config
    
    def generate_report(self, object_data, objects_excluded, pixel_size, unit, bit_depth):
        """Generate complete Excel report"""
        excel_file = os.path.join(self.config.DEBUG_DIR, "fluorescence_statistics.xlsx")
        wb = Workbook()
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Determine unit label
        excel_unit = 'um' if unit in ['um', 'µm', 'μm'] else unit
        has_calibration = pixel_size is not None
        
        # Generate plots
        print("  Creating plots...")
        errorbar_plot = self._create_error_bar_plot(object_data, excel_unit)
        barchart_plot = self._create_bar_chart(object_data, excel_unit)
        statistics_plot = self._create_statistics_plot(object_data, excel_unit)
        
        # Create sheets
        print("  Creating Excel sheets...")
        self._create_primary_sheet(wb, object_data, objects_excluded, excel_unit, 
                                   bit_depth, has_calibration, pixel_size,
                                   errorbar_plot, barchart_plot, statistics_plot)
        
        self._create_8bit_sheet(wb, object_data, excel_unit, bit_depth, has_calibration)
        
        self._create_enhanced_sheet(wb, object_data, excel_unit, has_calibration)
        
        self._create_comparison_sheet(wb, object_data, objects_excluded, excel_unit, 
                                      bit_depth, has_calibration, pixel_size)
        
        # Save
        wb.save(excel_file)
        print(f"  ✓ Excel file saved: {excel_file}")
    
    def _create_error_bar_plot(self, object_data, unit_str):
        """Create line plot with error bars"""
        particle_ids = [obj['object_id'] for obj in object_data]
        intensities = [obj['intensity_per_area_orig'] for obj in object_data]
        errors = [intensity * self.config.ERROR_PERCENTAGE for intensity in intensities]
        
        plt.figure(figsize=(12, 6))
        plt.errorbar(particle_ids, intensities, yerr=errors, 
                    fmt='o-', capsize=5, capthick=2, 
                    ecolor='red', markersize=6, color='blue',
                    linewidth=1.5, label=f'Intensity ± {int(self.config.ERROR_PERCENTAGE*100)}% error')
        
        plt.xlabel('Particle ID', fontsize=12, fontweight='bold')
        plt.ylabel(f'Total Intensity per {unit_str}²', fontsize=12, fontweight='bold')
        plt.title('Fluorescence Intensity by Particle with Error Bars', 
                 fontsize=14, fontweight='bold')
        plt.grid(True, alpha=0.3, linestyle='--')
        plt.legend(fontsize=10)
        plt.tight_layout()
        
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=self.config.PLOT_DPI, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        
        return img_buffer
    
    def _create_bar_chart(self, object_data, unit_str):
        """Create bar chart with error bars"""
        particle_ids = [obj['object_id'] for obj in object_data]
        intensities = [obj['intensity_per_area_orig'] for obj in object_data]
        errors = [intensity * self.config.ERROR_PERCENTAGE for intensity in intensities]
        
        plt.figure(figsize=(14, 6))
        bars = plt.bar(particle_ids, intensities, yerr=errors, 
                      capsize=5, color='skyblue', edgecolor='navy', linewidth=1.5,
                      error_kw={'elinewidth': 2, 'capthick': 2, 'ecolor': 'darkred'})
        
        colors = plt.cm.get_cmap('RdYlGn_r')(np.linspace(0.2, 0.8, len(bars)))
        for bar, color in zip(bars, colors):
            bar.set_color(color)
        
        plt.xlabel('Particle ID', fontsize=12, fontweight='bold')
        plt.ylabel(f'Total Intensity per {unit_str}²', fontsize=12, fontweight='bold')
        plt.title('Fluorescence Intensity Distribution with Error Bars', 
                 fontsize=14, fontweight='bold')
        plt.xticks(particle_ids)
        plt.grid(True, axis='y', alpha=0.3, linestyle='--')
        plt.tight_layout()
        
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=self.config.PLOT_DPI, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        
        return img_buffer
    
    def _create_statistics_plot(self, object_data, unit_str):
        """Create statistics summary plot"""
        intensities = np.array([obj['intensity_per_area_orig'] for obj in object_data])
        
        mean_intensity = np.mean(intensities)
        std_intensity = np.std(intensities, ddof=1)
        sem_intensity = std_intensity / np.sqrt(len(intensities))
        
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
        
        # Mean with SEM
        ax1.bar(['Mean Intensity'], [mean_intensity], 
               yerr=[sem_intensity], capsize=20, 
               color='lightcoral', edgecolor='darkred', linewidth=2,
               error_kw={'elinewidth': 3, 'capthick': 3, 'ecolor': 'black'})
        ax1.set_ylabel(f'Total Intensity per {unit_str}²', fontsize=11, fontweight='bold')
        ax1.set_title(f'Average Fluorescence Intensity\n(n={len(intensities)} particles)', 
                     fontsize=12, fontweight='bold')
        ax1.grid(True, axis='y', alpha=0.3)
        ax1.text(0, mean_intensity + sem_intensity + 200, 
                f'Mean: {mean_intensity:.2f}\nSEM: ±{sem_intensity:.2f}\nSD: ±{std_intensity:.2f}',
                ha='center', fontsize=9, 
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
        
        # Histogram
        ax2.hist(intensities, bins=15, color='steelblue', edgecolor='black', alpha=0.7)
        ax2.axvline(mean_intensity, color='red', linestyle='--', linewidth=2, label='Mean')
        ax2.axvline(mean_intensity + std_intensity, color='orange', linestyle=':', linewidth=2, label='Mean ± SD')
        ax2.axvline(mean_intensity - std_intensity, color='orange', linestyle=':', linewidth=2)
        ax2.set_xlabel(f'Total Intensity per {unit_str}²', fontsize=11, fontweight='bold')
        ax2.set_ylabel('Frequency', fontsize=11, fontweight='bold')
        ax2.set_title('Intensity Distribution', fontsize=12, fontweight='bold')
        ax2.legend()
        ax2.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=self.config.PLOT_DPI, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        
        return img_buffer
    
    def _create_primary_sheet(self, wb, object_data, objects_excluded, unit, bit_depth, 
                             has_calibration, pixel_size, errorbar_plot, barchart_plot, 
                             statistics_plot):
        """Create primary data sheet with embedded plots"""
        ws = wb.create_sheet(f"Primary Data ({bit_depth}-bit)")
        
        # Headers
        headers = [
            'Particle ID',
            f'Area ({unit}²)' if has_calibration else 'Area (pixels²)',
            f'Perimeter ({unit})' if has_calibration else 'Perimeter (pixels)',
            f'Total Intensity ({bit_depth}-bit)',
            f'Mean Intensity ({bit_depth}-bit)',
            f'Std Dev',
            f'Total Int. per {unit}²' if has_calibration else 'Total Int. per pixel²',
            f'Mean Int. per {unit}²' if has_calibration else 'Mean Int. per pixel²'
        ]
        ws.append(headers)
        self._style_header(ws, 1)
        
        # Data
        precision = 4 if has_calibration else 2
        for obj in object_data:
            ws.append([
                obj['object_id'],
                round(obj['area_physical'], precision),
                round(obj['perimeter_physical'], 2),
                round(obj['red_total_orig'], 1),
                round(obj['red_mean_orig'], 2),
                round(obj['red_std_orig'], 2),
                round(obj['intensity_per_area_orig'], 2),
                round(obj['mean_intensity_per_area_orig'], 2)
            ])
        
        # Summary statistics
        summary_row = len(object_data) + 3
        self._add_summary(ws, object_data, objects_excluded, unit, bit_depth, 
                         has_calibration, pixel_size, summary_row)
        
        # Embed plots
        img_errorbar = XLImage(errorbar_plot)
        img_errorbar.width = 900
        img_errorbar.height = 450
        ws.add_image(img_errorbar, 'K2')
        
        img_barchart = XLImage(barchart_plot)
        img_barchart.width = 1050
        img_barchart.height = 450
        ws.add_image(img_barchart, 'K28')
        
        img_statistics = XLImage(statistics_plot)
        img_statistics.width = 900
        img_statistics.height = 375
        ws.add_image(img_statistics, 'K54')
        
        self._auto_adjust_width(ws)
    
    def _create_8bit_sheet(self, wb, object_data, unit, bit_depth, has_calibration):
        """Create 8-bit converted data sheet"""
        ws = wb.create_sheet("8-bit Converted (Scaled)")
        
        headers = [
            'Particle ID',
            f'Area ({unit}²)' if has_calibration else 'Area (pixels²)',
            f'Perimeter ({unit})' if has_calibration else 'Perimeter (pixels)',
            'Total (8-bit)',
            'Mean (8-bit)',
            f'Total (scaled to {bit_depth}-bit)',
            f'Mean (scaled to {bit_depth}-bit)',
            f'Total Int. per {unit}²' if has_calibration else 'Total Int. per pixel²',
            'Recovery Accuracy (%)'
        ]
        ws.append(headers)
        self._style_header(ws, 1)
        
        precision = 4 if has_calibration else 2
        for obj in object_data:
            accuracy = (obj['red_total_8bit_scaled'] / obj['red_total_orig'] * 100) if obj['red_total_orig'] > 0 else 0
            ws.append([
                obj['object_id'],
                round(obj['area_physical'], precision),
                round(obj['perimeter_physical'], 2),
                round(obj['red_total_8bit'], 1),
                round(obj['red_mean_8bit'], 2),
                round(obj['red_total_8bit_scaled'], 1),
                round(obj['red_mean_8bit_scaled'], 2),
                round(obj['intensity_per_area_8bit_scaled'], 2),
                round(accuracy, 1)
            ])
        
        self._auto_adjust_width(ws)
    
    def _create_enhanced_sheet(self, wb, object_data, unit, has_calibration):
        """Create enhanced 8-bit data sheet"""
        ws = wb.create_sheet("Enhanced 8-bit (Visual)")
        
        headers = [
            'Particle ID',
            f'Area ({unit}²)' if has_calibration else 'Area (pixels²)',
            f'Perimeter ({unit})' if has_calibration else 'Perimeter (pixels)',
            'Total (enhanced)',
            'Mean (enhanced)',
            f'Total Int. per {unit}²' if has_calibration else 'Total Int. per pixel²'
        ]
        ws.append(headers)
        self._style_header(ws, 1)
        
        precision = 4 if has_calibration else 2
        for obj in object_data:
            ws.append([
                obj['object_id'],
                round(obj['area_physical'], precision),
                round(obj['perimeter_physical'], 2),
                round(obj['red_total_enh'], 1),
                round(obj['red_mean_enh'], 2),
                round(obj['intensity_per_area_enh'], 2)
            ])
        
        self._auto_adjust_width(ws)
    
    def _create_comparison_sheet(self, wb, object_data, objects_excluded, unit, 
                                 bit_depth, has_calibration, pixel_size):
        """Create comparison analysis sheet"""
        ws = wb.create_sheet("Comparison Analysis")
        
        ws.append(['', '', '', f'{bit_depth}-BIT ORIGINAL', '', '8-BIT SCALED', '', 'ENHANCED 8-BIT'])
        ws.merge_cells('D1:E1')
        ws.merge_cells('F1:G1')
        ws.merge_cells('H1:I1')
        self._style_header(ws, 1, 'ED7D31')
        
        headers = [
            'Particle ID',
            f'Area ({unit}²)' if has_calibration else 'Area (pixels²)',
            f'Perimeter ({unit})' if has_calibration else 'Perimeter (pixels)',
            'Total', 'Mean',
            'Total (scaled)', 'Recovery %',
            'Total (enhanced)', 'Visual Ratio'
        ]
        ws.append(headers)
        self._style_header(ws, 2)
        
        precision = 4 if has_calibration else 2
        for obj in object_data:
            recovery = (obj['red_total_8bit_scaled'] / obj['red_total_orig'] * 100) if obj['red_total_orig'] > 0 else 0
            visual_ratio = obj['red_total_enh'] / obj['red_total_orig'] if obj['red_total_orig'] > 0 else 0
            
            ws.append([
                obj['object_id'],
                round(obj['area_physical'], precision),
                round(obj['perimeter_physical'], 2),
                round(obj['red_total_orig'], 1),
                round(obj['red_mean_orig'], 2),
                round(obj['red_total_8bit_scaled'], 1),
                f"{recovery:.1f}%",
                round(obj['red_total_enh'], 1),
                f"{visual_ratio:.3f}x"
            ])
        
        self._auto_adjust_width(ws)
    
    def _add_summary(self, ws, object_data, objects_excluded, unit, bit_depth, 
                    has_calibration, pixel_size, start_row):
        """Add summary statistics to sheet"""
        ws.append([])
        ws.append(['SUMMARY STATISTICS'])
        self._style_summary_header(ws.cell(start_row, 1))
        
        ws.append(['Total Particles:', len(object_data)])
        ws.append(['Particles Excluded (no fluorescence):', objects_excluded])
        ws.append([])
        
        avg_area = np.mean([obj['area_physical'] for obj in object_data])
        avg_perimeter = np.mean([obj['perimeter_physical'] for obj in object_data])
        
        ws.append(['Avg Area:', round(avg_area, 4 if has_calibration else 2), 
                  f'{unit}²' if has_calibration else 'pixels²'])
        ws.append(['Avg Perimeter:', round(avg_perimeter, 2), 
                  f'{unit}' if has_calibration else 'pixels'])
        
        ws.append([])
        ws.append(['Avg Total Intensity:', round(np.mean([obj['red_total_orig'] for obj in object_data]), 1)])
        ws.append(['Avg Mean Intensity:', round(np.mean([obj['red_mean_orig'] for obj in object_data]), 2)])
        ws.append(['Avg Intensity per Area:', 
                  round(np.mean([obj['intensity_per_area_orig'] for obj in object_data]), 2),
                  f'per {unit}²' if has_calibration else 'per pixel²'])
    
    def _style_header(self, ws, row_num, fill_color='4472C4'):
        """Style header row"""
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        font = Font(bold=True, color='FFFFFF', size=11)
        alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[row_num]:
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
    
    def _style_summary_header(self, cell):
        """Style summary section header"""
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    def _auto_adjust_width(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width