# Standard library imports
import atexit
import csv
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import textwrap
import time as pytime
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional, Tuple, cast
from contextlib import nullcontext
import configparser

# Third-party data science imports
import numpy as np
import pandas as pd
from scipy import stats as scipy_stats
from scipy.spatial import KDTree
from scipy.stats import skew
from tqdm import tqdm

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None

# Computer vision imports
import cv2

# Plotting imports
import matplotlib
matplotlib.use("Agg")

import matplotlib.pyplot as plt
matplotlib.rcParams["pdf.compression"] = 9
matplotlib.rcParams["savefig.dpi"] = 110


from matplotlib.axes import Axes as MplAxes
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
from matplotlib.lines import Line2D
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages
from mpl_toolkits.axes_grid1.inset_locator import inset_axes

# Excel/Office imports
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab imports (for individual sample report additions)
try:
    from reportlab.lib.pagesizes import A4 as RL_A4
    from reportlab.lib.units import cm as rl_cm, mm as rl_mm
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image as RLImage, PageBreak, HRFlowable, KeepTogether,
    )
    from reportlab.graphics.shapes import Drawing, Line as RLLine
    from reportlab.platypus.flowables import Flowable
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# PDF merging (optional – used to combine matplotlib + ReportLab pages)
try:
    from pypdf import PdfWriter as _PdfWriter
    PYPDF_AVAILABLE = True
except ImportError:
    try:
        from PyPDF2 import PdfMerger as _PdfMerger   # legacy fallback
        PYPDF_AVAILABLE = True
        _PdfWriter = None
    except ImportError:
        PYPDF_AVAILABLE = False
        _PdfWriter = None
        _PdfMerger = None

# Local imports
from bacteria_configs import SegmentationConfig

# debug
import math
import io
from PIL import Image as PILImage

# ==================================================
# Configuration & Globals
# ==================================================

def get_project_root() -> Path:
    if getattr(sys, 'frozen', False):
        # If running as a PyInstaller bundle, return the folder containing the .exe
        return Path(sys.executable).resolve().parent
    else:
        # If running as a normal Python script, return the script's folder
        return Path(__file__).resolve().parent

_cfg = configparser.ConfigParser()

# CORRECTED: Use get_project_root() instead of Path(__file__).parent
_cfg.read(get_project_root() / "config.ini")

DEBUG_MODE = (
    _cfg.get("pipeline", "mode", fallback="default").strip().lower() == "debug"
)
print(f"[CONFIG] Running in {'DEBUG' if DEBUG_MODE else 'DEFAULT'} mode")



PROJECT_ROOT = get_project_root()
OUTPUTS_DIR = PROJECT_ROOT / "outputs"
OUTPUTS_DIR.mkdir(exist_ok=True)
_logs_dir = PROJECT_ROOT / "logs"
_logs_dir.mkdir(exist_ok=True)
_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
_script_name = Path(sys.argv[0]).stem
_log_path = _logs_dir / f"run_{_timestamp}_{_script_name}.txt"
_log_file: Optional[Any] = None

class Tee:
    def __init__(self, *streams):
        self.streams = streams
    def write(self, data: str) -> None:
        for s in self.streams:
            s.write(data)
            s.flush()
    def flush(self) -> None:
        for s in self.streams:
            s.flush()

try:
    _log_file = open(_log_path, "w", encoding="utf-8")
    sys.stdout = Tee(sys.stdout, _log_file)
    sys.stderr = Tee(sys.stderr, _log_file)
except Exception as e:
    print(f"Warning: Could not set up logging: {e}")

print(f"Saving output to: {_log_path}")
print(f"Project root: {PROJECT_ROOT.resolve()}")
print(f"Running as: {'EXECUTABLE' if getattr(sys, 'frozen', False) else 'SCRIPT'}")

@atexit.register
def _close_log_file() -> None:
    global _log_file
    if _log_file is not None:
        try:
            _log_file.close()
        except Exception:
            pass
        finally:
            _log_file = None

logger = logging.getLogger("particle_scout")
if not logger.handlers:
    logger.setLevel(logging.INFO)
    _sh = logging.StreamHandler()
    _sh.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
    logger.addHandler(_sh)

# Constants
SOURCE_DIR = Path("./source")
CONTROL_DIR = None
IMAGE_GLOB = "*_ch00.tif"
OUTPUT_DIR: Optional[Path] = None

SCALE_BAR_LENGTH_UM = 10
SCALE_BAR_HEIGHT = 4
SCALE_BAR_MARGIN = 15
SCALE_BAR_COLOR = (255, 255, 255)
SCALE_BAR_BG_COLOR = (0, 0, 0)
SCALE_BAR_TEXT_COLOR = (255, 255, 255)
SCALE_BAR_FONT_SCALE = 0.5
SCALE_BAR_FONT_THICKNESS = 1

GAUSSIAN_SIGMA = 15
MORPH_KERNEL_SIZE = 3
MORPH_ITERATIONS = 1
DILATE_ITERATIONS = 1
ERODE_ITERATIONS = 1

MIN_AREA_UM2 = 3.0
MAX_AREA_UM2 = 2000.0
MIN_CIRCULARITY = 0.0
MAX_FRACTION_OF_IMAGE_AREA = 0.25

FLUOR_GAUSSIAN_SIGMA = 1.5
FLUOR_MORPH_KERNEL_SIZE = 3
FLUOR_MIN_AREA_UM2 = 3.0
FLUOR_MATCH_MIN_INTERSECTION_PX = 5.0

CLEAR_OUTPUT_DIR_EACH_RUN = True
SEPARATE_OUTPUT_BY_GROUP = True
FALLBACK_UM_PER_PX: float = 0.109492


# ==================================================
# ReportLab Report Constants
# ==================================================
PALETTE = {
    "group1_bar":     "#F4A7A7",
    "group2_bar":     "#B5B87F",
    "group3_bar":     "#5BBCB5",
    "control_bar":    "#B5A7CC",
    "scatter_dot":    "#00BFFF",
    "ctrl_mean_line": "#1F77B4",
    "threshold_line": "#D62728",
    "edge_default":   "#555555",
    "text_dark":      "#1A1A2E",
    "header_bg":      "#1A2A4A",
    "header_fg":      "#FFFFFF",
    "row_alt":        "#EFF3F8",
    "detected_green": "#28A745",
    "ns_grey":        "#666666",
}

BAR_COLORS = [
    PALETTE["group1_bar"],
    PALETTE["group2_bar"],
    PALETTE["group3_bar"],
    PALETTE["control_bar"],
]

SIG_COLORS = {
    "***": "#B22222",
    "**":  "#CC5500",
    "*":   "#CC8800",
    "ns":  PALETTE["ns_grey"],
}

# ReportLab page layout dimensions (points)
if REPORTLAB_AVAILABLE:
    _RL_PAGE_W, _RL_PAGE_H = RL_A4
    _RL_MARGIN   = 2.0 * rl_cm
    _RL_CONTENT_W = _RL_PAGE_W - 2 * _RL_MARGIN
else:
    _RL_PAGE_W = _RL_PAGE_H = _RL_MARGIN = _RL_CONTENT_W = 0.0

# Sentinel keys used by is_page5_empty()
_EMPTY_PAGE5_SENTINEL_KEYS = {
    "methodology", "quality_control", "limitations",
    "approval", "performed_by", "reviewed_by",
}


# ==================================================
# Helper Functions
# ==================================================


def _compress_pdf_with_pypdf(pdf_path: Path) -> bool:
    """
    Rewrites a PDF with compressed page content streams using pypdf, if available.
    """
    if not PYPDF_AVAILABLE or _PdfWriter is None or not pdf_path.exists():
        return False

    try:
        from pypdf import PdfReader, PdfWriter
        reader = PdfReader(str(pdf_path))
        writer = PdfWriter()

        for page in reader.pages:
            try:
                page.compress_content_streams()
            except Exception:
                pass
            writer.add_page(page)

        tmp_path = pdf_path.with_name(pdf_path.stem + "_compressed.pdf")
        with open(tmp_path, "wb") as fh:
            writer.write(fh)

        if tmp_path.exists() and tmp_path.stat().st_size <= pdf_path.stat().st_size:
            pdf_path.unlink(missing_ok=True)
            tmp_path.replace(pdf_path)
        else:
            tmp_path.unlink(missing_ok=True)

        return True
    except Exception:
        return False

def _bytesio_size_mb(buf: io.BytesIO) -> float:
    pos = buf.tell()
    try:
        return len(buf.getbuffer()) / (1024 * 1024)
    finally:
        buf.seek(pos)


def _file_size_mb(path: Path) -> float:
    try:
        return path.stat().st_size / (1024 * 1024)
    except Exception:
        return 0.0


def _compress_image_for_pdf(
    image_source: Any,
    *,
    max_width_px: int = 1400,
    max_height_px: int = 1400,
    jpeg_quality: int = 65,
    convert_to_jpeg: bool = True,
) -> Any:
    """
    Return a compressed in-memory image suitable for ReportLab RLImage.

    - Downscales large raster images.
    - Converts RGB/RGBA images to JPEG where practical.
    - Falls back to original path/source if Pillow is unavailable or processing fails.
    """
    if PILImage is None:
        return image_source

    try:
        if isinstance(image_source, (str, Path)):
            im = PILImage.open(str(image_source))
        else:
            im = PILImage.open(image_source)

        im.load()

        if im.mode in ("RGBA", "LA"):
            bg = PILImage.new("RGB", im.size, (255, 255, 255))
            alpha = im.getchannel("A") if "A" in im.getbands() else None
            bg.paste(im.convert("RGB"), mask=alpha)
            im = bg
        elif im.mode != "RGB":
            im = im.convert("RGB")

        im.thumbnail((max_width_px, max_height_px), PILImage.Resampling.LANCZOS)

        out = io.BytesIO()
        if convert_to_jpeg:
            im.save(
                out,
                format="JPEG",
                quality=jpeg_quality,
                optimize=True,
                progressive=True,
            )
        else:
            im.save(out, format="PNG", optimize=True)

        out.seek(0)
        return out
    except Exception:
        return image_source





def _fmt_elapsed(seconds: float) -> str:
    if seconds >= 60:
        return f"{int(seconds // 60)}m {int(seconds % 60)}s"
    return f"{seconds:.1f}s"

def logged_input(prompt: str) -> str:
    print(prompt, end='', flush=True)
    user_input = input()
    if user_input.strip():
        print(user_input)
    else:
        print("(pressed Enter)")
    return user_input

def _compute_group_vs_control_stats(group_values: pd.Series, control_values: pd.Series, threshold: float) -> dict:
    n = len(group_values)
    mean_val = float(group_values.mean())
    std_val = float(group_values.std(ddof=1)) if n > 1 else 0.0

    if n >= 2 and std_val > 0:
        t_crit = float(scipy_stats.t.ppf(0.975, df=n - 1))
        sem = std_val / np.sqrt(n)
        ci_lower = mean_val - t_crit * sem
        ci_upper = mean_val + t_crit * sem
    else:
        ci_lower = mean_val
        ci_upper = mean_val

    ctrl_mean = float(control_values.mean())
    ctrl_std = float(control_values.std(ddof=1)) if len(control_values) > 1 else 0.0
    pooled_std = np.sqrt(((n - 1) * std_val ** 2 + (len(control_values) - 1) * ctrl_std ** 2) / (n + len(control_values) - 2)) if (n + len(control_values) - 2) > 0 else 1.0
    cohens_d = (mean_val - ctrl_mean) / pooled_std if pooled_std > 0 else 0.0

    if n >= 2 and len(control_values) >= 2:
        try:
            t_raw, p_raw = scipy_stats.ttest_ind(group_values.values, control_values.values, equal_var=False)
            p_value = float(np.asarray(p_raw).item())
        except Exception:
            p_value = np.nan
    else:
        p_value = np.nan

    if pd.isna(p_value):
        sig_label = "N/A (n<2)"
    elif p_value < 0.001:
        sig_label = "***"
    elif p_value < 0.01:
        sig_label = "**"
    elif p_value < 0.05:
        sig_label = "*"
    else:
        sig_label = "ns"

    if n < 3:
        confidence = "Low"
    elif (ci_upper < threshold) or (ci_lower > threshold):
        if abs(cohens_d) >= 0.8:
            confidence = "High"
        else:
            confidence = "Moderate"
    else:
        confidence = "Low"

    return {
        'CI_Lower': round(ci_lower, 2),
        'CI_Upper': round(ci_upper, 2),
        'Cohens_d': round(cohens_d, 3),
        'P_Value': round(p_value, 6) if not pd.isna(p_value) else np.nan,
        'Significance': sig_label,
        'Classification_Confidence': confidence,
    }

def validate_config(config: dict) -> bool:
    always_required = ['source_dir', 'batch_mode', 'dataset_id', 'dataset_id_base', 'percentile', 'threshold_pct']
    for key in always_required:
        if key not in config:
            print(f"[ERROR] Missing required config key: '{key}'")
            return False

    source_dir = config['source_dir']
    if source_dir is None or not source_dir.exists():
        print(f"[ERROR] Source directory does not exist: {source_dir}")
        return False

    if config['batch_mode']:
        batch_required = ['source_dir_positive', 'source_dir_negative']
        for key in batch_required:
            if key not in config:
                print(f"[ERROR] Batch mode requires config key: '{key}'")
                return False
        for key, label in [('source_dir_positive', 'G+'), ('source_dir_negative', 'G-')]:
            path = config[key]
            if path is None or not path.exists():
                print(f"[ERROR] {label} source directory does not exist: {path}")
                return False

    percentile = config['percentile']
    if not (0.0 < percentile <= 0.5):
        print(f"[ERROR] percentile out of range: {percentile:.4f} (expected 0 < p ≤ 0.5)")
        return False

    threshold_pct = config['threshold_pct']
    if not (0.0 < threshold_pct <= 0.5):
        print(f"[ERROR] threshold_pct out of range: {threshold_pct:.4f} (expected 0 < t ≤ 0.5)")
        return False

    return True

def setup_output_directory(config: Dict) -> Path:
    timestamp  = config.get('timestamp', datetime.now().strftime("%Y%m%d_%H%M%S"))
    dataset_id = config['dataset_id']
    folder_name = f"{timestamp}_{dataset_id}"

    if config.get('batch_mode', False):
        output_root = OUTPUTS_DIR / folder_name
        output_root.mkdir(parents=True, exist_ok=True)
        positive_dir = output_root / "Positive"
        negative_dir = output_root / "Negative"
        positive_dir.mkdir(exist_ok=True)
        negative_dir.mkdir(exist_ok=True)
        config['positive_output'] = positive_dir
        config['negative_output'] = negative_dir
        print(f"📁 Output directory: {output_root}\n   ├── Positive/\n   └── Negative/\n")
        return output_root
    else:
        output_dir = OUTPUTS_DIR / folder_name
        output_dir.mkdir(parents=True, exist_ok=True)
        print(f"📁 Output directory: {output_dir}\n")
        return output_dir

def load_bacteria_config_from_json(bacteria_key: Optional[str]) -> Optional['SegmentationConfig']:
    if not bacteria_key:
        return None
    config_file = Path("bacteria_configs") / f"{bacteria_key}.json"
    if not config_file.exists():
        print(f"[WARN] Config file not found: {config_file}")
        return None
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        config_data = json_data.get("config", json_data)

        def safe_float(value, default):
            if value is None: return default
            try: return float(value)
            except (TypeError, ValueError): return default

        def safe_int(value, default):
            if value is None: return default
            try: return int(value)
            except (TypeError, ValueError): return default

        def safe_bool(value, default):
            if value is None: return default
            return bool(value)

        config = SegmentationConfig(
            name=config_data.get('name', 'Unknown'),
            description=config_data.get('description', ''),
            gaussian_sigma=safe_float(config_data.get('gaussian_sigma'), 15.0),
            min_area_um2=safe_float(config_data.get('min_area_um2'), 3.0),
            max_area_um2=safe_float(config_data.get('max_area_um2'), 2000.0),
            dilate_iterations=safe_int(config_data.get('dilate_iterations'), 0),
            erode_iterations=safe_int(config_data.get('erode_iterations'), 0),
            morph_kernel_size=safe_int(config_data.get('morph_kernel_size'), 3),
            morph_iterations=safe_int(config_data.get('morph_iterations'), 1),
            min_circularity=safe_float(config_data.get('min_circularity'), 0.0),
            max_circularity=safe_float(config_data.get('max_circularity'), 1.0),
            min_aspect_ratio=safe_float(config_data.get('min_aspect_ratio'), 0.2),
            max_aspect_ratio=safe_float(config_data.get('max_aspect_ratio'), 10.0),
            min_mean_intensity_bf=safe_float(config_data.get('min_mean_intensity_bf', config_data.get('min_mean_intensity', 0.0)), 0.0),
            max_mean_intensity_bf=safe_float(config_data.get('max_mean_intensity_bf', config_data.get('max_mean_intensity', 255.0)), 255.0),
            object_mean_intensity_bf=config_data.get('object_mean_intensity_bf', None),
            min_solidity=safe_float(config_data.get('min_solidity'), 0.3),
            max_fraction_of_image=safe_float(config_data.get('max_fraction_of_image'), 0.25),
            fluor_min_area_um2=safe_float(config_data.get('fluor_min_area_um2'), 3.0),
            fluor_max_area_um2=safe_float(config_data.get('fluor_max_area_um2'), 2000.0),
            fluor_match_min_intersection_px=safe_float(config_data.get('fluor_match_min_intersection_px'), 5.0),
            invert_image=safe_bool(config_data.get('invert_image'), False),
            use_intensity_threshold=safe_bool(config_data.get('use_intensity_threshold'), False),
            intensity_threshold=safe_float(config_data.get('intensity_threshold'), 80.0),
            max_edge_gradient=safe_float(config_data.get('max_edge_gradient'), 999.0),
            pixel_size_um=safe_float(config_data.get('pixel_size_um'), 0.109492),
            last_modified=config_data.get('last_modified'),
            tuned_by=config_data.get('tuned_by')
        )
        return config
    except Exception as e:
        print(f"[ERROR] Failed to load config from {config_file}: {e}")
        return None

def _load_multi_scan_whitelist() -> list[str]:
    try:
        from bacteria_registry import registry as _reg
        wl = [k for k in _reg.get_whitelist() if k != 'default']
        if wl: return wl
    except Exception:
        pass
    return ['klebsiella_pneumoniae', 'proteus_mirabilis']

MULTI_SCAN_WHITELIST: list[str] = _load_multi_scan_whitelist()

def select_bacteria_config() -> dict:
    return select_bacteria_configuration()

def select_bacteria_configuration() -> dict:
    print("\n" + "=" * 80 + "\nBACTERIA CONFIGURATION SELECTION\n" + "=" * 80)
    config_dir = Path("bacteria_configs")
    if not config_dir.exists():
        return {'mode': 'single', 'bacteria_type': 'default', 'selected_config': None, 'configs_to_scan': [], 'config_names': {}, 'configs': {}}

    json_files = [f for f in config_dir.glob("*.json") if f.stem != 'registry']
    available_configs, config_names, _seen_names, _seen_modified = [], {}, {}, {}

    for json_file in sorted(json_files):
        bacteria_key = json_file.stem
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            config_data = data.get("config", data)
            name = config_data.get("name", bacteria_key)
            last_modified = config_data.get("last_modified", "")
        except Exception:
            name, last_modified = bacteria_key, ""

        if name in _seen_names:
            if last_modified > _seen_modified.get(name, ""):
                idx = available_configs.index(_seen_names[name])
                available_configs[idx] = bacteria_key
                config_names.pop(_seen_names[name])
                config_names[bacteria_key] = name
                _seen_names[name] = bacteria_key
                _seen_modified[name] = last_modified
            continue

        _seen_names[name] = bacteria_key
        _seen_modified[name] = last_modified
        available_configs.append(bacteria_key)
        config_names[bacteria_key] = name

    if not DEBUG_MODE:
        configs_to_scan = [k for k in available_configs if k in MULTI_SCAN_WHITELIST]
        if not configs_to_scan:
            raise SystemExit("Cannot run in DEFAULT mode without validated configurations. Run tuner.py first.")
        configs_dict = {}
        for k in configs_to_scan:
            cfg = load_bacteria_config_from_json(k)
            if cfg: configs_dict[k] = cfg
        return {'mode': 'multi_scan', 'bacteria_type': None, 'selected_config': None, 'configs_to_scan': configs_to_scan, 'config_names': config_names, 'configs': configs_dict}

    print("\nSelect processing mode:\n  [1] UNKNOWN SAMPLE (Multi-scan)\n  [2] KNOWN BACTERIA (Single config)\n")
    while True:
        choice = logged_input("Select mode [1-2]: ").strip()
        if choice in ['q', 'quit', 'exit']: raise SystemExit(0)
        if choice == "1":
            configs_to_scan = [k for k in available_configs if k in MULTI_SCAN_WHITELIST]
            configs_dict = {}
            for k in configs_to_scan:
                cfg = load_bacteria_config_from_json(k)
                if cfg is not None:
                    configs_dict[k] = cfg
            return {'mode': 'multi_scan', 'bacteria_type': None, 'selected_config': None, 'configs_to_scan': configs_to_scan, 'config_names': config_names, 'configs': configs_dict}
        elif choice == "2":
            for i, k in enumerate(available_configs, 1):
                print(f"  [{i}] {config_names[k]}")
            while True:
                b_choice = logged_input("Select bacteria configuration: ").strip()
                if b_choice.isdigit() and 1 <= int(b_choice) <= len(available_configs):
                    selected_key = available_configs[int(b_choice) - 1]
                    break
                elif b_choice in available_configs:
                    selected_key = b_choice
                    break
            cfg = load_bacteria_config_from_json(selected_key)
            return {'mode': 'single', 'bacteria_type': selected_key, 'selected_config': cfg, 'configs_to_scan': [selected_key], 'config_names': config_names, 'configs': {selected_key: cfg}}
        else:
            print("Invalid choice. Please enter 1 or 2.")

def select_source_directory(max_depth=2) -> Optional[Path]:
    root_dir = Path('source')
    if not root_dir.exists(): return None
    valid_dirs = []
    for item in root_dir.iterdir():
        if not item.is_dir(): continue
        if (item / 'G+').is_dir() and (item / 'G-').is_dir():
            valid_dirs.append(item.name)
            continue
        try:
            if any(d.name.lower().startswith('control') for d in item.iterdir() if d.is_dir()):
                valid_dirs.append(item.name)
        except OSError: pass
    if not valid_dirs: return None
    valid_dirs.sort()
    for i, d in enumerate(valid_dirs, 1):
        print(f"  [{i}] {d}")
    while True:
        sel = logged_input("\nEnter number or name: ").strip()
        if sel.lower() in {'q', 'quit', 'exit'}: raise SystemExit(0)
        if sel.isdigit() and 1 <= int(sel) <= len(valid_dirs): return root_dir / valid_dirs[int(sel) - 1]
        if sel in valid_dirs: return root_dir / sel

def configure_dataset() -> dict:
    config = {}
    source_dir = select_source_directory()
    if source_dir is None: raise SystemExit("No source directory selected")
    config['source_dir'] = source_dir
    gplus_path, gminus_path = source_dir / 'G+', source_dir / 'G-'
    if gplus_path.is_dir() and gminus_path.is_dir():
        config['batch_mode'], config['source_dir_positive'], config['source_dir_negative'] = True, gplus_path, gminus_path
    else:
        config['batch_mode'] = False

    default_id = source_dir.name
    dataset_id = logged_input(f"Dataset label (Enter for '{default_id}'): ").strip()
    config['dataset_id'] = dataset_id if dataset_id else default_id
    config['dataset_id_base'] = config['dataset_id']

    if not DEBUG_MODE:
        config['percentile'], config['threshold_pct'] = 0.30, 0.05
    else:
        pct = logged_input("Enter percentile (% or Enter for 30%): ").strip()
        config['percentile'] = float(pct)/100.0 if pct else 0.30
        thr = logged_input("Enter threshold (% or Enter for 5%): ").strip()
        config['threshold_pct'] = float(thr)/100.0 if thr else 0.05
    return config

def display_configuration_summary(config: dict) -> None:
    print("\n" + "=" * 80 + "\nCONFIGURATION SUMMARY\n" + "=" * 80)
    if not DEBUG_MODE: return
    if logged_input("Proceed? (y/n): ").strip().lower() not in ("", "y", "yes"):
        raise SystemExit("Cancelled.")

def validate_batch_structure(config: dict) -> bool:
    if not config.get('batch_mode'): return True
    issues = []
    for k in ['source_dir_positive', 'source_dir_negative']:
        if k not in config or not config[k].exists() or not list(config[k].rglob("*_ch00.tif")):
            issues.append(f"Missing or empty {k}")
    return len(issues) == 0


# ==================================================
# Image Processing & Math Helpers
# ==================================================

def safe_imread(path: Path, flags: int = cv2.IMREAD_UNCHANGED) -> Optional[np.ndarray]:
    try:
        with open(path, 'rb') as f:
            file_bytes = np.asarray(bytearray(f.read()), dtype=np.uint8)
        return cv2.imdecode(file_bytes, flags)
    except Exception: return None

def safe_imwrite(path: Path, img: np.ndarray, params: Optional[list] = None) -> bool:
    try:
        ext = path.suffix.lower() or '.png'
        is_success, buffer = cv2.imencode(ext, img, params) if params else cv2.imencode(ext, img)
        if is_success:
            with open(path, 'wb') as f: f.write(buffer.tobytes())
            return True
    except Exception: pass
    return False

def safe_xml_parse(xml_path: Path) -> Optional[ET.ElementTree]:
    try:
        with open(xml_path, 'r', encoding='utf-8') as f:
            return ET.ElementTree(ET.fromstring(f.read()))
    except Exception: return None

def validate_path_encoding(path: Path) -> bool:
    try:
        str(path.resolve()).encode(sys.getfilesystemencoding())
        return True
    except Exception: return False

def add_scale_bar(img: np.ndarray, pixel_size: float, unit: str = "um", length_um: float = 10) -> np.ndarray:
    if pixel_size <= 0: return img
    bar_length_px = int(round(length_um / pixel_size))
    if bar_length_px < 10: return img
    h, w = img.shape[:2]
    bar_x, bar_y = w - bar_length_px - SCALE_BAR_MARGIN, h - SCALE_BAR_HEIGHT - SCALE_BAR_MARGIN
    if bar_x < 0 or bar_y < 0: return img
    label = f"{int(length_um)} {unit}"
    font = cv2.FONT_HERSHEY_SIMPLEX
    (text_w, text_h), _ = cv2.getTextSize(label, font, SCALE_BAR_FONT_SCALE, SCALE_BAR_FONT_THICKNESS)
    text_x, text_y = bar_x + (bar_length_px - text_w) // 2, bar_y - 8
    img = img.copy()
    cv2.rectangle(img, (bar_x, bar_y), (bar_x + bar_length_px, bar_y + SCALE_BAR_HEIGHT), SCALE_BAR_COLOR, -1)
    cv2.putText(img, label, (text_x, text_y), font, SCALE_BAR_FONT_SCALE, SCALE_BAR_TEXT_COLOR, SCALE_BAR_FONT_THICKNESS, cv2.LINE_AA)
    return img

def ensure_dir(p: Path) -> None: p.mkdir(parents=True, exist_ok=True)

def save_debug(folder: Path, name: str, img: np.ndarray, pixel_size_um: Optional[float] = None, force: bool = False) -> None:
    if not DEBUG_MODE and not force: return
    out = folder / name
    img_to_save = add_scale_bar(img.copy(), float(pixel_size_um), "um", SCALE_BAR_LENGTH_UM) if pixel_size_um and pixel_size_um > 0 else img
    safe_imwrite(out, img_to_save)

def _display_group_name(name: str) -> str:
    return "Control" if name.lower().startswith("control") else name

def _group_order_key(g: str) -> tuple[int, int]:
    if g == "Control": return (1, 10**9)
    if g.isdigit(): return (0, int(g))
    return (0, 10**8)

def find_metadata_paths(img_path: Path) -> tuple[Optional[Path], Optional[Path]]:
    base = img_path.stem[:-5] if img_path.stem.endswith("_ch00") else img_path.stem
    md_dir = img_path.parent / "MetaData"
    xml_main, xml_props = md_dir / f"{base}.xml", md_dir / f"{base}_Properties.xml"
    return (xml_props if xml_props.exists() else None, xml_main if xml_main.exists() else None)

def _require_attr(elem: ET.Element, attr: str, context: str) -> str:
    v = elem.get(attr)
    if v is None: raise ValueError(f"Missing attribute '{attr}' in {context}")
    return v

def _parse_float(s: str) -> float: return float(s.strip().replace(",", "."))

def get_pixel_size_um(xml_props_path: Optional[Path], xml_main_path: Optional[Path]) -> Tuple[float, float]:
    for xml_path in filter(None, [xml_props_path, xml_main_path]):
        try:
            tree = safe_xml_parse(xml_path)
            if tree is None:
                continue
            root = tree.getroot()
            if root is None:
                continue
            dim_elems = root.findall(".//ImageDescription/Dimensions/DimensionDescription")
            if not dim_elems:
                continue
            dims = {str(d.get("DimID")): d for d in dim_elems}
            if "X" not in dims or "Y" not in dims:
                continue
            def read_dim(dim_id: str) -> tuple[float, int, str]:
                d = dims[dim_id]
                return (
                    _parse_float(str(d.get("Length", "0"))),
                    int(str(d.get("NumberOfElements", "1"))),
                    str(d.get("Unit", "")),
                )
            x_len, x_n, _ = read_dim("X")
            y_len, y_n, _ = read_dim("Y")
            return float(x_len / max(x_n, 1)), float(y_len / max(y_n, 1))
        except Exception:
            continue
    raise ValueError("Could not determine pixel size.")

def contour_perimeter_um(contour: np.ndarray, um_per_px_x: float, um_per_px_y: float) -> float:
    pts = contour.reshape(-1, 2).astype(np.float64)
    pts[:, 0] *= float(um_per_px_x)
    pts[:, 1] *= float(um_per_px_y)
    d = np.diff(np.vstack([pts, pts[0]]), axis=0)
    return float(np.sqrt((d[:, 0] ** 2) + (d[:, 1] ** 2)).sum())

def equivalent_diameter_from_area(area: float) -> float:
    return float(2.0 * np.sqrt(area / np.pi)) if area > 0 else 0.0

def normalize_to_8bit(img: np.ndarray) -> np.ndarray:
    if img.dtype == np.uint8: return img
    out = np.zeros_like(img, dtype=np.uint8)
    cv2.normalize(img, out, 0.0, 255.0, cv2.NORM_MINMAX)
    return out

def _put_text_outline(img: np.ndarray, text: str, org: tuple[int, int], font_scale: float = 0.5, color: tuple[int, int, int] = (255, 255, 255), thickness: int = 1) -> None:
    cv2.putText(img, text, org, cv2.FONT_HERSHEY_SIMPLEX, font_scale, (0, 0, 0), thickness + 2, cv2.LINE_AA)
    cv2.putText(img, text, org, cv2.FONT_HERSHEY_SIMPLEX, font_scale, color, thickness, cv2.LINE_AA)

def draw_object_ids(img_bgr: np.ndarray, contours: list[np.ndarray], labels: Optional[list[str]] = None) -> np.ndarray:
    out = img_bgr.copy()
    for i, c in enumerate(contours, 1):
        M = cv2.moments(c)
        if M["m00"] != 0:
            cx, cy = int(M["m10"] / M["m00"]), int(M["m01"] / M["m00"])
            text = labels[i - 1] if (labels and i - 1 < len(labels)) else str(i)
            _put_text_outline(out, text, (cx, cy), font_scale=0.5, color=(0, 255, 0), thickness=1)
    return out

def save_debug_ids(folder: Path, original_name: str, img_bgr: np.ndarray, accepted_contours: list[np.ndarray], object_ids: list[str], pixel_size_um: Optional[float] = None, force: bool = False) -> None:
    labeled = draw_object_ids(img_bgr, accepted_contours, labels=object_ids)
    save_debug(folder, original_name.replace(".png", "_ids.png"), labeled, pixel_size_um, force)

def _to_unit_float32(img: np.ndarray) -> np.ndarray:
    if img.ndim == 3: img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    out = np.zeros(img.shape, dtype=np.uint8)
    cv2.normalize(img, out, 0.0, 255.0, cv2.NORM_MINMAX)
    return out.astype(np.float32) / 255.0

def _dog_filter(img_f32: np.ndarray, sigma_fine: float = 1.5, sigma_coarse: float = 4.0) -> np.ndarray:
    fine = cv2.GaussianBlur(img_f32, (0, 0), sigma_fine)
    coarse = cv2.GaussianBlur(img_f32, (0, 0), sigma_coarse)
    dog = np.abs(fine - coarse)
    mx = float(dog.max())
    return (dog / mx).astype(np.float32) if mx > 0 else dog

def _phase_correlate_windowed(src: np.ndarray, dst: np.ndarray) -> tuple[tuple[float, float], float]:
    h, w = src.shape
    win = np.outer(np.hanning(h).astype(np.float32), np.hanning(w).astype(np.float32))
    shift_xy, response = cv2.phaseCorrelate(src * win, dst * win)
    return (float(shift_xy[1]), float(shift_xy[0])), float(response)

def _build_alignment_diagnostics(bf_f32: np.ndarray, fluor_original: np.ndarray, fluor_aligned: np.ndarray, resp_raw: float, resp_dog: float) -> dict:
    def _bgr_overlay(bf: np.ndarray, fluor: np.ndarray) -> np.ndarray:
        h, w = bf.shape
        canvas = np.zeros((h, w, 3), dtype=np.uint8)
        canvas[:, :, 1] = (bf * 255).clip(0, 255).astype(np.uint8)
        fl = cv2.cvtColor(fluor, cv2.COLOR_BGR2GRAY) if fluor.ndim == 3 else fluor
        fl_norm = np.zeros_like(fl, dtype=np.uint8)
        cv2.normalize(fl, fl_norm, 0.0, 255.0, cv2.NORM_MINMAX)
        canvas[:, :, 2] = fl_norm
        return canvas
    bf_bgr = cv2.cvtColor((bf_f32 * 255).clip(0, 255).astype(np.uint8), cv2.COLOR_GRAY2BGR)
    return {
        'overlay_none': _bgr_overlay(bf_f32, fluor_original),
        'overlay_pos': _bgr_overlay(bf_f32, fluor_aligned),
        'overlay_neg': bf_bgr,
        'resp_raw': resp_raw, 'resp_dog': resp_dog,
        'shift_pos': (0.0, 0.0), 'shift_neg': (0.0, 0.0),
        'error_direct': resp_raw, 'error_inverted': resp_dog,
    }

def align_fluorescence_channel(bf_img: np.ndarray, fluor_img: np.ndarray) -> tuple[np.ndarray, tuple[float, float], dict]:
    MAX_SHIFT_PX, MIN_RESPONSE, AGREEMENT_PX = 30.0, 0.02, 2.0
    bf_f, fluor_f = _to_unit_float32(bf_img), _to_unit_float32(fluor_img)
    bf_dog, fluor_dog = _dog_filter(bf_f), _dog_filter(fluor_f)

    try:
        shift_raw, resp_raw = _phase_correlate_windowed(bf_f, fluor_f)
        raw_ok = resp_raw >= MIN_RESPONSE and np.hypot(*shift_raw) <= MAX_SHIFT_PX
    except Exception: shift_raw, resp_raw, raw_ok = (0.0, 0.0), 0.0, False

    try:
        shift_dog, resp_dog = _phase_correlate_windowed(bf_dog, fluor_dog)
        dog_ok = resp_dog >= MIN_RESPONSE and np.hypot(*shift_dog) <= MAX_SHIFT_PX
    except Exception: shift_dog, resp_dog, dog_ok = (0.0, 0.0), 0.0, False

    if raw_ok and dog_ok and np.hypot(shift_raw[0] - shift_dog[0], shift_raw[1] - shift_dog[1]) <= AGREEMENT_PX:
        shift_y, shift_x = (shift_raw[0] + shift_dog[0]) / 2.0, (shift_raw[1] + shift_dog[1]) / 2.0
        method = "Raw+DoG averaged"
    elif dog_ok:
        shift_y, shift_x = shift_dog
        method = "DoG only"
    elif raw_ok:
        shift_y, shift_x = shift_raw
        method = "Raw only"
    else:
        diag = _build_alignment_diagnostics(bf_f, fluor_img, fluor_img, resp_raw, resp_dog)
        diag['method_used'] = 'none'
        return fluor_img.copy(), (0.0, 0.0), diag

    rows, cols = fluor_img.shape[:2]
    M = np.array([[1.0, 0.0, shift_x], [0.0, 1.0, shift_y]], dtype=np.float32)
    aligned = cv2.warpAffine(fluor_img, M, (cols, rows), flags=cv2.INTER_LINEAR, borderMode=cv2.BORDER_CONSTANT, borderValue=0)
    diag = _build_alignment_diagnostics(bf_f, fluor_img, aligned, resp_raw, resp_dog)
    diag['method_used'] = method
    return aligned, (float(shift_y), float(shift_x)), diag

def segment_fluorescence_global(fluor_img8: np.ndarray, bacteria_config: 'SegmentationConfig') -> np.ndarray:
    blur = cv2.GaussianBlur(fluor_img8, (0, 0), sigmaX=bacteria_config.gaussian_sigma * 0.1, sigmaY=bacteria_config.gaussian_sigma * 0.1)
    otsu_threshold = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[0]
    _, bw = cv2.threshold(blur, otsu_threshold * 0.5, 255, cv2.THRESH_BINARY)
    k = np.ones((bacteria_config.morph_kernel_size, bacteria_config.morph_kernel_size), np.uint8)
    bw = cv2.morphologyEx(bw, cv2.MORPH_OPEN, k, iterations=1)
    return cv2.morphologyEx(bw, cv2.MORPH_CLOSE, k, iterations=1)

def contour_intersection_area_px(c1: np.ndarray, c2: np.ndarray, shape_hw: tuple[int, int]) -> float:
    m1, m2 = np.zeros(shape_hw, dtype=np.uint8), np.zeros(shape_hw, dtype=np.uint8)
    cv2.drawContours(m1, [c1], -1, 255, thickness=-1)
    cv2.drawContours(m2, [c2], -1, 255, thickness=-1)
    return float(np.count_nonzero(cv2.bitwise_and(m1, m2)))

def match_fluor_to_bf_by_overlap(bf_contours: list[np.ndarray], fluor_contours: list[np.ndarray], img_shape_hw: tuple[int, int], min_intersection_px: float = FLUOR_MATCH_MIN_INTERSECTION_PX) -> list[Optional[int]]:
    matches = []
    fluor_boxes = [cv2.boundingRect(c) for c in fluor_contours]
    for bf in bf_contours:
        bx, by, bw, bh = cv2.boundingRect(bf)
        best_idx, best_inter = None, 0.0
        for j, (fx, fy, fw, fh) in enumerate(fluor_boxes):
            if (bx + bw < fx) or (fx + fw < bx) or (by + bh < fy) or (fy + fh < by): continue
            inter = contour_intersection_area_px(bf, fluor_contours[j], img_shape_hw)
            if inter > best_inter: best_inter, best_idx = inter, j
        matches.append(best_idx if best_inter >= min_intersection_px else None)
    return matches

def measure_fluorescence_intensity_with_global_area(fluor_img: np.ndarray, bf_contours: list[np.ndarray], fluor_contours: list[np.ndarray], bf_to_fluor_match: list[Optional[int]], um_per_px_x: float, um_per_px_y: float) -> list[dict]:
    um2_per_px2 = float(um_per_px_x) * float(um_per_px_y)
    measurements = []
    for i, bf in enumerate(bf_contours, 1):
        bf_mask = np.zeros(fluor_img.shape[:2], dtype=np.uint8)
        cv2.drawContours(bf_mask, [bf], -1, 255, thickness=-1)
        fluor_values = fluor_img[bf_mask > 0]
        j = bf_to_fluor_match[i - 1]
        s2_area_px = float(cv2.contourArea(fluor_contours[j])) if j is not None else 0.0
        if fluor_values.size > 0:
            measurements.append({"object_id": i, "fluor_area_px": s2_area_px, "fluor_area_um2": s2_area_px * um2_per_px2, "fluor_mean": float(np.mean(fluor_values)), "fluor_median": float(np.median(fluor_values)), "fluor_std": float(np.std(fluor_values)), "fluor_min": float(np.min(fluor_values)), "fluor_max": float(np.max(fluor_values)), "fluor_integrated_density": float(np.sum(fluor_values))})
        else:
            measurements.append({"object_id": i, "fluor_area_px": s2_area_px, "fluor_area_um2": s2_area_px * um2_per_px2, "fluor_mean": 0.0, "fluor_median": 0.0, "fluor_std": 0.0, "fluor_min": 0.0, "fluor_max": 0.0, "fluor_integrated_density": 0.0})
    return measurements

def visualize_fluorescence_measurements(fluor_img8: np.ndarray, contours: list[np.ndarray], measurements: list[dict]) -> np.ndarray:
    vis = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(vis, contours, -1, (0, 255, 0), 1)
    for m in measurements:
        M = cv2.moments(contours[m["object_id"] - 1])
        if M["m00"] != 0:
            cv2.putText(vis, f"{m['object_id']}: {m['fluor_mean']:.0f}", (int(M["m10"] / M["m00"]), int(M["m01"] / M["m00"])), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 255, 0), 1, cv2.LINE_AA)
    return vis

def segment_particles_brightfield(img8: np.ndarray, pixel_size_um: float, out_dir: Path, bacteria_config: 'SegmentationConfig') -> np.ndarray:
    blur = cv2.GaussianBlur(img8, (0, 0), sigmaX=bacteria_config.gaussian_sigma, sigmaY=bacteria_config.gaussian_sigma)
    if bacteria_config.use_intensity_threshold:
        _, thresh = cv2.threshold(blur, bacteria_config.intensity_threshold, 255, cv2.THRESH_BINARY_INV)
    else:
        enhanced = cv2.subtract(cv2.GaussianBlur(img8, (0, 0), sigmaX=bacteria_config.gaussian_sigma, sigmaY=bacteria_config.gaussian_sigma), img8)
        _, thresh = cv2.threshold(cv2.GaussianBlur(enhanced, (3, 3), 0), 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    kernel = np.ones((bacteria_config.morph_kernel_size, bacteria_config.morph_kernel_size), np.uint8)
    bw = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, iterations=bacteria_config.morph_iterations)
    bw = cv2.morphologyEx(bw, cv2.MORPH_CLOSE, kernel, iterations=bacteria_config.morph_iterations + 1)
    if bacteria_config.dilate_iterations > 0: bw = cv2.dilate(bw, kernel, iterations=bacteria_config.dilate_iterations)
    if bacteria_config.erode_iterations > 0: bw = cv2.erode(bw, kernel, iterations=bacteria_config.erode_iterations)
    return bw


# ==================================================
# Multi-Scan & Caching
# ==================================================

def get_cache_key(image_path: Path, bacteria_config: 'SegmentationConfig') -> str:
    import hashlib
    return hashlib.md5(f"{image_path}_{bacteria_config.name}_{bacteria_config.gaussian_sigma}_{bacteria_config.min_area_um2}_{bacteria_config.max_area_um2}".encode()).hexdigest()

def check_cache(cache_dir: Path, cache_key: str) -> Optional[dict]:
    cache_file = cache_dir / f"{cache_key}.json"
    if cache_file.exists():
        try:
            with open(cache_file, 'r', encoding='utf-8') as f: return json.load(f).get('results')
        except Exception: pass
    return None

def save_cache(cache_dir: Path, cache_key: str, results: dict) -> None:
    cache_dir.mkdir(parents=True, exist_ok=True)
    try:
        with open(cache_dir / f"{cache_key}.json", 'w', encoding='utf-8') as f:
            json.dump({'timestamp': datetime.now().isoformat(), 'results': results}, f, indent=2, default=str)
    except Exception: pass

def clear_old_cache(cache_dir: Path, max_age_days: int = 7) -> None:
    if not cache_dir.exists(): return
    from datetime import timedelta
    cutoff = datetime.now() - timedelta(days=max_age_days)
    for f in cache_dir.glob("*.json"):
        try:
            if datetime.fromtimestamp(f.stat().st_mtime) < cutoff: f.unlink()
        except Exception: pass

def collect_images_from_directory(source_dir: Path) -> Dict[str, Dict]:
    return {d.name: {'path': d, 'images': list(d.glob("*_ch00.tif"))} for d in sorted(source_dir.iterdir()) if d.is_dir() and list(d.glob("*_ch00.tif"))}

def process_multi_configuration(config: dict) -> dict:
    bacteria_config_info = config.get('bacteria_config_info', {})
    configs_to_scan = bacteria_config_info.get('configs_to_scan', [])
    config_names = bacteria_config_info.get('config_names', {})
    _empty = {'mode': 'multi_scan', 'all_results': {}, 'comparison_df': pd.DataFrame(), 'best_match': None, 'confidence_report': "No configurations available", 'comparison_path': None, 'report_path': None, 'plot_path': None}

    source_to_scan = config.get('current_source', config.get('source_dir'))
    if not configs_to_scan or source_to_scan is None: return _empty

    image_groups = collect_images_from_directory(source_to_scan)
    if not image_groups: return _empty

    cache_dir = config['output_dir'] / ".cache"
    cache_dir.mkdir(exist_ok=True)
    clear_old_cache(cache_dir)

    all_results = {}
    _ms_start = pytime.monotonic()
    total_imgs_all = sum(len(v['images']) for v in image_groups.values())
    print(f"\n  {len(configs_to_scan)} configuration(s) × "
          f"{len(image_groups)} group(s) × {total_imgs_all} image(s)")

    for _cfg_idx, bacteria_type in enumerate(configs_to_scan, 1):
        bacteria_config = load_bacteria_config_from_json(bacteria_type)
        if not bacteria_config: continue
        cfg_display = config_names.get(bacteria_type, bacteria_type)
        print(f"\n  [{_cfg_idx}/{len(configs_to_scan)}] {cfg_display}")
        config_output = config['output_dir'] / bacteria_type
        config_output.mkdir(exist_ok=True)

        scan_results = {
            'particles_detected': 0, 'fluorescence_sum': 0.0, 'mean_fluorescence': 0.0,
            'images_processed': 0, 'images_failed': 0, 'cached': False,
            'per_image_particles': [], 'per_image_fluorescence': [],
            'test_fluorescences': [], 'control_fluorescences': [],
            'test_particles': [], 'control_particles': [],
            'use_intensity_threshold': bacteria_config.use_intensity_threshold,
            'morph_circularity_list': [], 'morph_halo_ratio_list': [], 'morph_ripleys_ratio_list': [],
            'intensity_threshold': bacteria_config.intensity_threshold if bacteria_config.use_intensity_threshold else 80.0
        }

        for group_name, group_data in image_groups.items():
            is_control = group_name.lower().startswith('control')
            _grp_label = "Control" if is_control else group_name
            for img_path in tqdm(
                group_data['images'],
                desc=f"    {_grp_label:<12}",
                unit="img",
                leave=False,
                ncols=88,
            ):
                cache_key = get_cache_key(img_path, bacteria_config)
                cached_result = check_cache(cache_dir, cache_key)
                group_output = config_output / group_name

                if cached_result:
                    particle_count, mean_fluor = cached_result.get('particle_count', 0), cached_result.get('mean_fluorescence', 0.0)
                else:
                    try:
                        process_image(img_path, group_output, bacteria_config)
                        csv_path = group_output / img_path.stem / "object_stats.csv"
                        if csv_path.exists():
                            df = pd.read_csv(csv_path)
                            particle_count = len(df)
                            mean_fluor = float(pd.to_numeric(df['Fluor_Mean'], errors='coerce').dropna().mean()) if 'Fluor_Mean' in df.columns and len(df) > 0 else 0.0
                            save_cache(cache_dir, cache_key, {'particle_count': particle_count, 'mean_fluorescence': mean_fluor})
                        else: continue
                    except Exception:
                        scan_results['images_failed'] += 1
                        continue

                _morph_path = group_output / img_path.stem / "morphological_features.json"
                if _morph_path.exists():
                    try:
                        with open(_morph_path, 'r', encoding='utf-8') as _mf: _mf_data = json.load(_mf)
                        if _mf_data.get('mean_circularity', -1) > 0: scan_results['morph_circularity_list'].append(_mf_data['mean_circularity'])
                        if _mf_data.get('mean_halo_ratio', -1) > 0: scan_results['morph_halo_ratio_list'].append(_mf_data['mean_halo_ratio'])
                        if _mf_data.get('ripleys_ratio', -1) > 0: scan_results['morph_ripleys_ratio_list'].append(_mf_data['ripleys_ratio'])
                    except Exception: pass

                scan_results['particles_detected'] += particle_count
                scan_results['fluorescence_sum'] += mean_fluor
                scan_results['images_processed'] += 1
                scan_results['per_image_particles'].append(particle_count)
                scan_results['per_image_fluorescence'].append(mean_fluor)
                if cached_result: scan_results['cached'] = True

                if is_control:
                    scan_results['control_fluorescences'].append(mean_fluor)
                    scan_results['control_particles'].append(particle_count)
                else:
                    scan_results['test_fluorescences'].append(mean_fluor)
                    scan_results['test_particles'].append(particle_count)

        n_proc = scan_results['images_processed']
        if n_proc > 0:
            scan_results['mean_fluorescence'] = scan_results['fluorescence_sum'] / n_proc
            fa, pa = np.array(scan_results['per_image_fluorescence']), np.array(scan_results['per_image_particles'])
            scan_results['std_fluorescence'] = float(np.std(fa, ddof=1)) if len(fa) > 1 else 0.0
            scan_results['sem_fluorescence'] = float(scipy_stats.sem(fa)) if len(fa) > 1 else 0.0
            scan_results['mean_particles_per_image'] = float(np.mean(pa)) if len(pa) > 1 else float(scan_results['particles_detected'])
            scan_results['std_particles'] = float(np.std(pa, ddof=1)) if len(pa) > 1 else 0.0
            for k, lk in [('mean_circularity', 'morph_circularity_list'), ('mean_halo_ratio', 'morph_halo_ratio_list'), ('mean_ripleys_ratio', 'morph_ripleys_ratio_list')]:
                scan_results[k] = float(np.mean(scan_results[lk])) if scan_results.get(lk) else -1.0

            # ── Excel consolidation with visible status ───────────────
            _grps_xl = [gn for gn in image_groups if (config_output / gn).exists()]
            if _grps_xl:
                print(
                    f"    Consolidating {len(_grps_xl)} group(s) \u2192 Excel \u2026",
                    end="", flush=True,
                )
                _xl_t0 = pytime.monotonic()
                for group_name in _grps_xl:
                    consolidate_to_excel(
                        config_output / group_name, group_name, config['percentile']
                    )
                print(
                    f"  done  ({_fmt_elapsed(pytime.monotonic() - _xl_t0)})",
                    flush=True,
                )

        _cfg_elapsed = _fmt_elapsed(pytime.monotonic() - _ms_start)
        _disq = " ⛔ disqualified (overcount)" if scan_results.get('_overcount_disqualified') else ""
        print(
            f"    \u2713 {cfg_display}: "
            f"{scan_results['particles_detected']} particles  "
            f"fluor={scan_results.get('mean_fluorescence', 0.0):.3f}  "
            f"({scan_results['images_processed']} processed, "
            f"{scan_results['images_failed']} failed, "
            f"{_cfg_elapsed}){_disq}"
        )
        all_results[bacteria_type] = scan_results

    # ── post-loop: confidence report & comparison plot ────────────────
    print(f"  Generating confidence report \u2026", end="", flush=True)
    _rep_t0 = pytime.monotonic()
    comparison_df, confidence_report = generate_confidence_report(
        all_results, config_names, config['output_dir']
    )
    print(f"  done  ({_fmt_elapsed(pytime.monotonic() - _rep_t0)})", flush=True)

    print(f"  Generating comparison plot \u2026", end="", flush=True)
    _plt_t0 = pytime.monotonic()
    plot_path = generate_multi_config_comparison_plot(
        comparison_df, all_results, config['output_dir']
    )
    print(f"  done  ({_fmt_elapsed(pytime.monotonic() - _plt_t0)})", flush=True)

    comparison_path = config['output_dir'] / "configuration_comparison.csv"
    comparison_df.to_csv(comparison_path, index=False)
    report_path = config['output_dir'] / "confidence_report.txt"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(confidence_report)

    return {
        'mode': 'multi_scan', 'all_results': all_results,
        'comparison_df': comparison_df,
        'best_match': comparison_df.iloc[0] if not comparison_df.empty else None,
        'confidence_report': confidence_report,
        'comparison_path': comparison_path,
        'report_path': report_path, 'plot_path': plot_path,
    }


def run_multi_config_scan(config: dict, bacteria_config_info: dict) -> dict:
    if 'current_source' not in config: config['current_source'] = config.get('source_dir')
    source_to_scan = config['current_source']
    if source_to_scan is None or not source_to_scan.exists(): return {'ranked_results': [], 'all_results': {}, 'comparison_df': pd.DataFrame(), 'report_path': None, 'configs': {}}

    results = process_multi_configuration(config)
    configs_dict = {k: load_bacteria_config_from_json(k) for k in bacteria_config_info.get('configs_to_scan', []) if load_bacteria_config_from_json(k)}

    ranked_results = []
    if not results['comparison_df'].empty:
        for _, row in results['comparison_df'].iterrows():
            ranked_results.append({'rank': int(row['Rank']), 'config_key': str(row['Config_Key']), 'bacteria_name': str(row['Bacteria_Type']), 'confidence': float(row['Confidence_Score']), 'particles': int(row['Particles_Detected']), 'mean_fluorescence': float(row['Mean_Fluorescence'])})

    return {'ranked_results': ranked_results, 'all_results': results['all_results'], 'comparison_df': results['comparison_df'], 'report_path': results.get('report_path'), 'configs': configs_dict}

def promote_multiscan_output_as_clinical(config: dict, chosen_config_key: str) -> bool:
    output_dir, source_dir = config['output_dir'], config['current_source']
    src_base = output_dir / chosen_config_key
    if not src_base.exists(): return False

    image_groups = collect_images_from_directory(source_dir)
    promoted = 0
    for group_name in image_groups:
        src, dst = src_base / group_name, output_dir / group_name
        if not src.exists() or not (src / f"{group_name}_master.xlsx").exists(): return False
        if dst.exists(): shutil.rmtree(dst)
        shutil.move(str(src), str(dst))
        promoted += 1
    return promoted == len(image_groups)


# ==================================================
# Scoring & Classification
# ==================================================

def calculate_confidence_score(results: dict, default_fluor: Optional[float] = None) -> float:
    OVERCOUNT_HARD_LIMIT, OVERCOUNT_MID_LIMIT, OVERCOUNT_SOFT_LIMIT = 100.0, 50.0, 30.0
    NONSPECIFIC_FLUOR_CAP, NONSPECIFIC_FLUOR_STRICT = 2.0, 1.0

    results['_overcount_disqualified'], results['_mean_per_image_used'], results['_pooled_fluor_mean'] = False, 0.0, 0.0
    test_fluor, ctrl_fluor = np.array(results.get('test_fluorescences', []), dtype=float), np.array(results.get('control_fluorescences', []), dtype=float)
    n_test, n_ctrl = len(test_fluor), len(ctrl_fluor)

    per_image_particles = results.get('per_image_particles', [])
    if per_image_particles: mean_per_image = float(np.mean(per_image_particles))
    elif (n_test + n_ctrl) > 0: mean_per_image = float(np.sum(results.get('test_particles', [])) + np.sum(results.get('control_particles', []))) / max(n_test + n_ctrl, 1)
    else: mean_per_image = float(results.get('mean_particles_per_image', 0))

    if n_test < 2 or n_ctrl < 2:
        mean_fluor_fallback = float(results.get('mean_fluorescence', 0.0))
        results['_mean_per_image_used'], results['_pooled_fluor_mean'] = round(mean_per_image, 2), round(mean_fluor_fallback, 3)
        results['_overcount_disqualified'] = mean_per_image > OVERCOUNT_SOFT_LIMIT and mean_fluor_fallback < NONSPECIFIC_FLUOR_CAP
        n_particles = results.get('particles_detected', 0)
        if results['_overcount_disqualified']: return 3.0
        if 20 <= n_particles <= 200: return 15.0
        elif 5 <= n_particles < 20 or 200 < n_particles <= 500: return 8.0
        return 3.0 if n_particles > 0 else 0.0

    test_mean, ctrl_mean = float(np.mean(test_fluor)), float(np.mean(ctrl_fluor))
    test_std, ctrl_std = float(np.std(test_fluor, ddof=1)), float(np.std(ctrl_fluor, ddof=1))
    score = 0.0

    denom = n_test + n_ctrl - 2
    pooled_std = np.sqrt(((n_test - 1) * test_std ** 2 + (n_ctrl - 1) * ctrl_std ** 2) / denom) if denom > 0 else 1.0
    cohens_d = abs(test_mean - ctrl_mean) / pooled_std if pooled_std > 0 else 0.0

    if cohens_d >= 2.0: score += 35
    elif cohens_d >= 1.5: score += 30
    elif cohens_d >= 1.0: score += 25
    elif cohens_d >= 0.8: score += 20
    elif cohens_d >= 0.5: score += 12
    elif cohens_d >= 0.2: score += 5

    try: p_value = float(np.asarray(scipy_stats.ttest_ind(test_fluor, ctrl_fluor, equal_var=False)[1]).item())
    except Exception: p_value = 1.0

    if p_value < 0.001: score += 20
    elif p_value < 0.01: score += 15
    elif p_value < 0.05: score += 10
    elif p_value < 0.10: score += 3

    if 3 <= mean_per_image <= 30: score += 15
    elif 2 <= mean_per_image < 3: score += 10
    elif 30 < mean_per_image <= 60: score += 5
    elif 60 < mean_per_image <= 150: score += 2
    elif mean_per_image >= 1: score += 5

    avg_cv = ((test_std / test_mean if test_mean > 0 else 999.0) + (ctrl_std / ctrl_mean if ctrl_mean > 0 else 999.0)) / 2.0
    if avg_cv < 0.30: score += 15
    elif avg_cv < 0.50: score += 10
    elif avg_cv < 0.80: score += 5

    direction_bonus = 3.0 if test_mean < ctrl_mean and cohens_d >= 0.5 else (1.0 if test_mean < ctrl_mean else 0.0)
    sb_bonus = 0.0
    if default_fluor and default_fluor > 0:
        sb_ratio = float(np.mean(np.concatenate([test_fluor, ctrl_fluor]))) / default_fluor
        if sb_ratio >= 1.5: sb_bonus = 2.0
        elif sb_ratio >= 1.1: sb_bonus = 1.0
    score += min(direction_bonus + sb_bonus, 5.0)

    pooled_fluor_mean = float(np.mean(np.concatenate([test_fluor, ctrl_fluor])))
    is_disqualified = False

    if mean_per_image > OVERCOUNT_HARD_LIMIT and pooled_fluor_mean < NONSPECIFIC_FLUOR_CAP:
        score, is_disqualified = min(score, 10.0), True
    elif mean_per_image > OVERCOUNT_MID_LIMIT and pooled_fluor_mean < NONSPECIFIC_FLUOR_CAP:
        score, is_disqualified = score * 0.35, True
    elif mean_per_image > OVERCOUNT_SOFT_LIMIT and pooled_fluor_mean < NONSPECIFIC_FLUOR_STRICT:
        score, is_disqualified = score * 0.55, True

    results['_overcount_disqualified'], results['_mean_per_image_used'], results['_pooled_fluor_mean'] = is_disqualified, round(mean_per_image, 2), round(pooled_fluor_mean, 3)
    return min(score, 100.0)

def _apply_pairwise_discrimination_penalties(all_results: dict, raw_scores: dict[str, float]) -> tuple[dict[str, float], dict[str, dict]]:
    adjusted, penalty_log = dict(raw_scores), {}
    sorted_keys = sorted(raw_scores, key=lambda k: raw_scores[k], reverse=True)

    for i in range(len(sorted_keys) - 1):
        key_hi, key_lo = sorted_keys[i], sorted_keys[i + 1]
        fluor_hi = [float(x) for x in all_results[key_hi].get('per_image_fluorescence', []) if x is not None and not np.isnan(x)]
        fluor_lo = [float(x) for x in all_results[key_lo].get('per_image_fluorescence', []) if x is not None and not np.isnan(x)]
        if len(fluor_hi) < 3 or len(fluor_lo) < 3: continue
        try:
            p_val = float(np.asarray(scipy_stats.ttest_ind(fluor_hi, fluor_lo, equal_var=False)[1]).item())
            if p_val >= 0.05:
                score_hi, score_lo = adjusted[key_hi], adjusted[key_lo]
                factor = 0.30 + 0.60 * min((p_val - 0.05) / 0.95, 1.0)
                new_hi = round(max(score_hi - (score_hi - score_lo) * factor, score_lo), 1)
                adjusted[key_hi] = new_hi
                penalty_log[key_hi] = {'peer': key_lo, 'p_value': round(p_val, 6), 'factor': round(factor, 3), 'old_score': round(score_hi, 1), 'new_score': new_hi}
        except Exception: continue
    return adjusted, penalty_log

def generate_confidence_report(all_results: dict, config_names: dict, output_dir: Path) -> tuple[pd.DataFrame, str]:
    _default_fluor = max(float(all_results['default'].get('mean_fluorescence', 0.0)), 0.1) if 'default' in all_results else None
    raw_scores = {k: calculate_confidence_score(v, _default_fluor) for k, v in all_results.items()}
    adjusted_scores, penalty_log = _apply_pairwise_discrimination_penalties(all_results, raw_scores)

    for k, results in all_results.items():
        test_fl, ctrl_fl = np.array(results.get('test_fluorescences', []), dtype=float), np.array(results.get('control_fluorescences', []), dtype=float)
        if len(test_fl) >= 2 and len(ctrl_fl) >= 2:
            denom = len(test_fl) + len(ctrl_fl) - 2
            pooled_std = np.sqrt(((len(test_fl) - 1) * np.std(test_fl, ddof=1) ** 2 + (len(ctrl_fl) - 1) * np.std(ctrl_fl, ddof=1) ** 2) / denom) if denom > 0 else 1.0
            d_display = float((np.mean(test_fl) - np.mean(ctrl_fl)) / pooled_std) if pooled_std > 0 else 0.0
            try: p_display = float(np.asarray(scipy_stats.ttest_ind(test_fl, ctrl_fl, equal_var=False)[1]).item())
            except Exception: p_display = np.nan
        else:
            d_display, p_display = 0.0, np.nan
        results['_score_cohens_d'], results['_score_p_value'] = round(d_display, 3), round(p_display, 6) if not np.isnan(p_display) else np.nan

    comparison_data = []
    for k, results in all_results.items():
        sb_ratio = results['mean_fluorescence'] / _default_fluor if _default_fluor and _default_fluor > 0 and k != 'default' else 1.0
        test_fl, ctrl_fl = results.get('test_fluorescences', []), results.get('control_fluorescences', [])
        comparison_data.append({
            'Rank': 0, 'Bacteria_Type': config_names.get(k, k), 'Config_Key': k,
            'Particles_Detected': int(results['particles_detected']), 'Particles_Per_Image': float(results.get('mean_particles_per_image', 0)),
            'Particles_Std': float(results.get('std_particles', 0)), 'Mean_Fluorescence': float(results['mean_fluorescence']),
            'Fluor_Std': float(results.get('std_fluorescence', 0)), 'Fluor_SEM': float(results.get('sem_fluorescence', 0)),
            'Signal_to_Background': round(sb_ratio, 2), 'Test_Mean': round(float(np.mean(test_fl)) if test_fl else 0.0, 2),
            'Control_Mean': round(float(np.mean(ctrl_fl)) if ctrl_fl else 0.0, 2), 'Cohens_d': float(results.get('_score_cohens_d', 0)),
            'P_Value_TvC': float(results['_score_p_value']) if not pd.isna(results.get('_score_p_value', np.nan)) else np.nan,
            'Images_Processed': int(results['images_processed']), 'Images_Failed': int(results['images_failed']),
            'Raw_Score': round(raw_scores[k], 1), 'Confidence_Score': float(adjusted_scores[k]), 'Confidence_Percent': float(adjusted_scores[k]),
            'Penalty_Applied': k in penalty_log, 'Low_Fluor_Flag': float(np.mean(np.array(test_fl + ctrl_fl, dtype=float))) < 1.0 if (test_fl or ctrl_fl) else False,
            'Overcount_Disqualified': bool(results.get('_overcount_disqualified', False))
        })

    comparison_df = pd.DataFrame(comparison_data)
    if comparison_df.empty: return comparison_df, "No results to report"
    comparison_df = comparison_df.sort_values('Confidence_Percent', ascending=False).reset_index(drop=True)
    comparison_df['Rank'] = range(1, len(comparison_df) + 1)

    top1_key = str(comparison_df.iloc[0]['Config_Key'])
    top2_key = str(comparison_df.iloc[1]['Config_Key']) if len(comparison_df) > 1 else None
    stat_ambiguous, p_value_top2, t_stat_top2 = False, np.nan, np.nan

    if top2_key:
        top1_fluor = [float(x) for x in all_results[top1_key].get('per_image_fluorescence', []) if x is not None and not np.isnan(x)]
        top2_fluor = [float(x) for x in all_results[top2_key].get('per_image_fluorescence', []) if x is not None and not np.isnan(x)]
        if len(top1_fluor) >= 3 and len(top2_fluor) >= 3:
            try:
                t_stat_top2, p_value_top2 = [float(np.asarray(x).item()) for x in scipy_stats.ttest_ind(top1_fluor, top2_fluor, equal_var=False)]
                if p_value_top2 >= 0.05: stat_ambiguous = True
            except Exception: pass

    lines = ["=" * 80, "BACTERIA IDENTIFICATION CONFIDENCE REPORT", "=" * 80, "", f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", f"Output Directory: {output_dir.name}", "", "DETECTION SUMMARY:", "─" * 80]
    top_result = comparison_df.iloc[0]
    top_confidence, top_particles, top_name = float(top_result['Confidence_Percent']), int(top_result['Particles_Detected']), str(top_result['Bacteria_Type'])

    if top_particles == 0:
        lines.extend(["Status: NO BACTERIA DETECTED", "Confidence: High", "", "⚪ All configurations detected 0 particles", "   → Sample appears clean"])
    elif stat_ambiguous:
        lines.extend(["Status: BACTERIA DETECTED — AMBIGUOUS IDENTIFICATION", f"Confidence: Reduced ({top_confidence:.1f}% after convergence penalty, p={p_value_top2:.4f})", "", "⚠ Top two configurations are NOT statistically distinguishable", f"  Rank 1: {top_name}", f"  Rank 2: {str(comparison_df.iloc[1]['Bacteria_Type'])}", "  → Manual review or culture confirmation STRONGLY recommended"])
    elif top_confidence >= 70:
        lines.extend(["Status: BACTERIA DETECTED", f"Confidence: High ({top_confidence:.1f}%)", "", f"✓ Strong match to: {top_name}"])
    elif top_confidence >= 50:
        lines.extend(["Status: BACTERIA DETECTED", f"Confidence: Moderate ({top_confidence:.1f}%)", "", f"⚠ Possible match to: {top_name}", "   → Manual review recommended"])
    else:
        lines.extend(["Status: AMBIGUOUS", f"Confidence: Low ({top_confidence:.1f}%)", "", "⚠ MANUAL REVIEW REQUIRED", "   → Multiple configurations show similar results", "   → Consider culture-based confirmation"])
    lines.extend(["", ""])

    disqualified_keys = [k for k, v in all_results.items() if v.get('_overcount_disqualified', False)]
    if disqualified_keys:
        lines.extend(["⛔ DISQUALIFIED CONFIGURATIONS (overcount + low fluorescence):", "─" * 80])
        for key in disqualified_keys:
            lines.append(f"  ✗ {config_names.get(key, key)}: {all_results[key].get('_mean_per_image_used', 0):.0f} particles/image, mean fluor = {all_results[key].get('_pooled_fluor_mean', 0):.3f} a.u./µm²  → Non-specific detections (overcount artefact)")
        lines.append("")

    lines.extend(["TOP 3 CONFIGURATION MATCHES:", "─" * 80])
    for idx in range(min(3, len(comparison_df))):
        row = comparison_df.iloc[idx]
        lines.extend([
            f"{'🥇' if row['Rank'] == 1 else '🥈' if row['Rank'] == 2 else '🥉'} Rank {row['Rank']}: {row['Bacteria_Type']}",
            f"   Confidence: {float(row['Confidence_Percent']):.1f}%" + (f" (raw: {float(row.get('Raw_Score', row['Confidence_Percent'])):.1f}%, pairwise penalty applied)" if row.get('Penalty_Applied', False) else ""),
            *([f"   ⚠ Low fluorescence flag: mean < 1.0 a.u./µm² — detections may be non-specific"] if row.get('Low_Fluor_Flag', False) else []),
            f"   Total particles: {int(row['Particles_Detected'])}",
            f"   Particles/image: {float(row['Particles_Per_Image']):.1f} ± {float(row['Particles_Std']):.1f}",
            f"   Mean fluorescence: {float(row['Mean_Fluorescence']):.2f} ± {float(row['Fluor_Std']):.2f} a.u./µm² (SEM: ±{float(row['Fluor_SEM']):.2f})" if float(row['Mean_Fluorescence']) > 0 else "   Mean fluorescence: N/A",
            f"   Signal-to-background: {float(row['Signal_to_Background']):.1f}×",
            f"   Test-vs-control p-value: {float(row.get('P_Value_TvC', np.nan)):.4f}" if not pd.isna(row.get('P_Value_TvC', np.nan)) else "   Test-vs-control p-value: N/A"
        ])

    display_df = comparison_df.copy()
    for col, fmt in [('Particles_Per_Image', '{:.1f}'), ('Particles_Std', '±{:.1f}'), ('Mean_Fluorescence', '{:.2f}'), ('Fluor_Std', '±{:.2f}'), ('Fluor_SEM', '±{:.2f}'), ('Signal_to_Background', '{:.1f}×'), ('Confidence_Score', '{:.1f}'), ('Cohens_d', '{:.3f}')]:
        if col in display_df.columns: display_df[col] = display_df[col].apply(lambda x, f=fmt: f.format(float(x)))
    if 'P_Value_TvC' in display_df.columns: display_df['P_Value_TvC'] = display_df['P_Value_TvC'].apply(lambda x: f"{float(x):.4f}" if not pd.isna(x) else "N/A")

    lines.extend(["FULL CONFIGURATION COMPARISON WITH STATISTICS:", "─" * 80, display_df[[c for c in ['Rank', 'Bacteria_Type', 'Particles_Detected', 'Particles_Per_Image', 'Particles_Std', 'Mean_Fluorescence', 'Fluor_Std', 'Fluor_SEM', 'Signal_to_Background', 'Cohens_d', 'P_Value_TvC', 'Confidence_Score', 'Images_Processed', 'Low_Fluor_Flag'] if c in display_df.columns]].to_string(index=False), ""])

    if len(all_results) >= 2 and top2_key:
        lines.extend(["STATISTICAL SIGNIFICANCE ANALYSIS:", "─" * 80, ""])
        if not np.isnan(t_stat_top2):
            lines.extend([f"Comparison: {config_names[top1_key]} vs {config_names[top2_key]}", "  Independent t-test (Welch's) on per-image fluorescence:", f"    t-statistic: {t_stat_top2:.4f}", f"    p-value: {p_value_top2:.6f}", f"    Significance: {'*** Highly significant (p < 0.001)' if p_value_top2 < 0.001 else '** Very significant (p < 0.01)' if p_value_top2 < 0.01 else '* Significant (p < 0.05)' if p_value_top2 < 0.05 else 'Not significant (p ≥ 0.05)'}", "", "  Per-configuration test-vs-control statistics:"])
            for _, row in comparison_df.iterrows():
                p_val_raw = row.get('P_Value_TvC', np.nan)
                p_str = f"{float(p_val_raw):.4f}" if not pd.isna(p_val_raw) else 'N/A'
                lines.append(
                    f"    {str(row['Bacteria_Type']):<35}  "
                    f"d={float(row['Cohens_d']):+.3f}  "
                    f"p={p_str}  "
                    f"{'⚠ low-fluor' if row.get('Low_Fluor_Flag', False) else ''}"
                )
            lines.extend(["", f"  ✓ Top configuration ({config_names[top1_key]}) is statistically distinguishable from rank 2" if p_value_top2 < 0.05 else "  ⚠ Top configurations are NOT statistically different\n    → Convergence penalty applied to rank-1 score\n    → Consider both as viable candidates", ""])
        else: lines.extend(["  Insufficient data for statistical testing", ""])

    lines.extend(["CLINICAL RECOMMENDATIONS:", "─" * 80])
    if top_particles == 0: lines.extend(["✓ No bacteria detected across all configurations", "  → Sample appears clean", "  → No immediate action required"])
    elif stat_ambiguous:
        lines.append(f"⚠ AMBIGUOUS — {top_name} vs {str(comparison_df.iloc[1]['Bacteria_Type'])} not statistically separable")
        if bool(comparison_df.iloc[0].get('Low_Fluor_Flag', False)) and bool(comparison_df.iloc[1].get('Low_Fluor_Flag', False)): lines.extend(["  ⚠ Both top configs have low mean fluorescence (<1.0 a.u./µm²) — likely non-specific detections", "  → Re-tune the bacteria configuration parameters or check segmentation debug images"])
        elif bool(comparison_df.iloc[0].get('Low_Fluor_Flag', False)) or bool(comparison_df.iloc[1].get('Low_Fluor_Flag', False)): lines.append("  ⚠ One of the top configs has low mean fluorescence (<1.0 a.u./µm²) — it may be detecting background")
        lines.extend(["  → Culture-based identification REQUIRED before targeted therapy", "  → Consider empirical broad-spectrum coverage", "  → Infectious disease consultation recommended"])
    elif top_confidence >= 70: lines.extend([f"✓ High confidence match: {top_name}", "  → Proceed with targeted antimicrobial therapy", "  → Consider culture confirmation if treatment fails"])
    elif top_confidence >= 50: lines.extend([f"⚠ Moderate confidence match: {top_name}", "  → Recommend culture-based confirmation", "  → Consider broad-spectrum coverage initially"])
    else: lines.extend(["⚠ LOW CONFIDENCE - MANUAL REVIEW REQUIRED", "  → Multiple configurations show similar results", "  → STRONGLY recommend culture-based identification", "  → Consider infectious disease consultation"])
    lines.extend(["", "=" * 80])

    return comparison_df, "\n".join(lines)

def extract_discriminating_features(contours, centroids, gray_img):
    features = {}
    areas, aspect_ratios, circularities, equiv_diameters = [], [], [], []
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area < 1: continue
        perimeter = cv2.arcLength(cnt, True)
        circularities.append((4 * np.pi * area / (perimeter ** 2)) if perimeter > 0 else 0)
        if len(cnt) >= 5:
            axes = sorted(cv2.fitEllipse(cnt)[1])
            aspect_ratios.append(axes[1] / axes[0] if axes[0] > 0 else 1.0)
        else:
            _, _, w, h = cv2.boundingRect(cnt)
            aspect_ratios.append(max(w, h) / max(min(w, h), 1))
        areas.append(area)
        equiv_diameters.append(np.sqrt(4 * area / np.pi))

    features.update({'mean_area': np.mean(areas) if areas else 0, 'std_area': np.std(areas) if areas else 0, 'mean_circularity': np.mean(circularities) if circularities else 0, 'mean_aspect_ratio': np.mean(aspect_ratios) if aspect_ratios else 0, 'mean_equiv_diameter': np.mean(equiv_diameters) if equiv_diameters else 0})

    if len(centroids) >= 3:
        pts = np.array(centroids)
        nn_dists = KDTree(pts).query(pts, k=2)[0][:, 1]
        features.update({'mean_nn_distance': np.mean(nn_dists), 'std_nn_distance': np.std(nn_dists), 'clustering_index': np.std(nn_dists) / np.mean(nn_dists) if np.mean(nn_dists) > 0 else 0, 'ripleys_ratio': np.mean(nn_dists) / (0.5 * np.sqrt(gray_img.size / len(centroids)))})
    else: features.update({'mean_nn_distance': -1, 'std_nn_distance': -1, 'clustering_index': -1, 'ripleys_ratio': -1})

    halo_ratios = []
    h_img, w_img = gray_img.shape[:2]
    for cx, cy in centroids:
        inner_vals, outer_vals = [], []
        for angle in np.linspace(0, 2 * np.pi, 24, endpoint=False):
            for r, bucket in [(3, inner_vals), (7, outer_vals)]:
                px, py = int(cx + r * np.cos(angle)), int(cy + r * np.sin(angle))
                if 0 <= px < w_img and 0 <= py < h_img: bucket.append(float(gray_img[py, px]))
        if inner_vals and outer_vals:
            halo_mean = np.mean(outer_vals)
            halo_ratios.append(np.mean(inner_vals) / halo_mean if halo_mean > 0 else 1.0)

    features.update({'mean_halo_ratio': np.mean(halo_ratios) if halo_ratios else 1.0, 'std_halo_ratio': np.std(halo_ratios) if halo_ratios else 0.0, 'skew_halo_ratio': skew(halo_ratios) if len(halo_ratios) > 2 else 0.0})
    return features

def classify_species(morph_features, pipeline_stats):
    score, evidence = 0.0, []
    ppi, cv, d, p = pipeline_stats.get('particles_per_image', 0), pipeline_stats.get('fluor_cv', 0), pipeline_stats.get('cohens_d', 0), pipeline_stats.get('p_value', 1)

    if ppi < 2.5: score += 0.25; evidence.append(f"Low density ({ppi:.1f}/img) → Klebsiella +0.25")
    elif ppi > 4.0: score -= 0.25; evidence.append(f"High density ({ppi:.1f}/img) → Proteus -0.25")

    if cv > 0.75: score += 0.20; evidence.append(f"High CV ({cv:.2f}) → Klebsiella +0.20")
    elif cv < 0.55: score -= 0.20; evidence.append(f"Low CV ({cv:.2f}) → Proteus -0.20")

    if d > 1.0: score += 0.15; evidence.append(f"Large d ({d:.3f}) → Klebsiella +0.15")
    elif d < 0.8: score -= 0.10; evidence.append(f"Small d ({d:.3f}) → Proteus -0.10")

    if p < 0.05: score += 0.15; evidence.append(f"Significant p ({p:.4f}) on G+ → Klebsiella +0.15")

    circ = morph_features.get('mean_circularity', -1)
    if circ > 0.75: score += 0.15; evidence.append(f"High circularity ({circ:.2f}) → Klebsiella +0.15")
    elif 0 < circ < 0.60: score -= 0.15; evidence.append(f"Low circularity ({circ:.2f}) → Proteus -0.15")

    halo = morph_features.get('mean_halo_ratio', -1)
    if halo > 1.25: score += 0.15; evidence.append(f"Core-halo gradient ({halo:.2f}) → Klebsiella +0.15")
    elif halo < 1.05: score -= 0.10; evidence.append(f"Flat radial profile ({halo:.2f}) → Proteus -0.10")

    ripley = morph_features.get('ripleys_ratio', -1)
    if ripley > 1.2: score += 0.10; evidence.append(f"Dispersed spatial pattern ({ripley:.2f}) → Klebsiella +0.10")
    elif 0 < ripley < 0.8: score -= 0.10; evidence.append(f"Clustered spatial pattern ({ripley:.2f}) → Proteus -0.10")

    score = max(-1.0, min(1.0, score))
    kp_prob, pm_prob = (score + 1) / 2, 1 - ((score + 1) / 2)

    if kp_prob > 0.65: call, confidence = "Klebsiella pneumoniae", kp_prob
    elif pm_prob > 0.65: call, confidence = "Proteus mirabilis", pm_prob
    else: call, confidence = "INDETERMINATE", max(kp_prob, pm_prob)

    return {'classification': call, 'kp_probability': round(kp_prob, 3), 'pm_probability': round(pm_prob, 3), 'confidence': round(confidence * 100, 1), 'raw_score': round(score, 3), 'evidence_chain': evidence}

def classify_groups_clinical(output_root: Path, microgel_type: str = "negative", threshold_pct: float = 0.05) -> pd.DataFrame:
    control_folder = next((f for f in output_root.iterdir() if f.is_dir() and f.name.lower().startswith("control")), None)
    if not control_folder: return pd.DataFrame()
    control_master = control_folder / f"{control_folder.name}_master.xlsx"
    if not control_master.exists(): return pd.DataFrame()

    try:
        control_values = pd.to_numeric(pd.read_excel(control_master, sheet_name=f"{control_folder.name}_Typical_Particles")["Fluor_Density_per_BF_Area"], errors='coerce').dropna()
        control_mean, control_std = float(control_values.mean()), float(control_values.std(ddof=1))
    except Exception: return pd.DataFrame()

    threshold = control_mean * (1 - threshold_pct)
    results = []
    ctrl_n = len(control_values)
    ctrl_sem = (control_std / np.sqrt(ctrl_n)) if ctrl_n >= 2 else 0.0
    ctrl_ci_lo, ctrl_ci_hi = (control_mean - float(scipy_stats.t.ppf(0.975, df=ctrl_n - 1)) * ctrl_sem, control_mean + float(scipy_stats.t.ppf(0.975, df=ctrl_n - 1)) * ctrl_sem) if ctrl_n >= 2 else (control_mean, control_mean)

    results.append({'Group': 'Control', 'N': ctrl_n, 'Mean': round(control_mean, 2), 'Std_Dev': round(control_std, 2), 'SEM': round(ctrl_sem, 2), 'Median': round(float(control_values.median()) if ctrl_n > 0 else 0.0, 2), 'CI_Lower': round(ctrl_ci_lo, 2), 'CI_Upper': round(ctrl_ci_hi, 2), 'Control_Mean': round(control_mean, 2), 'Threshold': round(threshold, 2), 'Diff_from_Threshold': 0.0, 'Diff_from_Control': 0.0, 'Pct_Diff_from_Control': 0.0, 'Cohens_d': 0.0, 'd_CI_Lower': 0.0, 'd_CI_Upper': 0.0, 'Effect_Size_Label': '—', 'P_Value': np.nan, 'Significance': '—', 'Classification': 'CONTROL/Reference', 'Classification_Confidence': '—', 'Strength_of_Evidence': '—'})

    for excel_path in sorted(output_root.glob("*/*_master.xlsx")):
        group_name = excel_path.parent.name
        if group_name.lower().startswith("control"): continue
        try:
            values = pd.to_numeric(pd.read_excel(excel_path, sheet_name=f"{group_name}_Typical_Particles")["Fluor_Density_per_BF_Area"], errors='coerce').dropna()
            if values.empty: continue
        except Exception: continue

        mean_val, std_val, n = float(values.mean()), float(values.std(ddof=1)), len(values)
        classification = f"{'NEGATIVE' if microgel_type.lower() == 'negative' else 'POSITIVE'} / Bacteria Detected" if mean_val < threshold else f"{'POSITIVE' if microgel_type.lower() == 'negative' else 'NEGATIVE'} / No obvious bacteria"
        stats = _compute_group_vs_control_stats(values, control_values, threshold)
        d_val = stats['Cohens_d']

        if n + ctrl_n >= 4:
            d_se_term = (n + ctrl_n) / (n * ctrl_n) + (d_val ** 2) / (2.0 * (n + ctrl_n))
            d_se = math.sqrt(d_se_term)
            d_lo = round(d_val - 1.96 * d_se, 3)
            d_hi = round(d_val + 1.96 * d_se, 3)
        else:
            d_lo, d_hi = d_val, d_val

        sem_val = cast(float, values.sem()) if n >= 2 else 0.0
        median_val = cast(float, values.median()) if n > 0 else 0.0

        results.append({'Group': group_name, 'N': n, 'Mean': round(mean_val, 2), 'Std_Dev': round(std_val, 2), 'SEM': round(sem_val, 2), 'Median': round(median_val, 2), 'CI_Lower': stats['CI_Lower'], 'CI_Upper': stats['CI_Upper'], 'Control_Mean': round(control_mean, 2), 'Threshold': round(threshold, 2), 'Diff_from_Threshold': round(mean_val - threshold, 2), 'Diff_from_Control': round(mean_val - control_mean, 2), 'Pct_Diff_from_Control': round((mean_val - control_mean) / control_mean * 100, 1) if control_mean != 0 else 0, 'Cohens_d': d_val, 'd_CI_Lower': d_lo, 'd_CI_Upper': d_hi, 'Effect_Size_Label': "Negligible" if abs(d_val) < 0.2 else "Small" if abs(d_val) < 0.5 else "Medium" if abs(d_val) < 0.8 else "Large", 'P_Value': stats['P_Value'], 'Significance': stats['Significance'], 'Classification': classification, 'Classification_Confidence': stats['Classification_Confidence'], 'Strength_of_Evidence': "Insufficient (n<2)" if pd.isna(stats['P_Value']) else "Strong" if stats['P_Value'] < 0.01 and abs(d_val) >= 0.8 and n >= 5 else "Moderate" if stats['P_Value'] < 0.05 and abs(d_val) >= 0.5 and n >= 3 else "Weak" if stats['P_Value'] < 0.05 or abs(d_val) >= 0.5 else "Insufficient"})

    if not results: return pd.DataFrame()
    results_df = pd.DataFrame(results)
    results_df['sort_key'] = results_df['Group'].apply(lambda x: (1, 999) if x == 'Control' else (0, int(x) if x.isdigit() else 999))
    return results_df.sort_values('sort_key').drop('sort_key', axis=1)

def export_clinical_classification(output_root: Path, classification_df: pd.DataFrame, microgel_type: str = "negative") -> Optional[Path]:
    if classification_df.empty: return None
    csv_path = output_root / f"clinical_classification_{microgel_type}.csv"
    classification_df.to_csv(csv_path, index=False)
    try:
        with pd.ExcelWriter(output_root / f"clinical_classification_{microgel_type}.xlsx", engine='openpyxl') as writer:
            classification_df.to_excel(writer, sheet_name='Classification', index=False)
            ws = writer.sheets['Classification']
            header_fill, safe_fill, warning_fill = PatternFill("solid", fgColor="4472C4"), PatternFill("solid", fgColor="C6EFCE"), PatternFill("solid", fgColor="FFC7CE")
            for cell in ws[1]: cell.fill, cell.font, cell.alignment = header_fill, Font(bold=True, color="FFFFFF"), Alignment(horizontal="center", vertical="center")
            for r_idx in range(len(classification_df)):
                cls_val = classification_df.iloc[r_idx]['Classification']
                fill = safe_fill if "No obvious bacteria" in cls_val or ("NEGATIVE" not in cls_val and "POSITIVE" not in cls_val) else warning_fill
                for c_idx in range(1, len(classification_df.columns) + 1):
                    ws.cell(row=r_idx + 2, column=c_idx).fill = fill
                    ws.cell(row=r_idx + 2, column=c_idx).alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = min((max((len(str(c.value)) for c in col if c.value), default=0) + 2) * 1.1, 50)
    except Exception: pass
    return csv_path


# ==================================================
# ReportLab Individual Sample Report Helpers
# ==================================================

def plot_gm_typical_particles_comparison(
    group_means:     list,
    group_sds:       list,
    group_ns:        list,
    scatter_values:  list,
    control_mean:    float,
    threshold_pct:   float  = -0.05,
    significance:    list | None  = None,
    title:           str | None   = None,        # ← now truly optional; built dynamically below
    group_labels:    list[str] | None = None,    # ← NEW: e.g. ["1","2","3","Control"]
    channel_label:   str    = "G\u2212 Microgel (Gram-Negative)",
    y_label:         str    = "Fluorescence Density (a.u./\u03bcm\u00b2)",
    figsize:         tuple  = (7.2, 4.8),
    dpi:             int    = 110,
    highlight_group: int  | None  = None,
    output_path:     str  | None  = None,
) -> "io.BytesIO":
    import io as _io
    if significance is None:
        significance = ["ns"] * max(len(group_means) - 1, 0)

    threshold_val = control_mean * (1 + threshold_pct)
    n_groups = len(group_means)
    x_pos    = np.arange(n_groups)

    # ── dynamic group labels ────────────────────────────────────────────
    if group_labels is None:
        group_labels = [str(i) for i in range(1, n_groups)] + ["Control"]
    labels = [f"{g}\nn={group_ns[i]}" for i, g in enumerate(group_labels)]

    # ── dynamic title ───────────────────────────────────────────────────
    if title is None:
        title = (
            f"{channel_label} \u2014 "
            "Typical Particles: Middle 40% (Excluded top/bottom 30%)"
        )

    fig, ax = plt.subplots(figsize=figsize, dpi=dpi)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    bar_width = 0.55
    bars = ax.bar(
        x_pos, group_means,
        width=bar_width,
        color=BAR_COLORS[:n_groups],
        edgecolor=PALETTE["edge_default"],
        linewidth=0.8,
        zorder=2,
        alpha=0.88,
    )

    if highlight_group is not None and 0 <= highlight_group < n_groups:
        bars[highlight_group].set_edgecolor(BAR_COLORS[highlight_group % len(BAR_COLORS)])
        bars[highlight_group].set_linewidth(2.5)

    ax.errorbar(
        x_pos, group_means, yerr=group_sds,
        fmt="none", ecolor="#333333",
        elinewidth=1.2, capsize=5, capthick=1.2, zorder=4,
    )

    # ── scatter dots — match global seaborn stripplot style ─────────────
    rng = np.random.default_rng(42)
    for i, pts in enumerate(scatter_values):
        if not pts:
            continue
        
        pts_arr = np.asarray(list(pts) if not isinstance(pts, np.ndarray) else pts, dtype=float).ravel()
        pts_arr = pts_arr[np.isfinite(pts_arr)]

        if pts_arr.size == 0:
            continue        
        
        jitter = rng.uniform(-0.15, 0.15, pts_arr.size)
        ax.scatter(
            np.full(pts_arr.size, x_pos[i], dtype=float) + jitter,
            pts_arr,
            c="#55EAF2",
            s=52,
            zorder=6,
            edgecolors="black",
            linewidths=0.6,
            alpha=0.95,
            marker="o",
            clip_on=False,
        )

    ax.axhline(control_mean, color=PALETTE["ctrl_mean_line"], linewidth=1.4,
               linestyle=(0, (4, 3)), zorder=3,
               label=f"Control Mean ({control_mean:,.1f})")
    ax.axhline(threshold_val, color=PALETTE["threshold_line"], linewidth=1.4,
               linestyle=(0, (6, 3)), zorder=3,
               label=f"Lower Threshold ({threshold_pct*100:.0f}%: {threshold_val:,.1f})")

    _draw_significance_brackets(ax, x_pos, group_means, group_sds,
                                significance, n_groups - 1)

    y_top = max(group_means) + max(group_sds) * 1.55 + max(group_means) * 0.08
    ax.set_ylim(0, y_top)
    ax.set_xlim(-0.55, n_groups - 0.45)
    ax.set_xticks(x_pos)
    ax.set_xticklabels(labels, fontsize=9)
    ax.set_ylabel(y_label, fontsize=9, labelpad=6)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax.tick_params(axis="both", labelsize=8)
    ax.spines[["top", "right"]].set_visible(False)
    ax.spines[["left", "bottom"]].set_color("#AAAAAA")
    ax.yaxis.grid(True, color="#E5E5E5", linewidth=0.6, zorder=0)
    ax.set_axisbelow(True)
    ax.set_title(title, fontsize=9.5, color=PALETTE["text_dark"],
                 pad=10, fontweight="bold")

    legend_handles = [
        Line2D([0], [0], color=PALETTE["ctrl_mean_line"], lw=1.4,
               linestyle=(0, (4, 3)),
               label=f"Control Mean ({control_mean:,.1f})"),
        Line2D([0], [0], color=PALETTE["threshold_line"], lw=1.4,
               linestyle=(0, (6, 3)),
               label=f"Lower Threshold ({threshold_pct*100:.0f}%: {threshold_val:,.1f})"),
    ]
    ax.legend(handles=legend_handles, loc="upper right", fontsize=7.5,
              framealpha=0.85, edgecolor="#CCCCCC", handlelength=2.2)

    plt.tight_layout(pad=0.8)
    buf = _io.BytesIO()

    fig.savefig(
            buf,
            format="jpg",
            dpi=dpi,
            bbox_inches="tight",
            facecolor="white",
            pil_kwargs={"quality": 62, "optimize": True}
        )
    
    plt.close(fig)
    buf.seek(0)

    if output_path:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        with open(output_path, "wb") as fh:
            fh.write(buf.read())
        buf.seek(0)

    return buf



def _draw_significance_brackets(ax, x_pos, means, sds, sig_labels, ctrl_idx):
    """Draw comparison brackets from each test group to the control bar."""
    ctrl_top = means[ctrl_idx] + sds[ctrl_idx]
    y_step   = (ax.get_ylim()[1] - max(means)) * 0.07

    for i, label in enumerate(sig_labels):
        if i >= ctrl_idx:
            break
        g_top  = means[i] + sds[i]
        y_base = max(ctrl_top, g_top) + y_step * (i + 0.6)

        ax.plot([x_pos[i], x_pos[ctrl_idx]], [y_base, y_base],
                color=SIG_COLORS.get(label, "#555555"), linewidth=0.9, zorder=6)
        for xi in [x_pos[i], x_pos[ctrl_idx]]:
            ax.plot([xi, xi], [y_base - y_step * 0.25, y_base],
                    color=SIG_COLORS.get(label, "#555555"), linewidth=0.9, zorder=6)
        ax.text(
            (x_pos[i] + x_pos[ctrl_idx]) / 2, y_base + y_step * 0.1,
            label, ha="center", va="bottom", fontsize=8,
            color=SIG_COLORS.get(label, "#555555"),
            fontweight="bold" if label != "ns" else "normal", zorder=7,
        )


def _rl_base_styles() -> dict:
    """Return a dict of named ReportLab ParagraphStyles."""
    if not REPORTLAB_AVAILABLE:
        return {}
    ss = getSampleStyleSheet()
    return {
        "section_head": ParagraphStyle(
            "SectionHead", parent=ss["Heading2"],
            fontSize=11, textColor=rl_colors.HexColor(PALETTE["header_bg"]),
            spaceAfter=4, spaceBefore=10, fontName="Helvetica-Bold",
        ),
        "body": ParagraphStyle(
            "Body", parent=ss["Normal"],
            fontSize=8.5, leading=12,
            textColor=rl_colors.HexColor(PALETTE["text_dark"]),
        ),
        "caption": ParagraphStyle(
            "Caption", parent=ss["Normal"],
            fontSize=7.5, leading=10, textColor=rl_colors.grey,
            alignment=TA_CENTER,
        ),
        "table_head": ParagraphStyle(
            "TableHead", parent=ss["Normal"],
            fontSize=8, fontName="Helvetica-Bold",
            textColor=rl_colors.white, alignment=TA_CENTER,
        ),
        "table_cell": ParagraphStyle(
            "TableCell", parent=ss["Normal"],
            fontSize=7.5, leading=10, alignment=TA_CENTER,
        ),
        "table_cell_left": ParagraphStyle(
            "TableCellL", parent=ss["Normal"],
            fontSize=7.5, leading=10, alignment=TA_LEFT,
        ),
        "status_green": ParagraphStyle(
            "StatusGreen", parent=ss["Normal"],
            fontSize=8, fontName="Helvetica-Bold",
            textColor=rl_colors.HexColor(PALETTE["detected_green"]),
            alignment=TA_CENTER,
        ),
    }


def _rl_section_divider():
    """Thin horizontal rule for use between ReportLab report sections."""
    if not REPORTLAB_AVAILABLE:
        return None
    return HRFlowable(
        width="100%", thickness=1,
        color=rl_colors.HexColor("#CCCCCC"),
        spaceBefore=4, spaceAfter=8,
    )


def _find_all_image_pairs(
    group_dir: Path,
    max_pairs: Optional[int] = 2,
) -> list[tuple[Optional[Path], Optional[Path]]]:
    pairs: list[tuple[Optional[Path], Optional[Path]]] = []
    for img_subdir in sorted(group_dir.iterdir()):
        if not img_subdir.is_dir():
            continue
        c = img_subdir / "11_contours_rejected_orange_accepted_yellow_ids_green.png"
        f = img_subdir / "24_bf_fluor_matching_overlay_ids.png"
        if c.exists() or f.exists():
            pairs.append((c if c.exists() else None, f if f.exists() else None))
            if max_pairs is not None and len(pairs) >= max_pairs:
                break
    return pairs


def _find_representative_images(
    group_dir: Path,
) -> tuple[Optional[Path], Optional[Path]]:
    """Backward-compat shim — returns only the first pair."""
    pairs = _find_all_image_pairs(group_dir)
    return pairs[0] if pairs else (None, None)


def build_sample_images_section(
    path_contour_overlay:  str,          # kept for backward compat (ignored when all_pairs given)
    path_bf_fluor_overlay: str,
    styles:                dict | None   = None,
    image_height:          float | None  = None,
    all_image_pairs:       list | None   = None,   # NEW: list of (contour_path, fluor_path)
) -> list:
    """
    Render all (contour-map, BF/fluor-overlay) image pairs side-by-side.
    When all_image_pairs is supplied it is used; otherwise the single
    path_contour_overlay / path_bf_fluor_overlay pair is shown.
    Returns [] when ReportLab is not installed.
    """
    if not REPORTLAB_AVAILABLE:
        return []

    if image_height is None:
        image_height = 6.5 * rl_cm
    if styles is None:
        styles = _rl_base_styles()

    # Normalise to a list of (str-or-None, str-or-None)
    if all_image_pairs:
        pairs: list[tuple] = [
            (str(c) if c else None, str(f) if f else None)
            for c, f in all_image_pairs
        ]
    else:
        pairs = [(path_contour_overlay or None, path_bf_fluor_overlay or None)]

    half_w = (_RL_CONTENT_W - 0.5 * rl_cm) / 2
    story: list = []

    story.append(Spacer(1, 0.3 * rl_cm))
    story.append(_rl_section_divider())
    story.append(Paragraph(
    f"Representative Sample Overlay Images  ({len(pairs)} shown)",
        styles["section_head"],
    ))
    story.append(Spacer(1, 0.2 * rl_cm))

    def _safe_image(path, w, h):
        if path and os.path.isfile(str(path)):
            compressed = _compress_image_for_pdf(
                path,
                max_width_px=1200,
                max_height_px=1200,
                jpeg_quality=62,
                convert_to_jpeg=True,
            )
            return RLImage(compressed, width=w, height=h, kind="proportional")
        return RLImage(
            _rl_placeholder_image(
                int(w), int(h),
                label=os.path.basename(str(path)) if path else "Image not found"
            ),
            width=w, height=h,
        )

    for idx, (c_path, f_path) in enumerate(pairs, 1):
        # derive image sub-directory name for the caption
        subdir = ""
        for p in (c_path, f_path):
            if p:
                subdir = Path(str(p)).parent.name
                break

        img_contour  = _safe_image(c_path, half_w, image_height)
        img_bf_fluor = _safe_image(f_path, half_w, image_height)

        cap_contour = Paragraph(
            f"<b>Image {idx}{(' \u00b7 ' + subdir) if subdir else ''} \u00b7 Contour Map</b><br/>"
            "Rejected: <font color='#FF6600'><b>orange</b></font> \u00b7 "
            "Accepted: <font color='#CCCC00'><b>yellow</b></font> \u00b7 "
            "IDs: <font color='#00AA00'><b>green</b></font>",
            styles["caption"],
        )
        cap_bf_fluor = Paragraph(
            f"<b>Image {idx}{(' \u00b7 ' + subdir) if subdir else ''} \u00b7 BF / Fluorescence</b><br/>"
            "Brightfield (grey) co-registered with fluorescence (green); "
            "particle IDs labelled",
            styles["caption"],
        )

        image_table = Table(
            [[img_contour,  img_bf_fluor],
             [cap_contour,  cap_bf_fluor]],
            colWidths=[half_w, half_w],
            hAlign="LEFT",
        )
        image_table.setStyle(TableStyle([
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",        (0, 0), (-1,  0), "MIDDLE"),
            ("VALIGN",        (0, 1), (-1,  1), "TOP"),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 4),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
            ("LINEBELOW",     (0, 0), (-1,  0), 0.5,
             rl_colors.HexColor("#DDDDDD")),
        ]))
        story.append(KeepTogether([image_table]))
        if idx < len(pairs):
            story.append(Spacer(1, 0.15 * rl_cm))

    story.append(Spacer(1, 0.3 * rl_cm))
    return story





def _rl_placeholder_image(width_px: int, height_px: int,
                           label: str = "Image not found") -> "io.BytesIO":
    """Return a grey placeholder PNG as BytesIO when an image file is missing."""
    import io as _io
    fig, ax = plt.subplots(figsize=(width_px / 96, height_px / 96), dpi=96)
    ax.set_facecolor("#EEEEEE")
    ax.text(0.5, 0.5, label, ha="center", va="center",
            fontsize=7, color="#888888", wrap=True, transform=ax.transAxes)
    ax.axis("off")
    buf = _io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def build_typical_particles_section(
    typical_particles: list,
    group_summary:     dict,
    percentile_label:  str  = "Middle 40%",
    excluded_label:    str  = "Excluded top/bottom 30%",
    styles:            dict | None = None,
) -> list:
    """
    Returns ReportLab flowables for the Typical Particles section:
    a per-particle table ranked by Fluor Density, plus a summary stats box.
    Returns [] if ReportLab is not installed.

    typical_particles keys:
        object_id, source_image, bf_area_um2, equiv_diameter_um,
        circularity, fluor_mean, fluor_density_per_bf_area,
        bf_to_fluor_area_ratio, status
    """
    if not REPORTLAB_AVAILABLE:
        return []
    if styles is None:
        styles = _rl_base_styles()

    story = []
    story.append(Spacer(1, 0.3 * rl_cm))
    story.append(_rl_section_divider())
    story.append(Paragraph(
        f"Typical Particles Summary \u2014 {percentile_label} ({excluded_label})",
        styles["section_head"],
    ))
    story.append(Spacer(1, 0.15 * rl_cm))

    sorted_particles = sorted(
        typical_particles,
        key=lambda r: r.get("fluor_density_per_bf_area", 0),
        reverse=True,
    )

    col_widths_raw = [0.6, 3.2, 2.5, 1.5, 1.4, 1.3, 1.5, 2.5, 2.2, 1.4]
    scale = _RL_CONTENT_W / (sum(col_widths_raw) * rl_cm)
    col_widths = [w * rl_cm * scale for w in col_widths_raw]

    def _h(txt):  return Paragraph(txt, styles["table_head"])
    def _c(txt):  return Paragraph(str(txt), styles["table_cell"])
    def _cl(txt): return Paragraph(str(txt), styles["table_cell_left"])

    header_row = [
        _h("Rank"), _h("Object ID"), _h("Source Image"),
        _h("BF Area\n(\u03bcm\u00b2)"), _h("Eq. Diam.\n(\u03bcm)"),
        _h("Circularity"), _h("Fluor Mean\n(a.u.)"),
        _h("Fluor Density\n(a.u./\u03bcm\u00b2) \u2605"),
        _h("BF:Fluor\nArea Ratio"), _h("Status"),
    ]

    table_data = [header_row]
    for rank, p in enumerate(sorted_particles, 1):
        status_str = "\u2705" if str(p.get("status", "")).lower() == "accepted" else "\u274c"
        row = [
            _c(rank),
            _cl(p.get("object_id", "")),
            _cl(p.get("source_image", "")),
            _c(f"{p.get('bf_area_um2', 0):.2f}"),
            _c(f"{p.get('equiv_diameter_um', 0):.3f}"),
            _c(f"{p.get('circularity', 0):.4f}"),
            _c(f"{p.get('fluor_mean', 0):.2f}"),
            _c(f"{p.get('fluor_density_per_bf_area', 0):,.2f}"),
            _c(f"{p.get('bf_to_fluor_area_ratio', 0):.4f}"),
            Paragraph(status_str, styles["status_green"]
                      if status_str == "\u2705" else styles["table_cell"]),
        ]
        table_data.append(row)

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    header_color = rl_colors.HexColor(PALETTE["header_bg"])
    alt_color    = rl_colors.HexColor(PALETTE["row_alt"])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  header_color),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  rl_colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 7.5),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [rl_colors.white, alt_color]),
        ("GRID",          (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#CCCCCC")),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 3),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
        ("BACKGROUND",    (7, 1), (7, -1),  rl_colors.HexColor("#EAF4FF")),
        ("FONTNAME",      (7, 1), (7, -1),  "Helvetica-Bold"),
    ]))
    story.append(KeepTogether([tbl]))
    story.append(Spacer(1, 0.25 * rl_cm))

    story += _rl_build_summary_stats_box(group_summary, styles)
    story.append(Spacer(1, 0.3 * rl_cm))
    return story


def _rl_build_summary_stats_box(gs: dict, styles: dict) -> list:
    """Two-column summary-stats mini-table used below the particle table."""
    sig    = gs.get("sig", "ns")
    status = gs.get("status", "Unknown")
    pct    = gs.get("pct_vs_ctrl", 0.0)
    pct_str = f"{pct:+.1f}%"

    def _kv(k, v):
        return [
            Paragraph(f"<b>{k}</b>", styles["table_cell_left"]),
            Paragraph(str(v),         styles["table_cell"]),
        ]

    rows = [
        _kv("N (typical particles)",  gs.get("n", "\u2014")),
        _kv("Mean \u00b1 SD (a.u./\u03bcm\u00b2)",
            f"{gs.get('mean', 0):,.1f} \u00b1 {gs.get('sd', 0):,.1f}"),
        _kv("95% CI",
            f"[{gs.get('ci_low', 0):,.1f} \u2014 {gs.get('ci_high', 0):,.1f}]"),
        _kv("% vs. Control",   pct_str),
        _kv("Cohen\u2019s d",  f"{gs.get('cohens_d', 0):.3f}"),
        _kv("p-value",         f"{gs.get('p_value', 1.0):.4f}"),
        _kv("Significance",    sig),
        _kv("Confidence",      gs.get("confidence", "\u2014")),
        _kv("Detection Status", status),
    ]

    stat_tbl = Table(rows, colWidths=[5.0 * rl_cm, 5.0 * rl_cm], hAlign="LEFT")
    status_row_idx = 8
    status_color = (rl_colors.HexColor("#D4EDDA")
                    if "DETECTED" in status.upper()
                    else rl_colors.HexColor("#F8D7DA"))
    stat_tbl.setStyle(TableStyle([
        ("GRID",      (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#CCCCCC")),
        ("BACKGROUND",(0, 0), (0, -1),  rl_colors.HexColor("#EFF3F8")),
        ("FONTSIZE",  (0, 0), (-1, -1), 8),
        ("VALIGN",    (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
        ("BACKGROUND",
         (0, status_row_idx), (-1, status_row_idx), status_color),
    ]))

    return [
        Paragraph("Group Statistics", _rl_base_styles()["section_head"]),
        Spacer(1, 0.1 * rl_cm),
        KeepTogether([stat_tbl]),
    ]


def build_comparison_plot_section(
    group_means:     list,
    group_sds:       list,
    group_ns:        list,
    scatter_values:  list,
    control_mean:    float,
    threshold_pct:   float = -0.05,
    significance:    list | None = None,
    highlight_group: int | None = None,
    title:           str | None  = None,
    channel_label:   str   = "G\u2212 Microgel (Gram-Negative)",
    group_labels:    list[str] | None = None,  
    styles:          dict | None = None,
    plot_height:     float | None = None,
) -> list:
    """
    Generates the comparison plot PNG and returns it wrapped in ReportLab
    flowables ready to append to a story.
    Returns [] if ReportLab is not installed.
    """
    if not REPORTLAB_AVAILABLE:
        return []
    if plot_height is None:
        plot_height = 8.5 * rl_cm
    if styles is None:
        styles = _rl_base_styles()
    if title is None:
        title = (
            "K. pneumoniae Negative \u2014 G\u2212 Microgel \u2014 "
            "Typical Particles: Middle 40% (Excluded top/bottom 30%)"
        )

    buf = plot_gm_typical_particles_comparison(
        group_means=group_means, group_sds=group_sds,
        group_ns=group_ns, scatter_values=scatter_values,
        control_mean=control_mean, threshold_pct=threshold_pct,
        significance=significance, title=title,
        channel_label=channel_label, highlight_group=highlight_group,
        group_labels=group_labels,
    )

    plot_img = RLImage(buf, width=_RL_CONTENT_W, height=plot_height)

    story = []
    story.append(Spacer(1, 0.3 * rl_cm))
    story.append(_rl_section_divider())
    story.append(Paragraph(
        f"Sample Comparison \u2014 {channel_label}",
        styles["section_head"],
    ))
    story.append(Spacer(1, 0.15 * rl_cm))
    story.append(KeepTogether([plot_img]))
    story.append(Paragraph(
        "\u2605 Primary metric.  Error bars = \u00b11 SD.  "
        "Significance vs. Control: "
        "*** p<0.001 \u00b7 ** p<0.01 \u00b7 * p<0.05 \u00b7 ns p\u22650.05.",
        styles["caption"],
    ))
    story.append(Spacer(1, 0.3 * rl_cm))
    return story


def append_individual_sample_additions(
    existing_story:        list,
    path_contour_overlay:  str,
    path_bf_fluor_overlay: str,
    typical_particles:     list,
    group_summary:         dict,
    all_group_means:       list,
    all_group_sds:         list,
    all_group_ns:          list,
    all_scatter_values:    list,
    control_mean:          float,
    threshold_pct:         float  = -0.05,
    significance:          list | None  = None,
    highlight_group:       int | None   = 2,
    percentile_label:      str    = "Middle 40%",
    excluded_label:        str    = "Excluded top/bottom 30%",
    channel_label:         str    = "G\u2212 Microgel (Gram-Negative)",
) -> list:
    """
    Master orchestrator that appends three addition sections to an existing
    ReportLab story list:

      A  Sample Overlay Images
      B  Typical Particles Summary Table + Stats
      C  Sample Comparison Plot

    Parameters mirror save_individual_sample_report().
    Returns the extended story (same object as existing_story).
    """
    if not REPORTLAB_AVAILABLE:
        return existing_story

    styles = _rl_base_styles()

    existing_story += build_sample_images_section(
        path_contour_overlay=path_contour_overlay,
        path_bf_fluor_overlay=path_bf_fluor_overlay,
        styles=styles,
    )

    existing_story += build_typical_particles_section(
        typical_particles=typical_particles,
        group_summary=group_summary,
        percentile_label=percentile_label,
        excluded_label=excluded_label,
        styles=styles,
    )

    existing_story += build_comparison_plot_section(
        group_means=all_group_means,
        group_sds=all_group_sds,
        group_ns=all_group_ns,
        scatter_values=all_scatter_values,
        control_mean=control_mean,
        threshold_pct=threshold_pct,
        significance=significance,
        highlight_group=highlight_group if highlight_group is not None else 0,
        channel_label=channel_label,
        styles=styles,
    )

    return existing_story


def is_page5_empty(page5_data: dict) -> bool:
    """
    Return True when every Methodology/QC/Limitations/Approval field
    in page5_data is blank / None / empty string.
    """
    if not page5_data:
        return True
    return all(
        not str(v).strip()
        for k, v in page5_data.items()
        if k.lower().replace(" ", "_") in _EMPTY_PAGE5_SENTINEL_KEYS
        or "method" in k.lower()
        or "quality" in k.lower()
        or "limit"  in k.lower()
        or "approv" in k.lower()
        or "perform" in k.lower()
        or "review" in k.lower()
    )


def strip_empty_page5_from_story(story: list,
                                  page5_marker: str = "__PAGE5_START__") -> list:
    """
    Post-build cleanup: scans a pre-assembled ReportLab story backwards and
    removes the last PageBreak + its tail if every flowable in the tail is
    empty (Spacer, HRFlowable, or blank Paragraph).
    """
    if not REPORTLAB_AVAILABLE:
        return list(story)

    last_pb_idx = None
    for i in range(len(story) - 1, -1, -1):
        if isinstance(story[i], PageBreak):
            last_pb_idx = i
            break

    if last_pb_idx is None:
        return list(story)

    tail = story[last_pb_idx + 1:]

    def _is_empty(f):
        if isinstance(f, (Spacer, HRFlowable)):
            return True
        if isinstance(f, Paragraph):
            clean = re.sub(r"<[^>]+>", "", f.text or "").strip()
            return clean == "" or clean.upper() in {
                "METHODOLOGY", "QUALITY CONTROL", "LIMITATIONS", "APPROVAL",
            }
        return False

    if tail and all(_is_empty(f) for f in tail):
        return list(story[:last_pb_idx])

    return list(story)


# ==================================================
# Per-Sample Reporting  —  Data Collection Helpers
# ==================================================

def _collect_typical_particles_for_group(
    group_dir: Path,
    group_name: str,
    control_mean: float = 0.0,
    control_values: Optional[pd.Series] = None,
    threshold_pct: float = 0.05,
) -> tuple[list, dict]:
    """
    Reads the Typical_Particles sheet from a group master Excel and returns:
      (typical_particles_list, group_summary_dict)

    group_summary_dict keys:
        mean, sd, n, ci_low, ci_high, pct_vs_ctrl,
        cohens_d, p_value, sig, confidence, status
    """
    excel_path = group_dir / f"{group_name}_master.xlsx"
    if not excel_path.exists():
        return [], {}

    try:
        df = pd.read_excel(excel_path,
                           sheet_name=f"{group_name}_Typical_Particles")
    except Exception:
        return [], {}

    particles = []
    for _, row in df.iterrows():
        particles.append({
            'object_id':                str(row.get('Object_ID', '')),
            'source_image':             str(row.get('Source_Image', '')),
            'bf_area_um2':             float(row.get('BF_Area_um2', 0) or 0),
            'equiv_diameter_um':        float(row.get('EquivDiameter_um', 0) or 0),
            'circularity':             float(row.get('Circularity', 0) or 0),
            'fluor_mean':              float(row.get('Fluor_Mean', 0) or 0),
            'fluor_density_per_bf_area':
                float(row.get('Fluor_Density_per_BF_Area', 0) or 0),
            'bf_to_fluor_area_ratio':
                float(row.get('BF_to_Fluor_Area_Ratio', 0) or 0),
            'status': str(row.get('Status', 'Accepted')),
        })

    if not particles:
        return [], {}

    densities_s = pd.to_numeric(
        df.get('Fluor_Density_per_BF_Area', pd.Series(dtype=float)),
        errors='coerce',
    ).dropna()

    n        = len(densities_s)
    mean_val = float(densities_s.mean()) if n > 0 else 0.0
    std_val  = float(densities_s.std(ddof=1)) if n > 1 else 0.0

    if n >= 2:
        t_crit  = float(scipy_stats.t.ppf(0.975, df=n - 1))
        sem     = std_val / np.sqrt(n)
        ci_low  = mean_val - t_crit * sem
        ci_high = mean_val + t_crit * sem
    else:
        ci_low = ci_high = mean_val

    pct_vs_ctrl = (
        (mean_val - control_mean) / control_mean * 100.0
        if control_mean > 0 else 0.0
    )

    cohens_d = 0.0
    p_value  = np.nan
    sig      = "N/A"
    confidence = "Low"

    if control_values is not None and len(control_values) >= 2 and n >= 2:
        ctrl_std   = float(control_values.std(ddof=1))
        pooled_std = np.sqrt(
            ((n - 1) * std_val ** 2 + (len(control_values) - 1) * ctrl_std ** 2)
            / (n + len(control_values) - 2)
        )
        cohens_d = (mean_val - control_mean) / pooled_std if pooled_std > 0 else 0.0
        try:
            _, p_raw = scipy_stats.ttest_ind(
                densities_s.values, control_values.values, equal_var=False
            )
            p_value = float(np.asarray(p_raw).item())
        except Exception:
            p_value = np.nan

        if pd.isna(p_value):     sig = "N/A"
        elif p_value < 0.001:    sig = "***"
        elif p_value < 0.01:     sig = "**"
        elif p_value < 0.05:     sig = "*"
        else:                    sig = "ns"

        threshold = control_mean * (1.0 - threshold_pct)
        if n < 3:
            confidence = "Low"
        elif (ci_high < threshold) or (ci_low > threshold):
            confidence = "High" if abs(cohens_d) >= 0.8 else "Moderate"
        else:
            confidence = "Low"

    threshold_val = control_mean * (1.0 - threshold_pct)
    detected      = (control_mean > 0) and (mean_val < threshold_val)
    status_str    = (
        "DETECTED \u2014 bacteria present"
        if detected
        else "NOT DETECTED \u2014 no obvious bacteria"
    )

    group_summary = {
        'mean':        round(mean_val, 2),
        'sd':          round(std_val, 2),
        'n':           n,
        'ci_low':      round(ci_low, 2),
        'ci_high':     round(ci_high, 2),
        'pct_vs_ctrl': round(pct_vs_ctrl, 1),
        'cohens_d':    round(cohens_d, 3),
        'p_value':     round(p_value, 4) if not pd.isna(p_value) else np.nan,
        'sig':         sig,
        'confidence':  confidence,
        'status':      status_str,
    }

    return particles, group_summary


def _collect_all_groups_comparison_data(
    output_dir: Path,
) -> dict[str, list[float]]:
    """
    Walk all group master Excels under output_dir and return a dict mapping
    display group name → list of Fluor_Density_per_BF_Area values from the
    Typical_Particles sheet.  Control groups are included as "Control".
    """
    group_data: dict[str, list[float]] = {}

    for group_dir in sorted(output_dir.iterdir()):
        if not group_dir.is_dir():
            continue
        group_name = group_dir.name
        excel_path = group_dir / f"{group_name}_master.xlsx"
        if not excel_path.exists():
            continue
        display = _display_group_name(group_name)
        try:
            df = pd.read_excel(
                excel_path,
                sheet_name=f"{group_name}_Typical_Particles",
            )
            vals = pd.to_numeric(
                df.get("Fluor_Density_per_BF_Area", pd.Series(dtype=float)),
                errors="coerce",
            ).dropna().tolist()
            if vals:
                group_data[display] = vals
        except Exception:
            continue

    return group_data


def _merge_pdfs_to_file(input_paths: list[Path], output_path: Path) -> bool:
    """
    Attempt to merge a list of PDF files into output_path using pypdf.
    Returns True on success, False if pypdf is unavailable or an error occurs.
    """
    if not PYPDF_AVAILABLE:
        return False
    try:
        if _PdfWriter is not None:
            writer = _PdfWriter()
            for p in input_paths:
                if p.exists():
                    writer.append(str(p))
            with open(output_path, "wb") as fh:
                writer.write(fh)
        else:
            # PyPDF2 legacy path
            if _PdfMerger is None:
                return False
            merger = _PdfMerger()
            for p in input_paths:
                if p.exists():
                    merger.append(str(p))
            merger.write(str(output_path))
            merger.close()
        return True
    except Exception:
        return False


def _generate_individual_sample_additions_pdf(
    reports_dir:           Path,
    group_name:            str,
    output_dir:            Path,
    control_values:        Optional[pd.Series],
    control_mean:          float,
    threshold_pct:         float,
    all_group_data:        dict[str, list[float]],
    percentile:            float,
    channel_label:         str,
    config:                dict,
) -> Optional[Path]:
    if not REPORTLAB_AVAILABLE:
        return None

    group_dir = output_dir / group_name
    if not group_dir.exists():
        return None

    typical_particles, group_summary = _collect_typical_particles_for_group(
        group_dir, group_name,
        control_mean=control_mean,
        control_values=control_values,
        threshold_pct=threshold_pct,
    )
    if not typical_particles:
        return None

    # ── all image pairs (one per source image) ───────────────────────────
    all_pairs = _find_all_image_pairs(group_dir, max_pairs=2)

    # ── comparison plot data ─────────────────────────────────────────────
    ordered_keys = sorted(
        all_group_data.keys(),
        key=lambda g: (1, 10**9) if g == "Control"
                      else (0, int(g) if g.isdigit() else 10**8),
    )

    all_means    = [float(np.mean(all_group_data[g])) for g in ordered_keys]
    all_sds      = [float(np.std(all_group_data[g], ddof=1))
                    if len(all_group_data[g]) > 1 else 0.0
                    for g in ordered_keys]
    all_ns       = [len(all_group_data[g]) for g in ordered_keys]
    all_scatters = [
        pd.Series(pd.to_numeric(pd.Series(all_group_data.get(g, [])), errors="coerce")).dropna().tolist()
        for g in ordered_keys
    ]

    ctrl_mean_for_plot = (
        float(np.mean(all_group_data["Control"]))
        if "Control" in all_group_data else control_mean
    )
    ctrl_vals_arr = (
        np.array(all_group_data["Control"])
        if "Control" in all_group_data else np.array([])
    )

    significance: list[str] = []
    ctrl_idx = ordered_keys.index("Control") if "Control" in ordered_keys else -1
    for i, g in enumerate(ordered_keys):
        if i == ctrl_idx:
            continue
        if len(ctrl_vals_arr) >= 2 and len(all_group_data[g]) >= 2:
            try:
                _, p_raw = scipy_stats.ttest_ind(
                    np.array(all_group_data[g]), ctrl_vals_arr, equal_var=False
                )
                p = float(np.asarray(p_raw).item())
                significance.append(
                    "***" if p < 0.001 else "**" if p < 0.01
                    else "*" if p < 0.05 else "ns"
                )
            except Exception:
                significance.append("ns")
        else:
            significance.append("N/A")

    highlight_group = (ordered_keys.index(group_name)
                       if group_name in ordered_keys else None)

    percentile_pct   = int(percentile * 100)
    middle_pct       = 100 - 2 * percentile_pct
    percentile_label = f"Middle {middle_pct}%"
    excluded_label   = f"Excluded top/bottom {percentile_pct}%"

    # ── dynamic plot title ───────────────────────────────────────────────
    bacteria_name = config.get('dataset_id_base', config.get('dataset_id', ''))
    plot_title = (
        f"{bacteria_name} \u2014 {channel_label} \u2014 "
        f"Typical Particles: {percentile_label} ({excluded_label})"
    )

    # ── build ReportLab story directly (for full control) ────────────────
    styles = _rl_base_styles()
    story: list = []

    # Section A — all overlay images
    story += build_sample_images_section(
        path_contour_overlay="",   # ignored; all_image_pairs takes over
        path_bf_fluor_overlay="",
        styles=styles,
        all_image_pairs=all_pairs,
    )

    # Section B — typical particles table + stats
    story += build_typical_particles_section(
        typical_particles=typical_particles,
        group_summary=group_summary,
        percentile_label=percentile_label,
        excluded_label=excluded_label,
        styles=styles,
    )

    # Section C — comparison plot with dynamic title and group labels
    story += build_comparison_plot_section(
        group_means=all_means,
        group_sds=all_sds,
        group_ns=all_ns,
        scatter_values=all_scatters,
        control_mean=ctrl_mean_for_plot,
        threshold_pct=-threshold_pct,
        significance=significance,
        highlight_group=highlight_group,
        title=plot_title,
        channel_label=channel_label,
        group_labels=ordered_keys,   # passed through to the plot function
        styles=styles,
    )

    if not story:
        return None

    additions_path = reports_dir / f"sample_{group_name}_detailed.pdf"
    try:
        doc = SimpleDocTemplate(
            str(additions_path),
            pagesize=RL_A4,
            leftMargin=_RL_MARGIN, rightMargin=_RL_MARGIN,
            topMargin=_RL_MARGIN,  bottomMargin=_RL_MARGIN,
        )
        doc.build(story)
        return additions_path
    except Exception:
        return None




# ==================================================
# Per-Sample Reporting  —  Core Functions
# ==================================================

def _collect_per_image_data(output_dir: Path) -> Dict[str, Dict[str, Dict]]:
    """
    Returns {group_name: {image_name: {fluor_density, particle_count, particles_with_fluor}}}
    Reads from the Summary sheet of each group's master Excel.
    Skips control groups.
    """
    result: Dict[str, Dict[str, Dict]] = {}
    for group_dir in sorted(output_dir.iterdir()):
        if not group_dir.is_dir():
            continue
        group_name = group_dir.name
        excel_path = group_dir / f"{group_name}_master.xlsx"
        if not excel_path.exists():
            continue
        try:
            summary_df = pd.read_excel(excel_path, sheet_name="Summary")
            images: Dict[str, Dict] = {}
            for _, row in summary_df.iterrows():
                img_name = str(row.get("Image", "")).strip()
                if img_name:
                    images[img_name] = {
                        'fluor_density':        float(row.get("Avg_Fluor_Density", 0.0) or 0.0),
                        'particle_count':       int(row.get("Total_Particles_Detected", 0) or 0),
                        'particles_with_fluor': int(row.get("Particles_With_Fluor", 0) or 0),
                    }
            if images:
                result[group_name] = images
        except Exception:
            pass
    return result


def _get_control_stats_from_dir(output_dir: Path, threshold_pct: float) -> Dict:
    """
    Returns control statistics: mean, std, n, threshold, raw values.
    Reads from the Typical_Particles sheet of the control group master Excel.
    """
    empty = {'mean': 0.0, 'std': 0.0, 'n': 0, 'threshold': 0.0, 'values': []}
    control_folder = next(
        (f for f in output_dir.iterdir()
         if f.is_dir() and f.name.lower().startswith("control")),
        None,
    )
    if not control_folder:
        return empty
    excel_path = control_folder / f"{control_folder.name}_master.xlsx"
    if not excel_path.exists():
        return empty
    try:
        df = pd.read_excel(
            excel_path,
            sheet_name=f"{control_folder.name}_Typical_Particles",
        )
        values = pd.to_numeric(
            df["Fluor_Density_per_BF_Area"], errors='coerce'
        ).dropna()
        if values.empty:
            return empty
        mean_val = float(values.mean())
        std_val  = float(values.std(ddof=1)) if len(values) > 1 else 0.0
        return {
            'mean':      mean_val,
            'std':       std_val,
            'n':         len(values),
            'threshold': mean_val * (1.0 - threshold_pct),
            'values':    values.tolist(),
        }
    except Exception:
        return empty


def _draw_sample_channel_chart(
    ax: MplAxes,
    title: str,
    accent: str,
    sample_val: Optional[float],
    ctrl_mean: float,
    ctrl_std: float,
    threshold: float,
    status: str,
    control_points: Optional[list[float]] = None,
    sample_points: Optional[list[float]] = None,
) -> None:
    ax.set_facecolor('#FAFAFA')
    for spine in ax.spines.values():
        spine.set_linewidth(0.8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    bar_vals   = [ctrl_mean, sample_val if sample_val is not None else 0.0]
    bar_colors = ['#7EB5D6', accent]
    ax.bar([0, 1], bar_vals, color=bar_colors, edgecolor='black',
           linewidth=0.8, alpha=0.75, width=0.5, zorder=3)
    ax.errorbar(0, ctrl_mean, yerr=ctrl_std, fmt='none', ecolor='black',
                elinewidth=1.5, capsize=9, capthick=1.5, zorder=5)

    sample_std = (
        float(np.std(sample_points, ddof=1))
        if sample_points is not None and len(sample_points) > 1
        else 0.0
    )
    if sample_val is not None and sample_std > 0:
        ax.errorbar(1, sample_val, yerr=sample_std, fmt='none', ecolor='black',
                    elinewidth=1.5, capsize=9, capthick=1.5, zorder=5)

    y_max = max(ctrl_mean + ctrl_std, sample_val or 0.0) * 1.20 or 1.0
    ax.set_ylim(0, y_max)

    if ctrl_mean > 0:
        ax.axhline(ctrl_mean, color='#2E75B6', ls=':', lw=1.5,
                   label=f'Ctrl mean ({ctrl_mean:.2f})', zorder=4)
    if threshold > 0:
        ax.axhline(threshold, color='red', ls='--', lw=1.8,
                   label=f'Threshold ({threshold:.2f})', zorder=4)

    nudge = y_max * 0.02
    ax.text(0, ctrl_mean + ctrl_std + nudge, f"{ctrl_mean:.2f}",
            ha='center', va='bottom', fontsize=6.5, color='#333333')
    if sample_val is not None:
        ax.text(1, bar_vals[1] + nudge, f"{sample_val:.2f}",
                ha='center', va='bottom', fontsize=6.5, color='#333333')

    ax.set_xticks([0, 1])
    ax.set_xticklabels(['Control', 'Sample'], fontsize=7.5)
    ax.set_ylabel("Fluor Density (a.u./\u00b5m\u00b2)", fontsize=7)
    ax.set_title(title, fontsize=8.5, fontweight='bold', color=accent, pad=4)
    ax.tick_params(labelsize=7)
    ax.legend(fontsize=5.5, loc='upper right', framealpha=0.85,
              edgecolor='#CCCCCC')

    # ── colour logic covers both 701.png label sets ─────────────────────
    # "No obvious" → not detected → green
    # Anything else that is not N/A → bacteria found → red
    s_color = (
        '#9C0006' if 'No obvious' not in status and status not in ('N/A', '')
        else '#006100' if 'No obvious' in status
        else '#777777'
    )
    ax.text(0.5, -0.22, f"\u25ba {status}", transform=ax.transAxes,
            ha='center', fontsize=8.5, fontweight='bold', color=s_color)
    

    # ── scatter points (per-image values) ───────────────────────────────
    rng = np.random.default_rng(42)

    if control_points:
        ctrl_arr = np.asarray(control_points, dtype=float)
        ctrl_arr = ctrl_arr[np.isfinite(ctrl_arr)]
        if ctrl_arr.size > 0:
            jitter = rng.uniform(-0.08, 0.08, ctrl_arr.size)
            ax.scatter(
                np.full(ctrl_arr.size, 0.0) + jitter,
                ctrl_arr,
                c="#55EAF2",
                s=34,
                edgecolors="black",
                linewidths=0.5,
                alpha=0.9,
                zorder=6,
                rasterized=True,
            )

    if sample_points:
        samp_arr = np.asarray(sample_points, dtype=float)
        samp_arr = samp_arr[np.isfinite(samp_arr)]
        if samp_arr.size > 0:
            jitter = rng.uniform(-0.08, 0.08, samp_arr.size)
            ax.scatter(
                np.full(samp_arr.size, 1.0) + jitter,
                samp_arr,
                c="#55EAF2",
                s=34,
                edgecolors="black",
                linewidths=0.5,
                alpha=0.9,
                zorder=6,
                rasterized=True,
            )

from matplotlib.patches import FancyBboxPatch, Rectangle

def _draw_classification_reference(ax: MplAxes) -> None:
    """
    Draw a compact classification reference table showing how
    G+ and G− microgel statuses map to the final classification.
    Styled as a simplified rounded-box version of the 701 reference image.
    """
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis('off')

    # Palette
    border_c = '#8FD3F4'
    header_blue = '#1E9AD6'
    body_blue = '#D9F1FB'
    final_blue = '#1177C3'
    text_c = '#111111'
    dash_c = '#2F2F2F'

    # Outer rounded frame
    outer = FancyBboxPatch(
        (0.01, 0.03), 0.98, 0.94,
        boxstyle="round,pad=0.008,rounding_size=0.03",
        linewidth=1.4, edgecolor=border_c, facecolor='white'
    )
    ax.add_patch(outer)

    # Column layout
    left_x0 = 0.02
    left_x1 = 0.36
    mid_x0 = 0.41
    mid_x1 = 0.75
    right_x0 = 0.80
    right_x1 = 0.98

    header_y0 = 0.82
    header_h = 0.13
    body_y0 = 0.10
    body_y1 = 0.82
    body_h = body_y1 - body_y0

    # Body backgrounds
    ax.add_patch(Rectangle((left_x0, body_y0), left_x1 - left_x0, body_h,
                           facecolor=body_blue, edgecolor='none', alpha=0.95))
    ax.add_patch(Rectangle((mid_x0, body_y0), mid_x1 - mid_x0, body_h,
                           facecolor=body_blue, edgecolor='none', alpha=0.95))

    # Rounded headers
    ax.add_patch(FancyBboxPatch(
        (left_x0, header_y0), left_x1 - left_x0, header_h,
        boxstyle="round,pad=0.005,rounding_size=0.03",
        linewidth=0, facecolor=header_blue
    ))
    ax.add_patch(FancyBboxPatch(
        (mid_x0, header_y0), mid_x1 - mid_x0, header_h,
        boxstyle="round,pad=0.005,rounding_size=0.03",
        linewidth=0, facecolor=header_blue
    ))

    # Header labels
    ax.text((left_x0 + left_x1) / 2, header_y0 + header_h / 2,
            "Positive microgel result",
            ha='center', va='center', fontsize=8.8, fontweight='bold',
            color='#FFC999')

    ax.text((mid_x0 + mid_x1) / 2, header_y0 + header_h / 2,
            "Negative microgel result",
            ha='center', va='center', fontsize=8.8, fontweight='bold',
            color='white')

    ax.text((right_x0 + right_x1) / 2, header_y0 + header_h / 2 + 0.01,
            "Final result",
            ha='center', va='center', fontsize=9.0, fontweight='bold',
            color=header_blue)

    # Row contents
    rows = [
        ("Positive", "Positive/No obvious bacteria", "Positive"),
        ("Negative/No obvious bacteria", "Negative", "Negative"),
        ("Negative/No obvious bacteria", "Positive/No obvious bacteria", "No obvious bacteria"),
        ("Positive", "Negative", "Mixed/Contradictory"),
    ]

    n_rows = len(rows)
    row_h = body_h / n_rows

    for i, (gp_txt, gm_txt, final_txt) in enumerate(rows):
        y_top = body_y1 - i * row_h
        y_bot = y_top - row_h
        y_mid = (y_top + y_bot) / 2

        # Dashed separators
        if i > 0:
            ax.plot([0.02, 0.98], [y_top, y_top], ls='--', lw=0.9, color=dash_c)

        # Left / middle text
        ax.text((left_x0 + left_x1) / 2, y_mid, gp_txt,
                ha='center', va='center', fontsize=8.2, color=text_c)
        ax.text((mid_x0 + mid_x1) / 2, y_mid, gm_txt,
                ha='center', va='center', fontsize=8.2, color=text_c)

        # Final-result rounded pill
        pill_w = (right_x1 - right_x0) * 0.90
        pill_h = row_h * 0.62
        pill_x = right_x0 + ((right_x1 - right_x0) - pill_w) / 2
        pill_y = y_mid - pill_h / 2

        ax.add_patch(FancyBboxPatch(
            (pill_x, pill_y), pill_w, pill_h,
            boxstyle="round,pad=0.01,rounding_size=0.04",
            linewidth=0, facecolor=final_blue
        ))
        ax.text(pill_x + pill_w / 2, y_mid, final_txt,
                ha='center', va='center', fontsize=7.8,
                color='white', fontweight='bold')


def _generate_single_sample_report_pdf(
    reports_dir: Path,
    sample_id:   str,
    group_name:  str,
    gplus_val:   Optional[float],
    gminus_val:  Optional[float],
    gplus_ctrl:  Dict,
    gminus_ctrl: Dict,
    threshold_pct: float,
    config:      dict,
    gplus_points: Optional[list[float]] = None,
    gminus_points: Optional[list[float]] = None,
) -> Optional[Path]:
    A4_W, A4_H  = 8.27, 11.69
    HEADER_BG   = "#1B3A5C"
    ACCENT      = "#2E75B6"
    LIGHT_GRAY  = "#F2F2F2"
    RESULT_COLOURS = {
        'POSITIVE':            ("#FFC7CE", "#9C0006"),
        'NEGATIVE':            ("#C6EFCE", "#006100"),
        'NO OBVIOUS BACTERIA': ("#FFEB9C", "#9C6500"),
        'MIXED/CONTRADICTORY': ("#FCD5B4", "#974706"),
    }

    dataset_id = config.get('dataset_id_base', config.get('dataset_id', 'Unknown'))
    timestamp  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    gp_mean = gplus_ctrl.get('mean', 0.0)
    gm_mean = gminus_ctrl.get('mean', 0.0)
    gp_thr  = gplus_ctrl.get('threshold', 0.0)
    gm_thr  = gminus_ctrl.get('threshold', 0.0)

    gp_detected = (gplus_val  is not None) and (gp_mean > 0) and (gplus_val  < gp_thr)
    gm_detected = (gminus_val is not None) and (gm_mean > 0) and (gminus_val < gm_thr)

    # ── status labels follow 701.png terminology ─────────────────────────
    # G+ (positive microgel): "Positive" | "Negative / No obvious bacteria"
    # G- (negative microgel): "Negative" | "Positive / No obvious bacteria"
    gp_status = (
        "Positive"
        if gp_detected
        else "Negative / No obvious bacteria"
        if (gplus_val is not None and gp_mean > 0)
        else "N/A"
    )
    gm_status = (
        "Negative"
        if gm_detected
        else "Positive / No obvious bacteria"
        if (gminus_val is not None and gm_mean > 0)
        else "N/A"
    )

    # ── final classification follows 701.png decision table ──────────────
    _DECISION = {
        (True,  False): ('POSITIVE',            'Gram-positive bacteria DETECTED'),
        (False, True):  ('NEGATIVE',            'Gram-negative bacteria DETECTED'),
        (False, False): ('NO OBVIOUS BACTERIA', 'No obvious bacteria detected'),
        (True,  True):  ('MIXED/CONTRADICTORY', 'Contradictory result — manual review required'),
    }
    final_class, interpretation = _DECISION.get(
        (gp_detected, gm_detected),
        ('NO OBVIOUS BACTERIA', 'No obvious bacteria detected')
    )
    res_bg, res_tc = RESULT_COLOURS.get(final_class, ('#FFFFFF', '#333333'))

    pdf_path = reports_dir / f"sample_{sample_id}.pdf"

    # ── helper: status cell colour uses same rule as _draw_sample_channel_chart
    def _st_color(st: str) -> str:
        if 'No obvious' in st:
            return '#006100'
        if st not in ('N/A', ''):
            return '#9C0006'
        return '#777777'

    try:
        fig = plt.figure(figsize=(A4_W, A4_H))

        # Header
        ax_h = fig.add_axes((0, 0.935, 1, 0.065))
        ax_h.set_xlim(0, 1); ax_h.set_ylim(0, 1); ax_h.axis('off')
        ax_h.fill_between([0, 1], 0, 1, color=HEADER_BG)
        ax_h.fill_between([0, 1], 0, 0.07, color=ACCENT)
        ax_h.text(0.05, 0.68, "MICROGEL FLUORESCENCE ANALYSIS",
                  fontsize=13, fontweight='bold', color='white', va='center')
        ax_h.text(0.05, 0.28, "Individual Sample Laboratory Report",
                  fontsize=8.5, color='#A0C4E8', va='center')
        ax_h.text(0.97, 0.50, f"Sample  {sample_id}",
                  fontsize=14, fontweight='bold', color='white',
                  va='center', ha='right')

        # Info box
        ax_i = fig.add_axes((0.05, 0.865, 0.90, 0.062))
        ax_i.set_xlim(0, 1); ax_i.set_ylim(0, 1); ax_i.axis('off')
        ax_i.fill_between([0, 1], 0, 1, color=LIGHT_GRAY)
        ax_i.plot([0, 1, 1, 0, 0], [0, 0, 1, 1, 0], color='#CCCCCC', lw=0.5)
        for xi, (lbl, val) in enumerate([
            ("Dataset",   dataset_id),
            ("Group",     group_name),
            ("Sample ID", sample_id),
            ("Date",      timestamp.split()[0]),
            ("Threshold", f"−{threshold_pct*100:.0f}% vs control"),
        ]):
            xpos = 0.02 + xi * 0.195
            ax_i.text(xpos, 0.72, lbl + ":", fontsize=7, fontweight='bold',
                      color='#555555', va='center')
            ax_i.text(xpos, 0.28, val, fontsize=7.5, color='#222222', va='center')

        # G+ chart
        ax_gp = fig.add_axes((0.07, 0.575, 0.38, 0.250))
        _draw_sample_channel_chart(
            ax_gp, "G+  Microgel  (Positive)", ACCENT,
            gplus_val, gp_mean, gplus_ctrl.get('std', 0.0), gp_thr, gp_status,
            control_points=gplus_ctrl.get('values', []),
            sample_points=gplus_points or [],
        )

        # G− chart
        ax_gm = fig.add_axes((0.57, 0.575, 0.38, 0.250))
        _draw_sample_channel_chart(
            ax_gm, "G−  Microgel  (Negative)", "#C0504D",
            gminus_val, gm_mean, gminus_ctrl.get('std', 0.0), gm_thr, gm_status,
            control_points=gminus_ctrl.get('values', []),
            sample_points=gminus_points or [],
        )

        # Stats table
        ax_t = fig.add_axes((0.05, 0.430, 0.90, 0.115))
        ax_t.set_xlim(0, 1); ax_t.set_ylim(0, 1); ax_t.axis('off')

        col_x    = [0.00, 0.145, 0.290, 0.420, 0.545, 0.680, 0.820]
        col_w    = 0.145
        col_hdrs = ['Channel', 'Sample value', 'Control mean', 'Control SD',
                    'Threshold', '% vs control', 'Status']
        row_h_hdr = 0.22

        ax_t.fill_between([0, 1], 1 - row_h_hdr, 1.0, color=HEADER_BG)
        for j, h in enumerate(col_hdrs):
            ax_t.text(col_x[j] + col_w / 2, 1 - row_h_hdr / 2, h,
                      fontsize=6.4, fontweight='bold', color='white',
                      va='center', ha='center')

        data_rows = [
            ("G+  (Positive)",      gplus_val,  gp_mean, gplus_ctrl.get('std', 0.0),  gp_thr,  gp_status),
            ("G−  (Negative)",      gminus_val, gm_mean, gminus_ctrl.get('std', 0.0), gm_thr,  gm_status),
        ]

        row_h = (1 - row_h_hdr) / len(data_rows)

        for row_i, (ch_lbl, s_val, c_mean, c_std, thr, stat) in enumerate(data_rows):
            y_top = 1 - row_h_hdr - row_i * row_h
            y_bot = y_top - row_h
            y_mid = (y_top + y_bot) / 2

            ax_t.fill_between([0, 1], y_bot, y_top,
                              color="#EEF4FF" if row_i == 0 else "#FFFFFF")

            pct = (f"{(s_val / c_mean - 1) * 100:+.1f}%"
                   if s_val is not None and c_mean > 0 else "N/A")

            row_cells = [
                ch_lbl,
                f"{s_val:.3f}" if s_val is not None else "N/A",
                f"{c_mean:.3f}" if c_mean > 0 else "N/A",
                f"{c_std:.3f}",
                f"{thr:.3f}" if thr > 0 else "N/A",
                pct,
                stat,
            ]
            for j, cell in enumerate(row_cells):
                clr = _st_color(stat) if j == 6 else '#222222'
                wt  = 'bold' if j == 6 else 'normal'
                ax_t.text(col_x[j] + col_w / 2, y_mid, cell,
                          fontsize=6.3, va='center', ha='center',
                          color=clr, fontweight=wt)

        # Classification reference
        ax_r = fig.add_axes((0.05, 0.255, 0.90, 0.145))
        _draw_classification_reference(ax_r)

        # Verdict box
        ax_v = fig.add_axes((0.05, 0.140, 0.90, 0.095))
        ax_v.set_xlim(0, 1); ax_v.set_ylim(0, 1); ax_v.axis('off')
        ax_v.fill_between([0, 1], 0, 1, color=res_bg)
        ax_v.plot([0, 1, 1, 0, 0], [0, 0, 1, 1, 0], color='#AAAAAA', lw=1)
        ax_v.text(0.50, 0.78, "FINAL CLASSIFICATION",
                  fontsize=8, fontweight='bold', color='#555555',
                  va='center', ha='center')
        ax_v.text(0.50, 0.43, final_class,
                  fontsize=16, fontweight='bold', color=res_tc,
                  va='center', ha='center')
        ax_v.text(0.50, 0.12, interpretation,
                  fontsize=7.8, color='#333333', va='center', ha='center')

        # Note box
        ax_n = fig.add_axes((0.05, 0.050, 0.90, 0.075))
        ax_n.set_xlim(0, 1); ax_n.set_ylim(0, 1); ax_n.axis('off')
        ax_n.fill_between([0, 1], 0, 1, color='#F8F8F8')
        ax_n.plot([0, 1], [1, 1], color='#DDDDDD', lw=0.5)
        notes = [
            "Metric: Fluorescence Integrated Density / BF Area (a.u./µm²) from Typical Particles (middle percentile band).",
            "Decision reference: final classification is determined from the combination of G+ and G− channel status shown above.",
            "Research use only. Confirm with culture-based or orthogonal methods before clinical action.",
        ]
        for li, line in enumerate(notes):
            ax_n.text(0.02, 0.78 - li * 0.30, f"• {line}",
                      fontsize=5.8, color='#555555', va='center', wrap=True)

        # Footer
        ax_f = fig.add_axes((0, 0.00, 1, 0.040))
        ax_f.set_xlim(0, 1); ax_f.set_ylim(0, 1); ax_f.axis('off')
        ax_f.plot([0.05, 0.95], [0.72, 0.72], color='#CCCCCC', lw=0.5)
        ax_f.text(0.05, 0.30, f"Generated: {timestamp}",
                  fontsize=6.3, color='#999999', va='center')
        ax_f.text(0.50, 0.30, f"Dataset: {dataset_id}  |  Sample: {sample_id}",
                  fontsize=6.3, color='#999999', va='center', ha='center')
        ax_f.text(0.95, 0.30, "CONFIDENTIAL — Research Use Only",
                  fontsize=6.3, color='#CC0000', va='center', ha='right',
                  fontweight='bold')

        fig.savefig(str(pdf_path), format='pdf')
        plt.close(fig)
        return pdf_path

    except Exception:
        plt.close('all')
        return None


def generate_per_sample_reports(output_root: Path, config: dict) -> list[Path]:
    """
    Generate one PDF per group containing:
      • Page 1   – matplotlib overview  (_generate_single_sample_report_pdf)
      • Pages 2+ – ReportLab additions  (_generate_individual_sample_additions_pdf)
                   merged into a single file when pypdf is available,
                   otherwise saved as sample_{group}_detailed.pdf alongside.

    Returns the list of generated PDF paths.
    """
    positive_output = config.get('positive_output')
    negative_output = config.get('negative_output')

    if not isinstance(positive_output, Path) or not isinstance(negative_output, Path):
        print("  \u26a0 Skipping per-sample reports: missing positive/negative output directories")
        return []

    threshold_pct = config.get('threshold_pct', 0.05)
    percentile    = config.get('percentile', 0.30)

    gplus_per_image  = _collect_per_image_data(positive_output)
    gminus_per_image = _collect_per_image_data(negative_output)
    gplus_ctrl       = _get_control_stats_from_dir(positive_output,  threshold_pct)
    gminus_ctrl      = _get_control_stats_from_dir(negative_output, threshold_pct)

    # Pre-load control Series for stats computation
    gplus_ctrl_series  = (
        pd.Series(gplus_ctrl['values'])  if gplus_ctrl['values']  else None
    )
    gminus_ctrl_series = (
        pd.Series(gminus_ctrl['values']) if gminus_ctrl['values'] else None
    )

    # Collect comparison data across all groups for the two channels
    gplus_all_group_data  = _collect_all_groups_comparison_data(positive_output)
    gminus_all_group_data = _collect_all_groups_comparison_data(negative_output)

    reports_dir = output_root / "sample_reports"
    reports_dir.mkdir(exist_ok=True)

    all_groups = sorted(
        set(gplus_per_image.keys()) | set(gminus_per_image.keys()),
        key=_group_order_key,
    )
    all_groups = [g for g in all_groups if not g.lower().startswith('control')]

    generated: list[Path] = []

    for group_name in all_groups:
        gplus_imgs   = gplus_per_image.get(group_name, {})
        gminus_imgs  = gminus_per_image.get(group_name, {})

        gplus_densities  = [v['fluor_density'] for v in gplus_imgs.values()]
        gminus_densities = [v['fluor_density'] for v in gminus_imgs.values()]

        gplus_val  = float(np.mean(gplus_densities))  if gplus_densities  else None
        gminus_val = float(np.mean(gminus_densities)) if gminus_densities else None

        if gplus_val is None and gminus_val is None:
            continue

        # ── Page 1: matplotlib overview ──────────────────────────────────
        overview_path = _generate_single_sample_report_pdf(
            reports_dir,
            sample_id=group_name,
            group_name=group_name,
            gplus_val=gplus_val,
            gminus_val=gminus_val,
            gplus_ctrl=gplus_ctrl,
            gminus_ctrl=gminus_ctrl,
            threshold_pct=threshold_pct,
            config=config,
            gplus_points=gplus_densities,
            gminus_points=gminus_densities,
        )

        if overview_path is None:
            continue

        # ── Pages 2+: ReportLab additions (G+ channel detail) ────────────
        gplus_additions_path = _generate_individual_sample_additions_pdf(
            reports_dir=reports_dir,
            group_name=group_name,
            output_dir=positive_output,
            control_values=gplus_ctrl_series,
            control_mean=gplus_ctrl.get('mean', 0.0),
            threshold_pct=threshold_pct,
            all_group_data=gplus_all_group_data,
            percentile=percentile,
            channel_label="G+ Microgel (Gram-Positive)",
            config=config,
        )

        # ── Pages 2+: ReportLab additions (G− channel detail) ────────────
        gminus_additions_path = _generate_individual_sample_additions_pdf(
            reports_dir=reports_dir,
            group_name=group_name,
            output_dir=negative_output,
            control_values=gminus_ctrl_series,
            control_mean=gminus_ctrl.get('mean', 0.0),
            threshold_pct=threshold_pct,
            all_group_data=gminus_all_group_data,
            percentile=percentile,
            channel_label="G\u2212 Microgel (Gram-Negative)",
            config=config,
        )

        # ── Merge overview + additions into one PDF ───────────────────────
        parts_to_merge = [p for p in [overview_path,
                                       gplus_additions_path,
                                       gminus_additions_path]
                          if p is not None and p.exists()]

        final_path = reports_dir / f"sample_{group_name}.pdf"

        if len(parts_to_merge) > 1 and PYPDF_AVAILABLE:
            merged = _merge_pdfs_to_file(parts_to_merge, final_path)
            if merged:
                # Clean up intermediate files
                for tmp in parts_to_merge:
                    if tmp != final_path:
                        try:
                            tmp.unlink()
                        except Exception:
                            pass
            else:
                # Merge failed — keep overview as the final PDF
                if overview_path != final_path and overview_path.exists():
                    try:
                        shutil.copy2(overview_path, final_path)
                    except Exception:
                        pass
        else:
            # No merge available: rename overview as final; additions stay separate
            if overview_path != final_path and overview_path.exists():
                try:
                    shutil.copy2(overview_path, final_path)
                    overview_path.unlink()
                except Exception:
                    final_path = overview_path

        if final_path.exists():
            _compress_pdf_with_pypdf(final_path)
            generated.append(final_path)
            n_gp = len(gplus_densities)
            n_gm = len(gminus_densities)
            extra = ""
            if not PYPDF_AVAILABLE and (gplus_additions_path or gminus_additions_path):
                extra = " (detailed pages saved separately — install pypdf to merge)"
            print(
                f"  \u2713 Group {group_name}: report generated "
                f"(G+ mean over {n_gp} img{'s' if n_gp != 1 else ''}, "
                f"G\u2212 mean over {n_gm} img{'s' if n_gm != 1 else ''})"
                f"{extra}"
            )



    print(f"  \u2713 {len(generated)} sample report(s) \u2192 {reports_dir.name}/")
    return generated


def generate_final_clinical_matrix(output_root: Path, gplus_classification: pd.DataFrame, gminus_classification: pd.DataFrame, dataset_base_name: str) -> Optional[Path]:
    if gplus_classification.empty or gminus_classification.empty: return None
    gplus_classification['Group'], gminus_classification['Group'] = gplus_classification['Group'].astype(str), gminus_classification['Group'].astype(str)

    def _norm(c): return 'control' if 'CONTROL' in str(c).upper() else 'not_detected' if 'NO OBVIOUS' in str(c).upper() else 'detected' if 'DETECTED' in str(c).upper() else 'unknown'
    decision_matrix = {('detected', 'not_detected'): 'POSITIVE', ('not_detected', 'detected'): 'NEGATIVE', ('not_detected', 'not_detected'): 'NO OBVIOUS BACTERIA', ('detected', 'detected'): 'MIXED/CONTRADICTORY'}

    gplus_dict, gminus_dict = gplus_classification.set_index('Group').to_dict('index'), gminus_classification.set_index('Group').to_dict('index')
    all_groups = sorted(set(gplus_classification['Group']) | set(gminus_classification['Group']), key=_group_order_key)

    results = []
    for group in all_groups:
        gp, gm = gplus_dict.get(group, {}), gminus_dict.get(group, {})
        gp_norm, gm_norm = _norm(gp.get('Classification')), _norm(gm.get('Classification'))
        final_class = 'CONTROL (Reference)' if group == 'Control' else 'MISSING DATA' if not gp and not gm else 'MISSING G+' if not gp else 'MISSING G-' if not gm else decision_matrix.get((gp_norm, gm_norm), 'UNKNOWN COMBINATION')

        if group == 'Control':
            interp = final_class
        else:
            interp_parts = [{
                "NEGATIVE": "Gram-negative bacteria DETECTED",
                "POSITIVE": "Gram-positive bacteria DETECTED",
                "NO OBVIOUS BACTERIA": "No obvious bacteria detected",
                "MIXED/CONTRADICTORY": "Contradictory \u2014 manual review required",
            }.get(final_class, final_class)]
            for lbl, d in [("G+", gp), ("G-", gm)]:
                p_raw = d.get("P_Value")
                if p_raw is not None and not pd.isna(p_raw):
                    interp_parts.append(f"[{lbl} p={float(p_raw):.3f} {d.get('Significance', '')}]")
            interp = " ".join(interp_parts).strip()

        def _s(d, k, f="{:.2f}"): return f.format(float(d[k])) if d.get(k) is not None and not pd.isna(d.get(k)) else '\u2014'

        results.append({'Group': group, 'G+_N': _s(gp, 'N', "{:.0f}"), 'G+_Mean': _s(gp, 'Mean'), 'G+_Std': _s(gp, 'Std_Dev'), 'G+_SEM': _s(gp, 'SEM'), 'G+_Median': _s(gp, 'Median'), 'G+_CI': f"{_s(gp, 'CI_Lower')} \u2013 {_s(gp, 'CI_Upper')}", 'G+_Cohens_d': _s(gp, 'Cohens_d', "{:.3f}"), 'G+_d_CI': f"{_s(gp, 'd_CI_Lower', '{:.3f}')} \u2013 {_s(gp, 'd_CI_Upper', '{:.3f}')}", 'G+_Effect_Size': gp.get('Effect_Size_Label', '\u2014'), 'G+_P_Value': _s(gp, 'P_Value', "{:.4f}"), 'G+_Significance': gp.get('Significance', '\u2014'), 'G+_Detection': {'detected': 'Detected', 'not_detected': 'Not Detected'}.get(gp_norm, 'Unknown'), 'G+_Confidence': gp.get('Classification_Confidence', '\u2014'), 'G+_Strength': gp.get('Strength_of_Evidence', '\u2014'), 'G+_Pct_Diff': _s(gp, 'Pct_Diff_from_Control', "{:.1f}"), 'G+_Classification': gp.get('Classification', '\u2014'), 'G+_Control_Mean': _s(gp, 'Control_Mean'), 'G+_Threshold': _s(gp, 'Threshold'), 'G-_N': _s(gm, 'N', "{:.0f}"), 'G-_Mean': _s(gm, 'Mean'), 'G-_Std': _s(gm, 'Std_Dev'), 'G-_SEM': _s(gm, 'SEM'), 'G-_Median': _s(gm, 'Median'), 'G-_CI': f"{_s(gm, 'CI_Lower')} \u2013 {_s(gm, 'CI_Upper')}", 'G-_Cohens_d': _s(gm, 'Cohens_d', "{:.3f}"), 'G-_d_CI': f"{_s(gm, 'd_CI_Lower', '{:.3f}')} \u2013 {_s(gm, 'd_CI_Upper', '{:.3f}')}", 'G-_Effect_Size': gm.get('Effect_Size_Label', '\u2014'), 'G-_P_Value': _s(gm, 'P_Value', "{:.4f}"), 'G-_Significance': gm.get('Significance', '\u2014'), 'G-_Detection': {'detected': 'Detected', 'not_detected': 'Not Detected'}.get(gm_norm, 'Unknown'), 'G-_Confidence': gm.get('Classification_Confidence', '\u2014'), 'G-_Strength': gm.get('Strength_of_Evidence', '\u2014'), 'G-_Pct_Diff': _s(gm, 'Pct_Diff_from_Control', "{:.1f}"), 'G-_Classification': gm.get('Classification', '\u2014'), 'G-_Control_Mean': _s(gm, 'Control_Mean'), 'G-_Threshold': _s(gm, 'Threshold'), 'Final_Classification': final_class, 'Interpretation': interp})

    final_df = pd.DataFrame(results)
    final_df.to_csv(output_root / "final_clinical_results.csv", index=False)

    try:
        with pd.ExcelWriter(output_root / "final_clinical_results.xlsx", engine='openpyxl') as writer:
            final_df.to_excel(writer, sheet_name='Final Results', index=False)
            ws = writer.sheets['Final Results']
            fills = {'POSITIVE': PatternFill("solid", fgColor="FFC7CE"), 'NEGATIVE': PatternFill("solid", fgColor="C6EFCE"), 'NO OBVIOUS BACTERIA': PatternFill("solid", fgColor="FFEB9C"), 'MIXED/CONTRADICTORY': PatternFill("solid", fgColor="FCD5B4"), 'CONTROL (Reference)': PatternFill("solid", fgColor="E7E6E6")}
            for cell in ws[1]: cell.fill, cell.font, cell.alignment, cell.border = PatternFill("solid", fgColor="4472C4"), Font(bold=True, color="FFFFFF", size=10), Alignment(horizontal="center", vertical="center", wrap_text=True), Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for r_idx in range(len(final_df)):
                fill = fills.get(final_df.iloc[r_idx]['Final_Classification'], PatternFill("solid", fgColor="D9D9D9"))
                for c_idx in range(1, len(final_df.columns) + 1):
                    c = ws.cell(row=r_idx + 2, column=c_idx)
                    c.fill, c.alignment, c.border = fill, Alignment(horizontal="center", vertical="center", wrap_text=True), Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = min(max((len(str(c.value or '')) for c in col)) + 3, 28)
    except Exception: pass
    return output_root / "final_clinical_results.xlsx"

def generate_final_clinical_matrix_wrapper(output_root: Path, config: dict) -> None:
    print("\n" + "=" * 80 + "\nGENERATING FINAL CLINICAL MATRIX\n" + "=" * 80)

    positive_output = config.get('positive_output')
    negative_output = config.get('negative_output')

    if not isinstance(positive_output, Path) or not isinstance(negative_output, Path):
        print("  \u26a0 Missing positive/negative output directories")
        return

    gplus_csv  = positive_output / "clinical_classification_positive.csv"
    gminus_csv = negative_output / "clinical_classification_negative.csv"

    if not gplus_csv.exists() or not gminus_csv.exists():
        print("  \u26a0 Missing classification files")
        return

    gplus_df  = pd.read_csv(gplus_csv)
    gminus_df = pd.read_csv(gminus_csv)

    matrix_path = generate_final_clinical_matrix(
        output_root, gplus_df, gminus_df,
        config.get('dataset_id_base', config.get('dataset_id', 'Dataset'))
    )
    if matrix_path:
        print(f"  \u2713 Final matrix: {matrix_path.name}")

    print("\n" + "=" * 80 + "\nGENERATING PDF LABORATORY REPORT\n" + "=" * 80)
    final_csv = output_root / "final_clinical_results.csv"
    pdf_path  = generate_laboratory_report_pdf(
        output_root, config, gplus_df, gminus_df,
        pd.read_csv(final_csv) if final_csv.exists() else pd.DataFrame(),
    )
    if pdf_path:
        print(f"  \u2713 PDF report: {pdf_path.name}")

    print("\n" + "=" * 80 + "\nGENERATING INDIVIDUAL SAMPLE REPORTS\n" + "=" * 80)
    sample_reports = generate_per_sample_reports(output_root, config)
    if sample_reports:
        print(f"  \u2713 {len(sample_reports)} sample reports saved to: sample_reports/")
    else:
        print("  \u26a0 No individual sample reports generated")


# ==================================================
# Plotting & PDF Generation
# ==================================================

def generate_multi_config_comparison_plot(comparison_df: pd.DataFrame, all_results: dict, output_dir: Path) -> Optional[Path]:
    if comparison_df.empty: return None
    try:
        plt.figure(figsize=(14, 8))
        config_keys, bacteria_names = comparison_df["Config_Key"].astype(str).tolist(), comparison_df["Bacteria_Type"].astype(str).tolist()
        mean_fluor, confidence = comparison_df["Mean_Fluorescence"].to_numpy(dtype=float, na_value=0.0), comparison_df["Confidence_Percent"].to_numpy(dtype=float, na_value=0.0)
        std_fluor = np.array([float(all_results[k].get("std_fluorescence", 0.0)) for k in config_keys], dtype=float)
        x = np.arange(len(bacteria_names), dtype=float)
        bars = plt.bar(x, mean_fluor, width=0.6, alpha=0.8, edgecolor='black', linewidth=2)
        for bar, conf in zip(bars, confidence): bar.set_color('darkgreen' if conf >= 70 else 'gold' if conf >= 50 else 'orangered')
        plt.errorbar(x, mean_fluor, yerr=std_fluor, fmt='none', ecolor='black', elinewidth=2.5, capsize=10, capthick=2.5, zorder=10)
        plt.xlabel('Bacteria Configuration', fontsize=14, fontweight='bold')
        plt.ylabel('Mean Fluorescence (a.u./\u00b5m\u00b2)', fontsize=14, fontweight='bold')
        plt.title('Multi-Configuration Scan Results\n(Error bars: \u00b11 Standard Deviation)', fontsize=16, fontweight='bold', pad=20)
        plt.xticks(x, bacteria_names, rotation=45, ha='right', fontsize=11)
        plt.yticks(fontsize=11)
        plt.grid(axis='y', alpha=0.3, linestyle='--', linewidth=1)
        max_y = float(np.max(mean_fluor + std_fluor)) if len(mean_fluor) > 0 else 1.0
        for xi, yi, std_i, conf in zip(x, mean_fluor, std_fluor, confidence):
            plt.text(float(xi), float(yi + std_i + max_y * 0.03), f'{float(conf):.0f}%', ha='center', va='bottom', fontsize=11, fontweight='bold', bbox=dict(boxstyle='round,pad=0.4', facecolor='white', edgecolor='black', linewidth=1.5, alpha=0.9))
        plt.legend(handles=[mpatches.Patch(facecolor='darkgreen', edgecolor='black', label='High Confidence (\u226570%)'), mpatches.Patch(facecolor='gold', edgecolor='black', label='Moderate Confidence (50-69%)'), mpatches.Patch(facecolor='orangered', edgecolor='black', label='Low Confidence (<50%)')], loc='upper right', fontsize=10, framealpha=0.95)
        plt.tight_layout()
        plot_path = output_dir / "multi_config_comparison_with_statistics.png"
        plt.savefig(plot_path, dpi=300, bbox_inches='tight')
        plt.close()
        return plot_path
    except Exception:
        plt.close()
        return None

def generate_error_bar_comparison_with_threshold(output_dir: Path, percentile: float = 0.3, restrict_to_groups: Optional[list[str]] = None, output_path: Optional[Path] = None, title_suffix: str = "", dataset_id: str = "", threshold_pct: float = 0.05, microgel_type: str = "negative") -> Optional[Path]:
    excel_files = [p for p in output_dir.rglob("*_master.xlsx") if len(p.relative_to(output_dir).parts) == 2]
    if not excel_files: return None
    all_data_rows, group_stats, control_mean = [], {}, None
    for excel_path in sorted(excel_files):
        group_name_raw = excel_path.stem.replace("_master", "")
        display_name = _display_group_name(group_name_raw)
        if restrict_to_groups and display_name not in restrict_to_groups: continue
        try:
            df = pd.read_excel(excel_path, sheet_name=f"{group_name_raw}_Typical_Particles")
            if "Fluor_Density_per_BF_Area" not in df.columns: continue
            values = df["Fluor_Density_per_BF_Area"].dropna()
            if len(values) == 0: continue
            for v in values: all_data_rows.append({"Group": display_name, "Fluorescence Density": float(np.asarray(v).item())})
            mean_val, std_val, sem_val = float(np.asarray(values.mean()).item()), float(np.asarray(values.std()).item()), float(np.asarray(values.sem()).item())
            group_stats[display_name] = {"n": int(len(values)), "mean": mean_val, "std": std_val, "sem": sem_val}
            if display_name == "Control": control_mean = mean_val
        except Exception: continue
    if not all_data_rows: return None
    df_all = pd.DataFrame(all_data_rows)
    group_order = sorted(df_all["Group"].dropna().astype(str).drop_duplicates().tolist(), key=_group_order_key)
    if not group_order: return None
    palette_colors = ["silver", "violet"] if len(group_order) == 2 else ["skyblue"] if len(group_order) == 1 else list(sns.color_palette("husl", len(group_order)))
    plt.figure(figsize=(10, 7))
    sns.set_style("ticks")
    try: ax = sns.barplot(data=df_all, x="Group", y="Fluorescence Density", hue="Group", order=group_order, hue_order=group_order, palette=palette_colors, legend=False, errorbar=None, edgecolor="black", alpha=0.7)
    except TypeError: ax = sns.barplot(data=df_all, x="Group", y="Fluorescence Density", hue="Group", order=group_order, hue_order=group_order, palette=palette_colors, legend=False, ci=None, edgecolor="black", alpha=0.7)
    means, sds = df_all.groupby("Group")["Fluorescence Density"].mean(), df_all.groupby("Group")["Fluorescence Density"].std(ddof=1)
    for xi, g in enumerate(group_order): ax.errorbar(xi, float(means.get(g, 0.0)), yerr=float(sds.get(g, 0.0)), fmt="none", ecolor="black", elinewidth=1.5, capsize=14 if g == "Control" else 7, capthick=1.5, zorder=10)
    sns.stripplot(x="Group", y="Fluorescence Density", data=df_all, order=group_order, jitter=True, color="cyan", edgecolor="black", linewidth=0.5, size=6, alpha=0.6)
    legend_handles = []
    if control_mean is not None:
        legend_handles.append(ax.axhline(y=control_mean, color='blue', linestyle=':', linewidth=2.5, label=f'Control Mean ({control_mean:.1f})', zorder=5))
        threshold = control_mean * (1 - threshold_pct)
        legend_handles.append(ax.axhline(y=threshold, color='red', linestyle='--', linewidth=2.5, label=f'Lower Threshold (-{threshold_pct*100:.0f}%: {threshold:.1f})', zorder=5))
        ax.legend(handles=legend_handles, loc='upper right', fontsize=10, framealpha=0.95, edgecolor='black', fancybox=True)
    plt.ylabel("Fluorescence Density (a.u./\u00b5m\u00b2)", fontsize=12, fontweight="bold")
    plt.xlabel("")
    plt.xticks(fontsize=10, fontweight="bold")
    plt.yticks(fontsize=10, fontweight="bold")
    title_parts = [dataset_id] if dataset_id else []
    title_parts.append(f"{'G- Microgel' if microgel_type.lower() == 'negative' else 'G+ Microgel'} \u2014 Typical Particles: Middle {100 - 2*int(percentile * 100)}% (Excluded top/bottom {int(percentile * 100)}%)")
    if title_suffix: title_parts.append(title_suffix)
    plt.title("\n".join(textwrap.wrap(" \u2014 ".join(title_parts), width=80, break_long_words=False, break_on_hyphens=False)), fontsize=11, pad=10)
    for axis in ["top", "bottom", "left", "right"]: plt.gca().spines[axis].set_linewidth(1.5)
    plt.tight_layout()
    out_path = output_path or output_dir / f"comparison_{microgel_type}_{'_'.join([g.replace(' ', '_') for g in group_order]) if restrict_to_groups else 'all_groups'}.png"
    try:
        plt.savefig(out_path, dpi=300)
        plt.close()
        return out_path
    except Exception:
        plt.close()
        return None

def generate_pairwise_group_vs_control_plots(output_root: Path, percentile: float, dataset_id: str, threshold_pct: float, microgel_type: str) -> None:
    control_folder = next((f for f in output_root.iterdir() if f.is_dir() and f.name.lower().startswith("control")), None)
    if not control_folder or not (control_folder / f"{control_folder.name}_master.xlsx").exists(): return
    for group_dir in sorted(output_root.iterdir()):
        if group_dir.is_dir() and not group_dir.name.lower().startswith("control") and re.fullmatch(r"\d+", group_dir.name) and (group_dir / f"{group_dir.name}_master.xlsx").exists():
            generate_error_bar_comparison_with_threshold(output_root, percentile, [group_dir.name, _display_group_name(control_folder.name)], group_dir / f"Group_{group_dir.name}_vs_Control_threshold.png", f"Group {group_dir.name} vs Control", dataset_id, threshold_pct, microgel_type)

def embed_comparison_plots_into_all_excels(output_root: Path, percentile: float = 0.2, plot_path: Optional[Path] = None) -> None:
    if not plot_path or not plot_path.exists(): return
    for excel_path in [p for p in output_root.rglob("*_master.xlsx") if len(p.relative_to(output_root).parts) == 2]:
        try:
            wb = load_workbook(excel_path)
            ws_summary = wb["Summary"] if "Summary" in wb.sheetnames else wb["Error_Bar_Summary"] if "Error_Bar_Summary" in wb.sheetnames else wb.create_sheet("Summary")
            ws_summary.title = "Summary"
            if ws_summary["G1"].value != "COMPARISON_PLOTS_EMBEDDED":
                img = XLImage(str(plot_path))
                if getattr(img, "width", None) and getattr(img, "height", None): img.width, img.height = int(img.width * (620 / float(img.width))), int(img.height * (620 / float(img.width)))
                ws_summary.add_image(img, "G3")
                ws_summary["G1"].value = "COMPARISON_PLOTS_EMBEDDED"
                wb.save(excel_path)
        except Exception: pass

def _parse_confidence_report(report_path: Optional[Path]) -> dict:
    out = {'status': 'Unknown', 'confidence_pct': None, 'confidence_label': '', 'best_match': '', 'rank2_match': '', 'ambiguous': False, 'no_bacteria': False}
    if not report_path or not report_path.exists(): return out
    try:
        in_summary = False
        for raw in report_path.read_text(encoding='utf-8').splitlines():
            s = raw.strip()
            if 'DETECTION SUMMARY:' in s: in_summary = True; continue
            if any(k in s for k in ('TOP 3 CONFIGURATION', 'FULL CONFIGURATION', 'STATISTICAL SIGNIFICANCE', 'CLINICAL RECOMMENDATION')): in_summary = False
            if s.startswith('Status:'):
                out['status'] = s[7:].strip()
                out['ambiguous'], out['no_bacteria'] = 'AMBIGUOUS' in out['status'].upper(), 'NO BACTERIA' in out['status'].upper()
            elif s.startswith('Confidence:'):
                cs = s[11:].strip()
                m, lm = re.search(r'([\d.]+)%', cs), re.match(r'([A-Za-z]+)', cs)
                if m: out['confidence_pct'] = float(m.group(1))
                if lm: out['confidence_label'] = lm.group(1)
            if in_summary:
                if 'Strong match to:' in s: out['best_match'] = s.split('Strong match to:', 1)[-1].strip()
                elif 'Possible match to:' in s: out['best_match'] = s.split('Possible match to:', 1)[-1].strip()
                elif s.startswith('Rank 1:') and not out['best_match']: out['best_match'] = s[7:].strip()
                elif s.startswith('Rank 2:') and not out['rank2_match']: out['rank2_match'] = s[7:].strip()
    except Exception: pass
    return out

def _draw_chart_on_axis(ax: MplAxes, classification_df: pd.DataFrame, title: str, bar_colour: str, output_dir: Optional[Path] = None, threshold_pct: float = 0.05) -> None:
    if classification_df.empty:
        ax.text(0.5, 0.5, "No data", ha='center', va='center')
        ax.set_title(title, fontsize=9, fontweight='bold', color=bar_colour, pad=6)
        return
    df = classification_df.copy()
    df['sort_key'] = df['Group'].apply(lambda x: (1, 999) if x == 'Control' else (0, int(x) if x.isdigit() else 999))
    df = df.sort_values('sort_key').reset_index(drop=True)
    groups, means, stds, x = df['Group'].tolist(), np.array(df['Mean'].astype(float).tolist(), dtype=float), np.array(df['Std_Dev'].astype(float).tolist(), dtype=float), np.arange(len(df), dtype=float)

    individual_data = {}
    if output_dir and output_dir.exists():
        for g in groups:
            folder = next((f for f in output_dir.iterdir() if f.is_dir() and f.name.lower().startswith('control')), None) if g == 'Control' else output_dir / g
            if folder and (folder / f"{folder.name}_master.xlsx").exists():
                try:
                    vals = pd.to_numeric(pd.read_excel(folder / f"{folder.name}_master.xlsx", sheet_name=f"{folder.name}_Typical_Particles")['Fluor_Density_per_BF_Area'], errors='coerce').dropna().tolist()
                    if vals: individual_data[g] = vals
                except Exception: pass

    palette, colours, palette_idx = list(sns.color_palette("husl", max(sum(1 for g in groups if g != 'Control'), 1))), [], 0
    for g in groups:
        if g == 'Control': colours.append(bar_colour)
        else: colours.append(palette[palette_idx % len(palette)]); palette_idx += 1

    ax.bar(x, means, color=colours, edgecolor='black', linewidth=0.8, alpha=0.7, width=0.6)
    for xi, mean_val, sd_val, g in zip(x, means, stds, groups): ax.errorbar(float(xi), float(mean_val or 0.0), yerr=float(sd_val or 0.0), fmt='none', ecolor='black', elinewidth=1.5, capsize=14 if g == 'Control' else 7, capthick=1.5, zorder=10)

    if individual_data:
        rng = np.random.RandomState(42)
        for xi, g in zip(x, groups):
            if g in individual_data: ax.scatter(float(xi) + rng.uniform(-0.15, 0.15, size=len(individual_data[g])), individual_data[g], color='cyan', edgecolor='black', linewidth=0.5, s=30, alpha=0.6, zorder=11)

    ctrl_row = df[df['Group'] == 'Control']
    if not ctrl_row.empty:
        ctrl_mean = float(ctrl_row.iloc[0]['Mean'])
        threshold = float(ctrl_row.iloc[0].get('Threshold', ctrl_mean * (1 - threshold_pct)))
        ax.axhline(ctrl_mean, color='blue', ls=':', lw=2.0, label=f'Control Mean ({ctrl_mean:.1f})')
        ax.axhline(threshold, color='red', ls='--', lw=2.0, label=f'Threshold ({threshold:.1f})')
        ax.legend(fontsize=7, loc='upper right', framealpha=0.9)

    y_max = float(np.max(means + stds)) if len(means) > 0 else 1.0
    for i, (_, row_data) in enumerate(df.iterrows()):
        if row_data['Group'] != 'Control' and str(row_data.get('Significance', '')) not in ('', '\u2014', 'N/A (n<2)'): ax.text(float(x[i]), float(means[i]) + float(stds[i]) + y_max * 0.03, str(row_data['Significance']), ha='center', va='bottom', fontsize=9, fontweight='bold', color='#333333')

    ax.set_xticks(x)
    ax.set_xticklabels([f"{g}\nn={df[df['Group'] == g].iloc[0].get('N', '')}" if not df[df['Group'] == g].empty else g for g in groups], fontsize=7.5, fontweight='bold', multialignment='center', linespacing=1.4)
    ax.tick_params(axis='x', pad=6)
    ax.set_ylabel("Fluor Density (a.u./\u00b5m\u00b2)", fontsize=8)
    ax.set_title(title, fontsize=9, fontweight='bold', color=bar_colour, pad=6)

def _draw_forest_plot_on_axis(ax: MplAxes, gplus_df: pd.DataFrame, gminus_df: pd.DataFrame) -> None:
    rows = []
    for lbl, cdf, clr in [("G+", gplus_df, "#2E75B6"), ("G\u2212", gminus_df, "#C0504D")]:
        df = cdf.copy()
        df['sort_key'] = df['Group'].apply(lambda x: (1, 999) if x == 'Control' else (0, int(x) if x.isdigit() else 999))
        for _, r in df.sort_values('sort_key').iterrows():
            if r['Group'] != 'Control' and not pd.isna(r.get('Cohens_d', np.nan)):
                rows.append({'label': f"Grp {r['Group']} {lbl}", 'd': float(r['Cohens_d']), 'lo': float(r.get('d_CI_Lower', r['Cohens_d'])), 'hi': float(r.get('d_CI_Upper', r['Cohens_d'])), 'colour': clr, 'sig': str(r.get('Significance', ''))})
    if not rows:
        ax.text(0.5, 0.5, "No effect-size data", ha='center', va='center', fontsize=10)
        return
    rows = rows[::-1]
    for i, r in enumerate(rows):
        ax.plot([r['lo'], r['hi']], [i, i], color=r['colour'], lw=2.5, solid_capstyle='round')
        ax.plot(r['d'], i, 'D' if abs(r['d']) >= 0.8 else 'o', color=r['colour'], markersize=7, markeredgecolor='black', markeredgewidth=0.8, zorder=10)
        if r['sig'] and r['sig'] not in ('\u2014', 'ns', 'N/A (n<2)'): ax.text(r['hi'] + 0.15, i, r['sig'], fontsize=8, fontweight='bold', va='center', color='#006100')
        elif r['sig'] == 'ns': ax.text(r['hi'] + 0.15, i, 'ns', fontsize=7, va='center', color='#999999')
    ax.axvline(0, color='black', ls='--', lw=1, zorder=0)
    ax.axvspan(-0.2, 0.2, color='#E8E8E8', alpha=0.5, zorder=0, label='Negligible (|d|<0.2)')
    ax.set_yticks(np.arange(len(rows)))
    ax.set_yticklabels([r['label'] for r in rows], fontsize=7.5)
    ax.set_xlabel("Cohen\u2019s d (effect size vs Control)", fontsize=8, fontweight='bold')
    ax.set_title("Effect Size Forest Plot (95% CI)", fontsize=10, fontweight='bold', pad=8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='x', alpha=0.2, ls='--')
    ax.legend(handles=[Line2D([0], [0], color='#2E75B6', lw=2.5, label='G+ channel'), Line2D([0], [0], color='#C0504D', lw=2.5, label='G\u2212 channel'), Line2D([0], [0], color='black', ls='--', lw=1, label='No effect (d=0)')], fontsize=7, loc='lower right', framealpha=0.9)

def _draw_stats_table(ax: MplAxes, classification_df: pd.DataFrame, title: str, y_top: float, y_bottom_limit: float = 0.05) -> float:
    ax.text(0.05, y_top, title, fontsize=10, fontweight='bold', color="#1B3A5C")
    ax.plot([0.05, 0.95], [y_top - 0.01, y_top - 0.01], color="#2E75B6", lw=1)
    headers, col_x, row_h, y = ['Group', 'N', 'Mean', 'SD', '95% CI', "Cohen\u2019s d", 'p-value', 'Sig', 'Confidence'], [0.05, 0.13, 0.19, 0.28, 0.36, 0.52, 0.62, 0.73, 0.80], 0.022, y_top - 0.035
    if y - row_h / 2 >= y_bottom_limit:
        for j, h in enumerate(headers): ax.text(col_x[j], y, h, fontsize=6.5, fontweight='bold', color='white', va='center')
        ax.fill_between([0.05, 0.95], y - row_h / 2, y + row_h / 2, color="#1B3A5C")
    df = classification_df.copy()
    df['sort_key'] = df['Group'].apply(lambda x: (1, 999) if x == 'Control' else (0, int(x) if x.isdigit() else 999))
    last_row_bottom = y - row_h / 2
    for i, (_, row) in enumerate(df.sort_values('sort_key').reset_index(drop=True).iterrows()):
        y_row = y - (i + 1) * row_h
        if y_row - row_h / 2 < y_bottom_limit: break
        ax.fill_between([0.05, 0.95], y_row - row_h / 2, y_row + row_h / 2, color="#F2F2F2" if i % 2 == 0 else "#FFFFFF")
        p_val = row.get('P_Value', np.nan)
        cells = [str(row.get('Group', '')), str(row.get('N', '')), f"{float(row.get('Mean', 0)):.1f}", f"{float(row.get('Std_Dev', 0)):.1f}", f"{row.get('CI_Lower')}\u2013{row.get('CI_Upper')}" if row.get('CI_Lower') != '\u2014' else '\u2014', str(row.get('Cohens_d', '\u2014')), '\u2014' if pd.isna(p_val) else '<0.001' if p_val < 0.001 else f"{p_val:.4f}", str(row.get('Significance', '\u2014')), str(row.get('Classification_Confidence', '\u2014'))]
        for j, cell in enumerate(cells):
            colour, weight, sig_val = '#333333', 'normal', str(row.get('Significance', ''))
            if j == 7 and sig_val.startswith('*'): colour, weight = '#006100', 'bold'
            elif j == 7 and sig_val == 'ns': colour = '#999999'
            ax.text(col_x[j], y_row, cell, fontsize=6, va='center', color=colour, fontweight=weight)
        last_row_bottom = y_row - row_h / 2
    return float(last_row_bottom - 0.015)


def _draw_channel_gram_summary(
    ax: MplAxes,
    ch_df: pd.DataFrame,
    title: str,
    color: str,
    x0: float,
    x1: float,
    y_top: float,
) -> float:
    """
    Draw a compact per-group summary table for one microgel channel.
    Returns the y coordinate of the bottom edge.
    """
    ROW_H = 0.030
    span  = x1 - x0

    ax.text((x0 + x1) / 2, y_top, title,
            fontsize=8, fontweight='bold', color=color,
            va='center', ha='center')
    y = y_top - 0.026

    col_xf    = [0.00, 0.17, 0.34, 0.60, 0.80]
    col_heads = ['Group', 'N', 'Mean \u00b1 SD', '% vs ctrl', 'Status']
    ax.fill_between([x0, x1], y - ROW_H / 2, y + ROW_H / 2, color=color, alpha=0.85)
    for xf, h in zip(col_xf, col_heads):
        ax.text(x0 + xf * span + 0.005, y, h,
                fontsize=5.5, fontweight='bold', color='white', va='center')
    y -= ROW_H

    if ch_df.empty:
        ax.text((x0 + x1) / 2, y, "No data",
                ha='center', va='center', fontsize=7,
                color='#999999', style='italic')
        return y - ROW_H

    df = ch_df.copy()
    df['_sk'] = df['Group'].apply(
        lambda g: (1, 999) if str(g) == 'Control'
        else (0, int(g) if str(g).isdigit() else 999)
    )
    for i, (_, row) in enumerate(df.sort_values('_sk').iterrows()):
        is_ctrl = str(row['Group']) == 'Control'
        y_row   = y - i * ROW_H
        ax.fill_between([x0, x1], y_row - ROW_H / 2, y_row + ROW_H / 2,
                        color='#EEF4FF' if is_ctrl
                        else ('#F8F8F8' if i % 2 == 0 else '#FFFFFF'))

        cls_raw = str(row.get('Classification', ''))
        if is_ctrl:
            det, st_clr = 'CTRL',         '#555555'
        elif 'NO OBVIOUS' in cls_raw.upper() or 'NOT DETECTED' in cls_raw.upper():
            det, st_clr = 'Not Detected', '#006100'
        elif 'DETECTED' in cls_raw.upper():
            det, st_clr = 'Detected',     '#9C0006'
        else:
            det, st_clr = '\u2014',        '#777777'

        pct_raw = row.get('Pct_Diff_from_Control', '\u2014')
        try:
            pct_str = f"{float(pct_raw):+.1f}%"
        except (TypeError, ValueError):
            pct_str = '\u2014'

        cells       = [
            str(row.get('Group', '')),
            str(row.get('N', '')),
            f"{float(row.get('Mean', 0)):.2f} \u00b1 {float(row.get('Std_Dev', 0)):.2f}",
            pct_str,
            det,
        ]
        cell_colors = ['#333333', '#555555', '#333333', '#555555', st_clr]
        for xf, cell, cc in zip(col_xf, cells, cell_colors):
            ax.text(x0 + xf * span + 0.005, y_row, cell,
                    fontsize=5.5, va='center', color=cc,
                    fontweight='bold' if (cc == st_clr and not is_ctrl) else 'normal')

    return y - len(df) * ROW_H - 0.008


def generate_laboratory_report_pdf(
    output_root:          Path,
    config:               dict,
    gplus_classification: pd.DataFrame,
    gminus_classification: pd.DataFrame,
    final_df:             pd.DataFrame,
    include_page5:        bool = True,
    page5_data:           Optional[dict] = None,
) -> Optional[Path]:
    """
    Generate the main A4 laboratory report PDF (pages 1–5).

    Parameters
    ----------
    include_page5 : bool
        When False, page 5 (Methodology & QC) is always omitted.
        When True (default), page 5 is included unless page5_data is
        supplied and is_page5_empty() returns True for it.
    page5_data : dict | None
        Optional dict with keys such as 'methodology', 'quality_control',
        'limitations', 'performed_by', 'reviewed_by'.
        Pass an empty/blank dict to suppress page 5 dynamically.
        Pass None (default) to always include page 5.
    """
    # ── Decide page count ────────────────────────────────────────────────
    _emit_page5 = include_page5
    if _emit_page5 and page5_data is not None:
        _emit_page5 = not is_page5_empty(page5_data)
    total_pages = 5 if _emit_page5 else 4

    pdf_path   = output_root / "laboratory_report.pdf"
    timestamp  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    dataset_id = config.get('dataset_id_base', config.get('dataset_id', 'Unknown'))
    HEADER_BG  = "#1B3A5C"
    ACCENT     = "#2E75B6"
    LIGHT_GRAY = "#F2F2F2"
    A4_W, A4_H = 8.27, 11.69
    RESULT_COLOURS = {
        'POSITIVE':            ("#FFC7CE", "#9C0006"),
        'NEGATIVE':            ("#C6EFCE", "#006100"),
        'NO OBVIOUS BACTERIA': ("#FFEB9C", "#9C6500"),
        'MIXED/CONTRADICTORY': ("#FCD5B4", "#974706"),
        'CONTROL (Reference)': ("#E7E6E6", "#333333"),
        'MISSING DATA':        ("#D9D9D9", "#666666"),
    }

    pos_out = config.get('positive_output')
    neg_out = config.get('negative_output')
    gplus_report  = _parse_confidence_report((pos_out / 'confidence_report.txt') if isinstance(pos_out, Path) else None)
    gminus_report = _parse_confidence_report((neg_out / 'confidence_report.txt') if isinstance(neg_out, Path) else None)
    if not gplus_report['best_match'] and not gminus_report['best_match']:
        gplus_report = _parse_confidence_report(output_root / 'confidence_report.txt')

    bacteria_config = config.get('bacteria_config')
    chosen_config_name = (
        str(getattr(bacteria_config, 'name'))
        if getattr(bacteria_config, 'name', None) is not None
        else config.get('bacteria_config_info', {})
                   .get('config_names', {})
                   .get(config.get('bacteria_config_info', {})
                              .get('bacteria_type', ''),
                        config.get('bacteria_config_info', {})
                              .get('bacteria_type', ''))
    )

    def _read_comparison_csv(d):
        return (pd.read_csv(d / "configuration_comparison.csv")
                if isinstance(d, Path) and (d / "configuration_comparison.csv").exists()
                else pd.DataFrame())

    pos_comp_df = _read_comparison_csv(config.get('positive_output'))
    neg_comp_df = _read_comparison_csv(config.get('negative_output'))
    if pos_comp_df.empty and neg_comp_df.empty:
        pos_comp_df = _read_comparison_csv(output_root)

    def _make_ax(fig):
        ax = fig.add_axes((0, 0, 1, 1))
        ax.set_xlim(0, 1); ax.set_ylim(0, 1); ax.axis('off')
        return ax

    def _draw_header(ax, subtitle=""):
        ax.fill_between([0, 1], 0.92, 1.0, color=HEADER_BG)
        ax.text(0.05, 0.965, "MICROGEL FLUORESCENCE ANALYSIS",
                fontsize=16, fontweight='bold', color='white',
                va='center', family='sans-serif')
        ax.text(0.05, 0.935, "Laboratory Report",
                fontsize=11, color='#A0C4E8', va='center', family='sans-serif')
        if subtitle:
            ax.text(0.95, 0.935, subtitle, fontsize=9, color='#A0C4E8',
                    va='center', ha='right', family='sans-serif')
        ax.fill_between([0, 1], 0.915, 0.92, color=ACCENT)

    def _draw_footer(ax, p, t):
        ax.plot([0.05, 0.95], [0.04, 0.04], color='#CCCCCC', linewidth=0.5)
        ax.text(0.05, 0.025, f"Generated: {timestamp}",
                fontsize=7, color='#999999', va='center')
        ax.text(0.5,  0.025, f"Page {p} of {t}",
                fontsize=7, color='#999999', va='center', ha='center')
        ax.text(0.95, 0.025, "CONFIDENTIAL",
                fontsize=7, color='#CC0000', va='center',
                ha='right', fontweight='bold')

    _TBL_COL_X = [0.05, 0.09, 0.32, 0.50, 0.67, 0.77, 0.87]
    _TBL_COL_W = [0.04, 0.23, 0.18, 0.17, 0.10, 0.10, 0.08]
    _TBL_HEADERS = ['#', 'Bacteria Type', 'Part./img \u00b1SD',
                    'Mean Fluor \u00b1SD', 'd', 'p-val', 'Score']
    _COMP_ROW_H = 0.022

    def _draw_comparison_table(ax, comp_df, title, t_color, rpt, y_start):
        st    = rpt.get('status', '')
        s_lbl = ("AMBIGUOUS"    if 'AMBIGUOUS'   in st.upper() else
                 "NO BACTERIA"  if 'NO BACTERIA' in st.upper() else
                 "DETECTED"     if st            else '')
        s_clr = ('#974706' if s_lbl == 'AMBIGUOUS'
                 else '#555555' if s_lbl == 'NO BACTERIA'
                 else '#006100')
        ax.text(0.07, y_start, title, fontsize=8.5,
                fontweight='bold', color=t_color, va='center')
        if s_lbl:
            ax.text(0.93, y_start, s_lbl, fontsize=7.5,
                    fontweight='bold', color=s_clr, va='center', ha='right')
        y = y_start - 0.020
        ax.fill_between([0.05, 0.95], y - _COMP_ROW_H / 2, y + _COMP_ROW_H / 2,
                        color=HEADER_BG)
        for j, h in enumerate(_TBL_HEADERS):
            ax.text(_TBL_COL_X[j] + _TBL_COL_W[j] / 2, y, h,
                    fontsize=6, fontweight='bold', color='white',
                    va='center', ha='center')
        y -= _COMP_ROW_H
        if comp_df.empty:
            ax.text(0.50, y, "No comparison data available",
                    fontsize=7, color='#999999', va='center',
                    ha='center', style='italic')
            return y - _COMP_ROW_H * 0.6

        df_sorted = (comp_df.sort_values('Rank').reset_index(drop=True)
                     if 'Rank' in comp_df.columns
                     else comp_df.reset_index(drop=True))

        for i, (_, row) in enumerate(df_sorted.iterrows()):
            ax.fill_between([0.05, 0.95], y - _COMP_ROW_H / 2, y + _COMP_ROW_H / 2,
                            color='#EEF4FF' if i == 0 else '#F8F8F8')

            def _fv(k, default=0):
                v = row.get(k, default)
                try: return float(str(v).replace('\u00b1','').strip())
                except Exception: return 0.0

            ppi_str  = f"{_fv('Particles_Per_Image'):.1f} \u00b1 {_fv('Particles_Std'):.1f}"
            mfl_str  = f"{_fv('Mean_Fluorescence'):.1f} \u00b1 {_fv('Fluor_Std'):.1f}"
            p_raw    = row.get('P_Value_TvC')
            try:    p_flt = float(str(p_raw))
            except Exception: p_flt = np.nan
            p_str    = ('N/A' if pd.isna(p_flt)
                        else '<0.001' if p_flt < 0.001
                        else f"{p_flt:.4f}")
            cs_raw   = row.get('Confidence_Score', 0)
            try:    cs_flt = float(str(cs_raw).strip())
            except Exception: cs_flt = 0.0
            pen_flag = bool(row.get('Penalty_Applied', False))
            score_str = f"{cs_flt:.1f}%" + (" \u2020" if pen_flag else "")
            d_flt    = _fv('Cohens_d')

            cells  = [str(int(row.get('Rank', i + 1))),
                      str(row.get('Bacteria_Type', '\u2014')),
                      ppi_str, mfl_str,
                      f"{d_flt:+.3f}", p_str, score_str]
            colors = [
                '#333333',
                '#1B3A5C' if i == 0 else '#333333',
                '#333333', '#333333',
                '#006100' if d_flt > 0.5 else '#9C0006' if d_flt < -0.5 else '#333333',
                '#333333',
                '#006100' if cs_flt >= 70 else '#9C6500' if cs_flt >= 50 else '#9C0006',
            ]
            for j, (cell, cc) in enumerate(zip(cells, colors)):
                ax.text(_TBL_COL_X[j] + _TBL_COL_W[j] / 2, y, cell,
                        fontsize=6, va='center', ha='center', color=cc,
                        fontweight='bold' if (i == 0 and j <= 1) else 'normal')
            y -= _COMP_ROW_H

        if 'Penalty_Applied' in df_sorted.columns and df_sorted['Penalty_Applied'].any():
            ax.text(0.07, y + _COMP_ROW_H * 0.35,
                    "\u2020 pairwise convergence penalty applied",
                    fontsize=5.5, color='#777777', va='center', style='italic')
            y -= 0.010
        return y

    try:
        with PdfPages(str(pdf_path)) as pdf:

            # ── PAGE 1: Gram-type clinical summary ─────────────────────
            fig1 = plt.figure(figsize=(A4_W, A4_H))
            ax   = _make_ax(fig1)
            _draw_header(ax, "Gram-Type Clinical Report")
            _draw_footer(ax, 1, total_pages)

            ax.fill_between([0.05, 0.95], 0.82, 0.90, color=LIGHT_GRAY)
            ax.plot([0.05, 0.95, 0.95, 0.05, 0.05],
                    [0.90, 0.90, 0.82, 0.82, 0.90], color='#999999', lw=0.5)
            for i, (lbl, val) in enumerate([
                ("Sample ID:",           dataset_id),
                ("Date:",                timestamp),
                ("Segmentation config:", chosen_config_name or "Auto-selected"),
                ("Percentile filter:",   f"{config.get('percentile', 0.3) * 100:.0f}%"),
                ("Clinical threshold:",  f"\u2212{config.get('threshold_pct', 0.05) * 100:.0f}% vs control"),
            ]):
                x_base = 0.07 if i < 3 else 0.52
                y_pos  = 0.885 - (i % 3) * 0.022
                ax.text(x_base,        y_pos, lbl,       fontsize=8, fontweight='bold', va='center', color='#333333')
                ax.text(x_base + 0.15, y_pos, str(val),  fontsize=8, va='center', color='#333333')

            ax.text(0.05, 0.800, "GRAM-TYPE CHANNEL ANALYSIS",
                    fontsize=11, fontweight='bold', color=HEADER_BG)
            ax.plot([0.05, 0.95], [0.788, 0.788], color=ACCENT, lw=1)

            _draw_channel_gram_summary(
                ax, gplus_classification,
                "G+  Microgel  (Gram-Positive)", ACCENT, 0.05, 0.475, 0.772,
            )
            _draw_channel_gram_summary(
                ax, gminus_classification,
                "G\u2212  Microgel  (Gram-Negative)", "#C0504D", 0.525, 0.95, 0.772,
            )

            RESULTS_Y = 0.480
            ax.text(0.05, RESULTS_Y, "FINAL RESULTS",
                    fontsize=13, fontweight='bold', color=HEADER_BG, va='center')
            ax.plot([0.05, 0.95], [RESULTS_Y - 0.012, RESULTS_Y - 0.012],
                    color=ACCENT, lw=1.5)

            col_headers = ['Group', 'G+ Mean\u00b1SD', 'G+ Detection',
                           'G\u2212 Mean\u00b1SD', 'G\u2212 Detection', 'Final']
            col_widths  = [0.10, 0.18, 0.14, 0.18, 0.14, 0.18]
            col_x = [0.05]
            for w in col_widths[:-1]:
                col_x.append(col_x[-1] + w)

            for j, h in enumerate(col_headers):
                ax.fill_between([col_x[j], col_x[j] + col_widths[j]],
                                RESULTS_Y - 0.058, RESULTS_Y - 0.030, color=HEADER_BG)
                ax.text(col_x[j] + col_widths[j] / 2, RESULTS_Y - 0.044, h,
                        fontsize=7, fontweight='bold', color='white',
                        va='center', ha='center')

            for i, (_, row) in enumerate(final_df.iterrows()):
                y  = RESULTS_Y - 0.030 - (i + 1) * 0.028
                fc = row.get('Final_Classification', '')
                bg, tc = RESULT_COLOURS.get(fc, ('#FFFFFF', '#333333'))
                cells = [
                    str(row.get('Group', '')),
                    f"{row.get('G+_Mean','\u2014')}\u00b1{row.get('G+_Std','\u2014')}" if row.get('G+_Mean', '\u2014') != '\u2014' else '\u2014',
                    str(row.get('G+_Detection', '')),
                    f"{row.get('G-_Mean','\u2014')}\u00b1{row.get('G-_Std','\u2014')}" if row.get('G-_Mean', '\u2014') != '\u2014' else '\u2014',
                    str(row.get('G-_Detection', '')),
                    fc,
                ]
                for j, ct in enumerate(cells):
                    ax.fill_between(
                        [col_x[j], col_x[j] + col_widths[j]], y - 0.028, y,
                        color=bg if j == len(cells) - 1 else LIGHT_GRAY,
                    )
                    ax.text(col_x[j] + col_widths[j] / 2, y - 0.014, ct,
                            fontsize=7, va='center', ha='center',
                            color=tc if j == len(cells) - 1 else '#333333')

            box_y = RESULTS_Y - 0.030 - (len(final_df) + 2) * 0.028
            ax.text(0.05, box_y, "INTERPRETATION",
                    fontsize=11, fontweight='bold', color=HEADER_BG)
            ax.plot([0.05, 0.95], [box_y - 0.010, box_y - 0.010], color=ACCENT, lw=1)
            interp_y = box_y - 0.035
            for _, row in final_df.iterrows():
                if row.get('Group', '') == 'Control':
                    continue
                if interp_y - 0.030 < 0.10:
                    break
                sym, clr = {
                    'NEGATIVE':            ("\u25cf", "#006100"),
                    'POSITIVE':            ("\u25cf", "#9C0006"),
                    'NO OBVIOUS BACTERIA': ("\u25cb", "#9C6500"),
                    'MIXED/CONTRADICTORY': ("\u25c6", "#974706"),
                }.get(row.get('Final_Classification', ''), ("?", "#666666"))
                ax.text(0.07, interp_y, sym, fontsize=10, color=clr,
                        va='center', ha='center')
                ax.text(0.09, interp_y,
                        f"Group {row.get('Group', '')}: "
                        f"{str(row.get('Interpretation', row.get('Final_Classification', ''))).split('[')[0].strip()}",
                        fontsize=8.5, color='#333333', va='center', fontweight='bold')
                interp_y -= 0.030

            pdf.savefig(fig1)
            plt.close(fig1)

            # ── PAGE 2: Comparison charts ───────────────────────────────
            fig2 = plt.figure(figsize=(A4_W, A4_H))
            _draw_chart_on_axis(
                fig2.add_axes((0.10, 0.55, 0.82, 0.33)),
                gplus_classification, "G+ Microgel (Positive)", ACCENT,
                config.get('positive_output'), config.get('threshold_pct', 0.05),
            )
            _draw_chart_on_axis(
                fig2.add_axes((0.10, 0.10, 0.82, 0.33)),
                gminus_classification, "G\u2212 Microgel (Negative)", "#C0504D",
                config.get('negative_output'), config.get('threshold_pct', 0.05),
            )
            ax2 = _make_ax(fig2)
            _draw_header(ax2, "Comparison Charts")
            _draw_footer(ax2, 2, total_pages)
            pdf.savefig(fig2)
            plt.close(fig2)

            # ── PAGE 3: Forest plot ─────────────────────────────────────
            fig3 = plt.figure(figsize=(A4_W, A4_H))
            _draw_forest_plot_on_axis(
                fig3.add_axes((0.12, 0.10, 0.78, 0.74)),
                gplus_classification, gminus_classification,
            )
            ax3 = _make_ax(fig3)
            _draw_header(ax3, "Effect Size Analysis")
            _draw_footer(ax3, 3, total_pages)
            pdf.savefig(fig3)
            plt.close(fig3)

            # ── PAGE 4: Statistical analysis table ─────────────────────
            fig4 = plt.figure(figsize=(A4_W, A4_H))
            ax4  = _make_ax(fig4)
            _draw_header(ax4, "Statistical Analysis")
            _draw_footer(ax4, 4, total_pages)
            _draw_stats_table(ax4, gplus_classification,  "G+ Microgel",      0.87, 0.55)
            _draw_stats_table(ax4, gminus_classification, "G\u2212 Microgel", 0.53, 0.26)
            ax4.plot([0.05, 0.95], [0.228, 0.228], color='#CCCCCC', lw=0.5)
            ax4.text(0.05, 0.22, "Significance Key", fontsize=9,
                     fontweight='bold', color=HEADER_BG)
            for i, line in enumerate([
                "*** p < 0.001 (highly significant)",
                "**  p < 0.01  (very significant)",
                "*   p < 0.05  (significant)",
                "ns  p \u2265 0.05  (not significant)",
                "Effect size: |d| < 0.2 negligible, 0.2\u20130.5 small, "
                "0.5\u20130.8 medium, > 0.8 large (Cohen\u2019s d)",
            ]):
                ax4.text(0.07, 0.198 - i * 0.018, line,
                         fontsize=7, color='#555555', family='monospace')
            pdf.savefig(fig4)
            plt.close(fig4)

            # ── PAGE 5: Methodology & QC  (conditional) ────────────────
            if _emit_page5:
                fig5 = plt.figure(figsize=(A4_W, A4_H))
                ax5  = _make_ax(fig5)
                _draw_header(ax5, "Methodology & Quality Control")
                _draw_footer(ax5, 5, total_pages)

                for t, ys, lines in [
                    ("METHODOLOGY", 0.88, [
                        "Brightfield and fluorescence images were acquired using a Leica microscope.",
                        "Particles were segmented from brightfield (ch00) using Gaussian blur and intensity thresholding.",
                        "Fluorescence intensity (ch01) was measured within brightfield contours.",
                        "The primary metric is Fluorescence Integrated Density / BF Area (a.u./\u00b5m\u00b2).",
                        f"Typical particles: middle {100 - 2 * int(config.get('percentile', 0.3) * 100)}% "
                        f"(top/bottom {int(config.get('percentile', 0.3) * 100)}% excluded).",
                        f"Clinical threshold = Control Mean \u00d7 (1 \u2212 {config.get('threshold_pct', 0.05) * 100:.0f}%).",
                        "Groups whose typical-particle mean falls below threshold are classified as \u2018Bacteria Detected\u2019.",
                        "Cohen\u2019s d effect size computed with pooled SD; "
                        "95% CI via Hedges\u2013Olkin approximation.",
                    ]),
                    ("QUALITY CONTROL", 0.56, [
                        f"Pixel size derived from Leica XML metadata "
                        f"(fallback: {FALLBACK_UM_PER_PX} \u00b5m/px).",
                        "Alignment: phase cross-correlation; fallback to no-shift if correlation error > 0.5.",
                        "Excluded objects logged in each group\u2019s master Excel (Excluded_Objects sheet).",
                    ]),
                    ("LIMITATIONS", 0.43, [
                        "This assay is for research use only.",
                        "Results should be confirmed by culture-based methods.",
                        "Classification depends on threshold chosen and control baseline.",
                        "Low particle counts (N < 5) may yield unreliable statistics.",
                    ]),
                ]:
                    ax5.text(0.05, ys, t, fontsize=11,
                             fontweight='bold', color=HEADER_BG)
                    ax5.plot([0.05, 0.95], [ys - 0.01, ys - 0.01],
                             color=ACCENT, lw=1)
                    for i, l in enumerate(lines):
                        ax5.text(0.07, ys - 0.03 - i * 0.02,
                                 f"\u2022 {l}", fontsize=7.5, color='#333333')

                ax5.plot([0.05, 0.95], [0.17, 0.17], color='#CCCCCC', lw=0.5)
                ax5.text(0.05, 0.15, "APPROVAL", fontsize=10,
                         fontweight='bold', color=HEADER_BG)
                for lbl, xp in [("Performed by:", 0.05),
                                 ("Reviewed by:",  0.38),
                                 ("Date:",         0.71)]:
                    ax5.text(xp, 0.12, lbl, fontsize=8,
                             fontweight='bold', color='#555555')
                    ax5.plot([xp, xp + 0.25], [0.08, 0.08],
                             color='#333333', lw=0.8)

                pdf.savefig(fig5)
                plt.close(fig5)

        return pdf_path
    except Exception:
        return None


# ==================================================
# Excel Export & Cleanup
# ==================================================

def consolidate_to_excel(output_dir: Path, group_name: str, percentile: float) -> None:
    csv_files = list(output_dir.glob("*/object_stats.csv"))
    if not csv_files: return
    excel_path = output_dir / f"{group_name}_master.xlsx"
    if excel_path.exists():
        try: excel_path.unlink()
        except PermissionError: return

    try:
        yellow_fill, red_fill, header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"), PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"), PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font, center_align = Font(bold=True, color="FFFFFF"), Alignment(horizontal="center", vertical="center")

        def adjust_column_widths(ws):
            for column in ws.columns:
                ws.column_dimensions[get_column_letter(column[0].column)].width = min((max((len(str(cell.value)) for cell in column if cell.value), default=0) + 2) * 1.1, 50)
        def format_numbers(ws):
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)): cell.number_format = '0.0000'

        all_valid_objects, all_excluded_objects, all_rejected_objects = [], [], []

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            pd.DataFrame({"Column Name": ["Object_ID", "BF_Area_px", "BF_Area_um2", "Perimeter_px", "Perimeter_um", "EquivDiameter_px", "EquivDiameter_um", "Circularity", "AspectRatio", "CentroidX_px", "CentroidY_px", "CentroidX_um", "CentroidY_um", "BBoxX_px", "BBoxY_px", "BBoxW_px", "BBoxH_px", "BBoxW_um", "BBoxH_um", "Fluor_Area_px", "Fluor_Area_um2", "Fluor_Mean", "Fluor_Median", "Fluor_Std", "Fluor_Min", "Fluor_Max", "Fluor_IntegratedDensity", "Fluor_Density_per_BF_Area", "BF_to_Fluor_Area_Ratio"], "Description": ["Unique particle identifier", "Brightfield particle area (pixels\u00b2)", "Brightfield particle area (\u00b5m\u00b2)", "Particle perimeter (pixels)", "Particle perimeter (\u00b5m)", "Diameter of equivalent circle (pixels)", "Diameter of equivalent circle (\u00b5m)", "Shape roundness (0-1)", "Bounding box width/height ratio", "Particle center X (px)", "Particle center Y (px)", "Particle center X (\u00b5m)", "Particle center Y (\u00b5m)", "BBox top-left X (px)", "BBox top-left Y (px)", "BBox width (px)", "BBox height (px)", "BBox width (\u00b5m)", "BBox height (\u00b5m)", "Fluorescent region area (pixels\u00b2)", "Fluorescent region area (\u00b5m\u00b2)", "Avg fluorescence intensity", "Median fluorescence intensity", "Std Dev fluorescence", "Min fluorescence", "Max fluorescence", "Total fluorescence signal", "Fluor density / BF Area (Primary Metric)", "Ratio of BF area to Fluor area"]}).to_excel(writer, sheet_name="README", index=False)
            ws_readme = writer.sheets["README"]
            for cell in ws_readme[1]: cell.fill, cell.font, cell.alignment = header_fill, header_font, center_align
            adjust_column_widths(ws_readme)

            for csv_file in sorted(csv_files):
                image_name = csv_file.parent.name
                df = pd.read_csv(csv_file)
                for col in ["Fluor_Area_px", "Fluor_IntegratedDensity", "BF_Area_um2", "Fluor_Area_um2"]:
                    if col in df.columns: df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
                df["Fluor_Density_per_BF_Area"] = df["Fluor_IntegratedDensity"] / df["BF_Area_um2"] if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns else 0.0
                df["BF_to_Fluor_Area_Ratio"] = np.where(df["BF_Area_um2"] > 0, df["BF_Area_um2"] / df["Fluor_Area_um2"], 0.0) if "Fluor_Area_um2" in df.columns and "BF_Area_um2" in df.columns else 0.0
                df = df.replace([np.inf, -np.inf], 0).fillna(0)

                for _, row in df.iterrows():
                    reason = "Zero fluorescence area with positive integrated density" if row["Fluor_Area_px"] == 0 and row["Fluor_IntegratedDensity"] > 0 else "Zero fluorescence area" if row["Fluor_Area_px"] == 0 else "Zero integrated density" if row["Fluor_IntegratedDensity"] == 0 else None
                    if reason: all_excluded_objects.append({"Object_ID": row["Object_ID"], "Source_Image": image_name, "BF_Area_um2": row["BF_Area_um2"], "Fluor_Area_px": row["Fluor_Area_px"], "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"], "Exclusion_Reason": reason})

                df_valid = df[(df["Fluor_IntegratedDensity"] > 0) & (df["Fluor_Area_px"] > 0)].copy()
                df_valid["Source_Image"] = image_name
                if not df_valid.empty: all_valid_objects.append(df_valid)

                if "Fluor_Density_per_BF_Area" in df.columns: df = df.sort_values("Fluor_Density_per_BF_Area", ascending=False)
                df.to_excel(writer, sheet_name=image_name[:31], index=False)
                ws = writer.sheets[image_name[:31]]
                for cell in ws[1]: cell.fill, cell.font, cell.alignment = header_fill, header_font, center_align
                format_numbers(ws); adjust_column_widths(ws); ws.auto_filter.ref = ws.dimensions

                rejected_csv = csv_file.parent / "rejected_objects.csv"
                if rejected_csv.exists():
                    try:
                        rej_df = pd.read_csv(rejected_csv)
                        if not rej_df.empty: rej_df['Source_Image'] = image_name; all_rejected_objects.append(rej_df)
                    except Exception: pass

            if all_valid_objects:
                merged_all = pd.concat(all_valid_objects, ignore_index=True).sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)
                merged_all.to_excel(writer, sheet_name=f"{group_name}_All_Valid_Objects", index=False)
                ws_all = writer.sheets[f"{group_name}_All_Valid_Objects"]
                for cell in ws_all[1]: cell.fill, cell.font, cell.alignment = header_fill, header_font, center_align
                format_numbers(ws_all); adjust_column_widths(ws_all); ws_all.auto_filter.ref = ws_all.dimensions

                n_cut = int(len(merged_all) * percentile)
                if n_cut < len(merged_all) - n_cut and len(merged_all) > 3:
                    typical_particles = merged_all.iloc[n_cut:len(merged_all) - n_cut].copy()
                    for _, row in merged_all.iloc[:n_cut].iterrows(): all_excluded_objects.append({"Object_ID": row["Object_ID"], "Source_Image": row["Source_Image"], "BF_Area_um2": row["BF_Area_um2"], "Fluor_Area_px": row["Fluor_Area_px"], "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"], "Exclusion_Reason": f"Outside typical particle range (top {int(percentile*100)}%)"})
                    for _, row in merged_all.iloc[len(merged_all) - n_cut:].iterrows(): all_excluded_objects.append({"Object_ID": row["Object_ID"], "Source_Image": row["Source_Image"], "BF_Area_um2": row["BF_Area_um2"], "Fluor_Area_px": row["Fluor_Area_px"], "Fluor_IntegratedDensity": row["Fluor_IntegratedDensity"], "Exclusion_Reason": f"Outside typical particle range (bottom {int(percentile*100)}%)"})
                else: typical_particles = merged_all.copy()

                typical_particles.to_excel(writer, sheet_name=f"{group_name}_Typical_Particles", index=False)
                ws_typ = writer.sheets[f"{group_name}_Typical_Particles"]
                for cell in ws_typ[1]: cell.fill, cell.font, cell.alignment = header_fill, header_font, center_align
                for row in ws_typ.iter_rows(min_row=2, max_row=ws_typ.max_row):
                    for cell in row: cell.fill = yellow_fill
                format_numbers(ws_typ); adjust_column_widths(ws_typ); ws_typ.auto_filter.ref = ws_typ.dimensions

            if all_excluded_objects:
                pd.DataFrame(all_excluded_objects).to_excel(writer, sheet_name="Excluded_Objects", index=False)
                ws_excluded = writer.sheets["Excluded_Objects"]
                for cell in ws_excluded[1]: cell.fill, cell.font, cell.alignment = header_fill, header_font, center_align
                for row in ws_excluded.iter_rows(min_row=2, max_row=ws_excluded.max_row):
                    for cell in row: cell.fill = red_fill
                adjust_column_widths(ws_excluded); ws_excluded.auto_filter.ref = ws_excluded.dimensions

            if all_rejected_objects:
                merged_rejected = pd.concat(all_rejected_objects, ignore_index=True)
                if 'BF_Area_px' in merged_rejected.columns: merged_rejected = merged_rejected.sort_values('BF_Area_px', ascending=False).reset_index(drop=True)
                merged_rejected.to_excel(writer, sheet_name=f"{group_name}_Rejected_Objects", index=False)
                ws_rej = writer.sheets[f"{group_name}_Rejected_Objects"]
                for cell in ws_rej[1]: cell.fill, cell.font, cell.alignment = header_fill, header_font, center_align
                for row in ws_rej.iter_rows(min_row=2, max_row=ws_rej.max_row):
                    for cell in row: cell.fill = red_fill
                format_numbers(ws_rej); adjust_column_widths(ws_rej); ws_rej.auto_filter.ref = ws_rej.dimensions

            try:
                summary_data = []
                for csv_file in sorted(csv_files):
                    df = pd.read_csv(csv_file)
                    df["Fluor_Density_per_BF_Area"] = pd.to_numeric(df["Fluor_IntegratedDensity"], errors='coerce') / pd.to_numeric(df["BF_Area_um2"], errors='coerce') if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns else 0
                    if "Fluor_Area_px" in df.columns: df["Fluor_Area_px"] = pd.to_numeric(df["Fluor_Area_px"], errors='coerce').fillna(0)
                    df = df.replace([np.inf, -np.inf], 0).fillna(0)
                    df_stats = df[(df["Fluor_Area_px"] > 0) & (df["Fluor_IntegratedDensity"] > 0)]
                    summary_data.append({"Image": csv_file.parent.name, "Total_Particles_Detected": len(df), "Particles_With_Fluor": len(df_stats), "Avg_BF_Area_um2": df["BF_Area_um2"].mean() if "BF_Area_um2" in df.columns else 0, "Avg_Fluor_Density": df_stats["Fluor_Density_per_BF_Area"].mean() if not df_stats.empty else 0.0, "Avg_BF_to_Fluor_Ratio": (df_stats["BF_Area_um2"] / df_stats["Fluor_Area_um2"]).replace([np.inf, -np.inf], 0).mean() if not df_stats.empty and "BF_Area_um2" in df_stats.columns and "Fluor_Area_um2" in df_stats.columns else 0.0})
                summary_df = pd.DataFrame(summary_data)
                if "Avg_Fluor_Density" in summary_df.columns: summary_df = summary_df.sort_values("Avg_Fluor_Density", ascending=False)
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                ws_summary = writer.sheets["Summary"]
                for cell in ws_summary[1]: cell.fill, cell.font, cell.alignment = header_fill, header_font, center_align
                format_numbers(ws_summary); adjust_column_widths(ws_summary)
            except Exception: pass

            try:
                ws_qa = writer.book.create_sheet("Ratios")
                ws_qa["A1"] = f"QA Ratios - Group: {group_name}"; ws_qa["A1"].font = Font(bold=True, size=14)
                row = 3
                for csv_file in sorted(csv_files):
                    df = pd.read_csv(csv_file)
                    df["Fluor_Density_per_BF_Area"] = pd.to_numeric(df["Fluor_IntegratedDensity"], errors='coerce') / pd.to_numeric(df["BF_Area_um2"], errors='coerce') if "Fluor_IntegratedDensity" in df.columns and "BF_Area_um2" in df.columns else 0.0
                    df["BF_to_Fluor_Area_Ratio"] = pd.to_numeric(df["BF_Area_um2"], errors='coerce') / pd.to_numeric(df["Fluor_Area_um2"], errors='coerce') if "BF_Area_um2" in df.columns and "Fluor_Area_um2" in df.columns else 0.0
                    df = df.replace([np.inf, -np.inf], 0).fillna(0)
                    if "Fluor_Density_per_BF_Area" in df.columns: df = df.sort_values("Fluor_Density_per_BF_Area", ascending=False).reset_index(drop=True)
                    ws_qa[f"A{row}"] = csv_file.parent.name; ws_qa[f"A{row}"].font = Font(bold=True, size=12)
                    for col_idx, header in enumerate(["Object_ID", "Fluor_Density_per_BF_Area", "Rank", "Object_ID", "BF_to_Fluor_Area_Ratio"], 1):
                        cell = ws_qa.cell(row=row+1, column=col_idx, value=header); cell.font, cell.alignment = Font(bold=True), center_align
                    for k, r in enumerate(df.itertuples(index=False), 0):
                        ws_qa.cell(row=row+2+k, column=1, value=getattr(r, "Object_ID")); ws_qa.cell(row=row+2+k, column=2, value=float(getattr(r, "Fluor_Density_per_BF_Area", 0.0))).number_format = '0.0000'; ws_qa.cell(row=row+2+k, column=3, value=k+1); ws_qa.cell(row=row+2+k, column=4, value=getattr(r, "Object_ID")); ws_qa.cell(row=row+2+k, column=5, value=float(getattr(r, "BF_to_Fluor_Area_Ratio", 0.0))).number_format = '0.0000'
                    ws_qa.column_dimensions["A"].width = ws_qa.column_dimensions["B"].width = 20
                    ws_qa.column_dimensions["C"].width = ws_qa.column_dimensions["D"].width = ws_qa.column_dimensions["E"].width = 15
                    if len(df) > 0:
                        ch1 = ScatterChart(); ch1.title, ch1.y_axis.title, ch1.x_axis.title = "Fluor Density", "a.u./\u00b5m\u00b2", "Object Rank (Sorted)"
                        s1 = SeriesFactory(Reference(ws_qa, min_col=2, min_row=row+2, max_row=row+2+len(df)-1), Reference(ws_qa, min_col=3, min_row=row+2, max_row=row+2+len(df)-1), title="Density"); s1.marker = Marker(symbol="triangle", size=5); ch1.series.append(s1); ws_qa.add_chart(ch1, f"G{row+1}")
                        ch2 = ScatterChart(); ch2.title, ch2.y_axis.title, ch2.x_axis.title = "Area Ratio", "Ratio", "Object Rank (Sorted)"
                        s2 = SeriesFactory(Reference(ws_qa, min_col=5, min_row=row+2, max_row=row+2+len(df)-1), Reference(ws_qa, min_col=3, min_row=row+2, max_row=row+2+len(df)-1), title="Ratio"); s2.marker = Marker(symbol="circle", size=5); ch2.series.append(s2); ws_qa.add_chart(ch2, f"G{row+18}")
                    for path, cell in [(csv_file.parent / "13_mask_accepted_ids.png", f"Q{row+1}"), (csv_file.parent / "24_bf_fluor_matching_overlay_ids.png", f"Q{row+1}")]:
                        if path.exists():
                            img = XLImage(str(path))
                            if getattr(img, "width", None) and getattr(img, "height", None): img.width, img.height = int(img.width * (330 / float(img.width))), int(img.height * (330 / float(img.width)))
                            ws_qa.add_image(img, cell)
                    row += 34
            except Exception: pass

        wb2 = load_workbook(excel_path)
        desired_order = ["Summary", "Ratios", "README", f"{group_name}_Typical_Particles", f"{group_name}_All_Valid_Objects", "Excluded_Objects"]
        if f"{group_name}_Rejected_Objects" in wb2.sheetnames: desired_order.insert(5, f"{group_name}_Rejected_Objects")
        for idx, sheet_name in enumerate(desired_order):
            if sheet_name in wb2.sheetnames:
                if wb2.sheetnames.index(sheet_name) != idx: wb2.move_sheet(wb2[sheet_name], offset=idx - wb2.sheetnames.index(sheet_name))
        wb2.save(excel_path)
    except Exception: pass

def export_group_statistics_to_csv(output_root: Path) -> None:
    stats_list = []
    for excel_path in sorted(output_root.glob("*/*_master.xlsx")):
        try:
            df = pd.read_excel(excel_path, sheet_name=f"{excel_path.parent.name}_Typical_Particles")
            if "Fluor_Density_per_BF_Area" in df.columns:
                values = pd.to_numeric(df["Fluor_Density_per_BF_Area"], errors='coerce').dropna()
                if not values.empty:
                    mean_val, std_val = float(values.mean()), float(values.std(ddof=1))
                    stats_list.append({'Group': "Control" if excel_path.parent.name.lower().startswith("control") else excel_path.parent.name, 'N': int(len(values)), 'Mean': mean_val, 'Std_Dev': std_val, 'SEM': values.sem(), 'Median': float(values.median()), 'Q30': float(values.quantile(0.30)), 'Q70': float(values.quantile(0.70)), 'Min': float(values.min()), 'Max': float(values.max()), 'CV_percent': (std_val / mean_val * 100) if mean_val > 0 else 0})
        except Exception: pass
    if stats_list:
        stats_df = pd.DataFrame(stats_list)
        stats_df['sort_key'] = stats_df['Group'].apply(lambda x: (0, int(x)) if x.isdigit() else (1, 999))
        stats_df = stats_df.sort_values('sort_key').drop('sort_key', axis=1)
        numeric_cols = stats_df.select_dtypes(include=[np.number]).columns
        stats_df[numeric_cols] = stats_df[numeric_cols].round(2)
        stats_df.to_csv(output_root / "group_statistics_summary.csv", index=False)

def cleanup_and_reorganize_output(output_root: Path, config: dict) -> None:
    print("\n" + "="*80 + "\nCLEANING UP OUTPUT FOLDER\n" + "="*80)
    try:
        files_copied = 0
        for key, name in [('positive_output', 'positive'), ('negative_output', 'negative')]:
            d = config.get(key)
            if isinstance(d, Path) and d.exists():
                for f in [f"clinical_classification_{name}.xlsx", f"comparison_{name}_all_groups.png", "confidence_report.txt"]:
                    if (d / f).exists() and not (output_root / f).exists(): shutil.copy2(d / f, output_root / f); files_copied += 1
        if (output_root / "confidence_report.txt").exists(): files_copied += 1

        sub_dirs = []
        for k in ('positive_output', 'negative_output'):
            d = config.get(k)
            if isinstance(d, Path) and d.exists():
                sub_dirs.append(d)
        if not sub_dirs:
            sub_dirs = [output_root]
        for sub_dir in sub_dirs:
            for f in ["configuration_comparison.csv", "multi_config_comparison_with_statistics.png"]:
                if (sub_dir / f).exists(): (sub_dir / f).unlink()
            if (sub_dir / ".cache").exists(): shutil.rmtree(sub_dir / ".cache")
            for cfg_key in config.get('bacteria_config_info', {}).get('configs_to_scan', []):
                if (sub_dir / cfg_key).exists(): shutil.rmtree(sub_dir / cfg_key)

        if not DEBUG_MODE:
            for diag in output_root.rglob("DIAG_*.png"):
                if diag.name in ("DIAG_A_no_shift.png", "DIAG_B_positive_shift.png", "DIAG_C_negative_shift.png"):
                    try: diag.unlink()
                    except Exception: pass
    except Exception: pass

def generate_rejection_analysis(output_root: Path) -> Optional[Path]:
    all_rejections = []
    for excel_path in [p for p in output_root.rglob("*_master.xlsx") if len(p.relative_to(output_root).parts) <= 3]:
        try:
            df = pd.read_excel(excel_path, sheet_name=f"{excel_path.parent.name}_Rejected_Objects")
            if not df.empty:
                rel_parts = excel_path.relative_to(output_root).parts
                df['Group'] = f"{rel_parts[0]}/{excel_path.parent.name}" if len(rel_parts) >= 3 else excel_path.parent.name
                all_rejections.append(df)
        except Exception: pass
    if not all_rejections: return None
    merged = pd.concat(all_rejections, ignore_index=True)
    analysis_data = []
    for group in merged['Group'].unique():
        group_data = merged[merged['Group'] == group]
        reasons = []
        if 'Rejection_Reasons' in group_data.columns:
            rejection_reasons = cast(pd.Series, group_data['Rejection_Reasons'])
            for rs in rejection_reasons.dropna().astype(str).tolist():
                reasons.extend(r.strip() for r in str(rs).split(';') if r.strip())
        rc = pd.Series(reasons).value_counts()
        analysis_data.append({'Group': group, 'Total_Rejected': len(group_data), 'Mean_Area_px': group_data['BF_Area_px'].mean(), 'Mean_Circularity': group_data['Circularity'].mean(), 'Mean_AspectRatio': group_data['AspectRatio'].mean(), 'Mean_Solidity': group_data['Solidity'].mean(), 'Top_Rejection_Reason': rc.index[0] if len(rc) > 0 else 'N/A', 'Top_Reason_Count': rc.iloc[0] if len(rc) > 0 else 0})
    analysis_path = output_root / "rejection_analysis_summary.csv"
    pd.DataFrame(analysis_data).to_csv(analysis_path, index=False)
    return analysis_path

def open_folder(folder_path: Path) -> None:
    try:
        folder_str = str(folder_path.resolve())
        if sys.platform == 'win32': os.startfile(folder_str)
        elif sys.platform == 'darwin': subprocess.run(['open', folder_str])
        else: subprocess.run(['xdg-open', folder_str])
    except Exception: pass

def copy_log_to_output(log_path: Path, output_dir: Path) -> Optional[Path]:
    try:
        if log_path and log_path.exists():
            dest_path = output_dir / log_path.name
            shutil.copy2(log_path, dest_path)
            return dest_path
    except Exception: pass
    return None


# ==================================================
# Core Processing Pipeline
# ==================================================

def process_image(img_path: Path, output_root: Path, bacteria_config: 'SegmentationConfig', profiler: Optional[Any] = None) -> None:
    _img_start, _alignment_method, _shift_detected, _fluor_otsu = pytime.monotonic(), "none", (0.0, 0.0), 0.0
    accepted, rejected, rejection_reasons = [], [], []

    try:
        if not validate_path_encoding(img_path): return
        parts = img_path.stem.split()
        group_id, sequence_num = parts[0] if parts else "unk", parts[-1].split("_")[0] if parts else "0"
        xml_props, xml_main = find_metadata_paths(img_path)
        try: um_per_px_x, um_per_px_y = get_pixel_size_um(xml_props, xml_main)
        except Exception: um_per_px_x = um_per_px_y = float(FALLBACK_UM_PER_PX)
        um_per_px_avg, um2_per_px2 = (um_per_px_x + um_per_px_y) / 2.0, float(um_per_px_x) * float(um_per_px_y)

        img_out = output_root / img_path.stem
        ensure_dir(img_out)
        try:
            shutil.copy2(img_path, img_out / img_path.name)
            if (img_path.parent / img_path.name.replace("_ch00", "_ch01")).exists(): shutil.copy2(img_path.parent / img_path.name.replace("_ch00", "_ch01"), img_out / img_path.name.replace("_ch00", "_ch01"))
        except Exception: pass

        img = safe_imread(img_path, cv2.IMREAD_UNCHANGED)
        if img is None: raise FileNotFoundError(f"Could not read image: {img_path}")
        if img.ndim == 3: img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        img8 = normalize_to_8bit(img)
        save_debug(img_out, "01_gray_8bit.png", img8, um_per_px_avg)

        gradient_mag = np.sqrt(cv2.Sobel(img8, cv2.CV_32F, 1, 0, ksize=3) ** 2 + cv2.Sobel(img8, cv2.CV_32F, 0, 1, ksize=3) ** 2)
        mask = segment_particles_brightfield(img8, float(um_per_px_avg), img_out, bacteria_config)
        contours = cast(list[np.ndarray], cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2])

        vis_all = cv2.cvtColor(img8, cv2.COLOR_GRAY2BGR)
        cv2.drawContours(vis_all, contours, -1, (0, 0, 255), 1)
        save_debug(img_out, "10_contours_all.png", vis_all, um_per_px_avg)

        min_area_px, max_area_px, max_big_area_px = bacteria_config.min_area_um2 / um2_per_px2, bacteria_config.max_area_um2 / um2_per_px2, bacteria_config.max_fraction_of_image * float(img8.shape[0] * img8.shape[1])
        _contour_mask = np.zeros(img8.shape[:2], dtype=np.uint8)

        for contour_idx, c in enumerate(contours, 1):
            area_px = float(cv2.contourArea(c))
            if area_px <= 0 or area_px >= max_big_area_px or area_px < min_area_px or area_px > max_area_px:
                rejected.append(c); rejection_reasons.append({'contour_index': contour_idx, 'area_px': area_px, 'reasons': ['Area out of bounds']})
                continue

            perim_px = float(cv2.arcLength(c, True))
            circ = (4.0 * np.pi * area_px / (perim_px ** 2)) if perim_px > 0 else 0.0
            x_bb, y_bb, w_bb, h_bb = cv2.boundingRect(c)
            aspect_ratio = max(float(w_bb), float(h_bb)) / min(float(w_bb), float(h_bb)) if min(float(w_bb), float(h_bb)) > 0 else 1.0
            hull_area = float(cv2.contourArea(cv2.convexHull(c)))
            solidity = area_px / hull_area if hull_area > 0 else 0.0

            _contour_mask[:] = 0
            cv2.drawContours(_contour_mask, [c], -1, 255, thickness=-1)
            _bf_pixels = img8[_contour_mask > 0]

            if _bf_pixels.size > 0:
                bf_mean, bf_median, bf_std, bf_min_val, bf_max_val, bf_integrated, mean_gradient = float(np.mean(_bf_pixels)), float(np.median(_bf_pixels)), float(np.std(_bf_pixels)), float(np.min(_bf_pixels)), float(np.max(_bf_pixels)), float(np.sum(_bf_pixels.astype(np.float64))), float(np.mean(gradient_mag[_contour_mask > 0]))
            else: bf_mean = bf_median = bf_std = bf_min_val = bf_max_val = bf_integrated = mean_gradient = 0.0

            rejection_info = {'contour_index': contour_idx, 'area_px': area_px, 'reasons': [], 'perim_px': perim_px, 'circularity': circ, 'aspect_ratio': aspect_ratio, 'solidity': solidity, 'bbox': (x_bb, y_bb, w_bb, h_bb), 'bf_mean': bf_mean, 'bf_median': bf_median, 'bf_std': bf_std, 'bf_min_val': bf_min_val, 'bf_max_val': bf_max_val, 'bf_integrated': bf_integrated, 'mean_gradient': mean_gradient}

            passed = True
            if circ < bacteria_config.min_circularity or circ > bacteria_config.max_circularity: passed = False; rejection_info['reasons'].append('Circularity out of bounds')
            if aspect_ratio > bacteria_config.max_aspect_ratio or (bacteria_config.min_aspect_ratio > 1.0 and aspect_ratio < bacteria_config.min_aspect_ratio): passed = False; rejection_info['reasons'].append('Aspect ratio out of bounds')
            if solidity < bacteria_config.min_solidity: passed = False; rejection_info['reasons'].append('Solidity too low')
            if bf_mean < bacteria_config.min_mean_intensity_bf or bf_mean > bacteria_config.max_mean_intensity_bf: passed = False; rejection_info['reasons'].append('BF mean intensity out of bounds')
            if mean_gradient > bacteria_config.max_edge_gradient: passed = False; rejection_info['reasons'].append('Mean edge gradient too high')

            if passed: accepted.append(c)
            else: rejected.append(c); rejection_reasons.append(rejection_info)

        accepted_ids = [f"{group_id}_{sequence_num}_{i}" for i in range(1, len(accepted) + 1)]
        rejected_ids = [f"{group_id}_{sequence_num}_REJ{i}" for i in range(1, len(rejected) + 1)]

        if accepted and img8 is not None:
            try:
                with open(img_out / "morphological_features.json", 'w', encoding='utf-8') as _mf:
                    json.dump(extract_discriminating_features(accepted, [(cv2.moments(_c)["m10"] / cv2.moments(_c)["m00"], cv2.moments(_c)["m01"] / cv2.moments(_c)["m00"]) if cv2.moments(_c)["m00"] != 0 else (0.0, 0.0) for _c in accepted], img8.astype(np.float32)), _mf, default=float)
            except Exception: pass

        vis_acc = cv2.cvtColor(img8, cv2.COLOR_GRAY2BGR)
        cv2.drawContours(vis_acc, rejected, -1, (0, 165, 255), 1)
        cv2.drawContours(vis_acc, accepted, -1, (0, 255, 255), 1)
        save_debug(img_out, "11_contours_rejected_orange_accepted_yellow_ids_green.png", draw_object_ids(vis_acc, accepted, labels=accepted_ids), um_per_px_avg, force=True)

        mask_acc = np.zeros_like(mask)
        cv2.drawContours(mask_acc, accepted, -1, 255, thickness=-1)
        save_debug(img_out, "13_mask_accepted.png", mask_acc)

        fluor_path = img_path.parent / img_path.name.replace("_ch00", "_ch01")
        fluor_measurements = None
        fluor_bw, vis_fluor, vis_match = None, None, None

        if fluor_path.exists():
            fluor_img = safe_imread(fluor_path, cv2.IMREAD_UNCHANGED)
            if fluor_img is not None:
                fluor_img_aligned, (sy, sx), diagnostics = align_fluorescence_channel(img, fluor_img)
                _shift_detected, _alignment_method = (float(sy), float(sx)), diagnostics.get('method_used', 'none')

                if DEBUG_MODE:
                    safe_imwrite(img_out / "DIAG_A_no_shift.png", diagnostics['overlay_none'])
                    safe_imwrite(img_out / "DIAG_B_positive_shift.png", diagnostics['overlay_pos'])
                    safe_imwrite(img_out / "DIAG_C_negative_shift.png", diagnostics['overlay_neg'])

                fluor_img8 = normalize_to_8bit(cv2.cvtColor(fluor_img_aligned, cv2.COLOR_BGR2GRAY) if fluor_img_aligned.ndim == 3 else fluor_img_aligned)
                try: _fluor_otsu = float(cv2.threshold(cv2.GaussianBlur(fluor_img8, (0, 0), sigmaX=bacteria_config.gaussian_sigma * 0.1, sigmaY=bacteria_config.gaussian_sigma * 0.1), 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[0])
                except Exception: _fluor_otsu = 0.0

                fluor_bw = segment_fluorescence_global(fluor_img8, bacteria_config)
                fluor_contours = [c for c in cast(list[np.ndarray], cv2.findContours(fluor_bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2]) if float(cv2.contourArea(c)) >= (bacteria_config.fluor_min_area_um2 / um2_per_px2 if um2_per_px2 > 0 else 0.0)]

                vis_fluor = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
                cv2.drawContours(vis_fluor, fluor_contours, -1, (0, 255, 0), 1)

                matches = match_fluor_to_bf_by_overlap(accepted, fluor_contours, fluor_img8.shape[:2], min_intersection_px=bacteria_config.fluor_match_min_intersection_px)
                fluor_measurements = measure_fluorescence_intensity_with_global_area(fluor_img_aligned, accepted, fluor_contours, matches, float(um_per_px_x), float(um_per_px_y))

                vis_match = cv2.cvtColor(fluor_img8, cv2.COLOR_GRAY2BGR)
                for _idx_m, _bf_c in enumerate(accepted):
                    cv2.drawContours(vis_match, [_bf_c], -1, (0, 255, 255), 1)
                    _match_idx = matches[_idx_m]
                    if _match_idx is not None:
                        cv2.drawContours(vis_match, [fluor_contours[int(_match_idx)]], -1, (0, 0, 255), 2)
                save_debug(img_out, "24_bf_fluor_matching_overlay.png", vis_match, um_per_px_avg)

        save_debug_ids(img_out, "13_mask_accepted.png", cv2.cvtColor(mask_acc, cv2.COLOR_GRAY2BGR), accepted, accepted_ids, um_per_px_avg)
        if fluor_bw is not None: save_debug_ids(img_out, "22_fluorescence_mask_global.png", cv2.cvtColor(fluor_bw, cv2.COLOR_GRAY2BGR), accepted, accepted_ids, um_per_px_avg)
        if vis_fluor is not None: save_debug_ids(img_out, "23_fluorescence_contours_global.png", vis_fluor, accepted, accepted_ids, um_per_px_avg)
        if vis_match is not None: save_debug_ids(img_out, "24_bf_fluor_matching_overlay.png", vis_match, accepted, accepted_ids, um_per_px_avg, force=True)

        with open(img_out / "rejected_objects.csv", "w", newline="", encoding="utf-8") as _f_rej:
            _w_rej = csv.writer(_f_rej)
            _w_rej.writerow(["Object_ID", "Rejection_Reasons", "BF_Area_px", "BF_Area_um2", "Perimeter_px", "Circularity", "AspectRatio", "Solidity", "BF_Mean_Intensity", "BF_Median_Intensity", "BF_Std_Intensity", "BF_Min_Intensity", "BF_Max_Intensity", "BF_IntegratedDensity", "Mean_Edge_Gradient", "BBoxX_px", "BBoxY_px", "BBoxW_px", "BBoxH_px", "CentroidX_px", "CentroidY_px"])
            for _i_rej, (_c_rej, _rej_info) in enumerate(zip(rejected, rejection_reasons)):
                _M_r = cv2.moments(_c_rej)
                _cx_r, _cy_r = (float(_M_r["m10"] / max(_M_r["m00"], 1.0)), float(_M_r["m01"] / max(_M_r["m00"], 1.0))) if _M_r["m00"] != 0 else (0.0, 0.0)
                _w_rej.writerow([rejected_ids[_i_rej] if _i_rej < len(rejected_ids) else f"{group_id}_{sequence_num}_REJ{_i_rej + 1}", "; ".join(_rej_info['reasons']), f"{_rej_info['area_px']:.2f}", f"{_rej_info['area_px'] * um2_per_px2:.4f}", f"{_rej_info.get('perim_px', 0.0):.2f}", f"{_rej_info.get('circularity', 0.0):.4f}", f"{_rej_info.get('aspect_ratio', 0.0):.4f}", f"{_rej_info.get('solidity', 0.0):.4f}", f"{_rej_info.get('bf_mean', 0.0):.2f}", f"{_rej_info.get('bf_median', 0.0):.2f}", f"{_rej_info.get('bf_std', 0.0):.2f}", f"{_rej_info.get('bf_min_val', 0.0):.2f}", f"{_rej_info.get('bf_max_val', 0.0):.2f}", f"{_rej_info.get('bf_integrated', 0.0):.1f}", f"{_rej_info.get('mean_gradient', 0.0):.2f}", *_rej_info.get('bbox', (0, 0, 0, 0)), f"{_cx_r:.2f}", f"{_cy_r:.2f}"])

        with open(img_out / "object_stats.csv", "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Object_ID", "Status", "BF_Area_px", "BF_Area_um2", "Perimeter_px", "Perimeter_um", "EquivDiameter_px", "EquivDiameter_um", "Circularity", "AspectRatio", "Solidity", "CentroidX_px", "CentroidY_px", "CentroidX_um", "CentroidY_um", "BBoxX_px", "BBoxY_px", "BBoxW_px", "BBoxH_px", "BBoxW_um", "BBoxH_um", "BF_Mean_Intensity", "BF_Median_Intensity", "BF_Std_Intensity", "BF_Min_Intensity", "BF_Max_Intensity", "BF_IntegratedDensity", "Fluor_Area_px", "Fluor_Area_um2", "Fluor_Mean", "Fluor_Median", "Fluor_Std", "Fluor_Min", "Fluor_Max", "Fluor_IntegratedDensity"])
            for i_acc, c_acc in enumerate(accepted, 1):
                a_area_px, a_perim_px = float(cv2.contourArea(c_acc)), float(cv2.arcLength(c_acc, True))
                _ax, _ay, _aw, _ah = cv2.boundingRect(c_acc)
                _a_hull_area = float(cv2.contourArea(cv2.convexHull(c_acc)))
                M_a = cv2.moments(c_acc)
                a_cx, a_cy = (float(M_a.get("m10", 0.0)) / float(max(M_a.get("m00", 1.0), 1.0)), float(M_a.get("m01", 0.0)) / float(max(M_a.get("m00", 1.0), 1.0))) if M_a.get("m00", 0.0) != 0 else (0.0, 0.0)

                _contour_mask[:] = 0; cv2.drawContours(_contour_mask, [c_acc], -1, 255, thickness=-1)
                _acc_bf_pixels = img8[_contour_mask > 0]
                a_bf_mean, a_bf_median, a_bf_std, a_bf_min, a_bf_max, a_bf_integ = (float(np.mean(_acc_bf_pixels)), float(np.median(_acc_bf_pixels)), float(np.std(_acc_bf_pixels)), float(np.min(_acc_bf_pixels)), float(np.max(_acc_bf_pixels)), float(np.sum(_acc_bf_pixels.astype(np.float64)))) if _acc_bf_pixels.size > 0 else (0.0, 0.0, 0.0, 0.0, 0.0, 0.0)
                fm = fluor_measurements[i_acc - 1] if fluor_measurements else {"fluor_area_px": 0.0, "fluor_area_um2": 0.0, "fluor_mean": 0.0, "fluor_median": 0.0, "fluor_std": 0.0, "fluor_min": 0.0, "fluor_max": 0.0, "fluor_integrated_density": 0.0}

                w.writerow([accepted_ids[i_acc - 1], "Accepted", f"{a_area_px:.2f}", f"{a_area_px * um2_per_px2:.4f}", f"{a_perim_px:.2f}", f"{contour_perimeter_um(c_acc, float(um_per_px_x), float(um_per_px_y)):.4f}", f"{equivalent_diameter_from_area(a_area_px):.2f}", f"{equivalent_diameter_from_area(a_area_px * um2_per_px2):.4f}", f"{(4.0 * np.pi * a_area_px / (a_perim_px ** 2)) if a_perim_px > 0 else 0.0:.4f}", f"{max(float(_aw), float(_ah)) / min(float(_aw), float(_ah)) if min(float(_aw), float(_ah)) > 0 else 1.0:.4f}", f"{a_area_px / _a_hull_area if _a_hull_area > 0 else 0.0:.4f}", f"{a_cx:.2f}", f"{a_cy:.2f}", f"{a_cx * float(um_per_px_x):.4f}", f"{a_cy * float(um_per_px_y):.4f}", _ax, _ay, _aw, _ah, f"{float(_aw) * float(um_per_px_x):.4f}", f"{float(_ah) * float(um_per_px_y):.4f}", f"{a_bf_mean:.2f}", f"{a_bf_median:.2f}", f"{a_bf_std:.2f}", f"{a_bf_min:.2f}", f"{a_bf_max:.2f}", f"{a_bf_integ:.1f}", f"{float(fm['fluor_area_px']):.2f}", f"{float(fm['fluor_area_um2']):.4f}", f"{float(fm['fluor_mean']):.2f}", f"{float(fm['fluor_median']):.2f}", f"{float(fm['fluor_std']):.2f}", f"{float(fm['fluor_min']):.2f}", f"{float(fm['fluor_max']):.2f}", f"{float(fm['fluor_integrated_density']):.2f}"])

    except Exception: pass
    finally:
        if profiler:
            try: profiler.record_image(image_name=img_path.name, group=output_root.name, bacteria_config=bacteria_config.name if bacteria_config else "Unknown", processing_time_s=pytime.monotonic() - _img_start, accepted=len(accepted), rejected=len(rejected), alignment_method=_alignment_method, shift_px=_shift_detected, fluor_threshold_otsu=_fluor_otsu)
            except Exception: pass
        import gc; gc.collect()


def run_single_config_analysis(config: dict, *, profiler=None, skip_processing: bool = False) -> None:
    bacteria_config = (
        load_bacteria_config_from_json(config.get('bacteria_config'))
        if isinstance(config.get('bacteria_config'), str)
        else config.get('bacteria_config')
    )
    if not isinstance(bacteria_config, SegmentationConfig):
        return

    source_dir    = config.get('current_source')
    output_dir    = config.get('output_dir')
    percentile    = config['percentile']
    threshold_pct = config['threshold_pct']

    if not isinstance(source_dir, Path) or not isinstance(output_dir, Path):
        return

    microgel_type = 'negative' if 'Negative' in str(output_dir) else 'positive'

    with open(output_dir / "segmentation_config.txt", 'w', encoding='utf-8') as f:
        f.write(
            f"Configuration: {bacteria_config.name}\n"
            f"Description: {bacteria_config.description}\n"
            f"Gaussian Sigma: {bacteria_config.gaussian_sigma}\n"
            f"Area Range: {bacteria_config.min_area_um2} - {bacteria_config.max_area_um2} \u00b5m\u00b2\n"
        )

    image_groups = collect_images_from_directory(source_dir)

    if not skip_processing:
        total_images = sum(len(v['images']) for v in image_groups.values())
        print(
            f"\n  Processing {len(image_groups)} group(s), "
            f"{total_images} image(s)  [{bacteria_config.name}]"
        )
        _rsa_start = pytime.monotonic()

        for group_name, group_data in image_groups.items():
            n_imgs = len(group_data['images'])
            for img_path in tqdm(
                group_data['images'],
                desc=f"  {group_name:<14} ({n_imgs} img{'s' if n_imgs != 1 else ''})",
                unit="img",
                ncols=88,
            ):
                try:
                    process_image(
                        img_path,
                        output_dir / group_name,
                        bacteria_config,
                        profiler=profiler,
                    )
                except Exception:
                    pass

        print(f"  \u2713 Images done  ({_fmt_elapsed(pytime.monotonic() - _rsa_start)})")

        print("  Consolidating to Excel\u2026")
        _xl_start = pytime.monotonic()
        for group_name in tqdm(
            list(image_groups.keys()),
            desc="  Excel",
            unit="group",
            ncols=88,
        ):
            group_output_dir = output_dir / group_name
            if group_output_dir.exists():
                consolidate_to_excel(group_output_dir, group_name, percentile)
        print(f"  \u2713 Excel done  ({_fmt_elapsed(pytime.monotonic() - _xl_start)})")

    export_group_statistics_to_csv(output_dir)

    classification_df = classify_groups_clinical(output_dir, microgel_type, threshold_pct)
    if not classification_df.empty:
        export_clinical_classification(output_dir, classification_df, microgel_type)

    plot_path = generate_error_bar_comparison_with_threshold(
        output_dir, percentile,
        threshold_pct=threshold_pct,
        microgel_type=microgel_type,
        dataset_id=config.get('dataset_id_current', ''),
    )

    generate_pairwise_group_vs_control_plots(
        output_dir, percentile,
        config.get('dataset_id_current', ''),
        threshold_pct, microgel_type,
    )

    if plot_path is not None and plot_path.exists():
        embed_comparison_plots_into_all_excels(output_dir, percentile, plot_path)


def _phase(profiler, name: str, **kwargs): return profiler.phase(name, **kwargs) if profiler else nullcontext()
def _safe_record(fn, /, *args, _label: str = "profiler call", **kwargs) -> None:
    try: fn(*args, **kwargs)
    except Exception: pass
def _record_final_classifications(profiler, output_root: Optional[Path]) -> None:
    if not profiler or not output_root or not (output_root / "final_clinical_results.csv").exists(): return
    try:
        df_final = pd.read_csv(output_root / "final_clinical_results.csv")
        for _, row in df_final.iterrows():
            _safe_record(
                profiler.record_final_classification,
                group=str(row.get("Group", "")),
                final_class=str(row.get("Final_Classification", "")),
                gp_class=str(row.get("G+_Classification", "")),
                gm_class=str(row.get("G-_Classification", "")),
                _label="record_final_classification",
            )
    except Exception: pass


# ==================================================
# Overall Progress Tracker
# ==================================================

class OverallProgressTracker:
    """
    Pipeline-level progress indicator.

    Prints a numbered phase header when work begins and a ✓ summary when it
    ends, so the terminal never appears idle between tqdm bars.

    Usage
    -----
    prog = OverallProgressTracker(["G+ Scan", "G− Scan", "Reports"])
    prog.begin("2 configs")   # ─── [1/3] ▶  G+ Scan  —  2 configs ───
    ...work...
    prog.done("top: K.p.")    # ✓  G+ Scan  (14.2 s)  ·  top: K.p.
    prog.status("selecting")  #  ·  selecting  [16.0 s elapsed]
    prog.final_summary()      # ═══ ✓  Pipeline complete  ·  total: 38s ═══
    """
    _W = 80

    def __init__(self, phases: list[str]) -> None:
        self.phases   = phases
        self._n       = len(phases)
        self._idx     = 0
        self._label   = ""
        self._t_phase = 0.0
        self._t_start = pytime.monotonic()

    def begin(self, detail: str = "") -> None:
        """Advance to the next phase and print its header."""
        lbl           = self.phases[self._idx] if self._idx < self._n else "(extra)"
        self._label   = lbl
        self._t_phase = pytime.monotonic()
        num           = f"[{self._idx + 1}/{self._n}]"
        det           = f"  \u2014  {detail}" if detail else ""
        print(
            f"\n{'─' * self._W}\n"
            f"  {num}  \u25b6  {lbl}{det}\n"
            f"{'─' * self._W}",
            flush=True,
        )
        self._idx += 1

    def done(self, note: str = "") -> None:
        """Mark the current phase complete with elapsed time."""
        t   = pytime.monotonic() - self._t_phase
        sfx = f"  \u00b7  {note}" if note else ""
        print(f"  \u2713  {self._label}  ({_fmt_elapsed(t)}){sfx}", flush=True)

    def status(self, msg: str) -> None:
        """Emit a non-phase status line (config selection, promotions, etc.)."""
        t = pytime.monotonic() - self._t_start
        print(f"  \u00b7  {msg}  [{_fmt_elapsed(t)} elapsed]", flush=True)

    def final_summary(self) -> None:
        """Print total pipeline elapsed time."""
        t = pytime.monotonic() - self._t_start
        print(
            f"\n{'═' * self._W}\n"
            f"  \u2713  Pipeline complete  \u00b7  total: {_fmt_elapsed(t)}\n"
            f"{'═' * self._W}",
            flush=True,
        )

# ==================================================
# Main Function
# ==================================================


def main() -> None:
    print("\n" + "=" * 80 + "\nMICROGEL FLUORESCENCE ANALYSIS PIPELINE\n" + "=" * 80 + "\n")
    try:
        from run_profiler import RunProfiler as _RunProfiler
        _profiler_available = True
    except ImportError:
        _profiler_available, _RunProfiler = False, None

    output_root, config, positive_results, negative_results, results = None, {}, None, None, None
    profiler = (
        _RunProfiler(run_id=datetime.now().strftime("%Y%m%d_%H%M%S"), project_root=PROJECT_ROOT)
        if _profiler_available and _RunProfiler else None
    )

    try:
        with _phase(profiler, "step1_config_selection"):
            bacteria_config_info = select_bacteria_config()
        mode = bacteria_config_info['mode']
        config['bacteria_config_info'], config['processing_mode'] = bacteria_config_info, mode
        if profiler:
            try:
                profiler.summary["mode"] = "DEBUG" if DEBUG_MODE else "DEFAULT"
                profiler.summary["processing_mode"] = mode
            except Exception:
                pass

        with _phase(profiler, "step2_dataset_config"):
            config.update(configure_dataset())
        if profiler:
            try:
                profiler.summary["dataset_id"]   = config.get("dataset_id", "")
                profiler.summary["dataset_mode"] = "batch" if config.get("batch_mode") else "single"
            except Exception:
                pass

        if not validate_config(config) or (config.get('batch_mode') and not validate_batch_structure(config)):
            return

        display_configuration_summary(config)
        config['timestamp'] = datetime.now().strftime("%Y%m%d_%H%M%S")

        with _phase(profiler, "setup_output_directory"):
            output_root = setup_output_directory(config)

        # ── Build phase list now that mode + batch_mode are known ────────
        _batch  = config.get('batch_mode', False)
        _n_cfg  = len(bacteria_config_info.get('configs_to_scan', []))
        _cfg_d  = f"{_n_cfg} config{'s' if _n_cfg != 1 else ''}"

        if mode == "multi_scan":
            if _batch:
                _phases = [
                    "G+  Multi-scan",
                    "G\u2212  Multi-scan",
                    "G+  Clinical Analysis",
                    "G\u2212  Clinical Analysis",
                    "Final Matrix \u0026 Reports",
                    "Cleanup",
                ]
            else:
                _phases = ["Multi-scan", "Clinical Analysis", "Rejection Analysis"]
        else:
            if _batch:
                _phases = [
                    "G+  Clinical Analysis",
                    "G\u2212  Clinical Analysis",
                    "Final Matrix \u0026 Reports",
                    "Cleanup",
                ]
            else:
                _phases = ["Clinical Analysis"]

        progress = OverallProgressTracker(_phases)

        # ════════════════════════════════════════════════════════════════
        if mode == "multi_scan":
            if _batch:
                # ── Phase 1: G+ Multi-scan ───────────────────────────
                config['current_source']     = config['source_dir_positive']
                config['output_dir']         = config['positive_output']
                config['dataset_id_current'] = f"{config['dataset_id']} Positive"
                progress.begin(_cfg_d)
                with _phase(profiler, "multi_scan_G+",
                            source=str(config['source_dir_positive']),
                            bacteria_configs=_n_cfg):
                    positive_results = run_multi_config_scan(config, bacteria_config_info)
                if profiler and positive_results:
                    pos_ranked = positive_results.get('ranked_results', [])
                    _safe_record(
                        profiler.record_multi_scan_result, channel="G+",
                        ranked_results=pos_ranked,
                        stat_ambiguous=(
                            isinstance(pos_ranked, list) and len(pos_ranked) >= 2
                            and abs(float(pos_ranked[0].get('confidence', 0))
                                    - float(pos_ranked[1].get('confidence', 0))) < 10.0
                        ),
                        _label="record_multi_scan_result G+",
                    )
                _top_pos = (
                    positive_results.get('ranked_results', [{}])[0].get('bacteria_name', '\u2014')
                    if positive_results and positive_results.get('ranked_results') else '\u2014'
                )
                progress.done(f"top match: {_top_pos}")

                # ── Phase 2: G− Multi-scan ───────────────────────────
                config['current_source']     = config['source_dir_negative']
                config['output_dir']         = config['negative_output']
                config['dataset_id_current'] = f"{config['dataset_id']} Negative"
                progress.begin(_cfg_d)
                with _phase(profiler, "multi_scan_G-",
                            source=str(config['source_dir_negative']),
                            bacteria_configs=_n_cfg):
                    negative_results = run_multi_config_scan(config, bacteria_config_info)
                if profiler and negative_results:
                    neg_ranked = negative_results.get('ranked_results', [])
                    _safe_record(
                        profiler.record_multi_scan_result, channel="G-",
                        ranked_results=neg_ranked,
                        stat_ambiguous=(
                            len(neg_ranked) >= 2
                            and abs(neg_ranked[0].get('confidence', 0)
                                    - neg_ranked[1].get('confidence', 0)) < 10.0
                        ),
                        _label="record_multi_scan_result G-",
                    )
                _top_neg = (
                    negative_results.get('ranked_results', [{}])[0].get('bacteria_name', '\u2014')
                    if negative_results and negative_results.get('ranked_results') else '\u2014'
                )
                progress.done(f"top match: {_top_neg}")

                # ── Config selection (status only, no phase tick) ────
                best_positive = (
                    positive_results.get('ranked_results', [])[0]
                    if positive_results and positive_results.get('ranked_results') else None
                )
                best_negative = (
                    negative_results.get('ranked_results', [])[0]
                    if negative_results and negative_results.get('ranked_results') else None
                )
                chosen_config_key, selection_rule = None, None

                if best_positive and best_negative:
                    if best_positive['config_key'] == best_negative['config_key']:
                        chosen_config_key = best_positive['config_key']
                        selection_rule    = "G+ and G\u2212 best matches agree"
                    elif best_positive['confidence'] >= best_negative['confidence']:
                        chosen_config_key = best_positive['config_key']
                        selection_rule    = (
                            f"G+ higher confidence "
                            f"({best_positive['confidence']:.1f}% vs "
                            f"{best_negative['confidence']:.1f}%)"
                        )
                    else:
                        chosen_config_key = best_negative['config_key']
                        selection_rule    = (
                            f"G\u2212 higher confidence "
                            f"({best_negative['confidence']:.1f}% vs "
                            f"{best_positive['confidence']:.1f}%)"
                        )
                elif best_positive:
                    chosen_config_key, selection_rule = (
                        best_positive['config_key'], "only G+ results available"
                    )
                elif best_negative:
                    chosen_config_key, selection_rule = (
                        best_negative['config_key'], "only G\u2212 results available"
                    )

                if chosen_config_key is None:
                    progress.status("config selection failed \u2014 no ranked results available")
                    if profiler:
                        _safe_record(profiler.record_decision, "config_selection_failed",
                                     reason="No ranked results available from either channel",
                                     _label="record_decision config_selection_failed")
                else:
                    progress.status(
                        f"selected config: {chosen_config_key}  ({selection_rule})"
                    )
                    if profiler:
                        try:
                            profiler.summary["chosen_config"] = chosen_config_key
                            profiler.summary["selection_rule"] = selection_rule or ""
                        except Exception:
                            pass
                        _safe_record(profiler.record_decision, "config_chosen",
                                     config_key=chosen_config_key, rule=selection_rule,
                                     pos_confidence=float(best_positive['confidence']) if best_positive else None,
                                     neg_confidence=float(best_negative['confidence']) if best_negative else None,
                                     _label="record_decision config_chosen")

                    chosen_bacteria_config = (
                        positive_results.get('configs', {}).get(chosen_config_key)
                        if positive_results else None
                    )
                    if not chosen_bacteria_config and negative_results:
                        chosen_bacteria_config = (
                            negative_results.get('configs', {}).get(chosen_config_key)
                        )
                    if not chosen_bacteria_config:
                        chosen_bacteria_config = load_bacteria_config_from_json(chosen_config_key)

                    if chosen_bacteria_config is None:
                        progress.status(f"could not resolve config '{chosen_config_key}'")
                        if profiler:
                            _safe_record(profiler.record_decision, "config_resolve_failed",
                                         config_key=chosen_config_key,
                                         _label="record_decision config_resolve_failed")
                    else:
                        config['bacteria_config'] = chosen_bacteria_config

                        # ── Phase 3: G+ Clinical Analysis ────────────
                        config['current_source']     = config['source_dir_positive']
                        config['output_dir']         = config['positive_output']
                        config['dataset_id_current'] = f"{config['dataset_id']} Positive"
                        promoted_pos = promote_multiscan_output_as_clinical(config, chosen_config_key)
                        progress.begin(
                            "promoted \u2014 skipping re-processing" if promoted_pos
                            else chosen_config_key
                        )
                        with _phase(profiler, "clinical_run_G+", config_key=chosen_config_key):
                            run_single_config_analysis(config, profiler=profiler,
                                                       skip_processing=promoted_pos)
                        progress.done()

                        # ── Phase 4: G− Clinical Analysis ────────────
                        config['current_source']     = config['source_dir_negative']
                        config['output_dir']         = config['negative_output']
                        config['dataset_id_current'] = f"{config['dataset_id']} Negative"
                        promoted_neg = promote_multiscan_output_as_clinical(config, chosen_config_key)
                        progress.begin(
                            "promoted \u2014 skipping re-processing" if promoted_neg
                            else chosen_config_key
                        )
                        with _phase(profiler, "clinical_run_G-", config_key=chosen_config_key):
                            run_single_config_analysis(config, profiler=profiler,
                                                       skip_processing=promoted_neg)
                        progress.done()

                        if output_root:
                            # ── Phase 5: Final Matrix & Reports ───────
                            progress.begin()
                            with _phase(profiler, "generate_final_matrix"):
                                generate_final_clinical_matrix_wrapper(output_root, config)
                            with _phase(profiler, "rejection_analysis"):
                                generate_rejection_analysis(output_root)
                            _record_final_classifications(profiler, output_root)
                            progress.done()

            else:
                # ── single-dir multi-scan ────────────────────────────
                config['current_source']     = config['source_dir']
                config['output_dir']         = output_root
                config['dataset_id_current'] = config['dataset_id']
                progress.begin(_cfg_d)
                with _phase(profiler, "multi_scan",
                            source=str(config['source_dir']),
                            bacteria_configs=_n_cfg):
                    results = run_multi_config_scan(config, bacteria_config_info)
                ranked_results = results.get('ranked_results', []) if results else []
                _top = (
                    ranked_results[0].get('bacteria_name', '\u2014')
                    if ranked_results else '\u2014'
                )
                progress.done(f"top match: {_top}")

                if ranked_results:
                    chosen_config_key = cast(str, ranked_results[0]['config_key'])
                    progress.status(
                        f"selected config: {chosen_config_key}  "
                        f"(confidence: {ranked_results[0].get('confidence', 0):.1f}%)"
                    )
                    if profiler:
                        try:
                            profiler.summary["chosen_config"]  = chosen_config_key
                            profiler.summary["selection_rule"] = "top-ranked single-dir scan"
                        except Exception:
                            pass
                        _safe_record(profiler.record_decision, "config_chosen",
                                     config_key=chosen_config_key,
                                     confidence=float(ranked_results[0]['confidence']),
                                     _label="record_decision config_chosen")
                        _safe_record(profiler.record_multi_scan_result, channel="single",
                                     ranked_results=ranked_results,
                                     stat_ambiguous=(
                                         len(ranked_results) >= 2
                                         and abs(ranked_results[0].get('confidence', 0)
                                                 - ranked_results[1].get('confidence', 0)) < 10.0
                                     ),
                                     _label="record_multi_scan_result single")

                    chosen_bacteria_config = (
                        bacteria_config_info.get('configs', {}).get(chosen_config_key)
                        or load_bacteria_config_from_json(chosen_config_key)
                    )
                    if chosen_bacteria_config is None:
                        progress.status(f"could not resolve config '{chosen_config_key}'")
                        if profiler:
                            _safe_record(profiler.record_decision, "config_resolve_failed",
                                         config_key=chosen_config_key,
                                         _label="record_decision config_resolve_failed")
                    else:
                        config['bacteria_config'] = chosen_bacteria_config

                        # Phase 2: Clinical Analysis
                        progress.begin(chosen_config_key)
                        with _phase(profiler, "clinical_run_single",
                                    config_key=chosen_config_key):
                            run_single_config_analysis(config, profiler=profiler)
                        progress.done()

                        if output_root:
                            # Phase 3: Rejection Analysis
                            progress.begin()
                            if config.get('batch_mode', False):
                                with _phase(profiler, "generate_final_matrix"):
                                    generate_final_clinical_matrix_wrapper(output_root, config)
                            with _phase(profiler, "rejection_analysis"):
                                generate_rejection_analysis(output_root)
                            _record_final_classifications(profiler, output_root)
                            progress.done()

        else:
            # ── single-config mode ───────────────────────────────────
            config['bacteria_config'] = bacteria_config_info['selected_config']
            if profiler:
                try:
                    profiler.summary["chosen_config"]  = cast(str, bacteria_config_info.get('bacteria_type', 'unknown'))
                    profiler.summary["selection_rule"] = "user-selected single config"
                except Exception:
                    pass
                _safe_record(profiler.record_decision, "config_chosen",
                             config_key=cast(str, bacteria_config_info.get('bacteria_type', 'unknown')),
                             mode="single_config",
                             _label="record_decision config_chosen single_config")

            if _batch:
                # Phase 1: G+ Clinical Analysis
                config['current_source']     = config['source_dir_positive']
                config['output_dir']         = config['positive_output']
                config['dataset_id_current'] = f"{config['dataset_id']} Positive"
                progress.begin()
                with _phase(profiler, "clinical_run_G+"):
                    run_single_config_analysis(config, profiler=profiler)
                progress.done()

                # Phase 2: G− Clinical Analysis
                config['current_source']     = config['source_dir_negative']
                config['output_dir']         = config['negative_output']
                config['dataset_id_current'] = f"{config['dataset_id']} Negative"
                progress.begin()
                with _phase(profiler, "clinical_run_G-"):
                    run_single_config_analysis(config, profiler=profiler)
                progress.done()

                if output_root:
                    # Phase 3: Final Matrix & Reports
                    progress.begin()
                    with _phase(profiler, "generate_final_matrix"):
                        generate_final_clinical_matrix_wrapper(output_root, config)
                    with _phase(profiler, "rejection_analysis"):
                        generate_rejection_analysis(output_root)
                    _record_final_classifications(profiler, output_root)
                    progress.done()
            else:
                # Phase 1: Clinical Analysis
                config['current_source']     = config['source_dir']
                config['output_dir']         = output_root
                config['dataset_id_current'] = config['dataset_id']
                progress.begin()
                with _phase(profiler, "clinical_run_single"):
                    run_single_config_analysis(config, profiler=profiler)
                progress.done()

        # ── Cleanup (batch modes only, final phase) ──────────────────
        if _batch and output_root:
            progress.begin()
            with _phase(profiler, "cleanup"):
                cleanup_and_reorganize_output(output_root, config)
            progress.done()

        progress.final_summary()

        if output_root:
            try:
                open_folder(output_root)
            except Exception:
                pass

    except KeyboardInterrupt:
        if profiler:
            _safe_record(profiler.record_decision, "interrupted",
                         reason="KeyboardInterrupt",
                         _label="record_decision interrupted")
    except Exception as exc:
        if profiler:
            _safe_record(profiler.record_decision, "error",
                         message=str(exc),
                         _label="record_decision error")
    finally:
        try:
            if _log_file is not None:
                _log_file.flush()
                if hasattr(_log_file, 'fileno'):
                    try:
                        os.fsync(_log_file.fileno())
                    except Exception:
                        pass
        except Exception:
            pass

        if profiler:
            try:
                profiler.finalise(output_dir=output_root)
            except Exception:
                pass

        try:
            if output_root is not None and _log_path is not None and _log_path.exists():
                log_destinations: list[Path] = [output_root]
                if config.get('batch_mode', False):
                    positive_output = config.get('positive_output')
                    if isinstance(positive_output, Path) and positive_output.exists():
                        log_destinations.append(positive_output)
                for dest in log_destinations:
                    copy_log_to_output(_log_path, dest)
        except Exception:
            pass



if __name__ == "__main__":
    start_time = pytime.time()
    main()
    print("="*80 + f"\n  \u2022 Total runtime: {pytime.time() - start_time:.1f} seconds")